"""Match workbook Counties / Municipality rows to per-country Overture parquets.

Outputs three candidate matches per workbook row so editorial choice is yours:

  1. Direct       - the row's own Overture admin record (its own name match)
  2. Parent County- the admin-level-2 county whose polygon contains the row's
                    centroid. Empty for countries that have no county tier
                    (Ireland, Brazil) where the region is the next level up.
  3. Parent Region- the admin-level-1 region whose polygon contains the row's
                    centroid. Useful for IE and BR where Ireland's "County
                    Kerry" or Brazil's "Sao Paulo (state)" is the appropriate
                    boundary unit.

Pick the column set per country based on your editorial preference for that
country's polygon granularity. Suggested defaults:

    IE: Parent Region (Ireland counties live at admin 1)
    BR: Direct (Brazilian municipios are localities)
    CN, AT, CH, BE, RU: Parent County

Sheet routing (matches build-metro-boundaries plans):
    Municipality:  IE, CN, AT, CH, BE
    Counties:      BR, RU

Output: Overture-Match-Suggestions/{ISO}.xlsx and {ISO}.csv per country.

Usage:
    python scripts/match-overture-to-workbook.py
    python scripts/match-overture-to-workbook.py IE BR
    python scripts/match-overture-to-workbook.py --metro-only
"""
from __future__ import annotations

import csv
import os
import sys
import time
import unicodedata
from collections import defaultdict
from difflib import SequenceMatcher
from pathlib import Path

import openpyxl
import pyarrow.parquet as pq
from openpyxl.styles import Font, PatternFill
from shapely import wkb as shapely_wkb
from shapely.strtree import STRtree


SOURCE_DIR = os.environ.get(
    "OVERTURE_PER_COUNTRY_DIR",
    r"C:\Users\ashwi\Desktop\Projects\MapData",
)
WORKBOOK = "MetroAreas.xlsx"
OUT_DIR = Path("Overture-Match-Suggestions")

COUNTRY_ROUTE = {
    "Ireland":     ("Municipality", "IE"),
    "China":       ("Municipality", "CN"),
    "Austria":     ("Municipality", "AT"),
    "Switzerland": ("Municipality", "CH"),
    "Belgium":     ("Municipality", "BE"),
    "Brazil":      ("Counties",     "BR"),
    "Russia":      ("Counties",     "RU"),
    # 2026-05-21 expansion (CSV dumps already in Overture-Per-Country-Raw;
    # workbook rows present in Counties but Subtype/Admin/Region/Primary blank):
    "Algeria":     ("Counties",     "DZ"),
    "Egypt":       ("Counties",     "EG"),
    "Finland":     ("Counties",     "FI"),
    "Ghana":       ("Counties",     "GH"),
    "Greece":      ("Counties",     "GR"),
    "Hungary":     ("Counties",     "HU"),
    "Indonesia":   ("Counties",     "ID"),
    "Israel":      ("Counties",     "IL"),
    "Kuwait":      ("Counties",     "KW"),
    "Moldova":     ("Counties",     "MD"),
    "Qatar":       ("Counties",     "QA"),
    "Sweden":      ("Counties",     "SE"),
    "Tunisia":     ("Counties",     "TN"),
    "Taiwan":      ("Counties",     "TW"),
    "Venezuela":   ("Counties",     "VE"),
    "Vietnam":     ("Counties",     "VN"),
    "Samoa":       ("Counties",     "WS"),
    # 2026-05-21 expansion (second wave): 142 additional entries spanning
    # territories, city-states, dependencies, and the long tail of UN
    # member states. Bonaire is the only BQ entry; Saba and Sint Eustatius
    # (single-row each in the workbook) are hand-filled from BQ.xlsx output.
    "French Guiana":                                ("Counties", "GF"),
    "Guadeloupe":                                   ("Counties", "GP"),
    "Martinique":                                   ("Counties", "MQ"),
    "Monaco":                                       ("Counties", "MC"),
    "Réunion":                                      ("Counties", "RE"),
    "Saint Pierre and Miquelon":                    ("Counties", "PM"),
    "Isle of Man":                                  ("Counties", "IM"),
    "Curaçao":                                      ("Counties", "CW"),
    "Cameroon":                                     ("Counties", "CM"),
    "Slovenia":                                     ("Counties", "SI"),
    "Iran":                                         ("Counties", "IR"),
    "Ukraine":                                      ("Counties", "UA"),
    "Tahiti":                                       ("Counties", "PF"),
    "Ethiopia":                                     ("Counties", "ET"),
    "Serbia":                                       ("Counties", "RS"),
    "Kazakhstan":                                   ("Counties", "KZ"),
    "Czech Republic":                               ("Counties", "CZ"),
    "Lithuania":                                    ("Counties", "LT"),
    "Uzbekistan":                                   ("Counties", "UZ"),
    "Bosnia-Herzegovina":                           ("Counties", "BA"),
    "North Korea":                                  ("Counties", "KP"),
    "Bolivia":                                      ("Counties", "BO"),
    "Wallis and Futuna":                            ("Counties", "WF"),
    "Iraq":                                         ("Counties", "IQ"),
    "Slovakia":                                     ("Counties", "SK"),
    "North Macedonia":                              ("Counties", "MK"),
    "South Africa":                                 ("Counties", "ZA"),
    "Chile":                                        ("Counties", "CL"),
    "Cyprus":                                       ("Counties", "CY"),
    "Saudi Arabia":                                 ("Counties", "SA"),
    "Sierra Leone":                                 ("Counties", "SL"),
    "Tanzania":                                     ("Counties", "TZ"),
    "Azerbaijan":                                   ("Counties", "AZ"),
    "Belarus":                                      ("Counties", "BY"),
    "Mozambique":                                   ("Counties", "MZ"),
    "Paraguay":                                     ("Counties", "PY"),
    "Myanmar":                                      ("Counties", "MM"),
    "Thailand":                                     ("Counties", "TH"),
    "Seychelles":                                   ("Counties", "SC"),
    "United Arab Emirates":                         ("Counties", "AE"),
    "Albania":                                      ("Counties", "AL"),
    "Estonia":                                      ("Counties", "EE"),
    "Latvia":                                       ("Counties", "LV"),
    "Morocco":                                      ("Counties", "MA"),
    "Philippines":                                  ("Counties", "PH"),
    "New Zealand":                                  ("Counties", "NZ"),
    "Pakistan":                                     ("Counties", "PK"),
    "Afghanistan":                                  ("Counties", "AF"),
    "Kosovo":                                       ("Counties", "XK"),
    "Georgia":                                      ("Counties", "GE"),
    "Guam":                                         ("Counties", "GU"),
    "Hong Kong":                                    ("Counties", "HK"),
    "Aruba":                                        ("Counties", "AW"),
    "Bangladesh":                                   ("Counties", "BD"),
    "Montenegro":                                   ("Counties", "ME"),
    "Armenia":                                      ("Counties", "AM"),
    "Mayotte":                                      ("Counties", "YT"),
    "Syria":                                        ("Counties", "SY"),
    "Faroe Islands":                                ("Counties", "FO"),
    "Iceland":                                      ("Counties", "IS"),
    "Madagascar":                                   ("Counties", "MG"),
    "Namibia":                                      ("Counties", "NA"),
    "Nauru":                                        ("Counties", "NR"),
    "New Caledonia":                                ("Counties", "NC"),
    "Fiji":                                         ("Counties", "FJ"),
    "Laos":                                         ("Counties", "LA"),
    "Kyrgyzstan":                                   ("Counties", "KG"),
    "Nepal":                                        ("Counties", "NP"),
    "Palau":                                        ("Counties", "PW"),
    "Tajikistan":                                   ("Counties", "TJ"),
    "Trinidad & Tobago":                            ("Counties", "TT"),
    "Lebanon":                                      ("Counties", "LB"),
    "Oman":                                         ("Counties", "OM"),
    "Senegal":                                      ("Counties", "SN"),
    "Sri Lanka":                                    ("Counties", "LK"),
    "Uruguay":                                      ("Counties", "UY"),
    "Bermuda":                                      ("Counties", "BM"),
    "Bhutan":                                       ("Counties", "BT"),
    "Cape Verde":                                   ("Counties", "CV"),
    "Mali":                                         ("Counties", "ML"),
    "Mongolia":                                     ("Counties", "MN"),
    "Saint Martin":                                 ("Counties", "MF"),
    "Palestine":                                    ("Counties", "PS"),
    "Turkmenistan":                                 ("Counties", "TM"),
    "Botswana":                                     ("Counties", "BW"),
    "Gibraltar":                                    ("Counties", "GI"),
    "Jordan":                                       ("Counties", "JO"),
    "Macau":                                        ("Counties", "MO"),
    "Tonga":                                        ("Counties", "TO"),
    "Jamaica":                                      ("Counties", "JM"),
    "Libya":                                        ("Counties", "LY"),
    "Niger":                                        ("Counties", "NE"),
    "Papua New Guinea":                             ("Counties", "PG"),
    "Bahamas":                                      ("Counties", "BS"),
    "Cayman Islands":                               ("Counties", "KY"),
    "Rwanda":                                       ("Counties", "RW"),
    "St. Vincent & the Grenadines":                 ("Counties", "VC"),
    "Bahrain":                                      ("Counties", "BH"),
    "Barbados":                                     ("Counties", "BB"),
    "Comoros":                                      ("Counties", "KM"),
    "Congo":                                        ("Counties", "CG"),
    "Mauritania":                                   ("Counties", "MR"),
    "Mauritius":                                    ("Counties", "MU"),
    "São Tomé and Príncipe":                        ("Counties", "ST"),
    "Côte d'Ivoire":                                ("Counties", "CI"),
    "Kiribati":                                     ("Counties", "KI"),
    "Montserrat":                                   ("Counties", "MS"),
    "Saint Barthélemy":                             ("Counties", "BL"),
    "Belize":                                       ("Counties", "BZ"),
    "Equatorial Guinea":                            ("Counties", "GQ"),
    "Federated States of Micronesia":               ("Counties", "FM"),
    "Guinea-Bissau":                                ("Counties", "GW"),
    "Liberia":                                      ("Counties", "LR"),
    "Maldives":                                     ("Counties", "MV"),
    "Togo":                                         ("Counties", "TG"),
    "Turks & Caicos Islands":                       ("Counties", "TC"),
    "US Virgin Islands":                            ("Counties", "VI"),
    "Western Sahara":                               ("Counties", "EH"),
    "American Samoa":                               ("Counties", "AS"),
    "Anguilla":                                     ("Counties", "AI"),
    "Antigua & Barbuda":                            ("Counties", "AG"),
    "Bonaire":                                      ("Counties", "BQ"),
    "British Virgin Islands":                       ("Counties", "VG"),
    "Chad":                                         ("Counties", "TD"),
    "Cook Islands":                                 ("Counties", "CK"),
    "Djibouti":                                     ("Counties", "DJ"),
    "Dominica":                                     ("Counties", "DM"),
    "East Timor":                                   ("Counties", "TL"),
    "Eswatini":                                     ("Counties", "SZ"),
    "Falkland Islands":                             ("Counties", "FK"),
    "Grenada":                                      ("Counties", "GD"),
    "Guyana":                                       ("Counties", "GY"),
    "Marshall Islands":                             ("Counties", "MH"),
    "Niue":                                         ("Counties", "NU"),
    "Northern Mariana Islands":                     ("Counties", "MP"),
    "Saint Helena, Ascension and Tristan da Cunha": ("Counties", "SH"),
    "Saint Lucia":                                  ("Counties", "LC"),
    "Sint Maarten":                                 ("Counties", "SX"),
    "Solomon Islands":                              ("Counties", "SB"),
    "Suriname":                                     ("Counties", "SR"),
    "Tokelau":                                      ("Counties", "TK"),
    "Tuvalu":                                       ("Counties", "TV"),
    # 2026-05-24 evening expansion (Bucket A tail):
    "Yemen":                                        ("Counties", "YE"),
    "Zambia":                                       ("Counties", "ZM"),
    "Zimbabwe":                                     ("Counties", "ZW"),
}

COUNTIES_COLS = {
    "country": 0, "name": 1, "state": 2, "metro": 7,
}
MUNI_COLS = {
    "country": 1, "name": 2, "alt_name": 3, "state": 4, "metro": 6,
}

ADMIN_SUFFIXES = (
    " county", " municipality", " city", " region", " province",
    " prefecture", " district", " oblast", " krai", " okrug",
    " raion", " rayon", " selsoviet", " selsovet", " selsovyet",
    " urban district", " urban okrug", " municipal okrug",
    " gemeinde", " commune", " comuna", " gmina", " kraj",
    " republic", " autonomous okrug", " autonomous region",
    " federal city", " federal district",
    " voivodeship", " voivodship", " wojewodztwo", " wojewodship",
    " state", " free state", " bundesland", " federation",
    " federal subject", " autonomous oblast",
    " autonomous community",
    " federal territory", " union territory",
    " department", " departement", " regional district",
    " administrative region", " metropolitan city",
    " regional unit", " regional council", " regional municipality",
    " county council", " borough", " unitary authority",
    " region of", " regione",
)
ADMIN_PREFIXES = ("county ", "city of ", "municipality of ", "region of ", "the ")

CONFIDENCE_HIGH = "HIGH"
CONFIDENCE_MED = "MED"
CONFIDENCE_LOW = "LOW"
CONFIDENCE_NONE = "NONE"

FUZZY_HIGH = 0.97
FUZZY_MED = 0.92
FUZZY_LOW = 0.78


def normalize(name):
    if name is None:
        return ""
    s = str(name).strip().lower()
    if not s:
        return ""
    if "(" in s:
        s = s.split("(", 1)[0].strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    changed = True
    while changed:
        changed = False
        for suf in ADMIN_SUFFIXES:
            if s.endswith(suf):
                s = s[: -len(suf)].strip()
                changed = True
        for pre in ADMIN_PREFIXES:
            if s.startswith(pre):
                s = s[len(pre):].strip()
                changed = True
    return s.strip(" .,'-")


def safe_int(v):
    if v is None:
        return None
    try:
        if v != v:
            return None
        return int(v)
    except Exception:
        return None


def load_parquet_full(country_iso, want_geometry=True):
    fname = f"overture-{country_iso}.parquet"
    path = Path(SOURCE_DIR) / fname
    if not path.exists():
        return None, f"parquet not found: {path}"
    cols = ["region", "subtype", "admin_level", "names", "country"]
    if want_geometry:
        cols.append("geometry")
    table = pq.read_table(str(path), columns=cols)
    df = table.to_pandas()
    df = df[df["country"] == country_iso].reset_index(drop=True)
    return df, None


def build_records(df):
    region_name_to_iso = {}
    by_region = defaultdict(list)
    counties_per_region = defaultdict(list)
    regions_per_iso = {}
    global_by_name = {}

    for row in df.itertuples():
        names = row.names
        if not isinstance(names, dict):
            continue
        primary = names.get("primary")
        if not primary:
            continue
        common = names.get("common")
        aliases = []
        if isinstance(common, list):
            for pair in common:
                if isinstance(pair, (list, tuple)) and len(pair) >= 2:
                    aliases.append(str(pair[1]))
        elif isinstance(common, dict):
            aliases = [str(v) for v in common.values()]
        norm_aliases = [normalize(a) for a in aliases if a]
        region = row.region or ""
        norm_primary = normalize(primary)
        admin_level = safe_int(row.admin_level)
        geom_data = getattr(row, "geometry", None)
        try:
            geom = shapely_wkb.loads(geom_data) if geom_data else None
        except Exception:
            geom = None
        rec = {
            "primary": primary,
            "normalized": norm_primary,
            "subtype": row.subtype,
            "admin_level": admin_level,
            "region_iso": region,
            "aliases": norm_aliases,
            "geom": geom,
        }
        if admin_level == 1 and row.subtype == "region":
            if norm_primary:
                region_name_to_iso[norm_primary] = region
            for a in norm_aliases:
                if a and a not in region_name_to_iso:
                    region_name_to_iso[a] = region
            if region:
                regions_per_iso[region] = rec
        if region:
            by_region[region].append(rec)
        # Global by-name index: normalized primary -> list of records.
        # Used for the nationwide fallback when workbook state value does
        # not resolve to a region (e.g., Ireland's historic provinces).
        if norm_primary:
            global_by_name.setdefault(norm_primary, []).append(rec)
        for a in norm_aliases:
            if a:
                global_by_name.setdefault(a, []).append(rec)
        if region:
            if row.subtype == "county" and rec["geom"] is not None:
                # Simplify the polygon for faster contains() during the
                # centroid containment test. Tolerance 0.001 deg (~100m)
                # keeps accuracy at metro zoom while dropping vertex counts
                # by ~10x on dense polygons (Russia raions, China counties).
                try:
                    simplified = rec["geom"].simplify(0.001, preserve_topology=True)
                    if simplified.is_valid and not simplified.is_empty:
                        rec["geom_simple"] = simplified
                    else:
                        rec["geom_simple"] = rec["geom"]
                except Exception:
                    rec["geom_simple"] = rec["geom"]
                counties_per_region[region].append(rec)

    county_trees = {}
    for region, recs in counties_per_region.items():
        geoms = [r.get("geom_simple", r["geom"]) for r in recs]
        if geoms:
            county_trees[region] = (STRtree(geoms), recs)
    return region_name_to_iso, by_region, county_trees, regions_per_iso, global_by_name


def resolve_region_iso(state_value, region_name_to_iso):
    if not state_value:
        return None
    s = str(state_value).strip()
    if "-" in s and len(s) <= 8 and s.split("-", 1)[0].isalpha() and s.split("-", 1)[0].isupper():
        return s
    return region_name_to_iso.get(normalize(s))


def find_by_name(query_norm, candidates):
    if not query_norm or not candidates:
        return None, 0.0, []
    exact = [c for c in candidates if c["normalized"] == query_norm]
    if exact:
        return exact[0], 1.0, [(c, 1.0) for c in exact[1:3]]
    alias = [c for c in candidates if query_norm in c["aliases"]]
    if alias:
        return alias[0], 1.0, [(c, 1.0) for c in alias[1:3]]
    scored = []
    for c in candidates:
        if not c["normalized"]:
            continue
        s = SequenceMatcher(None, query_norm, c["normalized"]).ratio()
        if c["normalized"].startswith(query_norm) or query_norm in c["normalized"]:
            s = min(1.0, s + 0.05)
        scored.append((c, s))
    scored.sort(key=lambda t: -t[1])
    if not scored or scored[0][1] < FUZZY_LOW:
        return None, scored[0][1] if scored else 0.0, [(c, sc) for c, sc in scored[:2]]
    return scored[0][0], scored[0][1], [(c, sc) for c, sc in scored[1:3]]


def find_containing_county(point, region_iso, county_trees):
    entry = county_trees.get(region_iso)
    if entry is None:
        return None
    tree, county_recs = entry
    candidate_idxs = tree.query(point)
    for idx in candidate_idxs:
        rec = county_recs[idx]
        try:
            geom = rec.get("geom_simple", rec["geom"])
            if geom.contains(point):
                return rec
        except Exception:
            continue
    return None


def confidence_for(score):
    if score >= FUZZY_HIGH:
        return CONFIDENCE_HIGH
    if score >= FUZZY_MED:
        return CONFIDENCE_MED
    if score >= FUZZY_LOW:
        return CONFIDENCE_LOW
    return CONFIDENCE_NONE


def match_row(row, sheet_kind, region_name_to_iso, by_region, county_trees, regions_per_iso, global_by_name):
    cols = MUNI_COLS if sheet_kind == "Municipality" else COUNTIES_COLS
    state_value = row[cols["state"]] if len(row) > cols["state"] else None
    region_iso = resolve_region_iso(state_value, region_name_to_iso)
    relevant_subtypes = {"locality", "localadmin", "county", "region"}
    if region_iso:
        full_candidates = by_region.get(region_iso, [])
        candidates = [c for c in full_candidates if c["subtype"] in relevant_subtypes]
        if not candidates:
            candidates = full_candidates
        used_global = False
    else:
        # State did not resolve to a region. Use the global name index
        # (O(1) hash lookup keyed on normalized name) instead of scanning
        # every country row, otherwise the fuzzy fallback is too slow.
        candidates = []
        used_global = True

    queries = []
    name_value = row[cols["name"]] if len(row) > cols["name"] else None
    if name_value:
        queries.append(name_value)
    if sheet_kind == "Municipality":
        alt = row[cols["alt_name"]] if len(row) > cols["alt_name"] else None
        if alt and alt != name_value:
            queries.append(alt)

    direct = None
    direct_score = 0.0
    for q in queries:
        qn = normalize(q)
        if not qn:
            continue
        if used_global:
            hits = global_by_name.get(qn, [])
            hits = [c for c in hits if c["subtype"] in relevant_subtypes]
            if hits:
                # Exact-name hit at admin level 4 wins over higher levels.
                hits.sort(key=lambda c: (c.get("admin_level") is None, c.get("admin_level") or 99))
                rec, score = hits[0], 1.0
                if score > direct_score:
                    direct = rec; direct_score = score
            continue
        if not candidates:
            continue
        rec, score, _alts = find_by_name(qn, candidates)
        if rec is not None and score > direct_score:
            direct = rec
            direct_score = score

    parent_county = None
    parent_region = regions_per_iso.get(region_iso) if region_iso else None
    if direct is not None and direct.get("geom") is not None and region_iso:
        try:
            point = direct["geom"].representative_point()
            parent_county = find_containing_county(point, region_iso, county_trees)
        except Exception:
            parent_county = None

    if region_iso is None and direct is not None:
        region_iso = direct.get("region_iso")
        if region_iso:
            parent_region = regions_per_iso.get(region_iso)
    return {
        "region_iso": region_iso,
        "direct": direct,
        "direct_score": direct_score,
        "parent_county": parent_county,
        "parent_region": parent_region,
    }


def rec_tuple(rec):
    if rec is None:
        return ("", "", "", "")
    return (
        rec.get("subtype") or "",
        rec.get("admin_level") if rec.get("admin_level") is not None else "",
        rec.get("region_iso") or "",
        rec.get("primary") or "",
    )


FILLS = {
    CONFIDENCE_HIGH: PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    CONFIDENCE_MED:  PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    CONFIDENCE_LOW:  PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    CONFIDENCE_NONE: PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
}


def write_combined_output(collected, out_dir):
    """Write one consolidated combined.xlsx + combined.csv across every
    country processed in this run.

    Each entry in ``collected`` is a dict with keys ``country`` (workbook
    name), ``header`` (per-country output header), ``rows`` (list of
    (full_row, conf) tuples), and ``counts``.

    The combined file uses the FIRST country's header as the canonical
    schema, prepending a ``Source Country`` column. Per-country headers
    align when every country routes to the same sheet (Counties vs
    Municipality); a future run that mixes sheet routings may diverge in
    the columns after Confidence -- scope to one sheet at a time if so.
    """
    if not collected:
        print("\n=== Combined output: no rows to write ===")
        return
    out_dir.mkdir(parents=True, exist_ok=True)
    xlsx_path = out_dir / "combined.xlsx"
    csv_path = out_dir / "combined.csv"

    base_header = ["Source Country"] + list(collected[0]["header"])

    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "Suggestions"
    out_ws.append(base_header)
    bold = Font(bold=True)
    for cell in out_ws[1]:
        cell.font = bold

    grand_counts = {CONFIDENCE_HIGH: 0, CONFIDENCE_MED: 0,
                    CONFIDENCE_LOW: 0,  CONFIDENCE_NONE: 0}
    csv_rows = [base_header]
    total_rows = 0
    for entry in collected:
        country = entry["country"]
        for full_row, conf in entry["rows"]:
            tagged = [country] + list(full_row)
            out_ws.append(tagged)
            for cell in out_ws[out_ws.max_row]:
                cell.fill = FILLS[conf]
            grand_counts[conf] += 1
            csv_rows.append(tagged)
            total_rows += 1

    with open(csv_path, "w", encoding="utf-8-sig", newline="") as cf:
        w = csv.writer(cf)
        for r in csv_rows:
            w.writerow(["" if v is None else v for v in r])
    print(f"\n=== Combined output ===")
    print(f"  wrote {csv_path}  ({total_rows:,} rows across {len(collected)} countries)")
    try:
        out_wb.save(xlsx_path)
        print(f"  wrote {xlsx_path}")
    except PermissionError:
        print(f"  SKIP {xlsx_path}: file is open in Excel")
    total = sum(grand_counts.values())
    for conf in (CONFIDENCE_HIGH, CONFIDENCE_MED, CONFIDENCE_LOW, CONFIDENCE_NONE):
        pct = (100.0 * grand_counts[conf] / total) if total else 0
        print(f"    {conf:<5} {grand_counts[conf]:>6,} ({pct:5.1f}%)")


def process_country(country_name, sheet_kind, country_iso, wb, args):
    print(f"\n=== {country_name} ({country_iso}) -> {sheet_kind} sheet ===")
    t0 = time.time()
    # Counties-sheet routing means each workbook row is already at the
    # county tier - direct name match is enough, skip geometry containment.
    want_geom = (sheet_kind == "Municipality")
    df, err = load_parquet_full(country_iso, want_geometry=want_geom)
    if df is None:
        print(f"  SKIP: {err}")
        return
    print(f"  parquet rows: {len(df):,}  ({time.time()-t0:.1f}s)")
    region_name_to_iso, by_region, county_trees, regions_per_iso, global_by_name = build_records(df)
    print(f"  regions: {len(by_region)}, county-trees: {len(county_trees)}, "
          f"region-polys: {len(regions_per_iso)}")

    ws = wb[sheet_kind]
    cols = MUNI_COLS if sheet_kind == "Municipality" else COUNTIES_COLS
    rows = []
    for raw_row in ws.iter_rows(min_row=2, values_only=True):
        if (raw_row[cols["country"]] or "") != country_name:
            continue
        if args.get("metro_only") and not raw_row[cols["metro"]]:
            continue
        rows.append(list(raw_row))
    print(f"  workbook rows: {len(rows):,}")
    if not rows:
        return

    src_header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    src_ncols = len(src_header)
    extra_header = [
        "Direct Subtype", "Direct Admin", "Direct Region ISO", "Direct Primary",
        "County Subtype", "County Admin", "County Region ISO", "County Primary",
        "Region Subtype", "Region Admin", "Region Region ISO", "Region Primary",
        "Confidence", "Score", "Resolved Region ISO",
    ]
    full_header = list(src_header) + extra_header

    counts = {CONFIDENCE_HIGH: 0, CONFIDENCE_MED: 0,
              CONFIDENCE_LOW: 0,  CONFIDENCE_NONE: 0}
    processed_rows = []
    for raw_row in rows:
        res = match_row(raw_row, sheet_kind, region_name_to_iso, by_region, county_trees, regions_per_iso, global_by_name)
        direct_t = rec_tuple(res["direct"])
        county_t = rec_tuple(res["parent_county"])
        region_t = rec_tuple(res["parent_region"])
        conf = confidence_for(res["direct_score"]) if res["direct"] is not None else CONFIDENCE_NONE
        extra = list(direct_t) + list(county_t) + list(region_t) + [
            conf, round(res["direct_score"], 3), res["region_iso"] or "",
        ]
        padded = list(raw_row[:src_ncols]) + [None] * max(0, src_ncols - len(raw_row[:src_ncols]))
        full_row = padded + extra
        processed_rows.append((full_row, conf))
        counts[conf] += 1

    collector = args.get("collector")
    if collector is not None:
        # Combined-output mode: defer file writing to write_combined_output
        # in main(). Each entry carries the country tag so the combined
        # file can prepend a Source Country column.
        collector.append({
            "country": country_name,
            "header":  full_header,
            "rows":    processed_rows,
            "counts":  counts,
        })
        elapsed = time.time() - t0
        print(f"  collected {len(processed_rows):,} rows ({elapsed:.1f}s)")
        total = sum(counts.values())
        for conf in (CONFIDENCE_HIGH, CONFIDENCE_MED, CONFIDENCE_LOW, CONFIDENCE_NONE):
            pct = (100.0 * counts[conf] / total) if total else 0
            print(f"    {conf:<5} {counts[conf]:>6,} ({pct:5.1f}%)")
        return

    # Per-country file mode (original behavior preserved).
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    xlsx_path = OUT_DIR / f"{country_iso}.xlsx"
    csv_path = OUT_DIR / f"{country_iso}.csv"

    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "Suggestions"
    out_ws.append(full_header)
    bold = Font(bold=True)
    for cell in out_ws[1]:
        cell.font = bold

    csv_rows = [full_header]
    for full_row, conf in processed_rows:
        out_ws.append(full_row)
        for cell in out_ws[out_ws.max_row]:
            cell.fill = FILLS[conf]
        csv_rows.append(full_row)

    # Write CSV first (cheap, never locked).
    with open(csv_path, "w", encoding="utf-8-sig", newline="") as cf:
        w = csv.writer(cf)
        for r in csv_rows:
            w.writerow(["" if v is None else v for v in r])
    print(f"  wrote {csv_path}")
    # xlsx may be locked if the user has it open in Excel; skip gracefully.
    try:
        out_wb.save(xlsx_path)
        print(f"  wrote {xlsx_path}  ({time.time()-t0:.1f}s total)")
    except PermissionError:
        print(f"  SKIP {xlsx_path}: file is open in Excel ({time.time()-t0:.1f}s total)")
    total = sum(counts.values())
    for conf in (CONFIDENCE_HIGH, CONFIDENCE_MED, CONFIDENCE_LOW, CONFIDENCE_NONE):
        pct = (100.0 * counts[conf] / total) if total else 0
        print(f"    {conf:<5} {counts[conf]:>6,} ({pct:5.1f}%)")


def main():
    args = {"metro_only": False, "collector": None}
    countries = []
    combined = False
    for tok in sys.argv[1:]:
        if tok == "--metro-only":
            args["metro_only"] = True
        elif tok == "--combined":
            combined = True
        elif tok in {iso for (_, iso) in COUNTRY_ROUTE.values()}:
            countries.append(tok)
        elif tok in COUNTRY_ROUTE:
            countries.append(COUNTRY_ROUTE[tok][1])
    if not countries:
        countries = [iso for (_, iso) in COUNTRY_ROUTE.values()]

    if combined:
        args["collector"] = []

    print(f"Workbook: {WORKBOOK}")
    print(f"Source dir: {SOURCE_DIR}")
    print(f"Output dir: {OUT_DIR}")
    print(f"Metro-only: {args['metro_only']}")
    print(f"Combined output: {combined}")
    print(f"Countries: {countries}")

    wb = openpyxl.load_workbook(WORKBOOK, read_only=True, data_only=True)
    iso_to_country = {iso: name for name, (_, iso) in COUNTRY_ROUTE.items()}
    for iso in countries:
        country_name = iso_to_country.get(iso)
        if country_name is None:
            print(f"  SKIP unknown ISO: {iso}")
            continue
        sheet, _ = COUNTRY_ROUTE[country_name]
        process_country(country_name, sheet, iso, wb, args)

    if combined and args["collector"]:
        write_combined_output(args["collector"], OUT_DIR)


if __name__ == "__main__":
    main()

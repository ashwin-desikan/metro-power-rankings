"""Per-country Overture division_area dumps for hand validation.

Reads the global Overture parquet (5.8 GB) once, partitions rows by country
on the fly, and writes one xlsx per country in the Overture-Per-Country-Raw/
folder. Each output file mirrors the schema of the existing NA raw dumps:

  Sheet: 'Overture'
  Cols : GERS ID, Country (ISO), Region (ISO 3166-2), Subtype, Admin Level,
         Class, Primary Name, Common Names (JSON), Is Land, Is Territorial,
         BBox xmin/ymin/xmax/ymax, Division ID

Usage:
  cd "C:\\Users\\ashwi\\Desktop\\Projects\\Metro Area Project"
  python scripts\\dump-overture-country.py

Override defaults via env:
  OVERTURE_DIVISION_AREA  - path to the global parquet
  OVERTURE_DUMP_DIR       - output folder (default: Overture-Per-Country-Raw)
  OVERTURE_DUMP_COUNTRIES - comma-sep ISO list (default: GB,DE,FR,IT,ES,NL,PL,JP)

Memory profile: scans the parquet via pyarrow batches and accumulates per-
country dict-of-lists. Peak RSS depends on the LARGEST country's row count;
for the eight defaults this stays under ~1 GB.
"""
from __future__ import annotations

import json
import os
import sys
import time
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


SOURCE_PARQUET = os.environ.get(
    "OVERTURE_DIVISION_AREA",
    r"C:\Users\ashwi\Desktop\Projects\MapData\global-division-area.parquet",
)
OUT_DIR = Path(os.environ.get(
    "OVERTURE_DUMP_DIR",
    "Overture-Per-Country-Raw",
))
DEFAULT_COUNTRIES = ("GB", "DE", "FR", "IT", "ES", "NL", "PL", "JP")
COUNTRIES = tuple(
    s.strip().upper()
    for s in os.environ.get(
        "OVERTURE_DUMP_COUNTRIES",
        ",".join(DEFAULT_COUNTRIES),
    ).split(",")
    if s.strip()
)

HEADER = [
    "GERS ID",
    "Country (ISO)",
    "Region (ISO 3166-2)",
    "Subtype",
    "Admin Level",
    "Class",
    "Primary Name",
    "Common Names (JSON)",
    "Is Land",
    "Is Territorial",
    "BBox xmin",
    "BBox ymin",
    "BBox xmax",
    "BBox ymax",
    "Division ID",
]


def _common_names_json(names_dict):
    """Pack the Overture `names` struct's `common` map into a stable JSON
    string identical in shape to the existing NA raw files: a JSON list of
    [lang, label] pairs, sorted by lang.
    """
    if not isinstance(names_dict, dict):
        return None
    common = names_dict.get("common")
    if not common:
        return None
    if isinstance(common, dict):
        items = sorted(common.items(), key=lambda kv: kv[0])
        return json.dumps([[k, v] for k, v in items], ensure_ascii=False)
    if isinstance(common, list):
        # already in pair form
        try:
            items = sorted(common, key=lambda kv: kv[0])
            return json.dumps([[k, v] for k, v in items], ensure_ascii=False)
        except Exception:
            return json.dumps(common, ensure_ascii=False)
    return None


def _bbox_to_xyxy(bbox):
    """Overture stores bbox as a struct {xmin, ymin, xmax, ymax}. Return as
    a 4-tuple (xmin, ymin, xmax, ymax) suitable for direct cell write."""
    if not isinstance(bbox, dict):
        return (None, None, None, None)
    return (
        bbox.get("xmin"),
        bbox.get("ymin"),
        bbox.get("xmax"),
        bbox.get("ymax"),
    )


def _bool_str(v):
    """Match the existing NA dump format: 'TRUE'/'FALSE' upper-case strings,
    or None if the source value is missing."""
    if v is None:
        return None
    return "TRUE" if bool(v) else "FALSE"


def collect_per_country(parquet_path: str, countries: tuple):
    """Single pass over the parquet, returning {iso: list[row_tuple]}."""
    print(f"[1/2] Scanning {parquet_path}")
    print(f"      countries: {countries}")
    t0 = time.time()
    import pyarrow.parquet as pq

    pf = pq.ParquetFile(parquet_path)
    cols = [
        "id",
        "country",
        "region",
        "subtype",
        "names",
        "class",
        "is_land",
        "is_territorial",
        "bbox",
        "division_id",
    ]
    # Some Overture releases use "primary_division_id" instead. Detect.
    schema_cols = set(pf.schema_arrow.names)
    if "division_id" not in schema_cols and "primary_division_id" in schema_cols:
        cols[cols.index("division_id")] = "primary_division_id"
    # Some releases also expose admin_level explicitly; if not, derive from subtype.
    has_admin_level = "admin_level" in schema_cols
    if has_admin_level:
        cols.append("admin_level")
    # All Overture release schemas carry these top-level cols. If your release
    # differs, edit `cols` and the row-build below in lock-step.

    wanted = set(countries)
    bucket = defaultdict(list)
    rows_scanned = 0
    rows_kept = 0
    for batch in pf.iter_batches(batch_size=10_000, columns=cols):
        ids = batch.column("id").to_pylist()
        ctry = batch.column("country").to_pylist()
        regs = batch.column("region").to_pylist()
        subs = batch.column("subtype").to_pylist()
        nms = batch.column("names").to_pylist()
        cls = batch.column("class").to_pylist()
        land = batch.column("is_land").to_pylist()
        terr = batch.column("is_territorial").to_pylist()
        bbox = batch.column("bbox").to_pylist()
        divid_field = "division_id" if "division_id" in cols else "primary_division_id"
        divid = batch.column(divid_field).to_pylist()
        admin_level_col = (batch.column("admin_level").to_pylist()
                           if has_admin_level else None)

        for i in range(len(ids)):
            rows_scanned += 1
            c = ctry[i]
            if c not in wanted:
                continue
            nm = nms[i] if isinstance(nms[i], dict) else {}
            primary_name = nm.get("primary") if isinstance(nm, dict) else None
            common_json = _common_names_json(nm)
            xmin, ymin, xmax, ymax = _bbox_to_xyxy(bbox[i])
            sub = subs[i]
            if has_admin_level:
                admin_lv = admin_level_col[i]
            else:
                # Overture's admin-level convention: country=0, region=1,
                # county=2, locality=3 (rough). Use a small map; if subtype
                # not in map, leave None.
                admin_lv = {
                    "country": 0,
                    "region": 1,
                    "county": 2,
                    "locality": 3,
                    "neighborhood": 4,
                }.get(sub)
            row = (
                ids[i],
                c,
                regs[i],
                sub,
                admin_lv,
                cls[i],
                primary_name,
                common_json,
                _bool_str(land[i]),
                _bool_str(terr[i]),
                xmin, ymin, xmax, ymax,
                divid[i],
            )
            bucket[c].append(row)
            rows_kept += 1

    print(f"      scanned {rows_scanned:,} rows, kept {rows_kept:,} matching countries "
          f"in {time.time()-t0:.1f}s")
    for iso in countries:
        print(f"      {iso}: {len(bucket[iso]):,} rows")
    return bucket


def write_country_xlsx(out_dir: Path, iso: str, rows: list):
    """Write one xlsx file in the same shape as the NA raw dumps."""
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"overture-{iso}-full.xlsx"

    wb = Workbook(write_only=False)
    ws = wb.active
    ws.title = "Overture"

    # Header row
    ws.append(HEADER)
    for r in rows:
        ws.append(list(r))

    # Freeze the header and add an autofilter
    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(HEADER))
    ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"

    # Reasonable column widths (the NA raw files use sensible defaults but
    # nothing exotic; keep things readable).
    widths = [38, 14, 22, 16, 12, 10, 38, 60, 10, 14, 14, 14, 14, 14, 38]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(out_path)
    print(f"      wrote {out_path}  ({ws.max_row - 1:,} data rows)")


def main():
    print(f"Per-country Overture xlsx generator")
    print(f"Output dir: {OUT_DIR.resolve()}")
    print(f"Source parquet: {SOURCE_PARQUET}")
    bucket = collect_per_country(SOURCE_PARQUET, COUNTRIES)
    print(f"[2/2] Writing per-country xlsx files")
    for iso in COUNTRIES:
        rows = bucket.get(iso, [])
        if not rows:
            print(f"      WARNING: no rows for {iso}, skipping")
            continue
        write_country_xlsx(OUT_DIR, iso, rows)
    print("Done.")


if __name__ == "__main__":
    main()

#!/usr/bin/env python3
"""Build the WNBA hub data from the WNBA sheet of OtherLeagues.xlsx.

The WNBA sheet has the header on row 2 and season-by-season rows with a
Canonical Name column (folds relocations/renames into the current franchise
identity, e.g. Utah Starzz -> Las Vegas Aces). Two side lists (cols Y and AA)
classify each franchise as Current or Defunct. Emits public/data/wnba/data.json
with franchises (current + defunct), a flat season list, and champions by year.

Usage:
  python scripts/build-wnba-data.py [SOURCE_XLSX]
Defaults to OtherLeagues.xlsx at the repo root.
"""
import json, os, re, sys, tempfile, unicodedata
from collections import defaultdict

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = sys.argv[1] if len(sys.argv) > 1 else os.path.join(ROOT, "OtherLeagues.xlsx")
OUT = os.path.join(ROOT, "public", "data", "wnba", "data.json")
METROS = os.path.join(ROOT, "public", "data", "metros.json")

import openpyxl

# Curated franchise metadata keyed by canonical name.
# (abbr, city, state, metro_slug, color)
FR = {
    "Atlanta Dream":          ("ATL", "Atlanta", "Georgia", "atlanta", "#C8102E"),
    "Chicago Sky":            ("CHI", "Chicago", "Illinois", "chicago", "#418FDE"),
    "Connecticut Sun":        ("CONN", "Uncasville", "Connecticut", "new-london", "#F05023"),
    "Indiana Fever":          ("IND", "Indianapolis", "Indiana", "indianapolis", "#E03A3E"),
    "New York Liberty":       ("NYL", "New York", "New York", "new-york", "#6ECEB2"),
    "Washington Mystics":     ("WAS", "Washington", "D.C.", "washington-baltimore", "#0C2340"),
    "Dallas Wings":           ("DAL", "Arlington", "Texas", "dallas", "#0C2340"),
    "Golden State Valkyries": ("GSV", "San Francisco", "California", "san-francisco-san-jose", "#6F2DA8"),
    "Las Vegas Aces":         ("LVA", "Las Vegas", "Nevada", "las-vegas", "#C8102E"),
    "Los Angeles Sparks":     ("LAS", "Los Angeles", "California", "los-angeles", "#552583"),
    "Minnesota Lynx":         ("MIN", "Minneapolis", "Minnesota", "minneapolis", "#236192"),
    "Phoenix Mercury":        ("PHX", "Phoenix", "Arizona", "phoenix", "#E56020"),
    "Portland Fire":          ("POR", "Portland", "Oregon", "portland", "#C8102E"),
    "Seattle Storm":          ("SEA", "Seattle", "Washington", "seattle", "#2C5234"),
    "Toronto Tempo":          ("TOR", "Toronto", "Ontario", "toronto", "#B81237"),
    "Sacramento Monarchs":    ("SAC", "Sacramento", "California", "sacramento", "#5C2D91"),
    "Houston Comets":         ("HOU", "Houston", "Texas", "houston", "#BA0C2F"),
    "Charlotte Sting":        ("CHA", "Charlotte", "North Carolina", "charlotte", "#00778B"),
    "Cleveland Rockers":      ("CLE", "Cleveland", "Ohio", "cleveland", "#E35205"),
    "Miami Sol":              ("MIA", "Miami", "Florida", "miami", "#F47920"),
}

def slugify(s):
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c)).lower()
    return re.sub(r"[^a-z0-9]+", "-", s).strip("-")

def valid_metros():
    try:
        d = json.load(open(METROS, encoding="utf-8"))
    except Exception:
        return None  # metros.json unreadable (e.g. mount truncation); skip validation
    rows = d if isinstance(d, list) else d.get("metros", d)
    return {r.get("slug") for r in rows if isinstance(r, dict)}

def wbasketball_metros():
    """Team name -> metro_slug for W Basketball, from the workbook-derived
    all-teams.json. The WNBA team metros live in Team List of MetroAreas.xlsx;
    all-teams.json is the slug-ready derived form. Falls back silently (curated
    FR metros) if the file is unreadable."""
    path = os.path.join(ROOT, "public", "data", "sports", "all-teams.json")
    try:
        d = json.load(open(path, encoding="utf-8"))
    except Exception:
        return {}
    rows = d if isinstance(d, list) else d.get("teams", d)
    out = {}
    for r in rows:
        if r.get("sport") == "W Basketball" and r.get("metro_slug"):
            out[r.get("team")] = r.get("metro_slug")
    return out

def num(v):
    return v if isinstance(v, (int, float)) else None

def main():
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    if "WNBA" not in wb.sheetnames:
        sys.exit(f"WNBA sheet not found in {SRC}")
    ws = wb["WNBA"]
    rows = list(ws.iter_rows(values_only=True))
    # row0 = labels, row1 = header, data from row2
    data = [r for r in rows[2:] if isinstance(r[0], int)]
    current = [r[24] for r in rows[1:] if r[24] and r[24] != "Current Teams"]
    defunct = [r[26] for r in rows[1:] if r[26] and r[26] != "Defunct Teams"]
    current_set, defunct_set = set(current), set(defunct)
    vmetros = valid_metros()
    at_metros = wbasketball_metros()

    def finish(r):
        if str(r[16]) == "Y": return ("Champion", 1)
        if str(r[15]) == "Y": return ("Runner-up", 2)
        if str(r[14]) == "Y": return ("Semifinals", 3)
        if str(r[9]) == "Y":  return ("Playoffs", 4)
        return ("Missed playoffs", 5)

    seasons = []
    agg = defaultdict(lambda: {"seasons": 0, "w": 0, "l": 0, "playoffs": 0, "div": 0,
                               "sf": 0, "finals": 0, "titles": 0, "title_years": [],
                               "final_years": [], "names": set(), "first": 9999, "last": 0})
    champions = {}
    for r in data:
        canon = r[17] or r[1]
        cslug = slugify(canon)
        lab, rank = finish(r)
        seasons.append({
            "year": r[0], "team": r[1], "canonical": canon, "slug": cslug,
            "conference": r[2], "w": num(r[3]) or 0, "l": num(r[4]) or 0,
            "win_pct": num(r[5]), "gb": r[6], "ps_g": num(r[7]), "pf_g": num(r[8]),
            "playoffs": str(r[9]) == "Y", "div_title": str(r[10]) == "Y",
            "p_wins": num(r[12]) or 0, "p_losses": num(r[13]) or 0,
            "sf_app": str(r[14]) == "Y", "finals_app": str(r[15]) == "Y",
            "champion": str(r[16]) == "Y", "finish": lab, "finish_rank": rank,
        })
        a = agg[canon]
        a["seasons"] += 1; a["w"] += num(r[3]) or 0; a["l"] += num(r[4]) or 0
        a["playoffs"] += 1 if str(r[9]) == "Y" else 0
        a["div"] += 1 if str(r[10]) == "Y" else 0
        a["sf"] += 1 if str(r[14]) == "Y" else 0
        a["finals"] += 1 if str(r[15]) == "Y" else 0
        if str(r[16]) == "Y":
            a["titles"] += 1; a["title_years"].append(r[0]); champions[r[0]] = (canon, cslug)
        if str(r[15]) == "Y":
            a["final_years"].append(r[0])
        a["names"].add(r[1]); a["first"] = min(a["first"], r[0]); a["last"] = max(a["last"], r[0])

    # union of all franchises: those with data + current/defunct lists
    all_names = set(agg) | current_set | defunct_set
    franchises = []
    for canon in all_names:
        a = agg.get(canon)
        abbr, city, state, metro, color = FR.get(canon, (slugify(canon)[:3].upper(), None, None, None, "#666666"))
        metro = at_metros.get(canon, metro)
        if metro and vmetros is not None and metro not in vmetros:
            metro = None
        w = a["w"] if a else 0
        l = a["l"] if a else 0
        aka = sorted((a["names"] - {canon})) if a else []
        franchises.append({
            "slug": slugify(canon), "name": canon, "abbr": abbr, "city": city, "state": state,
            "metro_slug": metro, "color": color,
            "active": canon in current_set or (canon not in defunct_set and bool(a)),
            "defunct": canon in defunct_set,
            "seasons": a["seasons"] if a else 0,
            "w": w, "l": l, "win_pct": round(w / (w + l), 3) if (w + l) else None,
            "playoff_appearances": a["playoffs"] if a else 0,
            "division_titles": a["div"] if a else 0,
            "semifinals": a["sf"] if a else 0,
            "finals": a["finals"] if a else 0,
            "titles": a["titles"] if a else 0,
            "title_years": sorted(a["title_years"]) if a else [],
            "final_years": sorted(a["final_years"]) if a else [],
            "last_title": max(a["title_years"]) if (a and a["title_years"]) else None,
            "first_season": a["first"] if (a and a["seasons"]) else None,
            "last_season": a["last"] if (a and a["seasons"]) else None,
            "aka": aka,
        })
    franchises.sort(key=lambda f: (-f["titles"], -f["finals"], -f["playoff_appearances"], f["name"]))

    champ_list = [{"year": y, "champion": champions[y][0], "champion_slug": champions[y][1]}
                  for y in sorted(champions, reverse=True)]
    years = [s["year"] for s in seasons]
    meta = {"league": "WNBA", "sport": "Basketball",
            "founded": min(years), "latest_season": max(years),
            "total_seasons": len(set(years)),
            "active_count": sum(1 for f in franchises if not f["defunct"]),
            "defunct_count": sum(1 for f in franchises if f["defunct"])}

    payload = {"meta": meta, "franchises": franchises, "seasons": seasons, "champions": champ_list}
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    fd, tmp = tempfile.mkstemp(dir=os.path.dirname(OUT), prefix=".wnba-", suffix=".tmp")
    with os.fdopen(fd, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, separators=(",", ":"))
    os.replace(tmp, OUT)
    print(f"wrote {OUT}: {len(franchises)} franchises "
          f"({meta['active_count']} current, {meta['defunct_count']} defunct), "
          f"{len(seasons)} season-rows, {meta['founded']}-{meta['latest_season']}")
    print("latest champion:", champ_list[0]["champion"], champ_list[0]["year"])

if __name__ == "__main__":
    main()

"""Build public/data/champions-history.json from Champions_History.xlsx.

Source of truth: ~/OneDrive/Excel Files/Champions_History.xlsx (the all-time
champions ledger; sheet "Champions", 16 cols incl. era name, canonical team,
era-correct metro + metro slug, and Date in YYYY-MM-DD where known).

Emits one record per champion row (raw passthrough + a stable competition
slug). Link resolution (team page, league hub) happens at render time in
lib/championsHistory.ts, reusing lib/championsHub. Run on Windows; the workbook
is cloud-only and unreadable from the sandbox.

    python scripts/build-champions-history.py
"""
import json, os, re, sys, datetime, unicodedata
from openpyxl import load_workbook

SRC = os.environ.get("CHAMPS_SRC", os.path.expanduser("~/OneDrive/Excel Files/Champions_History.xlsx"))
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.environ.get("CHAMPS_OUT", os.path.join(ROOT, "public", "data", "champions-history.json"))
METROS = os.path.join(ROOT, "public", "data", "metros.json")

def _normname(x):
    import unicodedata
    x = unicodedata.normalize("NFKD", str(x))
    x = "".join(c for c in x if not unicodedata.combining(c))
    return x.lower().strip()

def _metro_lookup():
    """slug set + unique normalized-name -> slug, from metros.json, so a
    champion's era-correct metro resolves to a real metro page even when the
    workbook's Metro Slug column mangled accents (Beziers, Saint-Etienne, ...)."""
    with open(METROS, encoding="utf-8") as f:
        data = json.load(f)
    arr = data if isinstance(data, list) else data.get("metros", data)
    slugset = {m.get("slug") for m in arr}
    byname = {}
    # slug -> display name, so a METRO_OVERRIDE can correct the NAME as well as
    # the slug. See the 🔴 note on METRO_OVERRIDE for why that matters.
    bysslug = {m.get("slug"): m.get("name") for m in arr if m.get("slug")}
    for m in arr:
        byname.setdefault(_normname(m.get("name") or ""), set()).add(m.get("slug"))
    return slugset, byname, bysslug

# Confirmed metro corrections the workbook still carries wrong (audited
# 2026-06-26, user-confirmed). Keyed (competition, year, canonical) -> slug.
# Durable: the build corrects these cells even if the workbook lags.
# Canonical-name unifications: franchise lineages the workbook still carries
# under two names. Applied to the canonical column so the club is one entity
# everywhere (one honour-roll entry, one defunct card). Era (champion) name is
# left untouched for era-correct display.
CANONICAL_ALIAS = {
    "London Wasps": "Wasps",
    "Oklahoma A&M": "Oklahoma State",
}

# Metro NAME variants the workbook uses that differ from metros.json names,
# so _fix_slug can still resolve them (normalized name -> slug).
METRO_NAME_ALIAS = {
    "rotterdam": "rotterdam-the-hague",
}

# 🔴 AN OVERRIDE CORRECTS THE NAME AND THE SLUG, NOT JUST THE SLUG. It used to
# fix `metroSlug` only, and `metro` kept whatever the workbook said, so every
# row in this table shipped with a corrected link and a wrong label: the 1992
# and 1993 World Series read "Buffalo", the 2019 NBA title read "Tampa" (both
# the Blue Jays' and the Raptors' COVID-era temporary venues, applied to titles
# won decades before), and Super Bowl XVIII read "San Francisco-San Jose" for a
# team that was the Los Angeles Raiders that season.
#
# It was not cosmetic, because consumers key on either field and get different
# answers. build_heartbreak.py's parade_drought() keys on the NAME, so Toronto
# was told it had not held a parade since the 1967 Stanley Cup while Buffalo
# was credited with two World Series and Tampa with an NBA title. Ashwin found
# it on the live board, 2026-08-14.
#
# The name is now resolved FROM the corrected slug against metros.json, so the
# two cannot disagree again by construction rather than by anyone remembering.
METRO_OVERRIDE = {
    ("MLB", 1992, "Toronto Blue Jays"): "toronto",
    ("MLB", 1993, "Toronto Blue Jays"): "toronto",
    ("NBA", 2019, "Toronto Raptors"): "toronto",
    ("NFL", 1983, "Las Vegas Raiders"): "los-angeles",
    ("MLB", 1896, "Baltimore Orioles"): "washington-baltimore",
    ("MLB", 1897, "Baltimore Orioles"): "washington-baltimore",
}

def _fix_slug(slugset, byname, slug, name):
    if slug and slug in slugset:
        return slug
    if name:
        alias = METRO_NAME_ALIAS.get(_normname(name))
        if alias:
            return alias
    cand = byname.get(_normname(name)) if name else None
    if cand and len(cand) == 1:
        return next(iter(cand))
    return slug  # leave as-is (flagged downstream / metro not yet in metros.json)

def _norm_date(d):
    """Normalize a Date cell to YYYY-MM-DD. Handles Excel datetimes, ISO strings,
    and American m/d/Y strings (the FA Cup early finals). Unparseable values are
    returned unchanged so they surface rather than silently vanish."""
    if d in (None, ""):
        return ""
    if hasattr(d, "strftime"):
        return d.strftime("%Y-%m-%d")
    s = str(d).strip()
    if not s:
        return ""
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", s):
        return s
    for fmt in ("%Y-%m-%d %H:%M:%S", "%m/%d/%Y", "%m/%d/%y", "%Y/%m/%d", "%m-%d-%Y", "%d/%m/%Y"):
        try:
            return datetime.datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return s

def slugify(s: str) -> str:
    # Fold accents FIRST. Without this, "Argentina Primera División" slugs as
    # "argentina-primera-divisi-n" (the accented char is not [a-z0-9], so it
    # becomes a separator) instead of "argentina-primera-division". That is a
    # live-URL change and it also breaks the competition redirect map, which
    # check:slug-drift gates on. Same for Brasileiro Série A and Copa América.
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.lower().replace("&", "and")
    s = re.sub(r"[‘’']", "", s)
    s = re.sub(r"[^a-z0-9]+", "-", s).strip("-")
    return s

def cell(v):
    return "" if v is None else (v if isinstance(v, (int, float)) else str(v).strip())

def main():
    wb = load_workbook(SRC, read_only=True, data_only=True)
    ws = wb["Champions"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(c).strip() if c is not None else "" for c in rows[0]]
    REQUIRED = ["Sport", "Competition", "Era Name", "Season", "Year", "Champion",
                "Champion (Canonical)", "Metro", "Metro Slug", "Date", "Scope Type"]
    OPTIONAL = ["Scope", "Tier", "TierGuide", "Is Current", "Date Awarded", "Next Awarded Date"]
    ix = {n: hdr.index(n) for n in REQUIRED}
    for n in OPTIONAL:
        if n in hdr:
            ix[n] = hdr.index(n)
    # The "next title" date -- when the current holder's crown is next
    # contested -- is column V, "Next Awarded Date", and is read by name above
    # like every other optional column.
    #
    # It used to be mapped BY POSITION as "the column right after Is Current,
    # if that column is header-less". That broke silently on 2026-08-11 when
    # the date-backfill session added "Date Source" / "Date Method" / "Date
    # Precision" at S/T/U: column S stopped being header-less, the positional
    # test failed, and every row shipped nextAwardedDate: null, so the "Next
    # title" column on /sports/champions went blank sitewide with no error
    # anywhere. The 92 dates were migrated out to their own named column on
    # 2026-08-12. Do not reintroduce a positional fallback -- if the column is
    # missing the guard below is meant to be loud.
    if "Next Awarded Date" not in ix:
        raise SystemExit(
            "Champions sheet has no 'Next Awarded Date' column. It carries the "
            "next-title date for the 92 current champions and feeds the 'Next "
            "title' column on /sports/champions. Restore the header rather "
            "than mapping it by position."
        )
    slugset, byname, bysslug = _metro_lookup()
    out = []
    for r in rows[1:]:
        comp = cell(r[ix["Competition"]])
        if not comp:
            continue
        try:
            yr = int(float(r[ix["Year"]]))
        except (TypeError, ValueError):
            yr = None
        date = _norm_date(r[ix["Date"]])
        def opt(name, default=None):
            i = ix.get(name)
            return cell(r[i]) if i is not None and i < len(r) else default
        tier_raw = opt("Tier")
        try:
            tier = int(float(tier_raw)) if tier_raw not in (None, "") else None
        except (TypeError, ValueError):
            tier = None
        guide_raw = opt("TierGuide")
        try:
            tier_guide = float(guide_raw) if guide_raw not in (None, "") else None
        except (TypeError, ValueError):
            tier_guide = None
        canon = CANONICAL_ALIAS.get(cell(r[ix["Champion (Canonical)"]]), cell(r[ix["Champion (Canonical)"]]))
        ovr = METRO_OVERRIDE.get((comp, yr, cell(r[ix["Champion (Canonical)"]])))
        metro_slug = ovr or _fix_slug(slugset, byname, cell(r[ix["Metro Slug"]]), cell(r[ix["Metro"]]))
        # The corrected slug wins the LABEL too. Falls back to the workbook's
        # own name if the slug is not in metros.json, so an override for a
        # metro the site does not yet carry degrades to today's behaviour
        # rather than blanking the column.
        metro_name = (bysslug.get(ovr) if ovr else None) or cell(r[ix["Metro"]])
        out.append({
            "sport": cell(r[ix["Sport"]]),
            "competition": comp,
            "compSlug": slugify(comp),
            "eraName": cell(r[ix["Era Name"]]),
            "season": cell(r[ix["Season"]]),
            "year": yr,
            "champion": cell(r[ix["Champion"]]),
            "canonical": canon,
            "metro": metro_name,
            "metroSlug": metro_slug,
            "date": date,
            "scope": opt("Scope"),
            "scopeType": cell(r[ix["Scope Type"]]),
            "tier":            tier,
            "tierGuide":       tier_guide,
            "isCurrent":       opt("Is Current") == "Y",
            "dateAwarded":     _norm_date(opt("Date Awarded")) or date or None,  # fallback to Date col
            "nextAwardedDate": _norm_date(opt("Next Awarded Date")) or None,
        })
    # 2026-08-08: the workbook no longer writes this file directly. It now
    # feeds public.champions, and the JSON is regenerated FROM the table, so
    # every champion the site tracks lives in one place. The workbook still
    # wins for these rows (push() is delete-then-insert, same contract as
    # sync_city_lookup.py), and the generator is proven byte-identical, so
    # this swap changes no site output.
    sys.path.insert(0, os.path.join(ROOT, "scripts", "champions"))
    from sync_history import push, emit
    pushed = push(out)
    print(f"champions: synced {pushed:,} rows to public.champions")
    emit()

    dated = sum(1 for x in out if x["date"])
    current = sum(1 for x in out if x["isCurrent"])
    comps = len({x["competition"] for x in out})
    print(f"champions-history.json: {len(out)} rows, {comps} competitions, {dated} dated, {current} current")

if __name__ == "__main__":
    main()

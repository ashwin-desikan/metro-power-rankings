#!/usr/bin/env python3
"""
Build MLB team-pages data from MLB.xlsx (the canonical workbook).

Reads the source workbook (`~/OneDrive/Excel Files/MLB.xlsx` on the user's
machine, or a path passed on the command line / the bindfs upload path in
session sandboxes) and emits a set of JSON files under public/data/mlb/.

Outputs:
  public/data/mlb/franchises.json              30 active franchises
  public/data/mlb/championships.json           one entry per WS title per franchise
  public/data/mlb/championship-appearances.json  WS appearances by franchise
  public/data/mlb/stadium-history.json         stadium-era rows grouped by canonical franchise
  public/data/mlb/award-winners.json           player awards grouped by franchise
  public/data/mlb/historical.json              defunct franchises for /teams/mlb/historical
  public/data/mlb/seasons-by-team.json         per-franchise season-by-season rows
  public/data/mlb/top-games-by-team.json       top 12 postseason games per franchise (Game Score)
  public/data/mlb/top-games-all-time.json      top 50 postseason games leaguewide
  public/data/mlb/top-games-by-decade.json     top 10 postseason games per decade

Canonical join key throughout: Year by Year col CI (Name). Every other
sheet's "Name"-equivalent column joins back to this. See the workbook's
"Claude Notes" sheet for full schema.

Lahman-derived player data on team pages is intentionally limited to Awards
in v1. Hall of Fame and All-Stars come in v1.5.

Usage:
  python scripts/build-mlb-data.py
  python scripts/build-mlb-data.py /path/to/MLB.xlsx
"""

import json
import os
import re
import sys
from pathlib import Path
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl",
                           "--quiet", "--break-system-packages"])
    import openpyxl


# -------- Constants --------

REPO_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_SOURCE_CANDIDATES = [
    # User's OneDrive (Windows) — preferred when present
    Path(os.path.expanduser("~/OneDrive/Excel Files/MLB.xlsx")),
    Path(os.path.expanduser("~/OneDrive/Excel Files/MLB - Copy.xlsx")),
    Path(os.path.expanduser("~/OneDrive/Excel Files/MLB_backup.xlsx")),
    # OneDrive via Linux bindfs mount (path varies by session)
    Path("/sessions/wizardly-relaxed-allen/mnt/Excel Files/MLB.xlsx"),
    Path("/sessions/wizardly-relaxed-allen/mnt/Excel Files/MLB - Copy.xlsx"),
    # Project-local fallback (if the user committed a frozen copy)
    REPO_ROOT / "data" / "mlb-source" / "MLB.xlsx",
    REPO_ROOT / "data" / "mlb-source" / "MLB - Copy.xlsx",
    # Bootstrap fallback to the most recent upload
    Path("/sessions/wizardly-relaxed-allen/mnt/uploads/MLB - Copy.xlsx"),
]

OUT_DIR = REPO_ROOT / "public" / "data" / "mlb"

# Era boundary: World Series begins in 1903 (modern era).
# Pre-1903 "championships" are pre-WS cups (Temple Cup, Chronicle-Telegraph,
# World's Series, NL pennants). Tracked as Oth Chmp / Oth Chmp App in Totals.
WS_ERA_START = 1903

# Curated award list for the team page. Order is editorial: headline single
# winners first, then category awards. Positional Silver Sluggers / Gold
# Gloves are intentionally NOT in v1 since they would push 600+ entries onto
# every franchise page.
AWARD_ORDER = [
    "AL Most Valuable Player",
    "NL Most Valuable Player",
    "AL Cy Young",
    "NL Cy Young",
    "AL Rookie of the Year",
    "NL Rookie of the Year",
    "AL Manager of the Year",
    "NL Manager of the Year",
    "World Series MVP",
    "ALCS MVP",
    "NLCS MVP",
    "All-Star Game MVP",
    "Hank Aaron Award AL",
    "Hank Aaron Award NL",
    "Roberto Clemente Award",
    "Triple Crown Batter",
    "Triple Crown Pitcher",
]

# Aliases: the workbook stores awards with assorted shorthand. Normalise on
# read so the page can use a stable display label.
AWARD_ALIASES = {
    "AL MVP": "AL Most Valuable Player",
    "NL MVP": "NL Most Valuable Player",
    "MVP NL": "NL Most Valuable Player",
    "MVP AL": "AL Most Valuable Player",
    "Cy Young AL": "AL Cy Young",
    "Cy Young NL": "NL Cy Young",
    "AL ROY": "AL Rookie of the Year",
    "NL ROY": "NL Rookie of the Year",
    "Rookie of the Year AL": "AL Rookie of the Year",
    "Rookie of the Year NL": "NL Rookie of the Year",
    "AL MOY": "AL Manager of the Year",
    "NL MOY": "NL Manager of the Year",
    "Manager of the Year AL": "AL Manager of the Year",
    "Manager of the Year NL": "NL Manager of the Year",
    "WS MVP": "World Series MVP",
    "WSMVP": "World Series MVP",
    "ALCS MVP": "ALCS MVP",
    "NLCS MVP": "NLCS MVP",
    "ASG MVP": "All-Star Game MVP",
    "All Star MVP": "All-Star Game MVP",
    "ASMVP": "All-Star Game MVP",
}


# -------- Helpers --------

def slugify(s):
    s = (s or "").lower().strip()
    s = re.sub(r"[àáâãäå]", "a", s)
    s = re.sub(r"[èéêë]", "e", s)
    s = re.sub(r"[ìíîï]", "i", s)
    s = re.sub(r"[òóôõö]", "o", s)
    s = re.sub(r"[ùúûü]", "u", s)
    s = re.sub(r"[ñ]", "n", s)
    s = re.sub(r"[^a-z0-9\s\-/]", "", s)
    s = re.sub(r"[/]", "-", s)
    s = re.sub(r"\s+", "-", s)
    s = re.sub(r"-+", "-", s)
    return s.strip("-")


def metro_slugify(metro_name):
    """Match the rankings-side metro slug convention."""
    if not metro_name:
        return None
    return slugify(metro_name)


def franchise_slug(canonical):
    """Slug for a franchise. Strips parenthetical disambiguators so
    'Athletics (1)' becomes 'athletics-1' and 'Brown Stockings (S)' becomes
    'brown-stockings-s'. Active 30 are clean."""
    return slugify(canonical.replace("(", " ").replace(")", " ")).rstrip("-")


def safe_int(val, default=0):
    if val is None or val == "":
        return default
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return default


def safe_float(val, default=0.0):
    if val is None or val == "":
        return default
    try:
        n = float(val)
        if n != n:  # NaN
            return default
        return n
    except (ValueError, TypeError):
        return default


def safe_str(val):
    if val is None:
        return ""
    return str(val).strip()


def is_truthy_yn(val):
    """The workbook uses 'Y' / 1 / None for many flags."""
    if val is None or val == "":
        return False
    s = str(val).strip().upper()
    return s in ("Y", "YES", "1", "TRUE")


def normalise_award(label):
    if not label:
        return None
    s = safe_str(label)
    return AWARD_ALIASES.get(s, s)


# -------- Workbook loader --------

def find_source(cli_path=None):
    if cli_path:
        p = Path(cli_path)
        if p.exists():
            return p
        print(f"Provided path not found: {p}", file=sys.stderr)
        sys.exit(1)
    for cand in DEFAULT_SOURCE_CANDIDATES:
        if cand.exists():
            return cand
    print("Could not find MLB.xlsx in any default location. Pass a path:",
          file=sys.stderr)
    print("  python scripts/build-mlb-data.py /path/to/MLB.xlsx", file=sys.stderr)
    sys.exit(1)


# -------- Sheet readers --------

def read_totals(wb):
    """One row per franchise. Returns dict keyed by canonical Name (col BA)."""
    ws = wb["Totals"]
    out = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        canonical = safe_str(row[52])  # BA
        if not canonical:
            continue
        out[canonical] = {
            "league_history": safe_str(row[0]),
            "city_history": safe_str(row[1]),
            "team_history": safe_str(row[2]),
            "all_time_w": safe_int(row[3]),
            "all_time_l": safe_int(row[4]),
            "all_time_t": safe_int(row[5]),
            "win_pct": safe_float(row[6]),
            "seasons": safe_int(row[7]),
            "five_hundred_seasons": safe_int(row[8]),
            "top_half_seasons": safe_int(row[9]),
            "playoff_appearances": safe_int(row[10]),
            "division_titles": safe_int(row[11]),
            "best_main_div": safe_int(row[12]),
            "best_record_seasons": safe_int(row[13]),
            "playoff_w": safe_int(row[14]),
            "playoff_l": safe_int(row[15]),
            "playoff_t": safe_int(row[16]),
            "playoff_win_pct": safe_float(row[17]),
            "series_w": safe_int(row[18]),
            "series_l": safe_int(row[19]),
            "lcs_appearances": safe_int(row[22]),
            "lcs_wins": safe_int(row[23]),
            "ws_appearances": safe_int(row[24]),
            "ws_champs": safe_int(row[25]),
            "oth_chp_app": safe_int(row[26]),
            "oth_chmp": safe_int(row[27]),
            "abs_win_pct": safe_float(row[28]),
            "last_champ_year": safe_int(row[29]) or None,
            "last_ws_app": safe_int(row[30]) or None,
            "last_lcs_app": safe_int(row[31]) or None,
            "last_division_title": safe_int(row[35]) or None,
            "last_playoff_app": safe_int(row[36]) or None,
            "last_season": safe_int(row[39]) or None,
            "is_current": is_truthy_yn(row[40]),
            "is_defunct": is_truthy_yn(row[41]),
            "total_championships": safe_int(row[42]),
            "reg_games": safe_int(row[43]),
            "play_games": safe_int(row[44]),
            "total_games": safe_int(row[45]),
            "home_w": safe_int(row[46]),
            "home_l": safe_int(row[47]),
            "home_win_pct": safe_float(row[48]),
            "road_w": safe_int(row[49]),
            "road_l": safe_int(row[50]),
            "road_win_pct": safe_float(row[51]),
        }
    return out


def read_year_by_year(wb):
    """Build season-by-season per canonical franchise from Year by Year.

    Returns dict { canonical_name: [season_row, ...] } sorted ascending by
    year. Also returns a side-channel { canonical_name: latest_meta } with
    the most recent row's city/team/metro/state/stadium for franchise hero
    fields.
    """
    ws = wb["Year by Year"]
    by_team = defaultdict(list)
    latest_meta = {}
    earliest_year = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        canonical = safe_str(row[86]) if len(row) > 86 else ""
        if not canonical:
            continue
        year = safe_int(row[3])
        if year == 0:
            continue
        league = safe_str(row[2])
        city = safe_str(row[4])
        team = safe_str(row[5])
        w = safe_int(row[6])
        l = safe_int(row[7])
        win_pct_raw = row[8]
        win_pct = safe_float(win_pct_raw) if not isinstance(win_pct_raw, str) else 0.0
        rs = safe_int(row[9])
        ra = safe_int(row[10])
        playoff_yn = is_truthy_yn(row[11])
        div_title_yn = is_truthy_yn(row[12])
        best_rec_leag = is_truthy_yn(row[14])
        lcs_app_yn = is_truthy_yn(row[15])
        ws_app_yn = is_truthy_yn(row[16])
        ws_champ_yn = is_truthy_yn(row[17])
        oth_chmp_app_yn = is_truthy_yn(row[30])
        oth_champ_yn = is_truthy_yn(row[31])
        division = safe_str(row[21])
        main_div = safe_str(row[22])
        place_raw = row[23]
        place = "" if place_raw is None else str(place_raw).strip()
        ballpark_season = safe_str(row[110]) if len(row) > 110 else ""
        home_city = safe_str(row[124]) if len(row) > 124 else ""
        metro = safe_str(row[125]) if len(row) > 125 else ""
        state = safe_str(row[126]) if len(row) > 126 else ""
        ballpark_canonical = safe_str(row[130]) if len(row) > 130 else ""

        # Skip in-progress (no W or L) rows for ranking purposes but keep
        # them in the season list as zero-filled. The team page filters in
        # the live ESPN row separately for the in-progress season.
        by_team[canonical].append({
            "year": year,
            "league": league,
            "city": city,
            "team": team,
            "w": w,
            "l": l,
            "t": 0,  # ties extinct in MLB after 1973-ish but kept for shape parity
            "win_pct": round(win_pct, 4) if win_pct else 0.0,
            "rs": rs,
            "ra": ra,
            "run_diff": rs - ra,
            "playoff": playoff_yn,
            "div_title": div_title_yn,
            "best_rec_leag": best_rec_leag,
            "lcs_app": lcs_app_yn,
            "ws_app": ws_app_yn,
            "champ": ws_champ_yn,
            "champ_app": ws_app_yn,  # WS app is the championship-appearance flag for modern era
            "oth_chmp_app": oth_chmp_app_yn,
            "oth_chmp": oth_champ_yn,
            "conf_final": lcs_app_yn,  # LCS is the conference final analog for MLB
            "division": division,
            "main_div": main_div,
            "place": place,
        })
        if canonical not in latest_meta or year > latest_meta[canonical]["year"]:
            latest_meta[canonical] = {
                "year": year,
                "city": city,
                "team": team,
                "league": league,
                "main_div": main_div,
                "division": division,
                "ballpark_season": ballpark_season,
                "ballpark_canonical": ballpark_canonical,
                "home_city": home_city,
                "metro": metro,
                "state": state,
            }
        earliest_year[canonical] = min(earliest_year.get(canonical, year), year)

    # Sort each franchise's seasons ascending
    for k in by_team:
        by_team[k].sort(key=lambda s: s["year"])

    return dict(by_team), latest_meta, earliest_year


def read_stadium_master(wb):
    """Returns { canonical_name: { city, metro, state, first, last } }."""
    ws = wb["Stadiums"]
    out = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        canonical = safe_str(row[0])
        if not canonical:
            continue
        sport = safe_str(row[9]) if len(row) > 9 else ""
        if sport and sport.upper() != "MLB":
            continue
        out[canonical] = {
            "canonical": canonical,
            "variants": safe_str(row[1]),
            "primary_teams": safe_str(row[2]),
            "city": safe_str(row[3]),
            "metro": safe_str(row[4]),
            "state": safe_str(row[5]),
            "first_season": safe_int(row[6]) or None,
            "last_season": safe_int(row[7]) or None,
            "total_seasons": safe_int(row[8]) or None,
            "status": safe_str(row[10]) if len(row) > 10 else "",
            "notes": safe_str(row[11]) if len(row) > 11 else "",
        }
    return out


def build_stadium_history(year_by_year, stadium_master):
    """Group stadium-eras by canonical franchise.

    For each franchise, walk its season rows in chronological order, group
    consecutive seasons by canonical ballpark name (EA), and emit a row per
    venue with first_year/last_year. Within each venue row, build an `eras`
    array using the season-name (DG) so renames (e.g., Mile High → Coors)
    show as multiple eras inside one canonical venue.
    """
    out = {}
    for canonical, rows in year_by_year.items():
        venues = []
        current = None
        for r in rows:
            yr = r["year"]
            # Need access to ballpark_season + ballpark_canonical from latest_meta
            # but those aren't on the per-season row. We need to re-read YbY for them.
            pass
    return out  # placeholder — replaced by build_stadium_history_full below


def build_stadium_history_full(wb, stadium_master):
    """Build stadium history per canonical franchise by re-walking Year by
    Year so we keep the per-season ballpark name + canonical mapping."""
    ws = wb["Year by Year"]
    # First pass: collect (year, season_name, canonical_name) tuples per franchise
    per_team_seasons = defaultdict(list)
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        canonical_fr = safe_str(row[86]) if len(row) > 86 else ""
        if not canonical_fr:
            continue
        year = safe_int(row[3])
        if year == 0:
            continue
        season_name = safe_str(row[110]) if len(row) > 110 else ""
        canon_park = safe_str(row[130]) if len(row) > 130 else season_name
        per_team_seasons[canonical_fr].append((year, season_name, canon_park))

    out = {}
    for franchise_canon, seasons in per_team_seasons.items():
        seasons.sort(key=lambda x: x[0])
        venues = []  # list of {canonical, first_year, last_year, eras: [{era_name,first,last}]}

        def push_era(venues, year, season_name, canon_park):
            if not canon_park:
                return
            if venues and venues[-1]["canonical"] == canon_park:
                v = venues[-1]
                v["last_year"] = year
                if v["eras"] and v["eras"][-1]["era_name"] == season_name:
                    v["eras"][-1]["last_year"] = year
                else:
                    v["eras"].append({
                        "era_name": season_name or canon_park,
                        "first_year": year,
                        "last_year": year,
                    })
            else:
                venues.append({
                    "canonical": canon_park,
                    "first_year": year,
                    "last_year": year,
                    "eras": [{
                        "era_name": season_name or canon_park,
                        "first_year": year,
                        "last_year": year,
                    }],
                })

        for (year, season_name, canon_park) in seasons:
            push_era(venues, year, season_name, canon_park)

        # Attach city/metro/state from stadium master where we have it
        for v in venues:
            sm = stadium_master.get(v["canonical"])
            if sm:
                v["city"] = sm["city"]
                v["metro"] = sm["metro"]
                v["state"] = sm["state"]
            else:
                v["city"] = ""
                v["metro"] = ""
                v["state"] = ""

        out[franchise_canon] = venues
    return out


def read_awards(wb):
    """Group player awards by canonical franchise name (col I).

    Returns { canonical: { award_label: [ {year, player, team_at_time, league}, ... ] } }
    """
    ws = wb["Awards"]
    out = defaultdict(lambda: defaultdict(list))
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        year = safe_int(row[1])
        player = safe_str(row[2])
        team_at_time = safe_str(row[4])
        award_raw = row[5] if len(row) > 5 else None
        award = normalise_award(award_raw)
        if not award:
            continue
        if award not in AWARD_ORDER:
            continue
        league = safe_str(row[6]) if len(row) > 6 else ""
        canonical = safe_str(row[8]) if len(row) > 8 else ""
        if not canonical or not player or not year:
            continue
        out[canonical][award].append({
            "year": year,
            "player": player,
            "team_at_time": team_at_time,
            "league": league,
        })
    # Sort each award list newest-first
    for canon in out:
        for award in out[canon]:
            out[canon][award].sort(key=lambda r: -r["year"])
    return out


def read_top_games(wb):
    """Read Detailed Playoffs and produce two views:
       - per-canonical-franchise top games (this-team perspective)
       - league-wide deduped top games
    Sort key is Game Score (col AS).
    """
    ws = wb["Detailed Playoffs"]
    by_team = defaultdict(list)
    all_games = []
    seen_game_ids = set()
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        canonical = safe_str(row[24]) if len(row) > 24 else ""  # Y = Name
        if not canonical:
            continue
        year = safe_int(row[1])  # B = Sea.
        round_label = safe_str(row[2])  # C = Round
        game_num = safe_int(row[3]) or None  # D = G #
        date_raw = row[4]  # E = Date
        date_iso = None
        try:
            if hasattr(date_raw, "strftime"):
                date_iso = date_raw.strftime("%Y-%m-%d")
        except Exception:
            date_iso = None
        city = safe_str(row[5])
        team = safe_str(row[6])
        result = safe_str(row[7])  # W/L/T
        opp_city = safe_str(row[8])
        opp_team = safe_str(row[9])
        rf = safe_int(row[10])
        ra = safe_int(row[11])
        fin_inn = safe_int(row[12]) or None
        ballpark_canon = safe_str(row[35]) if len(row) > 35 else safe_str(row[13])
        opponent_canon = safe_str(row[25]) if len(row) > 25 else ""
        home_away = safe_str(row[43]) if len(row) > 43 else ""
        game_score_raw = row[44] if len(row) > 44 else None
        try:
            game_score = float(game_score_raw)
        except (TypeError, ValueError):
            continue
        if game_score != game_score:  # NaN
            continue

        # This-team-perspective row for franchise pages
        by_team[canonical].append({
            "year": year,
            "date": date_iso,
            "round": round_label,
            "game_num": game_num,
            "team_city": city,
            "team": team,
            "team_canonical": canonical,
            "opp_city": opp_city,
            "opp_team": opp_team,
            "opp_canonical": opponent_canon,
            "rf": rf,
            "ra": ra,
            "result": result,
            "fin_inn": fin_inn,
            "extra_innings": (fin_inn is not None and fin_inn > 9),
            "stadium": ballpark_canon,
            "is_home": home_away.lower() in ("home", "h"),
            "game_score": round(game_score, 4),
        })

        # League-wide dedupe via symmetric key (date + pair of canonicals)
        if not opponent_canon:
            continue
        sym_key = f"{date_iso}|{min(canonical, opponent_canon)}|{max(canonical, opponent_canon)}|{game_num}"
        if sym_key in seen_game_ids:
            continue
        seen_game_ids.add(sym_key)

        if result == "W":
            winner_city, winner_team, winner_canon = city, team, canonical
            loser_city, loser_team, loser_canon = opp_city, opp_team, opponent_canon
            winner_score, loser_score = rf, ra
        elif result == "L":
            winner_city, winner_team, winner_canon = opp_city, opp_team, opponent_canon
            loser_city, loser_team, loser_canon = city, team, canonical
            winner_score, loser_score = ra, rf
        else:
            winner_city, winner_team, winner_canon = city, team, canonical
            loser_city, loser_team, loser_canon = opp_city, opp_team, opponent_canon
            winner_score, loser_score = rf, ra

        all_games.append({
            "year": year,
            "date": date_iso,
            "round": round_label,
            "game_num": game_num,
            "winner_city": winner_city,
            "winner_team": winner_team,
            "winner_canonical": winner_canon,
            "loser_city": loser_city,
            "loser_team": loser_team,
            "loser_canonical": loser_canon,
            "winner_score": winner_score,
            "loser_score": loser_score,
            "fin_inn": fin_inn,
            "extra_innings": (fin_inn is not None and fin_inn > 9),
            "is_tie": result == "T",
            "stadium": ballpark_canon,
            "game_score": round(game_score, 4),
        })

    return all_games, by_team


# -------- Builders --------

def build_franchises(totals, latest_meta, year_by_year, earliest_year):
    """Build the 30 active franchise rows."""
    out = []
    for canonical, t in totals.items():
        if not t["is_current"]:
            continue
        meta = latest_meta.get(canonical, {})
        seasons = year_by_year.get(canonical, [])
        founding = earliest_year.get(canonical) or 0
        prior_cities = []
        # Year by Year city column has historical city; build a distinct list
        # excluding the current city
        cur_city = meta.get("home_city") or meta.get("city") or ""
        for s in seasons:
            sc = s["city"]
            if sc and sc != cur_city and sc not in prior_cities:
                prior_cities.append(sc)
        out.append({
            "canonical": canonical,
            "slug": franchise_slug(canonical),
            "key": "",  # filled later from Team Lookup if needed
            "name": meta.get("team") or canonical,
            "display_name": f"{meta.get('home_city') or meta.get('city')} {meta.get('team') or canonical}".strip(),
            "city": cur_city,
            "team": meta.get("team") or canonical,
            "league": meta.get("league") or t["league_history"],
            "conf": meta.get("main_div") or "",
            "division": meta.get("division") or "",
            "metro": meta.get("metro") or "",
            "metro_slug": metro_slugify(meta.get("metro")),
            "state": meta.get("state") or "",
            "stadium": meta.get("ballpark_canonical") or "",
            "stadium_season_name": meta.get("ballpark_season") or "",
            "founding_year": founding,
            "championships": t["ws_champs"],
            "pre_ws_championships": t["oth_chmp"],
            "total_championships": t["total_championships"],
            "ws_appearances": t["ws_appearances"],
            "lcs_appearances": t["lcs_appearances"],
            "division_titles": t["division_titles"],
            "playoff_appearances": t["playoff_appearances"],
            "playoff_w": t["playoff_w"],
            "playoff_l": t["playoff_l"],
            "all_time_w": t["all_time_w"],
            "all_time_l": t["all_time_l"],
            "all_time_t": t["all_time_t"],
            "win_pct": round(t["win_pct"], 4),
            "seasons": t["seasons"],
            "five_hundred_seasons": t["five_hundred_seasons"],
            "best_rec_seasons": t["best_record_seasons"],
            "last_championship_year": t["last_champ_year"],
            "last_ws_app": t["last_ws_app"],
            "last_lcs_app": t["last_lcs_app"],
            "last_division_title": t["last_division_title"],
            "last_playoff_app": t["last_playoff_app"],
            "prior_cities": prior_cities,
        })
    out.sort(key=lambda f: f["name"])
    return out


def build_championships(year_by_year):
    """One entry per WS title per franchise."""
    out = defaultdict(list)
    for canonical, rows in year_by_year.items():
        for r in rows:
            if r["champ"]:
                out[canonical].append({
                    "year": r["year"],
                    "city": r["city"],
                    "team": r["team"],
                    "league": r["league"],
                    "era": "ws" if r["year"] >= WS_ERA_START else "pre_ws",
                })
            elif r["oth_chmp"] and r["year"] < WS_ERA_START:
                # Pre-1903 cup wins (Temple Cup, World's Series, etc.)
                out[canonical].append({
                    "year": r["year"],
                    "city": r["city"],
                    "team": r["team"],
                    "league": r["league"],
                    "era": "pre_ws",
                })
    return {k: sorted(v, key=lambda r: r["year"]) for k, v in out.items()}


def build_championship_appearances(year_by_year):
    """One entry per WS appearance per franchise, including losses."""
    out = defaultdict(list)
    for canonical, rows in year_by_year.items():
        for r in rows:
            if r["ws_app"]:
                out[canonical].append({
                    "year": r["year"],
                    "city": r["city"],
                    "team": r["team"],
                    "won": r["champ"],
                    "era": "ws",
                })
            elif r["oth_chmp_app"] and r["year"] < WS_ERA_START:
                out[canonical].append({
                    "year": r["year"],
                    "city": r["city"],
                    "team": r["team"],
                    "won": r["oth_chmp"],
                    "era": "pre_ws",
                })
    return {k: sorted(v, key=lambda r: r["year"]) for k, v in out.items()}


def build_historical(totals, year_by_year):
    """Defunct franchises for /teams/mlb/historical, with first/last year."""
    rows = []
    for canonical, t in totals.items():
        if not t["is_defunct"]:
            continue
        seasons = year_by_year.get(canonical, [])
        years = [s["year"] for s in seasons if s.get("year")]
        first_year = min(years) if years else None
        last_year = max(years) if years else None
        rows.append({
            "canonical": canonical,
            "name": canonical,
            "city": t["city_history"],
            "team_historical": t["team_history"],
            "league": t["league_history"],
            "seasons": t["seasons"],
            "first_year": first_year,
            "last_year": last_year,
            "w": t["all_time_w"],
            "l": t["all_time_l"],
            "t": t["all_time_t"],
            "win_pct": round(t["win_pct"], 4),
            "championships": t["total_championships"],
        })
    rows.sort(key=lambda r: (
        -r["championships"],
        -(r["last_year"] or 0),
        r["city"] or "",
    ))
    return rows


def build_top_games_by_team(by_team, franchises, top_n=12):
    out = {}
    for f in franchises:
        canon = f["canonical"]
        rows = sorted(by_team.get(canon, []), key=lambda g: -g["game_score"])[:top_n]
        out[f["slug"]] = rows
    return out


def build_top_games_all_time(all_games, top_n=50):
    return sorted(all_games, key=lambda g: -g["game_score"])[:top_n]


def build_top_games_by_decade(all_games, top_n_per_decade=10):
    by_decade = defaultdict(list)
    for g in all_games:
        if not g["year"]:
            continue
        decade = (g["year"] // 10) * 10
        by_decade[decade].append(g)
    out = {}
    for decade, games in by_decade.items():
        out[str(decade)] = sorted(games, key=lambda g: -g["game_score"])[:top_n_per_decade]
    return out


def build_seasons_by_team(franchises, year_by_year):
    """Pre-slugged season rows for the per-team page."""
    out = {}
    for f in franchises:
        canon = f["canonical"]
        rows = year_by_year.get(canon, [])
        out[f["slug"]] = rows
    return out


# -------- Main --------

def main():
    cli_path = sys.argv[1] if len(sys.argv) > 1 else None
    src = find_source(cli_path)
    print(f"Reading: {src}")
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    print(f"Sheets: {wb.sheetnames}")

    print("Reading Totals...")
    totals = read_totals(wb)
    print(f"  {len(totals)} franchise rows (active + defunct)")

    print("Reading Year by Year...")
    yby, latest_meta, earliest_year = read_year_by_year(wb)
    print(f"  {len(yby)} canonical names, {sum(len(v) for v in yby.values())} team-seasons")

    print("Reading Stadium master...")
    stadium_master = read_stadium_master(wb)
    print(f"  {len(stadium_master)} MLB ballparks in master table")

    print("Building stadium history...")
    stadiums = build_stadium_history_full(wb, stadium_master)
    print(f"  {len(stadiums)} franchises with stadium history")

    print("Reading Awards...")
    awards = read_awards(wb)
    print(f"  {len(awards)} franchises with curated awards")

    print("Reading Detailed Playoffs top games (Game Score)...")
    all_games, games_by_team = read_top_games(wb)
    print(f"  {len(all_games)} unique games scored, {len(games_by_team)} franchises with game data")

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    franchises = build_franchises(totals, latest_meta, yby, earliest_year)
    print(f"Built franchises: {len(franchises)}")
    (OUT_DIR / "franchises.json").write_text(json.dumps(franchises, indent=2, ensure_ascii=False))

    champs = build_championships(yby)
    (OUT_DIR / "championships.json").write_text(json.dumps(champs, indent=2, ensure_ascii=False))

    champ_apps = build_championship_appearances(yby)
    (OUT_DIR / "championship-appearances.json").write_text(json.dumps(champ_apps, indent=2, ensure_ascii=False))

    (OUT_DIR / "stadium-history.json").write_text(json.dumps(stadiums, indent=2, ensure_ascii=False))

    awards_plain = {team: dict(by_award) for team, by_award in awards.items()}
    (OUT_DIR / "award-winners.json").write_text(json.dumps(awards_plain, indent=2, ensure_ascii=False))

    historical = build_historical(totals, yby)
    print(f"Built historical: {len(historical)}")
    (OUT_DIR / "historical.json").write_text(json.dumps(historical, indent=2, ensure_ascii=False))

    seasons_out = build_seasons_by_team(franchises, yby)
    (OUT_DIR / "seasons-by-team.json").write_text(json.dumps(seasons_out, indent=2, ensure_ascii=False))

    top_by_team = build_top_games_by_team(games_by_team, franchises, top_n=12)
    (OUT_DIR / "top-games-by-team.json").write_text(json.dumps(top_by_team, indent=2, ensure_ascii=False))

    top_all_time = build_top_games_all_time(all_games, top_n=50)
    (OUT_DIR / "top-games-all-time.json").write_text(json.dumps(top_all_time, indent=2, ensure_ascii=False))

    top_by_decade = build_top_games_by_decade(all_games, top_n_per_decade=10)
    (OUT_DIR / "top-games-by-decade.json").write_text(json.dumps(top_by_decade, indent=2, ensure_ascii=False))

    print("\nWrote:")
    for f in sorted(OUT_DIR.glob("*.json")):
        print(f"  {f.relative_to(REPO_ROOT)}  ({f.stat().st_size:,} bytes)")


if __name__ == "__main__":
    main()

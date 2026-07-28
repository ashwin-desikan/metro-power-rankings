#!/usr/bin/env python3
"""Build unmatched_api_teams.xlsx from the api-football team scan, applying the SAME
collision guard as refresh.py so the "Resolvable" tab only lists teams that can ACTUALLY be
linked by a Lookup edit.

Why this exists: the earlier ad-hoc report resolved every unmatched api id by name and called
anything that matched a Lookup club "Resolvable (not yet linked)". That was misleading: many of
those ids are DUPLICATE api ids for a club already mapped to a single primary team_id. No Lookup
edit links them, because refresh.py's collision guard (one canonical club -> one api team_id)
rejects the second id. Those belong in football_team_alias, not on a to-do list. This generator
splits them out.

Tabs written:
  - "Unmatched (no Lookup club)"      : does not resolve to any Lookup club -> add the club to Lookup
  - "Resolvable (not yet linked)"     : resolves cleanly AND the club is unowned -> a real to-do
  - "Duplicate api id (already mapped)": resolves to a club already owned by another team_id, or
                                          already in football_team_alias -> add to the alias table

Inputs: scripts/apifootball/_scratch/unmatched_teams.json (refresh via audit_unmatched.py first),
and live football_lookup / football_team / football_team_alias from Supabase.

    python build_unmatched_report.py                 # -> unmatched_api_teams.xlsx (here)
    python build_unmatched_report.py /path/out.xlsx  # custom output path
Env: Supabase key (same resolution as refresh.py).
"""
import os, sys, json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from refresh import supa_get, supa_key, build_resolver

HERE = os.path.dirname(os.path.abspath(__file__))
UNMATCHED_JSON = os.path.join(HERE, "_scratch", "unmatched_teams.json")

RESOLVER_SELECT = ("/rest/v1/football_lookup?select="
                   "cur_name,team,lookup_name,uefa_name,uefa_name_2,efs_name,api_name,api_name_2,country,level")


def classify(unmatched, resolve, owners, alias_map, known_ids):
    """Split the scanned unmatched teams into three buckets.

    owners     : {(canonical_name, country): team_id}  from football_team
    alias_map  : {dup_team_id: primary_team_id}         from football_team_alias
    known_ids  : set of team_id already present in football_team (skip -- already linked)
    Returns (unmatched_rows, resolvable_rows, duplicate_rows).
    """
    unm, resolvable, dup = [], [], []
    for t in unmatched:
        tid = t.get("team_id")
        if tid in known_ids:
            continue  # already linked since the scan; not a gap
        if tid in alias_map:
            dup.append({**t, "resolves_to": None, "owner_team_id": alias_map[tid],
                        "reason": "already in football_team_alias"})
            continue
        rec = resolve(t.get("name"))
        if not rec:
            unm.append(t)
            continue
        canon, country = rec.get("team"), rec.get("country")
        owner = owners.get((canon, country))
        if owner is not None and owner != tid:
            dup.append({**t, "resolves_to": canon, "owner_team_id": owner,
                        "reason": "canonical club already owned by another team_id "
                                  "-> add %s to football_team_alias" % tid})
        else:
            resolvable.append({**t, "resolves_to": canon})
    return unm, resolvable, dup


def _seasons(t):
    return ", ".join(t.get("seasons") or [])


def write_xlsx(path, unm, resolvable, dup):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    head_fill = PatternFill("solid", fgColor="1F3864")

    def sheet(title, rows, cols):
        ws = wb.create_sheet(title[:31])
        ws.append([c[0] for c in cols])
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF"); c.fill = head_fill
            c.alignment = Alignment(wrap_text=True, vertical="center")
        for r in rows:
            ws.append([c[1](r) for c in cols])
        for i, c in enumerate(cols, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = c[2]
        ws.freeze_panes = "A2"
        return ws

    base = [
        ("API team_id", lambda r: r.get("team_id"), 12),
        ("API name",    lambda r: r.get("name"), 30),
        ("Country",     lambda r: r.get("country"), 16),
        ("Min lvl",     lambda r: r.get("min_level"), 9),
        ("In Europe",   lambda r: "Yes" if r.get("in_europe") else "", 10),
        ("Appear.",     lambda r: r.get("appearances"), 9),
        ("Seasons",     lambda r: _seasons(r), 46),
        ("Priority",    lambda r: r.get("priority"), 9),
    ]
    sheet("Unmatched (no Lookup club)", unm, base)
    sheet("Resolvable (not yet linked)", resolvable,
          base[:7] + [("Resolves to (Lookup)", lambda r: r.get("resolves_to"), 26), base[7]])
    sheet("Duplicate api id (already mapped)", dup,
          base[:7] + [("Resolves to (Lookup)", lambda r: r.get("resolves_to"), 26),
                      ("Primary team_id", lambda r: r.get("owner_team_id"), 15),
                      ("Action", lambda r: r.get("reason"), 60)])
    wb.save(path)


def main():
    out = sys.argv[1] if len(sys.argv) > 1 else os.path.join(HERE, "unmatched_api_teams.xlsx")
    unmatched = json.load(open(UNMATCHED_JSON, encoding="utf-8"))
    skey = supa_key()
    resolve = build_resolver(supa_get(RESOLVER_SELECT, skey))
    ft = supa_get("/rest/v1/football_team?select=team_id,canonical_name,country", skey)
    owners = {(r["canonical_name"], r["country"]): r["team_id"] for r in ft}
    known_ids = {r["team_id"] for r in ft}
    alias_map = {r["dup_team_id"]: r["primary_team_id"]
                 for r in supa_get("/rest/v1/football_team_alias?select=dup_team_id,primary_team_id", skey)}
    unm, resolvable, dup = classify(unmatched, resolve, owners, alias_map, known_ids)
    write_xlsx(out, unm, resolvable, dup)
    print("[report] %s  ->  unmatched=%d  resolvable=%d  duplicate=%d"
          % (os.path.basename(out), len(unm), len(resolvable), len(dup)))


if __name__ == "__main__":
    main()

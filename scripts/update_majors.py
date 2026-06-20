#!/usr/bin/env python3
"""Auto-update Majors.xlsx with newly-completed major champions.

Designed to be driven by a scheduled Cowork task. The script is deterministic;
the live champion lookup is done by the agent (web search) and fed in via --add.

  python scripts/update_majors.py --list-missing [--asof YYYY-MM-DD]
      -> JSON list of majors that should be complete by `asof` but are absent.
  python scripts/update_majors.py --add "SPORT|TOURNAMENT|GENDER|YEAR|CHAMPION|NATION|NOTE"
      -> appends one champion, recomputes (n/total) counters, saves.
      GENDER is M/W for tennis, blank for golf. NOTE optional.
  python scripts/update_majors.py --verify-counters
      -> regression: regenerate counters from rows, diff vs stored (must be clean).

After any --add, re-run scripts/build-majors-data.py to refresh the hub JSON.
"""
import argparse, json, os, sys, datetime
from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, "Majors.xlsx")

TENNIS_ORDER = {"Australian Open":1,"French Open":2,"Wimbledon":3,"US Open":4}
GOLF_ORDER   = {"Masters Tournament":1,"U.S. Open":2,"The Open Championship":3,"PGA Championship":4}

# event -> (approx conclusion month, day) used only to decide "should exist by now"
CAL = {
 ("Tennis","Australian Open"):(1,28),("Tennis","French Open"):(6,8),
 ("Tennis","Wimbledon"):(7,14),("Tennis","US Open"):(9,8),
 ("Golf","Masters Tournament"):(4,14),("Golf","PGA Championship"):(5,21),
 ("Golf","U.S. Open"):(6,21),("Golf","The Open Championship"):(7,21),
}
TENNIS_EVENTS=["Australian Open","French Open","Wimbledon","US Open"]
GOLF_EVENTS=["Masters Tournament","U.S. Open","The Open Championship","PGA Championship"]

def load():
    return load_workbook(XLSX)

def rows_of(ws):
    hdr=[c.value for c in ws[1]]
    out=[]
    for r in ws.iter_rows(min_row=2):
        out.append({hdr[i]:r[i].value for i in range(len(hdr))})
    return hdr,out

def list_missing(asof):
    wb=load(); miss=[]
    tw=wb["TennisMajors"]; _,trows=rows_of(tw)
    gw=wb["GolfMajors"]; _,grows=rows_of(gw)
    have_t={(int(r["Year"]),r["Tournament"],r["Gender"]) for r in trows if r["Year"]}
    have_g={(int(r["Year"]),r["Tournament"]) for r in grows if r["Year"]}
    yr=asof.year
    for y in (yr-1,yr):
        for ev in TENNIS_EVENTS:
            m,d=CAL[("Tennis",ev)]; end=datetime.date(y,m,d)
            if asof>=end:
                for g in ("Men","Women"):
                    if (y,ev,g) not in have_t:
                        miss.append({"sport":"Tennis","tournament":ev,"gender":"M" if g=="Men" else "W","year":y})
        for ev in GOLF_EVENTS:
            m,d=CAL[("Golf",ev)]; end=datetime.date(y,m,d)
            if asof>=end and (y,ev) not in have_g:
                miss.append({"sport":"Golf","tournament":ev,"gender":"","year":y})
    # Ryder (biennial, odd years, ~Sep 28) and Davis (annual, ~Nov 27) flagged as reminders
    rw=wb["RyderCup"]; _,rr=rows_of(rw); have_r={int(r["Year"]) for r in rr if r["Year"]}
    for y in (yr-1,yr):
        if y%2==1 and asof>=datetime.date(y,9,28) and y not in have_r:
            miss.append({"sport":"Golf","tournament":"Ryder Cup","gender":"","year":y,"sheet":"RyderCup"})
    print(json.dumps(miss,indent=2))
    return miss

def recompute(wb, force=None):
    """Renumber (n/total) preserving the source\'s existing order and its
    intentionally-uncounted rows (e.g. pre-1925 French, occupation-era). Only rows
    that already carry a counter, or are newly added (in `force`), are sequenced."""
    from collections import defaultdict
    force=force or set()
    changed=0
    for sheet,order,gk in [("TennisMajors",TENNIS_ORDER,("Gender",)),("GolfMajors",GOLF_ORDER,())]:
        ws=wb[sheet]; hdr=[c.value for c in ws[1]]; ci={h:i for i,h in enumerate(hdr)}
        rows=[(r[0].row,[c.value for c in r]) for r in ws.iter_rows(min_row=2)]
        groups=defaultdict(list)
        for ridx,vals in rows:
            if "unrecognized" in str(vals[ci["Note"]] or ""): continue
            if vals[ci["CareerTitleNo"]] is None and (sheet,ridx) not in force: continue
            groups[(vals[ci["Champion"]],tuple(vals[ci[k]] for k in gk))].append((ridx,vals))
        for key,items in groups.items():
            def sortkey(iv):
                ridx,vals=iv; yr=vals[ci["Year"]] or 0; sn=vals[ci["CareerTitleNo"]]
                return (yr,0,sn) if sn is not None else (yr,1,order.get(vals[ci["Tournament"]],9))
            items.sort(key=sortkey); total=len(items)
            for n,(ridx,vals) in enumerate(items,1):
                if vals[ci["CareerTitleNo"]]!=n: ws.cell(row=ridx,column=ci["CareerTitleNo"]+1,value=n); changed+=1
                if vals[ci["CareerMajorTotal"]]!=total: ws.cell(row=ridx,column=ci["CareerMajorTotal"]+1,value=total); changed+=1
    return changed

def verify_counters():
    wb=load()
    # snapshot stored
    before={}
    for sheet in ("TennisMajors","GolfMajors"):
        ws=wb[sheet]; hdr=[c.value for c in ws[1]]; ci={h:i for i,h in enumerate(hdr)}
        for r in ws.iter_rows(min_row=2):
            vals=[c.value for c in r]
            key=(sheet,vals[ci["Year"]],vals[ci["Tournament"]],vals[ci.get("Gender",ci["Champion"])] if "Gender" in ci else None,vals[ci["Champion"]])
            before[key]=(vals[ci["CareerTitleNo"]],vals[ci["CareerMajorTotal"]])
    changed=recompute(wb)
    print(f"cells that differ from a clean regeneration: {changed}")
    print("PASS — stored counters match regeneration" if changed==0 else "MISMATCH — see above")
    return changed

def add(spec):
    parts=(spec.split("|")+[""]*7)[:7]
    sport,tour,gender,year,champ,nation,note=[p.strip() for p in parts]
    year=int(year)
    wb=load()
    if sport=="Tennis":
        ws=wb["TennisMajors"]; hdr=[c.value for c in ws[1]]
        g="Men" if gender.upper()=="M" else "Women"
        for r in ws.iter_rows(min_row=2):
            v=[c.value for c in r]
            if v[0]==year and v[1]==tour and v[2]==g:
                print(f"SKIP duplicate: {year} {tour} {g}"); return
        ws.append([year,tour,g,champ,nation or None,None,None,note or None])
        new=("TennisMajors",ws.max_row)
    elif sport=="Golf":
        ws=wb["GolfMajors"]; 
        for r in ws.iter_rows(min_row=2):
            v=[c.value for c in r]
            if v[0]==year and v[1]==tour:
                print(f"SKIP duplicate: {year} {tour}"); return
        ws.append([year,tour,champ,nation or None,None,None,note or None])
        new=("GolfMajors",ws.max_row)
    else:
        print("Unknown sport (use Tennis|Golf; Ryder/Davis are edited manually for now)"); return
    n=recompute(wb, force={new})
    tmp=XLSX+".tmp"; wb.save(tmp); os.replace(tmp,XLSX)
    print(f"ADDED {sport} {year} {tour} {gender} {champ} ({nation}); recomputed {n} counter cells.")

if __name__=="__main__":
    ap=argparse.ArgumentParser()
    ap.add_argument("--list-missing",action="store_true")
    ap.add_argument("--asof")
    ap.add_argument("--add")
    ap.add_argument("--verify-counters",action="store_true")
    a=ap.parse_args()
    if a.verify_counters: verify_counters()
    elif a.list_missing:
        asof=datetime.date.fromisoformat(a.asof) if a.asof else datetime.date.today()
        list_missing(asof)
    elif a.add: add(a.add)
    else: ap.print_help()

#!/usr/bin/env python3
"""Build public/data/cfl/data.json from OtherLeagues.xlsx.

Two sheets:
  - "CFL Standings"        : one row per team-season (1945-).  Join key = "Name" (col 14, canonical).
  - "CFL Grey Cup Finals"  : two rows per game (winner + loser), full history (1909-).
                             Join key = "Name" (col 12); "Opponent" = col 13.

Season-by-season records come from Standings; Grey Cup final appearances + the
honor roll come from the Grey Cup Finals sheet (authoritative full history, so a
franchise's Cup count includes pre-1945 titles its standings rows don't cover).
"""
import json, os, re, sys, unicodedata
from collections import defaultdict
try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl required (pip install openpyxl)")

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT  = os.path.join(ROOT, "public", "data", "cfl", "data.json")

def find_wb():
    cands = [os.environ.get("CFL_XLSX"),
             sys.argv[1] if len(sys.argv) > 1 else None,
             os.path.expanduser("~/OneDrive/Excel Files/OtherLeagues.xlsx"),
             os.path.join(ROOT, "OtherLeagues.xlsx")]
    for c in cands:
        if c and os.path.exists(c):
            return c
    sys.exit("OtherLeagues.xlsx not found (set CFL_XLSX)")

def slugify(s):
    s = unicodedata.normalize("NFKD", str(s).lower())
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"^-+|-+$", "", re.sub(r"[^a-z0-9]+", "-", s))

def S(v):  return str(v).strip() if v is not None else ""
def I(v):
    try: return int(float(v))
    except (TypeError, ValueError): return 0
def YN(v): return S(v).upper() == "Y"

# Canonical franchise -> metro slug (all verified present in metros.json).
CFL_METRO = {
 "Toronto Argonauts":"toronto","Edmonton Elks":"edmonton","Winnipeg Blue Bombers":"winnipeg",
 "Hamilton Tiger-Cats":"hamilton","Montreal Alouettes":"montreal","Calgary Stampeders":"calgary",
 "BC Lions":"vancouver","Saskatchewan Roughriders":"regina","Ottawa RedBlacks":"ottawa",
 "Ottawa Rough Riders":"ottawa","Ottawa Renegades":"ottawa","Ottawa Trojans":"ottawa",
 "Baltimore CFLs":"washington-baltimore","San Antonio Texans":"san-antonio",
 "Birmingham Barracudas":"birmingham","Memphis Mad Dogs":"memphis","Shreveport Pirates":"shreveport",
 "Sacramento Gold Miners":"sacramento","Las Vegas Posse":"las-vegas",
 "Kitchener-Waterloo Dutchmen":"kitchener-waterloo","Toronto Balmy Beach":"toronto",
 "Sarnia Imperials":"sarnia","Hamilton Wildcats":"hamilton","Hamilton Tigers":"hamilton",
 "Montreal Hornets":"montreal",
}

wb = openpyxl.load_workbook(find_wb(), read_only=True, data_only=True)

# ── 1. Standings → per-canonical seasons ──────────────────────────────────────
st_rows = list(wb["CFL Standings"].iter_rows(values_only=True))[1:]
seasons = defaultdict(list)          # canonical -> [season dict]
era_names = defaultdict(set)         # canonical -> {team era names}
for r in st_rows:
    if not r or not S(r[14]):
        continue
    canon = S(r[14])
    yr = I(r[0])
    if not yr:
        continue
    era_names[canon].add(S(r[3]))
    seasons[canon].append({
        "year": yr, "division": S(r[2]), "team": S(r[3]),
        "w": I(r[4]), "l": I(r[5]), "t": I(r[6]),
        "pct": round(float(r[7]), 3) if isinstance(r[7], (int, float)) else 0.0,
        "pf": I(r[8]), "pa": I(r[9]),
        "play_app": YN(r[10]), "gc_final": YN(r[11]), "grey_cup": YN(r[12]),
        "playoff_result": S(r[13]),
    })

latest = max((s["year"] for rows in seasons.values() for s in rows), default=0)

# name (canonical OR era) -> slug, for joining the Grey Cup sheet + opponents
name2slug = {}
for canon in seasons:
    sl = slugify(canon)
    name2slug[canon.lower()] = sl
    for e in era_names[canon]:
        name2slug[e.lower()] = sl

# ── 2. Grey Cup Finals → per-team final appearances + honor roll ──────────────
gc_rows = list(wb["CFL Grey Cup Finals"].iter_rows(values_only=True))[1:]
gc_finals = defaultdict(list)        # slug -> [final dict]
honor = {}                           # game-year -> honor-roll entry (winner row)
unresolved = set()
for r in gc_rows:
    if not r or not S(r[12]):
        continue
    name = S(r[12]); date = r[1]
    yr = date.year if hasattr(date, "year") else I(r[1])
    sl = name2slug.get(name.lower())
    if sl is None:
        unresolved.add(name)
    opp = S(r[13]); opp_slug = name2slug.get(opp.lower())
    result = S(r[3]).upper()
    entry = {
        "game": S(r[0]), "year": yr, "result": result,
        "pf": I(r[4]), "pa": I(r[5]), "ot": YN(r[6]),
        "opponent": opp, "opponent_slug": opp_slug,
        "venue": S(r[8]), "city": S(r[9]), "attendance": I(r[11]),
    }
    if sl:
        gc_finals[sl].append(entry)
    if result == "W":
        honor[(yr, S(r[0]))] = {
            "year": yr, "game": S(r[0]),
            "champion": name, "champion_slug": sl,
            "runner_up": opp, "runner_up_slug": opp_slug,
            "score": f"{I(r[4])}–{I(r[5])}", "ot": YN(r[6]),
            "venue": S(r[8]), "city": S(r[9]), "attendance": I(r[11]),
        }
for sl in gc_finals:
    gc_finals[sl].sort(key=lambda e: -e["year"])
honor_roll = sorted(honor.values(), key=lambda h: -h["year"])

# ── 3. Franchises ─────────────────────────────────────────────────────────────
franchises = []
seasons_out = {}
for canon, rows in seasons.items():
    rows.sort(key=lambda s: s["year"])
    sl = slugify(canon)
    w = sum(s["w"] for s in rows); l = sum(s["l"] for s in rows); t = sum(s["t"] for s in rows)
    gp = w + l + t
    finals = gc_finals.get(sl, [])
    title_years = sorted(e["year"] for e in finals if e["result"] == "W")
    gcf_years   = sorted(e["year"] for e in finals)
    first_y, last_y = rows[0]["year"], rows[-1]["year"]
    eras = [e for e in sorted(era_names[canon]) if e and e != canon]
    franchises.append({
        "slug": sl, "name": canon, "metro_slug": CFL_METRO.get(canon),
        "active": last_y >= latest, "first_year": first_y, "last_year": last_y,
        "seasons": len(rows), "w": w, "l": l, "t": t,
        "win_pct": round((w + 0.5 * t) / gp, 3) if gp else 0.0,
        "playoff_apps": sum(1 for s in rows if s.get("play_app")),
        "grey_cups": len(title_years), "gc_finals": len(finals),
        "title_years": title_years, "gc_final_years": gcf_years,
        "aka": eras,
        "divisions": sorted({s["division"] for s in rows if s["division"]}),
    })
    seasons_out[sl] = rows
franchises.sort(key=lambda f: (-f["grey_cups"], -f["win_pct"], f["name"]))

data = {
    "meta": {
        "league": "CFL", "abbr": "CFL", "sport": "Canadian Football",
        "founded": min((f["first_year"] for f in franchises), default=0),
        "latest_season": latest,
        "total_seasons": latest - min((f["first_year"] for f in franchises), default=latest) + 1,
        "active_teams": sum(1 for f in franchises if f["active"]),
        "grey_cup_games": len(honor_roll),
        "grey_cup_first_year": min((h["year"] for h in honor_roll), default=0),
    },
    "franchises": franchises,
    "seasons_by_team": {sl: seasons_out[sl] for sl in seasons_out},
    "grey_cup_finals_by_team": {sl: gc_finals[sl] for sl in gc_finals},
    "grey_cups": honor_roll,
}

os.makedirs(os.path.dirname(OUT), exist_ok=True)
with open(OUT, "w", encoding="utf-8") as f:
    json.dump(data, f, ensure_ascii=False, indent=0)
print(f"cfl/data.json: {len(franchises)} franchises, {len(honor_roll)} Grey Cup games, "
      f"{sum(len(v) for v in seasons_out.values())} team-seasons | active {data['meta']['active_teams']}")
if unresolved:
    print("Grey Cup names with no CFL franchise (honor-roll only, pre-CFL/amateur):",
          sorted(unresolved))

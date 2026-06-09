#!/usr/bin/env python3
"""
build-afl-nrl-data.py
Builds public/data/afl/data.json and public/data/nrl/data.json from:
  - OtherLeagues.xlsx  sheets: "AFL-NRL Ladders" (all-time), "AFL-NRL Grand Finals"
  - MetroAreas.xlsx    sheet: "Team List" (canonical current team names + metro/state/QID)

Canonical grouping is the workbook's "Name" column (relocations folded, e.g. South
Melbourne->Sydney Swans; mergers kept separate, e.g. Balmain / Western Suburbs vs
Wests Tigers). Current clubs are named per the Team List; defunct clubs keep their
canonical Name. Output shape mirrors lib/cfl.ts's CflData.

Usage: build_afl_nrl_data.py <OtherLeagues.xlsx> <MetroAreas.xlsx> <out_dir>
"""
import json, re, sys, unicodedata
from collections import defaultdict
import openpyxl

def slugify(s):
    s = unicodedata.normalize("NFKD", str(s).lower())
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"^-+|-+$", "", re.sub(r"[^a-z0-9]+", "-", s))

def S(v): return str(v).strip() if v is not None else ""
def I(v):
    try: return int(float(v))
    except (TypeError, ValueError): return None
def YN(v): return S(v).upper() == "Y"

# Brand colors [primary, secondary] for known clubs; others fall back to a hash hue.
COLORS = {
  # AFL current
  "Adelaide":["#002b5c","#ffd200"],"Brisbane Lions":["#a30046","#ffd200"],
  "Carlton":["#002a5c","#ffffff"],"Collingwood":["#000000","#ffffff"],
  "Essendon":["#cc2031","#000000"],"Fremantle":["#2a0d54","#ffffff"],
  "Geelong":["#002b5c","#ffffff"],"Gold Coast":["#d4002a","#fcb826"],
  "Greater Western Sydney":["#f47920","#2b2b2b"],"Hawthorn":["#4d2004","#fbbf15"],
  "Melbourne":["#0f1131","#cc2031"],"North Melbourne":["#14437f","#ffffff"],
  "Port Adelaide":["#008aab","#000000"],"Richmond":["#ffd200","#000000"],
  "St Kilda":["#ed0f05","#000000"],"Sydney Swans":["#ed171f","#ffffff"],
  "West Coast":["#062f6e","#f2a900"],"Western Bulldogs":["#003478","#e1251b"],
  # AFL defunct
  "Fitzroy":["#862633","#ffd200"],"Brisbane Bears":["#7a1f3d","#f2a900"],
  "University":["#1c3f94","#000000"],
  # NRL current
  "Brisbane Broncos":["#6c1d45","#fcb813"],"Canberra Raiders":["#9bd732","#0b3d1f"],
  "Canterbury-Bankstown":["#00509f","#ffffff"],"Cronulla-Sutherland":["#00a9c7","#000000"],
  "Dolphins":["#dc0029","#ffd200"],"Gold Coast Titans":["#009fda","#fcb813"],
  "Manly Warringah":["#6e1a32","#ffffff"],"Melbourne Storm":["#4b2e83","#fcb813"],
  "Newcastle Knights":["#003d78","#ee3124"],"North Queensland":["#002b5c","#ffd200"],
  "Parramatta":["#006eb5","#ffd200"],"Penrith":["#000000","#00a99d"],
  "South Sydney":["#ce1141","#006a4d"],"St George Illawarra":["#e1342f","#ffffff"],
  "Sydney Roosters":["#00254c","#e82c2a"],"Wests Tigers":["#000000","#f68a1f"],
  "New Zealand Warriors":["#0a0a0a","#1f51a3"],
  # NRL defunct / historical
  "North Sydney":["#e62a32","#000000"],"Balmain":["#000000","#f57f29"],
  "Western Suburbs":["#000000","#ffffff"],"Newtown":["#1d4e9e","#ffffff"],
  "St George":["#e1342f","#ffffff"],"Illawarra":["#e1251b","#1f7a3d"],
  "Glebe":["#7c1c2c","#ffffff"],"Sydney University":["#1c3f94","#ffd200"],
  "South Queensland":["#5b2a86","#1f9e5a"],"Western Reds":["#c8102e","#000000"],
  "Adelaide Rams":["#c8102e","#000000"],"Gold Coast Chargers":["#e21937","#fcb813"],
  "Northern Eagles":["#6e1a32","#1d4e9e"],
}
def hue_color(slug):
    h = 0
    for ch in slug: h = (h*31 + ord(ch)) & 0xffffffff
    return "hsl(%d 55%% 32%%)" % (h % 360)

def abbr_for(name):
    words = re.sub(r"[^A-Za-z0-9 ]","",name).split()
    if len(words) == 1: return words[0][:3].upper()
    return "".join(w[0] for w in words[:3]).upper()

LEAGUE_META = {
  "afl": {"sport":"Aussie Rules","league":"AFL","abbr":"AFL","title":"premiership"},
  "nrl": {"sport":"Rugby League","league":"NRL","abbr":"NRL","title":"premiership"},
}
SPORT_TO_LEAGUE = {"Aussie Rules":"afl","Rugby League":"nrl"}
# Premierships officially stripped/vacated (excluded from titles, flagged in UI).
STRIPPED = {("nrl", "Melbourne Storm", 2007), ("nrl", "Melbourne Storm", 2009)}

def load_team_list(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Team List"]
    cur = {}  # (league, name) -> {metro, state}
    for r in ws.iter_rows(min_row=2, values_only=True):
        lg = S(r[1])
        if lg in ("AFL","NRL"):
            cur[(lg, S(r[2]))] = {"metro": slugify(S(r[6])), "metro_name": S(r[6]),
                                  "state": S(r[7]), "qid": S(r[16]) or None}
    return cur

def main():
    other, metro_xlsx, outdir = sys.argv[1], sys.argv[2], sys.argv[3]
    team_list = load_team_list(metro_xlsx)
    wb = openpyxl.load_workbook(other, read_only=True, data_only=True)

    # ---- Ladders ----
    ws = wb["AFL-NRL Ladders"]
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    ix = {h: i for i, h in enumerate(hdr)}
    seasons = defaultdict(list)   # (lg, canon) -> [season dict]
    for r in ws.iter_rows(min_row=2, values_only=True):
        sport = S(r[ix["Sport"]])
        lg = SPORT_TO_LEAGUE.get(sport)
        if not lg: continue
        canon = S(r[ix["Name"]]) or S(r[ix["Team"]])
        _strip = (lg, canon, I(r[ix["Season"]])) in STRIPPED
        seasons[(lg, canon)].append({
            "year": I(r[ix["Season"]]), "league": S(r[ix["League"]]),
            "team": S(r[ix["Team"]]), "rank": I(r[ix["Rank"]]),
            "played": I(r[ix["Played"]]), "w": I(r[ix["Wins"]]),
            "d": I(r[ix["Draws"]]), "l": I(r[ix["Losses"]]),
            "pts": I(r[ix["PremiershipPoints"]]),
            "pf": I(r[ix["PointsFor"]]), "pa": I(r[ix["PointsAgainst"]]),
            "minor": YN(r[ix["Minor Prem"]]), "finals": YN(r[ix["Finals"]]),
            "gf": YN(r[ix["Grand Final App"]]), "prem": YN(r[ix["Premiership"]]) and not _strip, "stripped": _strip,
            "metro": slugify(S(r[ix["Metro Area"]])) or None,
            "state": S(r[ix["State"]]) or None,
        })

    # ---- Grand Finals ----
    ws = wb["AFL-NRL Grand Finals"]
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    gx = {h: i for i, h in enumerate(hdr)}
    gfs = defaultdict(list)  # (lg, canon) -> [gf dict]
    for r in ws.iter_rows(min_row=2, values_only=True):
        sport = S(r[gx["Sport"]])
        lg = SPORT_TO_LEAGUE.get(sport)
        if not lg: continue
        canon = S(r[gx["Name"]]) or S(r[gx["Team"]])
        _gstrip = (lg, canon, I(r[gx["Year  "]])) in STRIPPED
        gfs[(lg, canon)].append({
            "year": I(r[gx["Year  "]]), "date": S(r[gx["Date (YYYYMMDD)"]]),
            "result": S(r[gx["W/L"]]), "team": S(r[gx["Team"]]), "opp_team": S(r[gx["Opp Team"]]), "opponent": S(r[gx["Opponent"]]) or S(r[gx["Opp Team"]]),
            "pf": I(r[gx["For"]]), "pa": I(r[gx["Ag"]]),
            "stadium": S(r[gx["Stadium"]]), "metro": slugify(S(r[gx["Metro Area"]])) or None,
            "state": S(r[gx["State"]]) or None,
            "premiership": YN(r[gx["Premiership won"]]) and not _gstrip, "stripped": _gstrip,
        })

    for lg in ("afl", "nrl"):
        meta = LEAGUE_META[lg]
        franchises = []
        teams = sorted({k[1] for k in seasons if k[0] == lg})
        all_years = []
        for canon in teams:
            rows = sorted(seasons[(lg, canon)], key=lambda s: s["year"])
            years = [s["year"] for s in rows]
            all_years += years
            first_y, last_y = min(years), max(years)
            latest_overall = None  # set after loop; compute active below
            w = sum(s["w"] or 0 for s in rows)
            d = sum(s["d"] or 0 for s in rows)
            l = sum(s["l"] or 0 for s in rows)
            gp = w + d + l
            title_years = [s["year"] for s in rows if s["prem"]]
            stripped_years = [s["year"] for s in rows if s.get("stripped")]
            minor_years = [s["year"] for s in rows if s["minor"]]
            gf_years = [s["year"] for s in rows if s["gf"]]
            finals_apps = sum(1 for s in rows if s["finals"])
            eras = sorted({s["team"] for s in rows if s["team"] != canon})
            leagues = sorted({s["league"] for s in rows if s["league"]})
            # canonical display name + metro: Team List for current clubs
            tl = team_list.get((meta["league"], canon))
            # current-team alias: Team List uses "Sydney" for the AFL "Sydney Swans"
            if not tl and canon == "Sydney Swans":
                tl = team_list.get(("AFL", "Sydney"))
            metro_slug = tl["metro"] if tl else (rows[-1]["metro"])
            state = tl["state"] if tl else (rows[-1]["state"])
            qid = tl["qid"] if tl else None
            franchises.append({
                "slug": slugify(canon), "name": canon,
                "metro_slug": metro_slug, "state": state, "qid": qid,
                "active": None,  # filled after we know latest season
                "first_year": first_y, "last_year": last_y, "seasons": len(rows),
                "w": w, "d": d, "l": l, "win_pct": round((w + 0.5*d)/gp, 3) if gp else 0.0,
                "premierships": len(title_years), "minor_premierships": len(minor_years),
                "stripped_premierships": len(stripped_years), "stripped_years": stripped_years,
                "gf_apps": len(gf_years), "finals_apps": finals_apps,
                "title_years": title_years, "minor_years": minor_years, "gf_years": gf_years,
                "aka": eras, "leagues": leagues,
                "color": (COLORS.get(canon, [None])[0] or hue_color(slugify(canon))),
                "color2": (COLORS.get(canon, [None, None])[1] if canon in COLORS else "#ffffff"),
                "abbr": abbr_for(canon),
            })
        latest = max(all_years)
        for f in franchises:
            f["active"] = f["last_year"] >= latest
        franchises.sort(key=lambda f: (-f["premierships"], -f["win_pct"], f["name"]))

        seasons_by_team = {slugify(c): sorted(seasons[(lg, c)], key=lambda s: -s["year"]) for c in teams}
        gf_by_team = {slugify(c): sorted(gfs.get((lg, c), []), key=lambda g: -(g["year"] or 0)) for c in teams}
        active_n = sum(1 for f in franchises if f["active"])
        data = {
            "meta": {**meta, "founded": min(all_years), "latest_season": latest,
                     "total_seasons": latest - min(all_years) + 1,
                     "active_teams": active_n, "total_teams": len(franchises),
                     "source": "afltables.com", "source_url": "https://afltables.com/"},
            "franchises": franchises,
            "seasons_by_team": seasons_by_team,
            "grand_finals_by_team": gf_by_team,
        }
        import os
        d = os.path.join(outdir, lg); os.makedirs(d, exist_ok=True)
        with open(os.path.join(d, "data.json"), "w", encoding="utf-8") as fh:
            json.dump(data, fh, ensure_ascii=False, indent=0)
        prem_total = sum(f["premierships"] for f in franchises)
        print(f"[{lg}] {len(franchises)} franchises ({active_n} active), seasons {min(all_years)}-{latest}, "
              f"{prem_total} premierships total -> {os.path.join(d,'data.json')}")
        top = sorted(franchises, key=lambda f:-f['premierships'])[:3]
        for f in top: print(f"    {f['name']}: prem={f['premierships']} minor={f['minor_premierships']} seas={f['seasons']} active={f['active']} metro={f['metro_slug']}")

if __name__ == "__main__":
    main()

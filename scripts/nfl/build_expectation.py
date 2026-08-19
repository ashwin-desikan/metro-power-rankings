"""Build the NFL expectation ledger from NFL_all.xlsx.

WHAT THIS IS. The workbook carries `ELO Prob (Pre)` on all 106 seasons: a
pre-game win probability for essentially every NFL game since 1920, plus the
closing `Spread (Pre-Game)` from 1966. This emits that in the same shape the
live `nfl-predictions.json` ledger uses, so ONE reader serves both the current
week and the whole century, and the Brier score the picks game gives a user sits
on the same axis as every season ever played.

OUTPUT (public/data/nfl/expectation/)
  index.json        meta, per-season summary, all-time boards, metro rollup
  teams.json        one row per team-season: expected wins vs actual
  season-YYYY.json  per-game ledger for that season

CONVENTIONS ESTABLISHED BY MEASUREMENT, NOT ASSUMPTION (2026-08-19):
  * `ELO Prob (Pre)` is the ROW TEAM's own win probability. Paired rows sum to
    1.0 on all but 23 of 18,188 games; those 23 are excluded from scoring and
    counted in meta.excluded.
  * `Spread (Pre-Game)` is POSITIVE FOR THE FAVOURITE, the opposite of the usual
    convention. Verified on 11,391 graded rows: win rate runs 13.5% at -14,
    through 50.0% at 0, to 86.5% at +14, monotonic throughout.
  * `Name` is the franchise key. All 78 distinct values resolve: 32 in
    franchises.json, 46 in historical.json, no remainder.
  * A tie scores 0.5, the standard Elo convention.

DISPLAY vs IDENTITY. `City` + `Team` is the ERA name (Decatur Staleys) and
`Name` is the franchise (Bears). Rows carry both: print the era name, link and
aggregate on the franchise. Same rule as the champions rolls.

DUPLICATE HEADER NAMES. `W/L/T`, `PF`, `PA` and `H/A` each appear TWICE in the
header row; the first is the game value and the second a season-to-date
restatement. This reads the FIRST deliberately and asserts the duplicate set has
not changed.
"""
import datetime as _dt
import io
import json
import math
import os
import sys
from collections import defaultdict

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
SRC = os.path.join(os.path.expanduser("~"), "OneDrive", "Excel Files", "NFL_all.xlsx")
OUTDIR = os.path.join(ROOT, "public", "data", "nfl", "expectation")
FRANCHISES = os.path.join(ROOT, "public", "data", "nfl", "franchises.json")
HISTORICAL = os.path.join(ROOT, "public", "data", "nfl", "historical.json")

EXPECTED_DUPES = ["H/A", "PA", "PF", "W/L/T"]


def load_json(p):
    d = json.load(io.open(p, encoding="utf-8"))
    return d if isinstance(d, list) else list(d.values())[0]


def build_team_index():
    out = {}
    for f in load_json(FRANCHISES):
        out[f["canonical"]] = {
            "slug": f["slug"], "display": f["name"], "metro": f.get("metro"),
            "metro_slug": f.get("metro_slug"), "active": True,
        }
    for h in load_json(HISTORICAL):
        out.setdefault(h["canonical"], {
            "slug": None, "display": h.get("display_name") or h["canonical"],
            "metro": h.get("metro"), "metro_slug": h.get("metro_slug"), "active": False,
        })
    return out


def header_index(hdr):
    idx, dupes = {}, []
    for i, h in enumerate(hdr):
        if h is None or h == "":
            continue
        if h in idx:
            dupes.append(h)
        else:
            idx[h] = i
    return idx, sorted(set(dupes))


def norm_cdf(x):
    return 0.5 * (1.0 + math.erf(x / math.sqrt(2.0)))


def fit_spread_sigma(samples):
    """MLE for sigma in p(win) = Phi(spread / sigma). Fitted on our own data so
    the market-implied probability is defensible on the page rather than an
    imported constant. Ternary search on a unimodal log-likelihood."""
    lo, hi = 6.0, 24.0
    for _ in range(60):
        m1, m2 = lo + (hi - lo) / 3.0, hi - (hi - lo) / 3.0
        lls = []
        for s in (m1, m2):
            tot = 0.0
            for spread, actual in samples:
                p = min(max(norm_cdf(spread / s), 1e-9), 1 - 1e-9)
                tot += actual * math.log(p) + (1 - actual) * math.log(1 - p)
            lls.append(tot)
        if lls[0] < lls[1]:
            lo = m1
        else:
            hi = m2
    return (lo + hi) / 2.0


def outcome_value(res):
    return {"W": 1.0, "L": 0.0, "T": 0.5}.get(res)


def fnum(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def read_rows():
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    ws = wb["Regular Season"]
    it = ws.iter_rows(values_only=True)
    hdr = list(next(it))
    idx, dupes = header_index(hdr)
    if dupes != EXPECTED_DUPES:
        raise SystemExit("header duplicate set changed: %s (expected %s)" % (dupes, EXPECTED_DUPES))

    def g(row, col):
        i = idx.get(col)
        return None if i is None else row[i]

    games = defaultdict(list)
    for row in it:
        gid = str(g(row, "GameID") or "")
        name = str(g(row, "Name") or "").strip()
        opp = str(g(row, "Opponent") or "").strip()
        if not gid or not name or not opp:
            continue
        try:
            season = int(str(g(row, "NFL Season"))[:4])
        except (TypeError, ValueError):
            continue
        d = g(row, "Date")
        d = d.date().isoformat() if hasattr(d, "date") else (str(d)[:10] if d else None)
        # GameID is per ROW ("35820BroncosPackers" / "35820PackersBroncos"); the
        # leading digits are shared by both rows of a game. Pair on date + the
        # sorted FRANCHISE pair, which yields 18,195 keys all with exactly two
        # rows and no exceptions. Pairing on the era `Team` name instead collides
        # (there were two different Bulldogs), which is why `Name` is used.
        serial = ""
        for ch in gid:
            if ch.isdigit():
                serial += ch
            else:
                break
        games[(d,) + tuple(sorted([name, opp]))].append({
            "game_id": serial or gid,
            "season": season,
            "week": g(row, "Week #"),
            "date": d,
            "playoff": str(g(row, "Reg/Play") or "").strip() == "Playoff",
            "round": str(g(row, "Play. Type") or "").strip() or None,
            "name": name,
            "era_city": str(g(row, "City") or "").strip(),
            "era_team": str(g(row, "Team") or "").strip(),
            "ha": str(g(row, "H/A") or "").strip(),
            "res": str(g(row, "W/L/T") or "").strip(),
            "pf": fnum(g(row, "PF")),
            "pa": fnum(g(row, "PA")),
            "p": fnum(g(row, "ELO Prob (Pre)")),
            "spread": fnum(g(row, "Spread (Pre-Game)")),
            "elo_pre": fnum(g(row, "ELO - Pre")),
            "elo_shift": fnum(g(row, "ELO Shift")),
            "venue": str(g(row, "Stadium") or "").strip(),
            "metro": str(g(row, "Stadium Area") or "").strip(),
            "state": str(g(row, "Stad. State") or "").strip(),
            "qb": str(g(row, "Starting QB") or "").strip() or None,
        })
    wb.close()
    return games


def read_year_by_year():
    """Per team-season: W/L/T for reconciliation, AND the ERA identity.

    🔴 THE ERA NAME AND THE ERA METRO BOTH LIVE HERE, and both must be used.
    `City` + `Team` is what the club was called that year and `Metro Area` is
    where it actually played. Taking either from the current franchise record
    instead prints "1994 Tennessee Titans, Nashville" for a season the Houston
    Oilers played in Houston, and — far worse on a site about place — credits
    Houston's wins to Nashville in the metro rollup. Same rule as the champions
    rolls: display the era, aggregate on the franchise.
    """
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    ws = wb["Year by Year"]
    it = ws.iter_rows(values_only=True)
    idx, _ = header_index(list(next(it)))
    out = {}
    for row in it:
        try:
            y = int(str(row[idx["Year"]])[:4])
        except (TypeError, ValueError, KeyError):
            continue
        nm = str(row[idx["Name"]] or "").strip() if "Name" in idx else ""
        if not nm:
            continue
        city = str(row[idx["City"]] or "").strip() if "City" in idx else ""
        team = str(row[idx["Team"]] or "").strip() if "Team" in idx else ""
        metro = str(row[idx["Metro Area"]] or "").strip() if "Metro Area" in idx else ""
        out[(y, nm)] = {
            "w": fnum(row[idx["Wins"]]) or 0.0,
            "l": fnum(row[idx["Losses"]]) or 0.0,
            "t": fnum(row[idx["Ties"]]) or 0.0,
            "era_name": (city + " " + team).strip(),
            "era_metro": metro,
        }
    wb.close()
    return out


def metro_slugs():
    """Metro display name -> slug, from the workbook-derived metro index."""
    p = os.path.join(ROOT, "public", "data", "metros.json")
    rows = load_json(p)
    return {r["name"]: r["slug"] for r in rows if r.get("name") and r.get("slug")}


def orient(rows):
    """Return (home_row, away_row, neutral). 'vs' is home, 'at' is away."""
    a, b = rows
    if a["ha"] == "vs" and b["ha"] == "at":
        return a, b, False
    if b["ha"] == "vs" and a["ha"] == "at":
        return b, a, False
    return a, b, True


def main():
    teams = build_team_index()
    games = read_rows()
    print("game ids read: %d" % len(games))

    unpaired = [k for k, v in games.items() if len(v) != 2]
    print("unpaired game ids (skipped): %d" % len(unpaired))

    # 🔴 PER-SEASON SPREAD ORIENTATION GATE.
    #
    # `Spread (Pre-Game)` is positive for the favourite, so the favourite must
    # win comfortably more than half the time. Measured across 1966-2022 that
    # rate never leaves the 60-74% band. In 2023 it reads 43.5% and in 2024
    # 47.1%, BELOW even money, which is the signature of the sign being flipped
    # for those two seasons in the source workbook.
    #
    # The build REFUSES those seasons rather than flipping them back. Silently
    # correcting a source defect here would leave the workbook wrong for every
    # other consumer and hide the problem. The seasons are dropped from the
    # market layer, counted in meta, and printed loudly for Ashwin to fix at
    # source. Raise the gate, never widen it to swallow a bad season.
    ORIENT_MIN = 0.55
    fav = defaultdict(lambda: [0, 0])
    for rows in games.values():
        for r in rows:
            o = outcome_value(r["res"])
            if r["spread"] in (None, 0) or o is None or o == 0.5:
                continue
            if r["spread"] > 0:
                fav[r["season"]][1] += 1
                fav[r["season"]][0] += int(o == 1.0)
    # Three outcomes, all recorded. A season is only trusted if it has enough
    # graded games to test the orientation AND passes the test. 1966-1978 carry
    # a spread on a handful of games each (the Super Bowls), not on full
    # seasons, so they are too sparse to orient and are reported as such rather
    # than quietly folded in.
    # 🔴 A FAILING SEASON IS NOT ALWAYS A FLIPPED SEASON, and the two want
    # different actions. Below 50% means the sign is genuinely reversed and the
    # SOURCE needs fixing. Between 50% and the threshold means the direction is
    # right but the sample is too thin to confirm, and telling someone to "fix
    # the sign" there would send them to flip a column that is already correct.
    # 2024 sat in the second bucket the moment 2023's real flip was repaired.
    market_ok, market_bad, market_weak, market_sparse = set(), [], [], []
    for season, (w, n) in sorted(fav.items()):
        if n < 30:
            market_sparse.append({"season": season, "graded": n})
            continue
        rate = w / n
        entry = {"season": season, "favourite_win_rate": round(rate, 3), "graded": n}
        if rate >= ORIENT_MIN:
            market_ok.add(season)
        elif rate < 0.50:
            market_bad.append(entry)
        else:
            market_weak.append(entry)
    if market_sparse:
        print("market layer: %d seasons too sparse to orient (%s)"
              % (len(market_sparse),
                 ", ".join(str(s["season"]) for s in market_sparse[:15])))
    if market_weak:
        print("market layer: %d season(s) point the right way but too weakly to confirm "
              "— held out, NOT a sign problem:" % len(market_weak))
        for b in market_weak:
            print("   %d: favourite won %.1f%% of %d graded games (needs >=%.0f%%)"
                  % (b["season"], 100 * b["favourite_win_rate"], b["graded"], 100 * ORIENT_MIN))
    if market_bad:
        print("\n🔴 SPREAD SIGN LOOKS REVERSED — market layer dropped for these seasons:")
        for b in market_bad:
            print("   %d: favourite won %.1f%% of %d graded games, below even money"
                  % (b["season"], 100 * b["favourite_win_rate"], b["graded"]))
        print("   Fix the sign at source: python scripts/nfl/fix_spread_sign.py --write\n")

    sigma_samples = []
    for rows in games.values():
        if len(rows) != 2:
            continue
        for r in rows:
            o = outcome_value(r["res"])
            if r["spread"] is not None and o is not None and o != 0.5 and r["season"] in market_ok:
                sigma_samples.append((r["spread"], o))
    sigma = fit_spread_sigma(sigma_samples)
    print("market sigma fitted on %d rows from %d trusted seasons: %.3f points"
          % (len(sigma_samples), len(market_ok), sigma))

    by_season = defaultdict(list)
    excluded_prob = 0
    unknown_names = set()
    for gkey in sorted(games, key=lambda k: tuple(str(x) for x in k)):
        rows = games[gkey]
        if len(rows) != 2:
            continue
        h, a, neutral = orient(rows)
        for r in (h, a):
            if r["name"] not in teams:
                unknown_names.add(r["name"])
        pH = h["p"]
        if pH is None or a["p"] is None or abs(pH + a["p"] - 1.0) > 0.02:
            pH = None
            excluded_prob += 1
        res = {"W": "H", "L": "A", "T": "T"}.get(h["res"])
        actual = outcome_value(h["res"])
        ti = teams.get(h["name"], {})
        ta = teams.get(a["name"], {})
        row = {
            "game_id": h["game_id"],
            "season": h["season"],
            "week": h["week"],
            "date": h["date"],
            "playoff": h["playoff"],
            "home": ti.get("display") or h["name"],
            "away": ta.get("display") or a["name"],
            "home_slug": ti.get("slug"),
            "away_slug": ta.get("slug"),
            "home_era": (h["era_city"] + " " + h["era_team"]).strip(),
            "away_era": (a["era_city"] + " " + a["era_team"]).strip(),
            "home_key": h["name"],
            "away_key": a["name"],
            "venue": h["venue"] or None,
            "metro": h["metro"] or None,
            "neutral": neutral,
        }
        if h["round"]:
            row["round"] = h["round"]
        if res:
            row["result"] = res
        if h["pf"] is not None and h["pa"] is not None:
            row["score"] = "%d-%d" % (int(h["pf"]), int(h["pa"]))
        if pH is not None:
            row["model"] = {"pH": round(pH, 4)}
            if actual is not None:
                row["model_brier"] = round((pH - actual) ** 2, 4)
                row["surprise"] = round(1.0 - (pH if actual == 1.0 else (1 - pH) if actual == 0.0 else 0.5), 4)
        if h["spread"] is not None and h["season"] in market_ok:
            mp = norm_cdf(h["spread"] / sigma)
            row["market"] = {"spread": h["spread"], "pH": round(mp, 4)}
            if actual is not None:
                row["market_brier"] = round((mp - actual) ** 2, 4)
        if h["elo_shift"] is not None:
            row["elo_shift"] = round(h["elo_shift"], 2)
        if h["qb"] or a["qb"]:
            row["qb"] = {"home": h["qb"], "away": a["qb"]}
        by_season[h["season"]].append(row)

    if unknown_names:
        raise SystemExit("unresolved franchise names: %s" % sorted(unknown_names))
    print("games with unusable probability (excluded from scoring): %d" % excluded_prob)

    os.makedirs(OUTDIR, exist_ok=True)
    for season, rows in by_season.items():
        rows.sort(key=lambda r: (r["date"] or "", r["game_id"]))
        json.dump({"season": season, "games": rows},
                  io.open(os.path.join(OUTDIR, "season-%d.json" % season), "w",
                          encoding="utf-8", newline=""),
                  separators=(",", ":"), ensure_ascii=False)

    # --- team-season expectation -------------------------------------------
    agg = defaultdict(lambda: {"exp": 0.0, "act": 0.0, "n": 0, "scored": 0,
                               "pexp": 0.0, "pact": 0.0, "pn": 0})
    for season, rows in by_season.items():
        for r in rows:
            for side in ("home", "away"):
                key = r[side + "_key"]
                p = r.get("model", {}).get("pH")
                if p is not None and side == "away":
                    p = 1.0 - p
                res = r.get("result")
                if res is None:
                    continue
                won = 1.0 if res == ("H" if side == "home" else "A") else (0.5 if res == "T" else 0.0)
                a = agg[(season, key)]
                bucket = "p" if r["playoff"] else ""
                if r["playoff"]:
                    a["pn"] += 1
                    a["pact"] += won
                    if p is not None:
                        a["pexp"] += p
                else:
                    a["n"] += 1
                    a["act"] += won
                    if p is not None:
                        a["exp"] += p
                        a["scored"] += 1

    yby = read_year_by_year()
    mslug = metro_slugs()
    mismatch, unmatched_era, unslugged = [], [], set()
    team_rows = []
    for (season, key), a in sorted(agg.items()):
        ti = teams.get(key, {})
        ref = yby.get((season, key))
        exp = a["exp"] if a["scored"] == a["n"] else None
        if ref is None:
            unmatched_era.append((season, key))
        era_name = (ref or {}).get("era_name") or ti.get("display") or key
        era_metro = (ref or {}).get("era_metro") or ti.get("metro")
        slug = mslug.get(era_metro) if era_metro else None
        if era_metro and not slug:
            unslugged.add(era_metro)
        row = {
            "season": season, "key": key,
            # `team` is the ERA name; `franchise`/`slug` are the identity that
            # owns the record and carries the link.
            "team": era_name,
            "franchise": ti.get("display") or key,
            "slug": ti.get("slug"),
            "metro": era_metro, "metro_slug": slug,
            "games": a["n"], "wins": round(a["act"], 1),
            "exp_wins": round(a["exp"], 2) if exp is not None else None,
            "wae": round(a["act"] - a["exp"], 2) if exp is not None else None,
            "playoff_games": a["pn"], "playoff_wins": round(a["pact"], 1),
        }
        team_rows.append(row)
        if ref:
            ref_pts = ref["w"] + 0.5 * ref["t"]
            if abs(ref_pts - a["act"]) > 0.01:
                mismatch.append((season, key, a["act"], ref_pts))

    print("team-seasons: %d | reconciled against Year by Year: %d | MISMATCH: %d"
          % (len(team_rows), len(team_rows) - len(mismatch), len(mismatch)))
    if unmatched_era:
        print("🔴 %d team-seasons have no Year by Year row, so they fall back to the "
              "CURRENT franchise name and metro: %s"
              % (len(unmatched_era), unmatched_era[:8]))
    if unslugged:
        print("🔴 era metros with no slug in metros.json (will not link): %s"
              % sorted(unslugged))
    for m in mismatch[:15]:
        print("   %s %-22s ledger=%.1f  YearByYear=%.1f" % m)

    json.dump({"rows": team_rows},
              io.open(os.path.join(OUTDIR, "teams.json"), "w", encoding="utf-8", newline=""),
              separators=(",", ":"), ensure_ascii=False)

    # --- boards -------------------------------------------------------------
    allg = [r for rows in by_season.values() for r in rows]
    scored = [r for r in allg if "surprise" in r]

    def side_view(r, winner_side):
        # `score` is stored home-first. Orient it to the winner so the board
        # never prints a winning team next to a losing scoreline.
        sc = r.get("score")
        if sc and winner_side == "A":
            a, b = sc.split("-")
            sc = "%s-%s" % (b, a)
        return {
            "season": r["season"], "date": r["date"], "game_id": r["game_id"],
            "winner": r["home_era"] if winner_side == "H" else r["away_era"],
            "winner_slug": r["home_slug"] if winner_side == "H" else r["away_slug"],
            "loser": r["away_era"] if winner_side == "H" else r["home_era"],
            "loser_slug": r["away_slug"] if winner_side == "H" else r["home_slug"],
            "p_winner": round(r["model"]["pH"] if winner_side == "H" else 1 - r["model"]["pH"], 4),
            "score": sc, "metro": r.get("metro"), "venue": r.get("venue"),
            "playoff": r["playoff"], "round": r.get("round"),
        }

    upsets = sorted([r for r in scored if r.get("result") in ("H", "A")],
                    key=lambda r: -r["surprise"])[:100]
    upset_rows = [side_view(r, r["result"]) for r in upsets]

    seasons_sum = []
    for season in sorted(by_season):
        rows = by_season[season]
        mb = [r["model_brier"] for r in rows if "model_brier" in r]
        kb = [r["market_brier"] for r in rows if "market_brier" in r]
        seasons_sum.append({
            "season": season, "games": len(rows),
            "model_brier": round(sum(mb) / len(mb), 4) if mb else None,
            "market_brier": round(sum(kb) / len(kb), 4) if kb else None,
            "market_games": len(kb),
        })

    metro = defaultdict(lambda: {"games": 0, "exp": 0.0, "act": 0.0, "seasons": set()})
    for r in team_rows:
        if not r["metro"] or r["exp_wins"] is None:
            continue
        m = metro[(r["metro"], r["metro_slug"])]
        m["games"] += r["games"]
        m["exp"] += r["exp_wins"]
        m["act"] += r["wins"]
        m["seasons"].add(r["season"])
    metro_rows = sorted(
        [{"metro": k[0], "metro_slug": k[1], "games": v["games"],
          "wins": round(v["act"], 1), "exp_wins": round(v["exp"], 1),
          "wae": round(v["act"] - v["exp"], 1), "seasons": len(v["seasons"])}
         for k, v in metro.items()],
        key=lambda r: -r["wae"])

    best = sorted([r for r in team_rows if r["wae"] is not None], key=lambda r: -r["wae"])[:50]
    worst = sorted([r for r in team_rows if r["wae"] is not None], key=lambda r: r["wae"])[:50]

    # Model against market on the games where BOTH have a view, so the two are
    # compared on identical fixtures rather than on different denominators.
    hh = [r for r in allg if "model_brier" in r and "market_brier" in r]
    h2h = {
        "games": len(hh),
        "model_brier": round(sum(r["model_brier"] for r in hh) / len(hh), 4) if hh else None,
        "market_brier": round(sum(r["market_brier"] for r in hh) / len(hh), 4) if hh else None,
        "seasons_model_better": sum(
            1 for s in seasons_sum
            if s["market_games"] and s["model_brier"] is not None
            and s["model_brier"] < s["market_brier"]),
        "seasons_compared": sum(1 for s in seasons_sum if s["market_games"]),
    }

    meta = {
        # Required by the ISR reader in lib/nflExpectation.ts, which compares a
        # remote copy against the build-time one and keeps the newer.
        "generated_at": _dt.date.today().isoformat(),
        "source": "NFL_all.xlsx :: Regular Season",
        "seasons": [min(by_season), max(by_season)],
        "games": len(allg),
        "games_scored": len(scored),
        "excluded_probability": excluded_prob,
        "unpaired_game_ids": len(unpaired),
        "market_seasons": [s["season"] for s in seasons_sum if s["market_games"]][:1]
                          + [s["season"] for s in seasons_sum if s["market_games"]][-1:],
        "market_sigma": round(sigma, 3),
        "market_sigma_fit_rows": len(sigma_samples),
        "market_excluded": market_bad,
        "market_unconfirmed": market_weak,
        "market_too_sparse": market_sparse,
        "head_to_head": h2h,
        "tie_scores": 0.5,
        "notes": "ELO Prob (Pre) is the row team's win probability; Spread (Pre-Game) "
                 "is positive for the favourite. Both verified by measurement.",
    }
    json.dump({"meta": meta, "seasons": seasons_sum, "upsets": upset_rows,
               "best_seasons": best, "worst_seasons": worst, "metros": metro_rows},
              io.open(os.path.join(OUTDIR, "index.json"), "w", encoding="utf-8", newline=""),
              separators=(",", ":"), ensure_ascii=False)

    print("\nwrote %d season files + teams.json + index.json to %s" % (len(by_season), OUTDIR))
    print("meta:", json.dumps(meta))
    print("\nTOP 10 UPSETS BY PRE-GAME PROBABILITY")
    for u in upset_rows[:10]:
        print("  %d %-10s %-28s beat %-28s p=%.3f %-8s %s" % (
            u["season"], u["date"] or "", u["winner"][:28], u["loser"][:28],
            u["p_winner"], u["score"] or "", u["metro"] or ""))
    print("\nBEST 10 TEAM-SEASONS vs EXPECTATION")
    for r in best[:10]:
        print("  %d %-24s %4.1f wins vs %5.2f expected  (+%.2f)  %s"
              % (r["season"], r["team"], r["wins"], r["exp_wins"], r["wae"], r["metro"] or ""))
    print("\nWORST 10 TEAM-SEASONS vs EXPECTATION")
    for r in worst[:10]:
        print("  %d %-24s %4.1f wins vs %5.2f expected  (%.2f)  %s"
              % (r["season"], r["team"], r["wins"], r["exp_wins"], r["wae"], r["metro"] or ""))


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    main()

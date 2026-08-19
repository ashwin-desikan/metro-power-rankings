"""Load 2024 and 2025 closing spreads and totals into NFL_all.xlsx.

SOURCE. `scripts/nfl/sources/covers-nfl-odds-2024-2025.txt`, pasted from
https://www.covers.com/sportsoddshistory/nfl-game-season/?y=2024 (and the 2025
page). 570 games: 272 regular plus 13 playoff per season, both seasons complete.

🔴 TWO CONVENTIONS THAT DISAGREE, AND THE WHOLE JOB IS GETTING THIS RIGHT.
  * covers.com lists the FAVOURITE FIRST and states its spread as a NEGATIVE
    number. All 570 rows are negative, with no pick'em, so team1 is always the
    favourite and there is no ambiguity to resolve.
  * NFL_all.xlsx stores the spread PER TEAM ROW and POSITIVE FOR THE FAVOURITE,
    the opposite sign. Verified across 1979-2022, where the favourite's win rate
    never leaves the 60-74% band.
  So team1's workbook cell = -(file spread) and team2's = +(file spread). The
  over/under is a total, not directional, and is written identically to both.

The build's orientation gate is the backstop: if this load got the sign wrong,
`build_expectation.py` will report the season below even money and refuse it.

Writes surgically into the sheet XML, like scripts/nfl/fix_spread_sign.py.

Run:  python scripts/nfl/load_odds.py            (dry run, reports the match)
      python scripts/nfl/load_odds.py --write
"""
import io
import json
import os
import re
import shutil
import sys
import zipfile
from collections import Counter, defaultdict
from datetime import datetime

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
SRC = os.path.join(os.path.expanduser("~"), "OneDrive", "Excel Files", "NFL_all.xlsx")
# Every covers.com paste in sources/ is read. Add a season by dropping its file
# in and adding the year to SEASONS; nothing else changes.
ODDS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "sources")
ODDS_GLOB = "covers-nfl-odds-*.txt"
FRANCHISES = os.path.join(ROOT, "public", "data", "nfl", "franchises.json")
SHEET = "Regular Season"
SEASONS = {2023, 2024, 2025}
STAMP = "20260819-odds-2023-2025"


def parse_odds():
    import glob
    files = sorted(glob.glob(os.path.join(ODDS_DIR, ODDS_GLOB)))
    if not files:
        raise SystemExit("REFUSING: no source files matching %s in %s" % (ODDS_GLOB, ODDS_DIR))
    print("sources: %s" % [os.path.basename(f) for f in files])
    rows, problems = [], []
    for path in files:
        season = mode = None
        for ln, line in enumerate(io.open(path, encoding="utf-8"), 1):
            rows_before = len(rows)
            line = line.rstrip("\n")
            m = re.match(r"^(\d{4})\s+(Regular Season|Playoffs)", line)
            if m:
                season = int(m.group(1))
                mode = "reg" if m.group(2) == "Regular Season" else "post"
                continue
            f = line.split("\t")
            if len(f) < 11 or season is None:
                continue
            _parse_line(f, mode, season, rows, problems, ln, path)
    return _finish(rows, problems)


def _parse_line(f, mode, season, rows, problems, ln, path):
    """One data line. Regular and playoff rows are both 11 fields, but a playoff
    row spends field 1 on the round name and has no trailing note, so every
    field after it shifts by one."""
    # Regular: Day|Date|Time|H1|Team1|Res|Spread|H2|Team2|OU|Note
    # Playoff: Round|Day|Date|Time|H1|Team1|Res|Spread|H2|Team2|OU
    o = 1 if mode == "post" else 0
    date_s, t1, res, spr, t2, ou = f[1 + o], f[4 + o], f[5 + o], f[6 + o], f[8 + o], f[9 + o]
    if not t1.strip() or not t2.strip():
        return
    try:
        d = datetime.strptime(date_s.strip(), "%b %d, %Y").date().isoformat()
    except ValueError:
        problems.append((os.path.basename(path), ln, "date", date_s))
        return
    # The cover marker is W, L or P (push). Only the number matters here.
    ms = re.search(r"(-?\d+(?:\.\d+)?)", spr)
    mo = re.search(r"(\d+(?:\.\d+)?)", ou)
    mr = re.search(r"([WLT])\s+(\d+)-(\d+)", res)
    if not (ms and mr):
        problems.append((os.path.basename(path), ln, "spread/result", spr + " | " + res))
        return
    rows.append({
        "season": season, "date": d,
        "t1": re.sub(r"\s*\(\d+\)\s*$", "", t1).strip(),
        "t2": re.sub(r"\s*\(\d+\)\s*$", "", t2).strip(),
        "s1": int(mr.group(2)), "s2": int(mr.group(3)),
        "spread_t1": float(ms.group(1)),
        "ou": float(mo.group(1)) if mo else None,
    })


def _finish(rows, problems):
    if problems:
        raise SystemExit("REFUSING: %d unparsed lines: %s" % (len(problems), problems[:5]))
    pos = [r for r in rows if r["spread_t1"] > 0]
    if pos:
        raise SystemExit("REFUSING: %d rows have a POSITIVE spread, so 'favourite first' "
                         "does not hold and the sign rule above is unsafe: %s"
                         % (len(pos), pos[:3]))
    return rows


def name_to_key():
    d = json.load(io.open(FRANCHISES, encoding="utf-8"))
    d = d if isinstance(d, list) else list(d.values())[0]
    return {f["name"]: f["canonical"] for f in d}


def header_index(hdr):
    idx = {}
    for i, h in enumerate(hdr):
        if h is not None and h not in idx:
            idx[h] = i
    return idx


def sheet_xml_path(z, name):
    wb = z.read("xl/workbook.xml").decode("utf-8", "replace")
    rels = z.read("xl/_rels/workbook.xml.rels").decode("utf-8", "replace")
    relmap = dict(re.findall(r'Id="([^"]+)"[^>]*Target="([^"]+)"', rels))
    for nm, rid in re.findall(r'<sheet name="([^"]+)"[^>]*r:id="([^"]+)"', wb):
        if nm == name:
            return "xl/" + relmap[rid].lstrip("/")
    raise SystemExit("sheet %r not found" % name)


def main():
    write = "--write" in sys.argv
    odds = parse_odds()
    print("parsed %d games (%s)" % (len(odds), dict(Counter(r["season"] for r in odds))))
    n2k = name_to_key()
    unknown = {t for r in odds for t in (r["t1"], r["t2"]) if t not in n2k}
    if unknown:
        raise SystemExit("REFUSING: team names not in franchises.json: %s" % sorted(unknown))

    # (season, date, frozenset{keyA,keyB}) -> game
    lookup = {}
    for r in odds:
        k = (r["season"], r["date"], frozenset([n2k[r["t1"]], n2k[r["t2"]]]))
        if k in lookup:
            raise SystemExit("REFUSING: duplicate source game %s" % (k,))
        lookup[k] = r

    wb = openpyxl.load_workbook(SRC, read_only=False, data_only=False)
    ws = wb[SHEET]
    idx = header_index([c.value for c in ws[1]])
    need = ["NFL Season", "Date", "Name", "Opponent", "Spread (Pre-Game)", "Over/Under"]
    miss = [c for c in need if c not in idx]
    if miss:
        raise SystemExit("REFUSING: missing columns %s" % miss)
    scol, ocol = idx["Spread (Pre-Game)"] + 1, idx["Over/Under"] + 1

    plan = {}          # row -> (spread, ou)
    matched, unmatched, agree, disagree = 0, [], 0, []
    for r in range(2, ws.max_row + 1):
        try:
            y = int(str(ws.cell(row=r, column=idx["NFL Season"] + 1).value)[:4])
        except (TypeError, ValueError):
            continue
        if y not in SEASONS:
            continue
        d = ws.cell(row=r, column=idx["Date"] + 1).value
        d = d.date().isoformat() if hasattr(d, "date") else str(d)[:10]
        me = str(ws.cell(row=r, column=idx["Name"] + 1).value or "").strip()
        opp = str(ws.cell(row=r, column=idx["Opponent"] + 1).value or "").strip()
        g = lookup.get((y, d, frozenset([me, opp])))
        if g is None:
            unmatched.append((y, d, me, opp, r))
            continue
        matched += 1
        # sign flip: file is negative-for-favourite, workbook positive.
        mine = -g["spread_t1"] if n2k[g["t1"]] == me else g["spread_t1"]
        cur = ws.cell(row=r, column=scol).value
        if isinstance(cur, (int, float)):
            if abs(float(cur) - mine) < 1e-9:
                agree += 1
            else:
                disagree.append((y, d, me, float(cur), mine))
        plan[r] = (mine, g["ou"])
    wb.close()

    print("workbook rows in %s: matched %d, unmatched %d" % (sorted(SEASONS), matched, len(unmatched)))
    for u in unmatched[:10]:
        print("   UNMATCHED %s %s %s vs %s (row %d)" % u)
    print("existing spread values: %d agree with the source, %d differ" % (agree, len(disagree)))
    for d0 in disagree[:10]:
        print("   DIFFERS %s %s %-14s workbook=%+.1f source=%+.1f" % d0)
    print("cells to write: %d spread + %d over/under"
          % (len(plan), sum(1 for v in plan.values() if v[1] is not None)))
    if not write:
        print("\nDRY RUN. Re-run with --write to apply.")
        return
    if unmatched:
        raise SystemExit("REFUSING to write with %d unmatched rows." % len(unmatched))

    from openpyxl.utils import get_column_letter
    sl, ol = get_column_letter(scol), get_column_letter(ocol)
    with zipfile.ZipFile(SRC) as z:
        names = z.namelist()
        infos = {n: z.getinfo(n) for n in names}
        path = sheet_xml_path(z, SHEET)
        blobs = {n: z.read(n) for n in names}
    xml = blobs[path].decode("utf-8", "replace")

    def fmt(v):
        return ("%d" % v) if float(v) == int(v) else repr(float(v))

    def col_index(ref):
        n = 0
        for ch in re.match(r"[A-Z]+", ref).group(0):
            n = n * 26 + (ord(ch) - 64)
        return n

    # 🔴 EXCEL OMITS EMPTY CELLS ENTIRELY. 2025 has no spread at all and 2024 is
    # half-populated, so 660 of the 1,140 target cells simply do not exist in
    # the XML. A replace-only pass silently covers 480 of them, which is why the
    # count check below is not optional. Missing cells are INSERTED in column
    # order inside their own <row>, and the row's `spans` widened if needed.
    targets = {}
    for row, (sp, ou) in plan.items():
        cells = []
        if sp is not None:
            cells.append((col_index(sl), sl, sp))
        if ou is not None:
            cells.append((col_index(ol), ol, ou))
        targets[row] = sorted(cells)

    starts = [(int(m.group(1)), m.start()) for m in re.finditer(r'<row r="(\d+)"', xml)]
    pos_by_row = {r: i for r, i in starts}
    order = [p for _, p in starts] + [len(xml)]
    next_pos = {starts[i][0]: order[i + 1] for i in range(len(starts))}

    hits = [0, 0]
    pieces, cursor = [], 0
    for row in sorted(targets):
        if row not in pos_by_row:
            raise SystemExit("REFUSING: <row r=\"%d\"> not found in the sheet XML." % row)
        a, b = pos_by_row[row], next_pos[row]
        block = xml[a:b]
        for cidx, letter, val in targets[row]:
            ref = "%s%d" % (letter, row)
            pat = re.compile(r'<c r="%s"([^>]*?)(/>|>.*?</c>)' % ref, re.S)
            m = pat.search(block)
            which = 0 if letter == sl else 1
            if m:
                style = re.search(r'\ss="(\d+)"', m.group(1))
                new = '<c r="%s"%s><v>%s</v></c>' % (
                    ref, (' s="%s"' % style.group(1)) if style else "", fmt(val))
                block = block[:m.start()] + new + block[m.end():]
            else:
                ins = None
                for mc in re.finditer(r'<c r="([A-Z]+%d)"' % row, block):
                    if col_index(mc.group(1)) > cidx:
                        ins = mc.start()
                        break
                if ins is None:
                    ins = block.rfind("</row>")
                    if ins < 0:
                        raise SystemExit("REFUSING: no </row> for row %d." % row)
                block = block[:ins] + '<c r="%s"><v>%s</v></c>' % (ref, fmt(val)) + block[ins:]
            hits[which] += 1
        # widen spans so the new columns are inside the declared range
        sm = re.search(r'(<row r="%d"[^>]*?spans=")(\d+):(\d+)(")' % row, block)
        if sm:
            lo, hi = int(sm.group(2)), int(sm.group(3))
            need_hi = max(hi, max(c[0] for c in targets[row]))
            if need_hi != hi:
                block = block[:sm.start()] + "%s%d:%d%s" % (sm.group(1), lo, need_hi, sm.group(4)) + block[sm.end():]
        pieces.append(xml[cursor:a])
        pieces.append(block)
        cursor = b
    pieces.append(xml[cursor:])
    xml = "".join(pieces)
    print("rewritten: %d spread cells, %d over/under cells" % tuple(hits))
    want_s = len(plan)
    want_o = sum(1 for v in plan.values() if v[1] is not None)
    if hits[0] != want_s or hits[1] != want_o:
        raise SystemExit("REFUSING: rewrote %s, expected (%d, %d). Some target cells are "
                         "absent from the XML (Excel omits empty cells) and need inserting."
                         % (tuple(hits), want_s, want_o))

    backup = SRC + ".bak-" + STAMP
    shutil.copyfile(SRC, backup)
    print("backup:", backup)
    blobs[path] = xml.encode("utf-8")
    tmp = SRC + ".tmp"
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as out:
        for n in names:
            out.writestr(infos[n], blobs[n])
    os.replace(tmp, SRC)
    print("written.")

    with zipfile.ZipFile(SRC) as z:
        assert z.testzip() is None
        assert z.namelist() == names
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    print("verify: opens, %d sheets" % len(wb.sheetnames))
    wb.close()


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    main()

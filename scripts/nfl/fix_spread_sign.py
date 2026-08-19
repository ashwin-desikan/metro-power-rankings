"""Negate `Spread (Pre-Game)` for the 2023 and 2024 NFL seasons in NFL_all.xlsx.

WHY. The column is positive-for-the-favourite everywhere else in the file. The
favourite's win rate sits between 60% and 74% in every season from 1979 to 2022,
then reads 43.5% in 2023 and 47.1% in 2024, below even money. Two checks settled
that this is a sign error and not corruption:

  * The 2023 calibration curve is a clean MIRROR of the reference era
    (-14 -> 88.9%, 0 -> 50.0%, +14 -> 11.1%), monotonic throughout.
  * Every game's two rows still sum to zero: 285 of 285 pairs in 2023 and 240 of
    240 in 2024. The values are attached to the right games; only the sign is
    reversed. 2024 is also incomplete (240 of 285 games carry a spread) and 2025
    carries none, but that is missing data, not wrong data.

METHOD. NFL_all.xlsx is 71 MB across 13 sheets and it is Ashwin's master, so this
is a SURGICAL edit of the sheet XML inside the zip: patch only the target <c>
elements in the Regular Season sheet and write every other zip entry through
byte-for-byte with its original ZipInfo. openpyxl would rewrite the whole file
from its own model and silently drop what that model does not carry. Same recipe
as _set_metro.py; see the surgical-xlsx-cell-edit reference.

Run:  python scripts/nfl/fix_spread_sign.py          (dry run, reports only)
      python scripts/nfl/fix_spread_sign.py --write  (backs up, then edits)
"""
import os
import re
import shutil
import sys
import zipfile

import openpyxl

SRC = os.path.join(os.path.expanduser("~"), "OneDrive", "Excel Files", "NFL_all.xlsx")
SHEET = "Regular Season"
COL = "DW"                      # Spread (Pre-Game)
SEASONS = {2023, 2024}
STAMP = "20260819-spreadsign"


def sheet_xml_path(z, name):
    wb = z.read("xl/workbook.xml").decode("utf-8", "replace")
    rels = z.read("xl/_rels/workbook.xml.rels").decode("utf-8", "replace")
    relmap = dict(re.findall(r'Id="([^"]+)"[^>]*Target="([^"]+)"', rels))
    for nm, rid in re.findall(r'<sheet name="([^"]+)"[^>]*r:id="([^"]+)"', wb):
        if nm == name:
            return "xl/" + relmap[rid].lstrip("/")
    raise SystemExit("sheet %r not found" % name)


def target_rows():
    """Row numbers whose season is in SEASONS and whose spread cell is numeric."""
    wbk = openpyxl.load_workbook(SRC, read_only=False, data_only=False)
    ws = wbk[SHEET]
    hdr = [c.value for c in ws[1]]
    scol = ycol = None
    for i, h in enumerate(hdr):
        if h == "Spread (Pre-Game)" and scol is None:
            scol = i + 1
        if h == "NFL Season" and ycol is None:
            ycol = i + 1
    if scol is None or ycol is None:
        raise SystemExit("could not locate the Spread / NFL Season columns")
    rows, formulas = {}, 0
    for r in range(2, ws.max_row + 1):
        try:
            y = int(str(ws.cell(row=r, column=ycol).value)[:4])
        except (TypeError, ValueError):
            continue
        if y not in SEASONS:
            continue
        c = ws.cell(row=r, column=scol)
        if c.data_type == "f":
            formulas += 1
            continue
        if isinstance(c.value, (int, float)):
            rows[r] = float(c.value)
    wbk.close()
    if formulas:
        raise SystemExit("REFUSING: %d target cells are FORMULAS. A literal write "
                         "would destroy them." % formulas)
    return rows


def patch(xml, rows):
    """Negate <v> inside <c r="DW{row}"> for every row in `rows`."""
    hits = [0]
    pat = re.compile(r'<c r="%s(\d+)"([^>]*)>(\s*)<v>([^<]*)</v>' % COL)

    def sub(m):
        row = int(m.group(1))
        if row not in rows:
            return m.group(0)
        try:
            val = float(m.group(4))
        except ValueError:
            return m.group(0)
        hits[0] += 1
        neg = -val
        txt = ("%d" % neg) if neg == int(neg) else repr(neg)
        return '<c r="%s%d"%s>%s<v>%s</v>' % (COL, row, m.group(2), m.group(3), txt)

    return pat.sub(sub, xml), hits[0]


def main():
    write = "--write" in sys.argv
    rows = target_rows()
    print("target cells (numeric spread, seasons %s): %d"
          % (sorted(SEASONS), len(rows)))

    with zipfile.ZipFile(SRC) as z:
        names = z.namelist()
        infos = {n: z.getinfo(n) for n in names}
        path = sheet_xml_path(z, SHEET)
        print("sheet XML: %s   zip entries: %d" % (path, len(names)))
        blobs = {n: z.read(n) for n in names}

    xml = blobs[path].decode("utf-8", "replace")
    new_xml, hits = patch(xml, rows)
    print("cells rewritten: %d of %d expected" % (hits, len(rows)))
    if hits != len(rows):
        raise SystemExit("REFUSING: rewrote %d cells but expected %d." % (hits, len(rows)))
    if not write:
        print("\nDRY RUN. Re-run with --write to apply.")
        return

    backup = SRC + ".bak-" + STAMP
    shutil.copyfile(SRC, backup)
    print("backup: %s" % backup)

    blobs[path] = new_xml.encode("utf-8")
    tmp = SRC + ".tmp"
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as out:
        for n in names:
            out.writestr(infos[n], blobs[n])
    os.replace(tmp, SRC)
    print("written.")

    # --- verification -----------------------------------------------------
    with zipfile.ZipFile(SRC) as z:
        assert z.testzip() is None, "zip failed testzip()"
        after = z.namelist()
        assert after == names, "zip entry order changed"
        diff = [n for n in names if z.read(n) != blobs[n]]
        assert not diff, "unexpected entry differences: %s" % diff
    with zipfile.ZipFile(backup) as zb:
        changed = [n for n in names if zb.read(n) != blobs[n]]
    print("verify: testzip OK, %d entries in the same order, only %s differs (%s)"
          % (len(after), len(changed), changed))
    wbk = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    print("verify: workbook opens, %d sheets: %s" % (len(wbk.sheetnames), wbk.sheetnames))
    wbk.close()


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    main()

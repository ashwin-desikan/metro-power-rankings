#!/usr/bin/env python3
"""
Build public/data/champions.json from Champions_History.xlsx.

SOURCE OF TRUTH for "current champions" badges + the /sports/champions board.
Reads the Champions sheet and emits all rows where Is Current = "Y".

Supersedes ZoneZero_Champions.xlsx, which can be retired once the merge is done.
Columns used from Champions_History (post-merge schema):
  Sport, Competition, Champion (Canonical), Scope Type, Year,
  Date Awarded, Next Awarded Date, Tier, Is Current

Run after editing the sheet:
    python scripts/build-champions-data.py

Set env var CHAMPS_SRC to override the default path.
"""
import os, json, datetime, re
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC  = os.environ.get("CHAMPS_SRC", os.path.expanduser("~/OneDrive/Excel Files/Champions_History.xlsx"))
OUT  = os.path.join(ROOT, "public", "data", "champions.json")

# Fallback: if Champions_History not yet merged, read legacy ZoneZero file
LEGACY_SRC = os.path.join(ROOT, "ZoneZero_Champions.xlsx")

def to_iso(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return (v.date() if isinstance(v, datetime.datetime) else v).isoformat()
    if isinstance(v, str) and v.strip():
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
            try:
                return datetime.datetime.strptime(v.strip(), fmt).date().isoformat()
            except ValueError:
                pass
    return None

def to_year(v):
    iso = to_iso(v)
    if iso:
        return int(iso[:4])
    try:
        return int(v)
    except (TypeError, ValueError):
        return None

def cell(v):
    return "" if v is None else (v if isinstance(v, (int, float)) else str(v).strip())

def build_from_history(src):
    """Read Champions_History.xlsx (post-merge) filtered to Is Current = Y."""
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    ws = wb["Champions"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(c).strip() if c is not None else "" for c in rows[0]]
    idx = {h.lower(): i for i, h in enumerate(hdr)}

    def col(*names):
        for n in names:
            if n.lower() in idx:
                return idx[n.lower()]
        return None

    iSport    = col("Sport")
    iComp     = col("Competition")
    iCanon    = col("Champion (Canonical)", "Canonical", "Champion")
    iEra      = col("Champion", "Era Name")
    iScope    = col("Scope Type")
    iYear     = col("Year")
    iDate     = col("Date Awarded", "Date")  # "Date" is used post-merge
    iNext     = col("Next Awarded Date")
    iTier     = col("Tier")
    iCurrent  = col("Is Current")
    iIntl     = col("International"); iCont = col("Continental"); iDom = col("Domestic")

    out = []
    missing_current_col = iCurrent is None
    if missing_current_col:
        print("WARNING: 'Is Current' column not found -- run merge-champions-sources.py first.")
        print("         Falling back to most-recent-year-per-comp logic.")

    # Competitions whose reigning champion belongs on the Current board but whose
    # "Is Current" flag is not maintained in Champions_History.xlsx (they dropped
    # off the board when it became the single source). Their latest-year row is
    # treated as current regardless of the flag.
    ALWAYS_CURRENT = {"Champions League", "Club World Cup", "SuperLega",
                      "Europa League", "Europa Conference League"}

    # Latest year per competition (used by the fallback and by ALWAYS_CURRENT).
    comp_max = {}
    for r in rows[1:]:
        def gv(i): return r[i] if i is not None and i < len(r) else None
        c = cell(gv(iComp)); y = to_year(gv(iYear))
        if c and y and (c not in comp_max or y > comp_max[c]):
            comp_max[c] = y

    for r in rows[1:]:
        def g(i): return r[i] if i is not None and i < len(r) else None

        comp = cell(g(iComp))
        if not comp:
            continue

        if not missing_current_col:
            is_cur = str(cell(g(iCurrent))).strip().upper() == "Y"
            latest_always = comp in ALWAYS_CURRENT and to_year(g(iYear)) == comp_max.get(comp)
            if not (is_cur or latest_always):
                continue
        else:
            if comp_max.get(comp) != to_year(g(iYear)):
                continue

        canon = cell(g(iCanon)) if iCanon is not None else ""
        era   = cell(g(iEra))   if iEra   is not None else ""
        team  = canon or era
        if not team:
            continue

        scope_type  = cell(g(iScope)) if iScope is not None else ""
        scope_label = ""
        if iIntl is not None and g(iIntl):
            scope_label = cell(g(iIntl)); scope_type = scope_type or "International"
        elif iCont is not None and g(iCont):
            scope_label = cell(g(iCont)); scope_type = scope_type or "Continental"
        elif iDom is not None and g(iDom):
            scope_label = cell(g(iDom));  scope_type = scope_type or "Domestic"

        date_val = to_iso(g(iDate)) if iDate is not None else None
        next_val = to_iso(g(iNext)) if iNext is not None else None
        tier_raw = g(iTier) if iTier is not None else None
        try:
            tier = int(tier_raw) if tier_raw is not None else None
        except (TypeError, ValueError):
            tier = None

        out.append({
            "sport":           cell(g(iSport)),
            "competition":     comp,
            "team":            team,
            "year":            to_year(date_val) or to_year(g(iYear)),
            "dateAwarded":     date_val,
            "scope":           scope_label,
            "scopeType":       scope_type,
            "nextAwarded":     to_year(next_val) if next_val else None,
            "nextAwardedDate": next_val,
            "tier":            tier,
        })

    return out

def build_from_zone_zero(src):
    """Legacy fallback: read ZoneZero_Champions.xlsx (original schema)."""
    print(f"Reading legacy source: {src}")
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(c).strip() if c is not None else "" for c in rows[0]]
    idx = {h.lower(): i for i, h in enumerate(hdr)}
    def col(*names):
        for n in names:
            if n in idx: return idx[n]
        return None
    iSport = col("sport"); iComp = col("competiton", "competition"); iTeam = col("team")
    iDate = col("date awarded", "date", "year")
    iIntl = col("international"); iCont = col("continental"); iDom = col("domestic")
    iNext = col("next awarded date", "next awarded", "nextawarded", "next")
    iTier = col("my tier rank", "tier", "my tier", "rank")
    out = []
    for r in rows[1:]:
        def g(i): return r[i] if i is not None and i < len(r) else None
        team, comp = g(iTeam), g(iComp)
        if not team or not comp:
            continue
        intl, cont, dom = g(iIntl), g(iCont), g(iDom)
        scope_type = "International" if intl else "Continental" if cont else "Domestic" if dom else None
        out.append({
            "sport":           str(g(iSport) or "").strip(),
            "competition":     str(comp).strip(),
            "team":            str(team).strip(),
            "year":            to_year(g(iDate)),
            "dateAwarded":     to_iso(g(iDate)),
            "scope":           str(intl or cont or dom or "").strip(),
            "scopeType":       scope_type,
            "nextAwarded":     to_year(g(iNext)),
            "nextAwardedDate": to_iso(g(iNext)),
            "tier":            int(g(iTier)) if g(iTier) is not None else None,
        })
    return out

def main():
    if os.path.exists(SRC):
        out = build_from_history(SRC)
    elif os.path.exists(LEGACY_SRC):
        out = build_from_zone_zero(LEGACY_SRC)
    else:
        raise FileNotFoundError(
            f"Neither Champions_History.xlsx ({SRC}) nor ZoneZero_Champions.xlsx ({LEGACY_SRC}) found."
        )

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=0)

    import collections
    tiers = collections.Counter(x["tier"] for x in out)
    print(f"champions.json: {len(out)} current champions")
    print(f"tiers: {dict(sorted(tiers.items(), key=lambda x: (x[0] is None, x[0])))}")
    print(f"with dateAwarded: {sum(1 for x in out if x['dateAwarded'])}  "
          f"with nextDate: {sum(1 for x in out if x['nextAwardedDate'])}")

if __name__ == "__main__":
    main()

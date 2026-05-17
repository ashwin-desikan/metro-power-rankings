#!/usr/bin/env python3
"""
Build unified /sports landing-page data from MetroAreas.xlsx.

Reads:
  Team List           — every team across every league (3795 rows)
  FootballClub_Data   — global football clubs (9854 rows, ~499 Major League)

Writes:
  public/data/sports/all-teams.json  — array of marker rows
  public/data/sports/league-summary.json — per-league counts for the cards

Scope: every Team List and FootballClub_Data row with valid coordinates.
  Each marker carries level='Major' (workbook ml=='Y' or 'Euroleague') or
  level='Other' (everything else). Notable Venues and Historic Venues are
  excluded (inert stadiums, not teams).
  The /sports UI gates the visible set with a preset chip:
    Gold Standard = level='Major' AND isGoldStandardLeague(sport, league)
    Major League  = level='Major'
    Other         = level='Other'
    All           = no level filter

team_page_url resolution:
  - NFL/MLB/NBA: cross-reference public/data/<league>/franchises.json by team name
  - Other leagues: null until per-league pages ship (NHL next)

Per project_sports_home_page_scope.md.

Usage:
  python scripts/build-sports-index.py
"""

import json
import os
import re
import sys
from collections import Counter
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
WORKBOOK = ROOT / "MetroAreas.xlsx"
OUT_DIR = ROOT / "public" / "data" / "sports"
OUT_DIR.mkdir(parents=True, exist_ok=True)

# Country -> ISO2 mini-map for sites we ship. Extends as global coverage grows;
# fallback uses None when the workbook country isn't mapped.
COUNTRY_ISO2 = {
    "United States": "US",
    "Canada": "CA",
    "Mexico": "MX",
    "United Kingdom": "GB",
    "England": "GB",
    "Scotland": "GB",
    "Wales": "GB",
    "Northern Ireland": "GB",
    "Ireland": "IE",
    "France": "FR",
    "Germany": "DE",
    "Spain": "ES",
    "Italy": "IT",
    "Netherlands": "NL",
    "Portugal": "PT",
    "Belgium": "BE",
    "Switzerland": "CH",
    "Austria": "AT",
    "Poland": "PL",
    "Czech Republic": "CZ",
    "Sweden": "SE",
    "Norway": "NO",
    "Denmark": "DK",
    "Finland": "FI",
    "Russia": "RU",
    "Turkey": "TR",
    "Greece": "GR",
    "Brazil": "BR",
    "Argentina": "AR",
    "Mexico": "MX",
    "Chile": "CL",
    "Colombia": "CO",
    "Uruguay": "UY",
    "Peru": "PE",
    "Japan": "JP",
    "South Korea": "KR",
    "China": "CN",
    "Hong Kong": "HK",
    "Taiwan": "TW",
    "Australia": "AU",
    "New Zealand": "NZ",
    "India": "IN",
    "Pakistan": "PK",
    "Bangladesh": "BD",
    "Sri Lanka": "LK",
    "South Africa": "ZA",
    "Egypt": "EG",
    "Morocco": "MA",
    "Algeria": "DZ",
    "Saudi Arabia": "SA",
    "UAE": "AE",
    "United Arab Emirates": "AE",
    "Qatar": "QA",
    "Israel": "IL",
}


# Workbook Level values that are actually league names (the workbook
# inadvertently doubles them up). These belong in the League filter row,
# not in the Level filter row.
_LEVEL_VALUES_THAT_ARE_LEAGUES = {"F1", "NASCAR"}

def _normalize_level(raw):
    """Return the workbook Level column value as a stable string, or None.

    Filters out empty / None / 0 / 'None' / 'none' so the Level filter chip
    row stays clean. Preserves real values: '1', '2', '3', 'College',
    'Junior', 'Independent', 'A' / 'AA' / 'AAA' / 'High-A', etc. Drops the
    league-name values (F1, NASCAR) which should surface only as League chips.
    """
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        if raw == 0:
            return None
        # Drop trailing .0 for ints stored as float.
        if isinstance(raw, float) and raw.is_integer():
            return str(int(raw))
        return str(raw)
    s = str(raw).strip()
    if not s or s.lower() == "none" or s == "0":
        return None
    if s in _LEVEL_VALUES_THAT_ARE_LEAGUES:
        return None
    return s


def slugify(s: str) -> str:
    if not s:
        return ""
    s = s.lower()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    s = s.strip("-")
    return s


def normalize_team_name(s: str) -> str:
    """Lowercase, strip diacritics-free punctuation, used as join key."""
    if not s:
        return ""
    s = s.lower()
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return s.strip()


def load_franchise_slugs() -> dict:
    """Return {league_lower: {normalized_team_name: slug}} for NFL/MLB/NBA.

    Match keys cover several plausible spellings emitted by the workbook
    (display_name, city + team, team alone) since Team List rows use the
    full marketing name ("Buffalo Bills") while franchises.json may carry
    either short ("Bills") or long forms.
    """
    out = {}
    for league in ("nfl", "mlb", "nba"):
        path = ROOT / "public" / "data" / league / "franchises.json"
        if not path.exists():
            out[league] = {}
            continue
        with open(path) as f:
            data = json.load(f)
        index = {}
        for fr in data:
            slug = fr.get("slug")
            if not slug:
                continue
            candidates = [
                fr.get("display_name"),
                fr.get("name"),
                fr.get("team"),
                f"{fr.get('city', '')} {fr.get('team', '')}".strip(),
            ]
            for c in candidates:
                k = normalize_team_name(c or "")
                if k and k not in index:
                    index[k] = slug
        out[league] = index
    return out


def lookup_team_page_url(league: str, team_name: str, franchise_index: dict) -> str | None:
    league_lower = (league or "").lower()
    if league_lower not in franchise_index:
        return None
    idx = franchise_index[league_lower]
    key = normalize_team_name(team_name)
    if key in idx:
        return f"/teams/{league_lower}/{idx[key]}"
    # try last-token match for cases like "Yankees" alone vs "New York Yankees"
    parts = key.split()
    if len(parts) > 1:
        last = " ".join(parts[-2:]) if len(parts) > 1 else parts[-1]
        if last in idx:
            return f"/teams/{league_lower}/{idx[last]}"
    return None


def main():
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True, data_only=True)
    franchise_index = load_franchise_slugs()

    markers: list[dict] = []

    # ---------- Team List ----------
    ws = wb["Team List"]
    rows = ws.iter_rows(min_row=2, values_only=True)
    excluded_no_coords = 0
    excluded_notable_venues = 0
    excluded_not_ml = 0
    team_list_count = 0

    # Leagues that should be re-labeled by Main Division instead of the
    # workbook's broad League bucket. "Int'l Basketball" is handled
    # separately because Euroleague clubs (ml='Euroleague') keep the
    # Euroleague brand while non-Euroleague Int'l Basketball rows use
    # main_div.
    MAIN_DIV_LABEL_LEAGUES = {
        "Minor Lg Base",
        "Int'l Volleyball",
        "Int'l W Basketball",
        "Int'l Handball",
        "Minor/Jr/Int'l Hockey",
        "Dom. Rugby Union",
        "FBS",
        "FCS",
        "NCAA W",
        "College Hockey",
        "Other Sports",
        "Other Women Sports",
    }
    # Inert venue rows that should never appear on the /sports map.
    NON_TEAM_LEAGUES = {
        "Notable Venues",   # event venues, not teams
        "Historic Venues",  # historic stadiums, not teams
    }

    # Schema as of 2026-05-17: 20 cols. Gold Standard col L (idx 11) inserted
    # between Metro Area (val) and Major League. All downstream indices
    # shifted by 1; the gold flag is unpacked alongside the rest.
    for r in rows:
        if not r or len(r) < 20:
            continue
        sport, league, team, main_div, division, city, metro, state, country, level, _metro_val, gold_flag, ml, season, _affil, _annual, qid, wiki, lat, lng = r[:20]

        # Every row admits; tier is encoded via marker_level ('Major'
        # for ml in {Y, Euroleague}, 'Other' otherwise). The /sports UI
        # gates visibility via the preset chip.
        is_major_league_row = ml == "Y" or ml == "Euroleague"
        # Sport-specific overrides: F1 and NASCAR are workbook ml='' but
        # are unambiguously Major League scope in their sport. Promote them
        # to Major so they ship in the default first paint and the Gold
        # Standard set (F1 only) can resolve correctly.
        if sport == "Auto Racing" and league in ("F1", "NASCAR"):
            is_major_league_row = True
        is_ncaa_basketball = sport == "Basketball" and league == "NCAA"
        if league in NON_TEAM_LEAGUES:
            excluded_notable_venues += 1
            continue
        if lat is None or lng is None or (lat == 0 and lng == 0):
            excluded_no_coords += 1
            continue

        # Fold the four UK home nations into a single 'United Kingdom'
        # country chip on the /sports map. Workbook keeps them separate
        # for legitimate editorial reasons (e.g. the FootballClub_Data
        # league field uses 'England' as the Premier-League marker), but
        # at the country-filter facet level they read as one geography.
        if country in ("England", "Scotland", "Wales", "Northern Ireland"):
            country = "United Kingdom"

        country_iso2 = COUNTRY_ISO2.get(country)
        team_page_url = lookup_team_page_url(league or "", team or "", franchise_index)

        # Display-league override. Some workbook League values are broad
        # buckets ("Int'l Volleyball", "FBS") and the actual top-flight
        # competition lives in Main Division. Re-bucket those rows so the
        # chip filter exposes the league the user actually recognizes.
        display_league = league
        if league == "Int'l Basketball":
            display_league = "Euroleague" if ml == "Euroleague" else (main_div or league)
        elif league in MAIN_DIV_LABEL_LEAGUES or is_ncaa_basketball:
            display_league = main_div or league

        # Level coding: gold (Major) for ml='Y'/'Euroleague' rows, slate
        # (Other) for the NCAA / Minor / Junior / second-flight admissions.
        marker_level = "Major" if is_major_league_row else "Other"

        markers.append({
            "sport": sport,
            "league": display_league,
            "league_raw": league,
            "team": team,
            "main_div": main_div or None,
            "division": division or None,
            "city": city,
            "metro": metro,
            "metro_slug": slugify(metro) if metro else None,
            "state": state or None,
            "country": country,
            "country_iso2": country_iso2,
            "level": marker_level,
            "lat": float(lat),
            "lng": float(lng),
            "wikidata_qid": qid or None,
            "wikipedia_url": wiki or None,
            "team_page_url": team_page_url,
            "source": "team_list",
            "workbook_level": _normalize_level(level),
        })
        team_list_count += 1

    # ---------- FootballClub_Data ----------
    ws = wb["FootballClub_Data"]
    rows = ws.iter_rows(min_row=2, values_only=True)
    fc_count = 0
    fc_no_coords = 0

    # National-team metadata is sourced from public/data/national-teams.tsv
    # (definitive reference uploaded 2026-05-16). Columns: National Team,
    # Federation, FIFA, Active/Defunct. Federation includes 'Unaffiliated'.
    # FIFA is either 'FIFA' or blank. Status is 'Active' or 'Defunct';
    # Defunct rows are NEVER emitted as markers.
    #
    # We deliberately do NOT alias workbook names to NT names. The workbook
    # carries both legacy and modern entries for two countries (Macedonia +
    # North Macedonia, Swaziland + Eswatini); only the modern names match
    # NT directly, which keeps the FIFA count exact. The legacy duplicates
    # silently fall through to the no-NT-entry skip.
    NT_COUNTRY_ALIASES: dict[str, str] = {}
    nt_path = ROOT / "public" / "data" / "national-teams.tsv"
    nt_lookup: dict[str, dict] = {}
    if nt_path.exists():
        with open(nt_path, "r") as nf:
            next(nf)  # header
            for line in nf:
                parts = line.rstrip("\r\n").split("\t")
                if len(parts) < 4 or not parts[0]:
                    continue
                name, fed, fifa, status = (p.strip() for p in parts[:4])
                nt_lookup[name] = {
                    "federation": fed or None,
                    "fifa": fifa == "FIFA",
                    "active": status == "Active",
                }

    fc_excluded_not_ml = 0  # retained for parity but no longer used to drop rows
    for r in rows:
        if not r or len(r) < 11:
            continue
        team, city, metro, county, country, league, level, club, ml, lat, lng = r[:11]
        if lat is None or lng is None or (lat == 0 and lng == 0):
            fc_no_coords += 1
            continue

        is_national_team = isinstance(club, str) and club.strip().lower() == "country"

        # UK home-nations consolidation applies to CLUB rows only. National-
        # team rows must stay distinct (England, Scotland, Wales, Northern
        # Ireland are separate UEFA / FIFA national football teams; there is
        # no 'United Kingdom' international team).
        if not is_national_team and country in ("England", "Scotland", "Wales", "Northern Ireland"):
            country = "United Kingdom"

        country_iso2 = COUNTRY_ISO2.get(country)
        nt_fifa = None
        nt_active = None
        if is_national_team:
            # International Teams ship as level=Other so they only enter
            # the visible set via the Special Filter opt-in (not in the
            # default Major League preset). Federation / FIFA / Active
            # come from public/data/national-teams.tsv via nt_lookup.
            # Lookup by team name (which equals the country name for sovereign
            # nations and the territory name for dependencies whose 'country'
            # column is the parent sovereign — Christmas Island under Australia,
            # Saba / Sint Eustatius under Bonaire, etc).
            lookup_key = NT_COUNTRY_ALIASES.get(team, team)
            nt_entry = nt_lookup.get(lookup_key)
            if nt_entry is None or not nt_entry["active"]:
                # Skip rows that are defunct or absent from the reference.
                continue
            display_league = "International Teams"
            fc_marker_level = "Other"
            federation = nt_entry["federation"]
            nt_fifa = nt_entry["fifa"]
            nt_active = nt_entry["active"]
        else:
            display_league = league
            fc_marker_level = "Major" if ml == "Y" else "Other"
            federation = None

        markers.append({
            "sport": "Football",  # top-flight soccer
            "league": display_league,
            "league_raw": league,
            "team": team,
            "main_div": None,
            "division": None,
            "city": city,
            "metro": metro,
            "metro_slug": slugify(metro) if metro else None,
            "state": None,
            "country": country,
            "country_iso2": country_iso2,
            "level": fc_marker_level,
            "lat": float(lat),
            "lng": float(lng),
            "wikidata_qid": None,
            "wikipedia_url": None,
            "team_page_url": None,  # no per-club pages yet
            "source": "football_club_data",
            "federation": federation,
            "fifa": nt_fifa,
            "active": nt_active,
            "workbook_level": _normalize_level(level),
        })
        fc_count += 1

    # ---------- Sort + write ----------
    markers.sort(key=lambda m: (m["sport"] or "", m["league"] or "", m["team"] or ""))

    out_path = OUT_DIR / "all-teams.json"
    with open(out_path, "w") as f:
        json.dump(markers, f, ensure_ascii=False, indent=0, separators=(",", ":"))

    # ---------- Summary ----------
    by_sport = Counter(m["sport"] for m in markers)
    by_league = Counter(m["league"] for m in markers)
    by_country = Counter(m["country"] for m in markers)
    pageable = sum(1 for m in markers if m["team_page_url"])

    by_level = Counter(m["level"] for m in markers)
    summary = {
        "total_markers": len(markers),
        "major_markers": by_level.get("Major", 0),
        "other_markers": by_level.get("Other", 0),
        "by_sport": dict(by_sport.most_common()),
        "by_league_top": dict(by_league.most_common(40)),
        "by_country_top": dict(by_country.most_common(15)),
        "markers_with_team_page": pageable,
        "sources": {
            "team_list": team_list_count,
            "football_club_data": fc_count,
            "team_list_excluded_no_coords": excluded_no_coords,
            "team_list_excluded_not_ml": excluded_not_ml,
            "team_list_excluded_notable_venues": excluded_notable_venues,
            "football_club_data_excluded_no_coords": fc_no_coords,
        },
        # League cards rendered on /sports — order encodes "live" page status.
        # FootballClub_Data uses country names as league values ("England",
        # "Spain") rather than brand names, so the lookup keys reflect that.
        # Card labels stay brand-friendly. Static for now; ETL flips 'status'
        # once each per-league page ships.
        "league_cards": [
            {"league": "NFL",  "label": "NFL",            "sport": "American Football",  "status": "live",   "page": "/teams/nfl", "team_count": by_league.get("NFL", 0)},
            {"league": "MLB",  "label": "MLB",            "sport": "Baseball",           "status": "live",   "page": "/teams/mlb", "team_count": by_league.get("MLB", 0)},
            {"league": "NBA",  "label": "NBA",            "sport": "Basketball",         "status": "live",   "page": "/teams/nba", "team_count": by_league.get("NBA", 0)},
            {"league": "NHL",  "label": "NHL",            "sport": "Hockey",             "status": "live",   "page": "/teams/nhl", "team_count": by_league.get("NHL", 0)},
            {"league": "Euroleague", "label": "Euroleague", "sport": "Basketball", "status": "coming", "page": None,         "team_count": by_league.get("Euroleague", 0)},
            {"league": "England",  "label": "Premier League", "sport": "Football",       "status": "coming", "page": None,         "team_count": by_league.get("England", 0)},
            {"league": "Spain",    "label": "La Liga",        "sport": "Football",       "status": "coming", "page": None,         "team_count": by_league.get("Spain", 0)},
            {"league": "Italy",    "label": "Serie A",        "sport": "Football",       "status": "coming", "page": None,         "team_count": by_league.get("Italy", 0)},
            {"league": "Germany",  "label": "Bundesliga",     "sport": "Football",       "status": "coming", "page": None,         "team_count": by_league.get("Germany", 0)},
            {"league": "France",   "label": "Ligue 1",        "sport": "Football",       "status": "coming", "page": None,         "team_count": by_league.get("France", 0)},
            {"league": "AFL",  "label": "AFL",            "sport": "Aussie Rules",       "status": "coming", "page": None,         "team_count": by_league.get("AFL", 0)},
            {"league": "NRL",  "label": "NRL",            "sport": "Rugby League",       "status": "coming", "page": None,         "team_count": by_league.get("NRL", 0)},
            {"league": "IPL",  "label": "IPL",            "sport": "T20 Cricket",        "status": "coming", "page": None,         "team_count": by_league.get("IPL", 0)},
            {"league": "CFL",  "label": "CFL",            "sport": "Canadian Football",  "status": "coming", "page": None,         "team_count": by_league.get("CFL", 0)},
            {"league": "NWSL", "label": "NWSL",           "sport": "W Football",         "status": "coming", "page": None,         "team_count": by_league.get("NWSL", 0)},
            {"league": "WSL",  "label": "WSL",            "sport": "W Football",         "status": "coming", "page": None,         "team_count": by_league.get("WSL", 0)},
        ],
    }

    summary_path = OUT_DIR / "league-summary.json"
    with open(summary_path, "w") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)

    print(f"Wrote {out_path} ({len(markers)} markers)")
    print(f"  Team List markers: {team_list_count}")
    print(f"  FootballClub_Data markers: {fc_count}")
    print(f"  With team page URL: {pageable}")
    print(f"Wrote {summary_path}")
    print()
    print("Sports breakdown:")
    for s, c in by_sport.most_common(15):
        print(f"  {s}: {c}")


if __name__ == "__main__":
    main()

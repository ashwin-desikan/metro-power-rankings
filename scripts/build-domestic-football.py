#!/usr/bin/env python3
"""Build public/data/football/domestic-clubs.json: every club that has ever
played in a tracked first division across all tracked leagues, the nine marquee
hub leagues (from Leagues History; MLS via World) plus the long tail (StandOth +
World). The dedicated league hubs still exist for depth; this is the single
all-leagues master table.

Club identity is the Cur. Name alone — it is unique across the sheets, so a club
keeps one row no matter how many cities or countries it has played under. Metro
is just the latest season's city (a display attribute), so relocations don't
split a club. Honours are computed per club-and-COUNTRY (era) from the season
rows, with an all-time roll-up that is their sum, so a club that played under
more than one state (Dynamo Kyiv: Soviet Union then Ukraine; Red Star:
Yugoslavia, Serbia & Montenegro, Serbia) carries a breakdown for each.

A row is top flight if Level == 1 OR First Division == Y (the union catches both
pre-unified top divisions with a blank flag and national championships recorded
with no league level, e.g. Brazil's Taca Brasil).
"""
import json
import os
import openpyxl
from collections import defaultdict

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
WB = os.path.join(ROOT, "workbooks", "Champions League-201516.xlsx")
OUT = os.path.join(ROOT, "public", "data", "football", "domestic-clubs.json")
STANDINGS_SHEETS = ["Leagues History", "StandOth", "World"]


def to_year(v):
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


def truthy(v):
    return str(v or "").strip().upper() not in ("", "0", "N", "NO", "NONE", "#N/A", "FALSE")


def clean_metro(v):
    s = str(v).strip() if v is not None else ""
    return s if s and s != "#N/A" else None


def is_top_flight(level, first_division):
    try:
        if int(float(level)) == 1:
            return True
    except (TypeError, ValueError):
        pass
    return truthy(first_division)


def headers(ws):
    return [c if c is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]


def blank_honours():
    return {
        "titles": 0, "majorCups": 0, "minorCups": 0,
        "contTitles": 0, "contApps": 0, "clTitles": 0, "clApps": 0,
        "lastTitle": None, "lastYear": None, "firstYear": None,
    }


def main():
    wb = openpyxl.load_workbook(WB, read_only=True, data_only=True)

    clubs = {}            # Cur. Name -> club dict
    country_latest = {}   # country -> latest top-flight End Year

    for sheet in STANDINGS_SHEETS:
        ws = wb[sheet]
        ix = {h: i for i, h in enumerate(headers(ws))}

        def g(row, name):
            j = ix.get(name)
            return row[j] if j is not None else None

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not is_top_flight(g(row, "Level"), g(row, "First Division")):
                continue
            ey = to_year(g(row, "End Year"))
            if ey is None:
                continue
            name = (g(row, "Cur. Name") or g(row, "Team") or "").strip()
            country = g(row, "Country (Leag)")
            if not name or not country:
                continue
            metro = clean_metro(g(row, "Metro Area"))
            c = clubs.get(name)
            if c is None:
                c = clubs[name] = {
                    "name": name, "metro": None, "byCountry": {},
                    "lastFdYear": None, "currentCountry": None,
                }
            h = c["byCountry"].get(country)
            if h is None:
                h = c["byCountry"][country] = blank_honours()
            if truthy(g(row, "Champions")):
                h["titles"] += 1
                h["lastTitle"] = max(h["lastTitle"] or 0, ey)
            if truthy(g(row, "Cup (Major Domestic)")):
                h["majorCups"] += 1
            if truthy(g(row, "Cup (Minor Domestic)")):
                h["minorCups"] += 1
            if truthy(g(row, "Eur.Trophy")):
                h["contTitles"] += 1
            if truthy(g(row, "Eur. App")):
                h["contApps"] += 1
            if truthy(g(row, "CL Champ")):
                h["clTitles"] += 1
            if truthy(g(row, "CL App")):
                h["clApps"] += 1
            h["lastYear"] = max(h["lastYear"] or 0, ey)
            h["firstYear"] = ey if h["firstYear"] is None else min(h["firstYear"], ey)
            # Latest top-flight season sets the current country and city.
            if c["lastFdYear"] is None or ey >= c["lastFdYear"]:
                c["lastFdYear"] = ey
                c["currentCountry"] = country
                if metro:
                    c["metro"] = metro
            country_latest[country] = max(country_latest.get(country, 0), ey)

    global_latest = max(country_latest.values()) if country_latest else 0

    def era_record(h):
        return {
            "titles": h["titles"],
            "lastTitle": h["lastTitle"],
            "cups": h["majorCups"] + h["minorCups"],
            "majorTrophies": h["titles"] + h["majorCups"] + h["contTitles"],
            "contTitles": h["contTitles"],
            "contApps": h["contApps"],
            "clTitles": h["clTitles"],
            "clApps": h["clApps"],
            "lastYear": h["lastYear"],
            "firstYear": h["firstYear"],
        }

    out = []
    for c in clubs.values():
        by_country = {ctry: era_record(h) for ctry, h in c["byCountry"].items()}
        allt = {"titles": 0, "cups": 0, "majorTrophies": 0, "contTitles": 0,
                "contApps": 0, "clTitles": 0, "clApps": 0, "lastTitle": None}
        for r in by_country.values():
            for k in ("titles", "cups", "majorTrophies", "contTitles", "contApps", "clTitles", "clApps"):
                allt[k] += r[k]
            if r["lastTitle"]:
                allt["lastTitle"] = max(allt["lastTitle"] or 0, r["lastTitle"])
        cur = c["currentCountry"]
        is_current = bool(cur) and c["lastFdYear"] == country_latest.get(cur) and country_latest.get(cur, 0) >= global_latest - 2
        out.append({
            "name": c["name"],
            "metro": c["metro"],
            "country": cur,
            "status": "current" if is_current else "former",
            "lastTopFlight": c["lastFdYear"],
            "allTime": allt,
            "byCountry": by_country,
        })

    # Reconcile every honour to the authoritative Totals aggregate. The season
    # rows only carry cup / continental markers for the European sheets, so
    # non-European clubs (and a few renamed clubs whose old honours the author
    # credited to the current name, e.g. Dukla Prague -> FK Marila Pribram) come
    # up empty from the season recompute. Totals is complete for every club, so
    # we trust it for the all-time figures and add any shortfall to the club's
    # current-country era. Eur Trophy / EC-CL Cup double as continental for
    # non-European clubs (Copa Libertadores, etc.).
    ws_t = wb["Totals"]
    tix = {h: i for i, h in enumerate(headers(ws_t))}

    def tint(row, col):
        if col not in tix:
            return 0
        try:
            return int(row[tix[col]] or 0)
        except (TypeError, ValueError):
            return 0

    totals = {}
    for row in ws_t.iter_rows(min_row=2, values_only=True):
        nm = (row[tix["Cur. Name"]] if "Cur. Name" in tix else None)
        nm = (nm or "").strip()
        if not nm:
            continue
        agg = {
            "titles": tint(row, "# Title (1 Div)"),
            "cups": tint(row, "FA Cups") + tint(row, "Lg. Cups"),
            "contTitles": tint(row, "Eur Trophy"),
            "contApps": tint(row, "Eur App"),
            "clTitles": tint(row, "CL/EC Cup"),
            "clApps": tint(row, "EC/CL App"),
            "majorTrophies": tint(row, "# Maj Trophies"),
        }
        prev = totals.get(nm)
        if prev is None:
            totals[nm] = agg
        else:
            for k in agg:
                prev[k] = max(prev[k], agg[k])

    FIELDS = ("titles", "cups", "contTitles", "contApps", "clTitles", "clApps", "majorTrophies")
    for rec in out:
        t = totals.get(rec["name"])
        if not t or not rec["country"]:
            continue
        era = rec["byCountry"].setdefault(rec["country"], {
            "titles": 0, "lastTitle": None, "cups": 0, "majorTrophies": 0,
            "contTitles": 0, "contApps": 0, "clTitles": 0, "clApps": 0, "lastYear": None,
        })
        for f in FIELDS:
            gap = t[f] - rec["allTime"][f]
            if gap > 0:
                rec["allTime"][f] = t[f]
                era[f] += gap

    out.sort(key=lambda r: (-r["allTime"]["titles"], -r["allTime"]["majorTrophies"], r["name"]))
    countries = sorted({ctry for r in out for ctry in r["byCountry"]})
    data = {
        "_meta": {
            "source": "Champions League-201516.xlsx (Leagues History + StandOth + World)",
            "clubs": len(out),
            "current": sum(1 for r in out if r["status"] == "current"),
            "countries": len(countries),
        },
        "countries": countries,
        "clubs": out,
    }
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, separators=(",", ":"))
    print("wrote", OUT)
    print("clubs:", len(out), "| current:", data["_meta"]["current"], "| countries:", len(countries))


if __name__ == "__main__":
    main()

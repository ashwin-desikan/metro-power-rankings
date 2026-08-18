#!/usr/bin/env python3
"""Resolve every F1 factory town in bases.py to a Metro Area, or refuse.

  scripts/f1/curation/base_metros.json

MetroAreas.xlsx `Municipality` is the only authority for city-to-metro on this
site, and it does not make this easy. England is not stored as towns: it is
stored as MSOAs, the census geography, so Mercedes' home appears as "Brackley
North" and "Brackley South" and Alpine's as "Kingham, Enstone & Middle Barton".
A town therefore resolves only if it appears as a WORD inside one or more MSOA
names and those MSOAs agree on a single non-blank Metro Area.

Everything else is REFUSED and listed, in line with the rule the rankings board
already runs on: an unresolved place is a question for the workbook, not a
licence to infer. Two of the most important towns in the sport fail this test.
Brackley and Silverstone both exist in the workbook with a BLANK Metro Area, so
Mercedes' and Aston Martin's factories currently resolve to nothing at all.

  python scripts/f1/resolve_base_metros.py
  python scripts/f1/resolve_base_metros.py --self-test
"""
import argparse, json, os, re, sys, unicodedata
from collections import defaultdict

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(os.path.dirname(HERE))
sys.path.insert(0, HERE)

from bases import BASES  # noqa: E402

WORKBOOK = os.path.join(ROOT, "MetroAreas.xlsx")
DEST = os.path.join(HERE, "curation", "base_metros.json")


def fold(s):
    """Accent- and case-insensitive key. The workbook writes Viry-Chatillon with
    a circumflex and Fussgonheim with an eszett; bases.py writes both ASCII, so
    one side has to fold to meet the other."""
    s = (s or "").replace("ß", "ss")
    s = unicodedata.normalize("NFKD", s)
    return "".join(c for c in s if not unicodedata.combining(c)).lower().strip()


def load_sheet(wb, sheet, name_col):
    """One sheet, as {folded country: [(name, metro, (regions...))]}.

    🔴 THE TWO SHEETS DO NOT AGREE ON WHAT A COUNTRY IS. `Municipality` files
    the UK as England (6,856), Scotland (656), Wales (408) and Northern Ireland
    (80). `Counties` files all of it as "United Kingdom" (361 rows, one per
    local authority). Reading only Municipality, as this script did until
    2026-08-18, means a village too small to be its own MSOA has no answer at
    all, when its district has one sitting in the next sheet along.
    """
    ws = wb[sheet]
    it = ws.iter_rows(values_only=True)
    hdr = [str(c).strip() if c is not None else "" for c in next(it)]
    ix = {h: i for i, h in enumerate(hdr)}
    for col in ("Country", name_col, "Metro Area"):
        if col not in ix:
            sys.exit(f"FATAL: {sheet} sheet has no {col!r} column: {hdr}")
    # 🔴 THE COLUMN IS NOT CALLED "Region". The Municipality sheet's county-level
    # column is headed "Distri rrondissement/County" (sic), with a separate
    # "State/Region (ISO 3166-2)". Looking for a column called "Region" finds
    # "Region (ISO 3166-2)", which is blank for England, and every England row
    # then silently loses its county. That is not a crash, it is worse: it
    # quietly disables the disambiguation and Bourne in Lincolnshire resolves to
    # London via Bourne End in Buckinghamshire. Both columns are read and either
    # may satisfy a region hint.
    region_cols = [c for c in ("Distri rrondissement/County",
                               "State/Region (ISO 3166-2)") if c in ix]
    if not region_cols:
        sys.exit(f"FATAL: no county or state column on the {sheet} sheet: {hdr}")
    rows = defaultdict(list)
    for r in it:
        def g(k):
            i = ix.get(k)
            return r[i] if i is not None and i < len(r) else None
        country = str(g("Country") or "").strip()
        if not country:
            continue
        regions = tuple(sorted({str(g(c)).strip() for c in region_cols
                                if g(c) not in (None, "")}))
        rows[fold(country)].append((
            str(g(name_col) or "").strip(),
            str(g("Metro Area") or "").strip(),
            regions,
        ))
    return rows


def load_workbook_rows():
    from openpyxl import load_workbook
    wb = load_workbook(WORKBOOK, read_only=True, data_only=True)
    muni = load_sheet(wb, "Municipality", "Municipality")
    cnty = load_sheet(wb, "Counties", "Distri rrondissement/County")
    wb.close()
    return muni, cnty


def resolve(town, region, rows):
    """(metro, how, detail) or (None, reason, detail).

    Passes run exact-then-word, region-scoped first. Two rules do the real work:

    A pass that finds candidates but cannot agree on one metro STOPS rather than
    falling through to a looser pass, because a looser match on an ambiguous
    name is how Williams ends up in Leeds.

    And if the region hint names a county the workbook actually knows, the
    unscoped fallback is FORBIDDEN. Bourne in Lincolnshire has no metro; Bourne
    End in Buckinghamshire is in the London metro. Falling back past a real
    county silently swaps one for the other."""
    ftown = fold(town)
    fregion = fold(region)

    def verdict(cands, how):
        metros = {m for _n, m, _rg in cands if m}
        blanks = sum(1 for _n, m, _rg in cands if not m)
        if len(metros) == 1:
            return sorted(metros)[0], how, f"{len(cands)} row(s), {blanks} blank"
        if not metros:
            return None, "no-metro", f"{len(cands)} row(s), all blank"
        return None, "ambiguous", ", ".join(sorted(metros)[:5])

    exact = [(n, m, rg) for n, m, rg in rows if fold(n) == ftown]
    pat = re.compile(r"(?<![a-z])" + re.escape(ftown) + r"(?![a-z])")
    word = [(n, m, rg) for n, m, rg in rows if pat.search(fold(n))]

    def scope(cands):
        if not fregion:
            return []
        return [h for h in cands if any(fold(x) == fregion for x in h[2])]

    known_region = bool(fregion) and any(
        any(fold(x) == fregion for x in rg) for _n, _m, rg in rows)

    passes = [(scope(exact), "exact+region"), (scope(word), "word+region")]
    if not known_region:
        passes += [(exact, "exact"), (word, "word")]
    for cands, how in passes:
        if cands:
            return verdict(cands, how)

    if exact or word:
        if known_region:
            return None, "region-mismatch", (
                f"{len(exact) + len(word)} row(s) match the name, none in {region}")
        return None, "no-place", ""

    # Last resort, and not an inference: the workbook names some places only as
    # metros. Northampton has no municipality row and is a Metro Area, so a
    # factory in Northampton is in the Northampton metro by the workbook's own
    # word rather than by ours.
    metro_names = {fold(m) for _n, m, _rg in rows if m}
    if ftown in metro_names:
        return town, "metro-name", "named as a Metro Area, not as a municipality"
    return None, "no-place", ""


def resolve_in_counties(town, region, district, rows):
    """The fallback, one administrative level up.

    A factory village is often too small to be its own census area but its
    DISTRICT is always in the Counties sheet. Ockham has no Municipality row at
    all; Guildford, the borough containing it, resolves to London. Neuburg an
    der Donau is not a district, but Neuburg-Schrobenhausen is, and it resolves
    to Munich.

    Three ways in, most explicit first:
      1. an explicit `district` from the curation, matched exactly. Naming the
         district a village sits in is a checkable fact, and it is the only way
         to get from Ockham to Guildford, which share no letters.
      2. the town's own name against district names, exact then whole-word.
      3. the town's name against the STATE/REGION column, which is how a
         city-region like Tokyo resolves: 54 ward rows, all agreeing on Tokyo.
    Each way still refuses unless the matches agree on one non-blank metro."""
    if district:
        fd = fold(district)
        hits = [(n, m, rg) for n, m, rg in rows if fold(n) == fd]
        if hits:
            metros = {m for _n, m, _rg in hits if m}
            if len(metros) == 1:
                return sorted(metros)[0], "county-district", district
            if not metros:
                return None, "no-metro", f"{district} district, blank in Counties"
            return None, "ambiguous", ", ".join(sorted(metros)[:5])
        return None, "no-place", f"no district called {district!r}"

    metro, how, detail = resolve(town, region, rows)
    if metro:
        return metro, "county-" + how, detail

    fr = fold(town)
    byregion = [(n, m, rg) for n, m, rg in rows
                if any(fold(x) == fr for x in rg)]
    if byregion:
        metros = {m for _n, m, _rg in byregion if m}
        blanks = sum(1 for _n, m, _rg in byregion if not m)
        if len(metros) == 1:
            return (sorted(metros)[0], "county-region",
                    f"{len(byregion)} row(s), {blanks} blank")
        if metros:
            return None, "ambiguous", ", ".join(sorted(metros)[:5])
    return None, how, detail


def self_test():
    ok = True

    def check(label, got, want):
        nonlocal ok
        if got != want:
            ok = False; print(f"  FAIL {label}: got {got!r}, want {want!r}")
        else:
            print(f"  ok   {label}")

    check("fold strips accents", fold("Viry-Châtillon"), "viry-chatillon")
    check("fold handles eszett", fold("Fußgönheim"), "fussgonheim")
    fake = [("Grove Park", "Leeds-Bradford", ("West Yorkshire",)),
            ("Grove & Wantage", "Oxford", ("Oxfordshire",)),
            ("Brackley North", "", ("West Northamptonshire",)),
            ("Brackley South", "", ("West Northamptonshire",)),
            ("Bourne End", "London", ("Buckinghamshire",)),
            ("Bourne West", "", ("Lincolnshire",)),
            ("Towcester", "Northampton", ("West Northamptonshire",)),
            ("Woking Central", "London", ("Surrey",))]
    check("region narrows an ambiguous name",
          resolve("Grove", "Oxfordshire", fake)[0], "Oxford")
    check("no region, ambiguous name refuses",
          resolve("Grove", "", fake)[:2], (None, "ambiguous"))
    check("blank metro refuses",
          resolve("Brackley", "West Northamptonshire", fake)[:2], (None, "no-metro"))
    check("word match inside an MSOA name",
          resolve("Woking", "Surrey", fake)[0], "London")
    check("absent town refuses", resolve("Maranello", "Modena", fake)[1], "no-place")
    check("an exact row beats a word match",
          resolve("Grove Park", "West Yorkshire", fake)[1], "exact+region")
    # The bug this gate exists for: Bourne in Lincolnshire has no metro, and
    # Bourne End in Buckinghamshire is in London's.
    check("a real county forbids the unscoped fallback",
          resolve("Bourne", "Lincolnshire", fake)[:2], (None, "no-metro"))
    check("without a usable county the loose match wins, which is the hazard "
          "the gate exists to close",
          resolve("Bourne", "Ruritania", fake)[:2], ("London", "word"))
    check("a metro name with no municipality row still resolves",
          resolve("Northampton", "Northamptonshire", fake)[:2],
          ("Northampton", "metro-name"))
    # The Counties fallback, one level up from the village.
    counties = [("Guildford", "London", ("Surrey",)),
                ("West Northamptonshire", "", ("West Northamptonshire",)),
                ("Neuburg-Schrobenhausen", "Munich", ("Bavaria",)),
                ("Setagaya", "Tokyo", ("Tokyo",)),
                ("Nerima", "Tokyo", ("Tokyo",)),
                ("Kawasaki", "Yokohama", ("Kanagawa",)),
                ("Sagamihara", "Yokohama", ("Kanagawa",))]
    check("a district hint reaches a village's borough",
          resolve_in_counties("Ockham", "Surrey", "Guildford", counties)[:2],
          ("London", "county-district"))
    check("a blank district refuses and names itself",
          resolve_in_counties("Brackley", "West Northamptonshire",
                              "West Northamptonshire", counties)[:2],
          (None, "no-metro"))
    check("a district named after the town needs the hint",
          resolve_in_counties("Neuburg an der Donau", "Bavaria",
                              "Neuburg-Schrobenhausen", counties)[0], "Munich")
    check("a city-region resolves without a district hint",
          resolve_in_counties("Tokyo", "Tokyo", "", counties)[0], "Tokyo")
    check("a region name that is not a metro name still resolves off the "
          "region column",
          resolve_in_counties("Kanagawa", "", "", counties)[:2],
          ("Yokohama", "county-region"))
    check("the UK is one country on Counties and four on Municipality",
          [country_in("counties", "England"), country_in("muni", "England")],
          ["United Kingdom", "England"])
    print("resolver self-test:", "PASS" if ok else "FAIL")
    return ok


# 🔴 THE TWO SHEETS LABEL THE UK DIFFERENTLY. Municipality splits it into
# England, Scotland, Wales and Northern Ireland; Counties files the lot as
# "United Kingdom". Everywhere else the two agree, so only the UK needs a map.
UK = {"England", "Scotland", "Wales", "Northern Ireland"}


def country_in(sheet, country):
    if sheet == "counties" and country in UK:
        return "United Kingdom"
    return country


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--self-test", action="store_true")
    a = ap.parse_args()
    if a.self_test:
        return 0 if self_test() else 1
    if not self_test():
        return 1

    muni, cnty = load_workbook_rows()
    print(f"\nworkbook: {sum(len(v) for v in muni.values())} municipality rows "
          f"across {len(muni)} countries; "
          f"{sum(len(v) for v in cnty.values())} county rows "
          f"across {len(cnty)} countries")

    places = {}
    for b in BASES:
        key = (b["country"], b["town"], b["region"], b.get("district", ""))
        places.setdefault(key, []).append(b["lineage"])

    resolved, refused = {}, []
    for (country, town, region, district), lineages in sorted(places.items()):
        mrows = muni.get(fold(country_in("muni", country)), [])
        crows = cnty.get(fold(country_in("counties", country)), [])
        if not mrows and not crows:
            refused.append((country, town, region, "no-country",
                            f"neither sheet has rows for {country!r}",
                            sorted(set(lineages))))
            continue
        metro, how, detail = (resolve(town, region, mrows) if mrows
                              else (None, "no-place", "not on the Municipality sheet"))
        # Municipality first because it is the finer geography. Counties is the
        # fallback, not the default: a village's own row beats its district's.
        if not metro and crows:
            c_metro, c_how, c_detail = resolve_in_counties(town, region, district, crows)
            if c_metro:
                metro, how, detail = c_metro, c_how, c_detail
            elif how == "no-place":
                # Keep whichever refusal actually says something. "Brackley has
                # two census areas and both are blank" is a ruling someone can
                # act on; "no district called Brackley" is noise.
                how, detail = c_how, c_detail
            elif c_how != "no-place":
                detail = f"{detail}; in Counties: {c_how}" + (f" - {c_detail}" if c_detail else "")
        if metro:
            resolved[f"{country}|{town}"] = {"metro": metro, "how": how,
                                             "detail": detail, "region": region}
        else:
            refused.append((country, town, region, how, detail,
                            sorted(set(lineages))))

    print(f"\n{'PLACE':38s} {'METRO':22s} HOW")
    for k, v in sorted(resolved.items()):
        print(f"  {k:36s} {v['metro'][:20]:22s} {v['how']} ({v['detail']})")

    print(f"\nresolved {len(resolved)} of {len(places)} places\n")
    print("NEEDS A WORKBOOK RULING (these towns show no metro on the pages):")
    for country, town, region, why, detail, lineages in refused:
        print(f"  {town} ({region}, {country}): {why}"
              f"{' - ' + detail if detail else ''}")
        print(f"      affects: {', '.join(lineages)}")

    doc = {
        "meta": {
            "source": "MetroAreas.xlsx, Municipality sheet",
            "rule": ("A town resolves only when the workbook rows matching it "
                     "agree on one non-blank Metro Area. Anything else is "
                     "refused and listed, never inferred."),
            "resolved": len(resolved), "places": len(places),
            "refused": len(refused),
        },
        "metros": resolved,
        "unresolved": [{"country": c, "town": t, "region": r, "why": w,
                        "detail": d, "lineages": l}
                       for c, t, r, w, d, l in refused],
    }
    os.makedirs(os.path.dirname(DEST), exist_ok=True)
    with open(DEST, "w", encoding="utf-8") as f:
        json.dump(doc, f, indent=1, ensure_ascii=False, sort_keys=True)
    print(f"\n-> {DEST}")
    return 0


if __name__ == "__main__":
    sys.exit(main())

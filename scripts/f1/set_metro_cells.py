"""Set Metro Area cells in MetroAreas.xlsx surgically, on an explicit ruling.

MetroAreas.xlsx is 32 MB across 32 sheets and it is Ashwin's master, so this
patches only the target <c> elements inside the sheet XML and writes every other
zip entry through byte-for-byte with its original ZipInfo. openpyxl would rewrite
the whole file from its own model and silently drop what that model does not
carry. See the surgical-xlsx-cell-edit reference.

REFUSES rather than guesses:
  * if the value is not already in sharedStrings (adding one means renumbering
    count/uniqueCount, a different job),
  * if any target cell holds a FORMULA,
  * if a target cell's current value is not what the caller said to expect.

Current ruling (2026-08-19, Ashwin): Brackley belongs to the LONDON metro, "just
barely", overriding the West Northamptonshire district default of Northampton
set on 2026-08-18. Only the two Brackley MSOAs move. Silverstone and Towcester,
in the same district, stay Northampton.

Run:  python scripts/f1/set_metro_cells.py           (dry run)
      python scripts/f1/set_metro_cells.py --write
"""
import os
import re
import shutil
import sys
import zipfile

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
SRC = os.path.join(ROOT, "MetroAreas.xlsx")
STAMP = "20260819-brackley-london"

# (sheet, cell, expected current value, new value, label)
#
# 🔴 THIS LIST IS THE APPLIED STATE, NOT A HISTORY. Edit it for the ruling you
# are executing now; git carries the trail.
#
# APPLIED — ruling 1 (2026-08-19, Ashwin): Brackley belongs to LONDON, "just
# barely", overriding the West Northamptonshire district default of Northampton
# set on 2026-08-18. Re-running this now is a no-op: the expected-current check
# will refuse, which is the guard working.
#
# WITHDRAWN — ruling 2 (same day): "Silverstone and Towcester should be in
# London, given that circumstance", then "actually stop, keep Northampton as the
# metro area for both". It was applied and reverted from the backup
# MetroAreas.xlsx.bak-20260819-silverstone-towcester-london, which is why that
# file exists. Silverstone, Towcester and Brixworth all remain NORTHAMPTON.
# Do not re-apply it.
EDITS = [
    ("Municipality", "G17430", "Northampton", "London", "Brackley North"),
    ("Municipality", "G17432", "Northampton", "London", "Brackley South"),
]


def sheet_xml_path(z, name):
    wb = z.read("xl/workbook.xml").decode("utf-8", "replace")
    rels = z.read("xl/_rels/workbook.xml.rels").decode("utf-8", "replace")
    relmap = dict(re.findall(r'Id="([^"]+)"[^>]*Target="([^"]+)"', rels))
    for nm, rid in re.findall(r'<sheet name="([^"]+)"[^>]*r:id="([^"]+)"', wb):
        if nm == name:
            return "xl/" + relmap[rid].lstrip("/")
    raise SystemExit("sheet %r not found" % name)


def shared_index(z, value):
    xml = z.read("xl/sharedStrings.xml").decode("utf-8", "replace")
    items = re.findall(r"<si>(.*?)</si>", xml, re.S)
    for i, si in enumerate(items):
        text = "".join(re.findall(r"<t[^>]*>(.*?)</t>", si, re.S))
        if text == value:
            return i
    return None


def check_current():
    wb = openpyxl.load_workbook(SRC, read_only=False, data_only=False)
    for sheet, ref, expect, _new, label in EDITS:
        c = wb[sheet][ref]
        if c.data_type == "f":
            raise SystemExit("REFUSING: %s!%s is a FORMULA." % (sheet, ref))
        if c.value != expect:
            raise SystemExit("REFUSING: %s!%s (%s) holds %r, expected %r."
                             % (sheet, ref, label, c.value, expect))
        print("  %s!%s  %-22s %r -> %r" % (sheet, ref, label, c.value, _new))
    wb.close()


def patch(xml, ref, sidx):
    """Replace the cell at `ref` with a shared-string cell, keeping its style."""
    pat = re.compile(r'<c r="%s"([^>]*?)(/>|>.*?</c>)' % ref, re.S)
    m = pat.search(xml)
    if not m:
        raise SystemExit("REFUSING: cell %s not present in the sheet XML." % ref)
    attrs = m.group(1)
    style = re.search(r'\ss="(\d+)"', attrs)
    new = '<c r="%s"%s t="s"><v>%d</v></c>' % (
        ref, (' s="%s"' % style.group(1)) if style else "", sidx)
    return xml[:m.start()] + new + xml[m.end():]


def main():
    write = "--write" in sys.argv
    print("planned edits:")
    check_current()

    with zipfile.ZipFile(SRC) as z:
        names = z.namelist()
        infos = {n: z.getinfo(n) for n in names}
        idx = {}
        for _s, _r, _e, new, _l in EDITS:
            if new not in idx:
                i = shared_index(z, new)
                if i is None:
                    raise SystemExit("REFUSING: %r is not in sharedStrings." % new)
                idx[new] = i
        print("sharedStrings:", {k: v for k, v in idx.items()})
        blobs = {n: z.read(n) for n in names}

    touched = set()
    for sheet, ref, _e, new, _l in EDITS:
        with zipfile.ZipFile(SRC) as z:
            path = sheet_xml_path(z, sheet)
        xml = blobs[path].decode("utf-8", "replace")
        blobs[path] = patch(xml, ref, idx[new]).encode("utf-8")
        touched.add(path)
    print("sheet XML touched:", sorted(touched))

    if not write:
        print("\nDRY RUN. Re-run with --write to apply.")
        return

    backup = SRC + ".bak-" + STAMP
    shutil.copyfile(SRC, backup)
    print("backup:", backup)
    tmp = SRC + ".tmp"
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as out:
        for n in names:
            out.writestr(infos[n], blobs[n])
    os.replace(tmp, SRC)
    print("written.")

    with zipfile.ZipFile(SRC) as z:
        assert z.testzip() is None, "zip failed testzip()"
        assert z.namelist() == names, "zip entry order changed"
    with zipfile.ZipFile(backup) as zb:
        changed = [n for n in names if zb.read(n) != blobs[n]]
    print("verify: testzip OK, %d entries same order, changed: %s" % (len(names), changed))
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    print("verify: opens, %d sheets" % len(wb.sheetnames))
    wb.close()
    wb = openpyxl.load_workbook(SRC, read_only=False, data_only=False)
    for sheet, ref, _e, new, label in EDITS:
        got = wb[sheet][ref].value
        print("verify: %s!%s %-22s now %r %s" % (sheet, ref, label, got,
                                                 "OK" if got == new else "MISMATCH"))
    wb.close()


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    main()

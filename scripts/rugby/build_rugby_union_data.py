#!/usr/bin/env python3
"""Build /teams/rugby-union portal JSONs.

Sources:
  - OtherLeagues.xlsx sheets "Rugby Union - Intl Results" (6,392 perspective
    rows, 1871->2026) and "Rugby Union - Intl Tables" (943 season-standings
    rows with champion / Grand Slam / Triple Crown / RWC knockout flags).
  - wrurankings.txt: weekly Men's World Rugby rankings since 2003-10-06, wide
    TSV from https://commons.wikimedia.org/wiki/Data:Men%27s_World_Rugby_rankings.tab

Emits:
  public/data/rugby-union/teams.json
  public/data/rugby-union/hub.json
  public/data/rugby-union/team-detail/<slug>.json

Run: python build_rugby_v1.py <OtherLeagues.xlsx> <wrurankings.txt> <out_dir>
"""
import json
import re
import sys
import unicodedata
from collections import defaultdict
from datetime import date
from openpyxl import load_workbook

SIX_NATIONS = {"England", "France", "Ireland", "Italy", "Scotland", "Wales"}
SANZAAR = {"New Zealand", "Australia", "South Africa", "Argentina"}

# Western Samoa rebranded to Samoa in 1997; merge under Samoa. The others are
# the user's canonical countries.json display names.
NAME_CANON = {
    "Western Samoa": "Samoa",
    "Ivory Coast": "Côte d'Ivoire",
    "Bosnia and Herzegovina": "Bosnia-Herzegovina",
}
# Keep existing URLs stable when the canonical name would change the slug.
SLUG_OVERRIDE = {"Côte d'Ivoire": "ivory-coast", "Bosnia-Herzegovina": "bosnia-and-herzegovina"}


def canon(n):
    if n is None:
        return None
    n = str(n).strip()
    return NAME_CANON.get(n, n)


def slugify(name):
    s = unicodedata.normalize("NFKD", name).encode("ascii", "ignore").decode()
    return re.sub(r"[^a-zA-Z0-9]+", "-", s).strip("-").lower()


def iso8(v):
    s = str(v).strip()
    if len(s) >= 8 and s[:8].isdigit():
        return f"{s[:4]}-{s[4:6]}-{s[6:8]}"
    return None


def comp_family(comp):
    c = (comp or "").lower()
    if "six nations" in c:
        return "6N"
    if "five nations" in c or "home nations" in c:
        return "5N"
    if "rugby championship" in c:
        return "TRC"
    if "tri nations" in c or "tri-nations" in c:
        return "TRI"
    if "rugby world cup" in c and "warm" not in c:
        return "RWC"
    if "nations championship" in c:
        return "NC"
    return None


def main(xlsx_path, rank_path, out_dir):
    import os

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)

    # ---------------- Results ----------------
    ws = wb["Rugby Union - Intl Results"]
    rrows = list(ws.iter_rows(values_only=True))[1:]
    rec = defaultdict(lambda: {"m": 0, "w": 0, "l": 0, "d": 0, "pf": 0, "pa": 0,
                               "first": None, "last": None})
    h2h = defaultdict(lambda: defaultdict(lambda: {"m": 0, "w": 0, "l": 0, "d": 0}))
    matches_by_team = defaultdict(list)
    rwc_final_results = {}  # year -> {winner, runner_up, score, venue}
    fixtures = 0

    for r in rrows:
        d, team, wld, opp, pf, pa = r[0], canon(r[1]), r[2], canon(r[3]), r[4], r[5]
        comp, stage = r[7], (str(r[8]).strip() if r[8] else "")
        stadium, city, country = r[10], r[11], r[12]
        if not team or not d:
            continue
        dstr = iso8(d)
        if pf is None or pa is None:
            fixtures += 1
            continue  # scheduled fixture, no result yet
        pf, pa = int(pf), int(pa)
        res = "W" if pf > pa else ("L" if pf < pa else "D")
        t = rec[team]
        t["m"] += 1
        t["w" if res == "W" else "l" if res == "L" else "d"] += 1
        t["pf"] += pf
        t["pa"] += pa
        if t["first"] is None or dstr < t["first"]:
            t["first"] = dstr
        if t["last"] is None or dstr > t["last"]:
            t["last"] = dstr
        if opp:
            o = h2h[team][opp]
            o["m"] += 1
            o["w" if res == "W" else "l" if res == "L" else "d"] += 1
        matches_by_team[team].append({
            "date": dstr, "opp": opp, "result": res, "score": f"{pf}-{pa}",
            "comp": str(comp or ""), "stage": stage,
            "venue": str(stadium or ""), "city": str(city or ""), "country": str(country or ""),
        })
        fam = comp_family(str(comp or ""))
        if fam == "RWC" and stage.lower() == "final" and res == "W":
            m = re.match(r"^(\d{4})", str(comp))
            year = int(m.group(1)) if m else int(dstr[:4])
            rwc_final_results[year] = {
                "year": year, "winner": team, "runner_up": opp,
                "score": f"{pf}-{pa}", "venue": str(stadium or ""), "city": str(city or ""),
            }

    # ---------------- Tables (standings + flags) ----------------
    ws = wb["Rugby Union - Intl Tables"]
    trows = list(ws.iter_rows(values_only=True))[1:]
    seasons_by_team = defaultdict(list)
    champ = defaultdict(lambda: {"5N6N_titles": 0, "5N6N_years": [], "grand_slams": 0,
                                 "triple_crowns": 0, "trc_titles": 0, "trc_years": []})
    rwc = defaultdict(lambda: {"apps": set(), "qf": 0, "sf": 0, "f": 0,
                               "titles": 0, "title_years": []})
    roll_5n6n = {}
    roll_trc = {}

    for r in trows:
        season, comp, pool, place, team = r[0], str(r[1] or ""), r[2], r[3], canon(r[4])
        if not team or season is None:
            continue
        season = int(str(season)[:4])
        fam = comp_family(comp)
        flags = {
            "rwc_qf": r[17] == "Y", "rwc_sf": r[18] == "Y", "rwc_f": r[19] == "Y",
            "trophy": r[20] == "Y", "triple_crown": r[21] == "Y", "grand_slam": r[22] == "Y",
        }
        seasons_by_team[team].append({
            "season": season, "comp": comp, "pool": str(pool or ""),
            "place": int(place) if place is not None and str(place).isdigit() else None,
            **{k: v for k, v in flags.items() if v},
        })
        if fam in ("5N", "6N"):
            c = champ[team]
            if flags["trophy"]:
                c["5N6N_titles"] += 1
                c["5N6N_years"].append(season)
                # shared titles: multiple Trophy=Y rows in one season are possible
                roll_5n6n.setdefault(season, {"season": season, "comp": comp, "champions": [],
                                              "grand_slam": None, "triple_crown": None})
                roll_5n6n[season]["champions"].append(team)
            if flags["grand_slam"]:
                c["grand_slams"] += 1
                roll_5n6n.setdefault(season, {"season": season, "comp": comp, "champions": [],
                                              "grand_slam": None, "triple_crown": None})
                roll_5n6n[season]["grand_slam"] = team
            if flags["triple_crown"]:
                c["triple_crowns"] += 1
                roll_5n6n.setdefault(season, {"season": season, "comp": comp, "champions": [],
                                              "grand_slam": None, "triple_crown": None})
                roll_5n6n[season]["triple_crown"] = team
        elif fam in ("TRI", "TRC"):
            if flags["trophy"]:
                c = champ[team]
                c["trc_titles"] += 1
                c["trc_years"].append(season)
                roll_trc.setdefault(season, {"season": season, "comp": comp, "champions": []})
                roll_trc[season]["champions"].append(team)
        elif fam == "RWC":
            w = rwc[team]
            w["apps"].add(season)
            if flags["rwc_qf"]:
                w["qf"] += 1
            if flags["rwc_sf"]:
                w["sf"] += 1
            if flags["rwc_f"]:
                w["f"] += 1
            if flags["trophy"]:
                w["titles"] += 1
                w["title_years"].append(season)

    # ---------------- World Rugby rankings (weekly, wide TSV) ----------------
    with open(rank_path, encoding="utf-8") as f:
        lines = [ln.rstrip("\n") for ln in f]
    header = None
    data = []
    for ln in lines:
        cells = ln.split("\t")
        if cells and cells[0].strip() == "Date":
            header = [canon(c) for c in cells[1:] if c.strip()]
            continue
        if header and cells[0].strip() and re.match(r"^\d{1,2}/\d{1,2}/\d{4}$", cells[0].strip()):
            mm, dd, yyyy = cells[0].strip().split("/")
            dstr = f"{yyyy}-{int(mm):02d}-{int(dd):02d}"
            ranks = {}
            for i, t in enumerate(header):
                v = cells[1 + i].strip() if 1 + i < len(cells) else ""
                if v.isdigit():
                    ranks[t] = int(v)
            data.append((dstr, ranks))
    data.sort(key=lambda x: x[0])
    rank_asof, current_ranks = data[-1]
    peaks = {}
    weeks_at_1 = defaultdict(int)
    reigns = []
    prev1 = None
    for dstr, ranks in data:
        one = min(ranks, key=lambda t: ranks[t]) if ranks else None
        if one and ranks[one] == 1:
            weeks_at_1[one] += 1
            if reigns and prev1 == one:
                reigns[-1]["end"] = dstr
                reigns[-1]["weeks"] += 1
            else:
                reigns.append({"team": one, "start": dstr, "end": dstr, "weeks": 1})
            prev1 = one
        for t, rk in ranks.items():
            p = peaks.setdefault(t, {"peak": rk, "peak_first": dstr})
            if rk < p["peak"]:
                p["peak"], p["peak_first"] = rk, dstr

    number_ones = []
    agg = defaultdict(lambda: {"weeks": 0, "reigns": 0, "longest": 0, "last": None})
    for rg in reigns:
        a = agg[rg["team"]]
        a["weeks"] += rg["weeks"]
        a["reigns"] += 1
        a["longest"] = max(a["longest"], rg["weeks"])
        a["last"] = max(a["last"] or "", rg["end"])
    number_ones = sorted([{"team": t, **v} for t, v in agg.items()], key=lambda x: -x["weeks"])

    # ---------------- Assemble teams ----------------
    all_names = sorted(set(rec) | set(seasons_by_team))
    slugs = {n: SLUG_OVERRIDE.get(n, slugify(n)) for n in all_names}
    teams = []
    for name in all_names:
        t = rec.get(name)
        w = rwc.get(name)
        c = champ.get(name)
        teams.append({
            "slug": slugs[name], "name": name,
            "six_nations": name in SIX_NATIONS,
            "sanzaar": name in SANZAAR,
            "record": t,
            "rwc": ({
                "apps": len(w["apps"]), "titles": w["titles"], "title_years": sorted(w["title_years"]),
                "finals": w["f"], "sf": w["sf"], "qf": w["qf"],
            } if w else None),
            "championships": ({
                "five_six_titles": c["5N6N_titles"], "five_six_years": sorted(c["5N6N_years"]),
                "grand_slams": c["grand_slams"], "triple_crowns": c["triple_crowns"],
                "trc_titles": c["trc_titles"], "trc_years": sorted(c["trc_years"]),
            } if c else None),
            "ranking": ({
                "current": current_ranks.get(name),
                "peak": peaks.get(name, {}).get("peak"),
                "peak_first": peaks.get(name, {}).get("peak_first"),
                "weeks_at_1": weeks_at_1.get(name, 0),
            } if name in peaks or name in current_ranks else None),
        })

    n_matches = sum(t["record"]["m"] for t in teams if t["record"]) // 2
    firsts = [t["record"]["first"] for t in teams if t["record"] and t["record"]["first"]]
    lasts = [t["record"]["last"] for t in teams if t["record"] and t["record"]["last"]]

    rwc_finals = []
    for year in sorted(rwc_final_results):
        rwc_finals.append(rwc_final_results[year])

    hub = {
        "as_of": rank_asof,
        "rankings_as_of": rank_asof,
        "totals": {"matches": n_matches, "teams": len(teams),
                   "first": min(firsts), "last": max(lasts),
                   "scheduled": fixtures // 2},
        "world_rankings": sorted(
            [{"team": t2, "rank": rk} for t2, rk in current_ranks.items()],
            key=lambda x: x["rank"]),
        "number_ones": number_ones,
        "six_nations_roll": [roll_5n6n[s] for s in sorted(roll_5n6n, reverse=True)],
        "trc_roll": [roll_trc[s] for s in sorted(roll_trc, reverse=True)],
        "rwc_finals": rwc_finals,
    }

    os.makedirs(os.path.join(out_dir, "team-detail"), exist_ok=True)
    with open(os.path.join(out_dir, "teams.json"), "w") as f:
        json.dump(teams, f, separators=(",", ":"))
    with open(os.path.join(out_dir, "hub.json"), "w") as f:
        json.dump(hub, f, separators=(",", ":"))
    for name in all_names:
        ml = sorted(matches_by_team.get(name, []), key=lambda m: m["date"] or "")
        detail = {
            "slug": slugs[name], "name": name,
            "recent": ml[-12:][::-1],
            "seasons": sorted(seasons_by_team.get(name, []),
                              key=lambda s: (-s["season"], s["comp"])),
            "h2h": {
                opp: v for opp, v in sorted(h2h[name].items(), key=lambda kv: -kv[1]["m"]) if opp
            },
        }
        with open(os.path.join(out_dir, "team-detail", slugs[name] + ".json"), "w") as f:
            json.dump(detail, f, separators=(",", ":"))

    print("teams:", len(teams), "matches:", n_matches, "fixtures(2-row):", fixtures // 2)
    print("rankings as-of:", rank_asof, "current #1:", hub["world_rankings"][0])
    print("RWC finals:", len(rwc_finals), rwc_finals[-1] if rwc_finals else None)
    print("weeks at 1:", number_ones[:4])


if __name__ == "__main__":
    main(sys.argv[1], sys.argv[2], sys.argv[3])

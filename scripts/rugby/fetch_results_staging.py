#!/usr/bin/env python3
"""Stage new international rugby union results + fixtures for the workbook.

Pulls men's internationals from World Rugby's own feed (the Pulselive API
behind world.rugby/fixtures-results, verified 2026-06-11):

  GET https://api.wr-rims-prod.pulselive.com/rugby/v3/match
      ?states=C|U&sport=mru&startDate=&endDate=&page=&pageSize=&sort=asc

and emits paste-ready rows for the "Rugby Union - Intl Results" sheet in
OtherLeagues.xlsx, in the sheet's exact column order. This is deliberately a
STAGING tool, not an auto-writer: the workbook is the source of truth and the
competition labels / flag columns are editorial, so you review the TSV, paste
it into the OneDrive master, copy the formula columns (W/L/D, Diff, LBP)
down, then sync the repo copy as usual.

The output has two sections:
  NEW ROWS        - matches not in the sheet at all (results and fixtures)
  SCORE UPDATES   - completed matches whose fixture rows already exist in the
                    sheet with blank scores; paste only the PF/PA values

Scope filter: only matches where BOTH sides are nations already tracked in
public/data/rugby-union/teams.json (full tests; excludes "XV" / "A" sides and
club rugby). Matches where exactly one side is tracked are listed separately
as candidates so new opponents are a conscious decision, never a guess.

Usage (from the repo root, host python):
  python scripts/rugby/fetch_results_staging.py            # since last completed row
  python scripts/rugby/fetch_results_staging.py --since 2026-03-15
Output: %TEMP%/rugby-staging.tsv (path printed; nothing is committed).
"""
import argparse
import io
import json
import os
import sys
import urllib.parse
import urllib.request
from datetime import date, datetime, timedelta, timezone

API = "https://api.wr-rims-prod.pulselive.com/rugby/v3/match"
UA = ("MetroPowerRankingsBot/1.0 (rugby results staging; "
      "https://github.com/ashwin-desikan) python-urllib")

REPO = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
TEAMS_JSON = os.path.join(REPO, "public", "data", "rugby-union", "teams.json")
WORKBOOK = os.path.join(REPO, "OtherLeagues.xlsx")
RESULTS_SHEET = "Rugby Union - Intl Results"

# Pulselive name -> workbook name where they diverge.
NAME_ALIASES = {
    "USA": "United States",
    "United States of America": "United States",
    "Cote d'Ivoire": "Ivory Coast",
}

# Competition label -> (workbook competition template, flag column).
# {y} is the season year. Anything not matched stays as "{y} <label>" with
# Test Match=Y and a REVIEW note so naming conventions remain yours.
COMP_RULES = [
    ("six nations", "{y} Six Nations Championship", "HN"),
    ("rugby championship", "{y} Rugby Championship", "TRC"),
    ("tri nations", "{y} Tri Nations Series", "TRC"),
    ("rugby world cup", "{y} Rugby World Cup", "RWC"),
    ("nations championship", "{y} Nations Championship", "NC"),
    # SA v NZ standalone series in NC years (replaces their paused Rugby
    # Championship meetings; user-approved in scope 2026-06-12).
    ("greatest rivalry", "{y} Rugby's Greatest Rivalry", "TEST"),
]

# Tracked competitions only (user scope, 2026-06-12): Six Nations, Rugby
# Championship / Tri Nations, summer tours, autumn internationals, Rugby World
# Cup (incl. warm-ups), and the Nations Championship. Everything else (Rugby
# Europe Championship, World Rugby Nations Cup, Pacific Nations Cup, ...) is
# reported as skipped, never staged.
IN_SCOPE = [
    "six nations", "rugby championship", "tri nations", "rugby world cup",
    "nations championship", "men's internationals", "autumn", "summer",
    "end-of-year", "tour of", "greatest rivalry",
]


def comp_in_scope(label):
    low = (label or "").lower()
    return any(p in low for p in IN_SCOPE)


# Sheet columns (23): date, Team, W/L/D(formula), Opp, PF, PA, Diff(formula),
# competition, Stage, Pool, stadium, city, country, Home/Away, HN, TriRC, NC,
# RWC, Test, Round, LBP(formula), blank, blank
HEADER = ["date (YYYYMMDD)", "Team", "W/L/D", "Opp Team", "PF", "PA", "Diff",
          "competition", "Stage", "Pool", "stadium", "city", "country",
          "Home/Away", "Home/Five/Six Nations", "Tri Nations/Rugby Champ",
          "Nations Championship", "Rugby World Cup", "Test Match", "Round",
          "Losing Bonus Point", "", ""]


def canon(name):
    name = (name or "").strip()
    return NAME_ALIASES.get(name, name)


def api_get(params):
    url = API + "?" + urllib.parse.urlencode(params)
    req = urllib.request.Request(url, headers={"User-Agent": UA})
    with urllib.request.urlopen(req, timeout=60) as resp:
        return json.loads(resp.read().decode("utf-8"))


def fetch_matches(states, start, end):
    if start > end:
        return []
    out, page = [], 0
    while True:
        j = api_get({"states": states, "sport": "mru", "startDate": start,
                     "endDate": end, "page": page, "pageSize": 100, "sort": "asc"})
        out.extend(j.get("content", []))
        info = j.get("pageInfo", {})
        page += 1
        if page >= int(info.get("numPages", 1)):
            return out


def read_sheet_state():
    """(last completed date YYYYMMDD, {(date, team, opp): has_scores})."""
    from openpyxl import load_workbook
    wb = load_workbook(WORKBOOK, read_only=True, data_only=True)
    ws = wb[RESULTS_SHEET]
    last_completed, existing = None, {}
    for r in ws.iter_rows(values_only=True):
        v = str(r[0] or "").strip()
        if len(v) < 8 or not v[:8].isdigit():
            continue
        d, team, opp = v[:8], str(r[1] or "").strip(), str(r[3] or "").strip()
        has_scores = r[4] is not None and str(r[4]).strip() != ""
        existing[(d, team, opp)] = has_scores
        if has_scores and (last_completed is None or d > last_completed):
            last_completed = d
    return last_completed, existing


def comp_fields(label):
    low = (label or "").lower()
    year = ""
    for tok in (label or "").split():
        if tok[:4].isdigit() and len(tok) >= 4:
            year = tok[:4]
    for needle, template, flag in COMP_RULES:
        if needle in low:
            return template.format(y=year), flag, False
    base = " ".join(t for t in (label or "").split() if t[:4] != year)
    return (f"{year} {base}".strip() if year else (label or "")), "TEST", True


def to_date(m):
    millis = (m.get("time") or {}).get("millis")
    if not millis:
        return None
    return datetime.fromtimestamp(millis / 1000, tz=timezone.utc).strftime("%Y%m%d")


def match_rows(m, tracked):
    """Two perspective rows (workbook convention) or None if out of scope."""
    teams = [canon(t.get("name")) for t in (m.get("teams") or [])]
    if len(teams) != 2 or not all(teams):
        return None, teams
    if not all(t in tracked for t in teams):
        return None, teams
    d = to_date(m)
    if not d:
        return None, teams
    scores = m.get("scores") or [None, None]
    completed = (m.get("status") == "C" and scores[0] is not None
                 and scores[1] is not None)
    venue = m.get("venue") or {}
    comp, flag, review = comp_fields(m.get("competition")
                                     or ";".join(e.get("label", "") for e in m.get("events") or []))
    stage = (m.get("eventPhase") or "").strip()
    rows = []
    for i in (0, 1):
        pf = int(scores[i]) if completed else ""
        pa = int(scores[1 - i]) if completed else ""
        rows.append([
            d, teams[i], "", teams[1 - i], pf, pa, "",
            comp, stage, "",
            venue.get("name") or "", venue.get("city") or "", venue.get("country") or "",
            "Home" if i == 0 else "Away",
            "Y" if flag == "HN" else "",
            "Y" if flag == "TRC" else "",
            "Y" if flag == "NC" else "",
            "Y" if flag == "RWC" else "",
            "Y" if flag == "TEST" else "",
            "", "", "", "",
        ])
    return (rows, review, completed, d, teams), teams


def main():
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass
    ap = argparse.ArgumentParser()
    ap.add_argument("--since", help="YYYY-MM-DD (default: day after last completed row)")
    ap.add_argument("--horizon", type=int, default=180,
                    help="days ahead to pull fixtures (default 180)")
    args = ap.parse_args()

    teams = json.load(io.open(TEAMS_JSON, encoding="utf-8"))
    tracked = {t["name"] for t in teams}
    last_completed, existing = read_sheet_state()

    if args.since:
        start = args.since
    elif last_completed:
        start = (datetime.strptime(last_completed, "%Y%m%d")
                 + timedelta(days=1)).strftime("%Y-%m-%d")
    else:
        sys.exit("Could not read the sheet; pass --since.")
    today = date.today().isoformat()
    horizon = (date.today() + timedelta(days=args.horizon)).isoformat()

    print(f"Completed results {start} -> {today}; fixtures {today} -> {horizon}")
    completed_ms = fetch_matches("C", start, today)
    upcoming_ms = fetch_matches("U", today, horizon)

    new_rows, update_rows, reviews, candidates, out_of_scope = [], [], 0, [], []
    seen = set()
    for m in completed_ms + upcoming_ms:
        if not comp_in_scope(m.get("competition")):
            names_ = [canon(t.get("name")) for t in (m.get("teams") or [])]
            if all(n in tracked for n in names_) and len(names_) == 2:
                out_of_scope.append(f"  {to_date(m) or '?'}  {' v '.join(names_)}"
                                    f"  [{m.get('competition')}]")
            continue
        res, names = match_rows(m, tracked)
        if not res:
            if sum(1 for t in names if t in tracked) == 1:
                d = to_date(m) or "?"
                candidates.append(f"  {d}  {' v '.join(str(n) for n in names)}"
                                  f"  [{m.get('competition')}]")
            continue
        rows, review, is_completed, d, pair = res
        key = (d, pair[0], pair[1])
        if key in seen:
            continue
        seen.add(key)
        in_sheet = key in existing
        if in_sheet and existing[key]:
            continue  # already entered with scores
        if in_sheet and not existing[key]:
            if is_completed:
                update_rows.extend(rows)  # fixture row exists; paste scores
                reviews += 1 if review else 0
            continue  # still a future fixture, already in sheet
        new_rows.extend(rows)
        reviews += 1 if review else 0

    out_path = os.path.join(os.environ.get("TEMP", "/tmp"), "rugby-staging.tsv")
    with io.open(out_path, "w", encoding="utf-8", newline="\n") as f:
        f.write("### NEW ROWS (paste as new lines)\n")
        f.write("\t".join(HEADER) + "\n")
        for row in new_rows:
            f.write("\t".join(str(c) for c in row) + "\n")
        if update_rows:
            f.write("\n### SCORE UPDATES (fixture rows already in the sheet; "
                    "fill PF/PA on the existing rows)\n")
            f.write("\t".join(HEADER) + "\n")
            for row in update_rows:
                f.write("\t".join(str(c) for c in row) + "\n")

    print(f"Staged {len(new_rows) // 2} new matches and "
          f"{len(update_rows) // 2} score updates -> {out_path}")
    if reviews:
        print(f"REVIEW: {reviews} match(es) use an unmapped competition label "
              "(kept as-is with Test Match=Y); align naming before pasting.")
    if candidates:
        print("Candidates involving one tracked nation (not staged; add the "
              "opponent to the sheet first if you want them):")
        for c in candidates[:20]:
            print(c)
    if out_of_scope:
        print("Skipped (tracked nations, but competition out of scope):")
        for c in out_of_scope[:20]:
            print(c)
    print("Paste into the OneDrive master Results sheet, copy the W/L/D, "
          "Diff and LBP formulas down, set Home/Away for neutral venues, "
          "then sync the repo workbook copy.")


if __name__ == "__main__":
    main()

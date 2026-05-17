#!/usr/bin/env python3
"""
Build NHL team-pages data from NHL.xlsx.

Mirrors scripts/build-nba-data.py and scripts/build-mlb-data.py. NHL-specific
notes inline below.

Outputs:
  public/data/nhl/franchises.json
  public/data/nhl/historical.json                  defunct franchises
  public/data/nhl/championships.json               one entry per Stanley Cup
                                                   or Avco Cup win per franchise
                                                   (era: 'stanley' | 'avco')
  public/data/nhl/championship-appearances.json    Stanley Cup Finals (and pre-
                                                   1926 Cup Final) appearances
  public/data/nhl/stadium-history.json             arena-era rows per franchise
  public/data/nhl/award-winners.json               Hart / Norris / Vezina /
                                                   Calder / Conn Smythe / Adams /
                                                   Selke / Lady Byng
  public/data/nhl/all-star-team-counts.json        career 1st / 2nd team All-Star
                                                   totals per franchise
  public/data/nhl/presidents-trophies.json         Best Reg. Record list per
                                                   franchise (workbook col R)
  public/data/nhl/seasons-by-team.json             per-franchise season-by-season
  public/data/nhl/historical-seasons.json          per-franchise seasons for defunct

Championship era split for chip rendering:
  - Stanley Cup (NHA + PCHA + WCHL + NHL, 1910 onwards): gold
  - Avco Cup / WHA: slate (rival league championship, parallel to ABA in NBA)
  - 1893-1909 Stanley Cup challenge era: NOT INCLUDED IN V1 by user instruction

Canonical join key: Name column in Year by Year (col BE / index 56).
Totals sheet uses 'Cur. Name' (col AL / index 37) for the same key.

Trophy spelling normalization (per workbook Claude Notes):
  - 'Adams ' (trailing space)  -> 'Adams'
  - 'Conn Symthe' (typo)       -> 'Conn Smythe'

OTL / T convention: workbook combines OTL + shootout loss in col I (OTL/SO).
Ties (col H) only appear pre-2005 (and in mixed 1999-2004 era). The
season-by-season output emits both columns; the team page picks which to
render per era.

Usage:
  python scripts/build-nhl-data.py
  python scripts/build-nhl-data.py /path/to/NHL.xlsx
"""

import json
import os
import re
import sys
from pathlib import Path
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl",
                           "--quiet", "--break-system-packages"])
    import openpyxl


# -------- Constants --------

REPO_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_SOURCE_CANDIDATES = [
    Path(os.path.expanduser("~/OneDrive/Excel Files/NHL.xlsx")),
    REPO_ROOT / "NHL.xlsx",
    REPO_ROOT / "data" / "nhl-source" / "NHL.xlsx",
]

OUT_DIR = REPO_ROOT / "public" / "data" / "nhl"

# Curated award list. Order is career-arc first, then specialty awards.
AWARD_ORDER = [
    "Hart",          # MVP
    "Norris",        # top defenseman
    "Vezina",        # top goalie
    "Calder",        # rookie
    "Conn Smythe",   # playoff MVP
    "Adams",         # coach of the year (Jack Adams)
    "Selke",         # defensive forward
    "Lady Byng",     # sportsmanship
]

# Normalize the typos flagged in Claude Notes. Map workbook string to clean
# display label. Anything not in this map plus AWARD_ORDER list is excluded.
TROPHY_NORMALIZE = {
    "hart": "Hart",
    "norris": "Norris",
    "vezina": "Vezina",
    "calder": "Calder",
    "conn smythe": "Conn Smythe",
    "conn symthe": "Conn Smythe",   # workbook typo
    "adams": "Adams",
    "adams ": "Adams",              # workbook trailing-space variant
    "selke": "Selke",
    "lady byng": "Lady Byng",
}

# All-Star team labels in Trophies sheet.
ALL_STAR_TEAM_NORMALIZE = {
    "all-star 1st team": "1st",
    "all-star 2nd team": "2nd",
}

# Display-name overrides (city-as-brand). Workbook tends to be honest here,
# but a few brand strings are conventional.
DISPLAY_NAME_OVERRIDES = {
    # canonical -> (display_city, display_team)
    # Most NHL franchises display as (workbook city, canonical team name).
    # Add overrides only when the marketing name diverges meaningfully.
}


# Championship era classifier driven by the league string on the
# championship row.
def championship_era(league):
    if not league:
        return "stanley"
    s = str(league).strip().upper()
    if s == "WHA":
        return "avco"
    return "stanley"


# -------- Helpers --------

def slugify(s):
    s = (s or "").lower().strip()
    s = re.sub(r"[àáâãäå]", "a", s)
    s = re.sub(r"[èéêë]", "e", s)
    s = re.sub(r"[ìíîï]", "i", s)
    s = re.sub(r"[òóôõö]", "o", s)
    s = re.sub(r"[ùúûü]", "u", s)
    s = re.sub(r"[ñ]", "n", s)
    s = re.sub(r"[^a-z0-9\s\-/()]", "", s)
    s = re.sub(r"[()/]+", "-", s)
    s = re.sub(r"\s+", "-", s)
    s = re.sub(r"-+", "-", s)
    return s.strip("-")


def franchise_slug(canonical):
    """Slugify a canonical Name. Parens (e.g. 'Maroons (V)') collapse to
    the alpha portion only so the URL stays readable."""
    return slugify((canonical or "").replace("(", " ").replace(")", " ")).rstrip("-")


def safe_int(val, default=0):
    if val is None or val == "":
        return default
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return default


def safe_float(val, default=0.0):
    if val is None or val == "":
        return default
    try:
        n = float(val)
        if n != n:
            return default
        return n
    except (ValueError, TypeError):
        return default


def safe_str(val):
    if val is None:
        return ""
    return str(val).strip()


def is_truthy_yn(val):
    if val is None or val == "":
        return False
    s = str(val).strip().upper()
    return s in ("Y", "YES", "1", "TRUE")


def normalize_trophy(raw):
    if not raw:
        return None
    low = str(raw).strip().lower()
    return TROPHY_NORMALIZE.get(low)


def normalize_all_star_team(raw):
    if not raw:
        return None
    low = str(raw).strip().lower()
    return ALL_STAR_TEAM_NORMALIZE.get(low)


# -------- Workbook loader --------

def find_source(cli_path=None):
    if cli_path:
        p = Path(cli_path)
        if p.exists():
            return p
        print(f"Provided path not found: {p}", file=sys.stderr)
        sys.exit(1)
    for cand in DEFAULT_SOURCE_CANDIDATES:
        if cand.exists():
            return cand
    print("Could not find NHL.xlsx. Pass an explicit path.", file=sys.stderr)
    sys.exit(1)


# -------- Sheet readers --------

def read_totals(wb):
    """One row per franchise. Keyed by Cur. Name (col AL / index 37)."""
    ws = wb["Totals"]
    out = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0 or not row:
            continue
        canonical = safe_str(row[37]) if len(row) > 37 else ""
        if not canonical:
            continue
        out[canonical] = {
            "league_history": safe_str(row[0]),
            "city_history": safe_str(row[1]),
            "team_history": safe_str(row[2]),
            "all_time_w": safe_int(row[3]),
            "all_time_l": safe_int(row[4]),
            "all_time_t": safe_int(row[5]),
            "all_time_otl": safe_int(row[6]),
            "all_time_pts": safe_int(row[7]),
            "pts_pct": safe_float(row[8]),
            "seasons": safe_int(row[9]),
            "five_hundred_seasons": safe_int(row[10]),
            "playoff_appearances": safe_int(row[11]),
            "division_titles": safe_int(row[12]),
            "best_main_div_seasons": safe_int(row[13]),
            "best_record_seasons": safe_int(row[14]),
            "playoff_w": safe_int(row[15]),
            "playoff_l": safe_int(row[16]),
            "playoff_t": safe_int(row[17]),
            "playoff_win_pct": safe_float(row[18]),
            "series_w": safe_int(row[19]),
            "series_l": safe_int(row[20]),
            "series_win_pct": safe_float(row[21]),
            "sf_appearances": safe_int(row[22]),
            "champ_appearances": safe_int(row[23]),
            "championships": safe_int(row[24]),
            "last_championship": safe_int(row[25]) or None,
            "last_champ_app": safe_int(row[26]) or None,
            "last_sf_app": safe_int(row[27]) or None,
            "last_series_win": safe_int(row[28]) or None,
            "last_best_rec": safe_int(row[29]) or None,
            "last_best_main_div": safe_int(row[30]) or None,
            "last_division_title": safe_int(row[31]) or None,
            "last_playoff_app": safe_int(row[32]) or None,
            "last_500_season": safe_int(row[33]) or None,
            "reg_games": safe_int(row[34]),
            "play_games": safe_int(row[35]),
            "total_games": safe_int(row[36]),
            "is_current": is_truthy_yn(row[38]) if len(row) > 38 else False,
            "is_defunct": is_truthy_yn(row[39]) if len(row) > 39 else False,
        }
    return out


def read_year_by_year(wb):
    """Build season-by-season per canonical franchise.

    Year by Year column map (0-indexed) — confirmed against workbook:
      1  League (NHA / PCHA / WCHL / WHA / NHL)
      2  Year (end year)
      3  City
      4  Team
      5  W
      6  L
      7  T (ties; pre-2005 + 1999-2004 mixed era)
      8  OTL/SO (combined overtime + shootout loss; post-1999, dominant post-2005)
      9  Points
      10 Pt. %
      11 GF
      12 GA
      13 .500+
      14 Playoff Appearance Y/N
      15 Div. Title Y/N
      16 Best Main Div Y/N
      17 Best Rec. Y/N        (Presidents' Trophy proxy across all eras)
      18 P. Wins
      19 P. Loss.
      20 P. Ties
      21 SF/CF App Y/N
      22 Cham App Y/N         (Stanley Cup Final appearance)
      23 Champs. Y/N          (Stanley Cup or Avco Cup win)
      24 Playoff Seed
      25 Division
      26 Place #
      28 Main Div             (conference / main division)
      29 Home Arena
      56 Name (canonical join key)
      76 Home City
      77 Metro Area
      78 Home State/Prov.
      90 Final/Current Arena Name
    """
    ws = wb["Year by Year"]
    by_team = defaultdict(list)
    latest_meta = {}
    earliest_year = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if not row or len(row) < 57:
            continue
        canonical = safe_str(row[56])
        if not canonical:
            continue
        year = safe_int(row[2])
        if year == 0:
            continue
        league = safe_str(row[1])
        city = safe_str(row[3])
        team = safe_str(row[4])
        w = safe_int(row[5])
        l = safe_int(row[6])
        t = safe_int(row[7])
        otl = safe_int(row[8])
        pts = safe_int(row[9])
        pts_pct = safe_float(row[10])
        gf = safe_int(row[11])
        ga = safe_int(row[12])
        playoff_yn = is_truthy_yn(row[14])
        div_title_yn = is_truthy_yn(row[15])
        best_div_yn = is_truthy_yn(row[16])
        best_rec_yn = is_truthy_yn(row[17])
        p_wins = safe_int(row[18])
        p_loss = safe_int(row[19])
        sf_cf_app_yn = is_truthy_yn(row[21])
        cham_app_yn = is_truthy_yn(row[22])
        champ_yn = is_truthy_yn(row[23])
        playoff_seed = safe_str(row[24]) or None
        division = safe_str(row[25])
        place_raw = row[26]
        place = "" if place_raw is None else str(place_raw).strip()
        main_div = safe_str(row[28]) if len(row) > 28 else ""
        home_arena_season = safe_str(row[29]) if len(row) > 29 else ""
        home_city = safe_str(row[76]) if len(row) > 76 else ""
        metro = safe_str(row[77]) if len(row) > 77 else ""
        home_state = safe_str(row[78]) if len(row) > 78 else ""
        home_arena_canonical = safe_str(row[90]) if len(row) > 90 else ""

        # Era flag: T column was meaningful pre-2005; OTL post-1999.
        # 1999-2004 had both. Post-2005 the user combined OTL+SO into one
        # column so OTL alone represents non-regulation losses.
        if year <= 1999:
            era = "T_only"
        elif year <= 2004:
            era = "T_OTL_mix"
        else:
            era = "OTL_only"

        by_team[canonical].append({
            "year": year,
            "league": league,
            "city": city,
            "team": team,
            "w": w,
            "l": l,
            "t": t,
            "otl": otl,
            "pts": pts,
            "pts_pct": round(pts_pct, 4) if pts_pct else 0.0,
            "gf": gf,
            "ga": ga,
            "gd": gf - ga,
            "playoff": playoff_yn,
            "div_title": div_title_yn,
            "best_main_div": best_div_yn,
            "best_rec_leag": best_rec_yn,
            "p_wins": p_wins,
            "p_loss": p_loss,
            "sf_cf_app": sf_cf_app_yn,
            "champ_app": cham_app_yn,
            "champ": champ_yn,
            "playoff_seed": playoff_seed,
            "division": division,
            "main_div": main_div,
            "place": place,
            "home_arena_season": home_arena_season,
            "home_arena_canonical": home_arena_canonical,
            "home_city": home_city,
            "metro": metro,
            "home_state": home_state,
            "era": era,
        })

        if canonical not in latest_meta or year > latest_meta[canonical]["year"]:
            latest_meta[canonical] = {
                "year": year,
                "city": city,
                "team": team,
                "league": league,
                "main_div": main_div,
                "division": division,
                "home_arena_season": home_arena_season,
                "home_arena_canonical": home_arena_canonical,
                "home_city": home_city,
                "metro": metro,
                "home_state": home_state,
            }
        earliest_year[canonical] = min(earliest_year.get(canonical, year), year)

    for k in by_team:
        by_team[k].sort(key=lambda s: s["year"])

    return dict(by_team), latest_meta, earliest_year


def read_trophies(wb):
    """Returns dict keyed by canonical Name -> {awards: [...], all_star_teams: {1st: int, 2nd: int}}.

    Trophies sheet columns:
      0=Year, 1=Player, 2=City, 3=Team, 4=Trophy, 5=Pos, 6=Name, 7=Tie
    """
    ws = wb["Trophies"]
    out_awards = defaultdict(list)
    out_as_counts = defaultdict(lambda: {"first": 0, "second": 0})
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0 or not row:
            continue
        year = safe_int(row[0])
        player = safe_str(row[1])
        team_workbook = safe_str(row[3])
        trophy_raw = safe_str(row[4])
        position = safe_str(row[5])
        canonical = safe_str(row[6])
        if not canonical or not player or not year:
            continue

        # All-Star team selection?
        as_tier = normalize_all_star_team(trophy_raw)
        if as_tier:
            if as_tier == "1st":
                out_as_counts[canonical]["first"] += 1
            elif as_tier == "2nd":
                out_as_counts[canonical]["second"] += 1
            continue

        # Major trophy?
        clean = normalize_trophy(trophy_raw)
        if clean and clean in AWARD_ORDER:
            out_awards[canonical].append({
                "year": year,
                "player": player,
                "trophy": clean,
                "position": position,
            })

    # Sort awards: by award order, then year descending.
    award_rank = {name: i for i, name in enumerate(AWARD_ORDER)}
    for k in out_awards:
        out_awards[k].sort(key=lambda a: (award_rank.get(a["trophy"], 99), -a["year"]))
    return dict(out_awards), dict(out_as_counts)


def build_stadium_history(year_by_year):
    """Construct arena-era rows by collapsing consecutive years at the same
    workbook-named home arena. Each row carries start_year, end_year, the
    as-of arena name, and the canonical / current name from col CM."""
    out = {}
    for canonical, seasons in year_by_year.items():
        rows = []
        cur = None
        for s in seasons:
            arena = s["home_arena_season"] or "(unknown)"
            canon = s["home_arena_canonical"] or arena
            home_city = s["home_city"] or s["city"]
            home_state = s["home_state"]
            if cur and cur["arena"] == arena and cur["arena_canonical"] == canon:
                cur["end_year"] = s["year"]
            else:
                if cur:
                    rows.append(cur)
                cur = {
                    "arena": arena,
                    "arena_canonical": canon,
                    "start_year": s["year"],
                    "end_year": s["year"],
                    "city": home_city,
                    "state": home_state,
                    "metro": s["metro"],
                }
        if cur:
            rows.append(cur)
        # Reverse-chronological for display.
        rows.sort(key=lambda r: r["start_year"], reverse=True)
        out[canonical] = rows
    return out


# -------- MetroAreas.xlsx Team List integration --------

def read_metro_team_list():
    """Read MetroAreas.xlsx Team List to pull NHL team -> metro slug + lat/lng.
    Returns dict keyed by team name (Team col, cleaned) -> meta."""
    metro_wb_path = REPO_ROOT / "MetroAreas.xlsx"
    if not metro_wb_path.exists():
        return {}
    wb = openpyxl.load_workbook(metro_wb_path, read_only=True, data_only=True)
    ws = wb["Team List"]
    out = {}
    # Cols shifted 2026-05-17: Gold Standard inserted at col L pushed Major
    # League to col M (idx 12), Wikidata QID to col Q (idx 16), Wikipedia
    # URL to col R (idx 17), Lat/Long to cols S/T (idx 18/19).
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0 or not row or len(row) < 20:
            continue
        sport = safe_str(row[0])
        league = safe_str(row[1])
        team = safe_str(row[2])
        city = safe_str(row[5])
        metro = safe_str(row[6])
        state = safe_str(row[7])
        country = safe_str(row[8])
        ml = safe_str(row[12])
        qid = safe_str(row[16])
        wiki = safe_str(row[17])
        lat = safe_float(row[18])
        lng = safe_float(row[19])
        if league != "NHL" or ml != "Y":
            continue
        # Extract the nickname (last word) so it matches canonical Name
        # column convention (e.g. 'Maple Leafs', 'Red Wings', 'Blue Jackets'
        # are >1 word so use the canonical key carefully).
        last_two = " ".join(team.split()[-2:]) if team.split() else ""
        last_one = team.split()[-1] if team.split() else ""
        out[team] = {
            "team_full": team,
            "city": city,
            "metro": metro,
            "metro_slug": slugify(metro) if metro else None,
            "state": state,
            "country": country,
            "wikidata_qid": qid or None,
            "wikipedia_url": wiki or None,
            "lat": lat or None,
            "lng": lng or None,
        }
    return out


# Manual fallback mapping for ambiguous canonical names. Maps canonical Name
# to the Team List 'team_full' string. Most canonical names match the last
# token / two tokens of the workbook Team string, but a few need an explicit
# pin.
CANONICAL_TO_TEAM_LIST = {
    # canonical -> Team List full string
    "Mammoth": "Utah Mammoth",
    "Hurricanes": "Carolina Hurricanes",
    "Avalanche": "Colorado Avalanche",
    "Oilers": "Edmonton Oilers",
    "Senators": "Ottawa Senators",
    "Maple Leafs": "Toronto Maple Leafs",
    "Red Wings": "Detroit Red Wings",
    "Blue Jackets": "Columbus Blue Jackets",
    "Golden Knights": "Vegas Golden Knights",
    "Maple Leafs": "Toronto Maple Leafs",
    "Kraken": "Seattle Kraken",
    "Bruins": "Boston Bruins",
    "Sabres": "Buffalo Sabres",
    "Wild": "Minnesota Wild",
    "Ducks": "Anaheim Ducks",
    "Flyers": "Philadelphia Flyers",
    "Penguins": "Pittsburgh Penguins",
    "Canadiens": "Montreal Canadiens",
    "Lightning": "Tampa Bay Lightning",
    "Stars": "Dallas Stars",
    "Rangers": "New York Rangers",
    "Islanders": "New York Islanders",
    "Devils": "New Jersey Devils",
    "Capitals": "Washington Capitals",
    "Panthers": "Florida Panthers",
    "Jets": "Winnipeg Jets",
    "Predators": "Nashville Predators",
    "Blues": "St. Louis Blues",
    "Blackhawks": "Chicago Blackhawks",
    "Canucks": "Vancouver Canucks",
    "Sharks": "San Jose Sharks",
    "Kings": "Los Angeles Kings",
    "Flames": "Calgary Flames",
}


def resolve_metro_meta(canonical, metro_team_list):
    """Find Team List meta for a canonical Name using the explicit map or
    a best-effort suffix match."""
    pin = CANONICAL_TO_TEAM_LIST.get(canonical)
    if pin and pin in metro_team_list:
        return metro_team_list[pin]
    # fallback: suffix match
    for team_full, meta in metro_team_list.items():
        if team_full.endswith(canonical):
            return meta
    return None


# -------- Builders --------

def build_franchises(totals, latest_meta, year_by_year, earliest_year,
                     metro_team_list):
    """Build the active-franchise list. Sorted by championship count desc,
    then by name."""
    out = []
    for canonical, t in totals.items():
        if not t["is_current"]:
            continue
        seasons = year_by_year.get(canonical, [])
        if not seasons:
            continue
        last = latest_meta.get(canonical, {})

        display_city = last.get("city") or ""
        display_team = last.get("team") or canonical
        override = DISPLAY_NAME_OVERRIDES.get(canonical)
        if override:
            display_city, display_team = override

        founded = earliest_year.get(canonical)
        slug = franchise_slug(canonical)

        metro_meta = resolve_metro_meta(canonical, metro_team_list)
        if metro_meta:
            metro = metro_meta["metro"]
            metro_slug = metro_meta["metro_slug"]
            state = metro_meta["state"]
            country = metro_meta["country"]
            wikidata_qid = metro_meta["wikidata_qid"]
            wikipedia_url = metro_meta["wikipedia_url"]
            lat = metro_meta["lat"]
            lng = metro_meta["lng"]
            home_city_marketing = metro_meta["city"]
        else:
            metro = last.get("metro", "")
            metro_slug = slugify(metro) if metro else None
            state = last.get("home_state", "")
            country = ""
            wikidata_qid = None
            wikipedia_url = None
            lat = None
            lng = None
            home_city_marketing = last.get("home_city") or display_city

        out.append({
            "slug": slug,
            "canonical": canonical,
            "name": canonical,
            "display_name": f"{display_city} {display_team}".strip(),
            "city": display_city,
            "team": display_team,
            "home_city": home_city_marketing,
            "metro": metro,
            "metro_slug": metro_slug,
            "state": state,
            "country": country,
            "founded": founded,
            "league_history": t["league_history"],
            "team_history": t["team_history"],
            "city_history": t["city_history"],
            "all_time_w": t["all_time_w"],
            "all_time_l": t["all_time_l"],
            "all_time_t": t["all_time_t"],
            "all_time_otl": t["all_time_otl"],
            "all_time_pts": t["all_time_pts"],
            "pts_pct": round(t["pts_pct"], 4) if t["pts_pct"] else 0.0,
            "seasons": t["seasons"],
            "playoff_appearances": t["playoff_appearances"],
            "division_titles": t["division_titles"],
            "best_record_seasons": t["best_record_seasons"],
            "best_main_div_seasons": t["best_main_div_seasons"],
            "sf_appearances": t["sf_appearances"],
            "champ_appearances": t["champ_appearances"],
            "championships": t["championships"],
            "last_championship": t["last_championship"],
            "last_champ_app": t["last_champ_app"],
            "last_sf_app": t["last_sf_app"],
            "last_best_rec": t["last_best_rec"],
            "last_division_title": t["last_division_title"],
            "last_playoff_app": t["last_playoff_app"],
            "current_arena_canonical": last.get("home_arena_canonical", ""),
            "current_arena_season": last.get("home_arena_season", ""),
            "current_main_div": last.get("main_div", ""),
            "current_division": last.get("division", ""),
            "wikidata_qid": wikidata_qid,
            "wikipedia_url": wikipedia_url,
            "lat": lat,
            "lng": lng,
        })

    out.sort(key=lambda f: (-f["championships"], f["name"]))
    return out


def build_historical(totals, year_by_year, earliest_year):
    """Defunct franchises. Includes pre-NHL league teams as well as defunct
    NHL franchises. Sorted by championship count desc, then by latest year
    desc."""
    out = []
    for canonical, t in totals.items():
        if t["is_current"]:
            continue
        seasons = year_by_year.get(canonical, [])
        if not seasons:
            continue
        last = seasons[-1]
        founded = earliest_year.get(canonical)
        ended = last["year"]
        out.append({
            "slug": franchise_slug(canonical),
            "canonical": canonical,
            "name": canonical,
            "last_city": last["city"],
            "last_team": last["team"],
            "founded": founded,
            "ended": ended,
            "league_history": t["league_history"],
            "team_history": t["team_history"],
            "city_history": t["city_history"],
            "all_time_w": t["all_time_w"],
            "all_time_l": t["all_time_l"],
            "all_time_t": t["all_time_t"],
            "all_time_otl": t["all_time_otl"],
            "all_time_pts": t["all_time_pts"],
            "championships": t["championships"],
            "champ_appearances": t["champ_appearances"],
            "sf_appearances": t["sf_appearances"],
            "seasons": t["seasons"],
            "last_championship": t["last_championship"],
        })
    out.sort(key=lambda f: (-f["championships"], -(f["ended"] or 0), f["name"]))
    return out


def build_championships(year_by_year):
    """One entry per Cup or Avco win per franchise."""
    out = defaultdict(list)
    for canonical, seasons in year_by_year.items():
        for s in seasons:
            if s["champ"]:
                era = championship_era(s["league"])
                out[canonical].append({
                    "year": s["year"],
                    "league": s["league"],
                    "era": era,
                    "city": s["city"],
                    "team": s["team"],
                })
        out[canonical].sort(key=lambda c: -c["year"])
    return dict(out)


def build_championship_appearances(year_by_year):
    """Stanley Cup Final (Cham App Y/N) appearances. Includes WHA Avco
    Final appearances tagged with era='avco'."""
    out = defaultdict(list)
    for canonical, seasons in year_by_year.items():
        for s in seasons:
            if s["champ_app"]:
                era = championship_era(s["league"])
                out[canonical].append({
                    "year": s["year"],
                    "league": s["league"],
                    "era": era,
                    "result": "Won" if s["champ"] else "Lost",
                    "city": s["city"],
                    "team": s["team"],
                })
        out[canonical].sort(key=lambda c: -c["year"])
    return dict(out)


def build_presidents_trophies(year_by_year):
    """Best Reg. Record (workbook col R) by franchise. This is the
    Presidents' Trophy proxy across all eras — the formal trophy started
    in 1985-86 but the workbook flags best regular-season record going
    further back, which the team page renders as 'Presidents' Trophy /
    Best Reg. Record' depending on era."""
    out = defaultdict(list)
    for canonical, seasons in year_by_year.items():
        for s in seasons:
            if s["best_rec_leag"]:
                out[canonical].append({
                    "year": s["year"],
                    "pts": s["pts"],
                    "w": s["w"],
                    "l": s["l"],
                    "t": s["t"],
                    "otl": s["otl"],
                    "league": s["league"],
                })
        out[canonical].sort(key=lambda r: -r["year"])
    return dict(out)


def build_seasons_by_team(franchises, year_by_year):
    out = {}
    slug_by_canonical = {f["canonical"]: f["slug"] for f in franchises}
    for canonical, seasons in year_by_year.items():
        slug = slug_by_canonical.get(canonical)
        if not slug:
            continue
        out[slug] = seasons
    return out


def build_historical_seasons(historical, year_by_year):
    out = {}
    for h in historical:
        canonical = h["canonical"]
        out[h["slug"]] = year_by_year.get(canonical, [])
    return out


# -------- Playoff state --------

def read_playoff_state(wb):
    """Read NHL.xlsx Year by Year sheet for the latest playoff year and
    infer each franchise's playoff state.

    Returns (year, is_postseason_complete, by_canonical) where by_canonical
    maps {canonical_team_name: {state, last_round, year}}. State values
    align with lib/nhl-playoffs.ts NhlPlayoffStateValue.

    Inference rules (NHL needs 4 series wins per round):
      - Champs Y                                                -> champion
      - Cham App Y, Champs None: in / lost Stanley Cup Final
          (lost_final if any champion exists this year, else active_final)
      - SF/CF App Y, Cham App None                              -> active_cf
      - Made playoffs, P.Wins < 4                               -> eliminated_qf
      - Made playoffs, P.Wins >= 4, SF/CF App None: still in R2.
          Use NHL bracket pairings (two R1 winners share a Division within
          the same Conference; the one with SF/CF App = Y won, the other is
          eliminated_semis). Two R1 winners in the same division with
          neither holding SF/CF App = Y -> both active_semis (series live).
    """
    ws = wb["Year by Year"]
    header = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    def col(name):
        return header.index(name) if name in header else -1
    idx = {n: col(n) for n in [
        "Year", "Team", "Play. Ap", "P. Wins", "P. Loss.",
        "SF/CF App", "Cham App", "Champs.", "Division", "Leag.",
    ]}

    # Use the latest year present with any playoff activity.
    latest = 0
    for r in ws.iter_rows(values_only=True):
        if not r: continue
        y = r[idx["Year"]]
        if isinstance(y, int) and y > latest and r[idx["Leag."]] == "NHL":
            latest = y
    year = latest

    teams = []
    for r in ws.iter_rows(values_only=True):
        if not r or r[idx["Year"]] != year or r[idx["Leag."]] != "NHL":
            continue
        teams.append({
            "canonical": r[idx["Team"]],
            "papp":  r[idx["Play. Ap"]],
            "pwin":  r[idx["P. Wins"]] or 0,
            "ploss": r[idx["P. Loss."]] or 0,
            "sf":    r[idx["SF/CF App"]],
            "ch":    r[idx["Cham App"]],
            "chmps": r[idx["Champs."]],
            "div":   r[idx["Division"]],
        })

    playoff_teams = [t for t in teams if (t["papp"] or "") == "Y"]
    r1_winners = [t for t in playoff_teams if t["pwin"] >= 4]

    from collections import defaultdict
    by_div = defaultdict(list)
    for t in r1_winners:
        by_div[t["div"]].append(t)

    results = {}
    for winners in by_div.values():
        if len(winners) != 2:
            for t in winners:
                results[t["canonical"]] = (
                    "active_cf" if (t["sf"] or "") == "Y" else "active_semis"
                )
            continue
        a, b = winners
        a_sf = (a["sf"] or "") == "Y"
        b_sf = (b["sf"] or "") == "Y"
        if a_sf and not b_sf:
            results[a["canonical"]] = "active_cf"
            results[b["canonical"]] = "eliminated_semis"
        elif b_sf and not a_sf:
            results[b["canonical"]] = "active_cf"
            results[a["canonical"]] = "eliminated_semis"
        elif a_sf and b_sf:
            results[a["canonical"]] = "active_cf"
            results[b["canonical"]] = "active_cf"
        else:
            results[a["canonical"]] = "active_semis"
            results[b["canonical"]] = "active_semis"

    for t in playoff_teams:
        if t["canonical"] not in results:
            results[t["canonical"]] = "eliminated_qf"

    any_champion = any((t["chmps"] or "") == "Y" for t in playoff_teams)
    for t in playoff_teams:
        if (t["chmps"] or "") == "Y":
            results[t["canonical"]] = "champion"
        elif (t["ch"] or "") == "Y":
            results[t["canonical"]] = "lost_final" if any_champion else "active_final"

    LABEL = {
        "champion":         "Stanley Cup Champion",
        "lost_final":       "Lost Stanley Cup Final",
        "active_final":     "In the Stanley Cup Final",
        "eliminated_cf":    "Eliminated Conference Finals",
        "active_cf":        "Conference Finals",
        "eliminated_semis": "Eliminated Conference Semifinals",
        "active_semis":     "Conference Semifinals",
        "eliminated_qf":    "Eliminated First Round",
    }
    STATE_RANK = {
        "champion": 0, "lost_final": 1, "active_final": 1,
        "eliminated_cf": 2, "active_cf": 2,
        "eliminated_semis": 3, "active_semis": 3,
        "eliminated_qf": 4,
    }
    sorted_keys = sorted(results.keys(), key=lambda k: (STATE_RANK[results[k]], k))
    by_canonical = {
        k: {"state": results[k], "last_round": LABEL[results[k]], "year": year}
        for k in sorted_keys
    }
    return year, any_champion, by_canonical


# -------- Main --------

def main():
    cli_path = sys.argv[1] if len(sys.argv) > 1 else None
    src = find_source(cli_path)
    print(f"Reading {src}")
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    totals = read_totals(wb)
    print(f"  Totals rows: {len(totals)}")

    year_by_year, latest_meta, earliest_year = read_year_by_year(wb)
    print(f"  Year by Year canonical franchises: {len(year_by_year)}")
    total_seasons = sum(len(s) for s in year_by_year.values())
    print(f"  Total season rows: {total_seasons}")

    awards_by_team, all_star_team_counts = read_trophies(wb)
    print(f"  Award rows distributed across {len(awards_by_team)} franchises")
    print(f"  All-Star Team rows distributed across {len(all_star_team_counts)} franchises")

    stadium_history = build_stadium_history(year_by_year)

    metro_team_list = read_metro_team_list()
    print(f"  MetroAreas Team List NHL rows: {len(metro_team_list)}")

    franchises = build_franchises(totals, latest_meta, year_by_year,
                                  earliest_year, metro_team_list)
    print(f"  Active franchises: {len(franchises)}")

    historical = build_historical(totals, year_by_year, earliest_year)
    print(f"  Defunct franchises: {len(historical)}")

    championships = build_championships(year_by_year)
    championship_appearances = build_championship_appearances(year_by_year)
    presidents_trophies = build_presidents_trophies(year_by_year)
    seasons_by_team = build_seasons_by_team(franchises, year_by_year)
    historical_seasons = build_historical_seasons(historical, year_by_year)

    # Re-key per-franchise outputs by slug for fast page lookup.
    def by_slug(d, franchises_list, historical_list):
        slug_by_canon = {f["canonical"]: f["slug"] for f in franchises_list}
        slug_by_canon.update({h["canonical"]: h["slug"] for h in historical_list})
        return {slug_by_canon[k]: v for k, v in d.items() if k in slug_by_canon}

    championships_slug = by_slug(championships, franchises, historical)
    apps_slug = by_slug(championship_appearances, franchises, historical)
    awards_slug = by_slug(awards_by_team, franchises, historical)
    pres_slug = by_slug(presidents_trophies, franchises, historical)
    all_star_team_slug = by_slug(all_star_team_counts, franchises, historical)
    stadium_history_slug = by_slug(stadium_history, franchises, historical)

    # Playoff state for the current season, inferred from Year by Year
    # P.Wins / P.Loss / SF/CF App / Cham App / Champs. columns. Refreshed on
    # every NHL workbook sync so the /teams/nhl bracket stays current.
    playoff_year, postseason_complete, playoff_by_canonical = read_playoff_state(wb)
    playoff_bundle = {
        "year": playoff_year,
        "is_postseason_complete": postseason_complete,
        "by_franchise": playoff_by_canonical,
    }
    print(f"  Playoff state: year={playoff_year}, {len(playoff_by_canonical)} teams flagged, complete={postseason_complete}")

    outputs = {
        "franchises.json": franchises,
        "historical.json": historical,
        "championships.json": championships_slug,
        "championship-appearances.json": apps_slug,
        "award-winners.json": awards_slug,
        "all-star-team-counts.json": all_star_team_slug,
        "presidents-trophies.json": pres_slug,
        "stadium-history.json": stadium_history_slug,
        "seasons-by-team.json": seasons_by_team,
        "historical-seasons.json": historical_seasons,
        "playoff-state.json": playoff_bundle,
    }

    for name, data in outputs.items():
        path = OUT_DIR / name
        with open(path, "w") as f:
            json.dump(data, f, ensure_ascii=False, indent=0, separators=(",", ":"))
        print(f"  Wrote {path}")

    # Summary stats
    cup_total = sum(len(v) for v in championships_slug.values())
    print(f"\nTotal championship rows: {cup_total}")
    pres_total = sum(len(v) for v in pres_slug.values())
    print(f"Total Best Reg. Record rows: {pres_total}")
    print("Done.")


if __name__ == "__main__":
    main()

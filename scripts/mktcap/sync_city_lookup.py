#!/usr/bin/env python3
"""
sync_city_lookup.py — one-way mirror: CompaniesMarketCap workbook -> Supabase.

City Lookup sheet   -> public.mktcap_geo          (source of truth for symbol -> metro)
Valid Metros sheet  -> public.mktcap_valid_metros (canonical metro names)

Modeled on the Champions League pattern (scripts/apifootball/sync_lookup.py):
the workbook always wins. Every run makes Supabase exactly reflect the sheet.

Differences from a blind truncate-and-reload, deliberate:
  1. auto-stub rows are PRESERVED. New companies that refresh.py queued for
     mapping (mapped_by='auto-stub') and that are not yet in City Lookup are
     the to-do queue; wiping them would lose it. As soon as a symbol appears
     in City Lookup, the workbook row replaces the stub.
  2. Provenance is kept useful. Unchanged rows are left alone (their original
     mapped_at/mapped_by survive); only added/changed rows are stamped
     mapped_at=today, mapped_by='excel-sync'. Pass --wipe to get the football
     behavior instead: delete everything (stubs included) and reload.
  3. Universal invariant, same as the football Lookup rule: every non-blank
     metro in City Lookup must exist in the Valid Metros sheet. Violations
     abort the sync (use --force to override, but fix the sheet instead).

House convention matches refresh.py: DRY RUN by default, --write to apply.

Usage (from scripts/mktcap on the Windows box, after the Excel ritual,
BEFORE refresh.py --write):

    python sync_city_lookup.py                 # dry run, shows the diff
    python sync_city_lookup.py --write         # apply
    python sync_city_lookup.py --write --wipe  # strict erase-and-reload

Requires: openpyxl, requests. Reads supabase_key.txt (service_role) from the
same directory, like refresh.py. IMPORTANT: the workbook must be SAVED after
the ritual; this reads Excel's cached values, not live formulas.
"""

import argparse
import csv
import datetime as dt
import os
import sys

import requests
import openpyxl

# ---------------------------------------------------------------- config ----

SUPABASE_URL = "https://nmprqkmymrdknffwnuur.supabase.co"
KEY_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "supabase_key.txt")

# Adjust if the workbook lives elsewhere / is renamed.
DEFAULT_WORKBOOK = r"C:\Users\ashwi\OneDrive\Excel Files\companiesmarketcap.com - Companies ranked by Market Cap - CompaniesMarketCap.com (1).xlsx"

GEO_TABLE = "mktcap_geo"
METROS_TABLE = "mktcap_valid_metros"
CITY_LOOKUP_SHEET = "City Lookup"
VALID_METROS_SHEET = "Valid Metros"
BATCH = 500

# ------------------------------------------------------------- postgrest ----


def _key():
    with open(KEY_FILE) as f:
        return f.read().strip()


def _headers(key):
    return {
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Content-Type": "application/json",
    }


def fetch_all(session, key, table, select="*"):
    rows, offset = [], 0
    while True:
        r = session.get(
            f"{SUPABASE_URL}/rest/v1/{table}",
            headers={**_headers(key), "Range": f"{offset}-{offset + 999}"},
            params={"select": select},
            timeout=60,
        )
        r.raise_for_status()
        page = r.json()
        rows.extend(page)
        if len(page) < 1000:
            return rows
        offset += 1000


def upsert(session, key, table, rows):
    for i in range(0, len(rows), BATCH):
        r = session.post(
            f"{SUPABASE_URL}/rest/v1/{table}",
            headers={**_headers(key), "Prefer": "resolution=merge-duplicates"},
            json=rows[i : i + BATCH],
            timeout=120,
        )
        r.raise_for_status()


def delete_symbols(session, key, table, symbols):
    for i in range(0, len(symbols), 100):
        chunk = symbols[i : i + 100]
        quoted = ",".join('"' + s.replace('"', '\\"') + '"' for s in chunk)
        r = session.delete(
            f"{SUPABASE_URL}/rest/v1/{table}",
            headers=_headers(key),
            params={"symbol": f"in.({quoted})"},
            timeout=120,
        )
        r.raise_for_status()


def delete_metros(session, key, metros):
    for i in range(0, len(metros), 100):
        chunk = metros[i : i + 100]
        quoted = ",".join('"' + m.replace('"', '\\"') + '"' for m in chunk)
        r = session.delete(
            f"{SUPABASE_URL}/rest/v1/{METROS_TABLE}",
            headers=_headers(key),
            params={"metro": f"in.({quoted})"},
            timeout=120,
        )
        r.raise_for_status()


# -------------------------------------------------------------- workbook ----


def _clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def read_city_lookup(path):
    """Returns {symbol: row-dict} mirroring the sheet verbatim (no renaming,
    no entity unescaping — the sheet is the source of truth, sync doesn't edit)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[CITY_LOOKUP_SHEET]
    out = {}
    for r in ws.iter_rows(min_row=2, max_col=8, values_only=True):
        name, sym, metro, city, state, country, active, source = r
        sym = _clean(sym)
        if sym is None:
            continue
        if sym in out:
            print(f"  WARNING: duplicate symbol in City Lookup: {sym} (keeping first)")
            continue
        out[sym] = {
            "symbol": sym,
            "name": _clean(name),
            "metro": _clean(metro),
            "city": _clean(city),
            "state": _clean(state),
            "country": _clean(country),
            "is_active": (_clean(active) == "Active"),
            "source": _clean(source),
        }
    return out


def read_valid_metros(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[VALID_METROS_SHEET]
    metros = set()
    for r in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        m = _clean(r[0])
        if m:
            metros.add(m)
    return metros


# ------------------------------------------------------------------ main ----

COMPARE_COLS = ["name", "metro", "city", "state", "country", "is_active", "source"]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--workbook", default=DEFAULT_WORKBOOK)
    ap.add_argument("--write", action="store_true", help="apply changes (default: dry run)")
    ap.add_argument("--wipe", action="store_true",
                    help="strict football-style erase-and-reload (drops auto-stub queue and provenance)")
    ap.add_argument("--force", action="store_true", help="proceed despite metro validation failures")
    args = ap.parse_args()

    today = dt.date.today().isoformat()
    print(f"sync_city_lookup — {today} — {'WRITE' if args.write else 'DRY RUN'}")
    print(f"workbook: {os.path.abspath(args.workbook)}")

    sheet = read_city_lookup(args.workbook)
    valid = read_valid_metros(args.workbook)
    print(f"City Lookup: {len(sheet)} symbols, "
          f"{sum(1 for v in sheet.values() if v['metro'])} mapped | Valid Metros: {len(valid)}")

    # -- invariant: every mapped metro must be a valid metro ------------------
    bad = sorted({(v["symbol"], v["metro"]) for v in sheet.values()
                  if v["metro"] and v["metro"] not in valid})
    if bad:
        print(f"\nINVARIANT VIOLATION: {len(bad)} City Lookup metros not in Valid Metros:")
        for sym, m in bad[:20]:
            print(f"  {sym}: {m!r}")
        if len(bad) > 20:
            print(f"  ... and {len(bad) - 20} more")
        if not args.force:
            print("Fix the sheet (or --force). Nothing written.")
            sys.exit(1)

    session = requests.Session()
    key = _key()

    # -- valid metros mirror --------------------------------------------------
    db_metros = {r["metro"] for r in fetch_all(session, key, METROS_TABLE, "metro")}
    metros_add = sorted(valid - db_metros)
    metros_del = sorted(db_metros - valid)
    print(f"\n{METROS_TABLE}: +{len(metros_add)} / -{len(metros_del)}")

    # -- geo mirror -----------------------------------------------------------
    db = {r["symbol"]: r for r in fetch_all(session, key, GEO_TABLE)}
    adds, updates, keeps = [], [], 0
    for sym, row in sheet.items():
        cur = db.get(sym)
        if cur is None:
            adds.append({**row, "mapped_at": today, "mapped_by": "excel-sync"})
        elif any(cur.get(c) != row[c] for c in COMPARE_COLS):
            updates.append({**row, "mapped_at": today, "mapped_by": "excel-sync"})
        else:
            keeps += 1

    orphans = [s for s in db if s not in sheet]
    stubs = [s for s in orphans if db[s].get("mapped_by") == "auto-stub"]
    deletes = orphans if args.wipe else [s for s in orphans if db[s].get("mapped_by") != "auto-stub"]

    print(f"{GEO_TABLE}: {len(db)} in DB | +{len(adds)} add, ~{len(updates)} update, "
          f"={keeps} unchanged, -{len(deletes)} delete"
          f" ({len(stubs)} auto-stub rows {'DELETED (--wipe)' if args.wipe else 'preserved as queue'})")

    if updates:
        print("  sample updates:")
        for u in updates[:10]:
            old = db[u["symbol"]]
            changed = [f"{c}: {old.get(c)!r} -> {u[c]!r}" for c in COMPARE_COLS if old.get(c) != u[c]]
            print(f"    {u['symbol']}: " + "; ".join(changed))
    if deletes:
        print(f"  deleting (not in City Lookup): {deletes[:15]}{' ...' if len(deletes) > 15 else ''}")

    report = f"sync_city_lookup_report_{today}.csv"
    with open(report, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["action", "symbol", "detail"])
        for a in adds:
            w.writerow(["add", a["symbol"], a["metro"] or ""])
        for u in updates:
            old = db[u["symbol"]]
            w.writerow(["update", u["symbol"],
                        "; ".join(f"{c}:{old.get(c)!r}->{u[c]!r}" for c in COMPARE_COLS
                                  if old.get(c) != u[c])])
        for s in deletes:
            w.writerow(["delete", s, db[s].get("mapped_by") or ""])
    print(f"  full diff written to {report}")

    if not args.write:
        print("\nDry run only. Re-run with --write to apply.")
        return

    # metros first, so geo rows never reference a metro the table lacks
    if metros_add:
        upsert(session, key, METROS_TABLE, [{"metro": m} for m in metros_add])
    if metros_del:
        delete_metros(session, key, metros_del)
    if deletes:
        delete_symbols(session, key, GEO_TABLE, deletes)
    if adds or updates:
        upsert(session, key, GEO_TABLE, adds + updates)

    # -- verify ---------------------------------------------------------------
    db_after = fetch_all(session, key, GEO_TABLE, "symbol,metro,mapped_by")
    mapped = sum(1 for r in db_after if r["metro"])
    stubs_after = sum(1 for r in db_after if r["mapped_by"] == "auto-stub")
    expect = len(sheet) + (0 if args.wipe else len(stubs))
    status = "OK" if len(db_after) == expect else "MISMATCH — investigate before refresh.py"
    print(f"\nDONE. {GEO_TABLE}: {len(db_after)} rows ({mapped} mapped, "
          f"{stubs_after} auto-stub queued), expected {expect}: {status}")


if __name__ == "__main__":
    main()
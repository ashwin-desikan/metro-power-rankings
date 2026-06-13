#!/usr/bin/env python3
"""Nippon Professional Baseball (NPB) portal data, nested under /teams/baseball.

Source: scripts/npb/npb_source.xlsx (a Wikipedia extract, 1950-2025; gitignored,
214KB). Sheets: Standings, Postseason_Games, Awards, Team_Name_Map,
Canonical_Teams. Historical franchises fold into the 12 current clubs via the
Canonical Current Team columns.

The Japan Series champion each year is derived robustly from the postseason
games BETWEEN THE TWO LEAGUE FINALISTS (one Central, one Pacific) — this ignores
the source's empty Winner column, its "Team - Runs" mashed cells, and its
mislabelling of Climax Series games as "Japan Series" (e.g. 2023). The winner is
read from the Winner column when present, else from game runs.

Outputs under public/data/npb/: teams.json, hub.json, team-detail/<slug>.json
Run from repo root: python scripts/npb/build_npb_data.py
"""
import io
import json
import os
import re
import unicodedata
from collections import Counter, defaultdict

from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(os.path.dirname(HERE))
XLSX = os.path.join(HERE, "npb_source.xlsx")
ALL_TEAMS = os.path.join(ROOT, "public", "data", "sports", "all-teams.json")
OUT = os.path.join(ROOT, "public", "data", "npb")

_RUNS = re.compile(r"^(.*?)\s*[−\-]\s*(\d+)\s*$")

# Verified Japan Series result (champion, runner-up) for the years the source
# mangled (empty team cells in 2009/2010; Climax Series games mislabelled as the
# Japan Series in 2005/2006/2014/2018). Only applied where the cross-league
# derivation below finds no champion. Canonical current-team names.
JS_OVERRIDE = {
    2005: ("Chiba Lotte Marines", "Hanshin Tigers"),
    2006: ("Hokkaido Nippon-Ham Fighters", "Chunichi Dragons"),
    2009: ("Yomiuri Giants", "Hokkaido Nippon-Ham Fighters"),
    2010: ("Chiba Lotte Marines", "Chunichi Dragons"),
    2014: ("Fukuoka SoftBank Hawks", "Hanshin Tigers"),
    2018: ("Fukuoka SoftBank Hawks", "Hiroshima Toyo Carp"),
}


def slugify(name):
    s = unicodedata.normalize("NFKD", name).encode("ascii", "ignore").decode()
    return re.sub(r"[^a-zA-Z0-9]+", "-", s).strip("-").lower()


def rows(ws):
    data = list(ws.iter_rows(values_only=True))
    hdr = [str(h).strip() if h is not None else "" for h in data[0]]
    return hdr, data[1:]


def main():
    import sys
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

    wb = load_workbook(XLSX, read_only=True, data_only=True)

    # Canonical teams -> circuit (Central / Pacific)
    chdr, crows = rows(wb["Canonical_Teams"])
    ci = {h: i for i, h in enumerate(chdr)}
    circuit = {}
    for r in crows:
        if r[ci["Canonical Current Team"]]:
            circuit[r[ci["Canonical Current Team"]]] = r[ci["Circuit"]]
    teams = list(circuit)
    # Kintetsu stayed SEPARATE from Orix (user choice): recognise it for the
    # Japan Series cross-league check and its own record, but keep it out of the
    # 12-club table (surfaced only as a defunct card on its Osaka metro).
    KINTETSU = "Kintetsu Buffaloes"
    circuit[KINTETSU] = "Pacific"

    # Metro for each club from the Team List (all-teams.json).
    metro_of = {}
    at = json.load(io.open(ALL_TEAMS, encoding="utf-8"))
    at = at if isinstance(at, list) else at.get("teams", [])
    for t in at:
        if t.get("league") == "NPB":
            nm = t.get("team") or t.get("name")
            metro_of[nm] = (t.get("city"), t.get("metro"), t.get("metro_slug"))

    # Name -> canonical (Team_Name_Map + Standings historical names)
    name2canon = {t: t for t in teams}
    mhdr, mrows = rows(wb["Team_Name_Map"])
    mi = {h: i for i, h in enumerate(mhdr)}
    hist_key = mhdr[0]
    for r in mrows:
        hist, can = r[0], r[mi["Canonical Current Team"]]
        if hist and can:
            name2canon[str(hist).strip()] = can

    def canon(name):
        if name is None:
            return None
        return name2canon.get(str(name).strip(), str(name).strip())

    # ---------------- Standings: pennants + per-team record ----------------
    shdr, srows = rows(wb["Standings"])
    si = {h: i for i, h in enumerate(shdr)}
    rec = defaultdict(lambda: {"w": 0, "l": 0, "t": 0, "seasons": set(), "first": None})
    pennants = defaultdict(list)        # team -> [years] (regular-season P1)
    standings_by_team = defaultdict(list)
    for r in srows:
        yr = r[si["Season"]]
        hist = r[si["Historical Team"]]
        team = canon(r[si["Canonical Current Team"]])
        # Kintetsu is kept separate from Orix (it folds to Orix in the source's
        # Canonical column, which would double-count Orix's 1950-2004 record).
        if hist and "Kintetsu" in str(hist):
            team = KINTETSU
        if not yr or team not in circuit:
            continue
        yr = int(yr)
        w = int(r[si["Wins"]] or 0); l = int(r[si["Losses"]] or 0); t = int(r[si["Ties"]] or 0)
        rc = rec[team]
        rc["w"] += w; rc["l"] += l; rc["t"] += t; rc["seasons"].add(yr)
        rc["first"] = yr if rc["first"] is None else min(rc["first"], yr)
        if str(r[si["Position"]]).strip() == "1":
            pennants[team].append(yr)
        standings_by_team[team].append({
            "year": yr, "team": (str(hist).strip() if hist else team),
            "league": r[si["League"]], "pos": r[si["Position"]],
            "w": w, "l": l, "t": t, "pct": r[si["Win Pct"]], "gb": r[si["GB"]],
        })

    # ---------------- Japan Series champions (cross-league games) ----------------
    phdr, prows = rows(wb["Postseason_Games"])
    pi = {h: i for i, h in enumerate(phdr)}

    def team_and_runs(name_cell, canon_cell, runs_cell):
        # Kintetsu kept separate from Orix even in postseason finalist parsing,
        # so its 4 Japan Series runner-up appearances are credited to Kintetsu.
        if name_cell and "Kintetsu" in str(name_cell):
            runs = int(runs_cell) if runs_cell not in (None, "") else None
            if runs is None:
                m = _RUNS.match(str(name_cell))
                if m:
                    runs = int(m.group(2))
            return KINTETSU, runs
        if canon_cell:
            c = canon(canon_cell)
            return c, (int(runs_cell) if runs_cell not in (None, "") else None)
        m = _RUNS.match(str(name_cell or ""))
        if m:
            return canon(m.group(1)), int(m.group(2))
        return canon(name_cell), (int(runs_cell) if runs_cell not in (None, "") else None)

    games_by_year = defaultdict(list)
    for r in prows:
        if str(r[pi["Series/Stage"]]).strip() != "Japan Series":
            continue
        yr = r[pi["Season"]]
        if not yr:
            continue
        c1, r1 = team_and_runs(r[pi["Team 1"]], r[pi["Team 1 Canonical"]], r[pi["Team 1 Runs"]])
        c2, r2 = team_and_runs(r[pi["Team 2"]], r[pi["Team 2 Canonical"]], r[pi["Team 2 Runs"]])
        if not c1 or not c2 or c1 not in circuit or c2 not in circuit:
            continue
        if circuit[c1] == circuit[c2]:
            continue  # same league -> a Climax Series game mislabelled; skip
        win = None
        if r[pi["Winner"]]:
            win = canon(r[pi["Winner"]])
        elif r1 is not None and r2 is not None and r1 != r2:
            win = c1 if r1 > r2 else c2
        games_by_year[int(yr)].append((c1, c2, win))

    japan_series = {}   # year -> (champion, runner_up)
    for yr, gs in games_by_year.items():
        wins = Counter()
        finalists = []
        for c1, c2, win in gs:
            for c in (c1, c2):
                if c not in finalists:
                    finalists.append(c)
            if win:
                wins[win] += 1
        champ = wins.most_common(1)[0][0] if wins else None
        if not champ:
            continue
        ru = None
        if len(finalists) == 2:
            ru = finalists[0] if finalists[1] == champ else finalists[1]
        japan_series[yr] = (champ, ru)
    # Fill the gap years the source mangled with the verified result.
    for yr, res in JS_OVERRIDE.items():
        japan_series.setdefault(yr, res)

    # The source folds Kintetsu into Orix in the postseason cells too, so some
    # Japan Series runner-up slots are mis-credited to Orix (e.g. 1979/1980/2001,
    # which were Kintetsu losses). Kintetsu never won the Series, so only the
    # runner-up needs correcting: if "Orix Buffaloes" is the runner-up in a year
    # it did NOT win its pennant but Kintetsu did, the real finalist was Kintetsu.
    orix_pen = set(pennants.get("Orix Buffaloes", []))
    kin_pen = set(pennants.get(KINTETSU, []))
    for yr, (champ, ru) in list(japan_series.items()):
        if ru == "Orix Buffaloes" and yr not in orix_pen and yr in kin_pen:
            japan_series[yr] = (champ, KINTETSU)

    js_title = defaultdict(list)
    js_ru = defaultdict(list)
    for yr, (champ, ru) in japan_series.items():
        if champ:
            js_title[champ].append(yr)
        if ru:
            js_ru[ru].append(yr)

    # ---------------- Defunct franchises (no modern club under that name) ----------------
    # These were excluded from the 12-club table; surfaced under the All filter.
    DEFUNCT = [
        {"name": "Kintetsu Buffaloes", "members": None, "division": "Pacific",
         "city": "Osaka", "metro": "Osaka-Kyoto-Kobe", "metro_slug": "osaka-kyoto-kobe",
         "successor": "Merged into Orix Buffaloes in 2005"},
        {"name": "Daiei Stars", "members": {"Daiei Stars", "Daiei Unions"}, "division": "Pacific",
         "city": "Tokyo", "metro": "Tokyo", "metro_slug": "tokyo",
         "successor": "Merged into the Mainichi Orions line (now Chiba Lotte Marines)"},
        {"name": "Takahashi Unions", "members": {"Takahashi Unions", "Tombo Unions"}, "division": "Pacific",
         "city": "Kawasaki", "metro": "Tokyo", "metro_slug": "tokyo",
         "successor": "Absorbed into the Daiei Unions in 1957"},
    ]
    defunct_rows = []
    for d in DEFUNCT:
        if d["members"] is None:  # Kintetsu — already aggregated via the remap
            rc = rec[KINTETSU]; yrs = set(rc["seasons"])
            w, l, t = rc["w"], rc["l"], rc["t"]
            pen = list(pennants[KINTETSU]); ru = list(js_ru[KINTETSU])
        else:
            yrs = set(); w = l = t = 0; pen = []; ru = []
            for r in srows:
                hist = r[si["Historical Team"]]; yr = r[si["Season"]]
                if not hist or not yr or str(hist).strip() not in d["members"]:
                    continue
                yr = int(yr); yrs.add(yr)
                w += int(r[si["Wins"]] or 0); l += int(r[si["Losses"]] or 0); t += int(r[si["Ties"]] or 0)
                if str(r[si["Position"]]).strip() == "1":
                    pen.append(yr)
        if not yrs:
            continue
        defunct_rows.append({
            "name": d["name"], "division": d["division"], "city": d["city"],
            "metro": d["metro"], "metro_slug": d["metro_slug"],
            "first_season": min(yrs), "last_season": max(yrs), "seasons": len(yrs),
            "js_titles": 0, "pennants": len(pen),
            "js_runnerup": len(ru), "js_ru_years": sorted(ru),
            "w": w, "l": l, "t": t,
            "win_pct": round(w / (w + l), 3) if w + l else 0,
            "successor": d["successor"], "defunct": True,
        })
    defunct_rows.sort(key=lambda r: (-r["seasons"], r["first_season"]))

    # ---------------- Assemble ----------------
    team_rows = []
    for team in teams:
        rc = rec[team]
        city, metro, metro_slug = metro_of.get(team, (None, None, None))
        team_rows.append({
            "slug": slugify(team), "name": team, "division": circuit[team],
            "city": city, "metro": metro, "metro_slug": metro_slug,
            "js_titles": len(js_title[team]), "js_title_years": sorted(js_title[team]),
            "js_runnerup": len(js_ru[team]), "js_ru_years": sorted(js_ru[team]),
            "pennants": len(pennants[team]), "pennant_years": sorted(pennants[team]),
            "seasons": len(rc["seasons"]), "first_season": rc["first"],
            "w": rc["w"], "l": rc["l"], "t": rc["t"],
            "win_pct": round(rc["w"] / (rc["w"] + rc["l"]), 3) if rc["w"] + rc["l"] else 0,
        })
    team_rows.sort(key=lambda r: (-r["js_titles"], -r["pennants"], -r["win_pct"], r["name"]))

    hub = {
        "japan_series": [{"year": y, "champion": japan_series[y][0],
                          "runner_up": japan_series[y][1]}
                         for y in sorted(japan_series, reverse=True)],
        "totals": {"teams": len(team_rows), "seasons": len({s for t in standings_by_team.values() for s in [x["year"] for x in t]}),
                   "js_editions": len(japan_series)},
    }

    os.makedirs(os.path.join(OUT, "team-detail"), exist_ok=True)
    json.dump(team_rows, io.open(os.path.join(OUT, "teams.json"), "w", encoding="utf-8", newline=""),
              separators=(",", ":"), ensure_ascii=False)
    json.dump(hub, io.open(os.path.join(OUT, "hub.json"), "w", encoding="utf-8", newline=""),
              separators=(",", ":"), ensure_ascii=False)
    json.dump(defunct_rows, io.open(os.path.join(OUT, "defunct.json"), "w", encoding="utf-8", newline=""),
              separators=(",", ":"), ensure_ascii=False)
    for t in team_rows:
        d = next(x for x in team_rows if x["slug"] == t["slug"])
        json.dump({
            "slug": t["slug"], "name": t["name"], "division": t["division"],
            "city": t["city"], "metro": t["metro"], "metro_slug": t["metro_slug"],
            "first_season": t["first_season"], "js_titles": t["js_titles"],
            "js_title_years": t["js_title_years"], "pennants": t["pennants"],
            "pennant_years": t["pennant_years"], "w": t["w"], "l": t["l"], "t": t["t"],
            "win_pct": t["win_pct"],
            "seasons": sorted(standings_by_team[t["name"]], key=lambda s: -s["year"]),
        }, io.open(os.path.join(OUT, "team-detail", t["slug"] + ".json"), "w",
                   encoding="utf-8", newline=""), separators=(",", ":"), ensure_ascii=False)

    print("NPB: %d teams | %d Japan Series | %d seasons" % (
        len(team_rows), len(japan_series), hub["totals"]["seasons"]))
    print("Defunct:", [(d["name"], "%d-%d" % (d["first_season"], d["last_season"]),
                        "%dG %dpen %dru" % (d["seasons"], d["pennants"], d["js_runnerup"]))
                       for d in defunct_rows])
    print("JS titles:", [(t["name"], t["js_titles"]) for t in team_rows])
    print("recent JS:", [(y, japan_series[y]) for y in sorted(japan_series, reverse=True)[:6]])


if __name__ == "__main__":
    main()

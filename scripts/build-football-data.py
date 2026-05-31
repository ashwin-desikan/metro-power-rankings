#!/usr/bin/env python3
"""
Build Football team-pages data from the grand Football workbook
(Champions League-201516.xlsx — the legacy filename hides a global database;
see the workbook's Claude Notes sheet for full schema).

V0 scope:
  - Level 1 top flights across England, Spain, Italy, Germany, France,
    Netherlands, Portugal, Scotland
  - English Levels 2-5 additionally (Championship, League One, League Two,
    National League and their historical predecessors)
  - Scottish Levels 2-4 additionally (Championship, League One, League Two
    and their historical predecessors in the SPFL pyramid)
  - One canonical page per distinct Cur. Name across the in-scope tiers
  - Season-by-season standings (P/W/D/L/Pts/GF/GA/GD/Place) + cup finals
    + European appearances + summary totals

Row-level format flag per season:
  - format = "league"   → normal round-robin standings; render all cells
  - format = "playoff"  → workbook only carries participant + finish; render
                          a pill with the finish position. Germany pre-1963
                          Deutsche Fußballmeisterschaft is purely this shape;
                          Italy pre-1929 and France pre-1929 are mixed.

Outputs (all under public/data/football/):
  index.json     - all canonical clubs with metadata + top-line totals
  seasons.json   - { slug: [season rows...] }
  cups.json      - { slug: [cup-final rows...] }
  europe.json    - { slug: [european-competition entries...] }
  leagues.json   - per-country league hub data (current standings + all-time champions)

Usage:
  python3 scripts/build-football-data.py
  python3 scripts/build-football-data.py /path/to/Champions\\ League-201516.xlsx
"""

import json
import os
import re
import sys
from pathlib import Path
from collections import defaultdict, Counter

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl",
                           "--quiet", "--break-system-packages"])
    import openpyxl


# ---------- Source resolution ----------

REPO_ROOT = Path(__file__).resolve().parent.parent

DEFAULT_SOURCE_CANDIDATES = [
    Path(os.path.expanduser("~/OneDrive/Excel Files/Champions League-201516.xlsx")),
    Path("/mnt/c/Users/ashwi/OneDrive/Excel Files/Champions League-201516.xlsx"),
]
# Add every current uploads mount dynamically (session token changes per run).
for sess in Path("/sessions").glob("*/mnt/uploads/Champions League-201516.xlsx"):
    DEFAULT_SOURCE_CANDIDATES.append(sess)
for sess in Path("/sessions").glob("*/mnt/Excel Files/Champions League-201516.xlsx"):
    DEFAULT_SOURCE_CANDIDATES.append(sess)

OUT_DIR = REPO_ROOT / "public" / "data" / "football"

IN_SCOPE_COUNTRIES = {"England", "Spain", "Italy", "Germany", "France",
                      "Netherlands", "Portugal", "Scotland"}

# Hand-curated coordinate overrides for clubs whose Lookup / FootballClub_Data
# rows carry wrong lat/lng (typically an Albany NY fallback from a US-centric
# geocoder). Keyed by slug. Workbook should be updated to match these values
# at some point; see BACKLOG.md "Workbook coordinate fixes".
CURATED_COORDINATE_OVERRIDES = {
    "siena": (43.321667, 11.326111),  # Stadio Artemio Franchi, Siena, Italy
}

# Mapping from country -> set of in-scope tier levels. England carries the
# full pyramid through National League (L1-5); Scotland carries the full
# SPFL pyramid (L1-4: Premiership, Championship, League One, League Two).
# Every other country is Level 1 only until a deliberate expansion lands.
COUNTRY_TIERS = {
    "England":     {1, 2, 3, 4, 5},
    "Spain":       {1},
    "Italy":       {1},
    "Germany":     {1},
    "France":      {1},
    "Netherlands": {1},
    "Portugal":    {1},
    "Scotland":    {1, 2, 3, 4},
}

# League hub slugs for the modern top-flight competitions across the in-scope
# countries.
LEAGUE_HUBS = [
    ("premier-league",       "England",     "Premier League",       "Premier League",       1),
    ("la-liga",              "Spain",       "La Liga",              "La Liga",              1),
    ("serie-a",              "Italy",       "Serie A",              "Serie A",              1),
    ("bundesliga",           "Germany",     "Bundesliga",           "Bundesliga",           1),
    ("ligue-1",              "France",      "Ligue 1",              "Ligue 1",              1),
    ("eredivisie",           "Netherlands", "Eredivisie",           "Eredivisie",           1),
    ("primeira-liga",        "Portugal",    "Primeira Liga",        "Primeira Liga",        1),
    ("scottish-premiership", "Scotland",    "Scottish Premiership", "Scottish Premiership", 1),
]

# Standings sheets share an identical 86-column schema.
STANDINGS_SHEETS = ["Leagues History", "Stand2nd"]


# ---------- Helpers ----------

def slugify(s):
    if s is None: return None
    s = str(s).strip().lower()
    repl = {"á":"a","à":"a","â":"a","ä":"a","ã":"a","å":"a",
            "é":"e","è":"e","ê":"e","ë":"e",
            "í":"i","ì":"i","î":"i","ï":"i",
            "ó":"o","ò":"o","ô":"o","ö":"o","õ":"o","ø":"o",
            "ú":"u","ù":"u","û":"u","ü":"u",
            "ñ":"n","ç":"c","ß":"ss",
            "ý":"y","ÿ":"y","ž":"z","š":"s","č":"c","ć":"c","ř":"r"}
    for k, v in repl.items():
        s = s.replace(k, v)
    s = re.sub(r"[^a-z0-9]+", "-", s).strip("-")
    return s or None


def header_map(ws):
    row1 = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    out = {}
    for i, h in enumerate(row1):
        if h is None: continue
        h = str(h).strip()
        # First wins for duplicate headers (sheets sometimes echo cols).
        if h not in out:
            out[h] = i
    return out, row1


def find_source():
    if len(sys.argv) >= 2:
        p = Path(sys.argv[1])
        if not p.exists(): sys.exit(f"FAIL: source not found: {p}")
        return p
    for c in DEFAULT_SOURCE_CANDIDATES:
        if c.exists(): return c
    sys.exit("FAIL: no source workbook found. Pass a path or place it under "
             "OneDrive/Excel Files/Champions League-201516.xlsx")


def to_int(v):
    if v is None: return None
    if isinstance(v, bool): return int(v)
    if isinstance(v, (int, float)):
        try: return int(v)
        except: return None
    try: return int(str(v).strip())
    except: return None


def to_float(v):
    if v is None: return None
    if isinstance(v, (int, float)): return float(v)
    try: return float(str(v).strip())
    except: return None


def detect_format(w, d, l, matches):
    """Row-level format detection. Pre-Bundesliga German rows have all of these
    blank → 'playoff'. Italian Football Championship rows often have small
    match counts but populated W → 'league' if matches looks round-robin,
    else 'playoff'. Rule: format=league iff matches is populated AND w is
    populated. Otherwise playoff."""
    if matches is None or matches == 0: return "playoff"
    if w is None: return "playoff"
    return "league"


# ---------- Wikidata QIDs from MetroAreas.xlsx ----------

def load_wikidata_qids():
    """Read Wikidata_QIDs sheet from MetroAreas.xlsx and return
    {team_name_lower: (qid, wikipedia_url)}. Prints a warning if the file
    cannot be opened (e.g. locked by Excel) and returns {} in that case."""
    candidates = [REPO_ROOT / "MetroAreas.xlsx",
                  Path(os.path.expanduser("~/OneDrive/Excel Files/MetroAreas.xlsx")),
                  Path("/mnt/c/Users/ashwi/OneDrive/Excel Files/MetroAreas.xlsx")]
    try:
        for sess in Path("/sessions").glob("*/mnt/uploads/MetroAreas.xlsx"):
            candidates.append(sess)
        for sess in Path("/sessions").glob("*/mnt/Excel Files/MetroAreas.xlsx"):
            candidates.append(sess)
    except Exception:
        pass
    xlsx = next((p for p in candidates if p.exists()), None)
    if not xlsx:
        print("  WARNING: MetroAreas.xlsx not found — Wikidata QIDs skipped.")
        return {}
    try:
        wb = openpyxl.load_workbook(str(xlsx), read_only=True, data_only=True)
    except Exception as e:
        print(f"  WARNING: Could not open MetroAreas.xlsx ({e}) — "
              "close it in Excel if open, then re-run. Wikidata QIDs skipped.")
        return {}
    if "Wikidata_QIDs" not in wb.sheetnames:
        print("  WARNING: Wikidata_QIDs sheet not found in MetroAreas.xlsx.")
        wb.close(); return {}
    ws = wb["Wikidata_QIDs"]
    lookup = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 4: continue
        team, qid, url = row[1], row[2], row[3]
        if not team or not qid: continue
        lookup[str(team).strip().lower()] = (
            str(qid).strip(),
            str(url).strip() if url else None,
        )
    wb.close()
    print(f"  Wikidata QIDs loaded: {len(lookup)} clubs from {xlsx.name}")
    return lookup


# ---------- ETL: clubs from Lookup ----------

def build_clubs_index(wb, in_scope_curnames, country_mode_by_cn=None):
    """Build the master club index from Lookup, restricted to in_scope_curnames.
    Collapses intra-country dupe rows (e.g. Hornchurch, Watford Rovers) by
    preferring the row with metro/lat/long populated. country_mode_by_cn,
    when provided, overrides Lookup's country with the mode of in-scope
    standings-row countries so wartime / cross-border clubs (Rapid Wien,
    AS Monaco, Cardiff, Swansea) sit under the country they actually played
    in within our Big 5 scope."""
    wikidata_lookup = load_wikidata_qids()
    ws = wb["Lookup"]
    hdr, _ = header_map(ws)
    # Per Claude Notes verified col indices: Cur. Name = M (12), Team = A (0),
    # City = B (1), Metro = C (2), County = D (3), Country = E (4),
    # Continent = K (10), Lat = U (20), Long = V (21), Club = P (15)
    idx_curname = hdr.get("Cur. Name", 12)
    idx_city = hdr.get("City", 1)
    idx_metro = hdr.get("Metro Area", 2)
    idx_county = hdr.get("County", 3)
    idx_country = hdr.get("Country", 4)
    idx_continent = hdr.get("Continent", 10)
    idx_club = hdr.get("Club", 15)
    idx_lat = 20
    idx_lng = 21

    # Lookup col M (Cur. Name) is a formula echo (=A typically) and the
    # workbook is often saved without cached values for these cells, so
    # read_only+data_only returns None. Fall back to col A (Team) which
    # the user maintains as the canonical current name for the primary
    # registry row of each club.
    candidates = defaultdict(list)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row: continue
        cn = row[idx_curname] if idx_curname < len(row) else None
        if not cn:
            cn = row[0] if len(row) > 0 else None  # col A: Team
        if not cn: continue
        cn = str(cn).strip()
        if cn not in in_scope_curnames: continue
        # Only Clubs (not National teams).
        club_flag = row[idx_club] if idx_club < len(row) else None
        if club_flag and str(club_flag).strip().lower() == "nat": continue
        candidates[cn].append(row)

    clubs = {}
    for cn, rows in candidates.items():
        # Pick the row with the most populated metro/lat fields.
        def score(r):
            s = 0
            for i in (idx_metro, idx_county, idx_lat, idx_lng, idx_city):
                if i < len(r) and r[i] not in (None, "", 0): s += 1
            return s
        rows.sort(key=score, reverse=True)
        r = rows[0]

        def cell(i):
            return r[i] if i < len(r) and r[i] not in ("",) else None

        # Country: standings-row mode wins (for wartime / cross-border
        # clubs); Lookup country is the fallback.
        lookup_country = cell(idx_country)
        country = (country_mode_by_cn or {}).get(cn, lookup_country)
        slug = slugify(cn)
        lat = to_float(cell(idx_lat))
        lng = to_float(cell(idx_lng))
        # Hand-curated overrides take precedence when the workbook coords
        # are known-bad (Siena was geocoded to Albany NY in both Lookup
        # and FootballClub_Data; permanent fix tracked in BACKLOG.md).
        override = CURATED_COORDINATE_OVERRIDES.get(slug)
        if override:
            lat, lng = override
        wd = wikidata_lookup.get(cn.lower(), (None, None))
        clubs[cn] = {
            "slug": slug,
            "cur_name": cn,
            "country": country,
            "federation_country": lookup_country if lookup_country != country else None,
            "city": cell(idx_city),
            "metro": cell(idx_metro),
            "county": cell(idx_county),
            "continent": cell(idx_continent),
            "lat": lat,
            "lng": lng,
            **({"wikidata_qid": wd[0]} if wd[0] else {}),
            **({"wikipedia_url": wd[1]} if wd[1] else {}),
        }
    return clubs


# ---------- ETL: standings rows ----------

def collect_standings_rows(wb):
    """Pull every in-scope standings row (Big 5 Level 1 + England 2-5) from
    Leagues History + Stand2nd. Returns (rows, in_scope_curnames)."""
    rows_out = []
    curnames = set()
    for sheet in STANDINGS_SHEETS:
        ws = wb[sheet]
        hdr, row1 = header_map(ws)
        # Cur. Name lives at col BV (idx 73) per Claude Notes; the in-sheet
        # header_map also catches it, but pin to 73 in case of header drift.
        idx_curname = 73
        idx_country = hdr.get("Country (Leag)")
        idx_league = hdr.get("League")
        idx_year = hdr.get("End Year")
        idx_team = hdr.get("Team")
        idx_place = hdr.get("Place #")
        idx_w = hdr.get("W")
        idx_d = hdr.get("D")
        idx_l = hdr.get("L")
        idx_pts = hdr.get("Points")
        idx_gf = hdr.get("GS")
        idx_ga = hdr.get("GA")
        idx_gdiff = hdr.get("G Diff")
        idx_matches = hdr.get("Matches")
        idx_level = None
        for i, h in enumerate(row1):
            if h == "Level":
                idx_level = i; break
        idx_eur_qual = hdr.get("Eur Qual")
        idx_relegated = hdr.get("Relegated")
        idx_div = hdr.get("Division")
        # Per Claude Notes BV-CA block: BW=Final (74), BX=Champions (75).
        # The Champions flag is the canonical national-champion signal; it
        # fires for both league-format Level 1 winners (place=1) AND for
        # pre-modern playoff/knockout champions where place is null. It
        # does NOT fire for Second Division winners (e.g. MU 1975) or
        # 2.Bundesliga winners (e.g. Schalke 2022), so it cleanly matches
        # the editorial spec: Champion pill = national champion only.
        idx_champion = 75
        idx_final = 74

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row: continue
            if idx_country is None or idx_country >= len(row): continue
            country = row[idx_country]
            if country not in IN_SCOPE_COUNTRIES: continue
            cn = row[idx_curname] if idx_curname < len(row) else None
            if not cn: continue
            cn = str(cn).strip()
            level = to_int(row[idx_level]) if idx_level is not None and idx_level < len(row) else None
            allowed = COUNTRY_TIERS.get(country, set())
            if level not in allowed: continue

            year = to_int(row[idx_year]) if idx_year < len(row) else None
            league = row[idx_league] if idx_league < len(row) else None
            team = row[idx_team] if idx_team < len(row) else None
            w = to_int(row[idx_w]) if idx_w < len(row) else None
            d = to_int(row[idx_d]) if idx_d < len(row) else None
            l = to_int(row[idx_l]) if idx_l < len(row) else None
            pts = to_int(row[idx_pts]) if idx_pts < len(row) else None
            gf = to_int(row[idx_gf]) if idx_gf < len(row) else None
            ga = to_int(row[idx_ga]) if idx_ga < len(row) else None
            gdiff = to_int(row[idx_gdiff]) if idx_gdiff < len(row) else None
            matches = to_int(row[idx_matches]) if idx_matches < len(row) else None
            place = to_int(row[idx_place]) if idx_place < len(row) else None
            division = row[idx_div] if idx_div is not None and idx_div < len(row) else None

            fmt = detect_format(w, d, l, matches)
            champion = True if (idx_champion < len(row) and row[idx_champion] == "Y") else False
            runner_up_final = True if (idx_final < len(row) and row[idx_final] == "Y" and not champion) else False
            # Workbook's Relegated column carries richer values than just
            # 'Y': 'Reg' (relegated), 'Prom' (promoted), 'Not Prom' (lost
            # in promotion playoff). Use it directly as a seed so the
            # latest completed season has a flag even when the next-year
            # tier transition hasn't been recorded yet (no year+1 data).
            wb_rel_raw = None
            if idx_relegated is not None and idx_relegated < len(row):
                wb_rel_raw = row[idx_relegated]
            wb_rel_str = str(wb_rel_raw).strip() if wb_rel_raw else ""
            wb_promoted = wb_rel_str in ("Prom", "Y-Prom")
            wb_relegated = wb_rel_str in ("Y", "Reg", "Y-Reg")
            rows_out.append({
                "slug": slugify(cn),
                "cur_name": cn,
                "year": year,
                "country": country,
                "league": str(league) if league else None,
                "division": str(division) if division else None,
                "level": level,
                "team": str(team) if team else cn,
                "place": place,
                "w": w, "d": d, "l": l,
                "pts": pts, "gf": gf, "ga": ga, "gd": gdiff,
                "matches": matches,
                "format": fmt,
                # Eur Qual carries the competition code ('CL', 'EL', 'EUCL',
                # 'CWC', etc.) for the season's UEFA qualification awarded
                # via the workbook. Empty / None means no qualification.
                "eur_qual": (str(row[idx_eur_qual]).strip() if (idx_eur_qual is not None and idx_eur_qual < len(row) and row[idx_eur_qual]) else None),
                # promoted / relegated start from the workbook's own flag.
                # The forward-scan pass below OVERRIDES this seed when the
                # next existing-year row for the same club shows a tier
                # transition (handles WWI / WWII gap years where Arsenal
                # 1915 sits at L2 with workbook 'Reg' but the next data
                # row 1920 puts them at L1 = actually promoted).
                "promoted": wb_promoted,
                "relegated": wb_relegated,
                # Workbook BX (Champions) flag = national champion only,
                # which is the right signal for the Champion pill since
                # second-division winners and playoff-format champs both
                # need consistent treatment.
                "champion": champion,
                "final": runner_up_final,
            })
            curnames.add(cn)
    return rows_out, curnames


# ---------- ETL: cup finals ----------

def collect_cup_finals(wb, in_scope_curnames):
    """Pull domestic cup finals for in-scope clubs from Cup History."""
    ws = wb["Cup History"]
    hdr, _ = header_map(ws)
    # Per Claude Notes: B=Key, C=League (country), D=End Year, E=Cur. Name,
    # F=Cup Major (Y win, blank lost, date if scheduled), G=Cup Final Major,
    # H=Cup Minor (Y win, blank lost), I=Cup Final Minor,
    # J=Super Cup (Y win), K=Super Cup Final, L=Continent
    idx_country = hdr.get("League", 2)
    idx_year = hdr.get("End Year", 3)
    idx_cn = hdr.get("Cur. Name", 4)
    idx_maj = hdr.get("Cup (Major Domestic)", 5)
    idx_maj_f = hdr.get("Cup Final (Major Domestic)", 6)
    idx_min = hdr.get("Cup (Minor Domestic)", 7)
    idx_min_f = hdr.get("Cup Final (Minor Domestic)", 8)
    idx_sup = hdr.get("Super Cup", 9)
    idx_sup_f = hdr.get("Super Cup Final", 10)

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row: continue
        country = row[idx_country] if idx_country < len(row) else None
        if country not in IN_SCOPE_COUNTRIES: continue
        cn = row[idx_cn] if idx_cn < len(row) else None
        if not cn or str(cn).strip() not in in_scope_curnames: continue
        cn = str(cn).strip()
        year = to_int(row[idx_year]) if idx_year < len(row) else None

        def status(win_cell, final_cell):
            """Return (kind, result) per Claude Notes col F convention.
            Y = won, blank = lost, date serial = scheduled (not yet played)."""
            if final_cell != "Y": return None
            if win_cell == "Y": return "won"
            if win_cell in (None, "", 0): return "lost"
            # Date serial → scheduled.
            if isinstance(win_cell, (int, float)) and win_cell > 1000:
                return "scheduled"
            return "lost"

        # Super cups happen at season start (August); the 2026 super cups
        # haven't been played yet as of May 2026, and the workbook can't
        # distinguish 'scheduled' from 'lost' for them (both finalists
        # show K='Y' but J=None on both sides). Drop super cup rows for
        # the upcoming season's year; bring them back when MAX_DISPLAYED_YEAR
        # ticks past 2026.
        for kind, win_idx, final_idx in (
            ("major", idx_maj, idx_maj_f),
            ("minor", idx_min, idx_min_f),
            ("super", idx_sup, idx_sup_f),
        ):
            if win_idx is None or final_idx is None: continue
            if win_idx >= len(row) or final_idx >= len(row): continue
            res = status(row[win_idx], row[final_idx])
            if not res: continue
            # Suppress 2026 super cups (not yet played as of May 2026).
            # Major and minor cups happen mid-to-late season and HAVE been
            # played for 2025-26.
            if kind == "super" and year and year >= 2026:
                continue
            rows.append({
                "slug": slugify(cn),
                "cur_name": cn,
                "year": year,
                "country": country,
                "kind": kind,
                "result": res,
            })
    return rows


# ---------- ETL: European appearances ----------

# Rnd# 1=Final, 2=SF, 3=QF, 4=R16, 5=Group Stage, 6+ qualifying / earlier rounds.
# Rnd Bin carries the per-competition stage code (CLF, CLBSF, ELQ, etc.). We
# label off the bin suffix so the mapping survives competition-name changes.
ROUND_LABEL_BY_RND = {
    1: "Final",
    2: "Semi-final",
    3: "Quarter-final",
    4: "Round of 16",
    5: "Group stage",
    6: "Round of 32 / qualifying",
    7: "Qualifying",
    8: "Qualifying",
    9: "Qualifying",
    10: "Qualifying",
    11: "Qualifying",
}

def collect_european(wb, in_scope_curnames):
    """For each in-scope club, aggregate European-competition results from
    Eur RndbyRnd into one row per (club, year, competition) with the
    farthest round reached and a trophy_won flag.

    The user-facing result label is derived from the minimum Rnd# (the
    deepest round) and the Trophy Won flag, not from row count, since a
    Trophy Won = Y row only exists on the eventual winner's final-round
    entry."""
    ws = wb["Eur RndbyRnd"]
    hdr, _ = header_map(ws)
    idx_season = hdr.get("Season")
    idx_comp_name = hdr.get("Leag/Comp.")
    idx_team_seas = hdr.get("Seas")
    idx_cn = hdr.get("Cur. Name")
    idx_rnd = hdr.get("Rnd#")
    idx_comp_code = hdr.get("Comp")
    idx_bin = hdr.get("Rnd Bin")
    idx_trophy = hdr.get("Trophy Won")

    # Aggregate: { (slug, year, code) -> {meta} }
    agg = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or idx_cn is None or idx_cn >= len(row): continue
        cn = row[idx_cn]
        if not cn or str(cn).strip() not in in_scope_curnames: continue
        cn = str(cn).strip()
        year = to_int(row[idx_team_seas]) if idx_team_seas is not None else None
        comp_name = row[idx_comp_name] if idx_comp_name is not None and idx_comp_name < len(row) else None
        code = row[idx_comp_code] if idx_comp_code is not None and idx_comp_code < len(row) else None
        rnd = to_int(row[idx_rnd]) if idx_rnd is not None and idx_rnd < len(row) else None
        bin_label = row[idx_bin] if idx_bin is not None and idx_bin < len(row) else None
        trophy = row[idx_trophy] if idx_trophy is not None and idx_trophy < len(row) else None
        season = row[idx_season] if idx_season is not None and idx_season < len(row) else None

        if year is None or not code: continue
        key = (slugify(cn), year, str(code))
        rec = agg.get(key)
        if rec is None:
            rec = {
                "year": year,
                "season": str(season) if season else None,
                "competition": str(comp_name) if comp_name else None,
                "code": str(code),
                "deepest_rnd": rnd,
                "deepest_bin": str(bin_label) if bin_label else None,
                "trophy_won": trophy == "Y",
            }
            agg[key] = rec
        else:
            # Track the deepest round (smallest Rnd#).
            if rnd is not None and (rec["deepest_rnd"] is None or rnd < rec["deepest_rnd"]):
                rec["deepest_rnd"] = rnd
                rec["deepest_bin"] = str(bin_label) if bin_label else None
            if trophy == "Y":
                rec["trophy_won"] = True

    # Group by slug and sort DESCENDING by year then competition.
    by_club = defaultdict(list)
    for (slug, _y, _c), rec in agg.items():
        # Final label combines trophy_won + deepest_rnd into one human string.
        if rec["trophy_won"]:
            rec["result_label"] = "Winner"
        elif rec["deepest_rnd"] is not None:
            rec["result_label"] = ROUND_LABEL_BY_RND.get(rec["deepest_rnd"], f"Round {rec['deepest_rnd']}")
        else:
            rec["result_label"] = "Entered"
        by_club[slug].append(rec)
    for slug in by_club:
        by_club[slug].sort(key=lambda r: (-(r["year"] or 0), r["competition"] or ""))
    return dict(by_club)


# ---------- ETL: European tournament hubs ----------

# Editorial mapping from user-facing tournament hub slug to the underlying
# Eur RndbyRnd codes and year ranges. Some competitions are defunct (CWC,
# ICFC); some are still active. The Europa League / UEFA Cup hub is filtered
# to year >= 1972 so the Inter-Cities Fairs Cup era (1955-1971) routes to its
# own dedicated hub, even though the workbook uses the same EL code for the
# whole span. Club World Cup and Intercontinental Cup share a hub because
# they overlapped briefly and merged after 2004; the workbook keeps separate
# codes (FCWC, IC).
EUROPEAN_TOURNAMENT_HUBS = [
    {
        "slug": "champions-league",
        "label": "UEFA Champions League / European Cup",
        "short_label": "Champions League",
        "codes": ["CL"],
        "year_min": None,
        "year_max": None,
        "active": True,
        "era_notes": "Founded 1955 as the European Cup; rebranded UEFA Champions League in 1992-93.",
    },
    {
        "slug": "europa-league",
        "label": "UEFA Europa League / UEFA Cup",
        "short_label": "Europa League",
        "codes": ["EL"],
        "year_min": 1972,
        "year_max": None,
        "active": True,
        "era_notes": "UEFA Cup 1972-2009; rebranded UEFA Europa League in 2009-10.",
    },
    {
        "slug": "conference-league",
        "label": "UEFA Conference League",
        "short_label": "Conference League",
        "codes": ["EUCL"],
        "year_min": None,
        "year_max": None,
        "active": True,
        "era_notes": "UEFA's tertiary European club competition since 2021-22.",
    },
    {
        "slug": "cup-winners-cup",
        "label": "UEFA Cup Winners' Cup",
        "short_label": "Cup Winners' Cup",
        "codes": ["CWC"],
        "year_min": None,
        "year_max": None,
        "active": False,
        "era_notes": "1960-61 to 1998-99. Discontinued; merged into the UEFA Cup after the 1998-99 final.",
    },
    {
        "slug": "inter-cities-fairs-cup",
        "label": "Inter-Cities Fairs Cup",
        "short_label": "Fairs Cup",
        "codes": ["EL"],
        "year_min": None,
        "year_max": 1971,
        "active": False,
        "era_notes": "1955-1971. Predecessor to the UEFA Cup. Never officially recognized by UEFA as one of its competitions.",
    },
    {
        "slug": "uefa-super-cup",
        "label": "UEFA Super Cup",
        "short_label": "Super Cup",
        "codes": ["USC"],
        "year_min": None,
        "year_max": None,
        "active": True,
        "era_notes": "Annual one-match final between the UEFA Champions League and UEFA Europa League winners.",
    },
    {
        "slug": "club-world-cup",
        "label": "FIFA Club World Cup / Intercontinental Cup",
        "short_label": "Club World Cup",
        "codes": ["FCWC", "IC"],
        "year_min": None,
        "year_max": None,
        "active": True,
        "era_notes": "Intercontinental Cup 1960-2004; FIFA Club World Cup since 2000. The two competitions overlapped, then merged after 2004 into a unified FIFA tournament.",
    },
]

# What deepest-round bucket a Rnd# belongs to, for the live-bracket widget.
# Anything Rnd# > 5 (qualifying rounds before group stage) is folded into the
# "qualifying" bucket so the visual stays compact.
BRACKET_ROUND_BUCKETS = [
    {"key": "final",        "label": "Final",         "rnd_match": lambda r: r == 1},
    {"key": "semifinal",    "label": "Semi-finals",   "rnd_match": lambda r: r == 2},
    {"key": "quarterfinal", "label": "Quarter-finals","rnd_match": lambda r: r == 3},
    {"key": "round_of_16",  "label": "Round of 16",   "rnd_match": lambda r: r == 4},
    {"key": "group_stage",  "label": "Group stage",   "rnd_match": lambda r: r == 5},
    {"key": "qualifying",   "label": "Qualifying",    "rnd_match": lambda r: r is not None and r >= 6},
]

# The current European season the workbook's 2026 rows refer to.
CURRENT_EURO_SEASON = None  # 2025-26 complete; set next season string when it begins
CURRENT_EURO_YEAR = None   # Set next season year when it begins


def collect_european_tournaments(wb, slug_for_curname):
    """Walk Eur RndbyRnd and aggregate by tournament hub instead of by club.
    Unlike collect_european(), this does NOT filter by in_scope_curnames:
    European tournament hubs need every club that ever participated, not
    just the in-scope Big 5 + extensions. Clubs without canonical pages
    are surfaced by name only; clubs with pages get a `slug` field that
    links to /teams/football/[slug].

    Returns a dict { hub_slug: {payload} } where payload includes year
    range, editions count, all-time champions, all-time finalists, most
    decorated, and per-club current-season records for the live bracket.
    """
    ws = wb["Eur RndbyRnd"]
    hdr, _ = header_map(ws)
    idx_season = hdr.get("Season")
    idx_comp_name = hdr.get("Leag/Comp.")
    idx_team_seas = hdr.get("Seas")
    idx_cn = hdr.get("Cur. Name")
    idx_rnd = hdr.get("Rnd#")
    idx_comp_code = hdr.get("Comp")
    idx_bin = hdr.get("Rnd Bin")
    idx_trophy = hdr.get("Trophy Won")

    # First pass: collect all rows with normalized fields.
    raw_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or idx_cn is None or idx_cn >= len(row): continue
        cn = row[idx_cn]
        if not cn: continue
        cn = str(cn).strip()
        year = to_int(row[idx_team_seas]) if idx_team_seas is not None else None
        code = row[idx_comp_code] if idx_comp_code is not None and idx_comp_code < len(row) else None
        rnd = to_int(row[idx_rnd]) if idx_rnd is not None and idx_rnd < len(row) else None
        trophy = row[idx_trophy] if idx_trophy is not None and idx_trophy < len(row) else None
        comp_name = row[idx_comp_name] if idx_comp_name is not None and idx_comp_name < len(row) else None
        season = row[idx_season] if idx_season is not None and idx_season < len(row) else None
        bin_label = row[idx_bin] if idx_bin is not None and idx_bin < len(row) else None
        if year is None or not code: continue
        raw_rows.append({
            "cur_name": cn,
            "year": year,
            "code": str(code),
            "rnd": rnd,
            "trophy": trophy == "Y",
            "competition": str(comp_name) if comp_name else None,
            "season": str(season) if season else None,
            "bin": str(bin_label) if bin_label else None,
        })

    hubs = {}
    for hub in EUROPEAN_TOURNAMENT_HUBS:
        codes = set(hub["codes"])
        year_min = hub.get("year_min")
        year_max = hub.get("year_max")
        matching = [
            r for r in raw_rows
            if r["code"] in codes
            and (year_min is None or r["year"] >= year_min)
            and (year_max is None or r["year"] <= year_max)
        ]
        if not matching:
            continue

        # Per (club, year) deepest-round aggregation.
        per_club_year = {}  # (cur_name, year) -> {deepest_rnd, trophy, season, competition}
        for r in matching:
            key = (r["cur_name"], r["year"])
            cur = per_club_year.get(key)
            if cur is None:
                per_club_year[key] = {
                    "cur_name": r["cur_name"],
                    "year": r["year"],
                    "deepest_rnd": r["rnd"],
                    "trophy": r["trophy"],
                    "season": r["season"],
                    "competition": r["competition"],
                }
            else:
                if r["rnd"] is not None and (cur["deepest_rnd"] is None or r["rnd"] < cur["deepest_rnd"]):
                    cur["deepest_rnd"] = r["rnd"]
                if r["trophy"]:
                    cur["trophy"] = True

        # All-time champions and finalists.
        champions_list = []   # list of {year, season, cur_name, slug}
        finalists_list = []   # list of {year, season, cur_name, slug} (runner-up rows)
        for (cn, year), v in per_club_year.items():
            if v["trophy"]:
                champions_list.append({
                    "year": year,
                    "season": v["season"],
                    "cur_name": cn,
                    "slug": slug_for_curname.get(cn),
                    "competition": v["competition"],
                })
            elif v["deepest_rnd"] == 1:
                finalists_list.append({
                    "year": year,
                    "season": v["season"],
                    "cur_name": cn,
                    "slug": slug_for_curname.get(cn),
                    "competition": v["competition"],
                })
        champions_list.sort(key=lambda x: -(x["year"] or 0))
        finalists_list.sort(key=lambda x: -(x["year"] or 0))

        # Most decorated. The base set is every club that has ever appeared
        # in a final of this competition (champion OR runner-up). Two
        # consequences worth noting:
        #  - Clubs that won at least one title surface with positive
        #    champion_count and (usually) some finals_lost.
        #  - Clubs that reached the final but never won surface with
        #    champion_count = 0 and finals_lost > 0. They sit at the bottom
        #    under the default cups-desc sort but become discoverable via the
        #    Finals column sort on the client.
        # Default order: champion_count desc, finals_count desc tiebreaker,
        # then alphabetical. The client component can re-sort by any column.
        from collections import Counter as _Counter
        win_ctr = _Counter(c["cur_name"] for c in champions_list)
        loss_ctr = _Counter(f["cur_name"] for f in finalists_list)
        all_finalist_names = set(win_ctr.keys()) | set(loss_ctr.keys())
        most_decorated = []
        for cn in all_finalist_names:
            wins_count = win_ctr.get(cn, 0)
            wins = [c for c in champions_list if c["cur_name"] == cn]
            losses = [f for f in finalists_list if f["cur_name"] == cn]
            last_win_year = max(w["year"] for w in wins) if wins else None
            last_final_year = max(
                (e["year"] for e in (wins + losses) if e.get("year") is not None),
                default=None,
            )
            finals_lost = loss_ctr.get(cn, 0)
            most_decorated.append({
                "cur_name": cn,
                "slug": slug_for_curname.get(cn),
                "champion_count": wins_count,
                "finals_lost": finals_lost,
                "finals_count": wins_count + finals_lost,
                "last_won": last_win_year,
                "last_final": last_final_year,
            })
        most_decorated.sort(key=lambda d: (
            -d["champion_count"],
            -d["finals_count"],
            d["cur_name"].lower(),
        ))

        # Year range and edition count (distinct years with any matching row).
        years_present = sorted({r["year"] for r in matching if r["year"]})

        # Per-club current-season state (year = CURRENT_EURO_YEAR). Each club
        # surfaces with its deepest round + winner flag. The page uses this
        # to render the NBA/NHL-style "alive vs eliminated at each round"
        # bracket. No live data emitted for hubs without a 2025-26 row.
        current_entries = []
        for (cn, year), v in per_club_year.items():
            if year != CURRENT_EURO_YEAR:
                continue
            current_entries.append({
                "cur_name": cn,
                "slug": slug_for_curname.get(cn),
                "deepest_rnd": v["deepest_rnd"],
                "trophy": v["trophy"],
            })

        hubs[hub["slug"]] = {
            "slug": hub["slug"],
            "label": hub["label"],
            "short_label": hub["short_label"],
            "active": hub["active"],
            "era_notes": hub["era_notes"],
            "year_min": years_present[0] if years_present else None,
            "year_max": years_present[-1] if years_present else None,
            "editions": len(years_present),
            "champions": champions_list,
            "finalists": finalists_list,
            "most_decorated": most_decorated,
            "current_season": CURRENT_EURO_SEASON if current_entries else None,
            "current_year": CURRENT_EURO_YEAR if current_entries else None,
            "current_entries": current_entries,
        }

    return hubs


# ---------- ETL: Totals roll-up (compact) ----------

def collect_totals(wb, in_scope_curnames):
    """Pull the top-line totals from Totals for header summaries.
    Compact: titles, finals, top-4 finishes, major cups, minor cups, european
    titles, european finals, european app, top-flight years."""
    ws = wb["Totals"]
    hdr, row1 = header_map(ws)
    # Cur. Name in col A (0); other indices via header lookup.
    idx_cn = hdr.get("Cur. Name", 0)
    idx_country = hdr.get("Country (Leag)", 1)

    # Best-effort header lookups; tolerate slight name drift.
    def col(*names):
        for n in names:
            if n in hdr: return hdr[n]
        return None

    out = {}
    keys = {
        "titles":         col("# Title (1 Div)"),
        "last_title":     col("Last Tit. (1 Div)"),
        "league_finals":  col("# Top 2"),
        "league_t4":      col("# Top 4"),
        "major_cups":     col("# Maj Trophies"),
        "minor_cups":     col("# Trophies (Maj. & Min)"),
        "last_trophy":    col("Last Trophy"),
        "career_years":   col("# Years", "#Years"),
    }
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or idx_cn >= len(row): continue
        cn = row[idx_cn]
        if not cn or str(cn).strip() not in in_scope_curnames: continue
        cn = str(cn).strip()
        rec = {}
        for k, i in keys.items():
            if i is None or i >= len(row): continue
            v = row[i]
            if k.startswith("last_"):
                rec[k] = to_int(v)
            else:
                iv = to_int(v)
                rec[k] = iv if iv is not None else 0
        out[cn] = rec
    return out


# ---------- ETL: league hub data ----------

def build_league_hubs(wb, standings_rows):
    """Build the five Big 5 league hub aggregates.
    Each hub gets: current season standings (latest year for that league name),
    all-time champions list (every year, every Level-1 league name in that country)."""
    hubs = {}
    # Index standings rows by (country, league_name, year).
    by_country_level1 = defaultdict(list)
    for r in standings_rows:
        if r["level"] != 1: continue
        by_country_level1[r["country"]].append(r)

    for slug, country, league_name, display, level in LEAGUE_HUBS:
        rows = by_country_level1.get(country, [])
        # Current standings: most recent year for this exact modern league name
        # with populated data. Skip forward-keyed placeholder rows where the
        # workbook has carried the team list but no results yet (typical of
        # the upcoming season before round 1 is played).
        modern = [r for r in rows if r["league"] == league_name]
        completed_years = sorted({r["year"] for r in modern
                                  if r["year"] and r["pts"] is not None}, reverse=True)
        latest_year = completed_years[0] if completed_years else None
        current = [r for r in modern if r["year"] == latest_year]
        current.sort(key=lambda r: (r["place"] if r["place"] is not None else 99))

        # All-time champions: workbook BX (Champions) flag is the source
        # of truth. It fires for league-format Level 1 winners AND for
        # pre-modern playoff/knockout champions where place is null
        # (Schalke pre-Bundesliga, the Italian Football Championship era,
        # the French amateur era). It does NOT fire for second-division
        # winners, so Schalke's 2.Bundesliga rows correctly drop out of
        # the all-time list while their 7 pre-Bundesliga German titles
        # appear regardless of their current league level.
        champs = [r for r in rows if r.get("champion")]
        champs.sort(key=lambda r: r["year"] or 0)

        hubs[slug] = {
            "slug": slug,
            "country": country,
            "league": display,
            "current_year": latest_year,
            "current_standings": current,
            "all_time_champions": [
                {
                    "year": r["year"],
                    "champion": r["cur_name"],
                    "champion_team": r["team"],
                    "champion_slug": r["slug"],
                    "league_name": r["league"],
                    "format": r["format"],
                }
                for r in champs
            ],
        }
    return hubs


# ---------- Main ----------

def main():
    src = find_source()
    print(f"Source: {src}")
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    print("Collecting in-scope standings rows...")
    standings_rows, in_scope_curnames = collect_standings_rows(wb)
    print(f"  {len(standings_rows)} rows, {len(in_scope_curnames)} distinct canonical clubs")

    # Country mode per club, derived from the in-scope standings rows.
    # Wartime annexations (SK Rapid Wien, First Vienna FC, FC Admira Wacker
    # Mödling under the German playoff during 1938-1945; DFC Prag pre-WWI
    # German Empire) and cross-border traditions (AS Monaco in Ligue 1,
    # Cardiff and Swansea in English football, AC Libertas in early Italian
    # football, Moghreb Tétouán during the Spanish protectorate) all put
    # the club in a Big 5 country that differs from their Lookup federation.
    # Use the standings-row country as the canonical for everything site-
    # facing so the index, league hubs, and per-club page header all reflect
    # the workbook's own country-of-play classification.
    country_mode_by_cn = {}
    for cn in in_scope_curnames:
        counts = Counter(r["country"] for r in standings_rows if r["cur_name"] == cn)
        if counts:
            country_mode_by_cn[cn] = counts.most_common(1)[0][0]

    print("Building club index from Lookup...")
    clubs = build_clubs_index(wb, in_scope_curnames, country_mode_by_cn)
    missing = in_scope_curnames - set(clubs.keys())
    if missing:
        print(f"  WARN: {len(missing)} in-scope clubs have NO Lookup entry; synthesizing minimal records")
        for cn in missing:
            # Infer country from the first standings row for this club.
            country = next((r["country"] for r in standings_rows if r["cur_name"] == cn), None)
            clubs[cn] = {
                "slug": slugify(cn),
                "cur_name": cn,
                "country": country,
                "federation_country": None,
                "city": None, "metro": None, "county": None,
                "continent": "Europe",
                "lat": None, "lng": None,
            }
    print(f"  index has {len(clubs)} clubs")

    # Slug collision check + merge. Two Cur. Names that slugify identically
    # are treated as the same canonical club (e.g. 'SPAL' vs 'Spal'). The
    # winning Cur. Name is the one with more standings rows. The losing
    # Cur. Name's standings/cup/europe rows are rewritten to point at the
    # winner's slug downstream.
    slug_groups = defaultdict(list)
    for cn, club in clubs.items():
        slug_groups[club["slug"]].append(cn)
    canonical_for_slug = {}
    aliased_cn = {}
    for slug, cns in slug_groups.items():
        if len(cns) == 1:
            canonical_for_slug[slug] = cns[0]; continue
        # Pick the spelling with the most standings rows; tie-break alpha.
        row_count = Counter(r["cur_name"] for r in standings_rows if r["slug"] == slug)
        cns_sorted = sorted(cns, key=lambda x: (-row_count.get(x, 0), x))
        winner = cns_sorted[0]
        canonical_for_slug[slug] = winner
        for loser in cns_sorted[1:]:
            aliased_cn[loser] = winner
            print(f"  merging slug-colliding Cur. Names: {loser!r} -> {winner!r} (slug={slug!r})")
    # Drop loser clubs from the index; rewrite standings/cups to winner.
    for loser, winner in aliased_cn.items():
        clubs.pop(loser, None)
    for r in standings_rows:
        if r["cur_name"] in aliased_cn:
            r["cur_name"] = aliased_cn[r["cur_name"]]
            r["slug"] = slugify(r["cur_name"])

    print("Collecting cup finals...")
    cups_rows = collect_cup_finals(wb, in_scope_curnames)
    # Rewrite slug-merged Cur. Names. (aliased_cn isn't defined yet here on
    # first call; defer the rewrite to after the slug-merge block runs below.)
    print(f"  {len(cups_rows)} cup-final rows")

    print("Collecting European appearances...")
    europe = collect_european(wb, in_scope_curnames)
    print(f"  {sum(len(v) for v in europe.values())} euro entries across {len(europe)} clubs")

    print("Collecting Totals roll-up...")
    totals = collect_totals(wb, in_scope_curnames)
    print(f"  {len(totals)} totals rows matched")

    # Merge totals + tier coverage into clubs index.
    for cn, club in clubs.items():
        club_rows = [r for r in standings_rows if r["cur_name"] == cn]
        tiers = sorted({r["level"] for r in club_rows if r["level"] is not None})
        years = [r["year"] for r in club_rows if r["year"]]
        # Exclude 2027 placeholder rows from the playoff count so non-German
        # clubs (Chelsea, etc.) that have a 2026-27 placeholder row don't
        # mistakenly trigger the German-era playoff label downstream.
        playoff_only_years = {r["year"] for r in club_rows
                              if r["format"] == "playoff" and r["year"] and r["year"] < 2027}
        # Top-flight (Level 1) league seasons -- the only thing that should
        # carry the "top-flight" label per editorial spec.
        level1_league_years = {r["year"] for r in club_rows if r["format"] == "league" and r["level"] == 1}
        lower_league_years = {r["year"] for r in club_rows if r["format"] == "league" and (r["level"] or 0) > 1}
        # Per-year level map for the index page filter UX. Keys are year
        # numbers as strings (JSON limitation); values are the lowest level
        # number (highest tier) the club played that year. Empty when the
        # club has no row for that year. Skips 2027 placeholders to match
        # the lib/football.ts MAX_DISPLAYED_YEAR clamp.
        tier_by_year: dict[str, int] = {}
        # Parallel per-year country map. Mulhouse 1941 must group under
        # Germany (Anschluss-era Alsace was annexed) even though the club's
        # overall mode country is France. When multiple rows exist for the
        # same year at different levels, prefer the higher-tier (lower
        # level number) row's country, matching tier_by_year semantics.
        country_by_year: dict[str, str] = {}
        for r in club_rows:
            y, lv, ctry = r["year"], r["level"], r["country"]
            if y is None or lv is None or y >= 2027: continue
            key = str(y)
            cur = tier_by_year.get(key)
            if cur is None or lv < cur:
                tier_by_year[key] = lv
                if ctry: country_by_year[key] = ctry
        club["tiers"] = tiers
        club["first_year"] = min(years) if years else None
        club["last_year"] = max(years) if years else None
        club["top_flight_seasons"] = len(level1_league_years)
        club["lower_tier_seasons"] = len(lower_league_years)
        # Kept for backwards compat with the index page filter logic.
        club["league_seasons"] = len(level1_league_years) + len(lower_league_years)
        club["playoff_appearances"] = len(playoff_only_years)
        club["totals"] = totals.get(cn, {})
        club["tier_by_year"] = tier_by_year
        club["country_by_year"] = country_by_year

    # Derive promoted / relegated per row from consecutive-season level
    # transitions. The forward-scan finds the NEXT EXISTING year for the
    # same club (not just y+1), which closes WWI and WWII gaps (Arsenal
    # 1915 L2 -> 1920 L1 = promoted) and correctly handles English clubs
    # that took a season off. When the forward-scan produces a verdict,
    # it OVERRIDES the workbook seed (the workbook's 1915 'Reg' for
    # Arsenal is wrong because Arsenal was actually promoted via the
    # post-war First Division expansion). When the forward-scan finds
    # no next year (latest completed season) OR the next year is at the
    # same level (e.g. Villarreal 2012 -> 2014 both L1 because their
    # 2013 Segunda season is outside our scope), the workbook seed wins.
    seasons_by_slug = defaultdict(list)
    for r in standings_rows:
        seasons_by_slug[r["slug"]].append(r)
    for slug, rows in seasons_by_slug.items():
        rows.sort(key=lambda r: (r["year"] or 0, r["level"] or 99))
        # Index by year, picking the lowest-tier (highest level number)
        # row when multiple rows share a year.
        by_year = {}
        years_sorted = []
        for r in rows:
            y, lv = r["year"], r["level"]
            if y is None or lv is None: continue
            cur = by_year.get(y)
            if cur is None or (cur["level"] or 0) < lv:
                by_year[y] = r
        years_sorted = sorted(by_year.keys())
        for r in rows:
            y, lv = r["year"], r["level"]
            if y is None or lv is None: continue
            # Find next existing year strictly greater than y.
            nxt = None
            for ny in years_sorted:
                if ny > y:
                    nxt = by_year[ny]; break
            if not nxt or nxt["level"] is None:
                continue  # latest season; workbook seed stands
            if nxt["level"] < lv:
                r["promoted"] = True
                r["relegated"] = False  # override stale workbook seed
            elif nxt["level"] > lv:
                r["relegated"] = True
                r["promoted"] = False
            # else: same level -- keep workbook seed (Villarreal 2012 case)
        # Final sort: descending by year (newest first), then by level so
        # multi-tier same-year rows show the higher tier first.
        rows.sort(key=lambda r: (-(r["year"] or 0), r["level"] or 99))

    cups_by_slug = defaultdict(list)
    for r in cups_rows:
        cups_by_slug[r["slug"]].append(r)
    for slug in cups_by_slug:
        cups_by_slug[slug].sort(key=lambda r: (r["year"] or 0, r["kind"]))

    print("Building league hubs...")
    league_hubs = build_league_hubs(wb, standings_rows)

    print("Building European tournament hubs...")
    # slug_for_curname covers in-scope clubs only; clubs outside scope
    # (Anderlecht, Galatasaray, etc.) surface in tournament rows with no slug
    # so the page renders them as plain text without a broken link.
    slug_for_curname = {cn: c["slug"] for cn, c in clubs.items()}
    european_hubs = collect_european_tournaments(wb, slug_for_curname)
    print(f"  {len(european_hubs)} hubs: {sorted(european_hubs.keys())}")
    for s, h in european_hubs.items():
        ch = len(h["champions"])
        decorated = h["most_decorated"][0]["cur_name"] if h["most_decorated"] else "—"
        decorated_n = h["most_decorated"][0]["champion_count"] if h["most_decorated"] else 0
        print(f"    {s}: {ch} champions, top {decorated} ({decorated_n})")

    # ---------- Write outputs ----------
    index_path = OUT_DIR / "index.json"
    seasons_path = OUT_DIR / "seasons.json"
    cups_path = OUT_DIR / "cups.json"
    europe_path = OUT_DIR / "europe.json"
    leagues_path = OUT_DIR / "leagues.json"
    european_tournaments_path = OUT_DIR / "european-tournaments.json"

    payload_index = {
        "generated_at": __import__("datetime").date.today().isoformat(),
        "source": str(src.name),
        "scope": {"countries": sorted(IN_SCOPE_COUNTRIES), "country_tiers": {k: sorted(v) for k, v in COUNTRY_TIERS.items()}},
        "clubs": [c for c in sorted(clubs.values(), key=lambda x: (x["country"] or "", x["cur_name"]))],
    }
    with open(index_path, "w", encoding="utf-8") as f:
        json.dump(payload_index, f, ensure_ascii=False, indent=2)
    print(f"Wrote {index_path}  ({index_path.stat().st_size:,} bytes)")

    with open(seasons_path, "w", encoding="utf-8") as f:
        json.dump(dict(seasons_by_slug), f, ensure_ascii=False)
    print(f"Wrote {seasons_path}  ({seasons_path.stat().st_size:,} bytes)")

    with open(cups_path, "w", encoding="utf-8") as f:
        json.dump(dict(cups_by_slug), f, ensure_ascii=False)
    print(f"Wrote {cups_path}  ({cups_path.stat().st_size:,} bytes)")

    with open(europe_path, "w", encoding="utf-8") as f:
        json.dump(europe, f, ensure_ascii=False)
    print(f"Wrote {europe_path}  ({europe_path.stat().st_size:,} bytes)")

    with open(leagues_path, "w", encoding="utf-8") as f:
        json.dump(league_hubs, f, ensure_ascii=False)
    print(f"Wrote {leagues_path}  ({leagues_path.stat().st_size:,} bytes)")

    with open(european_tournaments_path, "w", encoding="utf-8") as f:
        json.dump(european_hubs, f, ensure_ascii=False)
    print(f"Wrote {european_tournaments_path}  ({european_tournaments_path.stat().st_size:,} bytes)")

    # Normalized-name -> slug lookup for cross-data-source joins. Mirrors
    # build-sports-index.py's normalize_team_name (lowercase, alnum-only,
    # collapsed whitespace) so the join works against the team name as it
    # appears in MetroAreas.xlsx FootballClub_Data and Team List.
    def _norm(s: str) -> str:
        s = s.lower()
        s = re.sub(r"[^a-z0-9]+", " ", s)
        return s.strip()
    slug_lookup = {}
    for club in clubs.values():
        key = _norm(club["cur_name"])
        if key and key not in slug_lookup:
            slug_lookup[key] = club["slug"]
    slug_lookup_path = OUT_DIR / "slug-lookup.json"
    with open(slug_lookup_path, "w", encoding="utf-8") as f:
        json.dump(slug_lookup, f, ensure_ascii=False)
    print(f"Wrote {slug_lookup_path}  ({slug_lookup_path.stat().st_size:,} bytes, {len(slug_lookup)} entries)")

    print("\nDone.")


if __name__ == "__main__":
    main()

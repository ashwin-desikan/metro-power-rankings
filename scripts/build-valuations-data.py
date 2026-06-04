#!/usr/bin/env python3
"""Build cross-sport team valuations data from the Team Valuations sheet of
OtherLeagues.xlsx.

The sheet is a curated, non-exhaustive snapshot: latest available valuation per
team. US leagues (NFL/NBA/MLB/NHL) carry Forbes figures; global soccer clubs
carry Sportico figures, with the League column holding the club's country.

Emits public/data/valuations/valuations.json:
  { "generated": "<iso>", "rows": [ {year, team, league, value_m, source}, ... ] }
Link resolution to canonical /teams pages is done in lib/valuations.ts so the
shared resolveTeamLink() stays the single source of truth.

Usage: python scripts/build-valuations-data.py [SOURCE_XLSX]
"""
import json, os, sys, datetime
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = sys.argv[1] if len(sys.argv) > 1 else os.path.join(ROOT, "OtherLeagues.xlsx")
OUT = os.path.join(ROOT, "public", "data", "valuations", "valuations.json")
SHEET = "Team Valuations"

def main():
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        sys.exit(f"FAIL: sheet '{SHEET}' not found in {SRC}")
    ws = wb[SHEET]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    idx = {name: i for i, name in enumerate(header)}
    for req in ("Year", "Team", "League", "Value ($M)", "Source"):
        if req not in idx:
            sys.exit(f"FAIL: missing column '{req}' (have {header})")
    out = []
    for r in rows[1:]:
        team = r[idx["Team"]]
        val = r[idx["Value ($M)"]]
        if team is None or val is None:
            continue
        out.append({
            "year": int(r[idx["Year"]]) if r[idx["Year"]] is not None else None,
            "team": str(team).strip(),
            "league": str(r[idx["League"]]).strip() if r[idx["League"]] is not None else "",
            "value_m": float(val),
            "source": str(r[idx["Source"]]).strip() if r[idx["Source"]] is not None else "",
        })
    out.sort(key=lambda x: x["value_m"], reverse=True)
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    payload = {"generated": datetime.datetime.now().isoformat(timespec="seconds"), "rows": out}
    tmp = OUT + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=0)
    os.replace(tmp, OUT)
    print(f"wrote {len(out)} rows -> {os.path.relpath(OUT, ROOT)}")
    # quick league histogram
    from collections import Counter
    print("by league:", dict(Counter(x["league"] for x in out)))

if __name__ == "__main__":
    main()

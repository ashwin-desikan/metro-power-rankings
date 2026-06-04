#!/usr/bin/env python3
"""
Build NFL team-pages data from NFL_all.xlsx.

Reads the canonical NFL workbook (NFL_all_backup.xlsx in the user's
OneDrive Excel Files folder, or a path passed on the command line) and
emits four JSON files under public/data/nfl/.

Outputs:
  public/data/nfl/franchises.json       - 32 active franchises (index + per-page header)
  public/data/nfl/championships.json    - one entry per championship year per franchise
  public/data/nfl/stadium-history.json  - stadium-era rows grouped by canonical franchise
  public/data/nfl/award-winners.json    - MVP, COY, OPOY, etc. grouped by franchise
  public/data/nfl/historical.json       - defunct franchises for /teams/nfl/historical
  public/data/nfl/hall-of-fame.json     - HoF inductees by primary team (v1: primary only)
  public/data/nfl/pro-bowl-counts.json  - per-franchise Pro Bowl selection counts
  public/data/nfl/seasons-by-team.json  - per-franchise season-by-season rows

Canonical join key throughout: Year by Year col DN ("Name"). Every other
sheet's "Name"-equivalent column joins back to this. See the workbook's
"Claude Notes" sheet for full schema.

Usage:
  python scripts/build-nfl-data.py
  python scripts/build-nfl-data.py /path/to/NFL_all_backup.xlsx
"""

import json
import os as _ometro, sys as _smetro
_smetro.path.insert(0, _ometro.path.dirname(_ometro.path.abspath(__file__)))
from _defunct_metro import resolve_city as _rmcity, resolve_nba as _rmnba, resolve_nfl as _rmnfl
import os
import re
import sys
from pathlib import Path
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl",
                           "--quiet", "--break-system-packages"])
    import openpyxl


# -------- Constants --------

REPO_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_SOURCE_CANDIDATES = [
    # User's OneDrive (Windows) — preferred when present
    Path(os.path.expanduser("~/OneDrive/Excel Files/NFL_all_backup.xlsx")),
    Path(os.path.expanduser("~/OneDrive/Excel Files/NFL_all - Copy.xlsx")),
    Path(os.path.expanduser("~/OneDrive/Excel Files/NFL_all.xlsx")),
    # OneDrive (Windows) — preferred filenames
    Path(os.path.expanduser("~/OneDrive/Excel Files/NFL_all_copy.xlsx")),
    # OneDrive via Linux bindfs mount
    Path("/sessions/jolly-tender-turing/mnt/Excel Files/NFL_all_backup.xlsx"),
    Path("/sessions/jolly-tender-turing/mnt/Excel Files/NFL_all_copy.xlsx"),
    Path("/sessions/jolly-tender-turing/mnt/Excel Files/NFL_all - Copy.xlsx"),
    Path("/sessions/jolly-tender-turing/mnt/Excel Files/NFL_all.xlsx"),
    # Project-local fallback (if the user committed a frozen copy)
    REPO_ROOT / "data" / "nfl-source" / "NFL_all_backup.xlsx",
    REPO_ROOT / "data" / "nfl-source" / "NFL_all.xlsx",
    # Bootstrap fallback to the most recent upload
    Path("/sessions/jolly-tender-turing/mnt/uploads/NFL_all - Copy-5a2eae1a.xlsx"),
]

OUT_DIR = REPO_ROOT / "public" / "data" / "nfl"


def read_team_external_links():
    """Load team-wikidata.tsv (at project root, shared across all leagues)
    and return a dict keyed by display-name (e.g. "Buffalo Bills") ->
    {"wikipedia_url", "wikidata_qid"}.

    Match key for NFL is the display name = "{city} {team}", which matches
    the TSV's column 0 directly (Buffalo Bills, Las Vegas Raiders, etc.).
    Falls back to empty when the file is missing.
    """
    tsv_path = REPO_ROOT / "team-wikidata.tsv"
    out = {}
    if not tsv_path.exists():
        print(f"  (note: {tsv_path.name} not found at project root; team Wikipedia/Wikidata links will be blank)")
        return out
    with tsv_path.open(encoding="utf-8") as fh:
        for line in fh:
            parts = line.rstrip("\r\n").split("\t")
            if len(parts) < 3:
                continue
            team = parts[0].strip()
            qid = parts[1].strip()
            wiki = parts[2].strip()
            if not team:
                continue
            out[team] = {
                "wikipedia_url": wiki or None,
                "wikidata_qid": qid or None,
            }
    return out


# Stolen-championship editorial flag for the 1925 Pottsville Maroons season.
# The Pottsville Maroons / Boston Bulldogs lineage is one franchise in the
# workbook with canonical name "Bulldogs (Boston)" (Pottsville 1925-28, Boston
# 1929). The separate canonical "Maroons" is the Toledo/Kenosha franchise
# (1922-24) that folded before Pottsville existed. Attach the stolen title
# to the Pottsville lineage, not the Toledo/Kenosha one.
STOLEN_TITLES = {
    ("Bulldogs (Boston)", 1925):
        "Pottsville finished 1925 at 10-2 and beat the Chicago Cardinals 21-7 in the de facto "
        "championship game. Six days later they played a Notre Dame All-Stars exhibition at Shibe "
        "Park inside the Frankford Yellow Jackets' protected territory and were suspended. The "
        "Cardinals padded their record against two dissolved teams and were awarded the title. "
        "Pete Rozelle reviewed in 1963 and 1972. NFL owners voted 30-2 in 2003 to leave the title "
        "with the Cardinals. Pottsville faithful have never accepted it.",
}

# Editorial overrides for city_history. The workbook's Totals row for the
# Toledo/Kenosha Maroons franchise erroneously includes "Pottsville/Boston"
# in its city list; those cities belong to the separate Bulldogs (Boston)
# franchise. Apply override at ETL time rather than asking the user to
# touch the source workbook.
CITY_HISTORY_OVERRIDES = {
    "Maroons": "Toledo/Kenosha",
}

# Editorial round-label overrides for the Regular Season sheet. Keyed by
# (date_iso, frozenset(winner_canonical, loser_canonical)) so the override
# fires regardless of which team-row is being processed. Use only when the
# workbook label is verifiably wrong and a workbook fix is deferred.
ROUND_OVERRIDES = {
    # 2023 AFC Championship Game, Kansas City 17 at Baltimore 10 (M&T Bank).
    # The workbook currently labels this game NFC Champ; both franchises are
    # AFC. Override to AFC Champ at ETL time pending a workbook correction.
    ("2024-01-28", frozenset({"Chiefs", "Ravens"})): "AFC Champ",
}


def resolve_round(date_iso, team_a_canonical, team_b_canonical, workbook_label):
    """Apply ROUND_OVERRIDES if a matching (date, team-pair) entry exists,
    otherwise return the workbook label unchanged."""
    if date_iso and team_a_canonical and team_b_canonical:
        key = (date_iso, frozenset({team_a_canonical, team_b_canonical}))
        if key in ROUND_OVERRIDES:
            return ROUND_OVERRIDES[key]
    return workbook_label


# -------- Helpers --------

def slugify(s):
    s = (s or "").lower().strip()
    s = re.sub(r"[àáâãäå]", "a", s)
    s = re.sub(r"[èéêë]", "e", s)
    s = re.sub(r"[ìíîï]", "i", s)
    s = re.sub(r"[òóôõö]", "o", s)
    s = re.sub(r"[ùúûü]", "u", s)
    s = re.sub(r"[ñ]", "n", s)
    s = re.sub(r"[^a-z0-9\s-]", "", s)
    s = re.sub(r"\s+", "-", s)
    s = re.sub(r"-+", "-", s)
    return s.strip("-")


def safe_int(val, default=0):
    if val is None or val == "":
        return default
    try:
        return int(val)
    except (ValueError, TypeError):
        try:
            return int(float(val))
        except (ValueError, TypeError):
            return default


def safe_float(val, default=0.0):
    if val is None or val == "":
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def safe_str(val):
    if val is None:
        return ""
    return str(val).strip()


def metro_slug_from_metro_area(metro_area):
    """Map the workbook's 'Metro Area' string to the rankings-side metro slug."""
    if not metro_area:
        return None
    return slugify(metro_area)


def find_source_xlsx(argv):
    if len(argv) >= 2:
        p = Path(argv[1])
        if p.exists():
            return p
        print(f"WARN: explicit path {p} not found, falling back to defaults", file=sys.stderr)
    for c in DEFAULT_SOURCE_CANDIDATES:
        if c.exists():
            return c
    raise SystemExit(f"No NFL workbook found. Checked: {DEFAULT_SOURCE_CANDIDATES}")


# -------- Sheet readers --------

def read_lookup_block3(wb):
    """
    Lookup sheet, cols N-U rows 2-37: current 32-team rich detail.
    Returns dict keyed on canonical team short name (col O):
      {"Cardinals": {city, team, division, conf, stadium, stadium_city, metro, state}}
    """
    ws = wb["Lookup"]
    out = {}
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=40, values_only=True), start=2):
        # Cols N-U = indices 13-20 (0-based). Pad to length 22 just in case.
        row = list(row) + [None] * max(0, 22 - len(row))
        city = safe_str(row[13])
        team = safe_str(row[14])
        division = safe_str(row[15])
        conf = safe_str(row[16])
        stadium = safe_str(row[17])
        stadium_city = safe_str(row[18])
        metro = safe_str(row[19])
        state = safe_str(row[20])
        if not team:
            continue
        out[team] = {
            "city": city,
            "team": team,
            "division": division,
            "conf": conf,
            "stadium": stadium,
            "stadium_city": stadium_city,
            "metro": metro,
            "state": state,
        }
    return out


def read_totals(wb):
    """
    Totals sheet: one row per franchise. AK col is canonical name (join key).
    Returns dict keyed on canonical name -> franchise record.
    """
    ws = wb["Totals"]
    out = {}
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row = list(row) + [None] * max(0, 37 - len(row))
        league = safe_str(row[0])
        city = safe_str(row[1])
        team_hist = safe_str(row[2])
        # D-G: W L T Win%
        w = safe_int(row[3])
        l = safe_int(row[4])
        t = safe_int(row[5])
        winpct = safe_float(row[6])
        # H: # seasons; I: .500 seasons
        seasons = safe_int(row[7])
        sea500 = safe_int(row[8])
        # J-K: P. App, Div. Titles
        playoff_apps = safe_int(row[9])
        div_titles = safe_int(row[10])
        # L-M: Best Main Div, Best Rec.
        best_div_count = safe_int(row[11])
        best_rec_count = safe_int(row[12])
        # N-O: Chmp App, Chmps
        champ_apps = safe_int(row[13])
        champs = safe_int(row[14])
        # P-S: Playoff W L T Win%
        p_w = safe_int(row[15])
        p_l = safe_int(row[16])
        p_t = safe_int(row[17])
        p_winpct = safe_float(row[18])
        # T-U: CF App, CF Wins
        cf_app = safe_int(row[19])
        cf_wins = safe_int(row[20])
        # V: Ab Win% (combined)
        abs_winpct = safe_float(row[21])
        # W: Last Champ year
        last_champ = safe_int(row[22], None) if row[22] is not None else None
        # AF-AG: Current Y/N, Defunct
        current = safe_str(row[31]) == "Y"
        defunct = safe_str(row[32]) == "Y"
        # AH-AJ: Reg Games, Play Games, Total Games
        reg_games = safe_int(row[33])
        play_games = safe_int(row[34])
        total_games = safe_int(row[35])
        # AK: canonical Name
        name = safe_str(row[36])
        if not name:
            continue
        out[name] = {
            "league": league,
            "city_history": city,
            "team_history": team_hist,
            "all_time_w": w,
            "all_time_l": l,
            "all_time_t": t,
            "win_pct": winpct,
            "seasons": seasons,
            "seasons_500_plus": sea500,
            "playoff_appearances": playoff_apps,
            "division_titles": div_titles,
            "champ_appearances": champ_apps,
            "championships": champs,
            "playoff_w": p_w,
            "playoff_l": p_l,
            "playoff_t": p_t,
            "playoff_win_pct": p_winpct,
            "conference_finals_app": cf_app,
            "conference_finals_wins": cf_wins,
            "abs_win_pct": abs_winpct,
            "last_championship": last_champ,
            "is_current": current,
            "is_defunct": defunct,
            "reg_games": reg_games,
            "play_games": play_games,
            "total_games": total_games,
            "canonical": name,
        }
    return out


def read_year_by_year(wb):
    """
    Year by Year sheet: one row per team-season.
    Used for:
      - per-franchise championship years (T = "Y")
      - season-by-season tables
      - founding year (min Year per canonical name)
    Returns dict keyed on canonical name -> list of season dicts.
    """
    ws = wb["Year by Year"]
    rows_by_name = defaultdict(list)
    # Header is row 1. Data from row 2.
    for row in ws.iter_rows(min_row=2, values_only=True):
        row = list(row) + [None] * max(0, 136 - len(row))
        year = safe_int(row[2], None) if row[2] is not None else None
        if year is None:
            continue
        league = safe_str(row[1])
        city = safe_str(row[3])
        team_hist = safe_str(row[4])
        w = safe_int(row[5])
        l = safe_int(row[6])
        t = safe_int(row[7])
        winpct = safe_float(row[8])
        pf = safe_int(row[9])
        pa = safe_int(row[10])
        playoff_y = safe_str(row[11])
        div_title_y = safe_str(row[12])
        # R (17) = CF App, S (18) = Cham App, T (19) = Champs (Y/blank)
        cf_app_y = safe_str(row[17])
        cham_app_y = safe_str(row[18])
        champ_y = safe_str(row[19])
        # U (20) division, V (21) place
        division = safe_str(row[20])
        place = safe_str(row[21])
        # Cols around 26 onward have stadium era; AJ-AL = home city/metro/state (35-37)
        home_stadium_era = safe_str(row[26]) if len(row) > 26 else ""
        home_city = safe_str(row[35]) if len(row) > 35 else ""
        metro_area = safe_str(row[36]) if len(row) > 36 else ""
        home_state = safe_str(row[37]) if len(row) > 37 else ""
        # DN (col 117) = canonical name
        canonical = safe_str(row[117]) if len(row) > 117 else ""
        if not canonical:
            continue
        rows_by_name[canonical].append({
            "year": year,
            "league": league,
            "city": city,
            "team_historical": team_hist,
            "w": w, "l": l, "t": t, "win_pct": winpct,
            "pf": pf, "pa": pa,
            "playoff_appearance": playoff_y == "Y",
            "division_title": div_title_y == "Y",
            "conference_final": cf_app_y == "Y",
            "championship_appearance": cham_app_y == "Y",
            "championship": champ_y == "Y",
            "division": division,
            "place": place,
            "stadium_era": home_stadium_era,
            "home_city": home_city,
            "metro_area": metro_area,
            "home_state": home_state,
        })
    # Sort each franchise's seasons ascending
    for k in rows_by_name:
        rows_by_name[k].sort(key=lambda r: r["year"])
    return rows_by_name


def read_stadiums(wb):
    """
    Stadiums sheet: A canonical name, B era name, C team (canonical), D city,
    E metro, F state, G first year, H last year.
    Returns dict keyed on canonical team -> list of stadium-era rows.
    Era rows for the same canonical stadium name get grouped under one
    'building' record on output.
    """
    ws = wb["Stadiums"]
    by_team = defaultdict(list)
    for row in ws.iter_rows(min_row=2, values_only=True):
        row = list(row) + [None] * max(0, 8 - len(row))
        canonical_stadium = safe_str(row[0])
        era_name = safe_str(row[1])
        team = safe_str(row[2])
        city = safe_str(row[3])
        metro = safe_str(row[4])
        state = safe_str(row[5])
        first_year = safe_int(row[6], None)
        last_year = safe_int(row[7], None)
        if not team:
            # neutral-site / host-only venue, skip per-team rollup
            continue
        if not canonical_stadium:
            continue
        by_team[team].append({
            "canonical_stadium": canonical_stadium,
            "era_name": era_name,
            "city": city,
            "metro": metro,
            "state": state,
            "first_year": first_year,
            "last_year": last_year,
        })
    # For each team, group eras by canonical stadium name (one building = one group)
    grouped = {}
    for team, rows in by_team.items():
        buildings = defaultdict(lambda: {
            "canonical": None, "city": None, "metro": None, "state": None,
            "first_year": None, "last_year": None, "eras": [],
        })
        for r in rows:
            b = buildings[r["canonical_stadium"]]
            b["canonical"] = r["canonical_stadium"]
            b["city"] = b["city"] or r["city"]
            b["metro"] = b["metro"] or r["metro"]
            b["state"] = b["state"] or r["state"]
            fy, ly = r["first_year"], r["last_year"]
            if fy is not None:
                b["first_year"] = fy if b["first_year"] is None else min(b["first_year"], fy)
            if ly is not None:
                b["last_year"] = ly if b["last_year"] is None else max(b["last_year"], ly)
            b["eras"].append({
                "era_name": r["era_name"], "first_year": fy, "last_year": ly,
            })
        out_buildings = []
        for cname, b in buildings.items():
            b["eras"].sort(key=lambda e: (e["first_year"] or 0))
            out_buildings.append(b)
        # Sort buildings by most-recent first
        out_buildings.sort(key=lambda b: -(b["last_year"] or 0))
        grouped[team] = out_buildings
    return grouped


def read_awards(wb):
    """
    Awards sheet: 1920-2025, 14 award types. Keyed on canonical name (col J).
    Returns dict[canonical_name][award_type] -> list of {year, player, position}.
    """
    ws = wb["Awards"]
    by_team = defaultdict(lambda: defaultdict(list))
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        row = list(row) + [None] * max(0, 11 - len(row))
        year = safe_int(row[3], None)
        player = safe_str(row[4])
        award = safe_str(row[7])
        position = safe_str(row[8])
        canonical = safe_str(row[9])
        if not (player and award and canonical and year):
            continue
        by_team[canonical][award].append({
            "year": year, "player": player, "position": position,
        })
    # Sort each award list by year asc
    for team in by_team:
        for award in by_team[team]:
            by_team[team][award].sort(key=lambda r: r["year"])
    return by_team


def read_hall_of_fame(wb):
    """
    Hall of Fame: A player, B year inducted, C category, D position, E birth,
    F death, G age, H-M: 3 team slots (city/team pairs), N: cur. pos, O-Q: current
    franchise names for the 3 slots, R-S: first/last name.
    v1: attribute to primary team (#1 slot, col O = canonical name of slot 1).
    """
    ws = wb["Hall of Fame"]
    by_team = defaultdict(list)
    for row in ws.iter_rows(min_row=2, values_only=True):
        row = list(row) + [None] * max(0, 19 - len(row))
        player = safe_str(row[0])
        year_ind = safe_int(row[1], None)
        category = safe_str(row[2])
        position = safe_str(row[3])
        primary_canonical = safe_str(row[14])  # O: #1 Cur Nam
        if not (player and year_ind and primary_canonical):
            continue
        by_team[primary_canonical].append({
            "year": year_ind, "player": player, "category": category, "position": position,
        })
    for team in by_team:
        by_team[team].sort(key=lambda r: r["year"])
    return by_team


def read_top_games(wb):
    """
    Regular Season sheet: 36,391 rows (2 per game), 142 cols. Extract game-level
    metadata + DU Game Score (column index 124). Dedupe by Game ID (EE = col 134)
    so each game appears once league-wide, but ALSO keep per-team variants so a
    franchise page can show the game from that team's perspective.

    Returns:
      games: list of dicts, one per game (deduped by EE), with both teams labeled
      by_team_canonical: dict[canonical_name] -> list of game dicts, this-team
        perspective preserved (W/L from this team's row)
    """
    ws = wb["Regular Season"]
    # Column indices (0-based) from Claude Notes section 3:
    #   A=0 League, B=1 Season(Year), D=3 Week, E=4 Reg/Play, F=5 Play.Type,
    #   H=7 Date, K=10 City, L=11 Team, M=12 W/L/T,
    #   O=14 OppCity, P=15 OppTeam, Q=16 PF, R=17 PA, S=18 OT,
    #   T=19 Stadium era, U=20 StadArea, V=21 StadState, W=22 H/A
    #   DK=114 Name (this canonical), DL=115 Opp canonical,
    #   DU=124 Game Score (refined), EE=134 GameID
    IDX = dict(league=0, year=1, week=3, regplay=4, ptype=5, date=7,
               city=10, team=11, result=12, opp_city=14, opp_team=15,
               pf=16, pa=17, ot=18, stadium=19, ha=22,
               dk=114, dl=115, du=124, ee=134)
    seen_game_ids = set()
    all_games = []
    by_team = defaultdict(list)

    for row in ws.iter_rows(min_row=2, values_only=True):
        # Skip rows with no DU score (incomplete / projection rows).
        if len(row) <= IDX["du"]:
            continue
        du = row[IDX["du"]]
        if du is None or du == "":
            continue
        try:
            du_val = float(du)
        except (ValueError, TypeError):
            continue
        # Pull all the fields we need.
        year = safe_int(row[IDX["year"]], None)
        if year is None:
            continue
        week = row[IDX["week"]] if len(row) > IDX["week"] else None
        regplay = safe_str(row[IDX["regplay"]])
        ptype = safe_str(row[IDX["ptype"]])
        city = safe_str(row[IDX["city"]])
        team = safe_str(row[IDX["team"]])
        result = safe_str(row[IDX["result"]])  # W / L / T
        opp_city = safe_str(row[IDX["opp_city"]])
        opp_team = safe_str(row[IDX["opp_team"]])
        pf = safe_int(row[IDX["pf"]])
        pa = safe_int(row[IDX["pa"]])
        ot = safe_str(row[IDX["ot"]])
        stadium = safe_str(row[IDX["stadium"]])
        ha = safe_str(row[IDX["ha"]])
        dk = safe_str(row[IDX["dk"]]) if len(row) > IDX["dk"] else ""
        dl = safe_str(row[IDX["dl"]]) if len(row) > IDX["dl"] else ""
        ee = safe_str(row[IDX["ee"]]) if len(row) > IDX["ee"] else ""

        # Canonical date (YYYY-MM-DD) for display. openpyxl returns datetime
        # for Excel-formatted date columns; fall back to None if the cell is
        # blank or non-date (very early-era rows occasionally have just a year).
        date_raw = row[IDX["date"]] if len(row) > IDX["date"] else None
        date_iso = None
        try:
            if hasattr(date_raw, "strftime"):
                date_iso = date_raw.strftime("%Y-%m-%d")
        except Exception:
            date_iso = None

        # This-team-perspective row (always captured for the franchise page).
        if dk:
            by_team[dk].append({
                "year": year,
                "date": date_iso,
                "week": week if isinstance(week, (int, float)) else None,
                "round": resolve_round(date_iso, dk, dl, ptype or regplay),
                "team_city": city, "team": team, "team_canonical": dk,
                "opp_city": opp_city, "opp_team": opp_team, "opp_canonical": dl,
                "pf": pf, "pa": pa,
                "result": result,
                "ot": ot == "OT",
                "stadium": stadium,
                "is_home": ha == "vs",
                "du": round(du_val, 4),
            })

        # League-wide dedupe. EE = Date & this-team & opp, which is not
        # symmetric across the two rows of the same game (one row has
        # date+A+B, the other date+B+A). Build a symmetric key instead.
        date_val = row[IDX["date"]] if len(row) > IDX["date"] else None
        sym_key = f"{date_val}|{min(dk, dl)}|{max(dk, dl)}" if dk and dl else None
        if sym_key and sym_key in seen_game_ids:
            continue
        if sym_key:
            seen_game_ids.add(sym_key)
        # Determine winner (canonical name) for the league-wide row
        if result == "W":
            winner_canonical, loser_canonical = dk, dl
            winner_city, winner_team = city, team
            loser_city, loser_team = opp_city, opp_team
            winner_pf, loser_pa = pf, pa
        elif result == "L":
            winner_canonical, loser_canonical = dl, dk
            winner_city, winner_team = opp_city, opp_team
            loser_city, loser_team = city, team
            winner_pf, loser_pa = pa, pf
        else:
            # Tie: arbitrary "team A / team B" naming
            winner_canonical, loser_canonical = dk, dl
            winner_city, winner_team = city, team
            loser_city, loser_team = opp_city, opp_team
            winner_pf, loser_pa = pf, pa

        all_games.append({
            "year": year,
            "date": date_iso,
            "week": week if isinstance(week, (int, float)) else None,
            "round": resolve_round(date_iso, winner_canonical, loser_canonical, ptype or regplay),
            "winner_city": winner_city, "winner_team": winner_team, "winner_canonical": winner_canonical,
            "loser_city": loser_city, "loser_team": loser_team, "loser_canonical": loser_canonical,
            "winner_score": winner_pf,
            "loser_score": loser_pa,
            "ot": ot == "OT",
            "is_tie": result == "T",
            "stadium": stadium,
            "du": round(du_val, 4),
        })

    return all_games, by_team


def read_pro_bowl_counts(wb):
    """
    Pro Bowl: count selections per canonical franchise (col O Name).
    v1 surfaces counts only; full list is v2.
    """
    ws = wb["Pro Bowl"]
    counts = defaultdict(int)
    for row in ws.iter_rows(min_row=2, values_only=True):
        row = list(row) + [None] * max(0, 18 - len(row))
        canonical = safe_str(row[14])
        if canonical:
            counts[canonical] += 1
    return dict(counts)


# -------- Output builders --------

def build_franchises(lookup, totals, year_by_year, external_links):
    """
    Active 32 franchises. Combines Lookup (current city/team/metro/etc) with
    Totals (all-time stats). Slug = slugify(city + "-" + team).
    """
    franchises = []
    for team_name, lk in lookup.items():
        tot = totals.get(team_name)
        if not tot:
            print(f"WARN: no Totals row for active franchise '{team_name}'", file=sys.stderr)
            continue
        seasons = year_by_year.get(team_name, [])
        founding_year = seasons[0]["year"] if seasons else None
        # Detect prior cities by inspecting historical city values
        cities_seen = []
        for s in seasons:
            c = s.get("city")
            if c and c not in cities_seen:
                cities_seen.append(c)
        current_city = lk["city"]
        prior_cities = [c for c in cities_seen if c != current_city]
        # Slug from city + team
        slug = slugify(f"{lk['city']}-{lk['team']}")
        franchises.append({
            "slug": slug,
            "name": f"{lk['city']} {lk['team']}",
            "canonical": team_name,
            "city": lk["city"],
            "team": lk["team"],
            "league": "NFL",
            "conf": lk["conf"],
            "division": lk["division"],
            "stadium": lk["stadium"],
            "stadium_city": lk["stadium_city"],
            "metro": lk["metro"],
            "metro_slug": metro_slug_from_metro_area(lk["metro"]),
            "state": lk["state"],
            "founding_year": founding_year,
            "prior_cities": prior_cities,
            "wikipedia_url": (external_links.get(f"{lk['city']} {lk['team']}") or {}).get("wikipedia_url"),
            "wikidata_qid": (external_links.get(f"{lk['city']} {lk['team']}") or {}).get("wikidata_qid"),
            "championships": tot["championships"],
            "division_titles": tot["division_titles"],
            "playoff_appearances": tot["playoff_appearances"],
            "all_time_w": tot["all_time_w"],
            "all_time_l": tot["all_time_l"],
            "all_time_t": tot["all_time_t"],
            "win_pct": round(tot["win_pct"], 4),
            "playoff_w": tot["playoff_w"],
            "playoff_l": tot["playoff_l"],
            "playoff_t": tot["playoff_t"],
            "playoff_win_pct": round(tot["playoff_win_pct"], 4),
            "conf_finals_app": tot["conference_finals_app"],
            "conf_finals_wins": tot["conference_finals_wins"],
            "seasons": tot["seasons"],
            "seasons_500_plus": tot["seasons_500_plus"],
            "last_championship": tot["last_championship"],
            "reg_games": tot["reg_games"],
            "play_games": tot["play_games"],
            "total_games": tot["total_games"],
        })
    # Sort by championships desc, then win pct desc
    franchises.sort(key=lambda f: (-f["championships"], -f["win_pct"]))
    return franchises


def build_championships(year_by_year):
    """
    Per-franchise list of championship years, with era flag.
    Era flag: 'pre_sb' (1920-1965) or 'sb' (1966+).
    """
    out = defaultdict(list)
    for canonical, seasons in year_by_year.items():
        for s in seasons:
            if s["championship"]:
                era = "sb" if s["year"] >= 1966 else "pre_sb"
                entry = {
                    "year": s["year"],
                    "era": era,
                    "league": s["league"],
                    "record": f"{s['w']}-{s['l']}-{s['t']}",
                    "season_city": s["city"],
                    "season_team": s["team_historical"],
                }
                stolen_key = (canonical, s["year"])
                if stolen_key in STOLEN_TITLES:
                    entry["stolen"] = True
                    entry["stolen_note"] = STOLEN_TITLES[stolen_key]
                out[canonical].append(entry)
    # Sort by year asc within each franchise
    return {k: sorted(v, key=lambda r: r["year"]) for k, v in out.items()}


def build_championship_appearances(year_by_year):
    """
    Per-franchise list of championship-game appearances (won OR lost),
    with era flag and an 'is_winner' boolean. Mirrors build_championships
    in shape so the renderer can iterate either list with the same colors.
    """
    out = defaultdict(list)
    for canonical, seasons in year_by_year.items():
        for s in seasons:
            if s["championship_appearance"] or s["championship"]:
                era = "sb" if s["year"] >= 1966 else "pre_sb"
                out[canonical].append({
                    "year": s["year"],
                    "era": era,
                    "league": s["league"],
                    "is_winner": s["championship"],
                    "record": f"{s['w']}-{s['l']}-{s['t']}",
                    "season_city": s["city"],
                    "season_team": s["team_historical"],
                })
    return {k: sorted(v, key=lambda r: r["year"]) for k, v in out.items()}


# Approved display names for defunct NFL franchises. Keyed by the franchise
# canonical/`name` as it appears in public/data/nfl/historical.json. Maps the
# terse workbook short name (e.g. "Bulldogs (Canton)") to the city/team identity
# each club is most commonly known by. Slugs are unchanged.
DEFUNCT_DISPLAY_NAMES = {
    "Bulldogs (Canton)": "Canton Bulldogs",
    "Indians (Akron)": "Akron Pros",
    "Wolverines": "Detroit Wolverines",
    "Yellow Jackets": "Frankford Yellow Jackets",
    "Steam Roller": "Providence Steam Roller",
    "Bulldogs (Boston)": "Pottsville Maroons",
    "Colts (D)": "Baltimore Colts (1950)",
    "Texans (D)": "Dallas Texans (1952)",
    "Dodgers": "Brooklyn Dodgers (NFL)",
    "Stapletons": "Staten Island Stapletons",
    "Bills (D)": "Buffalo Bills (AAFC)",
    "Bisons": "Buffalo Bisons",
    "Hornets": "Chicago Hornets",
    "Tigers (Chi)": "Chicago Tigers",
    "Carpitts": "Card-Pitt (1944)",
    "Celts": "Cincinnati Celts",
    "Gunners": "Cincinnati Reds",
    "Tigers (Clev)": "Cleveland Tigers",
    "Indians (Clev)": "Cleveland Indians (NFL)",
    "Tigers (Colm)": "Columbus Panhandles",
    "Tigers": "Dayton Triangles",
    "Panthers (Det)": "Detroit Panthers",
    "Heralds": "Detroit Heralds",
    "Tornadoes (NJ)": "Duluth Eskimos",
    "Crimson Giants": "Evansville Crimson Giants",
    "Pros": "Hammond Pros",
    "Blues": "Hartford Blues",
    "Cowboys (KC)": "Kansas City Cowboys",
    "Buccaneers (LA)": "Los Angeles Buccaneers",
    "Dons": "Los Angeles Dons (AAFC)",
    "Colonels": "Louisville Colonels",
    "Seahawks (Mia)": "Miami Seahawks (AAFC)",
    "Badgers": "Milwaukee Badgers",
    "Redjackets": "Minneapolis Red Jackets",
    "Flyers": "Muncie Flyers",
    "Brickley's Giants": "New York Brickley's Giants",
    "Yankees": "New York Yankees (AAFC)",
    "Indians (Marion)": "Oorang Indians",
    "Steagles": "Phil-Pitt Steagles",
    "Tornadoes": "Racine Tornadoes",
    "Jeffersons": "Rochester Jeffersons",
    "Independents": "Rock Island Independents",
    "All-Stars": "St. Louis All-Stars",
    "Maroons": "Toledo Maroons",
    "Kardex": "Tonawanda Kardex",
    "Senators (W)": "Washington Senators (NFL)",
}


def build_historical(totals, year_by_year):
    """Defunct franchises for /teams/nfl/historical.

    Adds first_year and last_year computed from Year-by-Year so the page
    can show a true active-range column and sort on it. Also applies any
    CITY_HISTORY_OVERRIDES, and marks any franchise that has a stolen
    championship entry so the page can lift it into the champions tier
    of the default sort.
    """
    stolen_by_canon = {canon for (canon, _yr) in STOLEN_TITLES.keys()}
    rows = []
    for canonical, tot in totals.items():
        if tot["is_current"]:
            continue
        seasons = year_by_year.get(canonical, [])
        years = [s["year"] for s in seasons if s.get("year")]
        first_year = min(years) if years else None
        last_year = max(years) if years else None
        city = CITY_HISTORY_OVERRIDES.get(canonical, tot["city_history"])
        rows.append({
            "canonical": canonical,
            "name": canonical,
            "display_name": DEFUNCT_DISPLAY_NAMES.get(canonical) or (f"{city} {canonical}".strip() or canonical),
            "metro": _rmnfl(canonical, city)[0],
            "metro_slug": _rmnfl(canonical, city)[1],
            "city": city,
            "team_historical": tot["team_history"],
            "league": tot["league"],
            "seasons": tot["seasons"],
            "first_year": first_year,
            "last_year": last_year,
            "w": tot["all_time_w"],
            "l": tot["all_time_l"],
            "t": tot["all_time_t"],
            "win_pct": round(tot["win_pct"], 4),
            "championships": tot["championships"],
            "stolen_championships": 1 if canonical in stolen_by_canon else 0,
        })
    rows.sort(key=lambda r: (
        -r["championships"],
        -r["stolen_championships"],
        r["city"] or "",
    ))
    return rows


def build_top_games_by_team(by_team, franchises, top_n=12):
    """For each active franchise (slug-keyed), return top N games by DU score
    from that team's perspective."""
    out = {}
    for f in franchises:
        canonical = f["canonical"]
        rows = sorted(by_team.get(canonical, []), key=lambda g: -g["du"])[:top_n]
        out[f["slug"]] = rows
    return out


def build_top_games_all_time(all_games, top_n=50):
    """Top N games across the league by DU score, deduped by GameID."""
    return sorted(all_games, key=lambda g: -g["du"])[:top_n]


def build_top_games_by_decade(all_games, top_n_per_decade=10):
    """Top N games per decade. 1920s bucket = years 1920-1929, etc."""
    by_decade = defaultdict(list)
    for g in all_games:
        decade = (g["year"] // 10) * 10
        by_decade[decade].append(g)
    out = {}
    for decade, games in by_decade.items():
        out[str(decade)] = sorted(games, key=lambda g: -g["du"])[:top_n_per_decade]
    return out


def build_historical_championships(year_by_year, totals):
    """Championship years for defunct franchises with stolen-title flagging."""
    out = defaultdict(list)
    for canonical, seasons in year_by_year.items():
        tot = totals.get(canonical)
        if not tot or tot["is_current"]:
            continue
        for s in seasons:
            if s["championship"]:
                entry = {
                    "year": s["year"], "era": "pre_sb",
                    "league": s["league"],
                    "season_city": s["city"], "season_team": s["team_historical"],
                }
                stolen_key = (canonical, s["year"])
                if stolen_key in STOLEN_TITLES:
                    entry["stolen"] = True
                    entry["stolen_note"] = STOLEN_TITLES[stolen_key]
                out[canonical].append(entry)
        for (sname, syear), note in STOLEN_TITLES.items():
            if sname == canonical and not any(e["year"] == syear for e in out[canonical]):
                out[canonical].append({
                    "year": syear, "era": "pre_sb", "stolen": True, "stolen_note": note,
                })
    return {k: sorted(v, key=lambda r: r["year"]) for k, v in out.items()}


def main():
    src = find_source_xlsx(sys.argv)
    print(f"Reading: {src}")
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    print(f"Sheets: {wb.sheetnames}")

    print("Reading Lookup block 3 (current 32)...")
    lookup = read_lookup_block3(wb)
    print(f"  {len(lookup)} active teams")

    print("Reading Totals...")
    totals = read_totals(wb)
    print(f"  {len(totals)} franchise rows (active + defunct)")

    print("Reading Year by Year...")
    yby = read_year_by_year(wb)
    print(f"  {len(yby)} canonical names, {sum(len(v) for v in yby.values())} team-seasons")

    print("Reading Stadiums...")
    stadiums = read_stadiums(wb)
    print(f"  {len(stadiums)} teams with stadium history")

    print("Reading Awards...")
    awards = read_awards(wb)
    print(f"  {len(awards)} teams with awards")

    print("Reading Hall of Fame...")
    hof = read_hall_of_fame(wb)
    print(f"  {len(hof)} teams with HoF inductees")

    print("Reading Pro Bowl counts...")
    pb_counts = read_pro_bowl_counts(wb)
    print(f"  {sum(pb_counts.values())} total selections across {len(pb_counts)} teams")

    print("Reading Regular Season top games (DU score)...")
    all_games, games_by_team_canonical = read_top_games(wb)
    print(f"  {len(all_games)} unique games scored, {len(games_by_team_canonical)} franchises with game data")

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    print("Reading team-wikidata.tsv (Wikipedia/Wikidata cross-links)...")
    external_links = read_team_external_links()
    print(f"  {len(external_links)} teams with external links")

    franchises = build_franchises(lookup, totals, yby, external_links)
    print(f"Built franchises: {len(franchises)}")
    (OUT_DIR / "franchises.json").write_text(json.dumps(franchises, indent=2, ensure_ascii=False), encoding="utf-8")

    champs = build_championships(yby)
    (OUT_DIR / "championships.json").write_text(json.dumps(champs, indent=2, ensure_ascii=False), encoding="utf-8")

    champ_apps = build_championship_appearances(yby)
    (OUT_DIR / "championship-appearances.json").write_text(json.dumps(champ_apps, indent=2, ensure_ascii=False), encoding="utf-8")

    (OUT_DIR / "stadium-history.json").write_text(json.dumps(stadiums, indent=2, ensure_ascii=False), encoding="utf-8")

    awards_plain = {team: dict(awards_by_type) for team, awards_by_type in awards.items()}
    (OUT_DIR / "award-winners.json").write_text(json.dumps(awards_plain, indent=2, ensure_ascii=False), encoding="utf-8")

    historical = build_historical(totals, yby)
    print(f"Built historical: {len(historical)}")
    (OUT_DIR / "historical.json").write_text(json.dumps(historical, indent=2, ensure_ascii=False), encoding="utf-8")

    hist_champs = build_historical_championships(yby, totals)
    (OUT_DIR / "historical-championships.json").write_text(json.dumps(hist_champs, indent=2, ensure_ascii=False), encoding="utf-8")

    (OUT_DIR / "hall-of-fame.json").write_text(json.dumps(dict(hof), indent=2, ensure_ascii=False), encoding="utf-8")

    active_canonicals = {f["canonical"]: f["slug"] for f in franchises}
    seasons_out = {}
    for canonical, slug in active_canonicals.items():
        rows = yby.get(canonical, [])
        seasons_out[slug] = [
            {
                "year": r["year"], "league": r["league"], "city": r["city"],
                "team": r["team_historical"],
                "w": r["w"], "l": r["l"], "t": r["t"], "win_pct": round(r["win_pct"], 4),
                "pf": r["pf"], "pa": r["pa"],
                "division": r["division"], "place": r["place"],
                "playoff": r["playoff_appearance"],
                "div_title": r["division_title"],
                "conf_final": r["conference_final"],
                "champ_app": r["championship_appearance"],
                "champ": r["championship"],
            }
            for r in rows
        ]
    (OUT_DIR / "seasons-by-team.json").write_text(json.dumps(seasons_out, indent=2, ensure_ascii=False), encoding="utf-8")

    (OUT_DIR / "pro-bowl-counts.json").write_text(json.dumps(pb_counts, indent=2, ensure_ascii=False), encoding="utf-8")

    # Top Games (DU Game Score)
    top_by_team = build_top_games_by_team(games_by_team_canonical, franchises, top_n=12)
    (OUT_DIR / "top-games-by-team.json").write_text(json.dumps(top_by_team, indent=2, ensure_ascii=False), encoding="utf-8")

    top_all_time = build_top_games_all_time(all_games, top_n=50)
    (OUT_DIR / "top-games-all-time.json").write_text(json.dumps(top_all_time, indent=2, ensure_ascii=False), encoding="utf-8")

    top_by_decade = build_top_games_by_decade(all_games, top_n_per_decade=10)
    (OUT_DIR / "top-games-by-decade.json").write_text(json.dumps(top_by_decade, indent=2, ensure_ascii=False), encoding="utf-8")

    print("\nWrote:")
    for f in sorted(OUT_DIR.glob("*.json")):
        print(f"  {f.relative_to(REPO_ROOT)}  ({f.stat().st_size:,} bytes)")


if __name__ == "__main__":
    main()

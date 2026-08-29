#!/usr/bin/env python3
"""
Extract data from MetroAreas.xlsx and generate JSON files for the website.

Usage:
  python scripts/extract.py [path/to/MetroAreas.xlsx]

If no path is given, looks for MetroAreas.xlsx in the parent directory.

Outputs:
  public/data/metros.json - Main rankings (all 4,285 metros)
  public/data/regions.json - Regional aggregates
  public/data/details/<slug>.json - Per-metro detail files
"""

import json
import math
import os
import re
import sys
import unicodedata
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--quiet", "--break-system-packages"])
    import openpyxl


# Characters that Unicode decomposition does NOT split into base + combining
# mark. Without these, NFKD leaves them whole and the ASCII filter deletes
# them outright.
_SLUG_CHAR_MAP = {
    'ł': 'l',   # l with stroke  (Łódź)
    'ø': 'o',   # o with stroke  (Andøy)
    'đ': 'd',   # d with stroke
    'ß': 'ss',  # sharp s
    'æ': 'ae',
    'œ': 'oe',
    'þ': 'th',  # thorn
    'ð': 'd',   # eth
    'ı': 'i',   # dotless i     (Diyarbakır)
    'ħ': 'h',
    'ŋ': 'n',
    'ŧ': 't',
}

# Separators that must survive as a hyphen. Deleting them ran compound names
# together: "Bydgoszcz–Toruń" became "bydgoszcztoru", "Biel/Bienne" became
# "bielbienne".
_SLUG_SEPARATORS = '‐‑‒–—―−·•/'


def slugify(name):
    """Convert metro name to URL-safe slug.

    Decomposes Unicode and drops combining marks, so any accented Latin script
    transliterates instead of losing characters. The version this replaced
    hand-listed about thirty Latin-1 characters and deleted everything else,
    which silently mangled 133 metros — overwhelmingly Polish, Romanian,
    Turkish, Czech and Bosnian, because the hand-written list covered Western
    European orthography and stopped at the Oder. Łódź became "od" and Huế
    became "hu".

    Slugs ARE live URLs. Any change here moves indexed pages, so it is gated by
    scripts/check-slug-drift.mjs against lib/metroRedirects.json: a slug cannot
    leave the build without a redirect covering it.
    """
    s = name.lower().strip()
    for ch in _SLUG_SEPARATORS:
        s = s.replace(ch, ' ')
    for ch, repl in _SLUG_CHAR_MAP.items():
        s = s.replace(ch, repl)
    s = unicodedata.normalize('NFKD', s)
    s = ''.join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r'[^a-z0-9\s-]', '', s)
    s = re.sub(r'[\s]+', '-', s)
    s = re.sub(r'-+', '-', s)
    return s.strip('-')


def safe_int(val, default=0):
    if val is None:
        return default
    try:
        return int(val)
    except (ValueError, TypeError):
        return default


def safe_float(val, default=0.0):
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def safe_str(val):
    if val is None:
        return ''
    return str(val).strip()


def build_country_continent_map(wb):
    """Canonical country -> continent map from the Country Populations sheet.

    Country Populations has duplicate rows for some countries (UK has 16,
    France 12, etc.) corresponding to constituent countries / overseas
    territories. The FIRST row per country is the sovereign-state summary
    and carries the right continent (col 13). Subsequent rows for
    constituents may have stale or contradictory continent values, so we
    take only the first occurrence per country name.

    Verified by dry run 2026-08-04: 199 entries -> 247, the junk header entry
    gone, every territory keyed under its own name, 'United Kingdom' retained,
    and the metros relying on the col-41 fallback drop from 50 to 0. No metro
    changes continent today - col 41 happened to agree for all 50 - so this is
    a latent-bug fix, not a data correction.
    """
    ws = wb["Country Populations"]
    # The sheet's header is on ROW 3 (row 1 holds stray working cells, row 2 is
    # blank), so data starts at row 4. min_row=2 previously ingested the header
    # itself as data, leaving a junk {"Parent Country": "Continent"} entry.
    #
    # Col A is PARENT Country and col H is Country. They match for the 197
    # sovereign states that parent themselves, but differ for the 49
    # constituents and territories (England, Hong Kong, Puerto Rico, Cayman
    # Islands...). Keying on col A alone therefore never gave those territories
    # an entry under their own name and they fell through to the untrusted
    # Metro Areas col 41. Keying on col H ALONE is also wrong: the UK has no
    # self row, so 'United Kingdom' would vanish and every UK metro would fall
    # back instead.
    #
    # So: pass 1 keys on the real country name, pass 2 backfills any parent
    # that has no entry of its own. Two passes, not one, so a real country row
    # always beats a parent-derived one. First occurrence still wins within
    # each pass, which keeps the sovereign-state summary ahead of any stale
    # constituent rows.
    rows = [r for r in ws.iter_rows(min_row=4, values_only=True) if r]
    out = {}
    for row in rows:                       # pass 1: Country (col H)
        if len(row) <= 13:
            continue
        country, cont = row[7], row[13]
        if country and cont and country not in out:
            out[country] = str(cont).strip()
    for row in rows:                       # pass 2: backfill Parent (col A)
        if len(row) <= 13:
            continue
        parent, cont = row[0], row[13]
        if parent and cont and parent not in out:
            out[parent] = str(cont).strip()
    return out


def build_score_index(wb, xlsx_path):
    """Score every metro in Python, replacing the workbook's cached column BG.

    Until 2026-08-10 the score was read straight out of BG. That made Excel a
    hard dependency of every data refresh: the ETL read a CACHED value, so any
    input that changed without Excel opening left the site serving a score
    computed from data it was no longer displaying. It happened on 2026-08-09
    and went unnoticed for a day.

    Now the score is computed here from the same source sheets the formula
    reads. BG stays in the workbook as a cross-check, and check:score-parity
    reports when the two disagree - which, until Excel next recalculates, they
    legitimately will.

    METRO_SCORE_SOURCE=workbook restores the old behaviour. It exists so the
    cutover could be A/B'd against the previous output, and so there is a way
    back if the engine ever misbehaves in production. Remove it once the
    workbook's BG column is retired.
    """
    if os.environ.get("METRO_SCORE_SOURCE", "engine").lower() == "workbook":
        print("  score: reading cached BG from the workbook (METRO_SCORE_SOURCE=workbook)")
        return None
    # extract.py is always run as a script, so scripts/ is already sys.path[0]
    # and the package resolves. Belt and braces for anyone importing it.
    _here = str(Path(__file__).resolve().parent)
    if _here not in sys.path:
        sys.path.insert(0, _here)
    from metro_score import sources as _src, score as _score, weights as _w
    engine = _score.Engine(_src.from_openpyxl(wb, Path(xlsx_path)), _w.load())
    # Keyed the way extract_metros names a metro (safe_str, which strips), NOT
    # by the engine's internal join key, which is deliberately unstripped to
    # match Excel. Getting this wrong would silently fall back to the cached BG
    # for the affected metro, so a miss is reported rather than swallowed.
    index, drift = {}, []
    for name, _k, cached, computed, _terms, _cols in engine.rows():
        index[name.strip().lower()] = computed
        if abs(computed - cached) > 1e-9:
            drift.append((abs(computed - cached), name, cached, computed))
    drift.sort(reverse=True)
    print(f"  score: computed in Python for {len(index):,} metros")
    if drift:
        print(f"  score: {len(drift)} differ from the workbook's cached BG "
              f"(stale until Excel recalculates; the computed value is the one used)")
        for d, name, cached, computed in drift[:10]:
            print(f"           {name:<30} BG {cached:12.6f} -> {computed:12.6f}  (+{computed - cached:.6f})")
        if len(drift) > 10:
            print(f"           ... and {len(drift) - 10} more")
    return index


def extract_metros(wb, score_index=None):
    """Extract main metro data from the Metro Areas sheet."""
    ws = wb["Metro Areas"]
    metros = []
    unscored = []
    # Derive continent from the country join rather than trusting Metro
    # Areas col 41, which has hundreds of stale or wrong values (e.g.
    # Mangaluru / International Falls / Shima tagged 'Europe'). The
    # workbook column stays as a fallback when the country has no
    # continent in Country Populations.
    country_continent = build_country_continent_map(wb)

    for row in ws.iter_rows(min_row=4, values_only=True):
        v = list(row)
        name = safe_str(v[5])
        # Column BG is the workbook's cached score. The engine recomputes it
        # from the same source sheets; see build_score_index().
        score = safe_float(v[58])
        if not name:
            continue
        if score_index is not None:
            computed = score_index.get(name.strip().lower())
            if computed is None:
                unscored.append(name)
            else:
                score = computed

        pop = safe_int(v[9])
        lat = safe_float(v[63])
        lon = safe_float(v[64])
        region = safe_str(v[65]) or safe_str(v[41])
        slug = slugify(name)

        # Wikidata QID (col BT / idx 71) and Wikipedia URL (col BU / idx 72).
        # Populated for Top 25 metros as of 2026-04-24; ranks 26+ intentionally
        # blank. Omit keys entirely when absent so the JSON payload stays clean
        # and downstream sameAs arrays never emit null placeholders.
        qid = safe_str(v[71]) if len(v) > 71 else ''
        wiki_url = safe_str(v[72]) if len(v) > 72 else ''

        metro = {
            'slug': slug,
            'name': name,
            'country': safe_str(v[0]),
            'subCountry': safe_str(v[1]),
            'language': safe_str(v[2]),
            'capital': safe_str(v[3]),
            'gawcClass': safe_str(v[4]),
            'primaryState': safe_str(v[6]),
            'state2': safe_str(v[7]),
            'state3': safe_str(v[8]),
            'pop': pop,
            'region': region,
            'continent': country_continent.get(safe_str(v[0])) or safe_str(v[41]),
            'score': round(score, 1),
            '_scoreFull': score,  # sort key only; deleted after ranking
            'lat': lat,
            'lon': lon,
            'primaryCity': safe_str(v[61]),
            'primaryCityCountry': safe_str(v[62]),
            'gdp': round(safe_float(v[69]), 1),
            'gdpPerCapita': round(safe_float(v[70])),
            # Raw dimension data
            'dims': {
                'majorLeagueTeams': safe_int(v[43]),
                'totalTeams': safe_int(v[42]),
                'majorSportingEvents': safe_int(v[44]),
                'companies': safe_int(v[45]),
                'marketCap': safe_float(v[46]),
                'culturalEvents': safe_int(v[47]),
                'universities': safe_int(v[48]),
                'topUniHospResearch': safe_int(v[49]),
                'museumsLandmarks': safe_int(v[50]),
                'portsExchangesInfra': safe_int(v[51]),
                'airportScore': safe_float(v[52]),
                'luxuryStars': safe_float(v[53]),
                'metroStations': safe_int(v[54]),
                'suburbStations': safe_int(v[55]),
                'trainHubs': safe_int(v[56]),
                'skyscrapers': safe_int(v[57]),
            },
            'naRank': safe_int(v[59]) if v[59] else None,
            'pctOfCountry': round(safe_float(v[67]) * 100, 1) if v[67] else 0,
        }
        if qid:
            metro['qid'] = qid
        if wiki_url:
            metro['wikipediaUrl'] = wiki_url
        metros.append(metro)

    if unscored:
        # A miss means the engine and this loop disagree about a metro's name,
        # and the metro quietly kept its stale cached score. Loud on purpose.
        raise SystemExit(
            f"ERROR: the score engine returned no value for {len(unscored)} metro(s): "
            f"{unscored[:5]}{' ...' if len(unscored) > 5 else ''}"
        )

    # Sort by the FULL-PRECISION score descending and assign global rank; the
    # displayed score stays rounded to 1dp. Sorting on the rounded value made
    # ties out of every 0.1-wide bucket, so ~3,000 metros held whatever rank
    # the workbook's ROW ORDER gave them (stable sort), and re-sorting the
    # sheet in Excel silently reshuffled the site. Full precision has 4,311
    # distinct values across 4,314 metros, so rank no longer depends on row
    # order. Measured before the change (2026-08-10): 2,993 ranks move, max
    # displacement 892, zero slugs move, zero displayed scores change.
    metros.sort(key=lambda x: x['_scoreFull'], reverse=True)
    for i, m in enumerate(metros):
        m['rank'] = i + 1
        del m['_scoreFull']

    # Resolve slug collisions HERE, before anything downstream keys off a slug.
    # Two metros can transliterate to the same slug — Kochi/Kōchi,
    # Cordoba/Córdoba, Leon/León, Macon/Mâcon, Merida/Mérida, Beja/Béja — and
    # the rule is that the higher-ranked one keeps the bare slug while the other
    # takes a country suffix.
    #
    # This used to happen much later, inside the detail-writing loop, by which
    # point compute_dimension_ranks() had already built its dict keyed on the
    # UNRESOLVED slug. So both colliding metros shared one entry: the loser
    # silently overwrote the winner's dimension ranks, and the suffixed metro
    # got none at all. Every existing collision was shipping with that damage.
    #
    # NOTE the tie-break is RANK, so a ranking change alone can move a live
    # URL with no code change. scripts/check-slug-drift.mjs exists to catch it.
    seen = {}
    for m in metros:
        slug = m['slug']
        if slug in seen:
            slug = f"{slug}-{m['country'].lower().replace(' ', '-')}"
            m['slug'] = slug
        seen[slug] = m

    return metros


def _normalize_league(raw):
    """Map internal league labels to the names shown on the site."""
    s = safe_str(raw)
    if s == "Major Venues":
        return "Notable Venues"
    return s


# Manual venue name aliases so the same physical venue dedupes across sports.
# Extend here rather than letting the frontend guess. The first tuple entry is
# a normalized form matched as a substring; the value is the canonical name.
_VENUE_NAME_ALIASES = {
    "New Wembley Stadium": "Wembley Stadium",
    "The O2 Arena": "O2 Arena",
}


def _normalize_venue_name(league, name):
    """Canonicalize venue names so a venue appearing under multiple sports
    (e.g. combat sports + tennis at the O2) dedupes cleanly on the frontend."""
    if league != "Notable Venues":
        return name
    return _VENUE_NAME_ALIASES.get(name, name)


def extract_teams(wb):
    """Extract teams grouped by metro."""
    ws = wb["Team List"]
    teams = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        v = list(row)
        metro = safe_str(v[6])
        if not metro:
            continue
        league = _normalize_league(v[1])
        level = safe_str(v[9])
        # Main Division (col 3) holds the authoritative tag for NCAA minor-sport
        # rows the user has put under generic league labels like "Other Sports" or
        # "Other Women Sports" (e.g. NCAA Wrestling, NCAA W Soccer, NCAA Baseball).
        # Promote level to "College" whenever Main Division starts with "NCAA" so
        # downstream bucketing routes these teams into College/University Teams
        # regardless of league label.
        main_division = safe_str(v[3])
        if main_division.upper().startswith("NCAA") and level != "College":
            level = "College"
        # Col L (idx 11) = "Gold Standard" flag, inserted 2026-05-17 as a
        # workbook-driven source for the apex top-flight designation per
        # sport-league. Replaces the hardcoded lib/goldStandard.ts map; a
        # curated Football men's Big 5 override is layered on the TS side
        # because the workbook intentionally leaves men's Football blank.
        # Cell value 'Gold' (case-insensitive) flags the row; anything else
        # (including blank) does not.
        gold_marker = safe_str(v[11])
        is_gold = gold_marker.upper() == 'GOLD'
        # Col M (idx 12) = "Major League" marker, shifted from idx 11 when the
        # Gold Standard column was inserted. Same three-value pattern as before:
        #   ""           → not a major-league row
        #   "Y"          → major-league row, league label comes from col 1 as-is
        #   "<TierName>" → major-league row AND the tier name overrides the
        #                  generic col 1 label. This is how Euroleague basketball
        #                  is encoded: col 1 = "Int'l Basketball" (the bucket),
        #                  col M = "Euroleague" (the actual tier the team
        #                  competes in). Surface the tier so the metro page
        #                  shows "Euroleague" rather than "Int'l Basketball".
        ml_marker = safe_str(v[12])
        is_major = ml_marker != ''
        league_for_team = ml_marker if (is_major and ml_marker != 'Y') else league
        team_entry = {
            'sport': safe_str(v[0]),
            'league': league_for_team,
            'team': _normalize_venue_name(league_for_team, safe_str(v[2])),
            'city': safe_str(v[5]),
            'country': safe_str(v[8]),
            'level': level,
            'major': is_major,
        }
        if is_gold:
            team_entry['gold'] = True
        # Col P (idx 15) = "Annual Event" flag, marked 'Y' for recurring
        # event-type entries in Team List (F1 Grands Prix, NASCAR races, Sailing
        # regattas, Powerboat races). These are teams in the source data but
        # behave like events on the site, so they route exclusively into the
        # "Annual Sporting Events" category on the metro page rather than the
        # Major League Teams or Other Teams buckets. Shifted from idx 14 when
        # the Gold Standard column was inserted at col L (2026-05-17).
        annual_flag = safe_str(v[15]) if len(v) > 15 else ''
        if annual_flag.upper() == 'Y':
            team_entry['annual'] = True
        # Cols Q/R (idx 16/17) = Wikidata QID and Wikipedia URL. Populated
        # as of 2026-04-24 for all US major league franchises (NFL/MLB/NBA/NHL)
        # plus every Canadian NHL team and the Toronto MLB/NBA franchises.
        # Omit the keys when empty so JSON-LD sameAs arrays never emit nulls.
        # Shifted from idx 15/16 when the Gold Standard column was inserted.
        qid = safe_str(v[16]) if len(v) > 16 else ''
        wiki_url = safe_str(v[17]) if len(v) > 17 else ''
        if qid:
            team_entry['qid'] = qid
        if wiki_url:
            team_entry['wikipediaUrl'] = wiki_url
        # Cols S/T (idx 18/19) = Lat / Long for the team's home venue (or for
        # venue-class rows, the venue itself). Used to render team and venue
        # markers on the metro detail page map. Both must be present and numeric
        # for the marker to plot; otherwise the entry is rendered in the written
        # sections only. Shifted from idx 17/18 when the Gold Standard column
        # was inserted at col L (2026-05-17).
        lat = safe_float(v[18]) if len(v) > 18 else 0
        lng = safe_float(v[19]) if len(v) > 19 else 0
        if lat or lng:
            team_entry['lat'] = lat
            team_entry['lng'] = lng
        teams.setdefault(metro, []).append(team_entry)

    # Note: RegTeams is intentionally NOT read. Team List is the single
    # source of truth for teams. RegTeams holds stale/regional legacy rows
    # (e.g. London Irish 2023) that must not reach the site.

    return teams


def extract_gold_standard_leagues(wb):
    """Build the sport -> set of league names map from Team List col L.

    The workbook owns Gold Standard designation as of 2026-05-17. Each row
    whose col L value is 'Gold' (case-insensitive) contributes its (Sport,
    League) pair to the map. The Major League override is honored: if col M
    is a tier-name override (e.g. 'Euroleague' for an Int'l Basketball row),
    that tier becomes the league key on the gold side too, so the JSON map
    aligns with the league strings emitted by extract_teams.

    Returns:
      { sport_name: [league_name, ...] }  (sorted, deduplicated)

    The Football men's Big 5 override (England / Spain / Italy / France /
    Germany) is intentionally NOT applied here. The workbook leaves men's
    Football blank by design; the curated override is layered on the TS
    side in lib/goldStandard.ts so the JSON file remains a pure mirror of
    the workbook.
    """
    ws = wb["Team List"]
    # Display-league re-bucketing. MUST stay in sync with the identical
    # MAIN_DIV_LABEL_LEAGUES / Int'l Basketball logic in
    # scripts/build-sports-index.py, which relabels these broad workbook
    # 'League' buckets to their Main Division (col D) when it emits each
    # team's `league` field. lib/goldStandard.isGoldStandardLeague compares
    # the gold key against that emitted league string, so the gold key must
    # use the same display label or it will never match (e.g. Rugby Union
    # clubs ship as 'Top 14', not the 'Dom. Rugby Union' bucket).
    MAIN_DIV_LABEL_LEAGUES = {
        "Minor Lg Base",
        "Int'l Volleyball",
        "Int'l Handball",
        "Minor/Jr/Int'l Hockey",
        "Dom. Rugby Union",
        "FBS",
        "FCS",
        "NCAA W",
        "College Hockey",
        "Other Sports",
        "Other Women Sports",
    }
    by_sport = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        v = list(row)
        if len(v) < 13:
            continue
        gold_marker = safe_str(v[11])
        if gold_marker.upper() != 'GOLD':
            continue
        sport = safe_str(v[0])
        league_raw = _normalize_league(v[1])
        main_div = safe_str(v[3])
        ml_marker = safe_str(v[12])
        # Mirror build-sports-index.py's display_league resolution so the
        # gold key equals the team's emitted league string.
        if league_raw == "Int'l Basketball":
            league = "EuroLeague" if ml_marker == "Euroleague" else (main_div or league_raw)
        elif league_raw == "Int'l W Basketball":
            league = "EuroLeague Women" if ml_marker == "Euroleague" else (main_div or league_raw)
        elif league_raw in MAIN_DIV_LABEL_LEAGUES or (sport == "Basketball" and league_raw == "NCAA"):
            league = main_div or league_raw
        else:
            league = league_raw
        if not sport or not league:
            continue
        by_sport.setdefault(sport, set()).add(league)
    return {sport: sorted(leagues) for sport, leagues in sorted(by_sport.items())}


def extract_universities(wb):
    """Extract universities grouped by metro.

    Cols Q/R (idx 16/17) carry Lat / Long for the campus. Populated for
    the global top-500 as of 2026-05-09; ranks 501-2000 still empty.
    Attach lat/lng only when both are present so the metro page map
    layer can filter cleanly.
    """
    ws = wb["Universities"]
    unis = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        v = list(row)
        metro = safe_str(v[5])
        if not metro:
            continue
        # A row with a metro but no World Rank is an institution that fell out of
        # the CWUR top 2000. From the 2026 edition those rows STAY in the
        # workbook, rank blanked, so their curated metro, city and coordinates
        # survive for the year the institution returns. They must not reach a
        # metro page: safe_int defaults to 0, so without this guard they would
        # emit as rank 0 and sort AHEAD of the number one university.
        if not safe_int(v[2]):
            continue
        entry = {
            'rank': safe_int(v[2]),
            'name': safe_str(v[3]),
            'city': safe_str(v[4]),
            'country': safe_str(v[0]),
        }
        lat = safe_float(v[16]) if len(v) > 16 else 0
        lng = safe_float(v[17]) if len(v) > 17 else 0
        if lat or lng:
            entry['lat'] = lat
            entry['lng'] = lng
        unis.setdefault(metro, []).append(entry)
    # Sort each metro's universities by rank
    for metro in unis:
        unis[metro].sort(key=lambda x: x['rank'] if x['rank'] else 9999)
    return unis


def extract_culture(wb):
    """Extract cultural assets and infrastructure grouped by metro."""
    ws = wb["Culture-Infra"]
    culture = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        v = list(row)
        metro = safe_str(v[6])
        if not metro:
            continue
        entry = {
            'name': safe_str(v[4]),
            'city': safe_str(v[5]),
            'subtype': safe_str(v[8]),
            'type': safe_str(v[11]),
            'majorType': safe_str(v[10]),
        }
        # Columns T/U (19/20) = Lat/Long. Empty until 2026-08-09: airports and
        # stations resolve by IATA code against OurAirports, everything else by
        # name+category against FSQ OS Places. Emitted only when BOTH are
        # present. Provenance lives in Supabase public.place_ids.
        clat = safe_float(v[19]) if len(v) > 19 else 0
        clng = safe_float(v[20]) if len(v) > 20 else 0
        if clat and clng:
            entry['lat'] = clat
            entry['lng'] = clng
        # Column O (index 14) = "Annual Event" flag, marked "Y" for recurring events
        annual_flag = safe_str(v[14]) if len(v) > 14 else ''
        if annual_flag.upper() == 'Y':
            entry['annual'] = True
        # Column P (index 15) = "# Stations", used on Metro System / Suburban Rail
        # so each line shows its own station count instead of just the aggregate.
        if entry['type'] in ('Metro System', 'Suburban Rail'):
            stations = safe_int(v[15]) if len(v) > 15 else 0
            if stations:
                entry['stations'] = stations
        # Override: America's Cup editions are championship moments, not annual events.
        # The xlsx flags them Annual=Y because they recur on a multi-year cadence, but
        # semantically each edition belongs in Championship Finals (like Super Bowl
        # or World Series), not Annual Sporting Events.
        if entry['type'] == 'Sporting Event' and entry['name'].startswith("America's Cup"):
            entry.pop('annual', None)
        culture.setdefault(metro, []).append(entry)
    return culture


def load_skyscrapers(wb, metros):
    """Per-metro 150m+/200m+/300m+ counts, keyed by SLUG.

    Source of record is public/data/skyscrapers.json, built from SKYDB by
    scripts/build-skyscrapers.py. The hand-curated Skyscrapers sheet remains in
    the workbook as the regression test that check-skyscrapers.mjs runs the API
    pull against, and as the fallback here.

    Why the swap: the sheet holds 7,727 buildings at 150m+, SKYDB 9,863. The
    gap is almost entirely in the 150-200 m band, which is where a manually
    maintained list quietly falls behind - at 300m+ the two are within 14%,
    because supertall buildings are famous and get noticed. SKYDB also refreshes
    monthly instead of whenever someone remembers.

    The fallback is loud on purpose. Silently reverting to a stale sheet is how
    a data regression ships without anyone noticing, so if the generated file is
    missing this says so in capitals and check-skyscrapers.mjs fails the build.
    """
    path = Path(__file__).resolve().parent.parent / "public" / "data" / "skyscrapers.json"
    if path.exists():
        payload = json.loads(path.read_text(encoding="utf-8"))
        data = payload.get("metros") or {}
        known = {m['slug'] for m in metros}
        # A slug in the file that no longer exists as a metro means the boundary
        # set and the metro list have drifted apart. Report it; do not emit it.
        orphans = [s for s in data if s not in known]
        if orphans:
            print(f"  WARNING: {len(orphans)} slug(s) in skyscrapers.json match no "
                  f"metro: {orphans[:8]}")
        out = {s: dict(v) for s, v in data.items() if s in known}
        print(f"  source: SKYDB via public/data/skyscrapers.json "
              f"(generated {payload.get('generated', '?')})")
        return out

    print("  *** WARNING: public/data/skyscrapers.json IS MISSING ***")
    print("  *** falling back to the hand-curated Skyscrapers sheet.      ***")
    print("  *** Run: python scripts/build-skyscrapers.py                 ***")
    by_name = extract_skyscrapers_sheet(wb)
    name_to_slug = {m['name']: m['slug'] for m in metros}
    return {name_to_slug[n]: v for n, v in by_name.items() if n in name_to_slug}


def extract_skyscrapers_sheet(wb):
    """Extract skyscraper data grouped by metro, from the workbook sheet.

    Kept as the fallback for load_skyscrapers and as the input to the
    divergence guard. Returns a dict keyed by metro NAME.

    The Skyscrapers sheet has one row per municipality, and a single metro can
    span multiple municipalities (NYC + Jersey City + Fort Lee, Tokyo +
    Yokohama + Kawasaki, Bay Area's SF + San Jose + Oakland, etc.). We must
    SUM the 150m+/200m+/300m+ tier counts across all rows for the metro, not
    overwrite. The 'city' field stores the largest contributing municipality
    so the value remains meaningful if surfaced later.
    """
    ws = wb["Skyscrapers"]
    scrapers = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        v = list(row)
        metro = safe_str(v[4])
        if not metro:
            continue
        city = safe_str(v[2])
        over150 = safe_int(v[9])
        over200 = safe_int(v[10])
        over300 = safe_int(v[11])
        agg = scrapers.get(metro)
        if agg is None:
            scrapers[metro] = {
                'city': city,
                'over150m': over150,
                'over200m': over200,
                'over300m': over300,
                # Bookkeeping: track the largest contributor's 150m+ count so
                # the headline city stays the densest municipality even when
                # a smaller one is processed later in the sheet.
                '_top_city_over150m': over150,
            }
            continue
        agg['over150m'] += over150
        agg['over200m'] += over200
        agg['over300m'] += over300
        # Promote the largest contributor as the headline city. Ties broken
        # by row order (first occurrence wins).
        if over150 > agg.get('_top_city_over150m', 0):
            agg['city'] = city
            agg['_top_city_over150m'] = over150
    # Strip the bookkeeping field so the output JSON stays clean.
    for v in scrapers.values():
        v.pop('_top_city_over150m', None)
    return scrapers


def extract_luxury(wb):
    """Extract luxury hospitality grouped by metro."""
    ws = wb["Luxury Hospitality"]
    luxury = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        v = list(row)
        metro = safe_str(v[6])
        if not metro:
            continue
        entry = {
            'name': safe_str(v[4]),
            'city': safe_str(v[5]),
            'type': safe_str(v[8]),
        }
        # Columns P/Q (15/16) = Lat/Long. Empty until 2026-08-09, when they were
        # filled from FSQ OS Places. Emitted only when BOTH are present, so a
        # half-resolved row never renders as a pin at (0, 0) off West Africa.
        # Provenance for every value lives in Supabase public.place_ids.
        lat = safe_float(v[15]) if len(v) > 15 else 0
        lng = safe_float(v[16]) if len(v) > 16 else 0
        if lat and lng:
            entry['lat'] = lat
            entry['lng'] = lng
        luxury.setdefault(metro, []).append(entry)
    return luxury


def extract_events(wb):
    """Extract sporting events (Golf, Tennis, F1, Boxing) grouped by metro."""
    ws = wb["Golf-Tennis-F1"]
    events = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        v = list(row)
        metro = safe_str(v[6])
        if not metro:
            continue
        entry = {
            'sport': safe_str(v[0]),
            'event': safe_str(v[1]),
            'year': safe_str(v[2]),
            'venue': safe_str(v[3]),
        }
        # Boxing Event type (Column K / index 10): only applies to Boxing rows
        event_type = safe_str(v[10]) if len(v) > 10 else ''
        if event_type:
            entry['type'] = event_type
        events.setdefault(metro, []).append(entry)
    return events


def extract_mktcap(wb):
    """Extract market cap data grouped by metro, including company name and source.

    Primary source since the 2026-08-29 CMC-workbook sunset: the pipeline's
    committed CSV (scripts/mktcap/out/mktcap_export.csv, written by the mini's
    Saturday refresh.py run). It is the SAME feed the workbook's MktCap_Data
    Power Query table reads, so going straight to it means extract no longer
    depends on when the workbook was last opened/refreshed — a workbook-sync
    run can never publish a stale Top Companies section. Proven byte-identical
    to the sheet-cache path across all 555 mapped metros on 2026-08-29.
    Fallback: the MktCap_Data sheet's cached values (pre-cutover behaviour),
    kept only for a checkout that somehow lacks the CSV.

    scripts/mktcap/update_top_companies.py mirrors this logic for the
    workbook-free Saturday path — change the shape in one place, change it in
    the other."""
    csv_path = (Path(__file__).resolve().parent / "mktcap" / "out"
                / "mktcap_export.csv")
    if csv_path.exists():
        import csv as _csv
        mktcap = {}
        with open(csv_path, newline='', encoding='utf-8') as f:
            for r in _csv.DictReader(f):
                metro = (r.get("Metro Area") or '').strip()
                try:
                    val = float(r.get("Valuation") or 0)
                except ValueError:
                    val = 0.0
                if not metro or val == 0:
                    continue
                mktcap.setdefault(metro, []).append({
                    'valuation': val,
                    'name': (r.get("Company Name") or '').strip(),
                    'source': (r.get("Source") or '').strip(),
                })
        for metro in mktcap:
            mktcap[metro].sort(key=lambda x: x['valuation'], reverse=True)
        print("  (source: scripts/mktcap/out/mktcap_export.csv)")
        return mktcap

    print("  WARNING: mktcap_export.csv not found; falling back to the "
          "workbook's cached MktCap_Data sheet (may be stale)")
    ws = wb["MktCap_Data"]
    mktcap = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        v = list(row)
        metro = safe_str(v[0])
        val = safe_float(v[1])
        if not metro or val == 0:
            continue
        company_name = safe_str(v[2]) if len(v) > 2 else ''
        source = safe_str(v[3]) if len(v) > 3 else ''
        mktcap.setdefault(metro, []).append({
            'valuation': val,
            'name': company_name,
            'source': source,
        })
    # Sort each metro's companies by valuation descending
    for metro in mktcap:
        mktcap[metro].sort(key=lambda x: x['valuation'], reverse=True)
    return mktcap


def _slugify_state(name):
    """Slug-ify a state/admin division name for use in URLs."""
    if not name:
        return ""
    s = safe_str(name).lower()
    # Strip diacritics
    import unicodedata
    s = unicodedata.normalize('NFKD', s)
    s = ''.join(c for c in s if not unicodedata.combining(c))
    # Replace common punctuation/whitespace
    out = []
    for ch in s:
        if ch.isalnum():
            out.append(ch)
        elif ch in (' ', '-', '_', '/', '.', "'", '(', ')'):
            out.append('-')
    s = ''.join(out)
    # Collapse repeated dashes and trim
    while '--' in s:
        s = s.replace('--', '-')
    return s.strip('-')


# Per-country sheet routing is data-driven inside extract_metro_state_edges
# rather than driven by a static list. Notes lists 33 countries as
# "Municipality countries", but several (Brazil, Japan, Russia, India,
# Australia) actually have zero rows there in practice. Discovering which
# sovereigns have Municipality rows at runtime keeps coverage correct
# regardless of which long-tail countries get added or moved between sheets.

# Map UK constituent labels back to the sovereign so per-state lookups
# resolve consistently against the States sheet (which stores Greater
# London under Country='England', Edinburgh under Country='Scotland', etc.).
WORKBOOK_TO_CANONICAL_COUNTRY = {
    "England":          "United Kingdom",
    "Scotland":         "United Kingdom",
    "Wales":            "United Kingdom",
    "Northern Ireland": "United Kingdom",
}


def extract_metro_state_edges(wb, all_metros):
    """Aggregate (state, metro) edges from Counties + Municipality.

    Routing is data-driven, not based on a static list. We scan Municipality
    first to discover which countries actually have rows there. Those
    countries are sourced exclusively from Municipality (it's denser when
    populated). Every other country falls through to Counties — including
    countries Notes lists as "Municipality" but which have zero rows there
    in practice (Brazil, Japan, Russia, India, Australia all qualify as of
    Session 73).

    Sheet column layout (verified directly against the workbook, ignoring
    the Notes' off-by-one column-letter shorthand for Counties):
      Counties:     col 0 = Country, col 2 = State, col 7 = Metro, col 8 = Sub Country
      Municipality: col 1 = Country, col 4 = State, col 6 = Metro

    UK rows in Municipality use the constituent ("England", "Scotland",
    etc.) in col 1, never the sovereign. Those canonicalize to United
    Kingdom in the edge key while the constituent is preserved for the
    States-sheet matcher's fallback path.

    Returns:
      edges: dict[(country, sub_country, state_name)] -> set(metro_name)
    """
    edges = {}
    metros_by_name = {m['name']: m for m in all_metros}

    def add_edge(country, sub, state, metro_name):
        if not country or not state or not metro_name:
            return
        if metro_name not in metros_by_name:
            return
        canonical = WORKBOOK_TO_CANONICAL_COUNTRY.get(country, country)
        sub_label = (
            country if country in WORKBOOK_TO_CANONICAL_COUNTRY else (sub or "")
        )
        key = (canonical, sub_label, state)
        edges.setdefault(key, set()).add(metro_name)

    # Pass 1: scan Municipality. Track which sovereign countries have
    # actual rows so the Counties pass below knows which to skip.
    municipality_sovereigns = set()
    if "Municipality" in wb.sheetnames:
        ws = wb["Municipality"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            v = list(row)
            country = safe_str(v[1]) if len(v) > 1 else ''
            if not country:
                continue
            canonical = WORKBOOK_TO_CANONICAL_COUNTRY.get(country, country)
            municipality_sovereigns.add(canonical)
            state = safe_str(v[4]) if len(v) > 4 else ''
            metro = safe_str(v[6]) if len(v) > 6 else ''
            add_edge(country, '', state, metro)

    # Pass 2: scan Counties for countries NOT served by Municipality.
    # Falls back the long tail (Brazil, Japan, India, Russia, Australia,
    # plus the 200+ smaller countries that never had Municipality rows).
    if "Counties" in wb.sheetnames:
        ws = wb["Counties"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            v = list(row)
            country = safe_str(v[0]) if len(v) > 0 else ''
            if not country:
                continue
            canonical = WORKBOOK_TO_CANONICAL_COUNTRY.get(country, country)
            if canonical in municipality_sovereigns:
                continue
            state = safe_str(v[2]) if len(v) > 2 else ''
            metro = safe_str(v[7]) if len(v) > 7 else ''
            sub = safe_str(v[8]) if len(v) > 8 else ''
            add_edge(country, sub, state, metro)

    return edges


# Editorial overrides: metros whose real footprint spans more states/
# administrative areas than the workbook's primary/state2/state3 columns
# can carry. Each metro lists ADDITIONAL state names (in addition to its
# primaryState/state2/state3) to render on the metro page tables and to
# include in /states/[slug] pages for those states.
#
# Names must match the Administrative Division (col 2) in the States sheet
# under the metro's country (or subCountry for UK constituents). Add new
# entries here as editorial calls land.
METRO_ADDITIONAL_STATES = {
    # Greater London Built-Up Area extends well beyond the boundary of the
    # GLA into the home counties; commuter belt covers Surrey,
    # Hertfordshire, and Berkshire.
    "London": ["Surrey", "Hertfordshire", "Berkshire"],
}


def extract_states(wb, all_metros):
    """Extract states/provinces from the States (ISO 3166-2) sheet.

    Returns:
      states_list: list of state dicts ready for states.json
      metro_state_slugs: dict {metro_name: {primary?, state2?, state3?}}
        Each value is a dict with up to three keys mapping the metro's
        primaryState / state2 / state3 names to their resolved state slugs.
        Missing keys mean the lookup found no match (state isn't in the
        ISO sheet under either country or subCountry).

    Keying: (Country, Administrative Division) is the canonical pair. Cross-
    reference each metro's (country, primaryState) against this pair, falling
    back to (subCountry, primaryState) for UK constituents whose metros tag
    country='United Kingdom' but the States sheet stores them under
    Country='England' / 'Scotland' / 'Wales' / 'Northern Ireland'.

    Slug rule: kebab-cased Administrative Division. On collision (Punjab in
    India + Pakistan, Amazonas in Brazil/Venezuela/Colombia), append the
    country slug to disambiguate. Final fallback prepends ISO if available.
    """
    if "States (ISO 3166-2)" not in wb.sheetnames:
        return [], {}
    ws = wb["States (ISO 3166-2)"]

    # Pre-compute country slug map from metros so state slug collisions can
    # disambiguate using the same conventions as country pages. The canonical
    # slugify() helper strips ampersands, apostrophes, periods, commas, and
    # diacritics; using the naive lower+replace here previously produced
    # state countrySlug values like "antigua-&-barbuda" that did not match
    # any entry in countries.json and 404d on the country link.
    country_slug = {}
    for m in all_metros:
        country_slug[m['country']] = slugify(m['country'])
        if m.get('subCountry'):
            country_slug[m['subCountry']] = slugify(m['subCountry'])

    # Pass 1: collect raw state rows.
    raw = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        v = list(row)
        admin = safe_str(v[2]) if len(v) > 2 else ''
        country = safe_str(v[4]) if len(v) > 4 else ''
        if not admin or not country:
            continue
        raw.append({
            'name': admin,
            'type': safe_str(v[3]) if len(v) > 3 else '',
            'country': country,  # the immediate parent in the sheet (England, Anguilla, etc.)
            'pop': safe_int(v[5]) if len(v) > 5 else 0,
            'capital': safe_str(v[6]) if len(v) > 6 else '',
            'continent': safe_str(v[8]) if len(v) > 8 else '',
            'iso': safe_str(v[9]) if len(v) > 9 else '',
            'subRegion': safe_str(v[10]) if len(v) > 10 else '',
            'mainCountry': safe_str(v[11]) if len(v) > 11 else '',
            'languageAdmin': safe_str(v[12]) if len(v) > 12 else '',
            'languageSecondary': safe_str(v[13]) if len(v) > 13 else '',
            'languageDeFacto': safe_str(v[16]) if len(v) > 16 else '',
            'languageNational': safe_str(v[17]) if len(v) > 17 else '',
        })

    # Pass 2: assign collision-aware slugs.
    # Group by raw slug to detect collisions across countries.
    by_raw_slug = {}
    for s in raw:
        rs = _slugify_state(s['name'])
        if not rs:
            continue
        s['_raw_slug'] = rs
        by_raw_slug.setdefault(rs, []).append(s)

    states_by_key = {}  # (country, name) -> state dict with final slug
    for rs, group in by_raw_slug.items():
        if len(group) == 1:
            s = group[0]
            s['slug'] = rs
        else:
            # Disambiguate by appending the immediate parent country slug.
            for s in group:
                country_part = country_slug.get(s['country']) or _slugify_state(s['country'])
                # If state name == country name (e.g. Andorra/Andorra),
                # don't double the slug — fall back to ISO if present.
                if rs == country_part and s['iso']:
                    s['slug'] = s['iso'].lower()
                else:
                    s['slug'] = f"{rs}-{country_part}"
        for s in group:
            states_by_key[(s['country'], s['name'])] = s

    # Sanity: enforce slug uniqueness across the whole set; fall back to ISO
    # on any residual collision so the slug truly identifies one row.
    seen_slugs = {}
    for s in states_by_key.values():
        if s['slug'] in seen_slugs and seen_slugs[s['slug']] is not s:
            if s['iso']:
                s['slug'] = s['iso'].lower()
            else:
                s['slug'] = f"{s['slug']}-{_slugify_state(s['mainCountry'])}"
        seen_slugs[s['slug']] = s

    # Pass 3: resolve every metro's primary, #2 and #3 state slugs. The
    # homepage rankings table and country-page tables both read these three
    # slugs off metros.json so neither has to redo the lookup at render
    # time. Per-state metro counts (and the state page's metro list) come
    # from the cross-sheet aggregator below — which sees the FULL list of
    # states a metro spans rather than just the workbook's three slots.
    metro_state_slugs = {}  # metro_name -> {primary, state2, state3, additional}

    def _resolve(country, sub, name):
        if not name:
            return None
        return (
            states_by_key.get((country, name))
            or (states_by_key.get((sub, name)) if sub else None)
        )

    # Resolve the up-to-three workbook slots for the homepage / country
    # tables (these are space-constrained and only show 3 states inline).
    for m in all_metros:
        country = m['country']
        sub = m.get('subCountry') or ''
        primary = m.get('primaryState') or ''
        s2 = m.get('state2') or ''
        s3 = m.get('state3') or ''
        primary_state = _resolve(country, sub, primary)
        s2_state = _resolve(country, sub, s2)
        s3_state = _resolve(country, sub, s3)
        slugs = {}
        if primary_state:
            slugs['primary'] = primary_state['slug']
        if s2_state:
            slugs['state2'] = s2_state['slug']
        if s3_state:
            slugs['state3'] = s3_state['slug']
        extras = METRO_ADDITIONAL_STATES.get(m['name'], [])
        if extras:
            additional = []
            for name in extras:
                st = _resolve(country, sub, name)
                if st:
                    additional.append({'name': name, 'slug': st['slug']})
                else:
                    print(f"  [warn] additional-state override for {m['name']!r} did not match {(country, name)!r}")
                    additional.append({'name': name})
            if additional:
                slugs['additional'] = additional
        if slugs:
            metro_state_slugs[m['name']] = slugs

    # Pass 4: cross-sheet aggregation. Counties + Municipality together
    # capture every state a metro touches (a metro that spans 7 English
    # ceremonial counties has 7 edges in Municipality; primaryState only
    # records 1). For each (country, state) edge, accumulate the metro into
    # the matching State row's _metroSlugs set and recompute totals from
    # that set. This becomes the source of truth for the country-page chip
    # counts and the /states/[slug] metro list.
    edges = extract_metro_state_edges(wb, all_metros)
    metros_by_name = {m['name']: m for m in all_metros}
    metros_by_slug = {m['slug']: m for m in all_metros}
    state_by_slug = {s['slug']: s for s in states_by_key.values()}

    for (country, sub_label, state_name), metro_set in edges.items():
        st = _resolve(country, sub_label, state_name)
        if not st:
            # No matching row in States sheet. Common in countries whose
            # Counties data tags state names that aren't in the ISO sheet
            # under the same Country column (cross-listed via Sub-Country
            # is already tried by _resolve). Skip silently — the metro
            # still shows on its primary state page if that resolved.
            continue
        bucket = st.setdefault('_metroSlugs', set())
        for metro_name in metro_set:
            m = metros_by_name.get(metro_name)
            if m:
                bucket.add(m['slug'])

    # Always also include the metro's primary state slug — the workbook's
    # primaryState is the intentional editorial primary, and a metro
    # sometimes has zero rows in Counties/Municipality for its primary
    # state (e.g., when the "metro" is actually a single municipality and
    # the parent state isn't otherwise tagged). Belt-and-braces.
    for m in all_metros:
        slugs = metro_state_slugs.get(m['name'], {})
        primary = slugs.get('primary')
        if not primary:
            continue
        primary_state = state_by_slug.get(primary)
        if primary_state:
            primary_state.setdefault('_metroSlugs', set()).add(m['slug'])

    # Roll the metro slug set into final counts and a sorted list. Pop
    # and score totals follow the same set so they don't double-count
    # metros that span multiple workbook columns. O(N) overall thanks
    # to the metros_by_slug index built above.
    for st in states_by_key.values():
        slug_set = st.get('_metroSlugs') or set()
        st['_metroCount'] = len(slug_set)
        st['_metroPop'] = 0
        st['_scoreTotal'] = 0.0
        for sl in slug_set:
            mm = metros_by_slug.get(sl)
            if mm:
                st['_metroPop'] += mm.get('pop') or 0
                st['_scoreTotal'] += mm.get('score') or 0.0
        # Stable order: by metro rank ascending so the highest-ranked
        # metro in each state lands first in the state page table.
        st['_metroSlugList'] = sorted(
            slug_set,
            key=lambda sl: metros_by_slug.get(sl, {}).get('rank', 99999),
        )

    # Final shape for states.json — keep concise; promote internal keys.
    states_list = []
    for s in states_by_key.values():
        country_slug_val = country_slug.get(s['country']) or _slugify_state(s['country'])
        main_country_slug_val = (
            country_slug.get(s['mainCountry']) or _slugify_state(s['mainCountry'])
            if s['mainCountry'] else None
        )
        out = {
            'slug': s['slug'],
            'name': s['name'],
            'country': s['country'],
            'countrySlug': country_slug_val,
            'mainCountry': s['mainCountry'] or s['country'],
            'mainCountrySlug': main_country_slug_val or country_slug_val,
            'type': s['type'] or 'Administrative Area',
            'iso': s['iso'] or None,
            'pop': s['pop'] or None,
            'capital': s['capital'] or None,
            'continent': s['continent'] or None,
            'subRegion': s['subRegion'] or None,
            'languageAdmin': s['languageAdmin'] or None,
            'languageSecondary': s['languageSecondary'] or None,
            'languageDeFacto': s['languageDeFacto'] or None,
            'languageNational': s['languageNational'] or None,
            'metroCount': s.get('_metroCount', 0),
            'metroPop': s.get('_metroPop', 0),
            'scoreTotal': round(s.get('_scoreTotal', 0.0), 2),
            # The full list of metro slugs that touch this state (from
            # Counties/Municipality cross-sheet aggregation, plus each
            # metro's primary state). Drives the /states/[slug] metro
            # table; supersedes the older "filter metros by stateSlug"
            # client-side scan, which only saw the first-3 workbook slots.
            'metroSlugs': s.get('_metroSlugList', []),
        }
        states_list.append(out)

    # Stable sort: by main country, then descending metro count, then name.
    states_list.sort(key=lambda x: (
        x['mainCountrySlug'] or '',
        -(x['metroCount'] or 0),
        x['name'],
    ))
    return states_list, metro_state_slugs


def extract_football(wb):
    """Extract football club data grouped by metro.

    FootballClub_Data column layout (as of 2026-05-05):
      0: Team, 1: City, 2: Metro Area, 3: County, 4: Country, 5: League,
      6: Level, 7: Club, 8: Major League, 9: Latitude, 10: Longitude.

    Country aggregate rows (col 7 = "Country") are pure roll-ups and have
    no metro home; skip them. Cols 9/10 previously scaffolded for QID and
    Wikipedia URL; the user repurposed them for venue coordinates on
    2026-05-05. If a future pass adds club-level Wikidata linking, give
    those columns a new home rather than reusing 9/10.
    """
    ws = wb["FootballClub_Data"]
    football = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        v = list(row)
        metro = safe_str(v[2])  # Metro Area column
        team_name = safe_str(v[0])  # Team name column
        if not metro or not team_name:
            continue
        # Skip Country-level aggregate rows (~250 entries); they exist to
        # carry a country-wide pin and are not stadiums.
        if safe_str(v[7]) == 'Country':
            continue
        entry = {
            'team': team_name,
            'city': safe_str(v[1]),
            'country': safe_str(v[4]),
            'league': safe_str(v[5]),
            'level': safe_int(v[6]),
            'major': safe_str(v[8]) == 'Y',
        }
        # Cols 9/10 = Latitude / Longitude. Only attach when both are
        # present and numeric so the marker layer can filter cleanly.
        lat = safe_float(v[9]) if len(v) > 9 else 0
        lng = safe_float(v[10]) if len(v) > 10 else 0
        if lat or lng:
            entry['lat'] = lat
            entry['lng'] = lng
        football.setdefault(metro, []).append(entry)
    return football


def extract_towers(wb):
    """Extract supertall structures (350m+) grouped by metro.

    Tower_Data layout (Row 1 empty, Row 2 headers, Row 3+ data):
      0: Rank, 1: Name, 2: Height (m), 3: Height (ft),
      4: City, 5: Country, 6: Year Built, 7: Notes, 8: Metro Area
    The xlsx is already pre-filtered to 350m+, but we still gate on
    height as a defensive measure in case the threshold changes upstream.
    """
    if "Tower_Data" not in wb.sheetnames:
        return {}
    ws = wb["Tower_Data"]
    towers = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        v = list(row)
        if len(v) < 9:
            continue
        metro = safe_str(v[8])
        name = safe_str(v[1])
        height_m = safe_float(v[2])
        if not metro or not name or height_m < 350:
            continue
        towers.setdefault(metro, []).append({
            'name': name,
            'city': safe_str(v[4]),
            'heightM': round(height_m, 1),
            'yearBuilt': safe_int(v[6]) or None,
        })
    # Sort each metro's towers tallest-first so the UI gets a stable order.
    for metro in towers:
        towers[metro].sort(key=lambda x: -x['heightM'])
    return towers


def compute_regions(metros):
    """Compute regional aggregates."""
    regions = {}
    for m in metros:
        r = m['region']
        if not r:
            continue
        if r not in regions:
            regions[r] = {
                'name': r,
                'metros': 0,
                'above50': 0,
                'above20': 0,
                'totalScore': 0,
                'top3': [],
                'totalPop': 0,
                'totalMarketCap': 0,
                'scores': [],
            }
        reg = regions[r]
        reg['metros'] += 1
        reg['totalScore'] += m['score']
        reg['totalPop'] += m['pop']
        reg['totalMarketCap'] += m['dims']['marketCap']
        reg['scores'].append(m['score'])
        if m['score'] >= 50:
            reg['above50'] += 1
        if m['score'] >= 20:
            reg['above20'] += 1
        if len(reg['top3']) < 3:
            reg['top3'].append({
                'name': m['name'],
                'score': m['score'],
                'rank': m['rank'],
                'slug': m['slug'],
            })

    # Compute medians
    for r in regions.values():
        scores = sorted(r['scores'])
        n = len(scores)
        r['medianScore'] = round(scores[n // 2], 2) if n else 0
        r['totalMarketCap'] = round(r['totalMarketCap'])
        del r['scores']

    return list(regions.values())


def compute_dimension_ranks(metros):
    """Compute per-dimension ranks across all metros. Returns dict keyed by slug."""
    dim_keys = list(metros[0]['dims'].keys()) if metros else []
    ranks_by_slug = {}

    for key in dim_keys:
        # Collect (value, slug) pairs, sorted descending by value
        entries = [(m['dims'][key], m['slug']) for m in metros]
        entries.sort(key=lambda x: -x[0])

        # Assign ranks with tie handling
        rank_map = {}
        i = 0
        while i < len(entries):
            val = entries[i][0]
            # Find all entries with this same value
            j = i
            while j < len(entries) and entries[j][0] == val:
                j += 1
            tied_count = j - i
            rank_pos = i + 1  # 1-based rank
            is_tie = tied_count > 1 and val > 0
            for k in range(i, j):
                slug = entries[k][1]
                if val <= 0:
                    rank_map[slug] = None  # No rank for zero/negative values
                elif is_tie:
                    rank_map[slug] = f"T-{rank_pos}"
                else:
                    rank_map[slug] = str(rank_pos)
            i = j

        for slug, rank_str in rank_map.items():
            if slug not in ranks_by_slug:
                ranks_by_slug[slug] = {}
            ranks_by_slug[slug][key] = rank_str

    return ranks_by_slug


def fetch_mktcap_snapshot_date():
    """The 'Source data as of' date for the Top Companies tables, from the
    pipeline itself: public.mktcap_valuations' latest weekly snapshot (the
    Mac mini's Saturday run of scripts/mktcap/refresh.py --write).

    Since the 2026-08-29 cutover (HANDOFF: CMC-workbook sunset) the
    CompaniesMarketCap xlsx is retired -- its mtime, which used to feed this
    label, would pin it to 2026-08-23 forever. Supabase is the source of
    truth; MetroAreas.xlsx now pulls the pipeline's committed CSV via Power
    Query (refresh-on-open), so this date and the extracted numbers move
    together on any normal Saturday flow.

    Anon read, ~1KB response. Returns 'YYYY-MM-DD' or None on any failure so
    an offline ETL run still completes (caller falls back to the old path).
    """
    import urllib.request
    url = (os.environ.get("SUPABASE_URL")
           or "https://nmprqkmymrdknffwnuur.supabase.co").rstrip("/")
    anon = (os.environ.get("SUPABASE_ANON_KEY") or
            "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Im5tcHJxa215bXJka25mZndudXVyIiwicm9sZSI6ImFub24iLCJpYXQiOjE3ODMyMDkzNDMsImV4cCI6MjA5ODc4NTM0M30."
            "4RXU3mQ-Yl81ZqC2_a10aizKGu_87B4vt8OK5Pi_-sM")
    try:
        req = urllib.request.Request(
            url + "/rest/v1/mktcap_valuations?select=as_of&order=as_of.desc&limit=1",
            headers={"apikey": anon, "Authorization": "Bearer " + anon})
        with urllib.request.urlopen(req, timeout=15) as r:
            rows = json.load(r)
        as_of = (rows or [{}])[0].get("as_of")
        return as_of if as_of and re.fullmatch(r"\d{4}-\d{2}-\d{2}", str(as_of)) else None
    except Exception:
        return None


def find_companies_source(explicit_path=None):
    """Locate the upstream companiesmarketcap.com xlsx whose mtime tells us
    when the market cap data was last refreshed. Returns Path or None.

    Lookup order:
      1. COMPANIES_SOURCE_XLSX environment variable, if set
      2. explicit_path argument, if provided
      3. ~/OneDrive/Excel Files/companiesmarketcap.com - Companies ranked by Market Cap - CompaniesMarketCap.com (1).xlsx
      4. ~/Excel Files/...
      5. Sibling 'Excel Files' folder next to project root
    """
    filename = (
        "companiesmarketcap.com - Companies ranked by Market Cap - "
        "CompaniesMarketCap.com (1).xlsx"
    )
    candidates = []
    env_path = os.environ.get("COMPANIES_SOURCE_XLSX")
    if env_path:
        candidates.append(Path(env_path))
    if explicit_path:
        candidates.append(Path(explicit_path))
    home = Path.home()
    candidates.extend([
        home / "OneDrive" / "Excel Files" / filename,
        home / "Excel Files" / filename,
        Path(__file__).resolve().parent.parent.parent / "Excel Files" / filename,
    ])
    for c in candidates:
        try:
            if c.exists():
                return c
        except OSError:
            continue
    return None


def build_detail(metro_name, teams, unis, culture, scrapers, luxury, events, mktcap, football, towers, mktcap_as_of=None, metro_slug=None):
    """Build a detail JSON object for a single metro.

    Note the mixed keys. Everything the workbook produces is keyed by metro
    NAME, because that is what the sheets carry. `scrapers` is the exception:
    it comes from SKYDB via point-in-polygon, so it is keyed by SLUG, which is
    the actual identifier. The two are 1:1 today (4,314 metros, 4,314 distinct
    names), but slug is the one that survives a rename.
    """
    detail = {}

    if metro_name in teams:
        detail['teams'] = teams[metro_name]

    if metro_name in unis:
        detail['universities'] = unis[metro_name]

    if metro_name in culture:
        # Group by type
        by_type = {}
        for item in culture[metro_name]:
            t = item['type'] or 'Other'
            by_type.setdefault(t, []).append(item)
        detail['culture'] = by_type

    if metro_slug and metro_slug in scrapers:
        detail['skyscrapers'] = scrapers[metro_slug]

    if metro_name in luxury:
        detail['luxury'] = luxury[metro_name]

    if metro_name in events:
        detail['events'] = events[metro_name]

    if metro_name in mktcap:
        companies = mktcap[metro_name]
        total = sum(c['valuation'] for c in companies)
        detail['marketCap'] = {
            'total': round(total),
            'count': len(companies),
            'top12': [
                {
                    'name': c['name'],
                    'valuation': round(c['valuation']),
                    'source': c['source'],
                }
                for c in companies[:12]
            ],
        }
        if mktcap_as_of:
            detail['marketCap']['asOf'] = mktcap_as_of

    if metro_name in towers:
        detail['supertallStructures'] = towers[metro_name]

    if metro_name in football:
        clubs = football[metro_name]
        detail['football'] = {
            'total': len(clubs),
            'byLevel': {},
        }
        for c in clubs:
            lvl = str(c['level'])
            detail['football']['byLevel'][lvl] = detail['football']['byLevel'].get(lvl, 0) + 1

        # Merge football clubs into the teams array for site rendering
        # Major league clubs always included; non-major only if they have a valid level (> 0)
        if 'teams' not in detail:
            detail['teams'] = []
        for c in clubs:
            if not c['major'] and not c['level']:
                continue
            merged = {
                'sport': 'Soccer',
                'league': c['league'],
                'team': c['team'],
                'city': c['city'],
                'country': c['country'],
                'level': str(c['level']),
                'major': c['major'],
            }
            if c.get('qid'):
                merged['qid'] = c['qid']
            if c.get('wikipediaUrl'):
                merged['wikipediaUrl'] = c['wikipediaUrl']
            if c.get('lat') is not None and c.get('lng') is not None:
                merged['lat'] = c['lat']
                merged['lng'] = c['lng']
            detail['teams'].append(merged)

    return detail


def main():
    # Find the Excel file
    script_dir = Path(__file__).parent
    site_dir = script_dir.parent

    if len(sys.argv) > 1:
        xlsx_path = Path(sys.argv[1])
    else:
        # Prefer the project-root copy maintained by sync_source_xlsx.py;
        # fall back to a sibling 'MetroAreas.xlsx' next to the project root
        # for legacy layouts.
        primary = site_dir / "MetroAreas.xlsx"
        legacy = site_dir.parent / "MetroAreas.xlsx"
        xlsx_path = primary if primary.exists() else legacy

    if not xlsx_path.exists():
        print(f"ERROR: Cannot find {xlsx_path}")
        print(f"Usage: python {sys.argv[0]} [path/to/MetroAreas.xlsx]")
        sys.exit(1)

    print(f"Reading {xlsx_path}...")
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)

    # Extract all data
    print("Extracting metro data...")
    score_index = build_score_index(wb, xlsx_path)
    metros = extract_metros(wb, score_index)
    print(f"  {len(metros)} metros")

    print("Extracting teams...")
    teams = extract_teams(wb)
    print(f"  {sum(len(v) for v in teams.values())} teams across {len(teams)} metros")

    print("Extracting universities...")
    unis = extract_universities(wb)
    print(f"  {sum(len(v) for v in unis.values())} universities")

    print("Extracting cultural assets...")
    culture_data = extract_culture(wb)
    print(f"  {sum(len(v) for v in culture_data.values())} assets")

    print("Extracting skyscrapers...")
    scrapers = load_skyscrapers(wb, metros)
    print(f"  {len(scrapers)} metros with skyscrapers")

    print("Extracting luxury hospitality...")
    luxury = extract_luxury(wb)
    print(f"  {sum(len(v) for v in luxury.values())} entries")

    print("Extracting sporting events...")
    events = extract_events(wb)
    print(f"  {sum(len(v) for v in events.values())} events")

    print("Extracting market cap data...")
    mktcap = extract_mktcap(wb)
    print(f"  {sum(len(v) for v in mktcap.values())} companies")

    # "Source data as of {date}" for the Top Companies tables. Primary:
    # the pipeline's latest weekly snapshot in Supabase (the mini's Saturday
    # run) -- see fetch_mktcap_snapshot_date(). Fallback: the retired CMC
    # xlsx's mtime (pre-cutover behaviour), kept only so an offline ETL run
    # still stamps something rather than nothing.
    import datetime as _dt
    mktcap_as_of = fetch_mktcap_snapshot_date()
    if mktcap_as_of:
        print(f"  companies as-of: {mktcap_as_of} (mktcap_valuations latest snapshot)")
    else:
        companies_src = find_companies_source()
        if companies_src is not None:
            mktcap_as_of = _dt.datetime.fromtimestamp(
                os.path.getmtime(str(companies_src))
            ).strftime('%Y-%m-%d')
            print(f"  companies as-of FALLBACK (Supabase unreachable): "
                  f"{companies_src.name} mtime {mktcap_as_of}")
        else:
            mktcap_as_of = None
            print("  companies as-of: Supabase unreachable and no source xlsx; skipping stamp")

    print("Extracting football clubs...")
    football = extract_football(wb)
    print(f"  {sum(len(v) for v in football.values())} clubs")

    print("Extracting supertall structures...")
    towers = extract_towers(wb)
    print(f"  {sum(len(v) for v in towers.values())} supertalls across {len(towers)} metros")

    print("Extracting states/provinces...")
    states_list, metro_state_slugs = extract_states(wb, metros)
    print(f"  {len(states_list)} states; {len(metro_state_slugs)} metros resolved to state slug")

    print("Extracting Gold Standard leagues (workbook col L of Team List)...")
    gold_leagues = extract_gold_standard_leagues(wb)
    total_gold = sum(len(v) for v in gold_leagues.values())
    print(f"  {len(gold_leagues)} sports, {total_gold} leagues flagged Gold")

    wb.close()

    # Build a country-name -> slug lookup from public/data/countries.json so
    # every metro entry can carry a ready-to-use countrySlug field. Client
    # components on the homepage and elsewhere link metros to /countries/...
    # without a runtime lookup. Also captures subCountry slug for UK metros
    # so the constituent (England / Scotland / Wales / Northern Ireland)
    # link reads naturally rather than always pointing at /countries/united-kingdom.
    country_slug_by_name = {}
    countries_json_path = site_dir / "public" / "data" / "countries.json"
    if countries_json_path.exists():
        try:
            with open(countries_json_path, encoding='utf-8') as f:
                for c in json.load(f):
                    if c.get('name') and c.get('slug'):
                        country_slug_by_name[c['name']] = c['slug']
            print(f"  loaded {len(country_slug_by_name)} country slugs from countries.json")
        except Exception as e:
            print(f"  WARN: could not parse countries.json ({e}); countrySlug will be omitted")
    else:
        print("  WARN: countries.json not found; countrySlug will be omitted from metros.json")

    # Enrich every metro dict with resolved slug fields BEFORE both the
    # slim metros.json emission and the per-metro detail emission consume
    # them. This way detail.metro (used by app/rankings/[slug]/page.tsx)
    # and slim_metros (used by the homepage RankingsTable) both link to
    # /countries/[slug] and /states/[slug] without any runtime lookup.
    for m in metros:
        primary_cs = country_slug_by_name.get(m.get('country') or '')
        sub_cs = (
            country_slug_by_name.get(m.get('subCountry') or '')
            if m.get('subCountry')
            else None
        )
        if sub_cs:
            m['countrySlug'] = sub_cs
            if primary_cs and primary_cs != sub_cs:
                m['sovereignSlug'] = primary_cs
        elif primary_cs:
            m['countrySlug'] = primary_cs
        slugs = metro_state_slugs.get(m['name']) or {}
        if slugs.get('primary'):
            m['stateSlug'] = slugs['primary']
        if slugs.get('state2'):
            m['state2Slug'] = slugs['state2']
        if slugs.get('state3'):
            m['state3Slug'] = slugs['state3']
        if slugs.get('additional'):
            m['additionalStates'] = slugs['additional']

    # Compute regions
    print("Computing regional aggregates...")
    regions = compute_regions(metros)

    # Compute dimension ranks
    print("Computing dimension ranks...")
    dim_ranks = compute_dimension_ranks(metros)
    print(f"  Ranked {len(dim_ranks)} metros across {len(metros[0]['dims'])} dimensions")

    # Output directories
    data_dir = site_dir / "public" / "data"
    details_dir = data_dir / "details"
    data_dir.mkdir(parents=True, exist_ok=True)
    details_dir.mkdir(parents=True, exist_ok=True)

    # Write main metros.json (slim version for rankings table)
    print("Writing metros.json...")
    slim_metros = []
    for m in metros:
        entry = {
            'rank': m['rank'],
            'slug': m['slug'],
            'name': m['name'],
            'country': m['country'],
            'region': m['region'],
            'continent': m['continent'],
            'pop': m['pop'],
            'score': m['score'],
            'lat': m['lat'],
            'lon': m['lon'],
            'primaryCity': m['primaryCity'],
            'gdp': m['gdp'],
            # Key dimension summaries for the table
            'majorTeams': m['dims']['majorLeagueTeams'],
            'companies': m['dims']['companies'],
            'marketCap': m['dims']['marketCap'],
            'skyscrapers': m['dims']['skyscrapers'],
            'metroStations': m['dims']['metroStations'],
            'universities': m['dims']['universities'],
        }
        # Country slugs (resolved on the master metros list above).
        if m.get('countrySlug'):
            entry['countrySlug'] = m['countrySlug']
        if m.get('sovereignSlug'):
            entry['sovereignSlug'] = m['sovereignSlug']
        # Include subCountry for UK metros (for search)
        if m['country'] == 'United Kingdom' and m['subCountry']:
            entry['subCountry'] = m['subCountry']
        # Include states for search (only non-empty values)
        if m['primaryState']:
            entry['primaryState'] = m['primaryState']
        if m['state2']:
            entry['state2'] = m['state2']
        if m['state3']:
            entry['state3'] = m['state3']
        # Resolved state slugs (from extract_states matching). Includes
        # the primary state plus state2 / state3 when each name resolves
        # against the States sheet. Both the homepage rankings table and
        # the country-page metro table read these directly so neither
        # client component has to redo the (country, name) lookup.
        # State slugs (resolved on the master metros list above).
        if m.get('stateSlug'):
            entry['stateSlug'] = m['stateSlug']
        if m.get('state2Slug'):
            entry['state2Slug'] = m['state2Slug']
        if m.get('state3Slug'):
            entry['state3Slug'] = m['state3Slug']
        if m.get('additionalStates'):
            entry['additionalStates'] = m['additionalStates']
        slim_metros.append(entry)

    with open(data_dir / "metros.json", 'w', encoding='utf-8') as f:
        json.dump(slim_metros, f, separators=(',', ':'))
    size = os.path.getsize(data_dir / "metros.json")
    print(f"  metros.json: {size:,} bytes ({size/1024:.0f} KB)")

    # Write regions.json
    print("Writing regions.json...")
    with open(data_dir / "regions.json", 'w', encoding='utf-8') as f:
        json.dump(regions, f, separators=(',', ':'))

    # Write states.json (every state row in the ISO sheet, with metro counts).
    print("Writing states.json...")
    with open(data_dir / "states.json", 'w', encoding='utf-8') as f:
        json.dump(states_list, f, separators=(',', ':'))
    states_size = os.path.getsize(data_dir / "states.json")
    print(f"  states.json: {states_size:,} bytes ({states_size/1024:.0f} KB)")

    # Write per-metro detail files
    print("Writing detail files...")
    detail_count = 0
    total_detail_size = 0

    for m in metros:
        # Slugs were made final (and collision-resolved) in extract_metros, so
        # every downstream lookup — dim_ranks included — keys off the same
        # value the URL uses.
        slug = m['slug']

        detail = build_detail(
            m['name'], teams, unis, culture_data, scrapers,
            luxury, events, mktcap, football, towers,
            mktcap_as_of=mktcap_as_of, metro_slug=slug,
        )

        # Add the full metro data to the detail file
        detail['metro'] = m

        # Add dimension ranks
        if slug in dim_ranks:
            detail['dimRanks'] = dim_ranks[slug]

        detail_path = details_dir / f"{slug}.json"
        with open(detail_path, 'w', encoding='utf-8') as f:
            json.dump(detail, f, separators=(',', ':'))

        fsize = os.path.getsize(detail_path)
        total_detail_size += fsize
        detail_count += 1

    print(f"  {detail_count} detail files, {total_detail_size/1024/1024:.1f} MB total")

    # Rewrite metros.json with corrected slugs
    for sm in slim_metros:
        matched = next((m for m in metros if m['name'] == sm['name'] and m['country'] == sm['country']), None)
        if matched:
            sm['slug'] = matched['slug']

    with open(data_dir / "metros.json", 'w', encoding='utf-8') as f:
        json.dump(slim_metros, f, separators=(',', ':'))

    # Write meta.json with last update date from the Excel file
    import datetime
    xlsx_mtime = os.path.getmtime(str(xlsx_path))
    last_update = datetime.datetime.fromtimestamp(xlsx_mtime).strftime('%Y-%m-%d')
    meta = {'lastUpdate': last_update}
    if mktcap_as_of:
        meta['companiesAsOf'] = mktcap_as_of
    with open(data_dir / "meta.json", 'w', encoding='utf-8') as f:
        json.dump(meta, f, separators=(',', ':'))
    print(f"  meta.json: lastUpdate={last_update}" + (f", companiesAsOf={mktcap_as_of}" if mktcap_as_of else ""))

    # Gold Standard map (workbook col L of Team List, extracted above before
    # the workbook was closed). Consumed by lib/goldStandard.ts via static
    # JSON import alongside a curated Football men's Big 5 override on the
    # TS side. Emit even when empty so the file always exists for the import.
    print("Writing gold-standard-leagues.json...")
    with open(data_dir / "gold-standard-leagues.json", 'w', encoding='utf-8') as f:
        json.dump({'sports': gold_leagues}, f, separators=(',', ':'))
    gs_size = os.path.getsize(data_dir / "gold-standard-leagues.json")
    gs_total = sum(len(v) for v in gold_leagues.values())
    print(f"  gold-standard-leagues.json: {gs_size:,} bytes ({len(gold_leagues)} sports, {gs_total} leagues)")

    print("\nDone. Data files written to public/data/")
    print(f"  metros.json ({len(slim_metros)} metros)")
    print(f"  regions.json ({len(regions)} regions)")
    print(f"  details/ ({detail_count} files)")
    print(f"  meta.json (lastUpdate: {last_update})")
    # Regenerate quiz forward queue against the freshly written data.
    # Regenerate quiz forward queue against the freshly written data.
    # Locked issues are preserved by generator idempotency. Forward slots
    # are recomputed. The CI guard validates non-strict; tier-band slips
    # emit warnings rather than failing the ETL.
    import subprocess
    print("\n--- quiz queue ---")
    try:
        subprocess.run(["python3", "scripts/generate_quiz_questions.py", "--days", "30"],
                       check=True)
        subprocess.run(["python3", "scripts/check_quiz_queue.py"], check=False)
    except Exception as e:
        print(f"  quiz queue regeneration failed: {e}", file=sys.stderr)


if __name__ == "__main__":
    main()

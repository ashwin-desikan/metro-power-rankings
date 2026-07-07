#!/usr/bin/env python3
"""One-time loader: OtherLeagues.xlsx Women's Club Football + FACup-LgCup SF
(domestic cups) + Team Valuations sheets -> Supabase tables (source of truth).
Plain INSERT under a temporary anon-write RLS policy; tables are truncated
server-side immediately before this runs.

    python scripts/supabase/load_wfootball_cups_valuations.py            # all
    python scripts/supabase/load_wfootball_cups_valuations.py valuations # one group

Extraction is byte-parity-locked: it reproduces exactly the row shapes the
rewired build-wfootball-data.py / build-domestic-cups-data.py /
build-valuations-data.py expect (verified against committed JSON, ignoring the
daily generated timestamps).
"""
import os, sys, json, urllib.request, urllib.error
from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
XLSX = os.path.join(ROOT, "OtherLeagues.xlsx")
SB_URL = (os.environ.get("SUPABASE_URL") or "https://nmprqkmymrdknffwnuur.supabase.co").rstrip("/")
KEY = os.environ.get("SUPABASE_ANON_KEY") or "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Im5tcHJxa215bXJka25mZndudXVyIiwicm9sZSI6ImFub24iLCJpYXQiOjE3ODMyMDkzNDMsImV4cCI6MjA5ODc4NTM0M30.4RXU3mQ-Yl81ZqC2_a10aizKGu_87B4vt8OK5Pi_-sM"

def T(v):      return None if v is None else str(v)     # raw text, null for empty
def Iexact(v): return v if isinstance(v, int) else None  # preserve int, else null

def _hdr_index(ws):
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(c).strip() if c is not None else "" for c in rows[0]]
    return rows, hdr, {h: i for i, h in enumerate(hdr)}

def extract_wfootball(wb):
    out = []
    for r in list(wb["Women's Club Football"].iter_rows(values_only=True))[1:]:
        c = lambda i: r[i] if i < len(r) else None
        out.append({"competition": T(c(0)), "country": T(c(1)), "year": Iexact(c(2)),
                    "winner_disp": T(c(3)), "score": T(c(4)), "runner_disp": T(c(5)),
                    "winner_canon": T(c(6)), "runner_canon": T(c(7))})
    return {"womens_club_football": out}

def extract_valuations(wb):
    rows, hdr, ix = _hdr_index(wb["Team Valuations"])
    out = []
    for r in rows[1:]:
        g = lambda n: r[ix[n]] if n in ix and ix[n] < len(r) else None
        val = g("Value ($M)")
        out.append({"year": Iexact(g("Year")), "team": T(g("Team")), "league": T(g("League")),
                    "value_m": float(val) if val is not None else None, "source": T(g("Source"))})
    return {"team_valuations": out}

_DC_KEY = {"Team": "team", "Leag/Comp.": "leag_comp", "Year": "year", "Season": "season",
           "Comp. Rnd": "comp_rnd", "Cur. Name": "cur_name", "Opp. Name": "opp_name",
           "W/D/L": "wdl", "Trophy Won": "trophy_won", "Cup Final": "cup_final",
           "Metro Area": "metro_area", "County": "county", "Country": "country",
           "Continent": "continent", "YYYYMMDD": "yyyymmdd", "Date": "date_str",
           "For": "for_val", "Ag": "ag_val", "Comp Leg": "comp_leg", "Stadium": "stadium",
           "Stad. Metro Area": "stad_metro_area"}
_DC_INT = {"year", "for_val", "ag_val"}

def extract_domcups(wb):
    rows, hdr, ix = _hdr_index(wb["FACup-LgCup SF"])
    out = []
    for r in rows[1:]:
        d = {}
        for h, k in _DC_KEY.items():
            v = r[ix[h]] if h in ix and ix[h] < len(r) else None
            d[k] = Iexact(v) if k in _DC_INT else T(v)
        out.append(d)
    return {"domestic_cups": out}

GROUPS = {"wfootball": extract_wfootball, "domcups": extract_domcups, "valuations": extract_valuations}

def post(table, records, chunk=500):
    done = 0
    for i in range(0, len(records), chunk):
        batch = records[i:i + chunk]
        req = urllib.request.Request(
            f"{SB_URL}/rest/v1/{table}",
            data=json.dumps(batch).encode(), method="POST",
            headers={"apikey": KEY, "Authorization": f"Bearer {KEY}",
                     "Content-Type": "application/json", "Prefer": "return=minimal"})
        try:
            with urllib.request.urlopen(req, timeout=120):
                pass
        except urllib.error.HTTPError as e:
            sys.exit(f"HTTP {e.code} writing {table}: {e.read().decode(errors='replace')[:300]}\n"
                     f"(A 401/403 means the temporary anon-write policy isn't in place.)")
        done += len(batch)
    return done

if __name__ == "__main__":
    which = sys.argv[1:] or list(GROUPS)
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    for g in which:
        if g not in GROUPS:
            sys.exit(f"unknown group {g!r}; choose from {list(GROUPS)}")
        for table, recs in GROUPS[g](wb).items():
            print(f"{g} -> {table}: inserted {post(table, recs)} rows")

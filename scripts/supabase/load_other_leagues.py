#!/usr/bin/env python3
"""One-time / re-runnable historical loader: OtherLeagues.xlsx sheets -> Supabase
tables (source of truth). Idempotent upsert. Run where there is Supabase egress
(your machine or the mini; the Cowork sandbox has none).

Service key resolution order: SUPABASE_SERVICE_KEY env -> .env.local/.env file
-> secure prompt. Just run it and paste the key when asked:
    python scripts/supabase/load_other_leagues.py euroleague
The key is the service_role secret in Supabase -> Settings -> API.
"""
import os, sys, json, base64, getpass, urllib.request, urllib.error
from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
XLSX = os.path.join(ROOT, "OtherLeagues.xlsx")
SB_URL = (os.environ.get("SUPABASE_URL") or "https://nmprqkmymrdknffwnuur.supabase.co").rstrip("/")
# Public anon key fallback (used with the temporary anon-write policy when a
# service_role key is unavailable). Safe: this key is already public.
_ANON = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Im5tcHJxa215bXJka25mZndudXVyIiwicm9sZSI6ImFub24iLCJpYXQiOjE3ODMyMDkzNDMsImV4cCI6MjA5ODc4NTM0M30.4RXU3mQ-Yl81ZqC2_a10aizKGu_87B4vt8OK5Pi_-sM"

def _key():
    # TEMP: this project's service_role key is rejected (new key system), so the
    # one-time load uses the public anon key + a temporary anon-write RLS policy.
    print("Using the public anon key + temporary anon-write policy for this load.")
    return _ANON
    k = (os.environ.get("SUPABASE_SERVICE_KEY") or os.environ.get("SUPABASE_SERVICE_ROLE_KEY") or "").strip()
    if not k:
        for fn in (".env.local", ".env"):
            p = os.path.join(ROOT, fn)
            if os.path.exists(p):
                for line in open(p, encoding="utf-8"):
                    for name in ("SUPABASE_SERVICE_KEY", "SUPABASE_SERVICE_ROLE_KEY"):
                        if line.strip().startswith(name + "="):
                            k = line.strip().split("=", 1)[1].strip().strip('"').strip("'")
    if not k:
        print("No service key set; using the public anon key + the temporary anon-write policy.")
        k = _ANON
    # sanity: warn if this is obviously the anon/read key
    try:
        if k.count(".") == 2:
            role = json.loads(base64.urlsafe_b64decode(k.split(".")[1] + "==")).get("role")
            if role != "service_role":
                print(f"WARNING: this key's role is '{role}', not 'service_role'. Writes will 401.")
    except Exception:
        pass
    return k

def C(r, i): return r[i] if i < len(r) else None
def S(x): return None if x in (None, "") else str(x).strip()
def I(x):
    try: return int(x)
    except Exception: return None
def F(x):
    try: return float(x)
    except Exception: return None
def Y(x): return x == "Y"

def _euroleague(r):
    if S(C(r,0)) is None or S(C(r,2)) is None: return None
    return {"season": S(C(r,0)), "competition": S(C(r,1)), "team": S(C(r,2)),
            "canonical_name": S(C(r,20)), "country": S(C(r,3)), "grp": S(C(r,4)),
            "round_1": S(C(r,5)), "round_2": S(C(r,6)), "group_stage": S(C(r,7)),
            "w": I(C(r,8)), "l": I(C(r,9)), "win_pct": F(C(r,10)),
            "playoffs": Y(C(r,11)), "qf_app": Y(C(r,12)), "final_four_app": Y(C(r,13)),
            "final_app": Y(C(r,14)), "champion": Y(C(r,15)),
            "qf_w": I(C(r,16)), "qf_l": I(C(r,17)), "f4_w": I(C(r,18)), "f4_l": I(C(r,19))}

def _wnba_seasons(r):
    if not isinstance(C(r,0), int): return None
    return {"season": C(r,0), "team": S(C(r,1)), "conference": S(C(r,2)),
            "w": I(C(r,3)), "l": I(C(r,4)), "win_pct": F(C(r,5)), "gb": F(C(r,6)),
            "ps_g": F(C(r,7)), "pf_g": F(C(r,8)), "playoffs": Y(C(r,9)),
            "div_title": Y(C(r,10)), "best_rec": Y(C(r,11)), "p_wins": I(C(r,12)),
            "p_losses": I(C(r,13)), "sf_app": Y(C(r,14)), "champ_app": Y(C(r,15)),
            "champ": Y(C(r,16)), "canonical_name": S(C(r,17))}

def _ipl_standings(r):
    if not isinstance(C(r,0), int): return None
    return {"season": C(r,0), "pos": I(C(r,1)), "team": S(C(r,2)),
            "canonical_name": S(C(r,12)), "m": I(C(r,3)), "w": I(C(r,4)),
            "l": I(C(r,5)), "nr": I(C(r,6)), "pts": I(C(r,7)), "nrr": F(C(r,8)),
            "playoffs": Y(C(r,9)), "finalist": Y(C(r,10)), "champion": Y(C(r,11)),
            "active": Y(C(r,13)), "metro_area": S(C(r,14)), "state": S(C(r,15))}

def _ipl_playoff(r):
    if not isinstance(C(r,0), int) or S(C(r,1)) is None: return None
    return {"season": C(r,0), "round": S(C(r,1)), "team1": S(C(r,2)),
            "team2": S(C(r,3)), "result": S(C(r,4))}

REGISTRY = {
    "euroleague": [("Euroleague Table", "euroleague_seasons", "season,competition,team", _euroleague)],
    "wnba": [("WNBA", "wnba_seasons", "season,team", _wnba_seasons)],
    "ipl": [("IPL Table", "ipl_standings", "season,team", _ipl_standings),
            ("IPL Playoff Matches", "ipl_playoff_matches", "season,round", _ipl_playoff)],
}

def upsert(key, table, on_conflict, records, chunk=500):
    done = 0
    for i in range(0, len(records), chunk):
        batch = records[i:i+chunk]
        req = urllib.request.Request(
            f"{SB_URL}/rest/v1/{table}?on_conflict={on_conflict}",
            data=json.dumps(batch).encode(), method="POST",
            headers={"apikey": key, "Authorization": f"Bearer {key}",
                     "Content-Type": "application/json",
                     "Prefer": "resolution=merge-duplicates,return=minimal"})
        try:
            with urllib.request.urlopen(req, timeout=90):
                pass
        except urllib.error.HTTPError as e:
            body = e.read().decode(errors="replace")[:300]
            sys.exit(f"HTTP {e.code} writing {table}: {body}\n"
                     f"(A 401 means the key isn't the service_role secret.)")
        done += len(batch)
    return done

def load_one(key):
    k = _key()
    for sheet, table, oc, fn in REGISTRY[key]:
        ws = load_workbook(XLSX, read_only=True, data_only=True)[sheet]
        recs = [rec for rec in (fn(r) for r in list(ws.iter_rows(values_only=True))[1:]) if rec]
        print(f"{key} -> {table}: upserted {upsert(k, table, oc, recs)} rows")

if __name__ == "__main__":
    keys = list(REGISTRY) if (len(sys.argv) < 2 or sys.argv[1] == "all") else [sys.argv[1]]
    for k in keys: load_one(k)

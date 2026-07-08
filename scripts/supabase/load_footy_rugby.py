#!/usr/bin/env python3
"""One-time loader: OtherLeagues.xlsx footy (CFL, AFL/NRL) + rugby union sheets
-> Supabase tables (source of truth). Plain INSERT under a temporary anon-write
RLS policy; the target tables are truncated server-side immediately before this
runs, so a single clean pass loads the full history.

    python scripts/supabase/load_footy_rugby.py            # all
    python scripts/supabase/load_footy_rugby.py cfl        # one group

The extraction here is byte-parity-locked: it reproduces exactly the row shapes
that the rewired build-*.py scripts expect, verified against the committed JSON.
"""
import os, sys, json, urllib.request, urllib.error
from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
XLSX = os.path.join(ROOT, "OtherLeagues.xlsx")
SB_URL = (os.environ.get("SUPABASE_URL") or "https://nmprqkmymrdknffwnuur.supabase.co").rstrip("/")
# Public anon key (already public); used with the temporary anon-write policy.
KEY = os.environ.get("SUPABASE_ANON_KEY") or "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Im5tcHJxa215bXJka25mZndudXVyIiwicm9sZSI6ImFub24iLCJpYXQiOjE3ODMyMDkzNDMsImV4cCI6MjA5ODc4NTM0M30.4RXU3mQ-Yl81ZqC2_a10aizKGu_87B4vt8OK5Pi_-sM"

# ── helpers (semantics matched per sheet to the rewired build scripts) ──────────
def S(v):  return str(v).strip() if v is not None else ""          # text -> "" for empty
def I0(v):                                                         # int, 0 on fail (CFL)
    try: return int(float(v))
    except (TypeError, ValueError): return 0
def Inone(v):                                                      # int, None on fail (AFL/NRL)
    try: return int(float(v))
    except (TypeError, ValueError): return None
def YN(v): return S(v).upper() == "Y"
def T(v):  return None if v is None else str(v)                    # rugby text -> null for empty
def Iv(v):                                                         # rugby int -> None on fail
    if v is None: return None
    try: return int(v)
    except (TypeError, ValueError):
        try: return int(float(v))
        except (TypeError, ValueError): return None

# ── extractors -> {table: [records]} ───────────────────────────────────────────
def extract_cfl(wb):
    st = []
    for r in list(wb["CFL Standings"].iter_rows(values_only=True))[1:]:
        if not r or not S(r[14]): continue
        yr = I0(r[0])
        if not yr: continue
        st.append({"year": yr, "division": S(r[2]), "team": S(r[3]),
                   "w": I0(r[4]), "l": I0(r[5]), "t": I0(r[6]),
                   "pct": float(r[7]) if isinstance(r[7], (int, float)) else None,
                   "pf": I0(r[8]), "pa": I0(r[9]),
                   "play_app": YN(r[10]), "gc_final": YN(r[11]), "grey_cup": YN(r[12]),
                   "playoff_result": S(r[13]), "canonical": S(r[14])})
    gc = []
    for r in list(wb["CFL Grey Cup Finals"].iter_rows(values_only=True))[1:]:
        if not r or not S(r[12]): continue
        date = r[1]; yr = date.year if hasattr(date, "year") else I0(r[1])
        gc.append({"game": S(r[0]), "year": yr, "result": S(r[3]),
                   "pf": I0(r[4]), "pa": I0(r[5]), "ot": YN(r[6]),
                   "venue": S(r[8]), "city": S(r[9]), "attendance": I0(r[11]),
                   "name": S(r[12]), "opponent": S(r[13])})
    return {"cfl_standings": st, "cfl_grey_cup_finals": gc}

def extract_aflnrl(wb):
    ws = wb["AFL-NRL Ladders"]
    ix = {h: i for i, h in enumerate([c.value for c in next(ws.iter_rows(min_row=1, max_row=1))])}
    Lmap = {"sport": "Sport", "name": "Name", "team": "Team", "season": "Season",
            "league": "League", "rank": "Rank", "played": "Played", "wins": "Wins",
            "draws": "Draws", "losses": "Losses", "premiership_points": "PremiershipPoints",
            "points_for": "PointsFor", "points_against": "PointsAgainst",
            "minor_prem": "Minor Prem", "finals": "Finals", "grand_final_app": "Grand Final App",
            "premiership": "Premiership", "metro_area": "Metro Area", "state": "State"}
    Lbool = {"minor_prem", "finals", "grand_final_app", "premiership"}
    Lint = {"season", "rank", "played", "wins", "draws", "losses",
            "premiership_points", "points_for", "points_against"}
    lad = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        d = {}
        for k, h in Lmap.items():
            v = r[ix[h]]
            d[k] = YN(v) if k in Lbool else (Inone(v) if k in Lint else S(v))
        lad.append(d)
    ws = wb["AFL-NRL Grand Finals"]
    gx = {h: i for i, h in enumerate([c.value for c in next(ws.iter_rows(min_row=1, max_row=1))])}
    Gmap = {"sport": "Sport", "name": "Name", "team": "Team", "year": "Year  ",
            "date": "Date (YYYYMMDD)", "wl": "W/L", "opp_team": "Opp Team",
            "opponent": "Opponent", "pf": "For", "pa": "Ag", "stadium": "Stadium",
            "metro_area": "Metro Area", "state": "State", "premiership_won": "Premiership won"}
    Gbool = {"premiership_won"}
    Gint = {"year", "pf", "pa"}
    gf = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        d = {}
        for k, h in Gmap.items():
            v = r[gx[h]]
            d[k] = YN(v) if k in Gbool else (Inone(v) if k in Gint else S(v))
        gf.append(d)
    return {"afl_nrl_ladders": lad, "afl_nrl_grand_finals": gf}

def extract_rugby(wb):
    res = []
    for r in list(wb["Rugby Union - Intl Results"].iter_rows(values_only=True))[1:]:
        res.append({"date": T(r[0]), "team": T(r[1]), "wld": T(r[2]), "opp": T(r[3]),
                    "pf": Iv(r[4]), "pa": Iv(r[5]), "comp": T(r[7]), "stage": T(r[8]),
                    "stadium": T(r[10]), "city": T(r[11]), "country": T(r[12]),
                    "home_away": T(r[13]),
                    "home_five_six_nations": bool(r[14]), "tri_nations_rugby_champ": bool(r[15]),
                    "nations_championship": bool(r[16]), "rugby_world_cup": bool(r[17])})
    tab = []
    for r in list(wb["Rugby Union - Intl Tables"].iter_rows(values_only=True))[1:]:
        tab.append({"season": T(r[0]), "comp": T(r[1]), "pool": T(r[2]), "place": T(r[3]),
                    "team": T(r[4]), "rwc_qf": r[17] == "Y", "rwc_sf": r[18] == "Y",
                    "rwc_f": r[19] == "Y", "trophy": r[20] == "Y",
                    "triple_crown": r[21] == "Y", "grand_slam": r[22] == "Y"})
    return {"rugby_results": res, "rugby_tables": tab}

GROUPS = {"cfl": extract_cfl, "aflnrl": extract_aflnrl, "rugby": extract_rugby}

def post(table, records, chunk=500):
    done = 0
    for i in range(0, len(records), chunk):
        batch = records[i:i + chunk]
        req = urllib.request.Request(
            f"{SB_URL}/rest/v1/{table}",
            data=json.dumps(batch).encode(), method="POST",
            headers={"apikey": KEY, "Authorization": f"Bearer {KEY}",
                     "Content-Type": "application/json", "Prefer": "return=minimal"})
        try:
            with urllib.request.urlopen(req, timeout=120):
                pass
        except urllib.error.HTTPError as e:
            sys.exit(f"HTTP {e.code} writing {table}: {e.read().decode(errors='replace')[:300]}\n"
                     f"(A 401/403 means the temporary anon-write policy isn't in place.)")
        done += len(batch)
    return done

if __name__ == "__main__":
    which = sys.argv[1:] or list(GROUPS)
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    for g in which:
        if g not in GROUPS:
            sys.exit(f"unknown group {g!r}; choose from {list(GROUPS)}")
        for table, recs in GROUPS[g](wb).items():
            print(f"{g} -> {table}: inserted {post(table, recs)} rows")

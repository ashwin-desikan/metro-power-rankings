#!/usr/bin/env python3
"""Rebuild the OtherLeagues.xlsx "Team Valuations" sheet from Sportico's 2026
Most Valuable Sports Franchises list.

Why this replaces rather than supplements the old sheet: the board sorts every
row into ONE column, so a Forbes NFL figure sitting above a Sportico football
figure is not a ranking, it is two rankings interleaved. Sportico's 2026 list
covers all 32 NFL, 30 NBA, 30 MLB and 32 NHL clubs -- exactly the 124 rows
Forbes held -- so the whole board can move to one house and one vintage.

Rows the new list does NOT contain (clubs that fell out of Sportico's top 206)
are CARRIED at their existing year and source rather than dropped, which is what
"latest published valuation per team" has always meant on this page.

DRY RUN BY DEFAULT. Prints the full diff and writes nothing until --write.

    python scripts/valuations/load_sportico_2026.py            # show the diff
    python scripts/valuations/load_sportico_2026.py --write    # update the sheet
    python scripts/valuations/load_sportico_2026.py --self-test

The workbook is ground truth (see CLAUDE.md); Supabase is reloaded FROM it and
public/data/valuations/valuations.json is built from Supabase. This script owns
step one only.
"""
import argparse, os, re, sys, io, csv

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
XLSX = os.path.join(ROOT, "OtherLeagues.xlsx")
SHEET = "Team Valuations"
RAW = os.path.join(os.path.dirname(os.path.abspath(__file__)), "sources", "sportico-2026.txt")
REVIEW = os.path.join(os.path.dirname(os.path.abspath(__file__)), "out", "sportico_2026_diff.csv")
YEAR = 2026
SOURCE = "Sportico, The Most Valuable Sports Franchises 2026"

# --- League rosters. Explicit, not inferred: a fuzzy sport guess is exactly the
# kind of silent wrong answer this project keeps paying for. Any team on the raw
# list that appears in none of these AND in no existing sheet row is a hard error.
ROSTERS = {
    "NFL": """Dallas Cowboys|Los Angeles Rams|New York Giants|New England Patriots|New York Jets|
        Philadelphia Eagles|San Francisco 49ers|Miami Dolphins|Las Vegas Raiders|Atlanta Falcons|
        Washington Commanders|Seattle Seahawks|Houston Texans|Chicago Bears|Denver Broncos|
        Kansas City Chiefs|Pittsburgh Steelers|Tampa Bay Buccaneers|Los Angeles Chargers|
        Tennessee Titans|Cleveland Browns|Green Bay Packers|Minnesota Vikings|Baltimore Ravens|
        Buffalo Bills|Carolina Panthers|Detroit Lions|Arizona Cardinals|Jacksonville Jaguars|
        Indianapolis Colts|New Orleans Saints|Cincinnati Bengals""",
    "NBA": """Golden State Warriors|Los Angeles Lakers|New York Knicks|Los Angeles Clippers|
        Boston Celtics|Brooklyn Nets|Chicago Bulls|Miami Heat|Philadelphia 76ers|Houston Rockets|
        Dallas Mavericks|Toronto Raptors|Phoenix Suns|Atlanta Hawks|Sacramento Kings|
        Cleveland Cavaliers|Denver Nuggets|Washington Wizards|Indiana Pacers|San Antonio Spurs|
        Oklahoma City Thunder|Milwaukee Bucks|Portland Trail Blazers|Utah Jazz|Orlando Magic|
        Charlotte Hornets|Detroit Pistons|Minnesota Timberwolves|New Orleans Pelicans|
        Memphis Grizzlies""",
    "MLB": """New York Yankees|Los Angeles Dodgers|Boston Red Sox|Chicago Cubs|San Francisco Giants|
        New York Mets|Philadelphia Phillies|Atlanta Braves|Houston Astros|Los Angeles Angels|
        St. Louis Cardinals|Texas Rangers|Seattle Mariners|Toronto Blue Jays|Washington Nationals|
        Chicago White Sox|San Diego Padres|Baltimore Orioles|Athletics|Milwaukee Brewers|
        Arizona Diamondbacks|Detroit Tigers|Minnesota Twins|Colorado Rockies|Cleveland Guardians|
        Pittsburgh Pirates|Cincinnati Reds|Kansas City Royals|Tampa Bay Rays|Miami Marlins""",
    "NHL": """Toronto Maple Leafs|New York Rangers|Montreal Canadiens|Edmonton Oilers|
        Los Angeles Kings|Boston Bruins|Chicago Blackhawks|Philadelphia Flyers|Washington Capitals|
        Detroit Red Wings|New Jersey Devils|Dallas Stars|Vegas Golden Knights|Vancouver Canucks|
        New York Islanders|Tampa Bay Lightning|Carolina Hurricanes|Colorado Avalanche|
        Calgary Flames|Seattle Kraken|Minnesota Wild|Pittsburgh Penguins|Florida Panthers|
        Nashville Predators|St. Louis Blues|San Jose Sharks|Utah Mammoth|Anaheim Ducks|
        Ottawa Senators|Winnipeg Jets|Buffalo Sabres|Columbus Blue Jackets""",
    "F1": """Ferrari|Mercedes|McLaren|Red Bull Racing|Aston Martin|Alpine|Williams|Racing Bulls|
        Haas F1 Team|Kick Sauber|Audi|Cadillac""",
    "WNBA": """New York Liberty|Golden State Valkyries|Las Vegas Aces|Indiana Fever|Seattle Storm|
        Phoenix Mercury|Los Angeles Sparks|Minnesota Lynx|Dallas Wings|Chicago Sky|Atlanta Dream|
        Connecticut Sun|Washington Mystics""",
    "NWSL": """Angel City FC|Gotham FC|Kansas City Current|Bay FC|San Diego Wave FC|
        Portland Thorns FC|Washington Spirit|Racing Louisville FC|North Carolina Courage|
        Seattle Reign FC|Orlando Pride|Chicago Stars FC|Houston Dash|Utah Royals FC|
        Denver Summit FC|Boston Legacy FC""",
}
LEAGUE_OF = {}
for _lg, _blob in ROSTERS.items():
    for _t in _blob.replace("\n", "").split("|"):
        _t = _t.strip()
        if _t:
            LEAGUE_OF[_t] = _lg

# --- Sportico's rendering -> the name the sheet (and therefore resolveTeamLink)
# already uses. Only listed where the two genuinely differ; an unlisted name must
# match a roster or an existing sheet row exactly.
ALIAS = {
    "Atletico de Madrid": "Atlético de Madrid",
    "Inter Milan": "Internazionale",
    "Napoli": "SSC Napoli",
    "LA Galaxy": "Los Angeles Galaxy",
    "Brighton & Hove": "Brighton & Hove Albion",
    "America": "CF América",
    "Guadalajara": "Chivas Guadalajara",
    "D.C. United": "DC United",
    "AFC Ajax": "Ajax",
    "SL Benfica": "Benfica",
    "Chicago Fire FC": "Chicago Fire",
    "Minnesota United FC": "Minnesota United",
    "Houston Dynamo FC": "Houston Dynamo",
    # The accent is load-bearing: getFootballClubByName matches exactly, so the
    # unaccented spelling was the one row of 214 that rendered without a link.
    "CF Montreal": "CF Montréal",
}

# --- Football clubs new to the sheet. The League column holds the club's COUNTRY
# for football, and the existing sheet files Toronto FC under "United States"
# because the column really tracks the LEAGUE's country, not the club's. The two
# other Canadian MLS sides follow that established convention rather than
# splitting MLS across two country labels.
NEW_FOOTBALL_COUNTRY = {
    "Real Salt Lake": "United States",
    "Orlando City SC": "United States",
    "New England Revolution": "United States",
    "FC Dallas": "United States",
    "Colorado Rapids": "United States",
    "Vancouver Whitecaps FC": "United States",
    "CF Montréal": "United States",
}


def parse_raw(path):
    """The pasted table is rank / name / value<TAB>revenue<TAB>growth, one per line."""
    lines = [l.rstrip("\n") for l in io.open(path, encoding="utf-8")]
    out, i = [], 0
    while i < len(lines):
        if re.fullmatch(r"\d+", lines[i].strip()):
            rank = int(lines[i].strip())
            team = lines[i + 1].strip()
            cells = lines[i + 2].split("\t")
            out.append({"rank": rank, "team": team, "value_m": money(cells[0])})
            i += 3
        else:
            i += 1
    return out


def money(s):
    m = re.match(r"^\$([\d.]+)([BM])$", (s or "").strip())
    if not m:
        return None
    v = float(m.group(1))
    return v * 1000 if m.group(2) == "B" else v


def self_test():
    assert money("$15.50B") == 15500.0
    assert money("$945M") == 945.0
    assert money("$6.40B") == 6400.0
    assert money("") is None and money("21%") is None
    assert LEAGUE_OF["Dallas Cowboys"] == "NFL"
    assert LEAGUE_OF["Haas F1 Team"] == "F1"
    assert LEAGUE_OF["Angel City FC"] == "NWSL"
    assert LEAGUE_OF["Dallas Wings"] == "WNBA"
    # A club name must not sit in two rosters: Dallas has a Cowboys, a Mavericks,
    # a Stars, an FC and a Wings, and the earlier pass classified none of them by
    # city. Uniqueness is what makes the exact-name lookup safe.
    seen = {}
    for lg, blob in ROSTERS.items():
        for t in blob.replace("\n", "").split("|"):
            t = t.strip()
            if not t:
                continue
            assert t not in seen, f"{t!r} in both {seen[t]} and {lg}"
            seen[t] = lg
    assert len(seen) == len(LEAGUE_OF)
    # Aliases must point somewhere real, never at another alias key.
    for k, v in ALIAS.items():
        assert k != v and v not in ALIAS, f"alias {k!r} -> {v!r} chains"
    print("self-test OK")


def read_sheet():
    from openpyxl import load_workbook
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb[SHEET]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(c).strip() if c is not None else "" for c in rows[0]]
    ix = {h: i for i, h in enumerate(hdr)}
    out = []
    for r in rows[1:]:
        g = lambda n: r[ix[n]] if n in ix and ix[n] < len(r) else None
        if g("Team") is None:
            continue
        out.append({"year": g("Year"), "team": str(g("Team")).strip(),
                    "league": str(g("League")).strip() if g("League") is not None else "",
                    "value_m": float(g("Value ($M)")), "source": str(g("Source") or "").strip()})
    wb.close()
    return hdr, out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--write", action="store_true")
    ap.add_argument("--self-test", action="store_true")
    a = ap.parse_args()
    if a.self_test:
        return self_test()
    self_test()

    if not os.path.exists(RAW):
        sys.exit(f"FATAL: {RAW} missing.")
    raw = parse_raw(RAW)
    ranks = [r["rank"] for r in raw]
    if ranks != list(range(1, len(ranks) + 1)):
        sys.exit(f"FATAL: ranks are not 1..{len(ranks)}; the paste is truncated or interleaved.")
    if any(r["value_m"] is None for r in raw):
        sys.exit("FATAL: unparsed value cells; refusing to write a partial list.")
    print(f"parsed {len(raw)} Sportico rows, ranks 1-{len(raw)}")

    hdr, existing = read_sheet()
    by_team = {e["team"]: e for e in existing}
    print(f"existing sheet: {len(existing)} rows")

    new, unknown = [], []
    for r in raw:
        team = ALIAS.get(r["team"], r["team"])
        lg = LEAGUE_OF.get(team) or NEW_FOOTBALL_COUNTRY.get(team)
        if not lg and team in by_team:
            lg = by_team[team]["league"]           # football country already ruled on
        if not lg:
            unknown.append((r["rank"], r["team"], team))
            continue
        new.append({"year": YEAR, "team": team, "league": lg,
                    "value_m": r["value_m"], "source": SOURCE})
    if unknown:
        for rank, orig, mapped in unknown:
            print(f"  UNMAPPED #{rank}: {orig!r} (as {mapped!r})")
        sys.exit(f"FATAL: {len(unknown)} teams have no league and no existing row. Add them to "
                 f"ROSTERS, ALIAS or NEW_FOOTBALL_COUNTRY rather than letting them fall through.")

    fresh = {n["team"] for n in new}
    carried = [e for e in existing if e["team"] not in fresh]
    merged = sorted(new + carried, key=lambda x: -x["value_m"])


    from collections import Counter
    added = [n for n in new if n["team"] not in by_team]
    moved = [(n, by_team[n["team"]]) for n in new if n["team"] in by_team]
    print(f"\n{len(new)} refreshed to {YEAR} | {len(added)} NEW to the sheet | "
          f"{len(carried)} carried at their old year | {len(merged)} total")
    print("by league:", dict(Counter(m['league'] for m in merged)))

    print(f"\nNEW ROWS ({len(added)})")
    for n in sorted(added, key=lambda x: -x["value_m"]):
        print(f"  {n['league']:<14} {n['team']:<28} {n['value_m']:>8.0f}")
    print(f"\nCARRIED, NOT IN SPORTICO 2026 ({len(carried)})")
    for c in sorted(carried, key=lambda x: -x["value_m"]):
        print(f"  {c['league']:<14} {c['team']:<28} {c['value_m']:>8.0f}  ({c['year']})")
    print(f"\nBIGGEST MOVES vs the previous figure")
    deltas = sorted(moved, key=lambda p: -abs(p[0]["value_m"] - p[1]["value_m"]))[:15]
    for n, o in deltas:
        d = n["value_m"] - o["value_m"]
        print(f"  {n['team']:<28} {o['value_m']:>8.0f} -> {n['value_m']:>8.0f}  "
              f"{d:+8.0f}  ({d / o['value_m'] * 100:+.0f}%)")

    os.makedirs(os.path.dirname(REVIEW), exist_ok=True)
    with io.open(REVIEW, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["status", "league", "team", "old_value_m", "new_value_m", "old_year", "year"])
        for n in sorted(new, key=lambda x: -x["value_m"]):
            o = by_team.get(n["team"])
            w.writerow(["new" if not o else "refreshed", n["league"], n["team"],
                        o["value_m"] if o else "", n["value_m"], o["year"] if o else "", YEAR])
        for c in sorted(carried, key=lambda x: -x["value_m"]):
            w.writerow(["carried", c["league"], c["team"], c["value_m"], c["value_m"],
                        c["year"], c["year"]])
    print(f"\n-> {REVIEW}")

    if not a.write:
        print("\nDRY RUN. Nothing written. Re-run with --write once the diff above reads right.")
        return

    from openpyxl import load_workbook
    wb = load_workbook(XLSX)
    ws = wb[SHEET]
    ws.delete_rows(2, ws.max_row)
    for m in merged:
        ws.append([m["year"], m["team"], m["league"], m["value_m"], m["source"]])
    wb.save(XLSX)
    print(f"WROTE {len(merged)} rows to {SHEET} in {os.path.basename(XLSX)}")


if __name__ == "__main__":
    main()

#!/usr/bin/env python3
"""Push the OtherLeagues.xlsx "Team Valuations" sheet to Supabase, replacing it.

The workbook is ground truth; Supabase is a serving copy; build-valuations-data.py
reads Supabase. The original one-time loader
(scripts/supabase/load_wfootball_cups_valuations.py) assumed the table had been
truncated by hand under a temporary anon-write RLS policy. That is not a thing
anyone should have to remember, so this does the delete and the insert itself
with the SERVICE key from .env.local, and refuses to run without one rather than
half-writing under the anon key.

    python scripts/valuations/sync_team_valuations.py            # dry run
    python scripts/valuations/sync_team_valuations.py --write

Verifies the row count back out of PostgREST before reporting success: a silent
partial write here would ship a board with holes in it.
"""
import argparse, json, os, sys, urllib.request, urllib.error

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
XLSX = os.path.join(ROOT, "OtherLeagues.xlsx")
SHEET = "Team Valuations"
TABLE = "team_valuations"
SB_URL = (os.environ.get("SUPABASE_URL") or "https://nmprqkmymrdknffwnuur.supabase.co").rstrip("/")


def service_key():
    k = os.environ.get("SUPABASE_SERVICE_KEY")
    if k:
        return k
    env = os.path.join(ROOT, ".env.local")
    if os.path.exists(env):
        for line in open(env, encoding="utf-8"):
            if line.strip().startswith("SUPABASE_SERVICE_KEY="):
                return line.split("=", 1)[1].strip().strip('"').strip("'")
    sys.exit("FATAL: no SUPABASE_SERVICE_KEY in the environment or .env.local. "
             "Refusing to fall back to the anon key, which would fail half way.")


def call(method, path, key, body=None, extra=None):
    h = {"apikey": key, "Authorization": f"Bearer {key}",
         "Content-Type": "application/json", "Prefer": "return=minimal"}
    h.update(extra or {})
    req = urllib.request.Request(f"{SB_URL}{path}", method=method,
                                 data=json.dumps(body).encode() if body is not None else None,
                                 headers=h)
    try:
        with urllib.request.urlopen(req, timeout=120) as r:
            raw = r.read().decode() or ""
            return raw, dict(r.headers)
    except urllib.error.HTTPError as e:
        sys.exit(f"HTTP {e.code} on {method} {path}: {e.read().decode(errors='replace')[:400]}")


def read_sheet():
    from openpyxl import load_workbook
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    rows = list(wb[SHEET].iter_rows(values_only=True))
    wb.close()
    hdr = [str(c).strip() if c is not None else "" for c in rows[0]]
    ix = {h: i for i, h in enumerate(hdr)}
    for need in ("Year", "Team", "League", "Value ($M)", "Source"):
        if need not in ix:
            sys.exit(f"FATAL: the sheet has no {need!r} column.")
    out = []
    for r in rows[1:]:
        g = lambda n: r[ix[n]] if ix[n] < len(r) else None
        if g("Team") is None or g("Value ($M)") is None:
            continue
        out.append({"year": int(g("Year")) if g("Year") is not None else None,
                    "team": str(g("Team")).strip(),
                    "league": str(g("League")).strip() if g("League") is not None else "",
                    "value_m": float(g("Value ($M)")),
                    "source": str(g("Source") or "").strip()})
    return out


def count(key):
    _, h = call("GET", f"/rest/v1/{TABLE}?select=id&limit=1", key,
                extra={"Prefer": "count=exact", "Range-Unit": "items", "Range": "0-0"})
    cr = h.get("Content-Range") or h.get("content-range") or ""
    return int(cr.split("/")[-1]) if "/" in cr else -1


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--write", action="store_true")
    a = ap.parse_args()

    rows = read_sheet()
    from collections import Counter
    print(f"sheet: {len(rows)} rows")
    print("by league:", dict(Counter(r["league"] for r in rows)))
    print("by source:", dict(Counter(r["source"] for r in rows)))
    dupes = [t for t, n in Counter(r["team"] for r in rows).items() if n > 1]
    if dupes:
        sys.exit(f"FATAL: duplicate team rows would make the /teams lookup ambiguous: {dupes}")

    key = service_key()
    before = count(key)
    print(f"supabase {TABLE}: {before} rows now")
    if not a.write:
        print("DRY RUN. Nothing written. Re-run with --write.")
        return

    call("DELETE", f"/rest/v1/{TABLE}?id=gt.0", key)
    mid = count(key)
    if mid != 0:
        sys.exit(f"FATAL: delete left {mid} rows; refusing to insert on top of a partial table.")
    for i in range(0, len(rows), 500):
        call("POST", f"/rest/v1/{TABLE}", key, rows[i:i + 500])
    after = count(key)
    if after != len(rows):
        sys.exit(f"FATAL: wrote {after} rows, expected {len(rows)}.")
    print(f"OK: {TABLE} now holds {after} rows (was {before})")


if __name__ == "__main__":
    main()

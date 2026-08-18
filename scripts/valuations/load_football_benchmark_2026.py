#!/usr/bin/env python3
"""Merge Football Benchmark's "The European Elite 2026" into the Team Valuations
sheet, taking the HIGHER of the two published figures for any club both houses
cover, and recording which house each surviving figure came from.

Ashwin's ruling, 2026-08-18: Football Benchmark is the primary European football
source, the board shows the higher valuation regardless of source, and every row
says which source it is.

## What the two numbers actually are, because they are not the same quantity

Football Benchmark publishes ENTERPRISE VALUE, defined on its own methodology
page as "the sum of the market value of the owners' equity, plus total debt,
less cash and cash equivalents... regardless of the capital structure used to
finance its operations". It is built from statutory filings, revenue excludes
player trading, and the published figure is the MIDPOINT of a range. Sportico
publishes a franchise value and does not state its basis.

So the board is now explicitly "the highest published valuation we hold per
team, source shown", not "the latest". The page copy has to say that, and the
per-row source tag is what makes it checkable rather than a claim.

## The known cost of the higher-of rule, recorded so nobody rediscovers it

Only the 26 clubs BOTH houses cover can be uplifted by taking a maximum. The 11
European clubs Sportico alone covers cannot be. That moves clubs relative to
each other on coverage rather than on value: Everton at 1,000 (Sportico only)
falls below West Ham at 1,190 (FB), where a single ruler puts Everton ahead.
The source tag makes it visible. It does not make it go away.

    python scripts/valuations/load_football_benchmark_2026.py            # diff
    python scripts/valuations/load_football_benchmark_2026.py --write
    python scripts/valuations/load_football_benchmark_2026.py --self-test

Dry run by default. Then sync_team_valuations.py and build-valuations-data.py.
"""
import argparse, csv, io, os, re, sys

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
XLSX = os.path.join(ROOT, "OtherLeagues.xlsx")
SHEET = "Team Valuations"
HERE = os.path.dirname(os.path.abspath(__file__))
RAW = os.path.join(HERE, "sources", "football-benchmark-2026.txt")
REVIEW = os.path.join(HERE, "out", "football_benchmark_2026_diff.csv")
YEAR = 2026
SOURCE = "Football Benchmark, The European Elite 2026 (enterprise value, midpoint)"


# Football Benchmark's rendering -> the name the sheet already uses. A value of
# None means the club is NEW to the board and must appear in NEW_COUNTRY below,
# so a club can never be added without a country ruling.
ALIAS = {
    "Real Madrid CF": "Real Madrid",
    "FC Barcelona": "FC Barcelona",
    "Manchester City FC": "Manchester City",
    "Manchester United FC": "Manchester United",
    "Arsenal FC": "Arsenal",
    "FC Bayern München": "Bayern Munich",
    "Liverpool FC": "Liverpool",
    "Paris Saint-Germain FC": "Paris Saint-Germain",
    "Tottenham Hotspur FC": "Tottenham Hotspur",
    "Chelsea FC": "Chelsea",
    "Borussia Dortmund": "Borussia Dortmund",
    "FC Internazionale Milano": "Internazionale",
    "Atlético de Madrid": "Atlético de Madrid",
    "Juventus FC": "Juventus",
    "AC Milan": "AC Milan",
    "Newcastle United FC": "Newcastle United",
    "Aston Villa FC": "Aston Villa",
    "West Ham United FC": "West Ham United",
    "SSC Napoli": "SSC Napoli",
    "Eintracht Frankfurt": "Eintracht Frankfurt",
    "SL Benfica": "Benfica",
    "AS Roma": "AS Roma",
    "Galatasaray SK": None,
    "Olympique de Marseille": None,
    "Atalanta BC": "Atalanta",
    "PSV Eindhoven": None,
    "FC Porto": "FC Porto",
    "Feyenoord Rotterdam": None,
    "SS Lazio": "Lazio",
    "LOSC Lille": None,
    "AFC Ajax": "Ajax",
    "Real Sociedad de Fútbol": None,
}

# Clubs new to the board. The League column carries the country for football.
# Turkey is a new country here. It gets NO entry in lib/valuations.ts
# FOOTBALL_COUNTRY_HUB, deliberately: only nine league hubs are actually built,
# and pointing a label at an unbuilt slug produces a 404 rather than a link.
# The country label falls back to /teams/football until a Super Lig hub ships.
# Names are the CANONICAL ones the site's football data already uses, taken from
# public/data/football/index.json, not Football Benchmark's rendering. Getting
# this wrong does not error: getFootballClubByName returns null and the row
# renders unlinked, indistinguishable from a club we have no page for. That is
# the CF Montréal trap from earlier today, so these were checked against the
# data rather than typed from the report.
NEW_COUNTRY = {
    "Galatasaray SK": ("Galatasaray SK", "Turkey"),
    "Olympique de Marseille": ("Olympique Marseille", "France"),
    "PSV Eindhoven": ("PSV Eindhoven", "Netherlands"),
    "Feyenoord Rotterdam": ("Feyenoord", "Netherlands"),
    "LOSC Lille": ("Lille OSC", "France"),
    "Real Sociedad de Fútbol": ("Real Sociedad", "Spain"),
}

LINE = re.compile(
    r"^(\d+)\s+(=|NEW|[+-]\d+)\s+(.+?)\s+([\d,]+)\s+(=|NEW|-?\d+%)\s+([\d,]+)\s+([\d,]+)\s*$")


def parse_raw(path):
    """rank / rank-move / club / EUR m / YoY / GBP m / USD m, one club per line."""
    out = []
    for line in io.open(path, encoding="utf-8"):
        m = LINE.match(line.strip())
        if not m:
            continue
        num = lambda s: float(s.replace(",", ""))
        out.append({"rank": int(m.group(1)), "club": m.group(3).strip(),
                    "eur_m": num(m.group(4)), "yoy": m.group(5),
                    "gbp_m": num(m.group(6)), "usd_m": num(m.group(7))})
    return out


def self_test():
    ok = True

    def check(label, got, want):
        nonlocal ok
        if got != want:
            ok = False
            print(f"  FAIL {label}: got {got!r}, want {want!r}")
        else:
            print(f"  ok   {label}")

    row = LINE.match("1 = Real Madrid CF 7,725 23% 6,735 9,054")
    check("plain row parses", row.group(3) if row else None, "Real Madrid CF")
    check("USD is the last column", float(row.group(7).replace(",", "")), 9054.0)
    # A club entering the ranking has NEW in BOTH the move and the YoY column.
    nw = LINE.match("16 NEW Newcastle United FC 1,470 NEW 1,282 1,723")
    check("NEW entrant parses", nw.group(3) if nw else None, "Newcastle United FC")
    # An unchanged rank uses "=" in the move column and may use "=" for YoY too.
    eq = LINE.match("15 -2 AC Milan 1,807 = 1,576 2,118")
    check("'=' YoY parses", eq.group(3) if eq else None, "AC Milan")
    # A club name containing digits or a dash must not be eaten by the number
    # groups -- LOSC and 1899 Hoffenheim style names are the risk.
    ln = LINE.match("30 NEW LOSC Lille 509 NEW 444 597")
    check("name before the numbers", ln.group(3) if ln else None, "LOSC Lille")
    check("total line is ignored", LINE.match("Total 72,641 63,336 85,142"), None)
    check("header is ignored", LINE.match("RANKING CLUB MIDPOINT"), None)
    # Every alias either maps to a board name or is declared as a new club.
    for k, v in ALIAS.items():
        if v is None and k not in NEW_COUNTRY:
            ok = False
            print(f"  FAIL {k!r} maps to None but has no NEW_COUNTRY entry")
    check("every unmapped alias is declared new", True, True)
    print("self-test:", "PASS" if ok else "FAIL")
    return 0 if ok else 1


def read_sheet():
    from openpyxl import load_workbook
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    rows = list(wb[SHEET].iter_rows(values_only=True))
    wb.close()
    hdr = [str(c).strip() if c is not None else "" for c in rows[0]]
    ix = {h: i for i, h in enumerate(hdr)}
    out = []
    for r in rows[1:]:
        g = lambda n: r[ix[n]] if ix[n] < len(r) else None
        if g("Team") is None:
            continue
        out.append({"year": g("Year"), "team": str(g("Team")).strip(),
                    "league": str(g("League") or "").strip(),
                    "value_m": float(g("Value ($M)")),
                    "source": str(g("Source") or "").strip()})
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--write", action="store_true")
    ap.add_argument("--self-test", action="store_true")
    a = ap.parse_args()
    if a.self_test:
        return self_test()
    if self_test():
        return 1

    fb = parse_raw(RAW)
    ranks = [r["rank"] for r in fb]
    if ranks != list(range(1, len(ranks) + 1)):
        sys.exit(f"FATAL: ranks are not 1..{len(ranks)}; the paste is truncated.")
    # The report prints its own total. Checking against it is a free proof that
    # no club was dropped by the regex.
    stated = None
    for line in io.open(RAW, encoding="utf-8"):
        m = re.match(r"^Total\s+([\d,]+)", line.strip())
        if m:
            stated = float(m.group(1).replace(",", ""))
    got = sum(r["eur_m"] for r in fb)
    if stated is not None and abs(got - stated) > len(fb):
        sys.exit(f"FATAL: EUR total {got:,.0f} vs the report's {stated:,.0f}; a club is missing.")
    print(f"parsed {len(fb)} Football Benchmark clubs, EUR total {got:,.0f} "
          f"(report says {stated:,.0f})")

    sheet = read_sheet()
    by_team = {r["team"]: r for r in sheet}
    print(f"sheet: {len(sheet)} rows")

    took_fb, kept_sportico, added, unmapped = [], [], [], []
    for r in fb:
        if r["club"] not in ALIAS:
            unmapped.append(r["club"]); continue
        canon = ALIAS[r["club"]]
        if canon is None:
            name, country = NEW_COUNTRY[r["club"]]
            if name in by_team:
                sys.exit(f"FATAL: {name!r} is declared new but is already on the sheet.")
            added.append({"year": YEAR, "team": name, "league": country,
                          "value_m": r["usd_m"], "source": SOURCE, "fb": r})
            continue
        cur = by_team.get(canon)
        if cur is None:
            sys.exit(f"FATAL: {r['club']!r} maps to {canon!r}, which is not on the sheet. "
                     f"Either fix the alias or declare it in NEW_COUNTRY.")
        if r["usd_m"] > cur["value_m"]:
            took_fb.append((canon, cur, r))
        else:
            kept_sportico.append((canon, cur, r))
    if unmapped:
        sys.exit(f"FATAL: {len(unmapped)} clubs have no ALIAS entry: {unmapped}. "
                 f"Add them rather than letting them fall through.")

    print(f"\n{len(took_fb)} rows take the Football Benchmark figure, "
          f"{len(kept_sportico)} keep Sportico's, {len(added)} clubs are NEW")
    print(f"\nSWITCHING TO FOOTBALL BENCHMARK ({len(took_fb)})")
    for canon, cur, r in sorted(took_fb, key=lambda t: -(t[2]['usd_m'] - t[1]['value_m'])):
        d = r["usd_m"] - cur["value_m"]
        print(f"  {canon:<24} {cur['value_m']:>7,.0f} -> {r['usd_m']:>7,.0f}  "
              f"{d:>+7,.0f} {d / cur['value_m'] * 100:>+5.0f}%   YoY {r['yoy']}")
    print(f"\nKEEPING SPORTICO, ITS FIGURE IS HIGHER ({len(kept_sportico)})")
    for canon, cur, r in sorted(kept_sportico, key=lambda t: -t[1]['value_m']):
        print(f"  {canon:<24} {cur['value_m']:>7,.0f}  (FB had {r['usd_m']:>7,.0f})")
    print(f"\nNEW TO THE BOARD ({len(added)})")
    for n in sorted(added, key=lambda x: -x["value_m"]):
        print(f"  {n['league']:<12} {n['team']:<24} {n['value_m']:>7,.0f}")

    # The cost of the higher-of rule, printed every run so it is never a surprise.
    covered = {c for c, _, _ in took_fb} | {c for c, _, _ in kept_sportico}
    sport_only = [r for r in sheet
                  if r["league"] not in ("NFL", "NBA", "MLB", "NHL", "F1", "WNBA", "NWSL")
                  and r["league"] not in ("United States", "Mexico")
                  and r["team"] not in covered]
    # Plain ASCII on purpose: the Windows console is cp1252 and a warning glyph
    # here crashed the run, which is a silly way to lose a diff.
    print(f"\nWARNING: {len(sport_only)} European clubs are on ONE source only and "
          f"cannot be uplifted by the higher-of rule:")
    for r in sorted(sport_only, key=lambda r: -r["value_m"]):
        print(f"  {r['team']:<24} {r['league']:<10} {r['value_m']:>7,.0f} ({r['year']})")

    os.makedirs(os.path.dirname(REVIEW), exist_ok=True)
    with io.open(REVIEW, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["status", "league", "team", "sportico_usd_m", "fb_usd_m",
                    "chosen_usd_m", "chosen_source", "fb_yoy"])
        for canon, cur, r in took_fb:
            w.writerow(["switched", cur["league"], canon, cur["value_m"], r["usd_m"],
                        r["usd_m"], "Football Benchmark", r["yoy"]])
        for canon, cur, r in kept_sportico:
            w.writerow(["kept", cur["league"], canon, cur["value_m"], r["usd_m"],
                        cur["value_m"], "Sportico", r["yoy"]])
        for n in added:
            w.writerow(["new", n["league"], n["team"], "", n["value_m"],
                        n["value_m"], "Football Benchmark", n["fb"]["yoy"]])
        for r in sport_only:
            w.writerow(["single-source", r["league"], r["team"], r["value_m"], "",
                        r["value_m"], "Sportico", ""])
    print(f"\n-> {REVIEW}")

    if not a.write:
        print("\nDRY RUN. Nothing written. Re-run with --write once the diff reads right.")
        return 0

    from openpyxl import load_workbook
    wb = load_workbook(XLSX)
    ws = wb[SHEET]
    hdr = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    ix = {h: i + 1 for i, h in enumerate(hdr)}
    switch = {canon: r for canon, _, r in took_fb}
    changed = 0
    for row in range(2, ws.max_row + 1):
        team = ws.cell(row, ix["Team"]).value
        if team in switch:
            ws.cell(row, ix["Value ($M)"]).value = switch[team]["usd_m"]
            ws.cell(row, ix["Year"]).value = YEAR
            ws.cell(row, ix["Source"]).value = SOURCE
            changed += 1
    for n in added:
        ws.append([YEAR, n["team"], n["league"], n["value_m"], SOURCE])
    wb.save(XLSX)
    print(f"WROTE: {changed} rows switched to Football Benchmark, {len(added)} appended. "
          f"Sheet now holds {ws.max_row - 1} rows.")
    return 0


if __name__ == "__main__":
    sys.exit(main())

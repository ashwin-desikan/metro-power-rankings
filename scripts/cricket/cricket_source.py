#!/usr/bin/env python3
"""Data-source adapter for the cricket portal ETL and stagers.

Lets the pipeline read its sheets from EITHER the InternationalCricket.xlsx
workbook (unchanged legacy path) OR the Supabase mirror tables, behind one
function: open_source(spec).

  spec endswith '.xlsx'   -> openpyxl workbook (byte-identical legacy behaviour)
  spec == 'supabase'      -> Supabase tables via supabase-py (env: SUPABASE_URL + key)
  spec == 'local:<path>'  -> a local JSON dump of the mirror rows (parity tests)

Back-ends expose the slice of the openpyxl API the pipeline uses:
  wb[sheet].iter_rows(values_only=True) -> ORIGINAL header tuple first, then rows
  wb.close()

Row 0 carries the ORIGINAL workbook headers (e.g. 'Start Date'), so consumers
that look columns up by name (afghanistan_stage) work unchanged; consumers that
slice [1:] and index positionally (the ETL, cricsheet_stage) are unaffected.
Cells are text (dates normalised to 'YYYY-MM-DD'); every downstream consumer
passes them through iso()/int()/float()/str()/truthiness, so output is identical.
"""
from datetime import datetime, date

# sheet -> (supabase table, ordered DB columns matching the sheet left-to-right)
SHEET_TABLES = {
    "Matches": ("cricket_matches", [
        "format","start_date","end_date","match_no","team","t1","opponent","t2",
        "team_score","opp_score","winner","result","result_detail","tournament_series",
        "major","round","venue","venue_city","venue_country","host_country","source","ball_by_ball"]),
    "Other Internationals": ("cricket_other_internationals", [
        "date","match_no","team1","team2","team1_score","team2_score","winner","result",
        "result_detail","tournament_series","major","round","venue","city","host_country","source","ball_by_ball"]),
    "Number Ones": ("cricket_number_ones", [
        "month","test_no1","test_rating","odi_no1","odi_rating","t20i_no1","t20i_rating"]),
    "Test Rankings": ("cricket_test_rankings", ["month","rank","team","rating","weighted_matches"]),
    "ODI Rankings": ("cricket_odi_rankings", ["month","rank","team","rating","weighted_matches"]),
    "T20I Rankings": ("cricket_t20i_rankings", ["month","rank","team","rating","weighted_matches","table_name"]),
    "Honours": ("cricket_honours", [
        "country","cwc_titles","cwc_titles_years","cwc_ru","cwc_ru_years",
        "t20wc_titles","t20wc_titles_years","t20wc_ru","t20wc_ru_years",
        "ct_titles","ct_titles_years","ct_ru","ct_ru_years",
        "wtc_titles","wtc_titles_years","wtc_ru","wtc_ru_years",
        "asia_titles","asia_titles_years","asia_ru","asia_ru_years"]),
    "Series Trophies": ("cricket_series_trophies", [
        "trophy","contested_by","format","first_season","last_season","current_holder","series","notes"]),
}

# original workbook header labels, in sheet column order (row 0 of each sheet)
SHEET_HEADERS = {
    "Matches": ["Format","Start Date","End Date","Match #","Team","T1","Opponent","T2",
        "Team Score","Opp Score","Winner","Result","Result Detail","Tournament / Series",
        "Major","Round","Venue","Venue City","Venue Country","Host Country","Source","Ball-by-Ball"],
    "Other Internationals": ["Date","Match #","Team 1","Team 2","Team 1 Score","Team 2 Score",
        "Winner","Result","Result Detail","Tournament / Series","Major","Round","Venue","City",
        "Host Country","Source","Ball-by-Ball"],
    "Number Ones": ["Month","Test #1","Test Rating","ODI #1","ODI Rating","T20I #1","T20I Rating"],
    "Test Rankings": ["Month","Rank","Team","Rating","Weighted Matches"],
    "ODI Rankings": ["Month","Rank","Team","Rating","Weighted Matches"],
    "T20I Rankings": ["Month","Rank","Team","Rating","Weighted Matches","Table"],
    "Honours": ["Country","Cricket World Cup Titles","Years","Cricket World Cup Runner-up","Years",
        "T20 World Cup Titles","Years","T20 World Cup Runner-up","Years","Champions Trophy Titles","Years",
        "Champions Trophy Runner-up","Years","WTC Final Titles","Years","WTC Final Runner-up","Years",
        "Asia Cup Titles","Years","Asia Cup Runner-up","Years"],
    "Series Trophies": ["Trophy","Contested by","Format","First","Last","Current holder","Series","Notes"],
}


def normalize_cell(v):
    """openpyxl value -> text stored in Supabase. Dates become YYYY-MM-DD."""
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    return str(v)


def extract_sheet_rows(ws):
    """Read a live openpyxl sheet into normalised row tuples (drops all-empty rows)."""
    out = []
    it = ws.iter_rows(values_only=True)
    next(it, None)  # skip header
    for r in it:
        vals = [normalize_cell(x) for x in r]
        if all(x is None for x in vals):
            continue
        out.append(vals)
    return out


class _Cell:
    __slots__ = ("value",)
    def __init__(self, v):
        self.value = v


class _Sheet:
    """Mimics the slice of openpyxl's worksheet API the pipeline uses:
    iter_rows(min_row=, max_row=, values_only=). Row 1 is the original header."""
    def __init__(self, header, rows):
        self._header = list(header)
        self._w = len(self._header)
        self._rows = rows
    def _all(self):
        yield tuple(self._header)
        for r in self._rows:
            row = list(r) + [None] * (self._w - len(r))
            yield tuple(row[:self._w])
    def iter_rows(self, min_row=1, max_row=None, values_only=False):
        for i, row in enumerate(self._all(), start=1):
            if i < min_row:
                continue
            if max_row is not None and i > max_row:
                break
            yield tuple(row) if values_only else tuple(_Cell(v) for v in row)


class LocalWB:
    """Reads a JSON dump {sheet_name: [[cells...], ...]} of the mirror rows."""
    def __init__(self, path):
        import json
        with open(path, encoding="utf-8") as f:
            self._d = json.load(f)
    def __getitem__(self, name):
        return _Sheet(SHEET_HEADERS[name], self._d[name])
    def close(self):
        pass


class SupabaseWB:
    """Reads the mirror tables from Supabase via supabase-py."""
    def __init__(self):
        import os
        from supabase import create_client
        url = os.environ["SUPABASE_URL"]
        key = os.environ.get("SUPABASE_SERVICE_KEY") or os.environ["SUPABASE_KEY"]
        self._c = create_client(url, key)
    def __getitem__(self, name):
        table, cols = SHEET_TABLES[name]
        rows, start, page = [], 0, 1000
        sel = ",".join(cols)
        while True:
            r = self._c.table(table).select(sel).order("row_num").range(start, start + page - 1).execute()
            batch = r.data or []
            rows.extend([tuple(row.get(c) for c in cols) for row in batch])
            if len(batch) < page:
                break
            start += page
        return _Sheet(SHEET_HEADERS[name], rows)
    def close(self):
        pass


def open_source(spec):
    spec = str(spec)
    if spec.endswith(".xlsx"):
        from openpyxl import load_workbook
        return load_workbook(spec, read_only=True, data_only=True)
    if spec.startswith("local:"):
        return LocalWB(spec[len("local:"):])
    if spec == "supabase":
        return SupabaseWB()
    raise ValueError(f"Unknown source spec: {spec!r}")

#!/usr/bin/env python3
"""afghanistan_stage.py - stage Afghanistan men's internationals for the
InternationalCricket.xlsx Matches sheet, harvested from Wikipedia.

Cricsheet permanently withholds Afghanistan men's fixtures, so the monthly
Cricsheet stager can never see them. This tool fills that gap: it finds
Afghanistan's completed Test/ODI/T20I matches since the workbook's most recent
Afghanistan match, parses them from Wikipedia series pages via the MediaWiki
API, dedupes against the workbook on (Format, Start Date, {teams}), and writes
a paste-ready CSV in the same 22-column layout as cricsheet_stage.

It NEVER writes the workbook. Anything it cannot resolve (unknown team code,
unresolved venue city, ambiguous format) is left blank and flagged in the
summary for manual fix before pasting, exactly like the Cricsheet stager.
Unplayed / future fixtures (no result yet) are skipped.

Wikipedia match cards use {{Two-innings cricket match}} (Tests) and
{{Single-innings cricket match}} (limited-overs); format for the latter is read
from the enclosing section heading (e.g. "ODI series", "T20I series").

Usage:
  python afghanistan_stage.py --workbook InternationalCricket.xlsx --out delta.csv
  python afghanistan_stage.py --workbook ... --out ... --since 2026-05-01
"""
import argparse, csv, datetime, json, re, sys, urllib.parse, urllib.request

import openpyxl

API = "https://en.wikipedia.org/w/api.php"
UA = {"User-Agent": "MetroPowerRankings-AFG-stager/1.0 (cricket portal maintenance)"}
TODAY = datetime.date.today()

HEADER = [
    "Format","Start Date","End Date","Match #","Team","T1","Opponent","T2",
    "Team Score","Opp Score","Winner","Result","Result Detail",
    "Tournament / Series","Major","Round","Venue","Venue City",
    "Venue Country","Host Country","Source","Ball-by-Ball",
]

# ICC team codes used by {{cr|XXX}} / {{cr-rt|XXX}} on Wikipedia cricket cards.
CODE = {
    "AFG":"Afghanistan","IND":"India","PAK":"Pakistan","SL":"Sri Lanka","SRI":"Sri Lanka",
    "BAN":"Bangladesh","BDESH":"Bangladesh","ZIM":"Zimbabwe","IRE":"Ireland","ENG":"England",
    "AUS":"Australia","NZ":"New Zealand","RSA":"South Africa","SA":"South Africa",
    "WI":"West Indies","NED":"Netherlands","SCO":"Scotland","UAE":"United Arab Emirates",
    "NEP":"Nepal","OMA":"Oman","HK":"Hong Kong","USA":"United States","CAN":"Canada",
    "NAM":"Namibia","PNG":"Papua New Guinea","BER":"Bermuda","KEN":"Kenya",
}

def norm(s): return "".join(ch for ch in (s or "").lower() if ch.isalnum())

def api_get(params):
    params = dict(params); params.update(format="json", formatversion="2")
    url = API + "?" + urllib.parse.urlencode(params)
    req = urllib.request.Request(url, headers=UA)
    with urllib.request.urlopen(req, timeout=30) as r:
        return json.load(r)

def search_titles(years):
    titles = set()
    queries = []
    for y in years:
        queries += [f"Afghan cricket team in {y}", f"cricket team in Afghanistan in {y}",
                    f"Afghanistan tri-nation series {y}", f"Afghanistan against {y}"]
    for q in queries:
        try:
            d = api_get(dict(action="query", list="search", srsearch=q, srlimit=15))
        except Exception as e:
            print(f"  (search failed for {q!r}: {e})"); continue
        for hit in d.get("query", {}).get("search", []):
            t = hit["title"]
            tl = t.lower()
            if "cricket" in tl and "afghan" in tl and "women" not in tl and "records" not in tl:
                titles.add(t)
    return sorted(titles)

def get_wikitext(title):
    d = api_get(dict(action="query", prop="revisions", rvprop="content",
                     rvslots="main", titles=title))
    pages = d.get("query", {}).get("pages", [])
    if not pages or "revisions" not in pages[0]:
        return ""
    return pages[0]["revisions"][0]["slots"]["main"]["content"]

def top_templates(text):
    """Yield (start_index, name, body) for each top-level {{...}} template."""
    i = 0; n = len(text)
    while i < n:
        if text[i:i+2] == "{{":
            depth = 0; j = i
            while j < n:
                if text[j:j+2] == "{{": depth += 1; j += 2; continue
                if text[j:j+2] == "}}":
                    depth -= 1; j += 2
                    if depth == 0: break
                    continue
                j += 1
            body = text[i+2:j-2]
            yield i, body.split("|", 1)[0].strip(), body
            i = j
        else:
            i += 1

def split_params(body):
    parts = []; depth = 0; buf = ""; k = 0
    while k < len(body):
        if body[k:k+2] in ("{{","[["): depth += 1; buf += body[k:k+2]; k += 2; continue
        if body[k:k+2] in ("}}","]]"): depth -= 1; buf += body[k:k+2]; k += 2; continue
        if body[k] == "|" and depth == 0:
            parts.append(buf); buf = ""; k += 1; continue
        buf += body[k]; k += 1
    parts.append(buf)
    d = {}
    for p in parts[1:]:
        if "=" in p:
            key, val = p.split("=", 1)
            d[key.strip().lower()] = val.strip()
    return d

def resolve_team(raw):
    if not raw: return None
    m = re.search(r"\{\{\s*cr[a-z-]*\s*\|\s*([A-Za-z0-9]{2,6})", raw)
    if m:
        code = m.group(1).upper()
        return CODE.get(code, code)
    m = re.search(r"\[\[[^\]|]*\|([^\]]+)\]\]", raw) or re.search(r"\[\[([^\]]+)\]\]", raw)
    if m:
        return m.group(1).replace(" cricket team", "").strip()
    v = re.sub(r"[\[\]{}]", "", raw).strip()
    return v or None

def strip_markup(s):
    s = re.sub(r"\{\{[^{}]*\}\}", "", s or "")
    s = re.sub(r"\[\[[^\]|]*\|([^\]]+)\]\]", r"\1", s)
    s = re.sub(r"\[\[([^\]]+)\]\]", r"\1", s)
    s = re.sub(r"<[^>]+>", "", s)
    return re.sub(r"\s+", " ", s).strip()

def clean_score(s):
    s = strip_markup(s)
    s = re.sub(r"\([^)]*\)", "", s)          # drop overs parentheticals
    s = s.replace("&nbsp;", " ")
    return re.sub(r"\s+", " ", s).strip()

MONTHS = {m:i for i,m in enumerate(
    ["january","february","march","april","may","june","july","august",
     "september","october","november","december"], 1)}

def parse_date(s, want_start=True):
    s = strip_markup(s).replace("–","-").replace("—","-")
    # range "6-8 June 2026" -> start=6 June, end=8 June
    m = re.search(r"(\d{1,2})\s*-\s*(\d{1,2})\s+([A-Za-z]+)\s+(\d{4})", s)
    if m:
        day = int(m.group(1) if want_start else m.group(2))
        mon = MONTHS.get(m.group(3).lower())
        if mon: return datetime.date(int(m.group(4)), mon, day)
    m = re.search(r"(\d{1,2})\s+([A-Za-z]+)\s+(\d{4})", s)
    if m:
        mon = MONTHS.get(m.group(2).lower())
        if mon: return datetime.date(int(m.group(3)), mon, int(m.group(1)))
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", s)
    if m: return datetime.date(*map(int, m.groups()))
    return None

def fmt_from_heading(head):
    h = (head or "").lower()
    if "test" in h: return "Test"
    if "t20" in h or "twenty20" in h: return "T20I"
    if "odi" in h or "one day" in h or "one-day" in h: return "ODI"
    return None

def test_scores(p):
    """Compose Test scores 'inns1 & inns2' per team from innings params."""
    def side(n):
        parts = []
        for k in (1, 2):
            v = p.get(f"score-team{n}-inns{k}") or p.get(f"runs-team{n}-inns{k}")
            if v: parts.append(clean_score(v))
        return " & ".join(parts)
    return side(1), side(2)

def extract_matches(text, page_title):
    heads = [(m.start(), fmt_from_heading(m.group(1)))
             for m in re.finditer(r"^=+\s*(.+?)\s*=+\s*$", text, re.M)]
    def section_fmt(pos):
        f = None
        for hp, hf in heads:
            if hp > pos: break
            if hf: f = hf
        return f
    out = []
    for idx, name, body in top_templates(text):
        nl = name.lower()
        if "cricket match" not in nl:
            continue
        p = split_params(body)
        if "team1" not in p or "team2" not in p:
            continue
        t1 = resolve_team(p.get("team1")); t2 = resolve_team(p.get("team2"))
        if not t1 or not t2 or "Afghanistan" not in (t1, t2):
            continue
        if "two-innings" in nl:
            fmt = "Test"; s1, s2 = test_scores(p)
        else:
            fmt = section_fmt(idx)
            s1 = clean_score(p.get("score1", "")); s2 = clean_score(p.get("score2", ""))
        out.append(dict(fmt=fmt, date=parse_date(p.get("date", ""), True),
            t1=t1, t2=t2, score1=s1, score2=s2,
            result=strip_markup(p.get("result", "")),
            venue=strip_markup(p.get("venue", "")), page=page_title))
    return out

def is_played(m):
    if m["date"] and m["date"] > TODAY: return False
    return bool(m["result"].strip() or m["score1"].strip() or m["score2"].strip())

def load_workbook_index(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Matches"]
    H = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    ix = {n: H.index(n) for n in ["Format","Start Date","Team","Opponent",
          "Venue","Venue City","Venue Country","Host Country"]}
    def sd(v):
        if isinstance(v, (datetime.date, datetime.datetime)):
            return datetime.date(v.year, v.month, v.day)
        try: return datetime.date(*map(int, str(v)[:10].split("-")))
        except Exception: return None
    existing = set(); last_afg = None; venue_by_city = {}; teams = set()
    for r in ws.iter_rows(min_row=2, values_only=True):
        f = r[ix["Format"]]; d = sd(r[ix["Start Date"]])
        t, o = r[ix["Team"]], r[ix["Opponent"]]
        if f in ("Test","ODI","T20I") and d and t and o:
            existing.add((f, d.isoformat(), frozenset((str(t), str(o)))))
            if "Afghanistan" in (str(t), str(o)) and (last_afg is None or d > last_afg):
                last_afg = d
        if t: teams.add(str(t))
        city = r[ix["Venue City"]]
        if city:
            key = norm(str(city))
            if key not in venue_by_city:
                venue_by_city[key] = (str(r[ix["Venue"]] or ""), str(city),
                    str(r[ix["Venue Country"]] or ""), str(r[ix["Host Country"]] or ""))
    return dict(existing=existing, last_afg=last_afg,
                venue_by_city=venue_by_city, teams=teams)

def result_fields(result, t1, t2):
    r = (result or "").lower()
    if "no result" in r: return ("No result", "", "No result")
    if "abandon" in r:   return ("No result", "", "Match abandoned")
    if "draw" in r:      return ("Draw", "", "Match drawn")
    if "tie" in r or "tied" in r: return ("Tie", "", "Match tied")
    winner = ""
    for t in (t1, t2):
        if t and t.lower() in r: winner = t; break
    return ("Win", winner, (result or "").strip())

def venue_fields(venue, idx):
    city = venue.rsplit(",", 1)[-1].strip() if "," in venue else ""
    hit = idx["venue_by_city"].get(norm(city))
    if hit:
        return hit[0], hit[1], hit[2], hit[3], []
    return venue, city, "", "", ([f"venue city {city!r} not in workbook"] if city else ["venue unresolved"])

def emit_rows(m, idx):
    flags = []
    if m["fmt"] not in ("Test","ODI","T20I"): flags.append("format unresolved")
    if m["date"] is None: flags.append("date unparsed")
    for t in (m["t1"], m["t2"]):
        if t not in idx["teams"]: flags.append(f"team {t!r} not in workbook")
    res, winner, detail = result_fields(m["result"], m["t1"], m["t2"])
    if res == "Win" and not winner: flags.append("winner unresolved")
    venue, city, vc, hc, vflags = venue_fields(m["venue"], idx); flags += vflags
    sd = m["date"].isoformat() if m["date"] else ""
    end = sd if (m["date"] and m["fmt"] == "Test") else ""
    A = [m["fmt"], sd, end, "", m["t1"], "T1", m["t2"], "T2", m["score1"], m["score2"],
         winner, res, detail, "", "", "", venue, city, vc, hc, "Wikipedia", "No"]
    B = [m["fmt"], sd, end, "", m["t2"], "T2", m["t1"], "T1", m["score2"], m["score1"],
         winner, res, detail, "", "", "", venue, city, vc, hc, "Wikipedia", "No"]
    return [A, B], flags

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--workbook", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--since", default=None, help="YYYY-MM-DD; default = day after last AFG match")
    args = ap.parse_args()

    idx = load_workbook_index(args.workbook)
    since = (datetime.date(*map(int, args.since.split("-"))) if args.since
             else (idx["last_afg"] + datetime.timedelta(days=1) if idx["last_afg"] else datetime.date(2000,1,1)))
    print(f"Workbook last Afghanistan match: {idx['last_afg']}. Harvesting since {since}.")
    years = sorted({since.year, TODAY.year})
    titles = search_titles(years)
    print(f"Candidate Wikipedia pages: {len(titles)}")
    for t in titles: print(f"  - {t}")

    parsed = []
    for t in titles:
        try: wt = get_wikitext(t)
        except Exception as e: print(f"  (fetch failed {t}: {e})"); continue
        for m in extract_matches(wt, t):
            if m["date"] and m["date"] >= since and is_played(m):
                parsed.append(m)

    seen = set(); staged = []; flagged = []
    for m in sorted(parsed, key=lambda x: x["date"]):
        key = (m["fmt"], m["date"].isoformat(), frozenset((m["t1"], m["t2"])))
        if key in idx["existing"] or key in seen: continue
        seen.add(key)
        rows, flags = emit_rows(m, idx)
        staged.append((m, rows))
        if flags: flagged.append((m, flags))

    with open(args.out, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f); w.writerow(HEADER)
        for _, rows in staged:
            for row in rows: w.writerow(row)

    print(f"\nStaged {len(staged)} new Afghanistan internationals -> {len(staged)*2} rows")
    for m, _ in staged:
        print(f"  {m['date']} {m['fmt'] or '??':4} {m['t1']} v {m['t2']}  "
              f"[{m['score1']} / {m['score2']}]  {m['result']}")
    if flagged:
        print("\nREVIEW BEFORE PASTING:")
        for m, flags in flagged:
            print(f"  {m['date']} {m['fmt']} {m['t1']} v {m['t2']}: " + "; ".join(flags))
    print(f"\nWrote {args.out}")

if __name__ == "__main__":
    sys.exit(main())

#!/usr/bin/env python3
"""
cricsheet_stage.py - stage newly-added Cricsheet internationals as paste-ready
rows for the InternationalCricket.xlsx 'Matches' sheet.

Reads a Cricsheet JSON zip (e.g. recently_added_30_male_json.zip), keeps men's
Test/ODI/T20I internationals, computes innings scorelines, and emits two mirrored
rows per match (one per team's perspective) matching the 22-column Matches schema.
Dedupes against the existing workbook on (Format, Start Date, {teams}). Venue
country / host country and team-name spellings are inherited from the workbook
itself so staged rows match house vocabulary; anything it cannot resolve is left
blank and listed in the summary for manual fill before pasting.

This NEVER writes to the workbook. Output is a CSV the user reviews and pastes.

Usage:
  python cricsheet_stage.py --zip recent.zip --workbook InternationalCricket.xlsx --out delta.csv
"""
import argparse
import collections
import csv
import json
import sys
import zipfile

import openpyxl
from cricket_source import open_source

HEADER = [
    "Format", "Start Date", "End Date", "Match #", "Team", "T1", "Opponent", "T2",
    "Team Score", "Opp Score", "Winner", "Result", "Result Detail",
    "Tournament / Series", "Major", "Round", "Venue", "Venue City",
    "Venue Country", "Host Country", "Source", "Ball-by-Ball",
]

# Cricsheet match_type -> workbook Format. Internationals only; non-official
# IT20s and all domestic/multi-day-non-Test types are intentionally excluded.
TYPE_MAP = {"Test": "Test", "ODI": "ODI", "T20": "T20I"}

DL_METHODS = {"D/L", "DLS", "VJD", "Duckworth-Lewis", "Duckworth-Lewis-Stern"}

# Cricsheet spelling -> workbook spelling, where they diverge. Applied to both
# the dedup key and the emitted rows so a name difference never stages a match
# the workbook already has. Keyed by the lower-cased Cricsheet name. Extend as
# new mismatches surface (the stager flags any team it cannot resolve).
ALIAS = {
    "united states of america": "United States",
}


def norm_team(s):
    return "".join(ch for ch in (s or "").lower() if ch.isalnum())


def innings_runs_wkts(inn):
    runs = 0
    wkts = 0
    for ov in inn.get("overs", []):
        for d in ov.get("deliveries", []):
            runs += (d.get("runs") or {}).get("total", 0)
            wkts += len(d.get("wickets") or [])
    return runs, wkts


def fmt_score(innings_for_team):
    parts = []
    for runs, wkts in innings_for_team:
        parts.append(str(runs) if wkts >= 10 else f"{runs}/{wkts}")
    return " & ".join(parts)


def resolve_outcome(outcome, spell):
    """Return (Result, Winner, Result Detail), winner spelled per workbook."""
    res = (outcome.get("result") or "").lower()
    if res == "draw":
        return "Draw", "", "Match drawn"
    if res == "tie":
        return "Tie", "", "Match tied"
    if res in ("no result", "noresult"):
        return "No result", "", "No result"
    winner = spell(outcome.get("winner"))
    by = outcome.get("by") or {}
    method = outcome.get("method")
    tail = " (D/L method)" if method in DL_METHODS else ""
    if "runs" in by:
        n = by["runs"]
        detail = f"{winner} won by {n} run{'' if n == 1 else 's'}{tail}"
    elif "wickets" in by:
        n = by["wickets"]
        detail = f"{winner} won by {n} wicket{'' if n == 1 else 's'}{tail}"
    elif outcome.get("eliminator"):
        detail = f"{winner} won (Super Over)"
    else:
        detail = f"{winner} won{tail}".strip()
    return "Win", winner, detail


def load_workbook_context(path):
    wb = open_source(path)
    ws = wb["Matches"]
    it = ws.iter_rows(values_only=True)
    next(it)  # header
    existing = set()           # (Format, iso-date, frozenset(norm teams))
    venue_country = collections.defaultdict(collections.Counter)
    venue_host = collections.defaultdict(collections.Counter)
    city_country = collections.defaultdict(collections.Counter)
    team_spell = {}            # norm -> workbook spelling
    for r in it:
        fmt, sdate, team, opp = r[0], r[1], r[4], r[6]
        venue, vcity, vcountry, host = r[16], r[17], r[18], r[19]
        if team:
            team_spell.setdefault(norm_team(team), team)
        if opp:
            team_spell.setdefault(norm_team(opp), opp)
        iso = sdate.date().isoformat() if hasattr(sdate, "date") else str(sdate or "")
        if fmt and iso and team and opp:
            existing.add((fmt, iso, frozenset((norm_team(team), norm_team(opp)))))
        if venue and vcountry:
            venue_country[venue][vcountry] += 1
        if venue and host:
            venue_host[venue][host] += 1
        if vcity and vcountry:
            city_country[vcity][vcountry] += 1
    wb.close()

    def most(counter_map, key):
        c = counter_map.get(key)
        return c.most_common(1)[0][0] if c else ""

    def spell(n):
        if not n:
            return ""
        a = ALIAS.get(n.strip().lower())
        if a:
            return a
        return team_spell.get(norm_team(n), n)

    return {
        "existing": existing,
        "spell": spell,
        "venue_country": lambda v: most(venue_country, v),
        "venue_host": lambda v: most(venue_host, v),
        "city_country": lambda c: most(city_country, c),
        "known_team": lambda n: norm_team(spell(n)) in team_spell,
    }


def iter_matches(zip_path):
    with zipfile.ZipFile(zip_path) as z:
        for name in z.namelist():
            if not name.endswith(".json") or name in ("README.txt",):
                continue
            try:
                with z.open(name) as fh:
                    yield name[:-5].split("/")[-1], json.load(fh)
            except Exception as e:  # noqa: BLE001
                print(f"  ! skip {name}: {e!r}", file=sys.stderr)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--zip", required=True)
    ap.add_argument("--workbook", required=True)
    ap.add_argument("--out")
    ap.add_argument("--to-supabase", action="store_true",
                    help="insert staged rows straight into cricket_matches instead of writing a CSV")
    args = ap.parse_args()

    ctx = load_workbook_context(args.workbook)
    spell = ctx["spell"]

    staged = []          # list of match dicts (for summary + rows)
    flags = []           # human-readable review notes
    seen_keys = set()
    counts = collections.Counter()

    for mid, j in iter_matches(args.zip):
        info = j.get("info", {})
        if info.get("team_type") != "international":
            continue
        fmt = TYPE_MAP.get(info.get("match_type"))
        if not fmt:
            continue
        teams = info.get("teams") or []
        if len(teams) != 2:
            continue
        dates = info.get("dates") or []
        iso = dates[0] if dates else ""
        key = (fmt, iso, frozenset(norm_team(spell(t)) for t in teams))
        if key in ctx["existing"] or key in seen_keys:
            continue
        seen_keys.add(key)

        # innings (skip super overs for the headline scoreline)
        per_team = collections.defaultdict(list)
        first_bat = None
        for inn in j.get("innings", []):
            if inn.get("super_over"):
                continue
            t = inn.get("team")
            if first_bat is None:
                first_bat = t
            per_team[t].append(innings_runs_wkts(inn))

        outcome = info.get("outcome", {}) or {}
        result, winner, detail = resolve_outcome(outcome, spell)

        venue = info.get("venue") or ""
        city = info.get("city") or ""
        vcountry = ctx["venue_country"](venue) or ctx["city_country"](city)
        host = ctx["venue_host"](venue) or vcountry
        event = (info.get("event") or {}).get("name") or ""
        stage = (info.get("event") or {}).get("stage") or ""
        end_date = dates[-1] if (fmt == "Test" and len(dates) > 1) else ""

        for t in teams:
            unknown = not ctx["known_team"](t)
            if unknown:
                flags.append(f"{iso} {fmt}: team '{t}' not in workbook (check spelling)")
        if not vcountry:
            flags.append(f"{iso} {fmt} @ {venue or city or '?'}: venue/host country unmapped")

        staged.append({
            "mid": mid, "fmt": fmt, "iso": iso, "end": end_date,
            "teams": teams, "first_bat": first_bat, "per_team": per_team,
            "winner": winner, "result": result, "detail": detail,
            "event": event, "stage": stage,
            "venue": venue, "city": city, "vcountry": vcountry, "host": host,
        })
        counts[fmt] += 1

    staged.sort(key=lambda m: (m["iso"], m["mid"]))

    rows = []
    for m in staged:
        for team in m["teams"]:
            opp = m["teams"][1] if team == m["teams"][0] else m["teams"][0]
            t1 = "T1" if m["first_bat"] == team else "T2"
            t2 = "T2" if t1 == "T1" else "T1"
            rows.append([
                m["fmt"],
                m["iso"],
                m["end"],
                "",                                   # Match #
                spell(team),
                t1,
                spell(opp),
                t2,
                fmt_score(m["per_team"].get(team, [])),
                fmt_score(m["per_team"].get(opp, [])),
                m["winner"],
                m["result"],
                m["detail"],
                m["event"],
                "",                                   # Major (curation)
                m["stage"],
                m["venue"],
                m["city"],
                m["vcountry"],
                m["host"],
                "Cricsheet",
                "Yes",
            ])

    if args.to_supabase:
        import cricket_store
        n = cricket_store.append_matches(rows)
        print(f"Inserted {n} rows into cricket_matches (Supabase)")
    elif args.out:
        with open(args.out, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(HEADER)
            w.writerows(rows)
    else:
        sys.exit("provide --out <csv> or --to-supabase")

    print(f"Staged {len(staged)} new internationals ({dict(counts)}) -> {len(rows)} rows")
    if staged:
        print(f"Date range: {staged[0]['iso']} .. {staged[-1]['iso']}")
        print("Matches:")
        for m in staged:
            print(f"  {m['iso']} {m['fmt']:4} {m['teams'][0]} v {m['teams'][1]}  [{m['detail']}]")
    if flags:
        print("\nREVIEW BEFORE PASTING:")
        for fl in dict.fromkeys(flags):
            print(f"  - {fl}")
    print(f"\nWrote {args.out}")


if __name__ == "__main__":
    main()

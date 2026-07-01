#!/usr/bin/env python3
"""Monthly ICC ranking recompute driver.

Feeds the InternationalCricket.xlsx Matches spine through icc_engine, produces
month-end snapshots, and APPENDS the newly-completed month(s) to the four
ranking sheets (Number Ones, Test/ODI/T20I Rankings). It never rewrites
existing months (append-only policy, user decision 2026-07-01), so historical
published numbers are preserved; the small stale-gap discontinuity at the
boundary is accepted.

Validation gate: before writing, it recomputes the most recent existing month
and checks the well-connected Test/ODI teams reproduce within a tolerance. If
they diverge beyond tolerance it aborts and prints the diff.

T20I keeps the curated main/associate split: membership is inherited from the
last existing month's 'main' tag; each table is ranked independently. Retagged
closed-pool teams (per data/cricket/retag_t20i_tables.py) stay associate.

Usage:
  python build_icc_rankings.py            # dry-run: print the new month tables
  python build_icc_rankings.py --write    # append to the workbook (backup first)
"""
import sys, os, argparse, datetime, collections, shutil
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
import icc_engine as E

MASTER = r"C:\Users\ashwi\OneDrive\Excel Files\InternationalCricket.xlsx"
FMT_SHEET = {"Test": "Test Rankings", "ODI": "ODI Rankings", "T20I": "T20I Rankings"}
MIN_N = 8
# Cricsheet/long-form spellings -> canonical Matches-sheet spelling.
CANON = {"United States of America": "United States"}
def canon(t): return CANON.get(t, t)

def dt(v):
    if isinstance(v, (datetime.date, datetime.datetime)):
        return datetime.date(v.year, v.month, v.day)
    if v is None: return None
    s = str(v)[:10]
    try: return datetime.date(*map(int, s.split("-")))
    except Exception: return None

def month_end(y, m):
    ny, nm = (y, m + 1) if m < 12 else (y + 1, 1)
    return datetime.date(ny, nm, 1) - datetime.timedelta(days=1)

def next_month(ym):
    y, m = map(int, ym.split("-"))
    return f"{y+1}-01" if m == 12 else f"{y}-{m+1:02d}"

def load_matches(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Matches"]
    H = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    ix = {n: H.index(n) for n in ["Format","Start Date","Team","Opponent","Winner","Result","Tournament / Series"]}
    seen = set(); out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        fmt = r[ix["Format"]]
        if fmt not in ("Test","ODI","T20I"): continue
        sd = dt(r[ix["Start Date"]])
        if sd is None: continue
        team = r[ix["Team"]]; opp = r[ix["Opponent"]]
        if not team or not opp: continue
        team, opp = canon(str(team)), canon(str(opp))
        key = (fmt, sd, frozenset((team, opp)))
        if key in seen: continue
        seen.add(key)
        out.append(dict(fmt=fmt, sd=sd, team=team, opp=opp,
            winner=canon(str(r[ix["Winner"]]).strip()) if r[ix["Winner"]] else "",
            result=(str(r[ix["Result"]]).strip() if r[ix["Result"]] else "")))
    return out

def build_limited(matches, fmt):
    items = []
    for m in matches:
        if m["fmt"] != fmt: continue
        if m["result"].lower() in ("no result", "abandoned"): continue
        items.append(dict(date=m["sd"], t1=m["team"], t2=m["opp"], winner=m["winner"], result=m["result"]))
    items.sort(key=lambda x: x["date"])
    return items

def build_test_series(matches):
    tests = [m for m in matches if m["fmt"] == "Test"]
    by_pair = collections.defaultdict(list)
    for m in tests:
        by_pair[frozenset((m["team"], m["opp"]))].append(m)
    series = []
    for pair, ms in by_pair.items():
        ms.sort(key=lambda x: x["sd"])
        cur = []
        for m in ms:
            if cur and (m["sd"] - cur[-1]["sd"]).days > 45:
                series.append(_mk_series(cur)); cur = []
            cur.append(m)
        if cur: series.append(_mk_series(cur))
    series.sort(key=lambda s: s["end"])
    return series

def _mk_series(ms):
    A, B = sorted(frozenset((ms[0]["team"], ms[0]["opp"])))
    wA = sum(1 for m in ms if m["winner"] == A)
    wB = sum(1 for m in ms if m["winner"] == B)
    dr = sum(1 for m in ms if m["winner"] not in (A, B))
    return dict(start=ms[0]["sd"], end=ms[-1]["sd"], A=A, B=B, wA=wA, wB=wB, dr=dr, n=len(ms))

def compute(matches):
    return {"Test": E.run(build_test_series(matches), True),
            "ODI": E.run(build_limited(matches, "ODI"), False),
            "T20I": E.run(build_limited(matches, "T20I"), False)}

def table(L, fmt, ym):
    y, m = map(int, ym.split("-")); me = month_end(y, m)
    rows = []
    for team, led in L[fmt].items():
        r, n = led.rating(me)
        if r is not None and n >= MIN_N:
            rows.append((team, round(r, 1), round(n, 1)))
    rows.sort(key=lambda x: (-x[1], x[0]))
    return rows

def last_month(ws):
    last = None
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0]:
            s = str(r[0])
            if last is None or s > last: last = s
    return last

def main_set_from_sheet(ws):
    """Teams tagged 'main' in the most recent month (canonicalized)."""
    lm = last_month(ws)
    mains = set()
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0] and str(r[0]) == lm and len(r) >= 6 and r[5] and str(r[5]).strip() == "main":
            mains.add(canon(str(r[2]).strip()))
    return mains, lm

def validate(L, wb, tol=0.35):
    """Recompute the last existing month; check connected Test/ODI teams match."""
    problems = []
    for fmt in ("Test", "ODI"):
        ws = wb[FMT_SHEET[fmt]]
        lm = last_month(ws)
        tgt = {}
        for r in ws.iter_rows(min_row=2, values_only=True):
            if r[0] and str(r[0]) == lm:
                tgt[canon(str(r[2]))] = float(r[3])
        got = {t: rt for t, rt, n in table(L, fmt, lm)}
        big = [(t, got[t], tgt[t]) for t in got if t in tgt and abs(got[t]-tgt[t]) > tol]
        if big:
            problems.append((fmt, lm, big))
    return problems

def build_month_rows(L, ym, main_set):
    """Return {sheet: [rows]} for the month, plus the Number Ones row."""
    out = {}
    for fmt in ("Test", "ODI"):
        rows = table(L, fmt, ym)
        out[FMT_SHEET[fmt]] = [[ym, i+1, t, rt, n] for i, (t, rt, n) in enumerate(rows)]
    # T20I split
    t20 = table(L, "T20I", ym)
    main = [(t, rt, n) for t, rt, n in t20 if t in main_set]
    asso = [(t, rt, n) for t, rt, n in t20 if t not in main_set]
    trows = []
    for i, (t, rt, n) in enumerate(main): trows.append([ym, i+1, t, rt, n, "main"])
    for i, (t, rt, n) in enumerate(asso): trows.append([ym, i+1, t, rt, n, "associate"])
    out[FMT_SHEET["T20I"]] = trows
    # Number Ones: top of each format (T20I top overall = main #1)
    def top(fmt):
        r = table(L, fmt, ym)
        return (r[0][0], r[0][1]) if r else ("", "")
    t1, tr = top("Test"); o1, orr = top("ODI"); x1, xr = top("T20I")
    no_row = [ym, t1, tr, o1, orr, x1, xr]
    return out, no_row

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--write", action="store_true")
    ap.add_argument("--workbook", default=MASTER)
    args = ap.parse_args()

    matches = load_matches(args.workbook)
    L = compute(matches)

    # read-only pass for structure
    wbro = openpyxl.load_workbook(args.workbook, read_only=True, data_only=True)
    problems = validate(L, wbro)
    if problems:
        print("VALIDATION FAILED (connected Test/ODI teams diverge beyond tol):")
        for fmt, lm, big in problems:
            for t, g, x in big: print(f"  {fmt} {lm} {t}: got {g} stored {x}")
        print("Aborting; no write."); return 2
    print("Validation OK: connected Test/ODI teams reproduce within tolerance.")

    ws_t20 = wbro[FMT_SHEET["T20I"]]
    main_set, lm_t20 = main_set_from_sheet(ws_t20)
    lm = last_month(wbro["Number Ones"])
    # months to append: from next(lm) through last complete month
    today = datetime.date.today()
    last_complete = month_end(today.year, today.month) if False else None
    prev = today.replace(day=1) - datetime.timedelta(days=1)
    last_complete = f"{prev.year}-{prev.month:02d}"
    todo = []
    cur = next_month(lm)
    while cur <= last_complete:
        todo.append(cur); cur = next_month(cur)
    print(f"Last existing month: {lm}. Appending: {todo or 'none'}")
    if not todo:
        print("Nothing to append."); return 0

    all_no = []
    all_fmt = {s: [] for s in FMT_SHEET.values()}
    for ym in todo:
        rows, no_row = build_month_rows(L, ym, main_set)
        all_no.append(no_row)
        for s, rr in rows.items(): all_fmt[s].extend(rr)
        # print summary
        print(f"\n== {ym} ==  NumberOnes: Test {no_row[1]} {no_row[2]} | ODI {no_row[3]} {no_row[4]} | T20I {no_row[5]} {no_row[6]}")
        for fmt in ("Test","ODI"):
            top5 = rows[FMT_SHEET[fmt]][:5]
            print(f"  {fmt} top5: " + ", ".join(f"{r[2]} {r[3]}" for r in top5))
        tm = [r for r in rows[FMT_SHEET["T20I"]] if r[5]=="main"][:5]
        print("  T20I main top5: " + ", ".join(f"{r[2]} {r[3]}" for r in tm))

    if not args.write:
        print("\nDry-run only. Re-run with --write to append.")
        return 0

    # ---- write: backup, append, verify others unchanged, overwrite ----
    stamp = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
    backup = args.workbook.replace(".xlsx", f".backup-rank-{stamp}.xlsx")
    tmpout = os.path.join(os.environ.get("TEMP","."), "IC-rank-append.xlsx")
    readcopy = os.path.join(os.environ.get("TEMP","."), "IC-rank-src.xlsx")
    shutil.copyfile(args.workbook, readcopy)
    wb = openpyxl.load_workbook(readcopy)
    ranking_sheets = set(FMT_SHEET.values()) | {"Number Ones"}
    snapshot = {n: [[c.value for c in row] for row in wb[n].iter_rows()]
                for n in wb.sheetnames if n not in ranking_sheets}
    counts_before = {n: wb[n].max_row for n in ranking_sheets}
    for s, rr in all_fmt.items():
        ws = wb[s]
        for row in rr: ws.append(row)
    wsn = wb["Number Ones"]
    for row in all_no: wsn.append(row)
    wb.save(tmpout)

    wb2 = openpyxl.load_workbook(tmpout)
    for n, snap in snapshot.items():
        if [[c.value for c in row] for row in wb2[n].iter_rows()] != snap:
            print(f"ABORT: non-ranking sheet changed: {n}"); return 3
    for n in ranking_sheets:
        exp = counts_before[n] + (len(all_no) if n=="Number Ones" else len(all_fmt[n]))
        if wb2[n].max_row != exp:
            print(f"ABORT: {n} rowcount {wb2[n].max_row} != expected {exp}"); return 3
    shutil.copyfile(args.workbook, backup)
    shutil.copyfile(tmpout, args.workbook)
    print(f"\nBackup: {backup}\nWrote ranking rows to {args.workbook}\nOK_DONE")
    return 0

if __name__ == "__main__":
    sys.exit(main())

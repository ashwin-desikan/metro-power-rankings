"""Global Overture dump of admin levels 0, 1, 2 across every country.

Produces a single editorial workbook listing every country, region/state,
and county/district from Overture's global division_area parquet. Designed
for the metro project's manual mapping pass — sort by Country, scan the
Region rows, then drill into Counties for each region.

Levels filtered:
  0 = country     (subtype = "country")
  1 = region      (subtype = "region", e.g. US states, JP prefectures)
  2 = county      (subtype = "county", e.g. US counties, UK boroughs)

Levels deliberately excluded:
  3 = locality, 4 = neighborhood — these explode row counts (~1M+) without
  helping the editorial mapping pass. Pull a separate dump if needed.

Usage:
  cd "C:\\Users\\ashwi\\Desktop\\Projects\\Metro Area Project"
  python scripts\\dump-overture-admin-012.py

Override defaults via env:
  OVERTURE_DIVISION_AREA - path to the global parquet
  OVERTURE_ADMIN012_OUT  - output xlsx path

Output layout (single sheet "Overture 012"):
  Country (ISO 3166-1)  -- e.g. US, GB, DE
  Country Name          -- joined from this dump's own level-0 rows
  Admin Level           -- 0 / 1 / 2
  Subtype               -- country / region / county
  Primary Name          -- the entity's English / primary label
  Region (ISO 3166-2)   -- the entity's own region code (level 1) or its
                           parent region (level 2). Null for level 0.
  Class                 -- Overture's class field (usually "land")
  Common Names (JSON)   -- localized names as [[lang, label], ...]
  GERS ID               -- the entity's GERS identifier
  Division ID           -- parent division reference (or self for top-level)
  Is Land               -- TRUE / FALSE
  Is Territorial        -- TRUE / FALSE
  BBox xmin/ymin/xmax/ymax

Rows are sorted by Country ISO ascending, then by Admin Level ascending
(0, 1, 2), then by Region (so all entries under US-CA cluster together),
then by Primary Name. That order matches the manual-mapping workflow:
fix the country header, then walk down through each region's counties.

Memory profile: a single in-memory list of ~35-60K rows. Comfortably
under 1 GB peak RSS even on the 5.8 GB source parquet, since the parquet
is streamed in batches and only matching rows are retained.
"""
from __future__ import annotations

import json
import os
import sys
import time
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


SOURCE_PARQUET = os.environ.get(
    "OVERTURE_DIVISION_AREA",
    r"C:\Users\ashwi\Desktop\Projects\MapData\global-division-area.parquet",
)
OUT_PATH = Path(os.environ.get(
    "OVERTURE_ADMIN012_OUT",
    r"Overture-Per-Country-Raw\overture-admin-levels-012-all.xlsx",
))


HEADER = [
    "Country (ISO)",
    "Country Name",
    "Admin Level",
    "Subtype",
    "Primary Name",
    "Region (ISO 3166-2)",
    "Class",
    "Common Names (JSON)",
    "GERS ID",
    "Division ID",
    "Is Land",
    "Is Territorial",
    "BBox xmin",
    "BBox ymin",
    "BBox xmax",
    "BBox ymax",
]

# Subtypes we keep, mapped to the numeric admin level we surface in the
# output. Anything else (locality, neighborhood) is dropped on the way in.
KEEP_SUBTYPES = {
    "country": 0,
    "region": 1,
    "county": 2,
}


def _common_names_json(names_dict):
    """Pack the Overture `names` struct's `common` map into a stable JSON
    string identical to dump-overture-country.py. Sorted by language code
    so diffs across runs stay clean."""
    if not isinstance(names_dict, dict):
        return None
    common = names_dict.get("common")
    if not common:
        return None
    if isinstance(common, dict):
        items = sorted(common.items(), key=lambda kv: kv[0])
        return json.dumps([[k, v] for k, v in items], ensure_ascii=False)
    if isinstance(common, list):
        try:
            items = sorted(common, key=lambda kv: kv[0])
            return json.dumps([[k, v] for k, v in items], ensure_ascii=False)
        except Exception:
            return json.dumps(common, ensure_ascii=False)
    return None


def _bbox_to_xyxy(bbox):
    if not isinstance(bbox, dict):
        return (None, None, None, None)
    return (
        bbox.get("xmin"),
        bbox.get("ymin"),
        bbox.get("xmax"),
        bbox.get("ymax"),
    )


def _bool_str(v):
    if v is None:
        return None
    return "TRUE" if bool(v) else "FALSE"


def collect_admin_012(parquet_path: str):
    """Single pass over the parquet. Returns (rows, country_name_by_iso).

    `rows` is a list of tuples in HEADER order MINUS the Country Name slot
    (which we attach on the second pass after the country-name lookup is
    fully populated). `country_name_by_iso` maps ISO codes harvested from
    the level-0 rows to their English/primary labels.
    """
    print(f"[1/2] Scanning {parquet_path}")
    t0 = time.time()
    import pyarrow.parquet as pq

    pf = pq.ParquetFile(parquet_path)
    schema_cols = set(pf.schema_arrow.names)

    cols = [
        "id",
        "country",
        "region",
        "subtype",
        "names",
        "class",
        "is_land",
        "is_territorial",
        "bbox",
    ]
    # Some Overture releases use "primary_division_id" instead of
    # "division_id"; keep the same compatibility shim as the per-country
    # dump script.
    if "division_id" in schema_cols:
        cols.append("division_id")
        divid_field = "division_id"
    elif "primary_division_id" in schema_cols:
        cols.append("primary_division_id")
        divid_field = "primary_division_id"
    else:
        divid_field = None
    has_admin_level = "admin_level" in schema_cols
    if has_admin_level:
        cols.append("admin_level")

    rows = []
    country_name_by_iso = {}
    rows_scanned = 0
    rows_kept = 0

    # Per-batch pull keeps peak memory bounded; the row buffer accumulates
    # only the kept admin-level rows (typically ~35-60K total).
    for batch in pf.iter_batches(batch_size=10_000, columns=cols):
        ids = batch.column("id").to_pylist()
        ctry = batch.column("country").to_pylist()
        regs = batch.column("region").to_pylist()
        subs = batch.column("subtype").to_pylist()
        nms = batch.column("names").to_pylist()
        cls = batch.column("class").to_pylist()
        land = batch.column("is_land").to_pylist()
        terr = batch.column("is_territorial").to_pylist()
        bbox = batch.column("bbox").to_pylist()
        divid = batch.column(divid_field).to_pylist() if divid_field else [None] * len(ids)
        admin_level_col = (
            batch.column("admin_level").to_pylist() if has_admin_level else None
        )

        for i in range(len(ids)):
            rows_scanned += 1
            sub = subs[i]
            if sub not in KEEP_SUBTYPES:
                continue
            level_from_subtype = KEEP_SUBTYPES[sub]
            # Trust an explicit admin_level when present (newer releases),
            # otherwise fall back to the subtype-derived level. Both should
            # agree for levels 0/1/2.
            admin_lv = admin_level_col[i] if has_admin_level else None
            if admin_lv is None:
                admin_lv = level_from_subtype

            nm = nms[i] if isinstance(nms[i], dict) else {}
            primary_name = nm.get("primary") if isinstance(nm, dict) else None
            common_json = _common_names_json(nm)
            xmin, ymin, xmax, ymax = _bbox_to_xyxy(bbox[i])

            iso = ctry[i]
            # Capture the primary name for level-0 rows so we can attach
            # readable country names on the second pass.
            if level_from_subtype == 0 and iso and primary_name:
                country_name_by_iso[iso] = primary_name

            rows.append((
                iso,
                # Country Name placeholder; filled on second pass.
                None,
                admin_lv,
                sub,
                primary_name,
                regs[i],
                cls[i],
                common_json,
                ids[i],
                divid[i],
                _bool_str(land[i]),
                _bool_str(terr[i]),
                xmin, ymin, xmax, ymax,
            ))
            rows_kept += 1

    print(f"      scanned {rows_scanned:,} rows, kept {rows_kept:,} at admin levels 0/1/2 "
          f"in {time.time()-t0:.1f}s")
    print(f"      level-0 country names captured: {len(country_name_by_iso):,}")
    return rows, country_name_by_iso


def attach_country_names(rows, country_name_by_iso):
    """Replace the placeholder Country Name slot with the joined value."""
    out = []
    for r in rows:
        iso = r[0]
        country_name = country_name_by_iso.get(iso)
        out.append((r[0], country_name) + r[2:])
    return out


def sort_rows(rows):
    """Sort: Country ISO, then admin level, then region (so a state's
    counties cluster under it), then primary name. Nulls sort last."""
    def key(r):
        iso = r[0] or "￿"
        level = r[2] if r[2] is not None else 99
        region = r[5] or "￿"
        name = r[4] or "￿"
        return (iso, level, region, name)
    rows.sort(key=key)
    return rows


def write_xlsx(out_path: Path, rows: list):
    out_path.parent.mkdir(parents=True, exist_ok=True)
    print(f"[2/2] Writing {out_path}")
    wb = Workbook(write_only=False)
    ws = wb.active
    ws.title = "Overture 012"

    ws.append(HEADER)
    for r in rows:
        ws.append(list(r))

    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(HEADER))
    ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"

    # Column widths tuned for at-a-glance readability on a 1080p screen.
    widths = [12, 30, 11, 11, 38, 18, 10, 60, 38, 38, 9, 12, 12, 12, 12, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(out_path)
    print(f"      wrote {ws.max_row - 1:,} data rows")


def main():
    print("Global Overture admin-levels 0/1/2 dump")
    print(f"Source parquet: {SOURCE_PARQUET}")
    print(f"Output xlsx:    {OUT_PATH}")
    if not Path(SOURCE_PARQUET).exists():
        print(f"ERROR: parquet not found at {SOURCE_PARQUET}")
        print("Set OVERTURE_DIVISION_AREA or download the file via the existing pipeline.")
        sys.exit(1)

    rows, country_names = collect_admin_012(SOURCE_PARQUET)
    rows = attach_country_names(rows, country_names)
    rows = sort_rows(rows)

    # Quick distribution print so the operator sees if level-2 coverage is
    # surprisingly thin for any country before opening the file.
    by_country_level = {}
    for r in rows:
        iso, _, lvl = r[0], r[1], r[2]
        if iso is None:
            continue
        by_country_level.setdefault(iso, {0: 0, 1: 0, 2: 0})
        if lvl in (0, 1, 2):
            by_country_level[iso][lvl] += 1
    countries_count = len(by_country_level)
    total_lvl0 = sum(d[0] for d in by_country_level.values())
    total_lvl1 = sum(d[1] for d in by_country_level.values())
    total_lvl2 = sum(d[2] for d in by_country_level.values())
    print(
        f"      countries: {countries_count}, "
        f"level-0: {total_lvl0:,}, level-1: {total_lvl1:,}, level-2: {total_lvl2:,}"
    )

    write_xlsx(OUT_PATH, rows)
    print("Done.")


if __name__ == "__main__":
    main()

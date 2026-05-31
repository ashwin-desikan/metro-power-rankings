"""
Patch public/data/football/index.json with Wikidata QIDs from MetroAreas.xlsx.
Run from the project root: python scripts/patch-football-qids.py
"""
import json, sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    import subprocess; subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--quiet", "--break-system-packages"])
    import openpyxl

ROOT = Path(__file__).resolve().parent.parent
INDEX = ROOT / "public/data/football/index.json"
XLSX  = ROOT / "MetroAreas.xlsx"

if not XLSX.exists():
    sys.exit(f"FAIL: {XLSX} not found")

print(f"Loading {XLSX.name}...")
wb = openpyxl.load_workbook(str(XLSX), read_only=True, data_only=True)
if "Wikidata_QIDs" not in wb.sheetnames:
    sys.exit("FAIL: Wikidata_QIDs sheet not found in MetroAreas.xlsx")

lookup = {}
for row in wb["Wikidata_QIDs"].iter_rows(min_row=2, values_only=True):
    if not row or len(row) < 4 or not row[1] or not row[2]: continue
    lookup[str(row[1]).strip().lower()] = (str(row[2]).strip(), str(row[3]).strip() if row[3] else None)
wb.close()
print(f"  {len(lookup)} QIDs loaded")

print(f"Patching {INDEX.name}...")
with open(INDEX, encoding="utf-8") as f:
    data = json.load(f)

patched = 0
for club in data["clubs"]:
    key = club["cur_name"].lower()
    if key in lookup:
        qid, url = lookup[key]
        club["wikidata_qid"] = qid
        if url: club["wikipedia_url"] = url
        patched += 1

with open(INDEX, "w", encoding="utf-8") as f:
    json.dump(data, f, ensure_ascii=False, separators=(",", ":"))
print(f"  {patched} clubs updated. Done.")

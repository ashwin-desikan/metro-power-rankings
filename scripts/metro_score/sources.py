"""Read the workbook's source sheets into per-metro indexes.

Everything the Score depends on lives in ten sheets, all joined to a metro by
its NAME as a plain string. Two rules govern that join, and both are Excel's,
not ours:

  * COUNTIFS is case-INSENSITIVE, so we lower-case both sides;
  * COUNTIFS does NOT trim whitespace, so we must not either.

The second one looks like a bug worth fixing and is not — at least not here.
Three FootballClub_Data rows carry a trailing space in the metro name
('Osnabruck ', 'Panevezys '), and Excel counts zero for them. Trimming would
make this engine disagree with the workbook on exactly those rows and the
parity gate would fail for a reason that has nothing to do with the migration.
Fix the data in the workbook; keep the engine faithful. `suspicious_keys()`
below exists to surface them rather than let them stay invisible.

openpyxl with data_only=True is used deliberately, matching scripts/extract.py,
rather than python-calamine: calamine's column offsets differ PER SHEET (the
Skyscrapers sheet has an empty leading column and indexes one to the left),
which has already produced one false verdict on this workbook.
"""
from __future__ import annotations

import sys
from collections import Counter
from pathlib import Path
from typing import Any, Callable, Dict, Iterable, Iterator, List, Optional, Sequence

try:
    import openpyxl
except ImportError:  # pragma: no cover
    import subprocess
    print("Installing openpyxl...")
    subprocess.check_call(
        [sys.executable, "-m", "pip", "install", "openpyxl", "--quiet", "--break-system-packages"]
    )
    import openpyxl


# Column letters are how the workbook is discussed, so the code speaks the same
# language. A() converts to the 0-based index openpyxl's values_only tuples use.
def A(letter: str) -> int:
    n = 0
    for ch in letter.upper():
        n = n * 26 + (ord(ch) - 64)
    return n - 1


def key(v: Any) -> str:
    """Excel's COUNTIFS text key: case-insensitive, whitespace NOT trimmed."""
    return v.lower() if isinstance(v, str) else ("" if v is None else str(v).lower())


def num(v: Any) -> float:
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    if isinstance(v, str):
        try:
            return float(v.strip())
        except ValueError:
            return 0.0
    return 0.0


def has_num(v: Any) -> bool:
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return True
    if isinstance(v, str):
        try:
            float(v.strip())
            return True
        except ValueError:
            return False
    return False


def text(v: Any) -> str:
    return v.strip() if isinstance(v, str) else ("" if v is None else str(v))


class Sheet:
    """A materialised sheet: a list of raw row tuples, from `first_data_row` on."""

    def __init__(self, name: str, rows: List[Sequence[Any]]):
        self.name = name
        self.rows = rows

    def __len__(self) -> int:
        return len(self.rows)

    def count_by(
        self, key_col: str, *, where: Optional[Callable[[Sequence[Any]], bool]] = None,
        value_col: Optional[str] = None,
    ) -> Counter:
        """COUNTIFS / SUMIFS over this sheet, indexed by the metro-name column."""
        ki, vi = A(key_col), (A(value_col) if value_col else None)
        out: Counter = Counter()
        for r in self.rows:
            if ki >= len(r):
                continue
            k = key(r[ki])
            if not k:
                continue
            if where is not None and not where(r):
                continue
            out[k] += num(r[vi]) if vi is not None and vi < len(r) else 1
        return out

    def col(self, r: Sequence[Any], letter: str) -> Any:
        i = A(letter)
        return r[i] if i < len(r) else None


SHEET_FIRST_ROW = {
    "Metro Areas": 4,        # rows 1-2 are a category band, row 3 the headers
    "Municipality": 2,
    "Counties": 2,
    "Team List": 2,
    "FootballClub_Data": 2,
    "Culture-Infra": 2,
    "Universities": 2,
    "MktCap_Data": 2,
    "Luxury Hospitality": 2,
    "SKYDB_Counts": 2,
    "Sheet2": 4,             # GDP; the sheet's own dimension starts at A3
}

# Every sheet the Score reads. Named here so a missing one is a clear error
# rather than an empty Counter that quietly zeroes a whole term.
REQUIRED_SHEETS = tuple(SHEET_FIRST_ROW)


class Workbook:
    def __init__(self, sheets: Dict[str, Sheet], path: Path,
                 header_rows: Optional[List[Sequence[Any]]] = None):
        self.sheets = sheets
        self.path = path
        # Metro Areas rows 1-3. Row 3 holds the headers the COUNTIFS criteria
        # point at ($N$3 and friends) and row 1 the category band that AI uses
        # ($AI$1); the data window starts at row 4, so they are captured here.
        self.header_rows: List[Sequence[Any]] = header_rows or []

    def __getitem__(self, name: str) -> Sheet:
        try:
            return self.sheets[name]
        except KeyError as exc:
            raise KeyError(f"workbook {self.path.name} has no sheet {name!r}") from exc


def load(xlsx_path: Path, sheets: Iterable[str] = REQUIRED_SHEETS) -> Workbook:
    wanted = list(sheets)
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    try:
        missing = [s for s in wanted if s not in wb.sheetnames]
        if missing:
            raise KeyError(f"{xlsx_path.name}: missing sheet(s) {missing}")
        out: Dict[str, Sheet] = {}
        header_rows: List[Sequence[Any]] = []
        for name in wanted:
            ws = wb[name]
            first = SHEET_FIRST_ROW.get(name, 2)
            if name == "Metro Areas":
                header_rows = [r for r in ws.iter_rows(min_row=1, max_row=3, values_only=True)]
                # read_only sheets are forward-only cursors, so the data window
                # needs a fresh handle after reading the header rows.
                ws = wb[name]
            rows = [r for r in ws.iter_rows(min_row=first, values_only=True)]
            out[name] = Sheet(name, rows)
        return Workbook(out, Path(xlsx_path), header_rows)
    finally:
        wb.close()


def suspicious_keys(wb: Workbook, metro_keys: Iterable[str]) -> List[str]:
    """Source rows whose metro name only fails to join because of whitespace.

    Excel drops these silently. The engine reproduces that faithfully, so this
    is the only place they become visible. Reported by parity.py, never fixed
    automatically: the correction belongs in the workbook.
    """
    known = set(metro_keys)
    seen: Counter = Counter()
    for sheet, col in (
        ("FootballClub_Data", "C"), ("Team List", "K"), ("Culture-Infra", "G"),
        ("Universities", "F"), ("MktCap_Data", "A"), ("Luxury Hospitality", "G"),
        ("Sheet2", "B"),
    ):
        for r in wb[sheet].rows:
            raw = wb[sheet].col(r, col)
            if not isinstance(raw, str):
                continue
            k = raw.lower()
            if k and k not in known and k.strip() in known:
                seen[f"{sheet}!{col}  {raw!r}"] += 1
    return [f"{n} row(s)  {what}" for what, n in seen.most_common()]

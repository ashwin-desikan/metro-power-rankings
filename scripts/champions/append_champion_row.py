"""Append one champion row to Champions_History.xlsx, surgically.

Adding The Hundred 2026 (Manchester Super Giants beat Trent Rockets at Lord's,
16 Aug 2026). The T20 roll already carried it via the manual supplement; the
CHAMPIONS LEDGER did not, so /teams/cricket/t20 showed the title and
/rankings/manchester#championships did not. One site, one answer.

METHOD. Insert a new <row> before </sheetData> and widen <dimension>. Every
other zip entry is copied through byte-for-byte with its original ZipInfo.

🔴 INLINE STRINGS, NOT sharedStrings. "Manchester Super Giants" is a new string;
adding it to sharedStrings.xml means renumbering count/uniqueCount, which the
surgical-edit rule says to refuse. Writing t="inlineStr" sidesteps that entirely
and Excel reads it natively.

Styles are copied per column from an existing Hundred row so the new row
inherits the same formatting (the Date column carries a yyyy-mm-dd numFmt even
though the value is text).

Run:  python scripts/champions/append_champion_row.py           (dry run)
      python scripts/champions/append_champion_row.py --write
"""
import os
import re
import shutil
import sys
import zipfile

import openpyxl

SRC = os.path.expanduser(r"~\OneDrive\Excel Files\Champions_History.xlsx")
SHEET = "Champions"
MODEL_ROW = 5573          # The Hundred 2025 - style + convention template
STAMP = "20260819-hundred-2026"

DATE_SOURCE = (
    "https://www.espncricinfo.com/series/the-hundred-men-s-competition-2026-1521176/"
    "manchester-super-giants-men-vs-trent-rockets-men-final-1521264/full-scorecard"
    ' | "MSG-M vs TR-M, Final at Lord\'s, London, Men\'s Hundred, Aug 16 2026"'
    " | champion in source: Manchester Super Giants (Men)"
)

# column index -> value. Mirrors the 2021-2025 Hundred rows exactly.
NEW = {
    1: "Cricket",                      # Sport
    2: "The Hundred",                  # Competition
    3: "The Hundred",                  # Era Name
    4: 2026,                           # Season
    5: 2026,                           # Year
    6: "Manchester Super Giants",      # Champion (ERA name - they are MSG in 2026)
    7: "Manchester Super Giants",      # Champion (Canonical)
    8: "Manchester",                   # Metro
    9: "manchester",                   # Metro Slug
    10: "2026-08-16",                  # Date
    11: "OK",                          # Canonical Status
    12: "England",                     # Scope
    13: "Domestic",                    # Scope Type
    14: 4,                             # Tier
    16: "data/cricket/honours.json",   # Source
    18: "Y",                           # Is Current
    19: DATE_SOURCE,                   # Date Source
    20: "Men's competition.",          # Date Method
    21: "exact | champion_confirmed",  # Date Precision
}


def esc(s):
    return (str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))


def sheet_xml_path(z, name):
    wb = z.read("xl/workbook.xml").decode("utf-8", "replace")
    rels = z.read("xl/_rels/workbook.xml.rels").decode("utf-8", "replace")
    relmap = dict(re.findall(r'Id="([^"]+)"[^>]*Target="([^"]+)"', rels))
    for nm, rid in re.findall(r'<sheet name="([^"]+)"[^>]*r:id="([^"]+)"', wb):
        if nm == name:
            return "xl/" + relmap[rid].lstrip("/")
    raise SystemExit("sheet %r not found" % name)


def col_letter(i):
    s = ""
    while i:
        i, r = divmod(i - 1, 26)
        s = chr(65 + r) + s
    return s


def main():
    write = "--write" in sys.argv

    wb = openpyxl.load_workbook(SRC, read_only=False, data_only=False)
    ws = wb[SHEET]
    last = ws.max_row
    new_row = last + 1
    print("sheet %r: max_row=%d -> appending row %d" % (SHEET, last, new_row))
    for r in range(2, last + 1):
        if (str(ws.cell(row=r, column=2).value or "") == "The Hundred"
                and str(ws.cell(row=r, column=5).value or "") == "2026"):
            raise SystemExit("REFUSING: The Hundred 2026 is already on row %d." % r)
    # Styles are read straight from the model row's XML further down, which is
    # the only place the real `s=` index lives.
    wb.close()

    with zipfile.ZipFile(SRC) as z:
        names = z.namelist()
        infos = {n: z.getinfo(n) for n in names}
        path = sheet_xml_path(z, SHEET)
        blobs = {n: z.read(n) for n in names}
    xml = blobs[path].decode("utf-8", "replace")

    if '<row r="%d"' % new_row in xml:
        raise SystemExit("REFUSING: row %d already exists in the XML." % new_row)

    # style ids from the model row, per column
    m = re.search(r'<row r="%d"[^>]*>(.*?)</row>' % MODEL_ROW, xml, re.S)
    if not m:
        raise SystemExit("REFUSING: model row %d not found." % MODEL_ROW)
    model_styles = {}
    for cm in re.finditer(r'<c r="([A-Z]+)%d"([^>]*)' % MODEL_ROW, m.group(1)):
        sm = re.search(r'\ss="(\d+)"', cm.group(2))
        if sm:
            model_styles[cm.group(1)] = sm.group(1)
    print("model row styles: %d columns carry an explicit style" % len(model_styles))

    cells = []
    for i in sorted(NEW):
        L = col_letter(i)
        st = (' s="%s"' % model_styles[L]) if L in model_styles else ""
        v = NEW[i]
        if isinstance(v, (int, float)):
            cells.append('<c r="%s%d"%s><v>%s</v></c>' % (L, new_row, st, v))
        else:
            cells.append('<c r="%s%d"%s t="inlineStr"><is><t xml:space="preserve">%s</t></is></c>'
                         % (L, new_row, st, esc(v)))
    row_xml = '<row r="%d" spans="1:24">%s</row>' % (new_row, "".join(cells))

    if "</sheetData>" not in xml:
        raise SystemExit("REFUSING: no </sheetData> in the sheet XML.")
    new_xml = xml.replace("</sheetData>", row_xml + "</sheetData>", 1)
    # widen <dimension ref="A1:X6811"/>
    new_xml, n = re.subn(r'(<dimension ref="[A-Z]+\d+:[A-Z]+)(\d+)("\s*/>)',
                         lambda mm: mm.group(1) + str(new_row) + mm.group(3), new_xml, count=1)
    print("dimension updated: %d" % n)
    print("\nrow to append:")
    for i in sorted(NEW):
        print("   %-2s %-22s = %r" % (col_letter(i), ws_header(i), NEW[i]))

    if not write:
        print("\nDRY RUN. Re-run with --write to apply.")
        return

    backup = SRC + ".bak-" + STAMP
    shutil.copyfile(SRC, backup)
    print("\nbackup:", backup)
    blobs[path] = new_xml.encode("utf-8")
    tmp = SRC + ".tmp"
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as out:
        for n in names:
            out.writestr(infos[n], blobs[n])
    os.replace(tmp, SRC)
    print("written.")

    with zipfile.ZipFile(SRC) as z:
        assert z.testzip() is None, "zip failed testzip()"
        assert z.namelist() == names, "zip entry order changed"
    wb = openpyxl.load_workbook(SRC, read_only=False, data_only=False)
    ws2 = wb[SHEET]
    print("verify: opens, %d sheets, max_row=%d" % (len(wb.sheetnames), ws2.max_row))
    got = {i: ws2.cell(row=new_row, column=i).value for i in sorted(NEW)}
    bad = [i for i in NEW if str(got[i]) != str(NEW[i])]
    for i in sorted(NEW):
        print("   %-2s = %r %s" % (col_letter(i), got[i], "OK" if i not in bad else "MISMATCH"))
    wb.close()
    if bad:
        raise SystemExit("REFUSING to declare success: %d cells read back wrong." % len(bad))


HEADERS = {}


def ws_header(i):
    return HEADERS.get(i, "col %d" % i)


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    _wb = openpyxl.load_workbook(SRC, read_only=True)
    _ws = _wb[SHEET]
    for _i in range(1, _ws.max_column + 1):
        HEADERS[_i] = str(_ws.cell(row=1, column=_i).value or "")
    _wb.close()
    main()

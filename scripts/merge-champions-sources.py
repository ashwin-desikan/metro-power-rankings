#!/usr/bin/env python3
"""One-time migration: merge ZoneZero_Champions.xlsx into Champions_History.xlsx.

Adds THREE new columns to the Champions sheet:
  Is Current        — "Y" on the reigning champion row for each competition
  Date Awarded      — exact date the title was won (from ZoneZero)
  Next Awarded Date — when the next edition is expected (from ZoneZero)

NOTE: Tier is NOT copied from ZoneZero. Champions_History.xlsx carries its own
Tier column with historically-calibrated values that supersede ZoneZero.
The build scripts will read Tier directly from Champions_History.

Matching is by normalized competition name (lowercase, strip punctuation).
Unmatched ZoneZero competitions are printed as warnings.

Run on Windows (needs both xlsx files):
    python scripts/merge-champions-sources.py

Writes:  ~/OneDrive/Excel Files/Champions_History_merged.xlsx
Review it, then rename/replace Champions_History.xlsx.
After that, ZoneZero_Champions.xlsx can be retired.
"""
import os, re, unicodedata, datetime
import openpyxl
from openpyxl.styles import Font, PatternFill

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ZONE_ZERO_SRC = os.path.join(ROOT, "ZoneZero_Champions.xlsx")
HISTORY_SRC   = os.path.expanduser("~/OneDrive/Excel Files/Champions_History.xlsx")
OUT           = os.path.expanduser("~/OneDrive/Excel Files/Champions_History_merged.xlsx")

# Manual name overrides: ZoneZero competition name -> History competition name.
# Add entries here if the auto-match misses something (printed as warnings).
MANUAL_MAP = {
    "Champions League":          "UEFA Champions League",
    "Europa League":             "UEFA Europa League",
    "Conference League":         "UEFA Europa Conference League",
    "Club World Cup":            "FIFA Club World Cup",
    "Nations League":            "UEFA Nations League",
    "Gold Cup":                  "CONCACAF Gold Cup",
    "African Cup of Nations":    "Africa Cup of Nations",
    "Women's World Cup":         "FIFA Women's World Cup",
    "NWSL Championship":         "NWSL",
    "FA Women's Super League":   "Women's Super League",
    "Super Rugby Pacific":       "Super Rugby",
    "Pro14":                     "Pro14/URC",
    "United Rugby Championship": "Pro14/URC",
    "URC":                       "Pro14/URC",
}

def _norm(s):
    """Normalize for fuzzy matching: lowercase, strip accents + punctuation."""
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.lower()
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    for noise in ("the", "of", "and", "&"):
        s = re.sub(rf"\b{noise}\b", "", s)
    return re.sub(r"\s+", " ", s).strip()

def to_iso(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return (v.date() if isinstance(v, datetime.datetime) else v).isoformat()
    if isinstance(v, str) and v.strip():
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
            try:
                return datetime.datetime.strptime(v.strip(), fmt).date().isoformat()
            except ValueError:
                pass
    return None

# ── Load ZoneZero ─────────────────────────────────────────────────────────────

def load_zone_zero():
    """Returns dict: normalized_comp_name -> {dateAwarded, nextAwardedDate}"""
    wb = openpyxl.load_workbook(ZONE_ZERO_SRC, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    idx = {h: i for i, h in enumerate(hdr)}

    def col(*names):
        for n in names:
            if n in idx:
                return idx[n]
        return None

    iComp = col("competiton", "competition")
    iDate = col("date awarded", "date", "year")
    iNext = col("next awarded date", "next awarded", "nextawarded", "next")

    result = {}
    for r in rows[1:]:
        def g(i): return r[i] if i is not None and i < len(r) else None
        comp = g(iComp)
        if not comp:
            continue
        comp = str(comp).strip()
        hist_name = MANUAL_MAP.get(comp, comp)
        key = _norm(hist_name)
        result[key] = {
            "zz_name":         comp,
            "dateAwarded":     to_iso(g(iDate)),
            "nextAwardedDate": to_iso(g(iNext)),
        }

    print(f"ZoneZero: loaded {len(result)} competitions")
    return result

# ── Merge into Champions_History ──────────────────────────────────────────────

def main():
    zz = load_zone_zero()

    wb = openpyxl.load_workbook(HISTORY_SRC, data_only=True)
    ws = wb["Champions"]

    # Read existing header
    hdr_row = [c.value for c in ws[1]]
    hdr = [str(h).strip() if h is not None else "" for h in hdr_row]
    comp_col = next(i for i, h in enumerate(hdr) if h.lower() == "competition")
    year_col  = next(i for i, h in enumerate(hdr) if h.lower() == "year")

    # Locate or append new columns (idempotent)
    NEW_COLS = ["Is Current", "Date Awarded", "Next Awarded Date"]
    existing = {h: i for i, h in enumerate(hdr)}
    next_col = len(hdr)
    col_idx = {}
    for nc in NEW_COLS:
        if nc in existing:
            col_idx[nc] = existing[nc]
        else:
            col_idx[nc] = next_col
            next_col += 1
            ws.cell(row=1, column=col_idx[nc] + 1, value=nc)

    # Find the max year per competition (= reigning champion row)
    comp_max_year = {}
    all_rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        comp = row[comp_col].value
        yr   = row[year_col].value
        if not comp:
            continue
        comp_str = str(comp).strip()
        try:
            yr_int = int(float(yr)) if yr is not None else None
        except (TypeError, ValueError):
            yr_int = None
        all_rows.append((i, comp_str, yr_int))
        if yr_int is not None:
            key = _norm(comp_str)
            if key not in comp_max_year or yr_int > comp_max_year[key]:
                comp_max_year[key] = yr_int

    # Write new column values
    unmatched_zz = set(zz.keys())
    unmatched_hist = set()

    for (row_num, comp_str, yr_int) in all_rows:
        key = _norm(comp_str)
        zz_entry = zz.get(key)

        is_current = None
        date_awarded = None
        next_awarded_date = None

        if zz_entry:
            unmatched_zz.discard(key)
            if yr_int is not None and comp_max_year.get(key) == yr_int:
                is_current = "Y"
                date_awarded = zz_entry["dateAwarded"] or None
                next_awarded_date = zz_entry["nextAwardedDate"] or None
        else:
            unmatched_hist.add(comp_str)

        ws.cell(row=row_num, column=col_idx["Is Current"] + 1).value = is_current
        ws.cell(row=row_num, column=col_idx["Date Awarded"] + 1).value = date_awarded
        ws.cell(row=row_num, column=col_idx["Next Awarded Date"] + 1).value = next_awarded_date

    # Style new header cells
    header_fill = PatternFill("solid", fgColor="DDEEFF")
    for nc in NEW_COLS:
        cell = ws.cell(row=1, column=col_idx[nc] + 1)
        cell.font = Font(bold=True)
        cell.fill = header_fill

    wb.save(OUT)
    print(f"\n✓ Wrote {OUT}")
    print(f"  Rows processed: {len(all_rows)}")

    current_count = sum(
        1 for (_, comp, yr) in all_rows
        if zz.get(_norm(comp)) and comp_max_year.get(_norm(comp)) == yr
    )
    print(f"  Is Current = Y rows written: {current_count}  (expect ~87)")

    if unmatched_zz:
        print(f"\n⚠  ZoneZero competitions NOT matched to History ({len(unmatched_zz)}):")
        for k in sorted(unmatched_zz):
            print(f"   ZZ: '{zz[k]['zz_name']}'  (norm: '{k}')")
        print("   → Add MANUAL_MAP entries or add rows to Champions_History.")

    if unmatched_hist:
        unique = sorted(set(unmatched_hist))
        print(f"\nℹ  History competitions not in ZoneZero ({len(unique)}) — fine for historical-only entries:")
        for c in unique[:20]:
            print(f"   '{c}'")
        if len(unique) > 20:
            print(f"   ... and {len(unique) - 20} more")


if __name__ == "__main__":
    main()

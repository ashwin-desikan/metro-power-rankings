"""North America Overture audit.

Reads the Overture division_area parquet for NA (US/CA/MX) plus the workbook's
States / Counties / Municipality sheets, then maps each unique workbook
admin-unit row to its Overture feature. Emits:

  - A working-copy xlsx with new Overture columns appended to each sheet.
  - Per-sheet CSV summaries of match quality.
  - A console report of exact / fuzzy / unmatched counts per country.

Why: the existing build-metro-boundaries.py grew 600+ lines of country-specific
name-normalization, suffix-stripping, alias and type-inference logic because
the workbook never told the script which Overture feature each row pointed at.
This audit is the first step toward making the workbook self-describing: once
each row carries Overture Country / Region / Subtype / Name / GERS ID, the
boundary builder collapses to a thin lookup table.

Run:
  python3 scripts/overture-na-audit.py \
      --parquet /path/to/north-america-division-area.parquet \
      --workbook MetroAreas.xlsx \
      --out-xlsx outputs/MetroAreas-with-overture-NA.xlsx \
      --out-csv-dir outputs/

Defaults assume the project layout under
/sessions/zen-relaxed-ritchie/mnt/Projects--Metro Area Project.
"""
from __future__ import annotations

import argparse
import csv
import os
import re
import sys
import time
import unicodedata
from collections import defaultdict, Counter
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl import load_workbook


# ---------- ISO 3166-2 maps -----------------------------------------------

US_STATE_TO_ISO = {
    "Alabama": "US-AL", "Alaska": "US-AK", "Arizona": "US-AZ", "Arkansas": "US-AR",
    "California": "US-CA", "Colorado": "US-CO", "Connecticut": "US-CT",
    "Delaware": "US-DE", "DC": "US-DC", "District of Columbia": "US-DC",
    "Florida": "US-FL", "Georgia": "US-GA", "Hawaii": "US-HI", "Idaho": "US-ID",
    "Illinois": "US-IL", "Indiana": "US-IN", "Iowa": "US-IA", "Kansas": "US-KS",
    "Kentucky": "US-KY", "Louisiana": "US-LA", "Maine": "US-ME", "Maryland": "US-MD",
    "Massachusetts": "US-MA", "Michigan": "US-MI", "Minnesota": "US-MN",
    "Mississippi": "US-MS", "Missouri": "US-MO", "Montana": "US-MT",
    "Nebraska": "US-NE", "Nevada": "US-NV", "New Hampshire": "US-NH",
    "New Jersey": "US-NJ", "New Mexico": "US-NM", "New York": "US-NY",
    "North Carolina": "US-NC", "North Dakota": "US-ND", "Ohio": "US-OH",
    "Oklahoma": "US-OK", "Oregon": "US-OR", "Pennsylvania": "US-PA",
    "Rhode Island": "US-RI", "South Carolina": "US-SC", "South Dakota": "US-SD",
    "Tennessee": "US-TN", "Texas": "US-TX", "Utah": "US-UT", "Vermont": "US-VT",
    "Virginia": "US-VA", "Washington": "US-WA", "West Virginia": "US-WV",
    "Wisconsin": "US-WI", "Wyoming": "US-WY",
    "Puerto Rico": "US-PR", "Guam": "US-GU", "American Samoa": "US-AS",
    "U.S. Virgin Islands": "US-VI", "Northern Mariana Islands": "US-MP",
}

CA_STATE_TO_ISO = {
    "Alberta": "CA-AB", "British Columbia": "CA-BC", "Manitoba": "CA-MB",
    "New Brunswick": "CA-NB", "Newfoundland and Labrador": "CA-NL",
    "Newfoundland": "CA-NL", "Nova Scotia": "CA-NS",
    "Northwest Territories": "CA-NT", "Nunavut": "CA-NU", "Ontario": "CA-ON",
    "Prince Edward Island": "CA-PE", "Quebec": "CA-QC", "Québec": "CA-QC",
    "Saskatchewan": "CA-SK", "Yukon": "CA-YT",
}

MX_STATE_TO_ISO = {
    "Aguascalientes": "MX-AGU", "Baja California": "MX-BCN",
    "Baja California Sur": "MX-BCS", "Campeche": "MX-CAM",
    "Chiapas": "MX-CHP", "Chihuahua": "MX-CHH",
    "Coahuila": "MX-COA", "Coahuila de Zaragoza": "MX-COA",
    "Colima": "MX-COL", "Mexico City": "MX-CMX",
    "Ciudad de México": "MX-CMX", "Distrito Federal": "MX-CMX",
    "Durango": "MX-DUR", "Guanajuato": "MX-GUA", "Guerrero": "MX-GRO",
    "Hidalgo": "MX-HID", "Jalisco": "MX-JAL",
    "México": "MX-MEX", "Estado de México": "MX-MEX", "State of Mexico": "MX-MEX",
    "Michoacán": "MX-MIC", "Michoacan": "MX-MIC",
    "Michoacán de Ocampo": "MX-MIC",
    "Morelos": "MX-MOR", "Nayarit": "MX-NAY", "Nuevo León": "MX-NLE",
    "Nuevo Leon": "MX-NLE", "Oaxaca": "MX-OAX", "Puebla": "MX-PUE",
    "Querétaro": "MX-QUE", "Queretaro": "MX-QUE", "Quintana Roo": "MX-ROO",
    "San Luis Potosí": "MX-SLP", "San Luis Potosi": "MX-SLP",
    "Sinaloa": "MX-SIN", "Sonora": "MX-SON", "Tabasco": "MX-TAB",
    "Tamaulipas": "MX-TAM", "Tlaxcala": "MX-TLA",
    "Veracruz": "MX-VER", "Veracruz de Ignacio de la Llave": "MX-VER",
    "Yucatán": "MX-YUC", "Yucatan": "MX-YUC", "Zacatecas": "MX-ZAC",
}

COUNTRY_STATE_MAP = {
    "United States": US_STATE_TO_ISO,
    "Canada": CA_STATE_TO_ISO,
    "Mexico": MX_STATE_TO_ISO,
}

COUNTRY_TO_ISO = {"United States": "US", "Canada": "CA", "Mexico": "MX"}


# ---------- Normalization -------------------------------------------------

ADMIN_SUFFIXES = (
    " County", " Counties", " Parish", " Borough", " Census Area",
    " Municipality", " Municipio", " Planning Region",
    " Regional District", " Regional Municipality", " District Municipality",
    " United Counties", " Region", " Rural District", " District",
    " City and Borough",
)


def ascii_fold(s: str) -> str:
    """Strip accents and modifier-letter marks Overture's primary names use
    (Hawaiʻi -> Hawaii, Kauaʻi -> Kauai, Behchokǫ̀ -> Behchoko, Montréal ->
    Montreal). Also strips smart apostrophes (Hawaiʼi).
    """
    if not s:
        return ""
    # Hawaiian okina (modifier letter turned comma U+02BB) and modifier letter
    # apostrophe (U+02BC) and right single quotation mark (U+2019) are not
    # combining marks, so unicodedata.normalize would leave them. Strip first.
    for ch in ("\u02BB", "\u02BC", "\u2019", "\u2018"):
        s = s.replace(ch, "")
    return "".join(
        c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c)
    )


def norm_name(s: str) -> str:
    """Normalize for matching: ASCII fold, lowercase, collapse whitespace."""
    if not s:
        return ""
    s = ascii_fold(s).lower().strip()
    s = re.sub(r"\s+", " ", s)
    return s


def strip_admin_suffix(name: str) -> str:
    """Remove a trailing admin suffix (' County', ' Parish', etc.) from a name."""
    if not name:
        return name
    for suf in ADMIN_SUFFIXES:
        if name.endswith(suf):
            return name[: -len(suf)]
    return name


# ---------- Editorial aliases (carried over from build-metro-boundaries) -

# (region, normalized workbook name) -> normalized Overture name
COUNTY_ALIASES = {
    ("US-RI", "washington"): "south",  # RI Washington County -> South County
    ("CA-BC", "greater vancouver"): "metro vancouver",
    ("CA-QC", "le saguenay-et-son-fjord"): "le fjord-du-saguenay",
    ("CA-ON", "haldimand-norfolk"): "haldimand",
    ("MX-VER", "fortin"): "fortin de las flores",
}


# ---------- Overture index -------------------------------------------------

def load_overture_index(parquet_path: str, *, countries=("US", "CA", "MX")):
    """Load the parquet metadata and build a multi-tier index keyed by
    (country, region, subtype, normalized name).

    Returns a dict where each leaf is a list of feature records:
        {
            (country, region, subtype, norm_name): [
                {"id": ..., "primary": ..., "class": ..., "admin_level": ...},
                ...
            ]
        }
    Plus a mapping (country, region, subtype) -> list of features (for
    enumeration when name lookup fails).
    """
    print(f"[overture] reading {parquet_path}")
    t0 = time.time()
    import pyarrow.parquet as pq

    cols = ["id", "country", "region", "subtype", "admin_level", "class", "names"]
    table = pq.read_table(parquet_path, columns=cols)
    n = table.num_rows
    print(f"[overture]   loaded {n:,} rows in {time.time()-t0:.1f}s")

    # Pull as pandas for fast iteration
    df = table.to_pandas()
    df = df[df["country"].isin(countries)]
    print(f"[overture]   {len(df):,} rows after country filter")

    by_key = defaultdict(list)
    by_subset = defaultdict(list)
    for r in df.itertuples(index=False):
        names = r.names or {}
        primary = names.get("primary") if isinstance(names, dict) else None
        if not primary:
            continue
        # Skip class=maritime when the same (country, region, subtype, name)
        # has class=land — handled by dedupe below. Keep for now.
        rec = {
            "id": r.id,
            "primary": primary,
            "class": r._fields[5] if False else r.__class__,  # placeholder
            "admin_level": int(r.admin_level) if r.admin_level == r.admin_level else None,
            "country": r.country,
            "region": r.region,
            "subtype": r.subtype,
        }
        # cleaner: read class properly
        rec["class"] = getattr(r, "class") if hasattr(r, "class") else None
        if rec["class"] is None:
            # itertuples mangles 'class' due to keyword. Re-fetch via attr name.
            rec["class"] = getattr(r, "class_", None)
        norm = norm_name(primary)
        key = (r.country, r.region, r.subtype, norm)
        by_key[key].append(rec)
        by_subset[(r.country, r.region, r.subtype)].append(rec)

    # Dedupe: prefer class=land over class=maritime within the same key.
    deduped_key = {}
    for key, recs in by_key.items():
        land = [r for r in recs if r.get("class") == "land"]
        deduped_key[key] = land if land else recs
    print(f"[overture]   indexed {len(deduped_key):,} unique (country,region,subtype,norm) keys")
    return deduped_key, by_subset


def _attr_class(row):
    """itertuples renames 'class' to '_5' when present in column order — return
    via direct attribute fetch."""
    return getattr(row, "class", None) or getattr(row, "_5", None)


def load_overture_index_v2(parquet_path: str, *, countries=("US", "CA", "MX")):
    """Cleaner v2: avoids the itertuples 'class' attribute hazard."""
    print(f"[overture] reading {parquet_path}")
    t0 = time.time()
    import pyarrow.parquet as pq

    cols = ["id", "country", "region", "subtype", "admin_level", "class", "names"]
    table = pq.read_table(parquet_path, columns=cols)
    print(f"[overture]   loaded {table.num_rows:,} rows in {time.time()-t0:.1f}s")
    df = table.to_pandas()
    df = df[df["country"].isin(countries)].reset_index(drop=True)
    print(f"[overture]   {len(df):,} rows after country filter")

    by_key = defaultdict(list)
    by_subset = defaultdict(list)
    for i in range(len(df)):
        names = df.at[i, "names"]
        primary = names.get("primary") if isinstance(names, dict) else None
        if not primary:
            continue
        rec = {
            "id": df.at[i, "id"],
            "primary": primary,
            "class": df.at[i, "class"],
            "admin_level": df.at[i, "admin_level"],
            "country": df.at[i, "country"],
            "region": df.at[i, "region"],
            "subtype": df.at[i, "subtype"],
        }
        norm = norm_name(primary)
        key = (rec["country"], rec["region"], rec["subtype"], norm)
        by_key[key].append(rec)
        by_subset[(rec["country"], rec["region"], rec["subtype"])].append(rec)

    deduped_key = {}
    for key, recs in by_key.items():
        land = [r for r in recs if r.get("class") == "land"]
        deduped_key[key] = land if land else recs
    print(f"[overture]   indexed {len(deduped_key):,} unique (country,region,subtype,norm) keys")
    return deduped_key, by_subset


# ---------- Lookup --------------------------------------------------------

SUFFIX_ADD_LIST = (
    " County", " Parish", " Borough", " Census Area",
    " Planning Region", " Regional District", " Regional Municipality",
    " City and Borough",
)

PREFIX_ADD_LIST = ("Regional District of ",)


def st_saint_variants(name):
    """Yield 'St.' <-> 'Saint' two-way swaps. Overture uses both forms
    inconsistently (Saint Johns FL but St. Ann MO)."""
    yield name
    if "St. " in name:
        yield name.replace("St. ", "Saint ")
    if "Saint " in name:
        yield name.replace("Saint ", "St. ")


def strip_brackets(name):
    """Remove '[Alias]' and '(parenthetical)' annotations from workbook names.
    MX rows carry these as editorial markers (e.g., 'Playas de Rosarito
    [Rosarito Beach]', 'Mezcalapa (← Tecpatán)') that Overture does not."""
    name = re.sub(r"\s*\[[^\]]*\]\s*", " ", name)
    name = re.sub(r"\s*\([^)]*\)\s*", " ", name)
    return re.sub(r"\s+", " ", name).strip()


def lookup(country_iso, region_iso, subtype, raw_name, by_key, by_subset, *, alias_map=None):
    """Try a cascade of normalizations against the Overture index. Returns
    a tuple (status, record, candidates) where:

        status: 'exact' | 'suffix-strip' | 'suffix-add' | 'prefix-add' |
                'ascii-fold' | 'st-saint' | 'bracket-strip' | 'alias' |
                'multi' | 'unmatched'
        record: the matched record dict (or None)
        candidates: list of all candidate records considered (for review)
    """
    if not raw_name:
        return ("unmatched", None, [])

    raw = str(raw_name).strip()
    candidates_seen = []

    def try_key(k, status):
        recs = by_key.get(k)
        if not recs:
            return None
        candidates_seen.extend(recs)
        if len(recs) == 1:
            return (status, recs[0], list(recs))
        # Multi-candidate: pick the first land-class as the representative,
        # caller can read len(candidates) from the returned tuple.
        return ("multi", recs[0], list(recs))

    # 1. Exact normalized match
    norm = norm_name(raw)
    res = try_key((country_iso, region_iso, subtype, norm), "exact")
    if res:
        return res

    # 2. Strip admin suffix from workbook name and retry
    stripped = strip_admin_suffix(raw)
    if stripped != raw:
        n = norm_name(stripped)
        res = try_key((country_iso, region_iso, subtype, n), "suffix-strip")
        if res:
            return res

    # 3. Bracketed alias strip (MX cleanup)
    debr = strip_brackets(raw)
    if debr and debr != raw:
        n = norm_name(debr)
        res = try_key((country_iso, region_iso, subtype, n), "bracket-strip")
        if res:
            return res
        # Continue with the de-bracketed name as the base for further passes.
        raw_base = debr
    else:
        raw_base = raw

    # 4. Try ADDING admin suffix to the workbook name
    for suf in SUFFIX_ADD_LIST:
        n = norm_name(raw_base + suf)
        res = try_key((country_iso, region_iso, subtype, n), "suffix-add")
        if res:
            return res

    # 5. Try ADDING admin prefix (BC "Regional District of X" form)
    for pre in PREFIX_ADD_LIST:
        n = norm_name(pre + raw_base)
        res = try_key((country_iso, region_iso, subtype, n), "prefix-add")
        if res:
            return res

    # 6. St./Saint two-way swap
    for variant in st_saint_variants(raw_base):
        if variant == raw_base:
            continue
        n = norm_name(variant)
        res = try_key((country_iso, region_iso, subtype, n), "st-saint")
        if res:
            return res
        # Also try with admin suffixes appended
        for suf in SUFFIX_ADD_LIST:
            n2 = norm_name(variant + suf)
            res = try_key((country_iso, region_iso, subtype, n2), "st-saint")
            if res:
                return res

    # 7. ASCII-fold pass (covers Hawaii okina, Behchoko etc.)
    folded = ascii_fold(raw_base).lower().strip()
    if folded != norm:
        res = try_key((country_iso, region_iso, subtype, folded), "ascii-fold")
        if res:
            return res
        # Also try ascii-folded + suffixes
        for suf in SUFFIX_ADD_LIST:
            n = norm_name(ascii_fold(raw_base + suf))
            res = try_key((country_iso, region_iso, subtype, n), "ascii-fold")
            if res:
                return res

    # 8. Alias map
    if alias_map:
        alias_to = alias_map.get((region_iso, norm))
        if alias_to:
            res = try_key((country_iso, region_iso, subtype, alias_to), "alias")
            if res:
                return res

    return ("unmatched", None, candidates_seen)


# ---------- Sheet processors ----------------------------------------------

OVERTURE_OUT_HEADERS = [
    "Overture Country",
    "Overture Region",
    "Overture Subtype",
    "Overture Name",
    "Overture GERS ID",
    "Overture Match",
    "Overture Notes",
]


def process_states(ws_in, ws_out, by_key, by_subset, summary):
    """States sheet: row already carries ISO 3166-2 in col 9. Match against
    subtype=region. Country comes from col 4 (full name)."""
    rows = list(ws_in.iter_rows(values_only=True))
    if not rows:
        return
    # Copy header + add new columns
    header = list(rows[0]) + OVERTURE_OUT_HEADERS
    ws_out.append(header)
    for r in rows[1:]:
        if not r:
            ws_out.append([])
            continue
        country_full = r[4] if len(r) > 4 else None
        iso2 = r[9] if len(r) > 9 else None
        new_cols = ["", "", "", "", "", "", ""]
        if country_full in COUNTRY_TO_ISO and iso2:
            country_iso = COUNTRY_TO_ISO[country_full]
            recs = by_subset.get((country_iso, iso2, "region")) or []
            land = [x for x in recs if x.get("class") == "land"]
            recs = land if land else recs
            if len(recs) == 1:
                rec = recs[0]
                new_cols = [
                    country_iso, iso2, "region",
                    rec["primary"], rec["id"], "exact", "",
                ]
                summary[(country_iso, "States", "exact")] += 1
            elif len(recs) > 1:
                rec = recs[0]
                names = "; ".join(sorted(set(x["primary"] for x in recs)))
                new_cols = [
                    country_iso, iso2, "region",
                    rec["primary"], rec["id"], "multi",
                    f"candidates: {names}",
                ]
                summary[(country_iso, "States", "multi")] += 1
            else:
                new_cols = [country_iso, iso2, "region", "", "", "unmatched", ""]
                summary[(country_iso, "States", "unmatched")] += 1
        else:
            summary[("?", "States", "skipped")] += 1
        ws_out.append(list(r) + new_cols)


def process_counties(ws_in, ws_out, by_key, by_subset, summary):
    """Counties sheet: col 0 country, col 1 county name, col 2 state full,
    col 6 type, col 7 metro. Match against subtype=county for US/CA/MX, with
    DC and city-type rows handled separately (subtype=region for DC; subtype=
    locality for VA/MO/NV independent cities)."""
    rows = list(ws_in.iter_rows(values_only=True))
    if not rows:
        return
    header = list(rows[0]) + OVERTURE_OUT_HEADERS
    ws_out.append(header)
    for r in rows[1:]:
        if not r:
            ws_out.append([])
            continue
        country_full = r[0] if len(r) > 0 else None
        county_name = r[1] if len(r) > 1 else None
        state_full = r[2] if len(r) > 2 else None
        county_type = r[6] if len(r) > 6 else None
        new_cols = ["", "", "", "", "", "", ""]

        if country_full in COUNTRY_TO_ISO and county_name and state_full:
            country_iso = COUNTRY_TO_ISO[country_full]
            iso2 = COUNTRY_STATE_MAP[country_full].get(str(state_full).strip())

            # DC special: subtype=region, name="District of Columbia"
            type_l = str(county_type or "").lower().strip()
            if iso2 == "US-DC" or type_l == "federal district":
                # Try subtype=region with name "District of Columbia"
                status, rec, cands = lookup(
                    country_iso, "US-DC", "region",
                    "District of Columbia", by_key, by_subset
                )
                if rec:
                    new_cols = [
                        country_iso, "US-DC", "region",
                        rec["primary"], rec["id"], status, "DC special-case",
                    ]
                    summary[(country_iso, "Counties", status)] += 1
                else:
                    new_cols = [country_iso, "US-DC", "region", "", "", "unmatched",
                                "DC region not found"]
                    summary[(country_iso, "Counties", "unmatched")] += 1
            elif type_l == "city":
                # Independent city (VA/MO/NV/MD): try subtype=locality, then
                # subtype=county. For "Baltimore City" (workbook) try the
                # bare-name "Baltimore" at county subtype as a final pass.
                status, rec, cands = lookup(
                    country_iso, iso2, "locality",
                    county_name, by_key, by_subset, alias_map=COUNTY_ALIASES
                )
                if rec:
                    new_cols = [
                        country_iso, iso2 or "", "locality",
                        rec["primary"], rec["id"], status,
                        "type=City -> locality",
                    ]
                    summary[(country_iso, "Counties", status)] += 1
                else:
                    status2, rec2, cands2 = lookup(
                        country_iso, iso2, "county",
                        county_name, by_key, by_subset, alias_map=COUNTY_ALIASES
                    )
                    if rec2:
                        new_cols = [
                            country_iso, iso2 or "", "county",
                            rec2["primary"], rec2["id"], status2,
                            "type=City but Overture has county",
                        ]
                        summary[(country_iso, "Counties", status2)] += 1
                    else:
                        # Strip trailing " City" from workbook name and retry
                        # county lookup. Catches "Baltimore City" -> "Baltimore"
                        # which Overture publishes as a bare county-class entry.
                        bare = re.sub(r"\s+City$", "", str(county_name).strip())
                        if bare and bare != str(county_name).strip():
                            status3, rec3, cands3 = lookup(
                                country_iso, iso2, "county",
                                bare, by_key, by_subset, alias_map=COUNTY_ALIASES
                            )
                            if rec3:
                                new_cols = [
                                    country_iso, iso2 or "", "county",
                                    rec3["primary"], rec3["id"], status3,
                                    "type=City stripped -> county bare",
                                ]
                                summary[(country_iso, "Counties", status3)] += 1
                                ws_out.append(list(r) + new_cols)
                                continue
                        new_cols = [country_iso, iso2 or "", "locality",
                                    "", "", "unmatched", "tried locality + county + bare"]
                        summary[(country_iso, "Counties", "unmatched")] += 1
            else:
                # Default: subtype=county
                status, rec, cands = lookup(
                    country_iso, iso2, "county",
                    county_name, by_key, by_subset, alias_map=COUNTY_ALIASES
                )
                if rec:
                    new_cols = [
                        country_iso, iso2 or "", "county",
                        rec["primary"], rec["id"], status, "",
                    ]
                    summary[(country_iso, "Counties", status)] += 1
                else:
                    # Fallback: try subtype=locality (for Carson City NV which
                    # is technically county-equivalent but published as locality)
                    status2, rec2, cands2 = lookup(
                        country_iso, iso2, "locality",
                        county_name, by_key, by_subset, alias_map=COUNTY_ALIASES
                    )
                    if rec2:
                        new_cols = [
                            country_iso, iso2 or "", "locality",
                            rec2["primary"], rec2["id"], status2,
                            "fell back to locality",
                        ]
                        summary[(country_iso, "Counties", status2)] += 1
                    else:
                        new_cols = [country_iso, iso2 or "", "county",
                                    "", "", "unmatched", "tried county + locality"]
                        summary[(country_iso, "Counties", "unmatched")] += 1
        else:
            summary[("?", "Counties", "skipped")] += 1
        ws_out.append(list(r) + new_cols)


def process_municipality(ws_in, ws_out, by_key, by_subset, summary):
    """Municipality sheet: col 1 country, col 2 municipality name, col 3
    district, col 4 state full. Match against subtype=locality for the
    municipality (col 2)."""
    rows = list(ws_in.iter_rows(values_only=True))
    if not rows:
        return
    header = list(rows[0]) + OVERTURE_OUT_HEADERS
    ws_out.append(header)
    for r in rows[1:]:
        if not r:
            ws_out.append([])
            continue
        country_full = r[1] if len(r) > 1 else None
        mun_name = r[2] if len(r) > 2 else None
        district = r[3] if len(r) > 3 else None
        state_full = r[4] if len(r) > 4 else None
        new_cols = ["", "", "", "", "", "", ""]

        if country_full in COUNTRY_TO_ISO and mun_name and state_full:
            country_iso = COUNTRY_TO_ISO[country_full]
            iso2 = COUNTRY_STATE_MAP[country_full].get(str(state_full).strip())

            # Balance-of redirect: workbook rows like "Balance of Anne Arundel
            # County" or "Balance of Colonie town" represent the unincorporated
            # remainder of a parent county/town. Overture does not publish
            # "Balance of" features. Redirect the lookup to col 3 (parent
            # district) at subtype=county.
            mun_str = str(mun_name).strip()
            balance_match = re.match(r"^Balance of\s+(.+?)$", mun_str, re.IGNORECASE)
            if balance_match and district:
                # Use col 3 as the lookup target at county level.
                status, rec, cands = lookup(
                    country_iso, iso2, "county",
                    str(district).strip(), by_key, by_subset, alias_map=COUNTY_ALIASES
                )
                if rec:
                    new_cols = [
                        country_iso, iso2 or "", "county",
                        rec["primary"], rec["id"], status,
                        f"Balance-of redirect to col3={district!r}",
                    ]
                    summary[(country_iso, "Municipality", status)] += 1
                    ws_out.append(list(r) + new_cols)
                    continue
                # If county lookup fails, fall through to locality attempt below.

            # Municipality rows: workbook col 2 maps to Overture subtype=locality
            # in most cases. Strip type-suffixes like " city", " town",
            # " township", " village" first.
            stripped = re.sub(
                r"\s+(city|town|township|village|borough|cdp|municipality)\s*$",
                "", mun_str, flags=re.IGNORECASE
            )
            status, rec, cands = lookup(
                country_iso, iso2, "locality",
                stripped, by_key, by_subset, alias_map=COUNTY_ALIASES
            )
            note = ""
            if status == "unmatched" and stripped != mun_str:
                # Try original (un-stripped) name
                status, rec, cands = lookup(
                    country_iso, iso2, "locality",
                    mun_str, by_key, by_subset, alias_map=COUNTY_ALIASES
                )
                if rec:
                    note = "matched without type-suffix strip"
            if status == "multi" and district:
                # Multi-candidate disambiguation hint: record the workbook's
                # parent district alongside so editorial review can pick the
                # right candidate.
                note = f"multi via col3={district!r}; candidates={len(cands)}"
            if rec:
                new_cols = [
                    country_iso, iso2 or "", "locality",
                    rec["primary"], rec["id"], status, note,
                ]
                summary[(country_iso, "Municipality", status)] += 1
            else:
                new_cols = [
                    country_iso, iso2 or "", "locality",
                    "", "", "unmatched",
                    f"district={district!r}",
                ]
                summary[(country_iso, "Municipality", "unmatched")] += 1
        else:
            summary[("?", "Municipality", "skipped")] += 1
        ws_out.append(list(r) + new_cols)


# ---------- CSV emit ------------------------------------------------------

def emit_csv(ws_out, csv_path):
    rows = list(ws_out.iter_rows(values_only=True))
    if not rows:
        return
    with open(csv_path, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        for r in rows:
            w.writerow([("" if v is None else v) for v in r])


# ---------- Main ----------------------------------------------------------

def main():
    p = argparse.ArgumentParser()
    p.add_argument("--parquet", required=True)
    p.add_argument("--workbook", required=True)
    p.add_argument("--out-xlsx", required=True)
    p.add_argument("--out-csv-dir", required=True)
    args = p.parse_args()

    by_key, by_subset = load_overture_index_v2(args.parquet)

    print(f"[workbook] reading {args.workbook}")
    wb = load_workbook(args.workbook, read_only=True, data_only=True)
    out_wb = openpyxl.Workbook()
    out_wb.remove(out_wb.active)

    summary = Counter()

    for sheet_name in ("States (ISO 3166-2)", "Counties", "Municipality"):
        print(f"[process] {sheet_name}")
        ws_in = wb[sheet_name]
        ws_out = out_wb.create_sheet(title=sheet_name[:31])
        if sheet_name == "States (ISO 3166-2)":
            process_states(ws_in, ws_out, by_key, by_subset, summary)
        elif sheet_name == "Counties":
            process_counties(ws_in, ws_out, by_key, by_subset, summary)
        elif sheet_name == "Municipality":
            process_municipality(ws_in, ws_out, by_key, by_subset, summary)

    out_xlsx = Path(args.out_xlsx)
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    out_wb.save(out_xlsx)
    print(f"[saved] {out_xlsx}")

    csv_dir = Path(args.out_csv_dir)
    csv_dir.mkdir(parents=True, exist_ok=True)
    for sheet_name in ("States (ISO 3166-2)", "Counties", "Municipality"):
        csv_path = csv_dir / f"overture-audit-{sheet_name.split()[0].lower()}.csv"
        emit_csv(out_wb[sheet_name[:31]], csv_path)
        print(f"[saved] {csv_path}")

    # Console summary
    print("\n=== Match summary (country, sheet, status -> count) ===")
    by_country_sheet = defaultdict(Counter)
    for (cc, sn, status), n in summary.items():
        by_country_sheet[(cc, sn)][status] += n
    for (cc, sn) in sorted(by_country_sheet.keys()):
        c = by_country_sheet[(cc, sn)]
        total = sum(c.values())
        print(f"\n  {cc} / {sn} (total={total:,})")
        for k in ("exact", "suffix-strip", "suffix-add", "ascii-fold", "alias",
                  "multi", "unmatched", "skipped"):
            if c.get(k):
                print(f"    {k}: {c[k]:,}")


if __name__ == "__main__":
    main()

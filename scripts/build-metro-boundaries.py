"""Build metro boundary GeoJSON files using a per-country sheet routing.

The MetroAreas.xlsx workbook now carries the same four user-curated Overture
columns on BOTH the Counties sheet and the Municipality sheet:

  - Subtype           (Overture subtype, e.g. 'county', 'region', 'locality')
  - Admin Level       (e.g. 2)
  - Region            (ISO 3166-2, e.g. 'US-AL', 'CA-ON', 'MX-AGU', 'GB-SCT',
                       'FR-ARA')
  - Primary Name      (exact Overture primary name, e.g. 'Marshall County',
                       'Manicouagan', 'Aberdeen City', 'Abbeville')

Each country we support is mapped (via COUNTRY_SHEET_MAP) to exactly ONE
source sheet AND to one Overture parquet (via COUNTRY_PARQUET_MAP, which
falls back to SOURCE_PARQUET if a country isn't listed).

Initial routing:
  United States  -> counties      (SOURCE_PARQUET)
  Mexico         -> counties      (SOURCE_PARQUET)
  Japan          -> counties      (overture-JP.parquet)
  Netherlands    -> counties      (overture-NL.parquet)
  Canada         -> municipality  (SOURCE_PARQUET)
  United Kingdom -> municipality  (SOURCE_PARQUET)
  France         -> municipality  (overture-FR.parquet)
  Germany        -> municipality  (overture-DE.parquet)
  Italy          -> municipality  (overture-IT.parquet)
  Spain          -> municipality  (overture-ES.parquet)
  Poland         -> municipality  (overture-PL.parquet)
  Andorra        -> municipality  (SOURCE_PARQUET)
  San Marino     -> municipality  (SOURCE_PARQUET)
  Vatican City   -> counties      (overture-VA.parquet)

Incremental build (build cache):
  Each metro's polygon is the function of its sorted (region, subtype,
  primary) row set, its anchor (lat, lon) from metros.json, and a small
  set of script constants captured in SCRIPT_VERSION_HASH below. We hash
  those inputs per metro and store the hash in
  public/data/metro-boundaries/build-cache.json. (Filename intentionally
  has no leading dot - OneDrive and Defender on Windows treat dot-prefix
  files inconsistently and silently delete them in some configurations.)

  On each run, we compute the new input hashes BEFORE touching the
  parquet. Metros whose hash matches the cache AND whose GeoJSON is
  present on disk are skipped entirely. The parquet scan runs only over
  the keys needed by metros that actually need rebuilding. If no metros
  need rebuilding and no stale slugs need pruning, the script exits in
  seconds without scanning the parquet at all.

  Pass --force to bypass the cache and rebuild everything.

  Bump SCRIPT_VERSION_HASH manually (or change any constant it includes)
  to force a global rebuild from the next run on. The hash is derived
  from the constants automatically, so bumping OUTLIER_PART_MAX_KM (for
  example) invalidates all cached metros without manual intervention.

Refreshment protocol (age-based invalidation):
  Each cache entry stores a `built_at` ISO timestamp alongside its hash.
  A metro is rebuilt if any of these are true:
    - Its (region, subtype, primary) member set or anchor changed (hash
      mismatch).
    - Its GeoJSON file is missing on disk.
    - Its cached built_at is older than DEFAULT_MAX_AGE_DAYS.
    - Its cache entry is in the legacy plain-string format (no built_at
      recorded), which means it predates the refreshment protocol.

  Override the age threshold with --max-age-days N. Use --max-age-days 0
  to rebuild every metro whose cache entry is at least one second old,
  which is effectively the same as --force but cheaper (it preserves
  cache structure and built_at recording for the next run).

  Recommended usage:
    - Daily incremental: no flags. Rebuilds only metros with hash drift.
    - Weekly refresh: --max-age-days 7. Picks up stale builds caused by
      script-logic drift the SCRIPT_VERSION_HASH check missed.
    - Monthly deep refresh: --max-age-days 30. Catches anything quirky.
  The boundary build runs locally because the Overture parquets live on
  the user's workstation, not in CI. Helper: scripts/refresh-boundaries.ps1
  wraps the weekly variant for one-click invocation.

To extend to a new country:
  1. Pick the sheet that holds its rows (Counties or Municipality).
  2. Populate the four Overture columns by hand for those rows.
  3. Add one entry each to COUNTRY_SHEET_MAP, COUNTRY_TO_ISO,
     COUNTRY_PARQUET_MAP. If the workbook stores the country under
     multiple constituent names, add WORKBOOK_TO_CANONICAL_COUNTRY entries.
  4. Run scripts/extract-overture-parquet.py to produce the per-country
     runtime parquet.

Outlier-part trim:
  After unioning a metro's member polygons, parts of the resulting
  MultiPolygon whose minimum distance from the anchor exceeds
  OUTLIER_PART_MAX_KM are dropped. Trims off Honolulu's NWHI tail and
  Tokyo's Izu/Ogasawara without harming NYC-scale metros.

Behavior:
  - Reads build-cache.json; computes new hashes; skips unchanged metros.
  - Wipes only stale slugs (no longer in workbook) from
    public/data/metro-boundaries/.
  - Writes one GeoJSON per metro that actually rebuilt.
  - Writes the updated cache file at the end.

Dependencies:
  pip install geopandas openpyxl pyarrow

Source parquet path defaults to the user's local layout but is overridable
via the OVERTURE_DIVISION_AREA env var.
"""
from __future__ import annotations

import hashlib
import json
import os
import sys
import time

# Windows console default is cp1252, which cannot encode Cyrillic, Greek,
# Arabic, or many accented Latin characters that appear in Overture primary
# names and the workbook's diagnostic rows (unmatched-members summary, etc).
# Reconfigure stdout/stderr to utf-8 with errors=replace so prints never
# crash mid-summary on a non-encodable glyph. Python 3.7+ has reconfigure().
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
from collections import defaultdict
from math import asin, cos, radians, sin, sqrt
from pathlib import Path

import openpyxl
from shapely.geometry import MultiPolygon, Point, mapping
from shapely.geometry import shape as _shapely_shape
from shapely.ops import nearest_points, unary_union


SOURCE_PARQUET = os.environ.get(
    "OVERTURE_DIVISION_AREA",
    r"C:\Users\ashwi\Desktop\Projects\MapData\global-division-area.parquet",
)
WORKBOOK = "MetroAreas.xlsx"
METROS_JSON = "public/data/metros.json"
OUT_DIR = Path("public/data/metro-boundaries")
BUILD_CACHE_FILE = OUT_DIR / "build-cache.json"
SIMPLIFY_TOLERANCE_DEG = 0.005
# Each member polygon is simplified to this tolerance BEFORE the per-metro
# unary_union. Cuts vertex count 5-10x on dense commune sets (Paris 1,563
# communes, Bordeaux 534, etc.) without visible quality loss at metro zoom.
# Set to 0 to disable pre-simplification.
MEMBER_SIMPLIFY_TOLERANCE_DEG = 0.001
OUTLIER_PART_MAX_KM = 200.0

# Refreshment protocol: any cached polygon older than this is rebuilt even
# if its hash has not changed. Catches drift from earlier script logic that
# the SCRIPT_VERSION_HASH check missed (e.g., the 2026-05-06 Tokyo build that
# excluded the 23 special wards). Override at the CLI with --max-age-days N.
DEFAULT_MAX_AGE_DAYS = 30


# ---------- Per-country parquet routing ---------------------------------
COUNTRY_PARQUET_MAP = {
    # Per-country parquets keep heavy commune scans off the 5.8 GB global
    # file. Generate via scripts/extract-overture-parquet.py.
    "France":      r"C:\Users\ashwi\Desktop\Projects\MapData\overture-FR.parquet",
    "Germany":     r"C:\Users\ashwi\Desktop\Projects\MapData\overture-DE.parquet",
    "Italy":       r"C:\Users\ashwi\Desktop\Projects\MapData\overture-IT.parquet",
    "Spain":       r"C:\Users\ashwi\Desktop\Projects\MapData\overture-ES.parquet",
    "Poland":      r"C:\Users\ashwi\Desktop\Projects\MapData\overture-PL.parquet",
    "Japan":       r"C:\Users\ashwi\Desktop\Projects\MapData\overture-JP.parquet",
    "Netherlands": r"C:\Users\ashwi\Desktop\Projects\MapData\overture-NL.parquet",
    # 2026-05-08 expansion: high-volume countries that warrant their own
    # per-country parquet. Generate each via scripts/extract-overture-parquet.py
    # before the next boundary build.
    "Portugal":    r"C:\Users\ashwi\Desktop\Projects\MapData\overture-PT.parquet",
    "Romania":     r"C:\Users\ashwi\Desktop\Projects\MapData\overture-RO.parquet",
    "Colombia":    r"C:\Users\ashwi\Desktop\Projects\MapData\overture-CO.parquet",
    "Turkey":      r"C:\Users\ashwi\Desktop\Projects\MapData\overture-TR.parquet",
    "Nigeria":     r"C:\Users\ashwi\Desktop\Projects\MapData\overture-NG.parquet",
    "India":       r"C:\Users\ashwi\Desktop\Projects\MapData\overture-IN.parquet",
    "Croatia":     r"C:\Users\ashwi\Desktop\Projects\MapData\overture-HR.parquet",
    "Congo DR":    r"C:\Users\ashwi\Desktop\Projects\MapData\overture-CD.parquet",
    "Luxembourg":  r"C:\Users\ashwi\Desktop\Projects\MapData\overture-LU.parquet",
    # 2026-05-10 expansion (six countries with workbook fills complete):
    "Brazil":      r"C:\Users\ashwi\Desktop\Projects\MapData\overture-BR.parquet",
    "Austria":     r"C:\Users\ashwi\Desktop\Projects\MapData\overture-AT.parquet",
    "Switzerland": r"C:\Users\ashwi\Desktop\Projects\MapData\overture-CH.parquet",
    "China":       r"C:\Users\ashwi\Desktop\Projects\MapData\overture-CN.parquet",
    "Australia":   r"C:\Users\ashwi\Desktop\Projects\MapData\overture-AU.parquet",
    "South Korea": r"C:\Users\ashwi\Desktop\Projects\MapData\overture-KR.parquet",
    # 2026-05-19 expansion (three countries with workbook fills complete):
    "Russia":      r"C:\Users\ashwi\Desktop\Projects\MapData\overture-RU.parquet",
    "Ireland":     r"C:\Users\ashwi\Desktop\Projects\MapData\overture-IE.parquet",
    "Belgium":     r"C:\Users\ashwi\Desktop\Projects\MapData\overture-BE.parquet",
    # 2026-05-20 expansion (Singapore, Malta, Liechtenstein, Vatican City):
    "Singapore":     r"C:\Users\ashwi\Desktop\Projects\MapData\overture-SG.parquet",
    "Malta":         r"C:\Users\ashwi\Desktop\Projects\MapData\overture-MT.parquet",
    "Liechtenstein": r"C:\Users\ashwi\Desktop\Projects\MapData\overture-LI.parquet",
    "Vatican City":  r"C:\Users\ashwi\Desktop\Projects\MapData\overture-VA.parquet",
    # 2026-05-21 expansion - workbook fills complete:
    "Argentina":     r"C:\Users\ashwi\Desktop\Projects\MapData\overture-AR.parquet",
    "Bulgaria":      r"C:\Users\ashwi\Desktop\Projects\MapData\overture-BG.parquet",
    "Denmark":       r"C:\Users\ashwi\Desktop\Projects\MapData\overture-DK.parquet",
    "Puerto Rico":   r"C:\Users\ashwi\Desktop\Projects\MapData\overture-PR.parquet",
    # 2026-05-22 expansion - workbook fills complete via combined-matcher run:
    "American Samoa": r"C:\Users\ashwi\Desktop\Projects\MapData\overture-AS.parquet",
    "Bahamas":      r"C:\Users\ashwi\Desktop\Projects\MapData\overture-BS.parquet",
    "Bangladesh":   r"C:\Users\ashwi\Desktop\Projects\MapData\overture-BD.parquet",
    "Belize":       r"C:\Users\ashwi\Desktop\Projects\MapData\overture-BZ.parquet",
    "Cape Verde":   r"C:\Users\ashwi\Desktop\Projects\MapData\overture-CV.parquet",
    "Cayman Islands": r"C:\Users\ashwi\Desktop\Projects\MapData\overture-KY.parquet",
    "Chile":        r"C:\Users\ashwi\Desktop\Projects\MapData\overture-CL.parquet",
    "East Timor":   r"C:\Users\ashwi\Desktop\Projects\MapData\overture-TL.parquet",
    "Equatorial Guinea": r"C:\Users\ashwi\Desktop\Projects\MapData\overture-GQ.parquet",
    "Eswatini":     r"C:\Users\ashwi\Desktop\Projects\MapData\overture-SZ.parquet",
    "Federated States of Micronesia": r"C:\Users\ashwi\Desktop\Projects\MapData\overture-FM.parquet",
    "Guadeloupe":   r"C:\Users\ashwi\Desktop\Projects\MapData\overture-GP.parquet",
    "Guinea-Bissau": r"C:\Users\ashwi\Desktop\Projects\MapData\overture-GW.parquet",
    "Guyana":       r"C:\Users\ashwi\Desktop\Projects\MapData\overture-GY.parquet",
    "Hong Kong":    r"C:\Users\ashwi\Desktop\Projects\MapData\overture-HK.parquet",
    "Isle of Man":  r"C:\Users\ashwi\Desktop\Projects\MapData\overture-IM.parquet",
    "Jordan":       r"C:\Users\ashwi\Desktop\Projects\MapData\overture-JO.parquet",
    "Kiribati":     r"C:\Users\ashwi\Desktop\Projects\MapData\overture-KI.parquet",
    "Latvia":       r"C:\Users\ashwi\Desktop\Projects\MapData\overture-LV.parquet",
    "Liberia":      r"C:\Users\ashwi\Desktop\Projects\MapData\overture-LR.parquet",
    "Marshall Islands": r"C:\Users\ashwi\Desktop\Projects\MapData\overture-MH.parquet",
    "Martinique":   r"C:\Users\ashwi\Desktop\Projects\MapData\overture-MQ.parquet",
    "Mayotte":      r"C:\Users\ashwi\Desktop\Projects\MapData\overture-YT.parquet",
    "Montenegro":   r"C:\Users\ashwi\Desktop\Projects\MapData\overture-ME.parquet",
    "Nauru":        r"C:\Users\ashwi\Desktop\Projects\MapData\overture-NR.parquet",
    "Nepal":        r"C:\Users\ashwi\Desktop\Projects\MapData\overture-NP.parquet",
    "New Zealand":  r"C:\Users\ashwi\Desktop\Projects\MapData\overture-NZ.parquet",
    "Niue":         r"C:\Users\ashwi\Desktop\Projects\MapData\overture-NU.parquet",
    "Northern Mariana Islands": r"C:\Users\ashwi\Desktop\Projects\MapData\overture-MP.parquet",
    "Papua New Guinea": r"C:\Users\ashwi\Desktop\Projects\MapData\overture-PG.parquet",
    "Philippines":  r"C:\Users\ashwi\Desktop\Projects\MapData\overture-PH.parquet",
    "Réunion":      r"C:\Users\ashwi\Desktop\Projects\MapData\overture-RE.parquet",
    "São Tomé and Príncipe": r"C:\Users\ashwi\Desktop\Projects\MapData\overture-ST.parquet",
    "Senegal":      r"C:\Users\ashwi\Desktop\Projects\MapData\overture-SN.parquet",
    "Serbia":       r"C:\Users\ashwi\Desktop\Projects\MapData\overture-RS.parquet",
    "Sri Lanka":    r"C:\Users\ashwi\Desktop\Projects\MapData\overture-LK.parquet",
    "Suriname":     r"C:\Users\ashwi\Desktop\Projects\MapData\overture-SR.parquet",
    "Thailand":     r"C:\Users\ashwi\Desktop\Projects\MapData\overture-TH.parquet",
    "Tokelau":      r"C:\Users\ashwi\Desktop\Projects\MapData\overture-TK.parquet",
    "Tuvalu":       r"C:\Users\ashwi\Desktop\Projects\MapData\overture-TV.parquet",
    "Uruguay":      r"C:\Users\ashwi\Desktop\Projects\MapData\overture-UY.parquet",
    # 2026-05-21 expansion - workbook fills PENDING (matcher run required):
    "Algeria":       r"C:\Users\ashwi\Desktop\Projects\MapData\overture-DZ.parquet",
    "Egypt":         r"C:\Users\ashwi\Desktop\Projects\MapData\overture-EG.parquet",
    "Finland":       r"C:\Users\ashwi\Desktop\Projects\MapData\overture-FI.parquet",
    "Ghana":         r"C:\Users\ashwi\Desktop\Projects\MapData\overture-GH.parquet",
    "Greece":        r"C:\Users\ashwi\Desktop\Projects\MapData\overture-GR.parquet",
    "Hungary":       r"C:\Users\ashwi\Desktop\Projects\MapData\overture-HU.parquet",
    "Indonesia":     r"C:\Users\ashwi\Desktop\Projects\MapData\overture-ID.parquet",
    "Israel":        r"C:\Users\ashwi\Desktop\Projects\MapData\overture-IL.parquet",
    "Kuwait":        r"C:\Users\ashwi\Desktop\Projects\MapData\overture-KW.parquet",
    "Moldova":       r"C:\Users\ashwi\Desktop\Projects\MapData\overture-MD.parquet",
    "Qatar":         r"C:\Users\ashwi\Desktop\Projects\MapData\overture-QA.parquet",
    "Sweden":        r"C:\Users\ashwi\Desktop\Projects\MapData\overture-SE.parquet",
    "Tunisia":       r"C:\Users\ashwi\Desktop\Projects\MapData\overture-TN.parquet",
    "Taiwan":        r"C:\Users\ashwi\Desktop\Projects\MapData\overture-TW.parquet",
    "Venezuela":     r"C:\Users\ashwi\Desktop\Projects\MapData\overture-VE.parquet",
    "Vietnam":       r"C:\Users\ashwi\Desktop\Projects\MapData\overture-VN.parquet",
    "Samoa":         r"C:\Users\ashwi\Desktop\Projects\MapData\overture-WS.parquet",
    # 2026-05-23 expansion - workbook fills landed:
    "Afghanistan":   r"C:\Users\ashwi\Desktop\Projects\MapData\overture-AF.parquet",
    "Azerbaijan":    r"C:\Users\ashwi\Desktop\Projects\MapData\overture-AZ.parquet",
    "Iran":          r"C:\Users\ashwi\Desktop\Projects\MapData\overture-IR.parquet",
    "Iraq":          r"C:\Users\ashwi\Desktop\Projects\MapData\overture-IQ.parquet",
    # 2026-05-23 PM expansion - Middle East + UAE locality-first wiring:
    "Bahrain":               r"C:\Users\ashwi\Desktop\Projects\MapData\overture-BH.parquet",
    "Lebanon":               r"C:\Users\ashwi\Desktop\Projects\MapData\overture-LB.parquet",
    "Oman":                  r"C:\Users\ashwi\Desktop\Projects\MapData\overture-OM.parquet",
    "Palestine":             r"C:\Users\ashwi\Desktop\Projects\MapData\overture-PS.parquet",  # XW + XG combined
    "Saudi Arabia":          r"C:\Users\ashwi\Desktop\Projects\MapData\overture-SA.parquet",
    "Syria":                 r"C:\Users\ashwi\Desktop\Projects\MapData\overture-SY.parquet",
    "United Arab Emirates":  r"C:\Users\ashwi\Desktop\Projects\MapData\overture-AE.parquet",
    # 2026-05-24 expansion - Bonaire's two Dutch Caribbean cousins:
    "Saba":                  r"C:\Users\ashwi\Desktop\Projects\MapData\overture-XS.parquet",
    "Sint Eustatius":        r"C:\Users\ashwi\Desktop\Projects\MapData\overture-XE.parquet",
    # 2026-05-24 expansion - newly keyed countries from workbook sync:
    "South Africa":          r"C:\Users\ashwi\Desktop\Projects\MapData\overture-ZA.parquet",
    "Belarus":               r"C:\Users\ashwi\Desktop\Projects\MapData\overture-BY.parquet",
    "Ukraine":               r"C:\Users\ashwi\Desktop\Projects\MapData\overture-UA.parquet",
    # 2026-05-24 PM expansion - French overseas + Monaco + Channel:
    "French Guiana":             r"C:\Users\ashwi\Desktop\Projects\MapData\overture-GF.parquet",
    "Monaco":                    r"C:\Users\ashwi\Desktop\Projects\MapData\overture-MC.parquet",
    "Saint Pierre and Miquelon": r"C:\Users\ashwi\Desktop\Projects\MapData\overture-PM.parquet",
    # 2026-05-24 PM expansion - UK-linked territories + Macau + US territories:
    "Jersey":                                       r"C:\Users\ashwi\Desktop\Projects\MapData\overture-JE.parquet",
    "Guernsey":                                     r"C:\Users\ashwi\Desktop\Projects\MapData\overture-GG.parquet",
    "Bermuda":                                      r"C:\Users\ashwi\Desktop\Projects\MapData\overture-BM.parquet",
    "Turks & Caicos Islands":                       r"C:\Users\ashwi\Desktop\Projects\MapData\overture-TC.parquet",
    "British Virgin Islands":                       r"C:\Users\ashwi\Desktop\Projects\MapData\overture-VG.parquet",
    "Gibraltar":                                    r"C:\Users\ashwi\Desktop\Projects\MapData\overture-GI.parquet",
    "Anguilla":                                     r"C:\Users\ashwi\Desktop\Projects\MapData\overture-AI.parquet",
    "Saint Helena, Ascension and Tristan da Cunha": r"C:\Users\ashwi\Desktop\Projects\MapData\overture-SH.parquet",
    "Montserrat":                                   r"C:\Users\ashwi\Desktop\Projects\MapData\overture-MS.parquet",
    "Falkland Islands":                             r"C:\Users\ashwi\Desktop\Projects\MapData\overture-FK.parquet",
    "Macau":                                        r"C:\Users\ashwi\Desktop\Projects\MapData\overture-MO.parquet",
    "Guam":                                         r"C:\Users\ashwi\Desktop\Projects\MapData\overture-GU.parquet",
    "US Virgin Islands":                            r"C:\Users\ashwi\Desktop\Projects\MapData\overture-VI.parquet",
    # 2026-05-24 late - additional newly-filled small countries:
    "Saint Lucia":                                  r"C:\Users\ashwi\Desktop\Projects\MapData\overture-LC.parquet",
    "Saint Martin":                                 r"C:\Users\ashwi\Desktop\Projects\MapData\overture-MF.parquet",
    "Antigua & Barbuda":                            r"C:\Users\ashwi\Desktop\Projects\MapData\overture-AG.parquet",
    "Solomon Islands":                              r"C:\Users\ashwi\Desktop\Projects\MapData\overture-SB.parquet",
    "Saint Barthélemy":                             r"C:\Users\ashwi\Desktop\Projects\MapData\overture-BL.parquet",
    "Sint Maarten":                                 r"C:\Users\ashwi\Desktop\Projects\MapData\overture-SX.parquet",
    "Bonaire":                                      r"C:\Users\ashwi\Desktop\Projects\MapData\overture-BQ.parquet",
    # 2026-05-24 - Western Sahara files under MA in Overture (no per-country
    # EH parquet exists). Rows carry MA-11/MA-12/MA-13 regions; the
    # iso_pref routing already handles those. Es-Semara row has region=None
    # so the ISO override is needed for the parquet_iso filter.
    "Western Sahara":                               r"C:\Users\ashwi\Desktop\Projects\MapData\overture-MA.parquet",
    # 2026-05-24 evening - mass wiring of countries with workbook fills but missing builder entries:
    "Botswana":                       r"C:\Users\ashwi\Desktop\Projects\MapData\overture-BW.parquet",
    "Chad":                           r"C:\Users\ashwi\Desktop\Projects\MapData\overture-TD.parquet",
    "Comoros":                        r"C:\Users\ashwi\Desktop\Projects\MapData\overture-KM.parquet",
    "Cook Islands":                   r"C:\Users\ashwi\Desktop\Projects\MapData\overture-CK.parquet",
    "Curaçao":                        r"C:\Users\ashwi\Desktop\Projects\MapData\overture-CW.parquet",
    "Czech Republic":                 r"C:\Users\ashwi\Desktop\Projects\MapData\overture-CZ.parquet",
    "Djibouti":                       r"C:\Users\ashwi\Desktop\Projects\MapData\overture-DJ.parquet",
    "Dominica":                       r"C:\Users\ashwi\Desktop\Projects\MapData\overture-DM.parquet",
    "Estonia":                        r"C:\Users\ashwi\Desktop\Projects\MapData\overture-EE.parquet",
    "Faroe Islands":                  r"C:\Users\ashwi\Desktop\Projects\MapData\overture-FO.parquet",
    "Georgia":                        r"C:\Users\ashwi\Desktop\Projects\MapData\overture-GE.parquet",
    "Grenada":                        r"C:\Users\ashwi\Desktop\Projects\MapData\overture-GD.parquet",
    "Iceland":                        r"C:\Users\ashwi\Desktop\Projects\MapData\overture-IS.parquet",
    "Jamaica":                        r"C:\Users\ashwi\Desktop\Projects\MapData\overture-JM.parquet",
    "Kazakhstan":                     r"C:\Users\ashwi\Desktop\Projects\MapData\overture-KZ.parquet",
    "Kosovo":                         r"C:\Users\ashwi\Desktop\Projects\MapData\overture-XK.parquet",
    "Kyrgyzstan":                     r"C:\Users\ashwi\Desktop\Projects\MapData\overture-KG.parquet",
    "Laos":                           r"C:\Users\ashwi\Desktop\Projects\MapData\overture-LA.parquet",
    "Libya":                          r"C:\Users\ashwi\Desktop\Projects\MapData\overture-LY.parquet",
    "Lithuania":                      r"C:\Users\ashwi\Desktop\Projects\MapData\overture-LT.parquet",
    "Madagascar":                     r"C:\Users\ashwi\Desktop\Projects\MapData\overture-MG.parquet",
    "Maldives":                       r"C:\Users\ashwi\Desktop\Projects\MapData\overture-MV.parquet",
    "Mali":                           r"C:\Users\ashwi\Desktop\Projects\MapData\overture-ML.parquet",
    "Mauritania":                     r"C:\Users\ashwi\Desktop\Projects\MapData\overture-MR.parquet",
    "Mauritius":                      r"C:\Users\ashwi\Desktop\Projects\MapData\overture-MU.parquet",
    "Mongolia":                       r"C:\Users\ashwi\Desktop\Projects\MapData\overture-MN.parquet",
    "Morocco":                        r"C:\Users\ashwi\Desktop\Projects\MapData\overture-MA.parquet",
    "Mozambique":                     r"C:\Users\ashwi\Desktop\Projects\MapData\overture-MZ.parquet",
    "Myanmar":                        r"C:\Users\ashwi\Desktop\Projects\MapData\overture-MM.parquet",
    "Namibia":                        r"C:\Users\ashwi\Desktop\Projects\MapData\overture-NA.parquet",
    "New Caledonia":                  r"C:\Users\ashwi\Desktop\Projects\MapData\overture-NC.parquet",
    "North Macedonia":                r"C:\Users\ashwi\Desktop\Projects\MapData\overture-MK.parquet",
    "Palau":                          r"C:\Users\ashwi\Desktop\Projects\MapData\overture-PW.parquet",
    "Paraguay":                       r"C:\Users\ashwi\Desktop\Projects\MapData\overture-PY.parquet",
    "Rwanda":                         r"C:\Users\ashwi\Desktop\Projects\MapData\overture-RW.parquet",
    "Seychelles":                     r"C:\Users\ashwi\Desktop\Projects\MapData\overture-SC.parquet",
    "Slovakia":                       r"C:\Users\ashwi\Desktop\Projects\MapData\overture-SK.parquet",
    "Slovenia":                       r"C:\Users\ashwi\Desktop\Projects\MapData\overture-SI.parquet",
    "St. Vincent & the Grenadines":   r"C:\Users\ashwi\Desktop\Projects\MapData\overture-VC.parquet",
    "Tahiti":                         r"C:\Users\ashwi\Desktop\Projects\MapData\overture-PF.parquet",
    "Tajikistan":                     r"C:\Users\ashwi\Desktop\Projects\MapData\overture-TJ.parquet",
    "Tanzania":                       r"C:\Users\ashwi\Desktop\Projects\MapData\overture-TZ.parquet",
    "Togo":                           r"C:\Users\ashwi\Desktop\Projects\MapData\overture-TG.parquet",
    "Tonga":                          r"C:\Users\ashwi\Desktop\Projects\MapData\overture-TO.parquet",
    "Trinidad & Tobago":              r"C:\Users\ashwi\Desktop\Projects\MapData\overture-TT.parquet",
    "Turkmenistan":                   r"C:\Users\ashwi\Desktop\Projects\MapData\overture-TM.parquet",
    "Uzbekistan":                     r"C:\Users\ashwi\Desktop\Projects\MapData\overture-UZ.parquet",
    "Wallis and Futuna":              r"C:\Users\ashwi\Desktop\Projects\MapData\overture-WF.parquet",

    # 2026-05-24 evening expansion - workbook fills PENDING (matcher run required):
    "Albania":                       r"C:\Users\ashwi\Desktop\Projects\MapData\overture-AL.parquet",
    "Armenia":                       r"C:\Users\ashwi\Desktop\Projects\MapData\overture-AM.parquet",
    "Aruba":                         r"C:\Users\ashwi\Desktop\Projects\MapData\overture-AW.parquet",
    "Bolivia":                       r"C:\Users\ashwi\Desktop\Projects\MapData\overture-BO.parquet",
    "Bosnia-Herzegovina":            r"C:\Users\ashwi\Desktop\Projects\MapData\overture-BA.parquet",
    "Cameroon":                      r"C:\Users\ashwi\Desktop\Projects\MapData\overture-CM.parquet",
    "Congo":                         r"C:\Users\ashwi\Desktop\Projects\MapData\overture-CG.parquet",
    "Cyprus":                        r"C:\Users\ashwi\Desktop\Projects\MapData\overture-CY.parquet",
    "Ethiopia":                      r"C:\Users\ashwi\Desktop\Projects\MapData\overture-ET.parquet",
    "North Korea":                   r"C:\Users\ashwi\Desktop\Projects\MapData\overture-KP.parquet",
    "Yemen":                         r"C:\Users\ashwi\Desktop\Projects\MapData\overture-YE.parquet",
    "Zambia":                        r"C:\Users\ashwi\Desktop\Projects\MapData\overture-ZM.parquet",
    "Zimbabwe":                      r"C:\Users\ashwi\Desktop\Projects\MapData\overture-ZW.parquet",

    # Andorra, San Marino, and the remaining 2026-05-08 small countries
    # (most of Latin America, sub-Saharan Africa, the Channel Islands, etc.)
    # are tiny enough to fall through to SOURCE_PARQUET; no per-country
    # parquet needed.
}


# ---------- Per-country sheet routing -----------------------------------
# Each country routes to exactly one sheet. The Overture lookup columns
# (Subtype, Admin Level, Region ISO 3166-2, Primary Name) are filled on
# only that sheet for the country. Routing was audited 2026-05-08: zero
# countries have fills in both sheets.
COUNTRY_SHEET_MAP = {
    # Pre-2026-05-08 set (14)
    "United States":            "counties",
    "Mexico":                   "counties",
    "Japan":                    "counties",
    "Netherlands":              "counties",
    "Canada":                   "municipality",
    "United Kingdom":           "municipality",
    "France":                   "municipality",
    "Germany":                  "municipality",
    "Italy":                    "municipality",
    "Spain":                    "municipality",
    "Poland":                   "municipality",
    "Andorra":                  "municipality",
    "San Marino":               "municipality",
    "Vatican City":             "counties",
    # 2026-05-08 expansion - Counties (38)
    "Romania":                  "counties",
    "Colombia":                 "counties",
    "Turkey":                   "counties",
    "Nigeria":                  "counties",
    "India":                    "counties",
    "Croatia":                  "counties",
    "Guatemala":                "counties",
    "Honduras":                 "counties",
    "El Salvador":              "counties",
    "Norway":                   "counties",
    "Kenya":                    "counties",
    "Ecuador":                  "counties",
    "Peru":                     "counties",
    "Sudan":                    "counties",
    "Cambodia":                 "counties",
    "Cuba":                     "counties",
    "Angola":                   "counties",
    "Dominican Republic":       "counties",
    "Malaysia":                 "counties",
    "Uganda":                   "counties",
    "Nicaragua":                "counties",
    "Costa Rica":               "counties",
    "Benin":                    "counties",
    "Panama":                   "counties",
    "South Sudan":              "counties",
    "Central African Republic": "counties",
    "Vanuatu":                  "counties",
    "Burkina Faso":             "counties",
    "Gabon":                    "counties",
    "Gambia":                   "counties",
    "Brunei":                   "counties",
    "Eritrea":                  "counties",
    "Guinea":                   "counties",
    "Burundi":                  "counties",
    "Malawi":                   "counties",
    "Somalia":                  "counties",
    "St. Kitts & Nevis":        "counties",
    "Lesotho":                  "counties",
    # 2026-05-08 expansion - Municipality (5)
    "Portugal":                 "municipality",
    "Luxembourg":               "municipality",
    "Congo DR":                 "municipality",
    "Guernsey":                 "counties",
    "Jersey":                   "counties",
    # 2026-05-10 expansion - Counties (3)
    "Brazil":                   "counties",
    "Australia":                "counties",
    "South Korea":              "counties",
    # 2026-05-10 expansion - Municipality (3)
    "Austria":                  "municipality",
    "Switzerland":              "municipality",
    "China":                    "municipality",
    # 2026-05-19 expansion - Counties (1)
    "Russia":                   "counties",
    # 2026-05-19 expansion - Municipality (2)
    "Ireland":                  "municipality",
    "Belgium":                  "municipality",
    # 2026-05-20 expansion - Counties (1)
    "Singapore":                "counties",
    # 2026-05-20 expansion - Municipality (2)
    "Malta":                    "municipality",
    "Liechtenstein":            "municipality",
    # 2026-05-21 expansion - Counties (workbook fills complete: 4)
    "Argentina":                "counties",
    "Bulgaria":                 "counties",
    "Denmark":                  "counties",
    "Puerto Rico":              "counties",
    # 2026-05-22 expansion - Counties (42 countries)
    "American Samoa":           "counties",
    "Bahamas":                  "counties",
    "Bangladesh":               "counties",
    "Belize":                   "counties",
    "Cape Verde":               "counties",
    "Cayman Islands":           "counties",
    "Chile":                    "counties",
    "East Timor":               "counties",
    "Equatorial Guinea":        "counties",
    "Eswatini":                 "counties",
    "Federated States of Micronesia": "counties",
    "Greenland":                "counties",
    "Guadeloupe":               "counties",
    "Guinea-Bissau":            "counties",
    "Guyana":                   "counties",
    "Hong Kong":                "counties",
    "Isle of Man":              "counties",
    "Jordan":                   "counties",
    "Kiribati":                 "counties",
    "Latvia":                   "counties",
    "Liberia":                  "counties",
    "Marshall Islands":         "counties",
    "Martinique":               "counties",
    "Mayotte":                  "counties",
    "Montenegro":               "counties",
    "Nauru":                    "counties",
    "Nepal":                    "counties",
    "New Zealand":              "counties",
    "Niue":                     "counties",
    "Northern Mariana Islands": "counties",
    "Papua New Guinea":         "counties",
    "Philippines":              "counties",
    "Réunion":                  "counties",
    "São Tomé and Príncipe":    "counties",
    "Senegal":                  "counties",
    "Serbia":                   "counties",
    "Sri Lanka":                "counties",
    "Suriname":                 "counties",
    "Thailand":                 "counties",
    "Tokelau":                  "counties",
    "Tuvalu":                   "counties",
    "Uruguay":                  "counties",
    # 2026-05-21 expansion - Counties (matcher pending: 17)
    "Algeria":                  "counties",
    "Egypt":                    "counties",
    "Finland":                  "counties",
    "Ghana":                    "counties",
    "Greece":                   "counties",
    "Hungary":                  "counties",
    "Indonesia":                "counties",
    "Israel":                   "counties",
    "Kuwait":                   "counties",
    "Moldova":                  "counties",
    "Qatar":                    "counties",
    "Sweden":                   "counties",
    "Tunisia":                  "counties",
    "Taiwan":                   "counties",
    "Venezuela":                "counties",
    "Vietnam":                  "counties",
    "Samoa":                    "counties",
    # 2026-05-23 expansion
    "Afghanistan":              "counties",
    "Azerbaijan":               "counties",
    "Iran":                     "counties",
    "Iraq":                     "counties",
    # 2026-05-23 PM expansion
    "Bahrain":                  "counties",
    "Lebanon":                  "counties",
    "Oman":                     "counties",
    "Palestine":                "counties",
    "Saudi Arabia":             "counties",
    "Syria":                    "counties",
    "United Arab Emirates":     "counties",
    # 2026-05-24 expansion
    "Saba":                     "counties",
    "Sint Eustatius":           "counties",
    "South Africa":             "counties",
    "Belarus":                  "counties",
    "Ukraine":                  "counties",
    "French Guiana":             "counties",
    "Monaco":                    "counties",
    "Saint Pierre and Miquelon": "counties",
    "Bermuda":                                      "counties",
    "Turks & Caicos Islands":                       "counties",
    "British Virgin Islands":                       "counties",
    "Gibraltar":                                    "counties",
    "Anguilla":                                     "counties",
    "Saint Helena, Ascension and Tristan da Cunha": "counties",
    "Montserrat":                                   "counties",
    "Falkland Islands":                             "counties",
    "Macau":                                        "counties",
    "Guam":                                         "counties",
    "US Virgin Islands":                            "counties",
    "Saint Lucia":                                  "counties",
    "Saint Martin":                                 "counties",
    "Antigua & Barbuda":                            "counties",
    "Solomon Islands":                              "counties",
    "Saint Barthélemy":                             "counties",
    "Sint Maarten":                                 "counties",
    "Bonaire":                                      "counties",
    "Western Sahara":                               "counties",
    # 2026-05-24 evening mass wiring:
    "Botswana":                       "counties",
    "Chad":                           "counties",
    "Comoros":                        "counties",
    "Cook Islands":                   "counties",
    "Curaçao":                        "counties",
    "Czech Republic":                 "counties",
    "Djibouti":                       "counties",
    "Dominica":                       "counties",
    "Estonia":                        "counties",
    "Faroe Islands":                  "counties",
    "Georgia":                        "counties",
    "Grenada":                        "counties",
    "Iceland":                        "counties",
    "Jamaica":                        "counties",
    "Kazakhstan":                     "counties",
    "Kosovo":                         "counties",
    "Kyrgyzstan":                     "counties",
    "Laos":                           "counties",
    "Libya":                          "counties",
    "Lithuania":                      "counties",
    "Madagascar":                     "counties",
    "Maldives":                       "counties",
    "Mali":                           "counties",
    "Mauritania":                     "counties",
    "Mauritius":                      "counties",
    "Mongolia":                       "counties",
    "Morocco":                        "counties",
    "Mozambique":                     "counties",
    "Myanmar":                        "counties",
    "Namibia":                        "counties",
    "New Caledonia":                  "counties",
    "North Macedonia":                "counties",
    "Palau":                          "counties",
    "Paraguay":                       "counties",
    "Rwanda":                         "counties",
    "Seychelles":                     "counties",
    "Slovakia":                       "counties",
    "Slovenia":                       "counties",
    "St. Vincent & the Grenadines":   "counties",
    "Tahiti":                         "counties",
    "Tajikistan":                     "counties",
    "Tanzania":                       "counties",
    "Togo":                           "counties",
    "Tonga":                          "counties",
    "Trinidad & Tobago":              "counties",
    "Turkmenistan":                   "counties",
    "Uzbekistan":                     "counties",
    "Wallis and Futuna":              "counties",

    # 2026-05-24 evening expansion - workbook fills PENDING (matcher run required):
    "Albania":                       "counties",
    "Armenia":                       "counties",
    "Aruba":                         "counties",
    "Bolivia":                       "counties",
    "Bosnia-Herzegovina":            "counties",
    "Cameroon":                      "counties",
    "Congo":                         "counties",
    "Cyprus":                        "counties",
    "Ethiopia":                      "counties",
    "North Korea":                   "counties",
    "Yemen":                         "counties",
    "Zambia":                        "counties",
    "Zimbabwe":                      "counties",

}

COUNTRY_TO_ISO = {
    # Pre-2026-05-08 set
    "United States":            "US",
    "Mexico":                   "MX",
    "Japan":                    "JP",
    "Netherlands":              "NL",
    "Canada":                   "CA",
    "United Kingdom":           "GB",
    "France":                   "FR",
    "Germany":                  "DE",
    "Italy":                    "IT",
    "Spain":                    "ES",
    "Poland":                   "PL",
    "Andorra":                  "AD",
    "San Marino":               "SM",
    "Vatican City":             "VA",
    # 2026-05-08 expansion
    "Romania":                  "RO",
    "Colombia":                 "CO",
    "Turkey":                   "TR",
    "Nigeria":                  "NG",
    "India":                    "IN",
    "Croatia":                  "HR",
    "Guatemala":                "GT",
    "Honduras":                 "HN",
    "El Salvador":              "SV",
    "Norway":                   "NO",
    "Kenya":                    "KE",
    "Ecuador":                  "EC",
    "Peru":                     "PE",
    "Sudan":                    "SD",
    "Cambodia":                 "KH",
    "Cuba":                     "CU",
    "Angola":                   "AO",
    "Dominican Republic":       "DO",
    "Malaysia":                 "MY",
    "Uganda":                   "UG",
    "Nicaragua":                "NI",
    "Costa Rica":               "CR",
    "Benin":                    "BJ",
    "Panama":                   "PA",
    "South Sudan":              "SS",
    "Central African Republic": "CF",
    "Vanuatu":                  "VU",
    "Burkina Faso":             "BF",
    "Gabon":                    "GA",
    "Gambia":                   "GM",
    "Brunei":                   "BN",
    "Eritrea":                  "ER",
    "Guinea":                   "GN",
    "Burundi":                  "BI",
    "Malawi":                   "MW",
    "Somalia":                  "SO",
    "St. Kitts & Nevis":        "KN",
    "Lesotho":                  "LS",
    "Portugal":                 "PT",
    "Luxembourg":               "LU",
    "Congo DR":                 "CD",
    "Guernsey":                 "GG",
    "Jersey":                   "JE",
    # 2026-05-10 expansion (six countries)
    "Brazil":                   "BR",
    "Austria":                  "AT",
    "Switzerland":              "CH",
    "China":                    "CN",
    "Australia":                "AU",
    "South Korea":              "KR",
    # 2026-05-19 expansion (three countries)
    "Russia":                   "RU",
    "Ireland":                  "IE",
    "Belgium":                  "BE",
    # 2026-05-20 expansion (three countries)
    "Singapore":                "SG",
    "Malta":                    "MT",
    "Liechtenstein":            "LI",
    # 2026-05-21 expansion - workbook fills complete (4)
    "Argentina":                "AR",
    "Bulgaria":                 "BG",
    "Denmark":                  "DK",
    "Puerto Rico":              "PR",
    # 2026-05-22 expansion (42 countries)
    "American Samoa":           "AS",
    "Bahamas":                  "BS",
    "Bangladesh":               "BD",
    "Belize":                   "BZ",
    "Cape Verde":               "CV",
    "Cayman Islands":           "KY",
    "Chile":                    "CL",
    "East Timor":               "TL",
    "Equatorial Guinea":        "GQ",
    "Eswatini":                 "SZ",
    "Federated States of Micronesia": "FM",
    "Greenland":                "GL",
    "Guadeloupe":               "GP",
    "Guinea-Bissau":            "GW",
    "Guyana":                   "GY",
    "Hong Kong":                "HK",
    "Isle of Man":              "IM",
    "Jordan":                   "JO",
    "Kiribati":                 "KI",
    "Latvia":                   "LV",
    "Liberia":                  "LR",
    "Marshall Islands":         "MH",
    "Martinique":               "MQ",
    "Mayotte":                  "YT",
    "Montenegro":               "ME",
    "Nauru":                    "NR",
    "Nepal":                    "NP",
    "New Zealand":              "NZ",
    "Niue":                     "NU",
    "Northern Mariana Islands": "MP",
    "Papua New Guinea":         "PG",
    "Philippines":              "PH",
    "Réunion":                  "RE",
    "São Tomé and Príncipe":    "ST",
    "Senegal":                  "SN",
    "Serbia":                   "RS",
    "Sri Lanka":                "LK",
    "Suriname":                 "SR",
    "Thailand":                 "TH",
    "Tokelau":                  "TK",
    "Tuvalu":                   "TV",
    "Uruguay":                  "UY",
    # 2026-05-21 expansion - matcher pending (17)
    "Algeria":                  "DZ",
    "Egypt":                    "EG",
    "Finland":                  "FI",
    "Ghana":                    "GH",
    "Greece":                   "GR",
    "Hungary":                  "HU",
    "Indonesia":                "ID",
    "Israel":                   "IL",
    "Kuwait":                   "KW",
    "Moldova":                  "MD",
    "Qatar":                    "QA",
    "Sweden":                   "SE",
    "Tunisia":                  "TN",
    "Taiwan":                   "TW",
    "Venezuela":                "VE",
    "Vietnam":                  "VN",
    "Samoa":                    "WS",
    # 2026-05-23 expansion
    "Afghanistan":              "AF",
    "Azerbaijan":               "AZ",
    "Iran":                     "IR",
    "Iraq":                     "IQ",
    # 2026-05-23 PM expansion
    "Bahrain":                  "BH",
    "Lebanon":                  "LB",
    "Oman":                     "OM",
    "Palestine":                "PS",
    "Saudi Arabia":             "SA",
    "Syria":                    "SY",
    "United Arab Emirates":     "AE",
    # 2026-05-24 expansion
    "Saba":                     "XS",
    "Sint Eustatius":           "XE",
    "South Africa":             "ZA",
    "Belarus":                  "BY",
    "Ukraine":                  "UA",
    "French Guiana":             "GF",
    "Monaco":                    "MC",
    "Saint Pierre and Miquelon": "PM",
    "Bermuda":                                      "BM",
    "Turks & Caicos Islands":                       "TC",
    "British Virgin Islands":                       "VG",
    "Gibraltar":                                    "GI",
    "Anguilla":                                     "AI",
    "Saint Helena, Ascension and Tristan da Cunha": "SH",
    "Montserrat":                                   "MS",
    "Falkland Islands":                             "FK",
    "Macau":                                        "MO",
    "Guam":                                         "GU",
    "US Virgin Islands":                            "VI",
    "Saint Lucia":                                  "LC",
    "Saint Martin":                                 "MF",
    "Antigua & Barbuda":                            "AG",
    "Solomon Islands":                              "SB",
    "Saint Barthélemy":                             "BL",
    "Sint Maarten":                                 "SX",
    "Bonaire":                                      "BQ",
    "Western Sahara":                               "EH",
    # 2026-05-24 evening mass wiring:
    "Botswana":                       "BW",
    "Chad":                           "TD",
    "Comoros":                        "KM",
    "Cook Islands":                   "CK",
    "Curaçao":                        "CW",
    "Czech Republic":                 "CZ",
    "Djibouti":                       "DJ",
    "Dominica":                       "DM",
    "Estonia":                        "EE",
    "Faroe Islands":                  "FO",
    "Georgia":                        "GE",
    "Grenada":                        "GD",
    "Iceland":                        "IS",
    "Jamaica":                        "JM",
    "Kazakhstan":                     "KZ",
    "Kosovo":                         "XK",
    "Kyrgyzstan":                     "KG",
    "Laos":                           "LA",
    "Libya":                          "LY",
    "Lithuania":                      "LT",
    "Madagascar":                     "MG",
    "Maldives":                       "MV",
    "Mali":                           "ML",
    "Mauritania":                     "MR",
    "Mauritius":                      "MU",
    "Mongolia":                       "MN",
    "Morocco":                        "MA",
    "Mozambique":                     "MZ",
    "Myanmar":                        "MM",
    "Namibia":                        "NA",
    "New Caledonia":                  "NC",
    "North Macedonia":                "MK",
    "Palau":                          "PW",
    "Paraguay":                       "PY",
    "Rwanda":                         "RW",
    "Seychelles":                     "SC",
    "Slovakia":                       "SK",
    "Slovenia":                       "SI",
    "St. Vincent & the Grenadines":   "VC",
    "Tahiti":                         "PF",
    "Tajikistan":                     "TJ",
    "Tanzania":                       "TZ",
    "Togo":                           "TG",
    "Tonga":                          "TO",
    "Trinidad & Tobago":              "TT",
    "Turkmenistan":                   "TM",
    "Uzbekistan":                     "UZ",
    "Wallis and Futuna":              "WF",


    # 2026-05-24 evening expansion - workbook fills PENDING (matcher run required):
    "Albania":                       "AL",
    "Armenia":                       "AM",
    "Aruba":                         "AW",
    "Bolivia":                       "BO",
    "Bosnia-Herzegovina":            "BA",
    "Cameroon":                      "CM",
    "Congo":                         "CG",
    "Cyprus":                        "CY",
    "Ethiopia":                      "ET",
    "North Korea":                   "KP",
    "Yemen":                         "YE",
    "Zambia":                        "ZM",
    "Zimbabwe":                      "ZW",
}

WORKBOOK_TO_CANONICAL_COUNTRY = {
    "England":          "United Kingdom",
    "Scotland":         "United Kingdom",
    "Wales":            "United Kingdom",
    "Northern Ireland": "United Kingdom",
}

UK_CONSTITUENT_REGION = {
    "England":          "GB-ENG",
    "Scotland":         "GB-SCT",
    "Wales":            "GB-WLS",
    "Northern Ireland": "GB-NIR",
}

# Countries Overture treats as having no ISO 3166-2 subdivision (region=None
# on every parquet row). The workbook Region column may be blank or carry the
# country code; we normalize to None so the (region, subtype, primary) lookup
# key matches the parquet, and we relax the incomplete-row check accordingly.
# Hand-curated boundaries: these slugs have manually drawn polygons in
# public/data/metro-boundaries/<slug>.geojson. The builder skips them in
# every phase - it neither hashes them, rebuilds them, nor prunes their
# files. Used for metros where Overture's admin layer is mismatched against
# the actual metro footprint (e.g. multi-emirate conurbations, oasis cities
# without locality polygons). The on-disk geojson is the source of truth.
# Properties.input_hash is set to "manual-curated" by convention.
CURATED_BOUNDARY_SLUGS = frozenset([
    "dubai-sharjah",   # Dubai-Sharjah-Ajman urban band (3-emirate composite)
    "al-ain",          # Al Ain oasis city (no Overture locality polygon)
])


# When a country's per-country parquet contains rows tagged with country
# codes other than the country's standard ISO, the builder needs to know
# which codes to accept when scanning the parquet. Default behavior pulls
# parquet_iso from COUNTRY_TO_ISO (single value), which fails when the
# parquet is a multi-ISO composite. Add overrides here.
#
# overture-PS.parquet is the XW (West Bank) + XG (Gaza) combined extract.
# Overture does not use ISO PS at all - it tags rows with XW or XG - so the
# filter must accept both.
COUNTRY_PARQUET_ISO_OVERRIDE = {
    "Palestine": {"XW", "XG"},
    # 2026-05-24: Western Sahara workbook rows route to overture-MA.parquet
    # because Overture files Sahrawi territory under MA-11/12/13. The Es-Semara
    # row has region=None so we need this override to tell load_overture the
    # parquet country code is MA, not EH.
    "Western Sahara": {"MA"},
}


# Per-row cross-border routing. When a workbook row's Region (ISO 3166-2)
# starts with an ISO code that has its own dedicated parquet, route the row
# to that parquet instead of the country default. Used for the Israel /
# West Bank case: 23 Israeli settlements have reg_iso='XW' even though
# country='Israel' - they need the XW parquet, not the IL parquet, because
# Overture tags West Bank entities under XW regardless of administrating
# authority. Same mechanism handles Palestine rows tagged 'XW' (West Bank
# PA cities) and 'XG' (Gaza governorates).
CROSS_BORDER_PARQUET = {
    "XW": r"C:\Users\ashwi\Desktop\Projects\MapData\overture-XW.parquet",
    "XG": r"C:\Users\ashwi\Desktop\Projects\MapData\overture-XG.parquet",
    # 2026-05-24: Russia-Crimea metros (Feodosia, Kerch, Simferopol, Sevastopol,
    # Yalta, etc.) live in workbook as Russia rows but Overture files Crimea
    # under UA (ISO follows pre-2014 borders). reg_iso prefixes UA-43 (Crimea)
    # and UA-40 (Sevastopol) route those rows to the UA parquet.
    "UA": r"C:\Users\ashwi\Desktop\Projects\MapData\overture-UA.parquet",
    # 2026-05-24: Aland Islands - workbook rows are under Country='Finland'
    # with reg_iso='AX' since Aland is autonomous within Finland by ISO 3166-1
    # convention. Overture files them under country='AX'.
    "AX": r"C:\Users\ashwi\Desktop\Projects\MapData\overture-AX.parquet",
}


def _region_iso_prefix(reg_iso):
    """ISO 3166-1 alpha-2 prefix from a region code like 'IL-Z' or 'XW'."""
    if not reg_iso:
        return None
    return reg_iso.split("-", 1)[0] if "-" in reg_iso else reg_iso


def _effective_region(region_str):
    """Region value used for the (region, subtype, primary) parquet join key.

    CROSS_BORDER_PARQUET keys (XW, XG) name parquets whose rows all carry
    region=None because those Overture codes sit at the top of the hierarchy
    (Overture treats them as the country, not a subdivision). The workbook
    fills region='XW'/'XG' so CROSS_BORDER_PARQUET routing can pick the
    right parquet, but the join key has to be normalized back to None so it
    matches the parquet rows. Without this, Palestine's WB+Gaza metros and
    Israel's 23 WB settlement rows route to the right parquet but never
    match a key.
    """
    if region_str in CROSS_BORDER_PARQUET:
        return None
    return region_str


REGIONLESS_COUNTRIES = {
    "Singapore",
    "Puerto Rico",
    # 2026-05-22 expansion: Overture treats each of these as a country with
    # no internal ISO 3166-2 subdivision (region=None on every parquet row).
    "American Samoa",
    "Cayman Islands",
    "Guadeloupe",
    "Hong Kong",
    "Isle of Man",
    "Martinique",
    "Mayotte",
    "Nauru",
    "Niue",
    "Northern Mariana Islands",
    "Réunion",
    "Tokelau",
    # 2026-05-24 expansion: Saba (ISO XS) and Sint Eustatius (ISO XE)
    # both appear in Overture as single dependency rows with region=None.
    "Saba",
    "Sint Eustatius",
    # Vatican City - only polygon is subtype=country with region=None.
    "Vatican City",
    # 2026-05-24 PM expansion - all three carry region=None on every parquet row:
    "French Guiana",
    "Monaco",
    "Saint Pierre and Miquelon",
    # 2026-05-24 territories - single-island dependencies, region=None throughout
    # (Saint Helena is NOT in this set - it has SH-TA/SH-AC/SH-HL subregions):
    "Jersey",
    "Guernsey",
    "Bermuda",
    "Turks & Caicos Islands",
    "British Virgin Islands",
    "Gibraltar",
    "Anguilla",
    "Montserrat",
    "Falkland Islands",
    "Macau",
    "Guam",
    "US Virgin Islands",
    # 2026-05-24 late additions - single-island dependencies:
    "Saint Martin",
    "Saint Barthélemy",
    "Sint Maarten",
    "Bonaire",
    # 2026-05-24 evening mass wiring - regionless parquets:
    "Cook Islands",
    "Curaçao",
    "Faroe Islands",
    "Kosovo",
    "Madagascar",
    "New Caledonia",
    "Tahiti",
    # 2026-05-24 evening Bucket A: Cyprus and Aruba both have universal
    # region=None on every matcher-suggested row. Cyprus's locality-subtype
    # fills (Greek-script primary names) carry no ISO 3166-2 subdivision in
    # Overture; Aruba is a single-territory dependency. Without REGIONLESS,
    # every fill gets filtered as incomplete and zero polygons render.
    "Cyprus",
    "Aruba",
}

SHEET_SCHEMAS = {
    "counties": {
        "sheet_name":        "Counties",
        "col_country":       0,
        "col_state_full":    2,
        "col_metro_display": 7,
        "col_subtype":       12,
        "col_admin_level":   13,
        "col_region":        14,
        "col_primary":       15,
    },
    "municipality": {
        "sheet_name":        "Municipality",
        "col_country":       1,
        "col_state_full":    4,
        "col_metro_display": 6,
        "col_subtype":       13,
        "col_admin_level":   14,
        "col_region":        15,
        "col_primary":       16,
    },
}


# ---------- Build-cache versioning --------------------------------------
#
# Hash of the script constants that affect output geometry. If any of
# these change, ALL cached metros are invalidated automatically on the
# next run. To force a global rebuild without changing a constant, bump
# the literal "logic_version" string below.
SCRIPT_VERSION_HASH = hashlib.sha256(json.dumps({
    "outlier_max_km":  OUTLIER_PART_MAX_KM,
    "simplify_tol":    SIMPLIFY_TOLERANCE_DEG,
    "logic_version":   "v2",
}, sort_keys=True).encode()).hexdigest()[:12]


# ---------- Geometry helpers --------------------------------------------

def _haversine_km(lat1, lon1, lat2, lon2):
    R = 6371.0
    dlat = radians(lat2 - lat1)
    dlon = radians(lon2 - lon1)
    a = (sin(dlat / 2) ** 2
         + cos(radians(lat1)) * cos(radians(lat2)) * sin(dlon / 2) ** 2)
    return 2 * R * asin(sqrt(a))


def trim_outlier_parts(geom, anchor_lat, anchor_lon,
                       max_distance_km=OUTLIER_PART_MAX_KM):
    if geom is None or geom.is_empty:
        return geom, 0, []
    if geom.geom_type != "MultiPolygon":
        return geom, 0, []
    parts = list(geom.geoms)
    if len(parts) <= 1:
        return geom, 0, []
    anchor = Point(anchor_lon, anchor_lat)
    kept = []
    dropped = []
    for p in parts:
        try:
            near_on_part, _ = nearest_points(p, anchor)
            d_km = _haversine_km(anchor_lat, anchor_lon,
                                 near_on_part.y, near_on_part.x)
        except Exception:
            kept.append(p)
            continue
        if d_km <= max_distance_km:
            kept.append(p)
        else:
            dropped.append((p, d_km))
    if not kept:
        largest = max(parts, key=lambda p: p.area)
        return largest, len(parts) - 1, [d for _, d in dropped]
    if len(kept) == 1:
        result = kept[0]
    else:
        result = MultiPolygon(kept)
    return result, len(dropped), [d for _, d in dropped]


# ---------- Overture loader ----------------------------------------------

def load_overture(parquet_path: str, wanted_keys: set, wanted_iso_codes: set):
    print(f"      Reading parquet: {parquet_path}")
    print(f"      country filter: {sorted(wanted_iso_codes)}")
    t0 = time.time()
    import pyarrow.parquet as pq
    from shapely import wkb as shapely_wkb

    pf = pq.ParquetFile(parquet_path)
    cols = ["geometry", "country", "region", "subtype", "class", "names"]

    by_key_land = defaultdict(list)
    by_key_any = defaultdict(list)
    rows_scanned = 0
    rows_kept = 0
    for batch in pf.iter_batches(batch_size=10_000, columns=cols):
        countries = batch.column("country").to_pylist()
        regions = batch.column("region").to_pylist()
        subtypes = batch.column("subtype").to_pylist()
        classes = batch.column("class").to_pylist()
        names_col = batch.column("names").to_pylist()
        geoms_col = batch.column("geometry").to_pylist()
        for i in range(len(countries)):
            rows_scanned += 1
            if countries[i] not in wanted_iso_codes:
                continue
            nm = names_col[i]
            primary = nm.get("primary") if isinstance(nm, dict) else None
            if not primary:
                continue
            key = (regions[i], subtypes[i], primary)
            if key not in wanted_keys:
                continue
            geom_bytes = geoms_col[i]
            if not geom_bytes:
                continue
            try:
                geom = shapely_wkb.loads(geom_bytes)
            except Exception:
                continue
            if classes[i] == "land":
                by_key_land[key].append(geom)
            by_key_any[key].append(geom)
            rows_kept += 1

    print(f"      scanned {rows_scanned:,} rows, kept {rows_kept:,} matching keys "
          f"in {time.time()-t0:.1f}s")
    print(f"      indexed {len(by_key_any):,} unique (region, subtype, primary) keys")
    return by_key_land, by_key_any


# ---------- Workbook loader ----------------------------------------------

def _read_sheet_rows(wb, sheet_key: str):
    schema = SHEET_SCHEMAS[sheet_key]
    sheet_name = schema["sheet_name"]
    if sheet_name not in wb.sheetnames:
        print(f"      WARNING: sheet '{sheet_name}' not in workbook, skipping")
        return
    ws = wb[sheet_name]

    cc = schema["col_country"]
    cs = schema["col_state_full"]
    cm = schema["col_metro_display"]
    csub = schema["col_subtype"]
    creg = schema["col_region"]
    cpri = schema["col_primary"]
    max_needed = max(cc, cs, cm, csub, creg, cpri)

    kept = 0
    routed_elsewhere = 0
    not_routed = 0
    incomplete = 0
    uk_region_overrides = 0
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or len(r) <= max_needed:
            continue
        country_workbook = r[cc]
        country_canonical = WORKBOOK_TO_CANONICAL_COUNTRY.get(
            country_workbook, country_workbook)
        if country_canonical not in COUNTRY_SHEET_MAP:
            not_routed += 1
            continue
        if COUNTRY_SHEET_MAP[country_canonical] != sheet_key:
            routed_elsewhere += 1
            continue

        state_full = r[cs]
        metro_display = r[cm]
        subtype = r[csub]
        region = r[creg]
        primary = r[cpri]

        region_is_optional = country_canonical in REGIONLESS_COUNTRIES
        if region_is_optional:
            if not (subtype and primary and metro_display):
                incomplete += 1
                continue
        else:
            if not (region and subtype and primary and metro_display):
                incomplete += 1
                continue

        region_str = str(region).strip() if region else None
        subtype_str = str(subtype).strip()
        primary_str = str(primary).strip()

        if sheet_key == "counties":
            if region_str == "US-DC" and subtype_str == "county":
                subtype_str = "region"
            if (region_str == "US-NC" and subtype_str == "county"
                    and primary_str == "Nash County"):
                subtype_str = "neighborhood"
            if country_canonical in REGIONLESS_COUNTRIES:
                # Country has no ISO 3166-2 subdivision in Overture
                # (region=None on every parquet row). Whether the workbook
                # Region column is blank or carries the country code, we
                # normalize to None so the (region, subtype, primary) key
                # matches the parquet. Singapore (filled 'SG') and Puerto
                # Rico (left blank) both flow through this path.
                region_str = None

        if sheet_key == "municipality":
            if (country_workbook in UK_CONSTITUENT_REGION
                    and region_str == "GB-ENG"):
                corrected = UK_CONSTITUENT_REGION[country_workbook]
                if corrected != region_str:
                    uk_region_overrides += 1
                region_str = corrected

        kept += 1
        yield {
            "country":       country_canonical,
            "state_full":    str(state_full or "").strip(),
            "metro_display": str(metro_display).strip(),
            "region":        region_str,
            "subtype":       subtype_str,
            "primary":       primary_str,
            "sheet_key":     sheet_key,
        }

    extra = ""
    if uk_region_overrides:
        extra = f"  uk-region-overrides {uk_region_overrides:,}"
    print(f"      [{sheet_name}] kept {kept:,}  routed-elsewhere {routed_elsewhere:,}  "
          f"unrouted-country {not_routed:,}  incomplete {incomplete:,}{extra}")


def load_workbook_rows(path: str):
    print(f"[1/5] Reading {path} (sheets: "
          f"{', '.join(SHEET_SCHEMAS[s]['sheet_name'] for s in SHEET_SCHEMAS)})")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    sheets_in_use = {COUNTRY_SHEET_MAP[c] for c in COUNTRY_SHEET_MAP}
    print(f"      country routing: " + ", ".join(
        f"{c}->{COUNTRY_SHEET_MAP[c]}" for c in COUNTRY_SHEET_MAP))

    out = []
    for sheet_key in sheets_in_use:
        out.extend(_read_sheet_rows(wb, sheet_key))
    print(f"      total rows kept across all sheets: {len(out):,}")
    return out


# ---------- Slug resolver ------------------------------------------------

def load_metros_index(path: str):
    with open(path, "r", encoding="utf-8") as f:
        metros = json.load(f)
    idx = {}
    routed = set(COUNTRY_SHEET_MAP)
    for m in metros:
        if m.get("country") not in routed:
            continue
        name = m.get("name", "").strip().lower()
        country = m.get("country", "")
        idx[(name, country)] = {
            "slug": m["slug"],
            "lat":  m.get("lat"),
            "lon":  m.get("lon"),
        }
    return idx


def resolve_slug_info(row, metros_index):
    base = row["metro_display"].strip().lower()
    return metros_index.get((base, row["country"]))


# ---------- Build cache --------------------------------------------------

def compute_input_hash(members, anchor):
    """Stable hash of the inputs that determine a metro's polygon."""
    keys = sorted([
        (m["region"], m["subtype"], m["primary"]) for m in members
    ])
    payload = json.dumps({
        "version": SCRIPT_VERSION_HASH,
        "keys":    keys,
        "anchor":  list(anchor) if anchor and anchor[0] is not None else None,
    }, sort_keys=True)
    return hashlib.sha256(payload.encode()).hexdigest()[:16]


def _normalize_cache_entries(cache):
    """Migrate legacy cache entries (plain hash strings) to the structured
    {hash, built_at} form. Legacy entries get a missing built_at so the
    age-based refresh rebuilds them on the next eligible run.
    """
    hashes = cache.get("hashes", {})
    out = {}
    for slug, entry in hashes.items():
        if isinstance(entry, str):
            out[slug] = {"hash": entry, "built_at": None}
        elif isinstance(entry, dict):
            out[slug] = {
                "hash": entry.get("hash"),
                "built_at": entry.get("built_at"),
            }
        else:
            out[slug] = {"hash": None, "built_at": None}
    cache["hashes"] = out
    return cache


def load_build_cache():
    if not BUILD_CACHE_FILE.exists():
        return {"version": SCRIPT_VERSION_HASH, "hashes": {}}
    try:
        with open(BUILD_CACHE_FILE, "r", encoding="utf-8") as f:
            cache = json.load(f)
        if not isinstance(cache, dict) or "hashes" not in cache:
            return {"version": SCRIPT_VERSION_HASH, "hashes": {}}
        if cache.get("version") != SCRIPT_VERSION_HASH:
            print(f"      cache: script version changed, invalidating "
                  f"{len(cache.get('hashes', {})):,} entries")
            return {"version": SCRIPT_VERSION_HASH, "hashes": {}}
        return _normalize_cache_entries(cache)
    except Exception as e:
        print(f"      cache: failed to read ({e}), starting fresh")
        return {"version": SCRIPT_VERSION_HASH, "hashes": {}}


def save_build_cache(cache):
    abs_path = BUILD_CACHE_FILE.resolve()
    print(f"      cache: writing to {abs_path}")
    print(f"      cache: entries to write: {len(cache.get('hashes', {})):,}")
    try:
        BUILD_CACHE_FILE.parent.mkdir(parents=True, exist_ok=True)
        with open(BUILD_CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(cache, f, indent=2, sort_keys=True)
        # Verify on disk
        if BUILD_CACHE_FILE.exists():
            sz = BUILD_CACHE_FILE.stat().st_size
            print(f"      cache: WROTE ok ({sz:,} bytes on disk)")
        else:
            print(f"      cache: ERROR wrote without exception but file does not exist on disk")
    except Exception as e:
        print(f"      cache: WRITE FAILED ({type(e).__name__}: {e})")
        raise


# ---------- Main ---------------------------------------------------------

def _parse_max_age_days(argv):
    """Pull --max-age-days N out of argv. Default DEFAULT_MAX_AGE_DAYS."""
    for i, tok in enumerate(argv):
        if tok == "--max-age-days" and i + 1 < len(argv):
            return int(argv[i + 1])
        if tok.startswith("--max-age-days="):
            return int(tok.split("=", 1)[1])
    return DEFAULT_MAX_AGE_DAYS


def main():
    import datetime as _dt
    force = "--force" in sys.argv
    max_age_days = _parse_max_age_days(sys.argv)
    age_threshold = _dt.timedelta(days=max(0, max_age_days))
    now = _dt.datetime.now(_dt.timezone.utc)
    if force:
        print("FORCE rebuild requested; cache will be ignored.")
    else:
        print(f"Refreshment protocol: rebuild any polygon older than "
              f"{max_age_days} day(s).")

    rows = load_workbook_rows(WORKBOOK)
    metros_index = load_metros_index(METROS_JSON)

    print("[2/5] Resolving slugs and grouping members")
    by_slug = defaultdict(list)
    slug_anchor = {}
    unresolved_metros = set()
    for row in rows:
        info = resolve_slug_info(row, metros_index)
        if info is None:
            unresolved_metros.add(f"{row['metro_display']} ({row['country']})")
            continue
        slug = info["slug"]
        by_slug[slug].append(row)
        slug_anchor[slug] = (info["lat"], info["lon"])
    print(f"      metros resolved: {len(by_slug):,}")
    if unresolved_metros:
        print(f"      metros unresolved (display name not in metros.json): "
              f"{len(unresolved_metros)}")
        for m in sorted(unresolved_metros)[:10]:
            print(f"        - {m}")
        if len(unresolved_metros) > 10:
            print(f"        ... and {len(unresolved_metros) - 10} more")

    keep_slugs = set(by_slug.keys()) | CURATED_BOUNDARY_SLUGS

    print("[3/5] Computing input hashes and consulting cache")
    cache = load_build_cache() if not force else {
        "version": SCRIPT_VERSION_HASH, "hashes": {}
    }
    cached_hashes = cache.get("hashes", {})
    new_hashes = {}
    needs_rebuild = set()
    aged_out = 0
    legacy_entries = 0
    for slug, members in by_slug.items():
        if slug in CURATED_BOUNDARY_SLUGS:
            continue  # never touch curated polygons
        h = compute_input_hash(members, slug_anchor.get(slug))
        new_hashes[slug] = h
        entry = cached_hashes.get(slug) or {}
        cached_h = entry.get("hash") if isinstance(entry, dict) else entry
        if cached_h != h:
            needs_rebuild.add(slug)
            continue
        if not (OUT_DIR / f"{slug}.geojson").exists():
            needs_rebuild.add(slug)
            continue
        built_at_raw = entry.get("built_at") if isinstance(entry, dict) else None
        if not built_at_raw:
            needs_rebuild.add(slug)
            legacy_entries += 1
            continue
        try:
            built_at = _dt.datetime.fromisoformat(built_at_raw)
            if built_at.tzinfo is None:
                built_at = built_at.replace(tzinfo=_dt.timezone.utc)
            if (now - built_at) > age_threshold:
                needs_rebuild.add(slug)
                aged_out += 1
        except Exception:
            needs_rebuild.add(slug)
            legacy_entries += 1
    print(f"      cached entries: {len(cached_hashes):,}")
    print(f"      cache hits: {len(by_slug) - len(needs_rebuild):,}")
    print(f"      need rebuild: {len(needs_rebuild):,}")
    if aged_out:
        print(f"      aged-out (older than {max_age_days}d): {aged_out:,}")
    if legacy_entries:
        print(f"      legacy entries with no built_at: {legacy_entries:,}")

    print("[4/5] Pruning stale boundary files")
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    deleted = 0
    failed_to_delete = []
    if OUT_DIR.exists():
        for f in OUT_DIR.iterdir():
            if not (f.is_file() and f.suffix == ".geojson"):
                continue
            slug = f.stem
            if slug in keep_slugs:
                continue
            try:
                f.unlink()
                deleted += 1
            except PermissionError:
                failed_to_delete.append(f.name)
    print(f"      deleted {deleted:,} stale boundary files")
    if failed_to_delete:
        print(f"      could not delete {len(failed_to_delete):,} files (sandbox bind-mount)")
        manifest = OUT_DIR.parent / "stale-boundaries-to-delete.txt"
        with open(manifest, "w", encoding="utf-8") as mf:
            for name in sorted(failed_to_delete):
                mf.write(name + "\n")
        print(f"      manifest written to {manifest}")

    if not needs_rebuild:
        print("[5/5] Nothing to rebuild; skipping parquet scan.")
        # Drop cache entries for slugs no longer in workbook so the file
        # stays clean.
        cache["version"] = SCRIPT_VERSION_HASH
        cache["hashes"] = {s: new_hashes[s] for s in keep_slugs if s in new_hashes}
        save_build_cache(cache)
        print()
        print("=" * 60)
        print(f"Boundaries written: 0 (all {len(by_slug):,} metros up to date)")
        print(f"Stale files removed: {deleted:,}")
        print("=" * 60)
        return

    print(f"[5/5] Rebuilding {len(needs_rebuild):,} metro(s)")

    # Group rebuild rows by parquet path
    rebuild_rows = [r for slug in needs_rebuild for r in by_slug[slug]]
    rows_by_parquet = defaultdict(list)
    for r in rebuild_rows:
        iso_pref = _region_iso_prefix(r.get("region"))
        if iso_pref and iso_pref in CROSS_BORDER_PARQUET:
            parquet = CROSS_BORDER_PARQUET[iso_pref]
        else:
            parquet = COUNTRY_PARQUET_MAP.get(r["country"], SOURCE_PARQUET)
        rows_by_parquet[parquet].append(r)
    print(f"      parquet routing: {len(rows_by_parquet)} distinct parquet(s)")
    for p, rs in rows_by_parquet.items():
        countries = sorted({r["country"] for r in rs})
        print(f"        {p}")
        print(f"          {len(rs):,} rows, countries: {countries}")

    by_key_land = defaultdict(list)
    by_key_any = defaultdict(list)
    for parquet_path, parquet_rows in rows_by_parquet.items():
        parquet_keys = {(_effective_region(r["region"]), r["subtype"], r["primary"])
                        for r in parquet_rows}
        parquet_iso = set()
        for r in parquet_rows:
            iso_pref = _region_iso_prefix(r.get("region"))
            if iso_pref:
                parquet_iso.add(iso_pref)
            else:
                override = COUNTRY_PARQUET_ISO_OVERRIDE.get(r["country"])
                if override:
                    parquet_iso.update(override)
                elif r["country"] in COUNTRY_TO_ISO:
                    parquet_iso.add(COUNTRY_TO_ISO[r["country"]])
        print(f"      wanted keys for {parquet_path}: {len(parquet_keys):,}")
        pl, pa = load_overture(parquet_path, parquet_keys, parquet_iso)
        for k, v in pl.items():
            by_key_land[k].extend(v)
        for k, v in pa.items():
            by_key_any[k].extend(v)

    written = 0
    skipped_no_geom = 0
    skipped_no_anchor = 0
    derived_anchors = 0
    unmatched_per_metro = defaultdict(list)
    trim_audit = []
    successfully_built = set()
    for slug in needs_rebuild:
        members = by_slug[slug]
        polys = []
        for m in members:
            key = (_effective_region(m["region"]), m["subtype"], m["primary"])
            geoms = by_key_land.get(key) or by_key_any.get(key)
            if not geoms:
                unmatched_per_metro[slug].append(
                    f"{m['region']}/{m['subtype']}/{m['primary']!r}"
                )
                continue
            polys.extend(geoms)
        if not polys:
            skipped_no_geom += 1
            continue

        # Pre-simplify each member to cut vertex count before union.
        # Massive speedup on heavy metros (Paris 1,563 communes).
        if MEMBER_SIMPLIFY_TOLERANCE_DEG > 0:
            polys = [
                p.simplify(MEMBER_SIMPLIFY_TOLERANCE_DEG, preserve_topology=True)
                for p in polys
            ]

        merged = unary_union(polys)

        anchor_lat, anchor_lon = slug_anchor.get(slug, (None, None))
        anchor_valid = (
            anchor_lat is not None and anchor_lon is not None
            and isinstance(anchor_lat, (int, float))
            and isinstance(anchor_lon, (int, float))
            and not (anchor_lat == 0 and anchor_lon == 0)
        )
        if not anchor_valid:
            # Derive anchor from the largest polygon part. This guarantees
            # the anchor lies inside the urban core for chains-of-islands
            # cases (Honolulu, Tokyo) since the mainland part dominates.
            try:
                if merged.geom_type == "MultiPolygon":
                    largest = max(merged.geoms, key=lambda p: p.area)
                else:
                    largest = merged
                rp = largest.representative_point()
                anchor_lat, anchor_lon = float(rp.y), float(rp.x)
                anchor_valid = True
                derived_anchors += 1
            except Exception:
                pass

        if anchor_valid:
            merged, n_dropped, dropped_dists = trim_outlier_parts(
                merged, float(anchor_lat), float(anchor_lon))
            if n_dropped > 0:
                trim_audit.append((slug, n_dropped, max(dropped_dists)))
        else:
            skipped_no_anchor += 1

        try:
            simplified = merged.simplify(SIMPLIFY_TOLERANCE_DEG, preserve_topology=True)
            if simplified.is_valid and not simplified.is_empty:
                merged = simplified
        except Exception:
            pass
        feature = {
            "type": "Feature",
            "properties": {
                "slug": slug,
                "members": len(polys),
                "country": members[0]["country"],
                "input_hash": new_hashes[slug],
            },
            "geometry": mapping(merged),
        }
        out_path = OUT_DIR / f"{slug}.geojson"
        try:
            with open(out_path, "w", encoding="utf-8") as f:
                json.dump({"type": "FeatureCollection", "features": [feature]}, f)
            written += 1
            successfully_built.add(slug)
        except PermissionError:
            import tempfile as _tf, shutil as _sh
            tfd, tname = _tf.mkstemp(dir=OUT_DIR, suffix=".geojson")
            os.close(tfd)
            with open(tname, "w", encoding="utf-8") as f:
                json.dump({"type": "FeatureCollection", "features": [feature]}, f)
            _sh.move(tname, out_path)
            written += 1
            successfully_built.add(slug)

    # Persist cache: keep only slugs still in workbook AND either rebuilt
    # successfully OR previously cached (which means their existing GeoJSON
    # is still valid). Each entry carries the hash plus a built_at timestamp
    # so the refreshment protocol can age-out stale polygons next run.
    now_iso = now.isoformat()
    new_entries = {}
    for slug in keep_slugs:
        if slug in CURATED_BOUNDARY_SLUGS:
            continue  # no cache entry needed; curated polygons are always-on
        if slug in successfully_built:
            new_entries[slug] = {"hash": new_hashes[slug], "built_at": now_iso}
        elif slug not in needs_rebuild:
            prior = cached_hashes.get(slug) or {}
            prior_built = prior.get("built_at") if isinstance(prior, dict) else None
            new_entries[slug] = {"hash": new_hashes[slug], "built_at": prior_built}
    cache["version"] = SCRIPT_VERSION_HASH
    cache["hashes"] = new_entries
    save_build_cache(cache)

    # Emit a single combined + simplified boundaries file for the Expandable
    # Map's full-corpus view. Per-metro .geojson files stay as-is for detail
    # pages, the home rankings overlay, and any per-slug caller. This combined
    # file is one HTTP request instead of ~4,000, sized for one-shot edge cache
    # rather than thousands of small CDN hits. Simplification tolerance 0.002
    # degree is visually indistinguishable at country/continent zoom and cuts
    # raw size roughly 5x; gzip lands around 3 MB at full corpus.
    print()
    print("[*] Emitting combined + simplified boundaries.json ...")
    combined_path = OUT_DIR.parent / "boundaries-simplified.json"
    SIMPLIFY_TOL = 0.002
    combined_features = []
    combined_written = 0
    combined_skipped = 0
    for gj_file in sorted(OUT_DIR.glob("*.geojson")):
        slug = gj_file.stem
        if slug in CURATED_BOUNDARY_SLUGS:
            # Keep curated polygons un-simplified - they're already minimal
            # and editorial. Read and append as-is.
            pass
        try:
            with open(gj_file, "r", encoding="utf-8") as f:
                fc = json.load(f)
            for feat in (fc.get("features") or []):
                geom = feat.get("geometry")
                if not geom:
                    continue
                # Re-simplify via shapely. Curated polygons skip simplification.
                if slug not in CURATED_BOUNDARY_SLUGS:
                    try:
                        sh = _shapely_shape(geom)
                        sh = sh.simplify(SIMPLIFY_TOL, preserve_topology=True)
                        geom = mapping(sh)
                    except Exception:
                        # Fall back to original geometry on simplify failure
                        pass
                props = dict(feat.get("properties") or {})
                props["slug"] = slug
                combined_features.append({
                    "type": "Feature",
                    "properties": props,
                    "geometry": geom,
                })
                combined_written += 1
        except Exception as e:
            combined_skipped += 1
            print(f"  warn: failed to read {gj_file.name}: {e}")
    with open(combined_path, "w", encoding="utf-8") as f:
        json.dump({"type": "FeatureCollection", "features": combined_features}, f, separators=(',', ':'))
    combined_size = combined_path.stat().st_size
    print(f"      combined boundaries: {combined_written:,} features, {combined_size/1024/1024:.1f} MB raw")
    if combined_skipped:
        print(f"      (skipped {combined_skipped} file(s) with read errors)")

    print()
    print("=" * 60)
    print(f"Boundaries written: {written:,}")
    print(f"Cache entries persisted: {len(cache['hashes']):,}")
    print(f"Stale files removed: {deleted:,}")
    print(f"Metros skipped (no geometry resolved): {skipped_no_geom:,}")
    if derived_anchors:
        print(f"Metros with anchor derived from largest polygon part: {derived_anchors:,}")
    if skipped_no_anchor:
        print(f"Metros built without outlier-trim (no anchor at all): {skipped_no_anchor:,}")
    if trim_audit:
        print(f"Outlier-trim applied to {len(trim_audit)} metro(s):")
        for slug, n, max_d in sorted(trim_audit, key=lambda x: -x[2]):
            print(f"  {slug}: dropped {n} part(s), furthest {max_d:,.0f} km")
    if unmatched_per_metro:
        total_unmatched = sum(len(v) for v in unmatched_per_metro.values())
        print(f"Members unmatched in parquet: {total_unmatched:,} across "
              f"{len(unmatched_per_metro)} metros")
        if "--verbose" in sys.argv:
            for slug, items in sorted(unmatched_per_metro.items())[:20]:
                print(f"  {slug}:")
                for it in items[:5]:
                    print(f"    - {it}")
                if len(items) > 5:
                    print(f"    ... +{len(items) - 5} more")
    print("=" * 60)


if __name__ == "__main__":
    main()

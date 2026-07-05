#!/usr/bin/env python3
"""
Build International Football team-pages data from the grand Football workbook
(Champions League-201516.xlsx — Int Totals, Int Summary, Int Tournaments sheets).

V1 scope:
  - One canonical page per Cur. Name from Int Totals (~232 active national teams)
  - PLUS distinct Cur. Name values in Int Summary that aren't in Int Totals
    (defunct teams kept distinct by the workbook, e.g. East Germany)
  - Per-team tournament history from Int Summary (every appearance, round reached,
    champion flag, with "as [predecessor]" attribution where Team != Cur. Name)
  - Per-team tournament finals match detail from Int Tournaments
    (Comp. Rnd = "Final" only — 124 entries total in the workbook)
  - Eight tournament hubs: world-cup, euros, copa-america, afcon, asian-cup,
    gold-cup, ofc-nations-cup, intercontinental

Out of v1: women's, youth, Olympics, friendlies, qualifiers. The Int
Tournaments sheet conveniently contains zero friendlies and zero qualifier
rows, so no filtering is needed there.

Outputs (all under public/data/international/):
  index.json       - team profile rows
  appearances.json - { slug: [appearance rows...] }
  finals.json      - { slug: [final-match rows...] }
  tournaments.json - { hub_slug: { hub payload... } }
  slug-lookup.json - normalize team name -> slug

Usage:
  python3 scripts/build-international-data.py
  python3 scripts/build-international-data.py /path/to/Champions\\ League-201516.xlsx
"""

import json
import os
import re
import sys
import csv
from pathlib import Path
from collections import defaultdict, Counter

import openpyxl

# ---------- Paths ----------

REPO_ROOT = Path(__file__).resolve().parent.parent
HOME = Path(os.path.expanduser("~"))

DEFAULT_SOURCE_CANDIDATES = [
    HOME / "OneDrive" / "Excel Files" / "Champions League-201516.xlsx",
    HOME / "Library" / "CloudStorage" / "OneDrive-Personal" / "Excel Files" / "Champions League-201516.xlsx",
    Path("/sessions/nice-admiring-ramanujan/mnt/uploads/Champions League-201516.xlsx"),
]
# Match the football builder's session-mount discovery for parity in dev.
for sess in Path("/sessions").glob("*/mnt/Excel Files/Champions League-201516.xlsx"):
    DEFAULT_SOURCE_CANDIDATES.append(sess)

NATIONAL_TEAMS_TSV = REPO_ROOT / "public" / "data" / "national-teams.tsv"
OUT_DIR = REPO_ROOT / "public" / "data" / "international"


# ---------- Tournament hub configuration ----------

# Each hub: (slug, label, filter_lambda over an Int Summary row dict).
# Row dict keys are normalized header names (lowercased, simplified) plus the
# raw flag columns we care about.
TOURNAMENT_HUBS = [
    {
        "slug": "world-cup",
        "label": "FIFA World Cup",
        "category": "WC",  # short tag stored on each appearance row
        "filter": lambda r: r.get("world_cup") == "Y",
    },
    {
        "slug": "euros",
        "label": "UEFA European Championship",
        "category": "EUROS",
        "filter": lambda r: r.get("con_champ") == "Y" and r.get("continent") == "Europe",
    },
    {
        "slug": "copa-america",
        "label": "Copa América",
        "category": "COPA",
        "filter": lambda r: r.get("con_champ") == "Y" and r.get("continent") == "South America",
    },
    {
        "slug": "afcon",
        "label": "Africa Cup of Nations",
        "category": "AFCON",
        "filter": lambda r: r.get("con_champ") == "Y" and r.get("continent") == "Africa",
    },
    {
        "slug": "asian-cup",
        "label": "AFC Asian Cup",
        "category": "ASIAN",
        "filter": lambda r: r.get("con_champ") == "Y" and r.get("continent") == "Asia",
    },
    {
        "slug": "gold-cup",
        "label": "CONCACAF Championship / Gold Cup",
        "category": "GOLD",
        "filter": lambda r: r.get("con_champ") == "Y" and r.get("continent") == "North America",
    },
    {
        "slug": "ofc-nations-cup",
        "label": "OFC Nations Cup",
        "category": "OFC",
        "filter": lambda r: r.get("con_champ") == "Y" and r.get("continent") == "Oceania",
    },
    {
        "slug": "intercontinental",
        "label": "Intercontinental Tournaments",
        "category": "INTER",
        "filter": lambda r: r.get("intercont_champ") == "Y",
    },
    # Other Tournaments hub: surfaces the heterogeneous bucket the workbook
    # lumps under the "Other Tourna" flag — Olympic Football (pre-1930),
    # Central European International Cup, Pan-American Championship, the UEFA
    # and CONCACAF Nations League Finals, the Tournoi de France, and
    # the 1960 European Nations Group. The OTHER_TOURNAMENT_NAMES map drives
    # the per-edition label so the page never reads as a flat "Other 1936".
    {
        "slug": "other-tournaments",
        "label": "Other Tournaments",
        "category": "OTHER",
        "filter": lambda r: r.get("other_tourna") == "Y",
    },
]


# ---------- Helpers ----------

def slugify(s):
    """Match scripts/build-football-data.py slugify() for cross-source parity."""
    if s is None:
        return None
    s = str(s).strip().lower()
    repl = {
        "á": "a", "à": "a", "â": "a", "ä": "a", "ã": "a", "å": "a",
        "é": "e", "è": "e", "ê": "e", "ë": "e",
        "í": "i", "ì": "i", "î": "i", "ï": "i",
        "ó": "o", "ò": "o", "ô": "o", "ö": "o", "õ": "o", "ø": "o",
        "ú": "u", "ù": "u", "û": "u", "ü": "u",
        "ñ": "n", "ç": "c", "ß": "ss",
        "ý": "y", "ÿ": "y", "ž": "z", "š": "s", "č": "c", "ć": "c", "ř": "r",
        "&": "and",
    }
    for k, v in repl.items():
        s = s.replace(k, v)
    s = re.sub(r"[^a-z0-9]+", "-", s).strip("-")
    return s or None


def normalize_team_name(s):
    """For cross-source joins (national-teams.tsv lookup, etc.)."""
    if not s:
        return ""
    return re.sub(r"[^a-z0-9]+", " ", str(s).lower()).strip()


def find_source():
    if len(sys.argv) >= 2:
        p = Path(sys.argv[1])
        if not p.exists():
            sys.exit(f"FAIL: source not found: {p}")
        return p
    for c in DEFAULT_SOURCE_CANDIDATES:
        if c.exists():
            return c
    sys.exit("FAIL: no source workbook found. Pass a path or place it under "
             "OneDrive/Excel Files/Champions League-201516.xlsx")


def to_int(v):
    if v is None:
        return None
    if isinstance(v, bool):
        return int(v)
    if isinstance(v, (int, float)):
        try:
            return int(v)
        except Exception:
            return None
    try:
        return int(str(v).strip())
    except Exception:
        return None


def yflag(v):
    """Workbook flag is 'Y' / blank. Normalize to 'Y' or None."""
    if v is None:
        return None
    s = str(v).strip()
    return "Y" if s == "Y" else None


# ---------- Federation lookup from national-teams.tsv ----------

def load_federation_lookup():
    """Returns dict: normalized_team_name -> { federation, fifa_recognized, active }."""
    out = {}
    if not NATIONAL_TEAMS_TSV.exists():
        print(f"WARN: {NATIONAL_TEAMS_TSV} not found; federation field will be blank.")
        return out
    with open(NATIONAL_TEAMS_TSV, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f, delimiter="\t")
        for row in reader:
            name = row.get("National Team") or ""
            key = normalize_team_name(name)
            if not key:
                continue
            out[key] = {
                "federation": (row.get("Federation") or "").strip() or None,
                "fifa_recognized": (row.get("FIFA") or "").strip() == "FIFA",
                "active": (row.get("Active/Defunct") or "").strip().lower() == "active",
                "display_name": name.strip(),
            }
    return out


# ---------- Sheet readers ----------

def read_int_totals(wb):
    """Int Totals header is layered: row 0 is group labels (World Cup, Continental,
    Intercontinental, Other Tournaments), row 1 is empty, row 2 is the column
    name row. We treat row 2 as headers and rows 3+ as data."""
    ws = wb["Int Totals"]
    all_rows = list(ws.iter_rows(values_only=True))
    if len(all_rows) < 4:
        sys.exit("FAIL: Int Totals sheet has fewer than 4 rows")
    headers = [str(h).strip() if h else "" for h in all_rows[2]]
    h_idx = {h: i for i, h in enumerate(headers) if h}
    teams = []
    for r in all_rows[3:]:
        if not r:
            continue
        name = r[h_idx["Name"]] if "Name" in h_idx and h_idx["Name"] < len(r) else None
        if not name:
            continue
        cur_name = r[h_idx["Cur. Name"]] if "Cur. Name" in h_idx and h_idx["Cur. Name"] < len(r) else None
        cur_name = cur_name or name
        get = lambda key: r[h_idx[key]] if key in h_idx and h_idx[key] < len(r) else None
        teams.append({
            "name": str(name).strip(),
            "cur_name": str(cur_name).strip(),
            "continent": (get("Continent (Current)") or "").strip() or None,
            "elo_rank": to_int(get("ELO Rank (Cur)")),
            "fifa_rank": to_int(get("FIFA Rank (Cur)")),
            "tour_app": to_int(get("Tour App")) or 0,
            "trophies": to_int(get("# Trophies")) or 0,
            "major_trophies": to_int(get("# Maj. Trophies")) or 0,
            "last_app": to_int(get("Last App")),
            "last_sf": to_int(get("Last SF")),
            "last_finals": to_int(get("Last Finals")),
            "last_trophy": to_int(get("Last Trophy")),
            "last_major_trophy": to_int(get("Last Maj. Trophy")),
            # Per-category counts
            "wc_app": to_int(get("WC App")) or 0,
            "wc_sf": to_int(get("WC SF")) or 0,
            "wc_finals": to_int(get("WC Finals")) or 0,
            "wc_champ": to_int(get("WC Champ")) or 0,
            "last_wc_app": to_int(get("Last WC App")),
            "last_wc_finals": to_int(get("Last WC Finals")),
            "last_wc_champ": to_int(get("Last WC Champ")),
            "con_app": to_int(get("Con App")) or 0,
            "con_sf": to_int(get("Con SF")) or 0,
            "con_finals": to_int(get("Con Finals")) or 0,
            "con_champ": to_int(get("Con Champ")) or 0,
            "last_con_app": to_int(get("Last Con App")),
            "last_con_champ": to_int(get("Last Con Champ")),
            "int_app": to_int(get("Int App")) or 0,
            "int_finals": to_int(get("Int Fin")) or 0,
            "int_champ": to_int(get("Int Champ")) or 0,
            "ot_app": to_int(get("OT App")) or 0,
            "ot_champ": to_int(get("OT Champ")) or 0,
            "fifa_recognized": yflag(get("FIFA/Confed Recog")) == "Y",
            "subdivision": (get("Subdivision/Zone") or "").strip() or None,
            "active": yflag(get("Active")) == "Y",
        })
    return teams


# Last-known ELO/FIFA snapshot dates. Used only as a per-field fallback when
# the workbook header cell is empty, so the "as of" caption never goes blank.
FALLBACK_RANK_SNAPSHOTS = {"elo": "2026-05-14", "fifa": "2026-04-01"}


def _fmt_snapshot_date(v):
    """Normalize an Int Totals 'as of' date cell to YYYY-MM-DD. Returns None
    for empty/unparseable cells so the caller can fall back."""
    if v is None:
        return None
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    return s or None


def read_rank_snapshots(wb):
    """ELO (D1) and FIFA (E1) snapshot dates from the Int Totals header row.
    Falls back per-field to the last-known dates so a blank cell never blanks
    the 'as of' caption on /teams/national and the team pages."""
    ws = wb["Int Totals"]
    first = ()
    for r in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        first = r
        break
    cell = lambda i: first[i] if i < len(first) else None
    return {
        "elo": _fmt_snapshot_date(cell(3)) or FALLBACK_RANK_SNAPSHOTS["elo"],   # D1
        "fifa": _fmt_snapshot_date(cell(4)) or FALLBACK_RANK_SNAPSHOTS["fifa"],  # E1
    }


def read_int_summary(wb):
    """Per-team-per-tournament-edition appearances. One row per (Team, Year,
    tournament-category) tuple."""
    ws = wb["Int Summary"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        sys.exit("FAIL: Int Summary sheet empty")
    headers = [str(h).strip() if h else "" for h in rows[0]]
    h_idx = {h: i for i, h in enumerate(headers) if h}
    out = []
    for r in rows[1:]:
        if not r:
            continue
        get = lambda key: r[h_idx[key]] if key in h_idx and h_idx[key] < len(r) else None
        year = to_int(get("Year"))
        team = get("Team")
        cur_name = get("Cur. Name") or team
        if not team or not year:
            continue
        out.append({
            "year": year,
            "continent": (get("Continent") or "").strip() or None,
            "team": str(team).strip(),
            "cur_name": str(cur_name).strip(),
            # Round reached flags
            "group_stage": yflag(get("Group Stage/Prelim Round")),
            "round_of_16": yflag(get("Round of 16")),
            "qf": yflag(get("QF")),
            "sf": yflag(get("SF")),
            "finals": yflag(get("Finals")),
            "champions": yflag(get("Champions")),
            # Category flags
            "world_cup": yflag(get("World Cup")),
            "con_champ": yflag(get("Con. Champ")),
            "intercont_champ": yflag(get("Intercont. Champ")),
            "other_tourna": yflag(get("Other Tourna")),
            "group": (str(get("Group")).strip() if get("Group") is not None else None) or None,
        })
    return out


def read_finals(wb):
    """Pull all Int Tournaments rows where Comp. Rnd = 'Final'. Returns a list
    of dicts keyed by Cur. Name. Each row carries Year, opponent, score,
    stadium, tournament label, win/loss flag."""
    ws = wb["Int Tournaments"]
    rows_iter = ws.iter_rows(values_only=True)
    headers_raw = next(rows_iter, None)
    if not headers_raw:
        return []
    headers = [str(h).strip() if h else "" for h in headers_raw]
    h_idx = {h: i for i, h in enumerate(headers) if h}
    out = []
    for r in rows_iter:
        if not r:
            continue
        get = lambda key: r[h_idx[key]] if key in h_idx and h_idx[key] < len(r) else None
        rnd = get("Comp. Rnd")
        if not rnd or str(rnd).strip() != "Final":
            continue
        year = to_int(get("Year"))
        cur_name = get("Cur. Name") or get("Team")
        opp_cur_name = get("Opp. Name") or get("Opp Team")
        comp = get("Leag/Comp.") or get("Eur. Comp/World Comp.")
        if not year or not cur_name or not comp:
            continue
        out.append({
            "year": year,
            "season": (get("Season") or "").strip() or None,
            "cur_name": str(cur_name).strip(),
            "team_as": (str(get("Team")).strip() if get("Team") else None),
            "opp_cur_name": str(opp_cur_name).strip() if opp_cur_name else None,
            "opp_team_as": (str(get("Opp Team")).strip() if get("Opp Team") else None),
            "competition": str(comp).strip(),
            "result": str(get("W/D/L")).strip() if get("W/D/L") else None,
            "for_goals": to_int(get("For")),
            "against_goals": to_int(get("Ag")),
            "penalty_kicks": to_int(get("Penatly Kicks")),  # workbook typo preserved
            "stadium": (str(get("Stadium")).strip() if get("Stadium") else None),
            "stad_country": (str(get("Stad. Country")).strip() if get("Stad. Country") else None),
            "stad_metro": (str(get("Stad. Metro Area")).strip() if get("Stad. Metro Area") else None),
            "category_world_cup": yflag(get("World Cup")),
            "category_con_champ": yflag(get("Con. Champ")),
            "category_intercont": yflag(get("Intercont. Champ")),
            "category_other": yflag(get("Other Tourna")),
            "continent": (str(get("Cont.")).strip() if get("Cont.") else None),
        })
    return out


# ---------- Tournament categorization ----------

def appearance_category(appearance_row):
    """Return the hub category code (WC / EUROS / COPA / ...) for an Int
    Summary row. A row can be flagged across multiple categories; we pick
    the most specific in priority order: World Cup beats Continental beats
    Intercontinental beats Other."""
    for hub in TOURNAMENT_HUBS:
        if hub["filter"](appearance_row):
            return hub["category"]
    if appearance_row.get("other_tourna") == "Y":
        return "OTHER"
    return None


def deepest_round_label(row):
    """Short string describing the deepest round reached on an appearance row."""
    if row.get("champions") == "Y":
        return "Champion"
    if row.get("finals") == "Y":
        return "Final"
    if row.get("sf") == "Y":
        return "Semifinal"
    if row.get("qf") == "Y":
        return "Quarterfinal"
    if row.get("round_of_16") == "Y":
        return "Round of 16"
    if row.get("group_stage") == "Y":
        return "Group Stage"
    return "Appearance"


def tournament_label_for(year, category, continent=None):
    """Human-readable tournament name combining the category with the year.
    Intercontinental rows get year-specific names since the workbook lumps
    Confederations Cup, King Fahd Cup, Mundialito, and Finalissima all under
    the same Intercont. Champ flag. Other Tourna rows similarly need year +
    continent disambiguation since the workbook categorizes Olympic football,
    Central European International Cup, Pan-American Championship, and the
    Nations League Finals all under one flag."""
    if category == "INTER":
        return INTERCONTINENTAL_TOURNAMENT_NAMES.get(year, f"Intercontinental tournament {year}")
    if category == "OTHER":
        key = (year, continent)
        return OTHER_TOURNAMENT_NAMES.get(key, f"Other tournament {year}")
    label_by_cat = {
        "WC": "FIFA World Cup",
        "EUROS": "UEFA European Championship",
        "COPA": "Copa América",
        "AFCON": "Africa Cup of Nations",
        "ASIAN": "AFC Asian Cup",
        "GOLD": "CONCACAF Championship / Gold Cup",
        "OFC": "OFC Nations Cup",
    }
    base = label_by_cat.get(category, "Tournament")
    return f"{base} {year}" if year else base


# (Year, Continent) keyed map for "Other Tourna" rows. Editorial decisions:
#  - 1908, 1912, 1920, 1924, 1928, 1936 World-flagged rows are the Olympic
#    Football tournaments. FIFA retroactively recognized 1924 and 1928 as
#    de facto World Championships; we render them as Olympic Football
#    consistently so the per-team page narrative doesn't conflate them with
#    the FIFA World Cup which began in 1930.
#  - 1930s Europe-flagged rows are the Central European International Cup
#    (also called Coppa Internazionale / Dr. Gero Cup), the de facto
#    European championship before the Euros launched in 1960.
#  - 1952-1960 mixed-continent rows are the Pan-American Championship era.
#  - 1997 World-flagged is the Tournoi de France (France invitational).
#  - 2019-2025 Europe/North America rows are the UEFA / CONCACAF Nations
#    League Finals.
OTHER_TOURNAMENT_NAMES = {
    (1908, "World"): "1908 Olympic Football",
    (1912, "World"): "1912 Olympic Football",
    (1920, "World"): "1920 Olympic Football",
    (1924, "World"): "1924 Olympic Football",
    (1928, "World"): "1928 Olympic Football",
    (1936, "World"): "1936 Olympic Football",
    (1930, "Europe"): "1930 Central European International Cup",
    (1932, "Europe"): "1932 Central European International Cup",
    (1935, "Europe"): "1935 Central European International Cup",
    (1938, "Europe"): "1938 Central European International Cup",
    (1953, "Europe"): "1953 Central European International Cup",
    (1952, "World"): "1952 Pan-American Championship",
    (1956, "World"): "1956 Pan-American Championship",
    (1960, "World"): "1960 Pan-American Championship",
    (1960, "Europe"): "1960 European Nations Group",
    (1997, "World"): "1997 Tournoi de France",
    (2019, "Europe"): "2019 UEFA Nations League Finals",
    (2021, "Europe"): "2021 UEFA Nations League Finals",
    (2021, "North America"): "2021 CONCACAF Nations League Finals",
    (2023, "Europe"): "2023 UEFA Nations League Finals",
    (2023, "North America"): "2023 CONCACAF Nations League Finals",
    (2025, "Europe"): "2025 UEFA Nations League Finals",
    (2025, "North America"): "2025 CONCACAF Nations League Finals",
}


# Year-keyed mapping for intercontinental tournaments. The workbook only
# carries a single Intercont. Champ flag; the actual competition name is
# derived from the era. 1992-1995 were King Fahd Cup; from 1997 onward it
# was rebranded FIFA Confederations Cup; 2022 is the Finalissima. The 1981
# and 1985 events were one-off invitational tournaments (Mundialito and the
# Artemio Franchi Trophy respectively). 1993 was an off-cycle King Fahd Cup
# precursor (the small Saudi-hosted tournament).
INTERCONTINENTAL_TOURNAMENT_NAMES = {
    1981: "Mundialito 1981",
    1985: "Artemio Franchi Trophy 1985",
    1992: "King Fahd Cup 1992",
    1993: "King Fahd Cup 1993",
    1995: "King Fahd Cup 1995",
    1997: "FIFA Confederations Cup 1997",
    1999: "FIFA Confederations Cup 1999",
    2001: "FIFA Confederations Cup 2001",
    2003: "FIFA Confederations Cup 2003",
    2005: "FIFA Confederations Cup 2005",
    2009: "FIFA Confederations Cup 2009",
    2013: "FIFA Confederations Cup 2013",
    2017: "FIFA Confederations Cup 2017",
    2022: "Finalissima 2022",
}


# ---------- Build outputs ----------

# Workbook artifacts that should not get their own canonical page. France B
# is a development-team row; Lebanon/Yemen is a composite row that doesn't
# represent a real national team.
ARTIFACT_TEAM_NAMES = {"France B", "Lebanon/Yemen"}

# Entities that have NO continuous successor state in current form. These
# are the only Team values that get standalone pages. The principle:
#   - USSR's history -> Russia's page (Russia is the continuous successor).
#   - Yugoslavia's history -> Serbia's page (Serbia is the legal successor
#     in FIFA terms, even though the Yugoslav federation fractured).
#   - Czechoslovakia's history -> Czech Republic's page (the Czech state is
#     the continuous successor; Slovakia got its own seat).
#   - South Yemen + South Vietnam are KEPT defunct per Ashwin's prior
#     explicit ask (the unified Yemen / Vietnam represent the North-side
#     continuation; the southern partner ceased to exist as a footballing
#     entity). Flag to revisit if editorial position changes.
#
# East Germany is also defunct; the workbook already keeps it distinct via
# Cur. Name so it doesn't need an entry here.
#
# Country renames (Burma -> Myanmar, Zaire -> Congo DR, Ivory Coast -> Cote
# d'Ivoire, Upper Volta -> Burkina Faso, Czechia <-> Czech Republic, etc.)
# stay folded into the modern entry; "as Burma" attribution on the row
# preserves the historical signal without splitting the page.
TRULY_DEFUNCT_TEAMS = {
    "South Yemen",
    "South Vietnam",
}


def build_teams_index(totals_rows, summary_rows, federation_lookup):
    """Produce one team per Cur. Name plus standalone entries for every
    distinct defunct-Team value (Soviet Union, Yugoslavia, Czechoslovakia,
    South Yemen, South Vietnam, Zaire, etc.). Totals rows are the primary
    source for active teams; teams that appear only in Int Summary (Cur.
    Name not in Totals) get a derived minimal profile, as do defunct Team
    values that the workbook merges into successor Cur. Names."""
    by_cur_name = {}

    for t in totals_rows:
        cn = t["cur_name"]
        if cn in ARTIFACT_TEAM_NAMES:
            continue
        if cn in by_cur_name:
            continue  # first wins; shouldn't happen but defensive
        by_cur_name[cn] = t

    # Derive minimal profiles for Cur. Names present in summary but not in totals.
    summary_curnames = set(r["cur_name"] for r in summary_rows if r.get("cur_name"))
    for cn in summary_curnames - set(by_cur_name.keys()):
        if cn in ARTIFACT_TEAM_NAMES:
            continue
        # Aggregate from summary alone.
        team_rows = [r for r in summary_rows if r["cur_name"] == cn]
        if not team_rows:
            continue
        # Use most recent row's continent as the team's continent.
        latest = max(team_rows, key=lambda r: r["year"])
        wc_app = sum(1 for r in team_rows if r["world_cup"] == "Y")
        wc_champ = sum(1 for r in team_rows if r["world_cup"] == "Y" and r["champions"] == "Y")
        con_app = sum(1 for r in team_rows if r["con_champ"] == "Y")
        con_champ = sum(1 for r in team_rows if r["con_champ"] == "Y" and r["champions"] == "Y")
        int_app = sum(1 for r in team_rows if r["intercont_champ"] == "Y")
        int_champ = sum(1 for r in team_rows if r["intercont_champ"] == "Y" and r["champions"] == "Y")
        ot_app = sum(1 for r in team_rows if r["other_tourna"] == "Y")
        ot_champ = sum(1 for r in team_rows if r["other_tourna"] == "Y" and r["champions"] == "Y")
        all_champ = wc_champ + con_champ + int_champ + ot_champ
        major_champ = wc_champ + con_champ
        by_cur_name[cn] = {
            "name": cn,
            "cur_name": cn,
            "continent": latest.get("continent"),
            "elo_rank": None,
            "fifa_rank": None,
            "tour_app": len(team_rows),
            "trophies": all_champ,
            "major_trophies": major_champ,
            "last_app": latest["year"],
            "last_sf": max((r["year"] for r in team_rows if r["sf"] == "Y"), default=None),
            "last_finals": max((r["year"] for r in team_rows if r["finals"] == "Y"), default=None),
            "last_trophy": max((r["year"] for r in team_rows if r["champions"] == "Y"), default=None),
            "last_major_trophy": max((r["year"] for r in team_rows
                                       if r["champions"] == "Y" and (r["world_cup"] == "Y" or r["con_champ"] == "Y")), default=None),
            "wc_app": wc_app, "wc_sf": 0, "wc_finals": 0, "wc_champ": wc_champ,
            "last_wc_app": max((r["year"] for r in team_rows if r["world_cup"] == "Y"), default=None),
            "last_wc_finals": None, "last_wc_champ": None,
            "con_app": con_app, "con_sf": 0, "con_finals": 0, "con_champ": con_champ,
            "last_con_app": max((r["year"] for r in team_rows if r["con_champ"] == "Y"), default=None),
            "last_con_champ": None,
            "int_app": int_app, "int_finals": 0, "int_champ": int_champ,
            "ot_app": ot_app, "ot_champ": ot_champ,
            "fifa_recognized": False,  # not in Int Totals => not active in workbook's eyes
            "subdivision": None,
            "active": False,
        }

    # Standalone defunct-team entries. Only entities in TRULY_DEFUNCT_TEAMS
    # qualify: renames (Burma -> Myanmar, Zaire -> Congo DR, etc.) stay
    # folded into the modern country and do NOT get their own page; only
    # entities that ceased to exist as nations get standalone pages so
    # the editorial position is consistent.
    summary_team_values = set(r["team"] for r in summary_rows if r.get("team"))
    defunct_team_names = sorted(
        (summary_team_values & TRULY_DEFUNCT_TEAMS) - set(by_cur_name.keys())
    )
    for defunct_name in defunct_team_names:
        defunct_rows = [r for r in summary_rows if r["team"] == defunct_name]
        if not defunct_rows:
            continue
        latest = max(defunct_rows, key=lambda r: r["year"])
        wc_app = sum(1 for r in defunct_rows if r["world_cup"] == "Y")
        wc_champ = sum(1 for r in defunct_rows if r["world_cup"] == "Y" and r["champions"] == "Y")
        con_app = sum(1 for r in defunct_rows if r["con_champ"] == "Y")
        con_champ = sum(1 for r in defunct_rows if r["con_champ"] == "Y" and r["champions"] == "Y")
        int_app = sum(1 for r in defunct_rows if r["intercont_champ"] == "Y")
        int_champ = sum(1 for r in defunct_rows if r["intercont_champ"] == "Y" and r["champions"] == "Y")
        ot_app = sum(1 for r in defunct_rows if r["other_tourna"] == "Y")
        ot_champ = sum(1 for r in defunct_rows if r["other_tourna"] == "Y" and r["champions"] == "Y")
        all_champ = wc_champ + con_champ + int_champ + ot_champ
        major_champ = wc_champ + con_champ
        # Use the defunct name as both `name` and `cur_name` for the
        # standalone entry so the slug + display reflect the defunct entity.
        by_cur_name[defunct_name] = {
            "name": defunct_name,
            "cur_name": defunct_name,
            "continent": latest.get("continent"),
            "elo_rank": None,
            "fifa_rank": None,
            "tour_app": len(defunct_rows),
            "trophies": all_champ,
            "major_trophies": major_champ,
            "last_app": latest["year"],
            "last_sf": max((r["year"] for r in defunct_rows if r["sf"] == "Y"), default=None),
            "last_finals": max((r["year"] for r in defunct_rows if r["finals"] == "Y"), default=None),
            "last_trophy": max((r["year"] for r in defunct_rows if r["champions"] == "Y"), default=None),
            "last_major_trophy": max((r["year"] for r in defunct_rows
                                       if r["champions"] == "Y" and (r["world_cup"] == "Y" or r["con_champ"] == "Y")), default=None),
            "wc_app": wc_app, "wc_sf": 0, "wc_finals": 0, "wc_champ": wc_champ,
            "last_wc_app": max((r["year"] for r in defunct_rows if r["world_cup"] == "Y"), default=None),
            "last_wc_finals": None, "last_wc_champ": None,
            "con_app": con_app, "con_sf": 0, "con_finals": 0, "con_champ": con_champ,
            "last_con_app": max((r["year"] for r in defunct_rows if r["con_champ"] == "Y"), default=None),
            "last_con_champ": None,
            "int_app": int_app, "int_finals": 0, "int_champ": int_champ,
            "ot_app": ot_app, "ot_champ": ot_champ,
            "fifa_recognized": False,
            "subdivision": None,
            "active": False,
        }

    # Attach slugs + federation
    teams = []
    for cn, t in by_cur_name.items():
        slug = slugify(cn)
        if not slug:
            continue
        fed_lookup = federation_lookup.get(normalize_team_name(cn))
        # Try alt name keys if cur_name didn't hit
        if not fed_lookup and t.get("name") and t["name"] != cn:
            fed_lookup = federation_lookup.get(normalize_team_name(t["name"]))
        teams.append({
            "slug": slug,
            "name": t["name"],
            "cur_name": cn,
            "continent": t["continent"],
            "federation": (fed_lookup or {}).get("federation"),
            "elo_rank": t["elo_rank"],
            "fifa_rank": t["fifa_rank"],
            "totals": {
                "tour_app": t["tour_app"],
                "trophies": t["trophies"],
                "major_trophies": t["major_trophies"],
                "last_app": t["last_app"],
                "last_sf": t["last_sf"],
                "last_finals": t["last_finals"],
                "last_trophy": t["last_trophy"],
                "last_major_trophy": t["last_major_trophy"],
            },
            "world_cup": {
                "app": t["wc_app"], "sf": t["wc_sf"], "finals": t["wc_finals"], "champ": t["wc_champ"],
                "last_app": t["last_wc_app"], "last_finals": t["last_wc_finals"], "last_champ": t["last_wc_champ"],
            },
            "continental": {
                "app": t["con_app"], "sf": t["con_sf"], "finals": t["con_finals"], "champ": t["con_champ"],
                "last_app": t["last_con_app"], "last_champ": t["last_con_champ"],
            },
            "intercontinental": {
                "app": t["int_app"], "finals": t["int_finals"], "champ": t["int_champ"],
            },
            "other": {
                "app": t["ot_app"], "champ": t["ot_champ"],
            },
            "fifa_recognized": t["fifa_recognized"],
            "subdivision": t["subdivision"],
            "active": t["active"],
        })
    teams.sort(key=lambda x: x["cur_name"].lower())
    return teams


MAX_DISPLAYED_TOURNAMENT_YEAR = 2026  # Bump after the next major-tournament cycle (Euros 2028, Asian Cup 2027 etc.) actually concludes.


# ---------- Honors weighting + similar-teams configuration ----------
#
# Editorial weights for the honors index. Each starter weight reflects the
# defensible delta in prestige between achievement levels. Numbers are
# deliberately round so the methodology block on the site can defend them
# in two sentences:
#  - World Cup win is the apex (8). A Euros / Copa América win and a World
#    Cup final lost are co-equal (3 each).
#  - Continental final lost (1) sits clearly below an intercontinental win
#    (1.5) because reaching a final still beats winning a smaller cup.
#  - A World Cup semifinal that didn't reach the final (0.75) is the lightest
#    weighted entry: enough to differentiate teams that "got close" from
#    those that never did, without inflating the score.
HONORS_WEIGHTS = {
    "wc_win": 8.0,
    "wc_final_lost": 3.0,
    "continental_win": 3.0,            # before continent-tier multiplier
    "continental_final_lost": 1.0,     # before continent-tier multiplier
    "intercontinental_win": 1.5,
    "wc_sf_without_final": 0.75,
}

# Continent tiering. Any football fan will tell you a Gold Cup is not a Euros.
# Multiplier applied to continental_win and continental_final_lost weights so
# Brazil's Copa stays at full value while Mexico's Gold Cups get appropriately
# discounted. Tiers are deliberately three-tier (1.0 / 0.75 / 0.5 / 0.3) so
# the methodology block can defend them in one sentence: Euros and Copa are
# the top tier of continental competition; AFCON sits one notch below; Asian
# Cup and Gold Cup at half weight; OFC at one-third reflecting AUS/NZL
# dominance of a small field.
CONTINENT_TOURNAMENT_WEIGHTS = {
    "EUROS": 1.0,
    "COPA":  1.0,
    "AFCON": 0.75,
    "ASIAN": 0.5,
    "GOLD":  0.5,
    "OFC":   0.3,
}

# Similar-teams scoring: six honors-shape dimensions weighted higher (honors
# profile is the editorial signal we cluster on) plus two longevity-shape
# dimensions weighted lower (span + decade coverage break ties between teams
# with similar trophy hauls but very different historical arcs).
SIMILAR_DIMENSION_WEIGHTS = {
    "wc_champ": 2.0,
    "wc_finals_lost": 1.5,
    "continental_champ": 1.5,
    "continental_finals_lost": 1.0,
    "intercontinental_champ": 1.0,
    "wc_sf_without_final": 0.75,
    "tournament_span_years": 0.5,
    "decade_coverage": 0.5,
}

# How many nearest neighbors to surface per team.
SIMILAR_NEIGHBORS = 5

# How many teams to surface in the honors leaderboard payload.
LEADERBOARD_SIZE = 30


def compute_honors(team_obj, team_appearances):
    """Returns (honors_index_float, breakdown_dict) for a team. Walks the
    team's appearance rows to derive per-continent wins and runner-up counts
    so the continent-tier multiplier can apply cleanly. team_appearances is
    the list of appearance dicts for this team (the value out of
    build_appearances() keyed by slug).

    breakdown_dict is what the methodology block on /teams/national defends
    in plain English; it surfaces the per-category counts AND the points
    each category contributed."""
    wc = team_obj["world_cup"]

    wc_champ = wc["champ"] or 0
    wc_finals_lost = max((wc["finals"] or 0) - wc_champ, 0)
    wc_sf_only = max((wc["sf"] or 0) - (wc["finals"] or 0), 0)

    # Derive per-continental-tournament champion + runner-up counts from
    # appearances. Each appearance has a category (EUROS/COPA/AFCON/ASIAN/
    # GOLD/OFC/WC/INTER/OTHER). We tally champions and final-round losses
    # per continental category and apply the tier multiplier.
    per_continent_champ = {k: 0 for k in CONTINENT_TOURNAMENT_WEIGHTS}
    per_continent_runner_up = {k: 0 for k in CONTINENT_TOURNAMENT_WEIGHTS}
    for a in (team_appearances or []):
        cat = a.get("category")
        if cat not in CONTINENT_TOURNAMENT_WEIGHTS:
            continue
        if a.get("champion"):
            per_continent_champ[cat] += 1
        elif a.get("round_reached") == "Final":
            per_continent_runner_up[cat] += 1

    # Continental contribution to the index.
    continental_points = 0.0
    for cat, mult in CONTINENT_TOURNAMENT_WEIGHTS.items():
        continental_points += per_continent_champ[cat] * HONORS_WEIGHTS["continental_win"] * mult
        continental_points += per_continent_runner_up[cat] * HONORS_WEIGHTS["continental_final_lost"] * mult

    inter_champ = (team_obj["intercontinental"]["champ"] or 0)

    score = (
        wc_champ * HONORS_WEIGHTS["wc_win"]
        + wc_finals_lost * HONORS_WEIGHTS["wc_final_lost"]
        + continental_points
        + inter_champ * HONORS_WEIGHTS["intercontinental_win"]
        + wc_sf_only * HONORS_WEIGHTS["wc_sf_without_final"]
    )

    # Aggregate continental counts for the breakdown (sum across continents)
    # plus per-tournament-tier detail so the per-team panel can show "Won 13
    # Gold Cups (tier 0.5, contributed 19.5 points)".
    continental_champ_total = sum(per_continent_champ.values())
    continental_finals_lost_total = sum(per_continent_runner_up.values())

    breakdown = {
        "wc_champ": wc_champ,
        "wc_finals_lost": wc_finals_lost,
        "wc_sf_only": wc_sf_only,
        "continental_champ": continental_champ_total,
        "continental_finals_lost": continental_finals_lost_total,
        "intercontinental_champ": inter_champ,
        "per_continent_champ": per_continent_champ,
        "per_continent_runner_up": per_continent_runner_up,
    }
    return round(score, 2), breakdown


def compute_team_longevity(slug, summary_rows, slug_for_any):
    """Tournament span (last_year - first_year, in years) and decade coverage
    (count of distinct decades 1900-2030 with at least one tournament
    appearance). Used by the similar-teams engine. Returns (span, decades)."""
    years = []
    for r in summary_rows:
        if not r.get("year"):
            continue
        if r["year"] > MAX_DISPLAYED_TOURNAMENT_YEAR:
            continue
        # Route the row to its slug the same way build_appearances does so
        # the per-team longevity numbers match what shows up on the page.
        successor_slug = slug_for_any.get(r["cur_name"])
        is_for_this_slug = successor_slug == slug
        if (
            not is_for_this_slug
            and r.get("team") in TRULY_DEFUNCT_TEAMS
            and slug_for_any.get(r["team"]) == slug
        ):
            is_for_this_slug = True
        if not is_for_this_slug:
            continue
        years.append(r["year"])
    if not years:
        return 0, 0
    span = max(years) - min(years)
    decades = len({(y // 10) * 10 for y in years})
    return span, decades


def build_similar_teams(teams):
    """For each team, compute weighted-Euclidean z-score distance against
    every other team across the SIMILAR_DIMENSION_WEIGHTS axes. Surface the
    five nearest neighbors per slug. Returns a dict {slug: [neighbor_obj, ...]}.
    Each neighbor_obj carries slug, cur_name, distance (rounded for display)
    and the shared-strength axis (the dimension where the pair is closest
    relative to the cohort, used by the editorial caption)."""
    if not teams:
        return {}

    dims = list(SIMILAR_DIMENSION_WEIGHTS.keys())

    def vec(t):
        b = t["honors_breakdown"]
        return [
            float(b["wc_champ"]),
            float(b["wc_finals_lost"]),
            float(b["continental_champ"]),
            float(b["continental_finals_lost"]),
            float(b["intercontinental_champ"]),
            float(b["wc_sf_only"]),
            float(t["tournament_span_years"]),
            float(t["decade_coverage"]),
        ]

    matrix = [vec(t) for t in teams]
    # z-score each dimension across the cohort. Teams with zero across the
    # board still get a vector at the origin; the distance metric handles
    # them naturally (they cluster together as "never been to a tournament").
    means = []
    stds = []
    for i in range(len(dims)):
        col = [row[i] for row in matrix]
        mu = sum(col) / len(col)
        var = sum((v - mu) ** 2 for v in col) / len(col)
        sigma = var ** 0.5 or 1.0  # avoid divide-by-zero on degenerate dims
        means.append(mu)
        stds.append(sigma)

    z_matrix = []
    for row in matrix:
        z_matrix.append([(row[i] - means[i]) / stds[i] for i in range(len(dims))])

    weights = [SIMILAR_DIMENSION_WEIGHTS[d] for d in dims]

    out = {}
    for i, t in enumerate(teams):
        dists = []
        for j, other in enumerate(teams):
            if i == j:
                continue
            # Weighted Euclidean across z-scored axes.
            d2 = 0.0
            for k in range(len(dims)):
                diff = z_matrix[i][k] - z_matrix[j][k]
                d2 += weights[k] * diff * diff
            dists.append((d2 ** 0.5, j))
        dists.sort(key=lambda x: x[0])
        neighbors = []
        for dist, j in dists[:SIMILAR_NEIGHBORS]:
            other = teams[j]
            # Identify the axis where the pair is closest relative to its
            # weight contribution. Used as a hint for the editorial caption,
            # which is finalized in the UI layer.
            best_axis = None
            best_score = None
            for k, d in enumerate(dims):
                contrib = weights[k] * (z_matrix[i][k] - z_matrix[j][k]) ** 2
                # Lowest contribution = most-shared dimension. Skip dims where
                # both teams are at the cohort baseline (z near 0) since
                # "shared" there just means both are unremarkable on that axis.
                if abs(z_matrix[i][k]) < 0.25 and abs(z_matrix[j][k]) < 0.25:
                    continue
                if best_score is None or contrib < best_score:
                    best_score = contrib
                    best_axis = d
            neighbors.append({
                "slug": other["slug"],
                "cur_name": other["cur_name"],
                "continent": other.get("continent"),
                "distance": round(dist, 3),
                "shared_axis": best_axis,
                "honors_index": other.get("honors_index"),
            })
        out[t["slug"]] = neighbors
    return out


def build_honors_leaderboard(teams):
    """Sorted descending by honors_index. Ties broken by World Cup wins,
    then continental wins, then total trophies, then cur_name alphabetical."""
    ranked = sorted(
        teams,
        key=lambda t: (
            -(t.get("honors_index") or 0),
            -(t["world_cup"]["champ"] or 0),
            -(t["continental"]["champ"] or 0),
            -(t["totals"]["trophies"] or 0),
            t["cur_name"].lower(),
        ),
    )
    leaderboard = []
    for rank, t in enumerate(ranked[:LEADERBOARD_SIZE], start=1):
        leaderboard.append({
            "rank": rank,
            "slug": t["slug"],
            "cur_name": t["cur_name"],
            "continent": t.get("continent"),
            "honors_index": t.get("honors_index"),
            "honors_breakdown": t.get("honors_breakdown"),
            "elo_rank": t.get("elo_rank"),
            "fifa_rank": t.get("fifa_rank"),
            "active": t.get("active"),
        })
    return {
        "weights": HONORS_WEIGHTS,
        "leaderboard": leaderboard,
        "leaderboard_size": LEADERBOARD_SIZE,
    }


def build_appearances(summary_rows, slug_for_any):
    """{slug: [appearance rows...]}. Each summary row may emit up to two
    entries: one on the successor's page (with 'as Defunct' attribution),
    one on the defunct entity's standalone page (no attribution since the
    page IS the defunct entity). Artifact teams are filtered out entirely.
    Future tournaments (year > MAX_DISPLAYED_TOURNAMENT_YEAR) are excluded
    so placeholder rows in the workbook for upcoming editions don't surface
    on team pages before the event has happened."""
    out = defaultdict(list)
    for r in summary_rows:
        if r.get("team") in ARTIFACT_TEAM_NAMES or r.get("cur_name") in ARTIFACT_TEAM_NAMES:
            continue
        if r.get("year") and r["year"] > MAX_DISPLAYED_TOURNAMENT_YEAR:
            continue
        cat = appearance_category(r)
        if cat is None:
            continue
        base = {
            "year": r["year"],
            "continent": r["continent"],
            "category": cat,
            "tournament_label": tournament_label_for(r["year"], cat, r["continent"]),
            "round_reached": deepest_round_label(r),
            "champion": r["champions"] == "Y",
            "group": r["group"],
        }
        # Successor / canonical page (inherits the row).
        successor_slug = slug_for_any.get(r["cur_name"])
        if successor_slug:
            out[successor_slug].append({
                **base,
                "team_as": r["team"] if r["team"] != r["cur_name"] else None,
            })
        # Defunct standalone page, only when Team is a TRULY defunct entity
        # (not a rename). The artifact check above already filtered those.
        if r.get("team") in TRULY_DEFUNCT_TEAMS and r["team"] != r["cur_name"]:
            defunct_slug = slug_for_any.get(r["team"])
            if defunct_slug and defunct_slug != successor_slug:
                out[defunct_slug].append({**base, "team_as": None})
    for slug in out:
        out[slug].sort(key=lambda x: -(x["year"] or 0))
    return dict(out)


def build_finals(finals_rows, slug_for_any):
    """{slug: [final-match rows...]}. Each row may emit two entries (successor
    + defunct) following the same pattern as build_appearances."""
    out = defaultdict(list)
    for r in finals_rows:
        if r.get("cur_name") in ARTIFACT_TEAM_NAMES or r.get("team_as") in ARTIFACT_TEAM_NAMES:
            continue
        opp_slug = slug_for_any.get(r["opp_cur_name"]) if r["opp_cur_name"] else None
        category = (
            "WC" if r["category_world_cup"] == "Y"
            else "CON" if r["category_con_champ"] == "Y"
            else "INTER" if r["category_intercont"] == "Y"
            else "OTHER" if r["category_other"] == "Y"
            else None
        )
        base = {
            "year": r["year"],
            "season": r["season"],
            "competition": r["competition"],
            "category": category,
            "result": r["result"],
            "for_goals": r["for_goals"],
            "against_goals": r["against_goals"],
            "penalty_kicks": r["penalty_kicks"],
            "opp_cur_name": r["opp_cur_name"],
            "opp_slug": opp_slug,
            "opp_team_as": r["opp_team_as"] if r["opp_team_as"] != r["opp_cur_name"] else None,
            "stadium": r["stadium"],
            "stad_country": r["stad_country"],
            "stad_metro": r["stad_metro"],
        }
        successor_slug = slug_for_any.get(r["cur_name"])
        if successor_slug:
            out[successor_slug].append({
                **base,
                "team_as": r["team_as"] if r["team_as"] and r["team_as"] != r["cur_name"] else None,
            })
        if r.get("team_as") in TRULY_DEFUNCT_TEAMS and r["team_as"] != r["cur_name"]:
            defunct_slug = slug_for_any.get(r["team_as"])
            if defunct_slug and defunct_slug != successor_slug:
                out[defunct_slug].append({**base, "team_as": None})
    for slug in out:
        out[slug].sort(key=lambda x: -(x["year"] or 0))
    return dict(out)


def build_tournament_hubs(summary_rows, slug_for_cur_name):
    """One payload per hub. Includes all-time champions (year+team), most
    decorated teams (team+count), year range, and total editions count."""
    hubs = {}
    for hub in TOURNAMENT_HUBS:
        cat = hub["category"]
        matching = [
            r for r in summary_rows
            if hub["filter"](r)
            and r.get("year")
            and r["year"] <= MAX_DISPLAYED_TOURNAMENT_YEAR
        ]
        if not matching:
            continue
        # Year range + distinct editions (some categories have multiple
        # editions in a year if e.g. CONCACAF reorganized; treat distinct
        # year set as the canonical "editions" count for v1).
        years = sorted({r["year"] for r in matching if r["year"]})
        champ_rows = [r for r in matching if r["champions"] == "Y"]
        finalist_rows = [r for r in matching if r["finals"] == "Y" and r["champions"] != "Y"]

        # All-time champions: one row per (year, champion). For categories
        # like Gold Cup that had both NACC and Gold Cup editions, multiple
        # champions per year can happen; we surface them all.
        # Variable-name categories (INTER, OTHER) carry tournament_label per
        # row because the workbook lumps several distinct competitions under
        # one flag; fixed-name categories (WC, EUROS, etc.) leave the field
        # null since the hub label already names the tournament.
        variable_label = cat in ("INTER", "OTHER")
        champions_list = []
        for r in sorted(champ_rows, key=lambda x: -(x["year"] or 0)):
            slug = slug_for_cur_name.get(r["cur_name"])
            champions_list.append({
                "year": r["year"],
                "champion_cur_name": r["cur_name"],
                "champion_slug": slug,
                "champion_as": r["team"] if r["team"] != r["cur_name"] else None,
                "group": r["group"],
                "tournament_label": tournament_label_for(r["year"], cat, r["continent"]) if variable_label else None,
            })

        # Most decorated: count champions per Cur. Name.
        decorated = Counter(r["cur_name"] for r in champ_rows)
        most_decorated = []
        for cn, count in decorated.most_common():
            slug = slug_for_cur_name.get(cn)
            last = max(r["year"] for r in champ_rows if r["cur_name"] == cn)
            most_decorated.append({
                "cur_name": cn,
                "slug": slug,
                "champion_count": count,
                "last_won": last,
            })

        # Finalists (runners-up): for each year, who else was in the final?
        finalists_list = []
        for r in sorted(finalist_rows, key=lambda x: -(x["year"] or 0)):
            slug = slug_for_cur_name.get(r["cur_name"])
            finalists_list.append({
                "year": r["year"],
                "cur_name": r["cur_name"],
                "slug": slug,
                "tournament_label": tournament_label_for(r["year"], cat, r["continent"]) if variable_label else None,
            })

        hubs[hub["slug"]] = {
            "slug": hub["slug"],
            "label": hub["label"],
            "category": cat,
            "year_min": years[0] if years else None,
            "year_max": years[-1] if years else None,
            "editions": len(years),
            "champions": champions_list,
            "finalists": finalists_list,
            "most_decorated": most_decorated,
        }
    return hubs


def build_wc2026(wb, slug_for_cur_name):
    """Extract the 2026 World Cup group-stage standings and knockout-round
    match pairings from Int Tournaments. Output drives the live WC widget
    on the International Football index page. Standings populate
    automatically from the workbook once Ashwin records match results;
    pre-tournament every team sits at 0-0-0.

    Group stage: aggregate per-team-per-group W/D/L/GS/GA/Pts across all
    three matches and sort by Pts desc, GD desc, GS desc.

    Knockout: each match has two mirror rows in Int Tournaments (one per
    side). Collapse to a single match record keyed by (date, sorted-team-
    pair) so each fixture appears once."""
    ws = wb["Int Tournaments"]
    rows = list(ws.iter_rows(values_only=True))
    hdrs = [str(h).strip() if h else "" for h in rows[0]]
    def idx(name):
        return hdrs.index(name) if name in hdrs else None
    h = {
        "year": idx("Year"),
        "league": idx("Leag/Comp."),
        "round": idx("Comp. Rnd"),
        "team": idx("Team"),
        "opp_team": idx("Opp Team"),
        "cur": idx("Cur. Name"),
        "opp_cur": idx("Opp. Name"),
        "w": idx("W"),
        "d": idx("D"),
        "l": idx("L"),
        "gs": idx("GS"),
        "ga": idx("GA"),
        "pts": idx("Points"),
        "for": idx("For"),
        "ag": idx("Ag"),
        "wdl": idx("W/D/L"),
        "group": idx("Group"),
        "stadium": idx("Stadium"),
        "stad_country": idx("Stad. Country"),
        "stad_metro": idx("Stad. Metro Area"),
        "date": idx("Date"),
        "pk": idx("Penatly Kicks"),  # workbook typo preserved
    }

    def get(r, k):
        i = h[k]
        return r[i] if i is not None and i < len(r) else None

    wc_rows = []
    for r in rows[1:]:
        if not r:
            continue
        if get(r, "year") != 2026:
            continue
        if get(r, "league") != "World Cup":
            continue
        wc_rows.append(r)

    if not wc_rows:
        return None

    # ---- Group Stage aggregation ----
    group_stage = {}
    for r in wc_rows:
        if get(r, "round") != "Group Stage":
            continue
        team_cur = get(r, "cur") or get(r, "team")
        grp = get(r, "group")
        if not team_cur or not grp:
            continue
        grp_key = str(grp).strip()
        team_key = str(team_cur).strip()
        if grp_key not in group_stage:
            group_stage[grp_key] = {}
        bucket = group_stage[grp_key]
        if team_key not in bucket:
            bucket[team_key] = {
                "cur_name": team_key,
                "slug": slug_for_cur_name.get(team_key),
                "w": 0, "d": 0, "l": 0,
                "gs": 0, "ga": 0,
                "pts": 0,
                "matches": 0,
            }
        s = bucket[team_key]
        s["w"] += to_int(get(r, "w")) or 0
        s["d"] += to_int(get(r, "d")) or 0
        s["l"] += to_int(get(r, "l")) or 0
        s["gs"] += to_int(get(r, "gs")) or 0
        s["ga"] += to_int(get(r, "ga")) or 0
        s["pts"] += to_int(get(r, "pts")) or 0
        s["matches"] += 1

    group_stage_out = {}
    for grp_key, teams in sorted(group_stage.items()):
        team_list = list(teams.values())
        for t in team_list:
            t["gd"] = t["gs"] - t["ga"]
        team_list.sort(key=lambda x: (-x["pts"], -x["gd"], -x["gs"], x["cur_name"]))
        group_stage_out[grp_key] = team_list

    # ---- Knockout rounds ----
    # Pair up the two mirror rows per match by (date, sorted team pair).
    knockout_rounds = ["Round of 32", "Round of 16", "Quarterfinals",
                       "Semifinals", "Third Place Game", "Final"]
    knockout = {rn: [] for rn in knockout_rounds}
    seen_pairings = set()
    for r in wc_rows:
        rnd = get(r, "round")
        if rnd not in knockout_rounds:
            continue
        team_cur = str(get(r, "cur") or get(r, "team") or "").strip()
        opp_cur = str(get(r, "opp_cur") or get(r, "opp_team") or "").strip()
        if not team_cur or not opp_cur:
            continue
        date_val = get(r, "date")
        date_str = date_val.strftime("%Y-%m-%d") if hasattr(date_val, "strftime") else (str(date_val) if date_val else None)
        pairing = tuple(sorted([team_cur, opp_cur]))
        key = (pairing, date_str)
        if key in seen_pairings:
            continue
        seen_pairings.add(key)
        team_score = to_int(get(r, "for"))
        opp_score = to_int(get(r, "ag"))
        # Status: "played" if scores are set, else "scheduled"
        played = team_score is not None and opp_score is not None and (team_score != 0 or opp_score != 0)
        knockout[rnd].append({
            "team_cur_name": team_cur,
            "team_slug": slug_for_cur_name.get(team_cur),
            "opp_cur_name": opp_cur,
            "opp_slug": slug_for_cur_name.get(opp_cur),
            "team_score": team_score,
            "opp_score": opp_score,
            "penalty_kicks": to_int(get(r, "pk")),
            "result": str(get(r, "wdl") or "").strip() if get(r, "wdl") else None,
            "stadium": str(get(r, "stadium") or "").strip() or None,
            "stad_country": str(get(r, "stad_country") or "").strip() or None,
            "stad_metro": str(get(r, "stad_metro") or "").strip() or None,
            "date": date_str,
            "played": played,
        })

    for rn in knockout_rounds:
        knockout[rn].sort(key=lambda m: m["date"] or "")

    # Order the Round of 32 by official FIFA 2026 match number (73..88). The
    # read-time bracket advancement (lib/wc2026Standings.ts) maps R32 slot i to
    # match 73+i, so the slots MUST be in official order or winners land in the
    # wrong Round-of-16 slots. Same-day matches otherwise keep arbitrary workbook
    # order. Derived from the group winners/runners-up + FIFA's group-slot map;
    # falls back to date order pre-knockout or if teams are not yet set.
    def _wc_r32_official_order(r32, groups):
        if len(r32) != 16 or len(groups) != 12:
            return None
        norm = {str(k).replace("Group", "").strip(): v for k, v in groups.items()}
        W = {}; RU = {}
        for g, tl in norm.items():
            if len(tl) < 2:
                return None
            W[g] = tl[0]["cur_name"]; RU[g] = tl[1]["cur_name"]
        # match_no -> (roleA, roleB); role is ("W"|"RU", group) for the fixed
        # slot, None for the variable best-third slot (matched via the other side).
        T = {73: (("RU","A"),("RU","B")), 74: (("W","E"),None), 75: (("W","F"),("RU","C")),
             76: (("W","C"),("RU","F")), 77: (("W","I"),None), 78: (("RU","E"),("RU","I")),
             79: (("W","A"),None), 80: (("W","L"),None), 81: (("W","D"),None), 82: (("W","G"),None),
             83: (("RU","K"),("RU","L")), 84: (("W","H"),("RU","J")), 85: (("W","B"),None),
             86: (("W","J"),("RU","H")), 87: (("W","K"),None), 88: (("RU","D"),("RU","G"))}
        def fixed(role):
            if role is None:
                return None
            kind, grp = role
            return W.get(grp) if kind == "W" else RU.get(grp)
        used = set(); order = {}
        for mno, (ra, rb) in T.items():
            need = [x for x in (fixed(ra), fixed(rb)) if x]
            if not need:
                return None
            cand = None
            for idx, m in enumerate(r32):
                if idx in used:
                    continue
                pair = {m.get("team_cur_name"), m.get("opp_cur_name")}
                if all(f in pair for f in need):
                    cand = idx; break
            if cand is None:
                return None
            used.add(cand); order[mno] = r32[cand]
        if len(order) != 16:
            return None
        return [order[m] for m in range(73, 89)]

    _r32_official = _wc_r32_official_order(knockout["Round of 32"], group_stage_out)
    if _r32_official:
        knockout["Round of 32"] = _r32_official

    return {
        "tournament": {
            "name": "FIFA World Cup 2026",
            "year": 2026,
            "starts_iso": "2026-06-11",
        },
        "group_stage": group_stage_out,
        "knockout": knockout,
    }


def build_slug_lookup(teams):
    """Normalized name -> slug. Includes both Cur. Name and (workbook) Name
    so cross-source joins (e.g. national-teams.tsv) can resolve either."""
    out = {}
    for t in teams:
        for k in (t["cur_name"], t["name"]):
            if not k:
                continue
            out[normalize_team_name(k)] = t["slug"]
    return out


# ---------- Main ----------

def main():
    src = find_source()
    print(f"Reading: {src}")
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)

    print("\nLoading federation lookup from national-teams.tsv...")
    fed_lookup = load_federation_lookup()
    print(f"  {len(fed_lookup)} entries")

    print("\nReading Int Totals...")
    totals_rows = read_int_totals(wb)
    print(f"  {len(totals_rows)} active team rows")

    print("\nReading rank snapshot dates (Int Totals D1/E1)...")
    rank_snapshots = read_rank_snapshots(wb)
    print(f"  ELO as of {rank_snapshots['elo']}, FIFA as of {rank_snapshots['fifa']}")

    print("\nReading Int Summary...")
    summary_rows = read_int_summary(wb)
    print(f"  {len(summary_rows)} appearance rows")

    # Patch missing INTER champion flags for one-off early editions whose
    # workbook rows exist but lack the Champions flag. Keyed by (year, cur_name).
    INTER_CHAMPION_PATCHES = {
        (1981, "Uruguay"),   # Mundialito 1981
        (1985, "France"),    # Artemio Franchi Trophy 1985
    }
    patched = 0
    for r in summary_rows:
        if r.get("intercont_champ") == "Y" and (r.get("year"), r.get("cur_name")) in INTER_CHAMPION_PATCHES:
            r["champions"] = "Y"
            r["finals"] = "Y"
            patched += 1
    if patched:
        print(f"  Patched {patched} INTER champion row(s) (Mundialito/Artemio Franchi)")

    print("\nReading Int Tournaments (Comp. Rnd = Final only)...")
    finals_rows = read_finals(wb)
    print(f"  {len(finals_rows)} final-match rows")

    print("\nBuilding team index...")
    teams = build_teams_index(totals_rows, summary_rows, fed_lookup)
    print(f"  {len(teams)} canonical teams")

    slug_for_cur_name = {t["cur_name"]: t["slug"] for t in teams}
    # Extended map: also resolve the workbook's `name` field (which can
    # differ from cur_name for the active set, and which equals the defunct
    # name for standalone defunct entries). Lets build_appearances /
    # build_finals route a row to either the successor or the standalone
    # defunct page using a single lookup.
    slug_for_any = dict(slug_for_cur_name)
    for t in teams:
        if t.get("name") and t["name"] != t["cur_name"]:
            slug_for_any[t["name"]] = t["slug"]

    print("\nBuilding appearances...")
    appearances = build_appearances(summary_rows, slug_for_any)
    print(f"  {sum(len(v) for v in appearances.values())} appearance rows across {len(appearances)} teams")

    # Honors index and longevity per team. Honors uses the team's appearance
    # rows to apply continent-tier weights (Gold Cup is not a Euros). Longevity
    # walks summary_rows once per team; cost is negligible at this dataset size.
    print("\nComputing honors index and longevity...")
    for t in teams:
        honors_index, breakdown = compute_honors(t, appearances.get(t["slug"], []))
        span, decades = compute_team_longevity(t["slug"], summary_rows, slug_for_any)
        t["honors_index"] = honors_index
        t["honors_breakdown"] = breakdown
        t["tournament_span_years"] = span
        t["decade_coverage"] = decades
    top_honors = sorted(teams, key=lambda t: -(t.get("honors_index") or 0))[:5]
    print("  Top 5 by honors index: " + ", ".join(
        f"{t['cur_name']} ({t['honors_index']})" for t in top_honors
    ))

    print("\nBuilding similar-teams engine...")
    similar = build_similar_teams(teams)
    print(f"  {len(similar)} teams with neighbor lists")

    print("\nBuilding honors leaderboard...")
    leaderboard_payload = build_honors_leaderboard(teams)
    print(f"  top {len(leaderboard_payload['leaderboard'])} entries")

    print("\nBuilding finals...")
    finals_by_slug = build_finals(finals_rows, slug_for_any)
    print(f"  {sum(len(v) for v in finals_by_slug.values())} final-match rows across {len(finals_by_slug)} teams")

    print("\nBuilding tournament hubs...")
    tournament_hubs = build_tournament_hubs(summary_rows, slug_for_cur_name)
    print(f"  {len(tournament_hubs)} hubs: {sorted(tournament_hubs.keys())}")

    print("\nBuilding 2026 World Cup bundle...")
    wc2026 = build_wc2026(wb, slug_for_cur_name)
    if wc2026:
        groups = len(wc2026["group_stage"])
        knockout_total = sum(len(v) for v in wc2026["knockout"].values())
        print(f"  {groups} groups, {knockout_total} knockout matches")
    else:
        print("  No 2026 WC rows found in workbook")

    print("\nBuilding slug lookup...")
    slug_lookup = build_slug_lookup(teams)
    print(f"  {len(slug_lookup)} entries")

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    payload = {
        "generated_at": __import__("datetime").datetime.now().isoformat(timespec="seconds"),
        "source": str(src),
        "rank_snapshots": rank_snapshots,
        "teams": teams,
    }

    def dump(name, data):
        path = OUT_DIR / name
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, separators=(",", ":"))
        print(f"  Wrote {path}  ({path.stat().st_size:,} bytes)")

    print("\nWriting outputs...")
    dump("index.json", payload)
    dump("appearances.json", appearances)
    dump("finals.json", finals_by_slug)
    dump("tournaments.json", tournament_hubs)
    dump("slug-lookup.json", slug_lookup)
    dump("honors-leaderboard.json", leaderboard_payload)
    dump("similar-teams.json", similar)
    if wc2026 and os.environ.get("WRITE_WC2026") == "1":
        dump("wc2026.json", wc2026)
    elif wc2026:
        print("  Skipping wc2026.json (live WC simulator owns it; set WRITE_WC2026=1 to regenerate the draw)")

    print("\nDone.")


if __name__ == "__main__":
    main()

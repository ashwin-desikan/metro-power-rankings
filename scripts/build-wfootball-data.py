#!/usr/bin/env python3
"""Build the Women's Club Football portal data from OtherLeagues.xlsx.

Source sheet: "Women's Club Football" (honors / finals history).
Columns: Competition, Country, Season(year), Winner, Score, Runner-up,
         Winner (Canonical), Runner-up (Canonical).

Emits public/data/football/womens-football.json with:
  meta          — counts + latest year
  competitions  — one block per competition in the same shape the men's
                  European-tournaments hub uses (champions / most_decorated /
                  current), so the /teams/wfootball pages mirror the existing
                  tournament hub renderer.
  clubs         — per canonical club: honors across all competitions, plus a
                  best-effort metro_slug matched against the W Football markers
                  already in public/data/sports/all-teams.json.

Run after any edit to the Women's Club Football sheet:
  python scripts/build-wfootball-data.py
The source workbook lives at the project root (OtherLeagues.xlsx) so the
portal is reproducible; do not rely on the ephemeral uploads copy.
"""

import json
import os
import re
import sys
import tempfile
import unicodedata
from collections import defaultdict

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(ROOT, "OtherLeagues.xlsx")
SHEET = "Women's Club Football"
OUT = os.path.join(ROOT, "public", "data", "football", "womens-football.json")
ALL_TEAMS = os.path.join(ROOT, "public", "data", "sports", "all-teams.json")

# Per-competition editorial config. Order drives hub display order.
COMP_CONFIG = {
    "UEFA Women's Champions League": dict(
        slug="uwcl", short="UWCL", label="UEFA Women's Champions League",
        kind="cup", region="Europe", order=1,
        notes="Europe's premier women's club competition. Run as the UEFA Women's Cup from 2001-02, rebranded the UEFA Women's Champions League in 2009.",
    ),
    "FIFA Women's Champions Cup": dict(
        slug="fifa-womens-champions-cup", short="FIFA Women's Champions Cup",
        label="FIFA Women's Champions Cup", kind="cup", region="Global", order=2,
        notes="FIFA's new global women's club competition, contested for the first time in 2026.",
    ),
    "WSL / Premier League": dict(
        slug="wsl", short="WSL", label="Women's Super League (England)",
        kind="league", region="England", order=3,
        notes="England's top flight. The FA Women's Premier League National Division from 1991-92, replaced by the Women's Super League in 2011.",
    ),
    "Women's FA Cup": dict(
        slug="womens-fa-cup", short="Women's FA Cup", label="Women's FA Cup (England)",
        kind="cup", region="England", order=4,
        notes="England's primary women's knockout cup, first contested in 1970-71.",
    ),
    "Primera División / Liga F": dict(
        slug="liga-f", short="Liga F", label="Liga F / Primera División (Spain)",
        kind="league", region="Spain", order=5,
        notes="Spain's top flight. The Primera División (Superliga) from 2001-02, rebranded the professional Liga F in 2022-23.",
    ),
    "NWSL Championship": dict(
        slug="nwsl-championship", short="NWSL Championship", label="NWSL Championship (United States)",
        kind="cup", region="United States", order=6,
        notes="The end-of-season playoff final of the National Women's Soccer League, the United States top flight, since 2013.",
    ),
    "NWSL Shield": dict(
        slug="nwsl-shield", short="NWSL Shield", label="NWSL Shield (United States)",
        kind="league", region="United States", order=7,
        notes="Awarded to the NWSL regular-season points leader, distinct from the playoff Championship.",
    ),
}

# ---- helpers ----------------------------------------------------------------

def slugify(s):
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.lower()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    return s.strip("-")

# Tokens stripped when normalizing a club name for metro matching.
_STOP = {
    "women", "womens", "ladies", "feminin", "feminines", "feminine",
    "femenino", "femenina", "femeni", "femenil", "frauen", "feminil",
    "fc", "cf", "sc", "afc", "ac", "club", "the", "lfc", "wfc", "fk", "sk",
}

def norm_name(s):
    if not s:
        return ""
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.lower()
    s = re.sub(r"\(.*?\)", " ", s)          # drop "(women)" etc.
    s = re.sub(r"[^a-z0-9]+", " ", s)
    toks = [t for t in s.split() if t and t not in _STOP]
    return "".join(toks)

def safe_str(v):
    return "" if v is None else str(v).strip()

# ---- metro lookup from existing W Football markers --------------------------

def _valid_metro_slugs():
    path = os.path.join(ROOT, "public", "data", "metros.json")
    if not os.path.exists(path):
        return None  # cannot validate; accept all
    md = json.load(open(path, encoding="utf-8"))
    rows = md if isinstance(md, list) else md.get("metros", md)
    return {r.get("slug") for r in rows if isinstance(r, dict)}

def build_metro_index():
    if not os.path.exists(ALL_TEAMS):
        return {}
    valid = _valid_metro_slugs()
    data = json.load(open(ALL_TEAMS, encoding="utf-8"))
    rows = data if isinstance(data, list) else data.get("teams", data)
    idx = {}
    for r in rows:
        if r.get("sport") != "W Football":
            continue
        key = norm_name(r.get("team"))
        if not key or key in idx:
            continue
        slug = r.get("metro_slug")
        if valid is not None and slug not in valid:
            slug = None  # do not link to a metro page that does not exist
        idx[key] = {"metro_slug": slug, "metro": r.get("metro")}
    return idx

# ---- load source ------------------------------------------------------------

def load_rows():
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    ws = wb[SHEET]
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        comp = safe_str(r[0])
        if not comp:
            continue
        out.append({
            "competition": comp,
            "country": safe_str(r[1]),
            "year": r[2],
            "winner_disp": safe_str(r[3]),
            "score": safe_str(r[4]),
            "runner_disp": safe_str(r[5]),
            "winner": safe_str(r[6]) or safe_str(r[3]),
            "runner": safe_str(r[7]) or safe_str(r[5]),
        })
    return out

# ---- build ------------------------------------------------------------------

def main():
    rows = load_rows()
    metro_idx = build_metro_index()
    # metro_slug -> country, used to backfill the nation for clubs whose only
    # honors are continental (e.g. Lyon, Wolfsburg) and therefore have no
    # domestic-competition country.
    metro_country = {}
    _mp = os.path.join(ROOT, "public", "data", "metros.json")
    if os.path.exists(_mp):
        _md = json.load(open(_mp, encoding="utf-8"))
        _mrows = _md if isinstance(_md, list) else _md.get("metros", _md)
        metro_country = {r.get("slug"): r.get("country") for r in _mrows if isinstance(r, dict)}

    def metro_for(name):
        return metro_idx.get(norm_name(name), {"metro_slug": None, "metro": None})

    # club aggregation across all competitions
    clubs = {}  # slug -> record
    def club_rec(name):
        slug = slugify(name)
        if slug not in clubs:
            m = metro_for(name)
            clubs[slug] = {
                "slug": slug, "name": name,
                "metro_slug": m["metro_slug"], "metro": m["metro"],
                "country": None,
                "honors": {},  # comp_slug -> {label, kind, titles, runner_ups, title_years, runner_up_years}
                "total_titles": 0, "total_finals": 0,
            }
        return clubs[slug]

    competitions = []
    by_comp = defaultdict(list)
    for r in rows:
        by_comp[r["competition"]].append(r)

    for comp_name, cfg in sorted(COMP_CONFIG.items(), key=lambda kv: kv[1]["order"]):
        crows = by_comp.get(comp_name, [])
        years = [r["year"] for r in crows if isinstance(r["year"], int)]
        # decided editions = rows with a winner
        decided = [r for r in crows if r["winner"]]

        champions = []
        # per-club tallies within this competition
        won = defaultdict(list)      # slug -> [years]
        lost = defaultdict(list)     # slug -> [years]
        names = {}                   # slug -> display canonical name
        for r in sorted(crows, key=lambda x: (x["year"] or 0)):
            if not r["winner"]:
                continue
            wslug = slugify(r["winner"])
            names[wslug] = r["winner"]
            won[wslug].append(r["year"])
            rslug = slugify(r["runner"]) if r["runner"] else None
            if rslug:
                names[rslug] = r["runner"]
                lost[rslug].append(r["year"])
            champions.append({
                "year": r["year"],
                "cur_name": r["winner"],
                "slug": wslug,
                "score": r["score"] or None,
                "runner_up": r["runner"] or None,
                "runner_up_slug": rslug,
            })
            # feed club-level aggregation
            wc = club_rec(r["winner"])
            wc["country"] = wc["country"] or (r["country"] if r["country"] not in ("Europe", "Global") else None)
            h = wc["honors"].setdefault(cfg["slug"], {
                "competition_slug": cfg["slug"], "competition_label": cfg["short"],
                "kind": cfg["kind"], "region": cfg["region"],
                "titles": 0, "runner_ups": 0, "title_years": [], "runner_up_years": [],
            })
            h["titles"] += 1
            h["title_years"].append(r["year"])
            wc["total_titles"] += 1
            wc["total_finals"] += 1
            if r["runner"]:
                rc = club_rec(r["runner"])
                rc["country"] = rc["country"] or (r["country"] if r["country"] not in ("Europe", "Global") else None)
                rh = rc["honors"].setdefault(cfg["slug"], {
                    "competition_slug": cfg["slug"], "competition_label": cfg["short"],
                    "kind": cfg["kind"], "region": cfg["region"],
                    "titles": 0, "runner_ups": 0, "title_years": [], "runner_up_years": [],
                })
                rh["runner_ups"] += 1
                rh["runner_up_years"].append(r["year"])
                rc["total_finals"] += 1

        champions.sort(key=lambda c: c["year"], reverse=True)

        # most decorated within the competition
        all_slugs = set(won) | set(lost)
        most = []
        for s in all_slugs:
            cc = len(won[s])
            fl = len(lost[s])
            most.append({
                "cur_name": names[s], "slug": s,
                "champion_count": cc,
                "finals_count": cc + fl,
                "finals_lost": fl,
                "last_won": max(won[s]) if won[s] else None,
                "last_final": max((won[s] + lost[s])) if (won[s] or lost[s]) else None,
                "metro_slug": metro_for(names[s])["metro_slug"],
            })
        most.sort(key=lambda d: (-d["champion_count"], -d["finals_count"], d["cur_name"]))

        latest_year = max(years) if years else None
        latest_row = None
        for r in crows:
            if r["year"] == latest_year:
                latest_row = r
                break
        current = None
        if latest_row is not None:
            if latest_row["winner"]:
                current = {
                    "year": latest_year, "decided": True,
                    "champion": latest_row["winner"], "champion_slug": slugify(latest_row["winner"]),
                    "runner_up": latest_row["runner"] or None,
                    "runner_up_slug": slugify(latest_row["runner"]) if latest_row["runner"] else None,
                    "score": latest_row["score"] or None,
                }
            else:
                current = {"year": latest_year, "decided": False}

        competitions.append({
            "slug": cfg["slug"], "label": cfg["label"], "short_label": cfg["short"],
            "kind": cfg["kind"], "region": cfg["region"], "country": cfg["region"],
            "active": True, "notes": cfg["notes"],
            "year_min": min(years) if years else None,
            "year_max": latest_year,
            "editions": len(decided),
            "champions": champions,
            "most_decorated": most,
            "current": current,
        })

    # finalize clubs
    club_list = []
    for c in clubs.values():
        honors = sorted(c["honors"].values(), key=lambda h: (-h["titles"], -h["runner_ups"], h["competition_label"]))
        for h in honors:
            h["title_years"] = sorted(h["title_years"])
            h["runner_up_years"] = sorted(h["runner_up_years"])
        all_years = []
        for h in honors:
            all_years += h["title_years"]
        c2 = {
            "slug": c["slug"], "name": c["name"],
            "metro_slug": c["metro_slug"], "metro": c["metro"], "country": c["country"] or metro_country.get(c["metro_slug"]),
            "total_titles": c["total_titles"], "total_finals": c["total_finals"],
            "first_title": min(all_years) if all_years else None,
            "last_title": max(all_years) if all_years else None,
            "honors": honors,
        }
        club_list.append(c2)
    club_list.sort(key=lambda c: (-c["total_titles"], -c["total_finals"], c["name"]))

    matched = sum(1 for c in club_list if c["metro_slug"])
    latest = max((comp["year_max"] for comp in competitions if comp["year_max"]), default=None)

    payload = {
        "meta": {
            "competitions": len(competitions),
            "clubs": len(club_list),
            "clubs_with_metro": matched,
            "latest_year": latest,
        },
        "competitions": competitions,
        "clubs": club_list,
    }

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    fd, tmp = tempfile.mkstemp(dir=os.path.dirname(OUT), prefix=".wf-", suffix=".tmp")
    with os.fdopen(fd, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, separators=(",", ":"))
    os.replace(tmp, OUT)

    print(f"wrote {OUT}")
    print(f"  competitions: {len(competitions)}")
    print(f"  clubs: {len(club_list)}  (metro-matched: {matched})")
    for comp in competitions:
        cur = comp["current"]
        cur_s = (f"{cur['year']} {cur.get('champion','TBD')}" if cur else "—")
        print(f"  - {comp['short_label']:24} editions={comp['editions']:3} top={comp['most_decorated'][0]['cur_name'] if comp['most_decorated'] else '—'} current={cur_s}")

# ─── Women's World Cup (national teams) ──────────────────────────────────────
# Source: the "Int Tournaments" sheet of the grand men's football workbook
# (Champions League-201516.xlsx at the repo root), filtered to the Women's
# World Cup competition. Emits public/data/football/womens-world-cup.json with
# editions and per-nation records, mirroring the men's international data shape.

WC_SRC = os.path.join(ROOT, "Champions League-201516.xlsx")
WC_OUT = os.path.join(ROOT, "public", "data", "football", "womens-world-cup.json")
_WC_RANK = {"Group Stage": 1, "Round of 16": 2, "Quarterfinals": 3, "Semifinals": 4, "Third Place Match": 5, "Final": 6}


def _strip_w(s):
    return re.sub(r"\s*\(W\)\s*$", "", str(s)).strip() if s else None


def build_womens_world_cup():
    if not os.path.exists(WC_SRC):
        print(f"  [skip] {WC_SRC} not found; womens-world-cup.json not built")
        return
    wb = openpyxl.load_workbook(WC_SRC, read_only=True, data_only=True)
    ws = wb["Int Tournaments"]
    header = list(next(ws.iter_rows(values_only=True)))
    ix = {name: i for i, name in enumerate(header)}
    iComp = ix["Leag/Comp."]
    iY, iRnd, iTeam, iWDL, iOpp = ix["Year"], ix["Comp. Rnd"], ix["Team"], ix["W/D/L"], ix["Opp Team"]
    iFor, iAg, iHost, iTrophy, iCont = ix["For"], ix["Ag"], ix["Stad. Country"], ix["Trophy Won"], ix["Continent"]

    rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[iComp] == "Women's World Cup"]

    ed = {}
    nat_year_rounds = defaultdict(set)
    nat_continent = {}
    for r in rows:
        y = r[iY]
        team = _strip_w(r[iTeam])
        opp = _strip_w(r[iOpp])
        rnd = str(r[iRnd] or "")
        e = ed.setdefault(y, {"year": y, "host": None, "champion": None, "runner_up": None,
                              "third": None, "fourth": None, "final_score": None})
        if rnd == "Final":
            e["host"] = r[iHost]
            if str(r[iTrophy]) == "Y":
                e["champion"], e["runner_up"] = team, opp
                e["final_score"] = f"{r[iFor]}–{r[iAg]}"
        if rnd == "Third Place Match" and str(r[iWDL]) == "W":
            e["third"], e["fourth"] = team, opp
        if team:
            nat_year_rounds[(team, y)].add(rnd)
            if r[iCont] and team not in nat_continent:
                nat_continent[team] = r[iCont]

    def finish_for(team, y):
        e = ed[y]
        if e["champion"] == team:
            return ("Champion", 1)
        if e["runner_up"] == team:
            return ("Runner-up", 2)
        if e["third"] == team:
            return ("Third place", 3)
        if e["fourth"] == team:
            return ("Fourth place", 4)
        deepest = max((_WC_RANK.get(rr, 0) for rr in nat_year_rounds[(team, y)]), default=0)
        if deepest >= 3:
            return ("Quarter-finals", 5)
        if deepest == 2:
            return ("Round of 16", 6)
        return ("Group stage", 7)

    teams = sorted({t for (t, _y) in nat_year_rounds if t})
    nations = []
    for team in teams:
        recs = []
        titles, title_years, finals, final_years = 0, [], 0, []
        best_rank, best_finish = 99, None
        for y in sorted(ed):
            if (team, y) not in nat_year_rounds:
                continue
            label, rank = finish_for(team, y)
            recs.append({"year": y, "host": ed[y]["host"], "finish": label, "rank": rank,
                         "final_score": ed[y]["final_score"] if rank <= 2 else None})
            if rank == 1:
                titles += 1
                title_years.append(y)
            if rank <= 2:
                finals += 1
                final_years.append(y)
            if rank < best_rank:
                best_rank, best_finish = rank, label
        nations.append({
            "slug": slugify(team), "name": team, "continent": nat_continent.get(team),
            "appearances": len(recs), "titles": titles, "title_years": title_years,
            "finals": finals, "final_years": final_years,
            "best_rank": best_rank, "best_finish": best_finish,
            "first_appearance": recs[0]["year"] if recs else None,
            "last_appearance": recs[-1]["year"] if recs else None,
            "results": recs,
        })
    nations.sort(key=lambda n: (-n["titles"], -n["finals"], n["best_rank"], -n["appearances"], n["name"]))

    editions = []
    for y in sorted(ed, reverse=True):
        e = ed[y]
        editions.append({**e,
                         "champion_slug": slugify(e["champion"]) if e["champion"] else None,
                         "runner_up_slug": slugify(e["runner_up"]) if e["runner_up"] else None,
                         "third_slug": slugify(e["third"]) if e["third"] else None,
                         "fourth_slug": slugify(e["fourth"]) if e["fourth"] else None})

    payload = {
        "meta": {"label": "FIFA Women's World Cup", "slug": "womens-world-cup",
                 "editions": len(editions), "nations": len(nations),
                 "year_min": min(ed) if ed else None, "year_max": max(ed) if ed else None},
        "editions": editions,
        "nations": nations,
    }
    os.makedirs(os.path.dirname(WC_OUT), exist_ok=True)
    fd, tmp = tempfile.mkstemp(dir=os.path.dirname(WC_OUT), prefix=".wwc-", suffix=".tmp")
    with os.fdopen(fd, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, separators=(",", ":"))
    os.replace(tmp, WC_OUT)
    print(f"wrote {WC_OUT}: {len(editions)} editions, {len(nations)} nations; "
          f"latest champion {editions[0]['champion']} ({editions[0]['year']})")


if __name__ == "__main__":
    main()
    build_womens_world_cup()

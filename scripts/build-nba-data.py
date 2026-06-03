#!/usr/bin/env python3
"""
Build NBA team-pages data from NBA.xlsx (the canonical workbook).

Mirrors scripts/build-mlb-data.py and scripts/build-nfl-data.py. See those
scripts for the architectural patterns; NBA-specific notes inline below.

Outputs:
  public/data/nba/franchises.json
  public/data/nba/championships.json              one entry per title per franchise
                                                  (era: 'baa' | 'nba' | 'aba')
  public/data/nba/championship-appearances.json   Finals appearances by franchise
  public/data/nba/stadium-history.json            arena-era rows per franchise
  public/data/nba/award-winners.json              MVP / DPOY / ROY / COY / MIP /
                                                  6MOY / CPOY grouped by franchise
  public/data/nba/all-nba-selections.json         1st / 2nd / 3rd team All-NBA
                                                  selections grouped by franchise
                                                  with year + player + tier
  public/data/nba/all-star-counts.json            career player-season All-Star
                                                  selections per franchise
  public/data/nba/historical.json                 defunct franchises (incl. ABA-only)
  public/data/nba/historical-seasons.json         per-franchise seasons for defunct
  public/data/nba/seasons-by-team.json            per-franchise season-by-season rows
  public/data/nba/top-games-by-team.json          top postseason games per franchise
                                                  (Game Score column blank for now)
  public/data/nba/top-games-all-time.json         top playoff games leaguewide (blank
                                                  scores; surfaced ordering by round +
                                                  recency until Game Score lands)
  public/data/nba/top-games-by-decade.json        top playoff games per decade
  public/data/nba/playoff-state.json              current-postseason status per
                                                  franchise (R1 eliminated, semis,
                                                  CF, lost finals, NBA champion,
                                                  or still active)

Championship era split for chip rendering:
  - BAA (1947-49) folds into NBA gold (continuous lineage)
  - ABA (1968-76) renders in slate (distinct visual tier)
  - NBA (1949+) renders in gold

Canonical join key throughout: Name column in Year by Year (col AI / index 34),
Totals (col AR / index 43), Detailed Playoffs (col AU / index 46), Awards
(col H / index 7), All-Stars (col Q / index 16). See workbook's Claude Notes.

Usage:
  python scripts/build-nba-data.py
  python scripts/build-nba-data.py /path/to/NBA.xlsx
"""

import json
import os
import re
import sys
from pathlib import Path
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl",
                           "--quiet", "--break-system-packages"])
    import openpyxl


# -------- Constants --------

REPO_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_SOURCE_CANDIDATES = [
    Path(os.path.expanduser("~/OneDrive/Excel Files/NBA.xlsx")),
    Path(os.path.expanduser("~/OneDrive/Excel Files/NBA_backup.xlsx")),
    REPO_ROOT / "NBA.xlsx",
    REPO_ROOT / "NBA_backup.xlsx",
    REPO_ROOT / "data" / "nba-source" / "NBA.xlsx",
]

# Game-level data and the documented Game Score (col BU) come from a SEPARATE
# workbook (NBA_RegSeason.xlsx, "Regular Season" sheet) — NOT the Detailed
# Playoffs sheet in NBA.xlsx. Override with NBA_REGSEASON_PATH if needed.
REGSEASON_SOURCE_CANDIDATES = [
    Path(os.path.expanduser("~/OneDrive/Excel Files/NBA_RegSeason.xlsx")),
    REPO_ROOT / "NBA_RegSeason.xlsx",
    REPO_ROOT / "data" / "nba-source" / "NBA_RegSeason.xlsx",
]

OUT_DIR = Path(os.environ.get("NBA_OUT_DIR") or (REPO_ROOT / "public" / "data" / "nba"))

# Curated award list. Order mirrors NFL/MLB convention: career-arc awards
# first (MVP, DPOY, ROY), then leadership (COY), then improvement/specialty
# awards (MIP, 6MOY, CPOY). Finals MVP NOT in v1 per scope conversation but
# easy to add later.
AWARD_ORDER = [
    "MVP",
    "DPOY",
    "ROY",
    "COY",
    "MIP",
    "6MOY",
    "CPOY",
]

# Workbook stores varied phrasings in the Awards column. Normalize to the
# label we display. Anything outside this map is excluded from the team
# Awards block (All-NBA selections are surfaced in their own block, not
# here). Pattern: lowercased substring match.
AWARD_LABEL_NORMALIZE = {
    "mvp": "MVP",
    "most valuable player": "MVP",
    "defensive player of the year": "DPOY",
    "dpoy": "DPOY",
    "rookie of the year": "ROY",
    "roy": "ROY",
    "coach of the year": "COY",
    "coy": "COY",
    "most improved": "MIP",
    "mip": "MIP",
    "sixth man": "6MOY",
    "6moy": "6MOY",
    "6th man": "6MOY",
    "clutch player": "CPOY",
    "cpoy": "CPOY",
}

ALL_NBA_TIER_NORMALIZE = {
    "1st team all nba": "1st",
    "1st team all-nba": "1st",
    "first team all nba": "1st",
    "first team all-nba": "1st",
    "2nd team all nba": "2nd",
    "2nd team all-nba": "2nd",
    "second team all nba": "2nd",
    "second team all-nba": "2nd",
    "3rd team all nba": "3rd",
    "3rd team all-nba": "3rd",
    "third team all nba": "3rd",
    "third team all-nba": "3rd",
    "all-nba": "1st",  # rare bare form, default to 1st
}

# Display-name overrides (city-as-brand). NBA has fewer of these than MLB
# but Golden State is the obvious one (San Francisco is the workbook city,
# but the franchise brands as Golden State).
DISPLAY_NAME_OVERRIDES = {
    "Warriors": ("Golden State", "Warriors"),
    "Pelicans": ("New Orleans", "Pelicans"),
}

# Championship era classifier (League column on the championship-year row).
def championship_era(league):
    if not league:
        return "nba"
    s = str(league).strip().upper()
    if s == "ABA":
        return "aba"
    if s == "BAA":
        return "baa"
    return "nba"


# -------- Helpers --------

def slugify(s):
    s = (s or "").lower().strip()
    s = re.sub(r"[àáâãäå]", "a", s)
    s = re.sub(r"[èéêë]", "e", s)
    s = re.sub(r"[ìíîï]", "i", s)
    s = re.sub(r"[òóôõö]", "o", s)
    s = re.sub(r"[ùúûü]", "u", s)
    s = re.sub(r"[ñ]", "n", s)
    s = re.sub(r"[^a-z0-9\s\-/]", "", s)
    s = re.sub(r"[/]", "-", s)
    s = re.sub(r"\s+", "-", s)
    s = re.sub(r"-+", "-", s)
    return s.strip("-")


def metro_slugify(metro_name):
    if not metro_name:
        return None
    return slugify(metro_name)


def franchise_slug(canonical):
    return slugify((canonical or "").replace("(", " ").replace(")", " ")).rstrip("-")


def safe_int(val, default=0):
    if val is None or val == "":
        return default
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return default


def safe_float(val, default=0.0):
    if val is None or val == "":
        return default
    try:
        n = float(val)
        if n != n:  # NaN
            return default
        return n
    except (ValueError, TypeError):
        return default


def safe_str(val):
    if val is None:
        return ""
    return str(val).strip()


def is_truthy_yn(val):
    if val is None or val == "":
        return False
    s = str(val).strip().upper()
    return s in ("Y", "YES", "1", "TRUE")


def normalize_award_name(raw):
    if not raw:
        return None
    low = str(raw).strip().lower()
    for key, label in AWARD_LABEL_NORMALIZE.items():
        if key in low and "all nba" not in low and "all-nba" not in low:
            return label
    return None


def normalize_all_nba_tier(raw):
    if not raw:
        return None
    low = str(raw).strip().lower()
    for key, tier in ALL_NBA_TIER_NORMALIZE.items():
        if key in low:
            return tier
    return None


# -------- Workbook loader --------

def find_source(cli_path=None):
    if cli_path:
        p = Path(cli_path)
        if p.exists():
            return p
        print(f"Provided path not found: {p}", file=sys.stderr)
        sys.exit(1)
    for cand in DEFAULT_SOURCE_CANDIDATES:
        if cand.exists():
            return cand
    print("Could not find NBA.xlsx in any default location. Pass a path:",
          file=sys.stderr)
    print("  python scripts/build-nba-data.py /path/to/NBA.xlsx", file=sys.stderr)
    sys.exit(1)


def find_regseason_source():
    """Locate NBA_RegSeason.xlsx (the Game Score source for top games)."""
    env = os.environ.get("NBA_REGSEASON_PATH")
    if env:
        p = Path(env)
        if p.exists():
            return p
        print(f"NBA_REGSEASON_PATH not found: {p}", file=sys.stderr)
        sys.exit(1)
    for cand in REGSEASON_SOURCE_CANDIDATES:
        if cand.exists():
            return cand
    print("Could not find NBA_RegSeason.xlsx (Game Score source). Place it in",
          "~/OneDrive/Excel Files/ or the repo root, or set NBA_REGSEASON_PATH.",
          file=sys.stderr)
    sys.exit(1)


# -------- Sheet readers --------

def read_totals(wb):
    """One row per franchise. Keyed by canonical Name (col AR / index 43)."""
    ws = wb["Totals"]
    out = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        canonical = safe_str(row[43]) if len(row) > 43 else ""
        if not canonical:
            continue
        out[canonical] = {
            "league_history": safe_str(row[0]),
            "city_history": safe_str(row[1]),
            "team_history": safe_str(row[2]),
            "all_time_w": safe_int(row[3]),
            "all_time_l": safe_int(row[4]),
            "win_pct": safe_float(row[5]),
            "abs_win_pct": safe_float(row[6]),
            "seasons": safe_int(row[7]),
            "five_hundred_seasons": safe_int(row[8]),
            "playoff_appearances": safe_int(row[9]),
            "division_titles": safe_int(row[10]),
            "best_main_div_seasons": safe_int(row[11]),
            "best_record_seasons": safe_int(row[12]),
            "cf_appearances": safe_int(row[13]),
            "champ_appearances": safe_int(row[14]),
            "championships": safe_int(row[15]),
            "playoff_w": safe_int(row[16]),
            "playoff_l": safe_int(row[17]),
            "playoff_win_pct": safe_float(row[18]),
            "series_w": safe_int(row[19]),
            "series_l": safe_int(row[20]),
            "series_win_pct": safe_float(row[21]),
            "last_champ_year": safe_int(row[22]) or None,
            "last_champ_app": safe_int(row[23]) or None,
            "last_cf_app": safe_int(row[24]) or None,
            "last_series_win": safe_int(row[25]) or None,
            "last_best_rec": safe_int(row[26]) or None,
            "last_division_title": safe_int(row[28]) or None,
            "last_playoff_app": safe_int(row[29]) or None,
            "last_500_season": safe_int(row[30]) or None,
            "reg_games": safe_int(row[31]),
            "play_games": safe_int(row[32]),
            "total_games": safe_int(row[33]),
            "home_w": safe_int(row[34]),
            "home_l": safe_int(row[35]),
            "home_win_pct": safe_float(row[36]),
            "road_w": safe_int(row[37]),
            "road_l": safe_int(row[38]),
            "road_win_pct": safe_float(row[39]),
            "neutral_w": safe_int(row[40]),
            "neutral_l": safe_int(row[41]),
            "neutral_win_pct": safe_float(row[42]),
            "is_current": is_truthy_yn(row[44]),
            "is_defunct": is_truthy_yn(row[45]),
        }
    return out


def read_year_by_year(wb):
    """Build season-by-season per canonical franchise.

    Year by Year column map (0-indexed):
      1 League, 2 Year (end year), 3 City, 4 Team, 5 W, 6 L, 7 Win%,
      8 PF, 9 PA, 10 Rebds, 11 Asts, 14 .500+, 15 PlayAp, 16 DivTitle,
      17 BestConf, 18 BestRec, 19 PWins, 20 PLoss, 21 CFApp, 22 ChamApp,
      23 Champs, 24 PlayoffSeed, 25 Division, 26 Place, 28 HomeArena,
      34 Name (canonical), 35 #AllStars, 36 #AllNBA, 37 Conf,
      104 Canonical home arena (DA), ...
    """
    ws = wb["Year by Year"]
    by_team = defaultdict(list)
    latest_meta = {}
    earliest_year = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        canonical = safe_str(row[34]) if len(row) > 34 else ""
        if not canonical:
            continue
        year = safe_int(row[2])
        if year == 0:
            continue
        league = safe_str(row[1])
        city = safe_str(row[3])
        team = safe_str(row[4])
        w = safe_int(row[5])
        l = safe_int(row[6])
        win_pct = safe_float(row[7])
        pf = safe_int(row[8])
        pa = safe_int(row[9])
        playoff_yn = is_truthy_yn(row[15])
        div_title_yn = is_truthy_yn(row[16])
        best_conf_yn = is_truthy_yn(row[17])
        best_rec_yn = is_truthy_yn(row[18])
        p_wins = safe_int(row[19])
        p_loss = safe_int(row[20])
        cf_app_yn = is_truthy_yn(row[21])
        cham_app_yn = is_truthy_yn(row[22])
        champ_yn = is_truthy_yn(row[23])
        playoff_seed = safe_int(row[24]) or None
        division = safe_str(row[25])
        place_raw = row[26]
        place = "" if place_raw is None else str(place_raw).strip()
        home_arena_season = safe_str(row[28]) if len(row) > 28 else ""
        num_all_stars = safe_int(row[35]) if len(row) > 35 else 0
        num_all_nba = safe_int(row[36]) if len(row) > 36 else 0
        conf = safe_str(row[37]) if len(row) > 37 else ""
        # PF/G is col BB = index 53; PA/G is col BC = index 54. Earlier
        # version read 54/55 which surfaced PA/G and Reb/G respectively.
        pf_g = safe_float(row[53]) if len(row) > 53 else 0.0
        pa_g = safe_float(row[54]) if len(row) > 54 else 0.0
        # Canonical home arena lives in col DA = index 104
        home_arena_canonical = safe_str(row[104]) if len(row) > 104 else ""

        by_team[canonical].append({
            "year": year,
            "league": league,
            "city": city,
            "team": team,
            "w": w,
            "l": l,
            "win_pct": round(win_pct, 4) if win_pct else 0.0,
            "pf": pf,
            "pa": pa,
            "pf_g": round(pf_g, 2) if pf_g else 0.0,
            "pa_g": round(pa_g, 2) if pa_g else 0.0,
            "point_diff": pf - pa,
            "playoff": playoff_yn,
            "div_title": div_title_yn,
            "best_conf": best_conf_yn,
            "best_rec_leag": best_rec_yn,
            "p_wins": p_wins,
            "p_loss": p_loss,
            "cf_app": cf_app_yn,
            "champ_app": cham_app_yn,
            "champ": champ_yn,
            "playoff_seed": playoff_seed,
            "division": division,
            "main_div": conf,  # NBA's analog of MLB main_div is conference
            "place": place,
            "home_arena_season": home_arena_season,
            "home_arena_canonical": home_arena_canonical,
            "num_all_stars": num_all_stars,
            "num_all_nba": num_all_nba,
        })
        if canonical not in latest_meta or year > latest_meta[canonical]["year"]:
            latest_meta[canonical] = {
                "year": year,
                "city": city,
                "team": team,
                "league": league,
                "conf": conf,
                "division": division,
                "home_arena_season": home_arena_season,
                "home_arena_canonical": home_arena_canonical,
            }
        earliest_year[canonical] = min(earliest_year.get(canonical, year), year)

    for k in by_team:
        by_team[k].sort(key=lambda s: s["year"])

    return dict(by_team), latest_meta, earliest_year


def read_arenas(wb):
    """Returns { canonical_name: { city, metro, state, first, last, sport, status } }.
    Limits to Sport=='NBA' rows so NCAA neutrals don't contaminate franchise
    arena history."""
    ws = wb["Arenas"]
    out = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        canonical = safe_str(row[0])
        if not canonical:
            continue
        sport = safe_str(row[9]) if len(row) > 9 else ""
        # Sport may be multi-tagged (e.g., 'NHL / NBA', 'NHL/NBA/NCAA').
        # Match by token-set membership so cross-listed arenas survive.
        sport_tokens = {t.strip().upper() for t in sport.replace('/', ',').split(',') if t.strip()}
        if sport_tokens and "NBA" not in sport_tokens:
            continue
        out[canonical] = {
            "canonical": canonical,
            "variants": safe_str(row[1]),
            "primary_teams": safe_str(row[2]),
            "city": safe_str(row[3]),
            "metro": safe_str(row[4]),
            "state": safe_str(row[5]),
            "first_season": safe_int(row[6]) or None,
            "last_season": safe_int(row[7]) or None,
            "total_seasons": safe_int(row[8]) or None,
            "status": safe_str(row[10]) if len(row) > 10 else "",
            "notes": safe_str(row[11]) if len(row) > 11 else "",
        }
    return out


def build_stadium_history(year_by_year, arenas):
    """Group arena-eras by canonical franchise.

    Walks each franchise's season rows in chronological order, groups
    consecutive seasons by canonical home arena name (col DA), and emits a
    row per venue with first_year / last_year. Within each venue row,
    builds an `eras` array using the season-name (col AC) so renames
    (e.g., Vivint Arena → Delta Center) show as multiple eras inside
    one canonical venue.
    """
    out = {}
    for canonical, rows in year_by_year.items():
        if not rows:
            continue
        venues = []
        current = None
        for r in rows:
            yr = r["year"]
            canon_arena = r.get("home_arena_canonical") or r.get("home_arena_season") or ""
            season_name = r.get("home_arena_season") or canon_arena
            if not canon_arena:
                continue
            if current is None or current["canonical"] != canon_arena:
                if current is not None:
                    venues.append(current)
                meta = arenas.get(canon_arena, {})
                current = {
                    "canonical": canon_arena,
                    "city": meta.get("city", ""),
                    "metro": meta.get("metro", ""),
                    "state": meta.get("state", ""),
                    "first_year": yr,
                    "last_year": yr,
                    "eras": [],
                    "_era_stack": [{"era_name": season_name, "first_year": yr, "last_year": yr}],
                }
            else:
                current["last_year"] = yr
                stack = current["_era_stack"]
                if stack and stack[-1]["era_name"] == season_name:
                    stack[-1]["last_year"] = yr
                else:
                    stack.append({"era_name": season_name, "first_year": yr, "last_year": yr})
        if current is not None:
            venues.append(current)
        # Finalize: extract _era_stack into eras; reverse-chronological order
        cleaned = []
        for v in venues:
            v["eras"] = v.pop("_era_stack")
            cleaned.append(v)
        cleaned.sort(key=lambda v: v.get("first_year") or 0, reverse=True)
        out[canonical] = cleaned
    return out


def read_awards(wb):
    """Read Awards sheet, group by canonical franchise name.

    Returns two dicts:
      individual_awards: { canonical: { 'MVP': [ {year, player}, ... ], ... } }
      all_nba_by_team: { canonical: [ {year, player, tier}, ... ] }

    Awards sheet columns (0-indexed):
      1 Year, 2 Player, 3 City, 4 Team, 5 Awards (string), 6 Pos,
      7 Name (canonical), 8 All NBA flag (Y/blank)
    """
    ws = wb["Awards"]
    individual = defaultdict(lambda: defaultdict(list))
    all_nba = defaultdict(list)
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        year = safe_int(row[1])
        if year == 0:
            continue
        player = safe_str(row[2])
        canonical = safe_str(row[7]) if len(row) > 7 else ""
        if not canonical:
            continue
        award_raw = safe_str(row[5])
        is_all_nba = is_truthy_yn(row[8]) if len(row) > 8 else False

        # All-NBA selection
        if is_all_nba or "all nba" in award_raw.lower() or "all-nba" in award_raw.lower():
            tier = normalize_all_nba_tier(award_raw)
            if tier:
                all_nba[canonical].append({
                    "year": year,
                    "player": player,
                    "tier": tier,
                })
            continue

        # Individual award (MVP, DPOY, ROY, COY, MIP, 6MOY, CPOY)
        label = normalize_award_name(award_raw)
        if label:
            individual[canonical][label].append({
                "year": year,
                "player": player,
            })

    # Sort everything descending by year for display
    for canon in individual:
        for label in individual[canon]:
            individual[canon][label].sort(key=lambda r: -r["year"])
    for canon in all_nba:
        all_nba[canon].sort(key=lambda r: (-r["year"], r["tier"], r["player"]))

    return individual, all_nba


def read_all_star_counts(wb):
    """Count player-season All-Star selections per franchise.

    All-Stars sheet columns (0-indexed):
      1 Year, 5 Team (display), 16 Name (canonical join), 17 Final/Current Arena.

    Each row in the sheet is one player-season selection. Total per
    franchise = number of rows whose Name == canonical.
    """
    ws = wb["All-Stars"]
    counts = defaultdict(int)
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        canonical = safe_str(row[16]) if len(row) > 16 else ""
        if not canonical:
            continue
        counts[canonical] += 1
    return dict(counts)


def read_top_games(wb):
    """Read the Regular Season sheet of NBA_RegSeason.xlsx and emit game-level
    rows ranked by Game Score.

    Game data and the documented Game Score live in NBA_RegSeason.xlsx, NOT in
    the Detailed Playoffs sheet of NBA.xlsx. Game Score is column BU (index 72):
    a composite of team strength, competitiveness, stakes, and seeding. Rows
    with a blank Game Score (All-Star games, no-ELO rows, unplayed/scheduled
    games) are skipped, and rows are ranked by Game Score descending.

    Column map (0-indexed) — same layout as the legacy Detailed Playoffs sheet:
      1 Lge, 2 Season (end year), 6 Round, 7 Gm#, 8 Date,
      9 Seed, 10 City, 11 Team, 12 W/L, 13 OppSeed, 14 OtherCity,
      15 OtherTeam, 16 PF, 17 PA, 18 OT, 19 Arena (as-of),
      20 Arena Area (Metro), 21 Arena State,
      46 Cur Name (canonical, OWN team), 47 Opponent (canonical),
      50 Round# (1=Finals, 2=CF, 3=Semis, 4=QF, 4.5=play-in 8-seed,
                 5=play-in), 53 Final/Current Arena Name (canonical)
    """
    ws = wb["Regular Season"]
    # Build the home-team selection: each game has two rows (away then home,
    # by host venue). Group by (year, round, gm#, own city pair) and dedupe.
    games_by_team = defaultdict(list)
    all_games_keyed = {}  # key -> game record (one per game)
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        league = safe_str(row[1])
        year = safe_int(row[2])
        if year == 0:
            continue
        round_name = safe_str(row[6])
        gm_num = safe_int(row[7]) or None
        date_raw = row[8]
        date_str = None
        if hasattr(date_raw, "isoformat"):
            date_str = date_raw.isoformat()[:10]
        elif date_raw:
            try:
                # Excel serial date fallback
                serial = float(date_raw)
                from datetime import datetime, timedelta
                base = datetime(1899, 12, 30)
                date_str = (base + timedelta(days=int(serial))).date().isoformat()
            except (ValueError, TypeError):
                date_str = str(date_raw)
        own_city = safe_str(row[10])
        own_team = safe_str(row[11])
        wl = safe_str(row[12])
        opp_city = safe_str(row[14])
        opp_team = safe_str(row[15])
        pf = safe_int(row[16])
        pa = safe_int(row[17])
        ot_flag = is_truthy_yn(row[18])
        arena_as_of = safe_str(row[19])
        arena_metro = safe_str(row[20])
        arena_state = safe_str(row[21])
        canonical_own = safe_str(row[46]) if len(row) > 46 else ""
        canonical_opp = safe_str(row[47]) if len(row) > 47 else ""
        round_num = row[50] if len(row) > 50 else None
        try:
            round_num_val = float(round_num) if round_num is not None and round_num != "" else None
        except (ValueError, TypeError):
            round_num_val = None
        arena_canonical = safe_str(row[53]) if len(row) > 53 else ""
        gs_raw = row[72] if len(row) > 72 else None
        try:
            game_score = float(gs_raw) if gs_raw not in (None, "") else None
        except (ValueError, TypeError):
            game_score = None

        if not canonical_own:
            continue

        # Skip rows without a Game Score: All-Star games, rows with no ELO, and
        # unplayed/scheduled games (e.g. future Finals) all have a blank BU and
        # must never surface in the top-games lists.
        if game_score is None:
            continue
        if (pf or 0) == 0 and (pa or 0) == 0:
            continue

        # Dedupe key: pair-symmetric (year, round, gm#, canonical pair sorted)
        pair = tuple(sorted([canonical_own, canonical_opp])) if canonical_opp else (canonical_own, "")
        key = (year, round_name, gm_num, pair)

        # Build per-team row (used for per-franchise top games)
        per_team_row = {
            "year": year,
            "date": date_str,
            "round": round_name,
            "round_num": round_num_val,
            "game_num": gm_num,
            "team_city": own_city,
            "team_team": own_team,
            "team_canonical": canonical_own,
            "opp_city": opp_city,
            "opp_team": opp_team,
            "opp_canonical": canonical_opp,
            "result": wl,
            "team_pts": pf,
            "opp_pts": pa,
            "ot": ot_flag,
            "arena_as_of": arena_as_of,
            "arena_canonical": arena_canonical,
            "arena_metro": arena_metro,
            "arena_state": arena_state,
            "league": league,
            "game_score": game_score,
        }
        games_by_team[canonical_own].append(per_team_row)

        # Leaguewide game record (single row per game, winner-perspective)
        if key not in all_games_keyed or wl == "W":
            # If a W row is available, prefer it so the leaguewide row shows the winner first
            if all_games_keyed.get(key, {}).get("result") != "W" or key not in all_games_keyed:
                all_games_keyed[key] = {
                    "year": year,
                    "date": date_str,
                    "round": round_name,
                    "round_num": round_num_val,
                    "game_num": gm_num,
                    "winner_canonical": canonical_own if wl == "W" else canonical_opp,
                    "loser_canonical": canonical_opp if wl == "W" else canonical_own,
                    "winner_city": own_city if wl == "W" else opp_city,
                    "winner_team": own_team if wl == "W" else opp_team,
                    "loser_city": opp_city if wl == "W" else own_city,
                    "loser_team": opp_team if wl == "W" else own_team,
                    "winner_pts": pf if wl == "W" else pa,
                    "loser_pts": pa if wl == "W" else pf,
                    "ot": ot_flag,
                    "arena_canonical": arena_canonical,
                    "arena_metro": arena_metro,
                    "arena_state": arena_state,
                    "league": league,
                    "game_score": game_score,
                }

    all_games = list(all_games_keyed.values())
    return all_games, dict(games_by_team)


def read_playoff_state(wb, year_by_year):
    """Read Detailed Playoffs for the current postseason and emit
    per-franchise playoff status. Sourced from the workbook's hand-
    maintained current state, not ESPN.

    Returns { canonical: { 'state', 'last_round', 'eliminated_year', 'champ' } }
    where state is one of:
      'champion'      -- won the Finals this year
      'lost_finals'   -- lost the Finals this year
      'eliminated_cf' -- lost in Conference Finals
      'eliminated_semis' -- lost in Conference Semifinals
      'eliminated_qf' -- lost in Quarterfinals / First Round
      'eliminated_play_in' -- lost in Play-In
      'active'        -- still in the bracket (no Elim row yet)
      None            -- did not make the playoffs this year (or not yet)
    """
    ws = wb["Detailed Playoffs"]
    # Find the latest year in the sheet
    latest_year = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        yr = safe_int(row[2])
        if yr > latest_year:
            latest_year = yr
    if latest_year == 0:
        return {}, None

    # Walk rows for that year and track each team's furthest round + elim/champ flags
    # round_num: 5 = Play-In, 4.5 = 8-seed decider, 4 = QF/R1, 3 = Semis, 2 = CF, 1 = Finals
    ROUND_LABELS = {
        1: "Finals",
        2: "Conference Finals",
        3: "Conference Semifinals",
        4: "Quarterfinals",
        4.5: "Play-In",
        5: "Play-In",
    }
    state_by_team = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        yr = safe_int(row[2])
        if yr != latest_year:
            continue
        canonical = safe_str(row[46]) if len(row) > 46 else ""
        if not canonical or canonical == "0":
            continue
        try:
            round_num = float(row[50]) if len(row) > 50 and row[50] not in (None, "") else None
        except (ValueError, TypeError):
            round_num = None
        if round_num is None:
            continue
        elim = is_truthy_yn(row[25]) if len(row) > 25 else False
        clinch = is_truthy_yn(row[24]) if len(row) > 24 else False
        champ = is_truthy_yn(row[44]) if len(row) > 44 else False
        result = safe_str(row[12])
        series_l_val = safe_int(row[23]) if len(row) > 23 else 0

        st = state_by_team.setdefault(canonical, {
            "deepest_round": 5,  # higher number = earlier round
            "eliminated_at_round": None,
            "champion": False,
            "lost_finals": False,
            "rounds_played": set(),
            "losses_by_round": {},  # round_num -> max loss count seen
        })
        # "Deepest" means lowest round_num
        if round_num < st["deepest_round"]:
            st["deepest_round"] = round_num
        st["rounds_played"].add(round_num)
        # Track max series_l seen for this team in this round (winning team also has rows
        # at the closing game, but their result is W; we only record series_l on L rows
        # so the count reflects actual series losses by this team).
        if result == "L":
            prev = st["losses_by_round"].get(round_num, 0)
            if series_l_val > prev:
                st["losses_by_round"][round_num] = series_l_val
        if champ and result == "W":
            st["champion"] = True
        # Only a loss row should mark a team as eliminated. The workbook
        # pre-populates Clinch=Y / Elim=Y on shell rows for the clinching
        # game; the WINNING team's row also has those flags (their loser
        # opponent is the one eliminated). Gating on result == "L" prevents
        # an advancing team from being classified as eliminated when their
        # series-clinch game's Elim=Y leaks through.
        if elim and result == "L":
            st["eliminated_at_round"] = round_num
        if round_num == 1 and result == "L" and elim:
            st["lost_finals"] = True

    # Inference pass: if a team has 4+ losses in their deepest round (or 3+ in
    # pre-1984 best-of-5 R1), mark eliminated at that round even when the workbook
    # didn't tag Elim=Y on the loss row. Best-of-7 ends at 4 L's; best-of-5 at 3.
    for canon, st in state_by_team.items():
        if st["eliminated_at_round"] is not None or st["champion"]:
            continue
        deepest = st["deepest_round"]
        max_l_in_deepest = st["losses_by_round"].get(deepest, 0)
        # All modern rounds are best-of-7. Best-of-5 was the R1 format pre-1984; we
        # use 3 as the threshold for round 4 in those eras. Since latest_year is
        # always modern when this matters, the simple >=4 check covers the bulk.
        threshold = 3 if (latest_year < 1984 and deepest == 4) else 4
        if max_l_in_deepest >= threshold:
            st["eliminated_at_round"] = deepest

    out = {}
    for canon, st in state_by_team.items():
        # Determine state
        if st["champion"]:
            state = "champion"
            last_round = "NBA Champion"
        elif st["lost_finals"]:
            state = "lost_finals"
            last_round = "Lost Finals"
        elif st["eliminated_at_round"] is not None:
            r = st["eliminated_at_round"]
            if r == 1:
                state = "lost_finals"; last_round = "Lost Finals"
            elif r == 2:
                state = "eliminated_cf"; last_round = "Eliminated Conference Finals"
            elif r == 3:
                state = "eliminated_semis"; last_round = "Eliminated Conference Semifinals"
            elif r == 4:
                state = "eliminated_qf"; last_round = "Eliminated First Round"
            elif r in (4.5, 5):
                state = "eliminated_play_in"; last_round = "Eliminated Play-In"
            else:
                state = "eliminated_qf"; last_round = "Eliminated"
        else:
            # Still in the bracket — assign by deepest reached
            r = st["deepest_round"]
            if r == 1:
                state = "active_finals"; last_round = "In the Finals"
            elif r == 2:
                state = "active_cf"; last_round = "In the Conference Finals"
            elif r == 3:
                state = "active_semis"; last_round = "In the Conference Semifinals"
            elif r == 4:
                state = "active_qf"; last_round = "In the First Round"
            else:
                state = "active_play_in"; last_round = "In the Play-In"

        out[canon] = {
            "state": state,
            "last_round": last_round,
            "year": latest_year,
        }

    # Postseason is complete if a champion has been crowned OR every team has been
    # eliminated. Page renderers use this flag to drop the chip layer after the Finals
    # so stale "X eliminated in semifinals" badges don't linger through the offseason.
    has_active = any(o.get("state", "").startswith("active_") for o in out.values())
    has_champion = any(o.get("state") == "champion" for o in out.values())
    is_complete = has_champion or (len(out) > 0 and not has_active)
    return out, latest_year, is_complete


def read_team_external_links():
    """Load team-wikidata.tsv. Same schema as the MLB ETL."""
    tsv_path = REPO_ROOT / "team-wikidata.tsv"
    out = {}
    if not tsv_path.exists():
        print("  (note: team-wikidata.tsv not found; Wikipedia/Wikidata links blank)")
        return out
    with tsv_path.open(encoding="utf-8") as fh:
        for line in fh:
            parts = line.rstrip("\r\n").split("\t")
            if len(parts) < 3:
                continue
            team = parts[0].strip()
            qid = parts[1].strip()
            wiki = parts[2].strip()
            if not team:
                continue
            out[team] = {
                "wikipedia_url": wiki or None,
                "wikidata_qid": qid or None,
            }
    return out




# ----- MetroAreas.xlsx Team List integration -----

# Authoritative source for team -> metro/lat/lng mapping. Joins by canonical
# team display name (e.g. "Atlanta Hawks"). The NBA workbook's Arenas sheet
# carries its own metro fields and was normalized to MetroAreas canonical in
# a prior workbook session, but the Team List sheet remains the single source
# of truth across the rankings site (per feedback memory: "Team List is the
# only source of truth for teams"). Using it here keeps the NBA team pages
# aligned with the rest of the surface area.
METROAREAS_PATHS = [
    Path(os.path.expanduser("~/OneDrive/Excel Files/MetroAreas.xlsx")),
    REPO_ROOT / "MetroAreas.xlsx",
    Path("/sessions/nice-epic-hawking/mnt/Excel Files/MetroAreas.xlsx"),
]


def read_metro_team_list():
    """Read MetroAreas.xlsx Team List sheet, filter to NBA rows, return
    dict keyed by team display name: { 'Atlanta Hawks': {metro, city, state, lat, lng}, ... }.

    Falls back to empty dict if the workbook is unreachable so the build
    never blocks on it; the NBA workbook's arena join still fills metros
    in that case.
    """
    out = {}
    for p in METROAREAS_PATHS:
        if not p.exists():
            continue
        try:
            mwb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        except Exception as e:
            print(f"  (warning: could not open {p}: {e})")
            continue
        if "Team List" not in mwb.sheetnames:
            mwb.close()
            continue
        ws = mwb["Team List"]
        # Columns: 0=Sport, 1=League, 2=Team, 5=City, 6=Metro, 7=State,
        # 18=Lat, 19=Long. Shifted from 17/18 on 2026-05-17 when the Gold
        # Standard column was inserted at col L of Team List.
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            league = safe_str(row[1] if len(row) > 1 else "")
            if league.upper() != "NBA":
                continue
            team = safe_str(row[2] if len(row) > 2 else "")
            if not team:
                continue
            out[team] = {
                "city": safe_str(row[5] if len(row) > 5 else ""),
                "metro": safe_str(row[6] if len(row) > 6 else ""),
                "state": safe_str(row[7] if len(row) > 7 else ""),
                "lat": safe_float(row[18] if len(row) > 18 else 0) or None,
                "lng": safe_float(row[19] if len(row) > 19 else 0) or None,
            }
        mwb.close()
        if out:
            print(f"  Team List loaded from {p}: {len(out)} NBA team rows")
            return out
    print("  (warning: MetroAreas.xlsx Team List not found; falling back to NBA workbook arena join)")
    return out


# -------- Builders --------

def build_franchises(totals, latest_meta, year_by_year, earliest_year,
                     external_links, all_star_counts, arenas=None, metroareas_team_list=None):
    arenas = arenas or {}
    metroareas_team_list = metroareas_team_list or {}
    out = []
    for canonical, t in totals.items():
        if not t.get("is_current"):
            continue
        meta = latest_meta.get(canonical, {})
        city = meta.get("city") or ""
        team = canonical
        override = DISPLAY_NAME_OVERRIDES.get(canonical)
        if override:
            city = override[0] or city
            team = override[1]
        display_name = f"{city} {team}".strip()
        # Stripping any disambiguator parens
        display_name = re.sub(r"\s*\(.*\)\s*$", "", display_name).strip()
        founding_year = earliest_year.get(canonical)
        # Pull prior cities from city_history string
        prior_cities = []
        ch = (t.get("city_history") or "").strip()
        if ch and ch != city:
            parts = [p.strip() for p in re.split(r"[/]", ch) if p.strip()]
            for p in parts:
                if p != city and p not in prior_cities:
                    prior_cities.append(p)
        ext = external_links.get(display_name, {})

        out.append({
            "slug": franchise_slug(canonical),
            "name": team,
            "display_name": display_name,
            "canonical": canonical,
            "city": city,
            "team": team,
            "league": "NBA",
            "conf": meta.get("conf", ""),
            "division": meta.get("division", ""),
            # MetroAreas Team List override (authoritative source). Falls back to arena join when not found.
            "metro": (metroareas_team_list.get(display_name, {}) or {}).get("metro") or (arenas.get(meta.get("home_arena_canonical", ""), {}) or {}).get("metro", "") or (arenas.get(meta.get("home_arena_season", ""), {}) or {}).get("metro", ""),
            "metro_slug": metro_slugify((metroareas_team_list.get(display_name, {}) or {}).get("metro") or (arenas.get(meta.get("home_arena_canonical", ""), {}) or {}).get("metro", "") or (arenas.get(meta.get("home_arena_season", ""), {}) or {}).get("metro", "")),
            "state": (metroareas_team_list.get(display_name, {}) or {}).get("state") or (arenas.get(meta.get("home_arena_canonical", ""), {}) or {}).get("state", "") or (arenas.get(meta.get("home_arena_season", ""), {}) or {}).get("state", ""),
            "lat": (metroareas_team_list.get(display_name, {}) or {}).get("lat"),
            "lng": (metroareas_team_list.get(display_name, {}) or {}).get("lng"),
            "arena": meta.get("home_arena_canonical", "") or meta.get("home_arena_season", ""),
            "arena_season_name": meta.get("home_arena_season", ""),
            "founding_year": founding_year,
            "prior_cities": prior_cities,
            "wikipedia_url": ext.get("wikipedia_url"),
            "wikidata_qid": ext.get("wikidata_qid"),
            "championships": t.get("championships", 0),
            "championship_appearances": t.get("champ_appearances", 0),
            "cf_appearances": t.get("cf_appearances", 0),
            "division_titles": t.get("division_titles", 0),
            "playoff_appearances": t.get("playoff_appearances", 0),
            "playoff_w": t.get("playoff_w", 0),
            "playoff_l": t.get("playoff_l", 0),
            "all_time_w": t.get("all_time_w", 0),
            "all_time_l": t.get("all_time_l", 0),
            "win_pct": t.get("win_pct", 0.0),
            "seasons": t.get("seasons", 0),
            "five_hundred_seasons": t.get("five_hundred_seasons", 0),
            "best_rec_seasons": t.get("best_record_seasons", 0),
            "last_championship_year": t.get("last_champ_year"),
            "last_champ_app": t.get("last_champ_app"),
            "last_cf_app": t.get("last_cf_app"),
            "last_division_title": t.get("last_division_title"),
            "last_playoff_app": t.get("last_playoff_app"),
            "all_star_count": all_star_counts.get(canonical, 0),
            "league_history": t.get("league_history", ""),
        })
    out.sort(key=lambda f: f["display_name"])
    return out


def build_championships(year_by_year):
    """Per-franchise list of championships with era classification."""
    out = {}
    for canonical, rows in year_by_year.items():
        ch_rows = [r for r in rows if r.get("champ")]
        if not ch_rows:
            continue
        out[canonical] = [
            {
                "year": r["year"],
                "era": championship_era(r.get("league")),
                "league": r.get("league"),
                "city": r.get("city"),
                "team": r.get("team"),
            }
            for r in sorted(ch_rows, key=lambda r: r["year"])
        ]
    return out


def build_championship_appearances(year_by_year):
    """Per-franchise list of Finals appearances (won or lost)."""
    out = {}
    for canonical, rows in year_by_year.items():
        ap_rows = [r for r in rows if r.get("champ_app")]
        if not ap_rows:
            continue
        out[canonical] = [
            {
                "year": r["year"],
                "era": championship_era(r.get("league")),
                "won": r.get("champ", False),
                "city": r.get("city"),
                "team": r.get("team"),
            }
            for r in sorted(ap_rows, key=lambda r: r["year"])
        ]
    return out


def build_seasons_by_team(franchises, year_by_year):
    out = {}
    for f in franchises:
        canon = f["canonical"]
        out[f["slug"]] = year_by_year.get(canon, [])
    return out


def build_historical(totals, year_by_year):
    """Defunct franchises (including ABA-only) for /teams/nba/historical."""
    out = []
    for canonical, t in totals.items():
        if not t.get("is_defunct"):
            continue
        rows = year_by_year.get(canonical, [])
        # Pull the city/team history from the Totals row (combined "X/Y" string)
        first_year = min((r["year"] for r in rows), default=None) if rows else None
        last_year = max((r["year"] for r in rows), default=None) if rows else None
        # Determine ABA-only flag
        leagues_in_history = {r.get("league", "").upper() for r in rows}
        aba_only = leagues_in_history == {"ABA"} if leagues_in_history else False
        # Build human display name from city + team history columns
        ch = (t.get("city_history") or "").strip()
        th = (t.get("team_history") or "").strip()
        # Strip the "(1968-76)" style year-range suffix already in team_history
        th_clean = re.sub(r"\s*\(\d{4}-\d{2,4}\)\s*$", "", th).strip()
        if ch and th_clean:
            display_name = f"{ch} {th_clean}"
        elif ch:
            display_name = f"{ch} {canonical}"
        else:
            display_name = canonical
        out.append({
            "slug": franchise_slug(canonical),
            "canonical": canonical,
            "display_name": display_name,
            "city_history": t.get("city_history", ""),
            "team_history": t.get("team_history", ""),
            "league_history": t.get("league_history", ""),
            "seasons": t.get("seasons", 0),
            "first_year": first_year,
            "last_year": last_year,
            "championships": t.get("championships", 0),
            "championship_appearances": t.get("champ_appearances", 0),
            "cf_appearances": t.get("cf_appearances", 0),
            "playoff_appearances": t.get("playoff_appearances", 0),
            "all_time_w": t.get("all_time_w", 0),
            "all_time_l": t.get("all_time_l", 0),
            "win_pct": t.get("win_pct", 0.0),
            "aba_only": aba_only,
            "leagues": sorted(leagues_in_history),
        })
    out.sort(key=lambda f: (-(f.get("first_year") or 0), f["display_name"]))
    return out


def build_historical_seasons(historical, year_by_year):
    out = {}
    for h in historical:
        out[h["slug"]] = year_by_year.get(h["canonical"], [])
    return out


# Rank top games per team, per decade, and leaguewide by Game Score (col BU)
# descending, read from NBA_RegSeason.xlsx. Ties break by round, then recency.
def _game_sort_key(g):
    # Primary: Game Score (BU) descending. Rows without a score sink to the
    # bottom; ties break by round importance, then recency.
    gs = g.get("game_score")
    gs_key = -gs if isinstance(gs, (int, float)) else float("inf")
    rn = g.get("round_num")
    if rn is None or rn <= 0:
        rn = 99
    return (gs_key, rn, -g.get("year", 0), -(int(g.get("date", "0").replace("-", "")) if g.get("date") else 0))


def build_top_games_by_team(games_by_team, franchises, top_n=12):
    out = {}
    franchise_canons = {f["canonical"]: f["slug"] for f in franchises}
    for canon, slug in franchise_canons.items():
        games = games_by_team.get(canon, [])
        # Only winning team's perspective for the franchise's top-games list,
        # to match MLB convention. Still include losses if the franchise's
        # game count would otherwise be empty.
        sorted_games = sorted(games, key=_game_sort_key)
        out[slug] = sorted_games[:top_n]
    return out


def build_top_games_all_time(all_games, top_n=50):
    return sorted(all_games, key=_game_sort_key)[:top_n]


def build_top_games_by_decade(all_games, top_n_per_decade=10):
    by_dec = defaultdict(list)
    for g in all_games:
        yr = g.get("year") or 0
        if not yr:
            continue
        dec = (yr // 10) * 10
        by_dec[dec].append(g)
    out = {}
    for dec, games in sorted(by_dec.items()):
        out[str(dec)] = sorted(games, key=_game_sort_key)[:top_n_per_decade]
    return out


# -------- Main --------

def main():
    cli_path = sys.argv[1] if len(sys.argv) > 1 else None
    src = find_source(cli_path)
    print(f"Reading: {src}")
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    print(f"Sheets: {wb.sheetnames}")

    print("Reading Totals...")
    totals = read_totals(wb)
    print(f"  {len(totals)} franchise rows (active + defunct)")

    print("Reading Year by Year...")
    yby, latest_meta, earliest_year = read_year_by_year(wb)
    print(f"  {len(yby)} canonical names, {sum(len(v) for v in yby.values())} team-seasons")

    print("Reading Arenas...")
    arenas = read_arenas(wb)
    print(f"  {len(arenas)} NBA arenas (NCAA neutrals filtered out)")

    print("Building stadium history...")
    stadiums = build_stadium_history(yby, arenas)
    print(f"  {len(stadiums)} franchises with arena history")

    print("Reading Awards (individual + All-NBA)...")
    individual_awards, all_nba_by_team = read_awards(wb)
    print(f"  {len(individual_awards)} franchises with curated individual awards")
    print(f"  {len(all_nba_by_team)} franchises with All-NBA selections")

    print("Reading All-Star counts...")
    all_star_counts = read_all_star_counts(wb)
    print(f"  {len(all_star_counts)} franchises with All-Star selections, "
          f"{sum(all_star_counts.values())} total player-seasons")

    print("Reading top games from NBA_RegSeason.xlsx (Game Score col BU)...")
    rs_src = find_regseason_source()
    rs_wb = openpyxl.load_workbook(rs_src, read_only=True, data_only=True)
    print(f"  Game-score source: {rs_src}")
    all_games, games_by_team = read_top_games(rs_wb)
    print(f"  {len(all_games)} unique scored games, {len(games_by_team)} franchises with game data")

    print("Reading playoff state (current postseason)...")
    playoff_state, latest_playoff_year, postseason_complete = read_playoff_state(wb, yby)
    print(f"  Year {latest_playoff_year}: {len(playoff_state)} franchises with state")

    print("Reading MetroAreas.xlsx Team List (authoritative metro + lat/lng)...")
    metro_team_list = read_metro_team_list()

    print("Reading TeamQIDs (Wikipedia/Wikidata cross-links)...")
    external_links = read_team_external_links()
    print(f"  {len(external_links)} teams with external links")

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    franchises = build_franchises(totals, latest_meta, yby, earliest_year,
                                  external_links, all_star_counts, arenas=arenas,
                                  metroareas_team_list=metro_team_list)
    print(f"Built franchises: {len(franchises)}")
    (OUT_DIR / "franchises.json").write_text(json.dumps(franchises, indent=2, ensure_ascii=False), encoding="utf-8")

    champs = build_championships(yby)
    (OUT_DIR / "championships.json").write_text(json.dumps(champs, indent=2, ensure_ascii=False), encoding="utf-8")

    champ_apps = build_championship_appearances(yby)
    (OUT_DIR / "championship-appearances.json").write_text(json.dumps(champ_apps, indent=2, ensure_ascii=False), encoding="utf-8")

    (OUT_DIR / "stadium-history.json").write_text(json.dumps(stadiums, indent=2, ensure_ascii=False), encoding="utf-8")

    awards_plain = {team: dict(by_award) for team, by_award in individual_awards.items()}
    (OUT_DIR / "award-winners.json").write_text(json.dumps(awards_plain, indent=2, ensure_ascii=False), encoding="utf-8")

    (OUT_DIR / "all-nba-selections.json").write_text(json.dumps(all_nba_by_team, indent=2, ensure_ascii=False), encoding="utf-8")

    (OUT_DIR / "all-star-counts.json").write_text(json.dumps(all_star_counts, indent=2, ensure_ascii=False), encoding="utf-8")

    historical = build_historical(totals, yby)
    print(f"Built historical: {len(historical)}")
    (OUT_DIR / "historical.json").write_text(json.dumps(historical, indent=2, ensure_ascii=False), encoding="utf-8")

    historical_seasons = build_historical_seasons(historical, yby)
    print(f"Built historical-seasons: {len(historical_seasons)} franchises, "
          f"{sum(len(v) for v in historical_seasons.values())} season-rows")
    (OUT_DIR / "historical-seasons.json").write_text(json.dumps(historical_seasons, indent=2, ensure_ascii=False), encoding="utf-8")

    seasons_out = build_seasons_by_team(franchises, yby)
    (OUT_DIR / "seasons-by-team.json").write_text(json.dumps(seasons_out, indent=2, ensure_ascii=False), encoding="utf-8")

    top_by_team = build_top_games_by_team(games_by_team, franchises, top_n=12)
    (OUT_DIR / "top-games-by-team.json").write_text(json.dumps(top_by_team, indent=2, ensure_ascii=False), encoding="utf-8")

    top_all_time = build_top_games_all_time(all_games, top_n=50)
    (OUT_DIR / "top-games-all-time.json").write_text(json.dumps(top_all_time, indent=2, ensure_ascii=False), encoding="utf-8")

    top_by_decade = build_top_games_by_decade(all_games, top_n_per_decade=10)
    (OUT_DIR / "top-games-by-decade.json").write_text(json.dumps(top_by_decade, indent=2, ensure_ascii=False), encoding="utf-8")

    (OUT_DIR / "playoff-state.json").write_text(
        json.dumps({"year": latest_playoff_year, "is_postseason_complete": postseason_complete, "by_franchise": playoff_state}, indent=2, ensure_ascii=False),
    encoding="utf-8",
    )

    print("\nWrote:")
    for f in sorted(OUT_DIR.glob("*.json")):
        try:
            shown = f.relative_to(REPO_ROOT)
        except ValueError:
            shown = f  # OUT_DIR overridden outside the repo (e.g. QA temp dir)
        print(f"  {shown}  ({f.stat().st_size:,} bytes)")


if __name__ == "__main__":
    main()

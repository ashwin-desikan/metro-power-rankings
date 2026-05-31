"""
Patch public/data/football/index.json with city + metro from MetroAreas.xlsx FootballClub_Data.
Run from project root: python scripts/patch-football-cities.py
"""
import json, sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    import subprocess; subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--quiet", "--break-system-packages"])
    import openpyxl

ROOT  = Path(__file__).resolve().parent.parent
INDEX = ROOT / "public/data/football/index.json"
XLSX  = ROOT / "MetroAreas.xlsx"

if not XLSX.exists():
    sys.exit(f"FAIL: {XLSX} not found")

print(f"Loading {XLSX.name}...")
wb = openpyxl.load_workbook(str(XLSX), read_only=True, data_only=True)
if "FootballClub_Data" not in wb.sheetnames:
    sys.exit("FAIL: FootballClub_Data sheet not found")

# Col A=Team, B=City, C=Metro Area
lookup = {}  # {name_lower: (city, metro)}
for row in wb["FootballClub_Data"].iter_rows(min_row=2, values_only=True):
    if not row or not row[0]: continue
    name = str(row[0]).strip()
    city  = str(row[1]).strip() if row[1] else None
    metro = str(row[2]).strip() if row[2] else None
    if city or metro:
        lookup[name.lower()] = (city, metro)
wb.close()
print(f"  {len(lookup)} clubs in FootballClub_Data")

print(f"Patching {INDEX.name}...")
with open(INDEX, encoding="utf-8") as f:
    data = json.load(f)

patched = 0
for club in data["clubs"]:
    key = club["cur_name"].lower()
    if key not in lookup: continue
    city, metro = lookup[key]
    changed = False
    if city and club.get("city") != city:
        club["city"] = city
        changed = True
    if metro and club.get("metro") != metro:
        club["metro"] = metro
        changed = True
    if changed:
        patched += 1
        print(f"  {club['cur_name']}: city={city}, metro={metro}")

with open(INDEX, "w", encoding="utf-8") as f:
    json.dump(data, f, ensure_ascii=False, separators=(",", ":"))
print(f"Done — {patched} clubs updated.")

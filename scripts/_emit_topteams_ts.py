#!/usr/bin/env python3
"""Regenerate lib/topTeams.ts from the "Top Sports Teams" sheet in MetroAreas.xlsx.

Reads the sheet, joins each row to the parent metro's rank from
public/data/metros.json, and rewrites the TOP_TEAMS array in lib/topTeams.ts
in place. Preserves the file header comment and helper functions
(normalizeTopTeamMetroName, METRO_NAME_ALIASES, getTopTeamByMetroName).

Usage:
  python scripts/_emit_topteams_ts.py [path/to/MetroAreas.xlsx]

If no path is given, looks for MetroAreas.xlsx in the project root, then
falls back to a sibling MetroAreas.xlsx next to the project root.

Run scripts/extract.py first if metros.json is stale, since rank assignment
relies on it.
"""

import json
import os
import re
import sys
import unicodedata
from pathlib import Path

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call(
        [sys.executable, "-m", "pip", "install", "openpyxl",
         "--quiet", "--break-system-packages"]
    )
    import openpyxl


# Mirrors METRO_NAME_ALIASES in lib/topTeams.ts. Keep in sync if you add new
# aliases there. Maps normalized SHEET name -> normalized DATASET name.
SHEET_TO_DATASET = {
    "minho-braga": "minho",
}


def norm(s: str) -> str:
    """Normalize a metro name to a stable token. Mirrors
    normalizeTopTeamMetroName in lib/topTeams.ts."""
    s = s.lower().strip()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r"[.'’]", "", s)
    s = re.sub(r"[\s_/]+", "-", s)
    s = re.sub(r"[^a-z0-9-]", "", s)
    s = re.sub(r"-+", "-", s)
    return s.strip("-")


def js_escape(s: str) -> str:
    """Escape a Python string for embedding in a JS double-quoted literal."""
    s = s.replace("\\", "\\\\").replace('"', '\\"')
    s = s.replace("\n", "\\n").replace("\r", "\\r").replace("\t", "\\t")
    return s


def find_xlsx(explicit: str | None) -> Path:
    here = Path(__file__).resolve().parent
    project = here.parent
    if explicit:
        return Path(explicit)
    primary = project / "MetroAreas.xlsx"
    if primary.exists():
        return primary
    legacy = project.parent / "MetroAreas.xlsx"
    if legacy.exists():
        return legacy
    raise FileNotFoundError(
        f"MetroAreas.xlsx not found at {primary} or {legacy}"
    )


def load_sheet(xlsx_path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    if "Top Sports Teams" not in wb.sheetnames:
        raise RuntimeError(
            f"'Top Sports Teams' sheet missing from {xlsx_path.name}"
        )
    ws = wb["Top Sports Teams"]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # header
        metro = "" if row[0] is None else str(row[0]).strip()
        if not metro:
            continue
        rows.append({
            "metro": metro,
            "sport": "" if row[1] is None else str(row[1]).strip(),
            "team": "" if row[2] is None else str(row[2]).strip(),
            "rationale": "" if row[3] is None else str(row[3]).strip(),
        })
    wb.close()
    return rows


def main():
    project = Path(__file__).resolve().parent.parent

    xlsx_path = find_xlsx(sys.argv[1] if len(sys.argv) > 1 else None)
    print(f"Reading {xlsx_path}...")

    sheet_rows = load_sheet(xlsx_path)
    print(f"  {len(sheet_rows)} sheet rows")

    metros_json = project / "public" / "data" / "metros.json"
    if not metros_json.exists():
        print(f"ERROR: {metros_json} missing. Run scripts/extract.py first.")
        sys.exit(1)
    metros = json.loads(metros_json.read_text(encoding="utf-8"))
    metro_by = {norm(m["name"]): m for m in metros}

    ranked = []
    orphans = []
    for row in sheet_rows:
        n = norm(row["metro"])
        target = SHEET_TO_DATASET.get(n, n)
        m = metro_by.get(target)
        if not m:
            orphans.append(row)
            continue
        ranked.append((m["rank"], row))

    if orphans:
        print(f"WARNING: {len(orphans)} sheet rows have no metros.json match:")
        for o in orphans:
            print(f"  - {o['metro']} (sport={o['sport']!r})")

    ranked.sort(key=lambda x: x[0])
    print(f"  Resolved {len(ranked)} entries against metros.json")

    # Build TOP_TEAMS array body
    entry_blocks = []
    for rank, row in ranked:
        entry_blocks.append(
            "  {\n"
            f"    rank: {rank},\n"
            f'    metro: "{js_escape(row["metro"])}",\n'
            f'    sport: "{js_escape(row["sport"])}",\n'
            f'    team: "{js_escape(row["team"])}",\n'
            f'    rationale: "{js_escape(row["rationale"])}",\n'
            "  }"
        )
    array_body = ",\n".join(entry_blocks)

    target_ts = project / "lib" / "topTeams.ts"
    src = target_ts.read_text(encoding="utf-8")
    start_marker = "export const TOP_TEAMS: TopTeamPick[] = ["
    end_marker = "\n];"
    i0 = src.find(start_marker)
    i1 = src.find(end_marker, i0)
    if i0 == -1 or i1 == -1:
        print(f"ERROR: cannot locate TOP_TEAMS array in {target_ts}")
        sys.exit(1)

    new_src = (
        src[:i0]
        + start_marker + "\n"
        + array_body + ",\n"
        + "];"
        + src[i1 + len(end_marker):]
    )

    target_ts.write_text(new_src, encoding="utf-8")
    print(f"Wrote {target_ts} ({len(new_src):,} bytes)")
    print("Done. Run `npx tsc --noEmit` to verify.")


if __name__ == "__main__":
    main()

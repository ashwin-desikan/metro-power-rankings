#!/usr/bin/env python3
"""Build F1_Comprehensive.xlsx from the canonical CSVs in ./data.
Usage: python3 f1_build.py
Reads ./data/*.csv -> writes ./F1_Comprehensive.xlsx (multi-sheet, formatted)."""
import pandas as pd
from f1_source import read_df, os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE=os.path.dirname(os.path.abspath(__file__)); D=os.path.join(BASE,'data'); OUT=os.path.join(BASE,'F1_Comprehensive.xlsx')
def L(f): return read_df(f, csv_dir=D)

rt=L('race_tracks.csv'); meta=L('race_meta.csv'); res=L('results.csv'); spr=L('sprint_results.csv')
poles=L('poles.csv'); ds=L('driver_standings.csv'); cs=L('constructor_standings.csv')
drivers=L('drivers.csv'); constructors=L('constructors.csv'); circuits=L('circuits.csv')

# ---- Races (spine) + Circuit ID ----
rt['Race ID']=rt['Season'].astype(int).astype(str)+'_'+rt['Race'].astype(int).astype(str)
meta['Race ID']=meta['season'].astype(int).astype(str)+'_'+meta['round'].astype(int).astype(str)
rt=rt.merge(meta[['Race ID','circuit_id']],on='Race ID',how='left').rename(columns={'circuit_id':'Circuit ID'})
front=['Race ID','Season','Race','Grand Prix','Date','Circuit','Circuit ID','City','Metro Area','Country']
front=[c for c in front if c in rt.columns]
races_sheet=rt[front+[c for c in rt.columns if c not in front]]

def rid(df): return df['season'].astype(int).astype(str)+'_'+df['round'].astype(int).astype(str)
res['Race ID']=rid(res); spr['Race ID']=rid(spr)
ren={'season':'Season','round':'Round','race_name':'Grand Prix','driver':'Driver','constructor':'Constructor',
     'grid':'Grid','position':'Pos','finish_order':'Finish Order','points':'Points','laps':'Laps',
     'time_gap':'Time/Gap','status':'Status','fastest_lap_time':'Fastest Lap','fastest_lap_speed':'FL Speed (kph)'}
rcols=['Race ID','Season','Round','Grand Prix','Pos','Finish Order','Driver','Constructor','Grid','Laps','Time/Gap','Points','Status','Fastest Lap','FL Speed (kph)']
results_sheet=res.rename(columns=ren)[ [c for c in rcols if c in res.rename(columns=ren).columns] ].sort_values(['Season','Round','Finish Order'])
scols=[c for c in rcols if c!='FL Speed (kph)']
sprint_sheet=spr.rename(columns=ren)[ [c for c in scols if c in spr.rename(columns=ren).columns] ].sort_values(['Season','Round','Finish Order'])

# ---- standings + champions ----
dname=drivers.set_index('driver_id')['driver'].to_dict() if 'driver' in drivers.columns else {}
dnat=drivers.set_index('driver_id')['nationality'].to_dict()
cname=constructors.set_index('constructor_id')['constructor'].to_dict()
def std(df,idc,nm,nt,idl,nl):
    d=df.copy(); d['position']=pd.to_numeric(d['position'],errors='coerce')
    d[nl]=d[idc].map(nm); d['Nationality']=d[idc].map(nt) if nt else None
    d=d.rename(columns={'season':'Season','round':'Final Round','position':'Position','points':'Points','wins':'Wins',idc:idl})
    keep=['Season','Position',idl,nl,'Nationality','Points','Wins','Final Round']
    return d[[c for c in keep if c in d.columns]].sort_values(['Season','Position'])
ds_sheet=std(ds,'driver_id',dname,dnat,'Driver ID','Driver')
cs_sheet=std(cs,'constructor_id',cname,None,'Constructor ID','Constructor')
dch=ds_sheet[ds_sheet['Position']==1][['Season','Driver','Points','Wins']].rename(columns={'Driver':'Drivers Champion','Points':'Driver Points','Wins':'Driver Wins'})
cch=cs_sheet[cs_sheet['Position']==1][['Season','Constructor','Points','Wins']].rename(columns={'Constructor':'Constructors Champion','Points':'Constructor Points','Wins':'Constructor Wins'})
champ_sheet=dch.merge(cch,on='Season',how='outer').sort_values('Season')
champ_sheet=champ_sheet[champ_sheet['Season']<datetime.now().year]  # exclude in-progress season

ry=sorted(results_sheet['Season'].unique()); sy=sorted(sprint_sheet['Season'].unique())
built=pd.Timestamp.today().strftime('%Y-%m-%d')
notes=[['F1 Comprehensive Workbook',''],['Built (auto)',built],
 ['Spine key','Race ID = Season_Round. Links Races, Race Results, Sprint Results.'],
 ['Refresh','Regenerated from ./data CSVs by f1_build.py. New rounds added by f1_update.py from the Jolpica-F1 API.'],['',''],
 ['SHEET','CONTENTS / COVERAGE'],
 ['Races',f'Editorial spine, one row per GP, 1950-{int(races_sheet.Season.max())}. Metro Area, circuit detail, Pole & Fastest Lap (2015 on).'],
 ['Circuits','Circuit reference with coordinates and Wikipedia.'],['Drivers','Driver reference (code, number, dob, nationality).'],
 ['Constructors','Constructor reference.'],['Season Champions','Drivers & constructors champions per season.'],
 ['Driver Standings','Season-end standings (current-year row updates live).'],
 ['Constructor Standings','Season-end standings. 1954 absent; title from 1958.'],
 ['Race Results',f'Full finishing order, every race {ry[0]}-{ry[-1]}.'],
 ['Sprint Results',f'Sprint results {sy[0]}-{sy[-1]}.'],['',''],
 ['SOURCE','F1.xlsx editorial spine + Ergast dump (1950-2024) + Jolpica-F1 API (2025+). See README.md.']]
notes_df=pd.DataFrame(notes)

sheets=[('Read Me',notes_df,False),('Races',races_sheet,True),('Circuits',circuits.rename(columns=str.title),True),
 ('Drivers',drivers.rename(columns=str.title),True),('Constructors',constructors.rename(columns=str.title),True),
 ('Season Champions',champ_sheet,True),('Driver Standings',ds_sheet,True),('Constructor Standings',cs_sheet,True),
 ('Race Results',results_sheet,True),('Sprint Results',sprint_sheet,True)]
with pd.ExcelWriter(OUT,engine='openpyxl') as xw:
    for n,df,h in sheets: df.to_excel(xw,sheet_name=n,index=False,header=h)

# ---- formatting ----
wb=load_workbook(OUT); HEAD=PatternFill('solid',fgColor='1F2A44'); HF=Font(name='Arial',bold=True,color='FFFFFF',size=11); BF=Font(name='Arial',size=10)
for ws in wb.worksheets:
    ws.sheet_view.showGridLines=False
    if ws.title=='Read Me':
        ws.column_dimensions['A'].width=24; ws.column_dimensions['B'].width=100
        for row in ws.iter_rows():
            for c in row: c.font=BF; c.alignment=Alignment(vertical='top',wrap_text=True)
        ws['A1'].font=Font(name='Arial',bold=True,size=15,color='1F2A44'); continue
    for c in ws[1]: c.fill=HEAD; c.font=HF; c.alignment=Alignment(horizontal='left',vertical='center')
    ws.freeze_panes='A2'; ws.auto_filter.ref=ws.dimensions; ws.row_dimensions[1].height=20
    for col in ws.iter_cols(min_row=2):
        for c in col: c.font=BF
    for i,col in enumerate(ws.iter_cols(),1):
        ml=max((len(str(c.value)) for c in col if c.value is not None),default=10)
        ws.column_dimensions[get_column_letter(i)].width=min(max(ml+2,9),46)
wb.save(OUT)
print(f'Built {OUT}')
print(f'  Race Results {len(results_sheet)} rows ({ry[0]}-{ry[-1]}), Sprint {len(sprint_sheet)} rows, Races spine {len(races_sheet)} rows')

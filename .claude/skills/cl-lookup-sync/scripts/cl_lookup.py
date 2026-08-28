#!/usr/bin/env python3
"""cl_lookup.py - workbook-to-Supabase diff engine for public.football_lookup.

Three subcommands, run in order. Each one prints the SQL for the next step, so a
session only ever moves small summaries between the workbook and the database
instead of ten thousand rows.

  extract    read the Lookup sheet -> lookup.json, print the per-country query
  countries  compare per-country hashes -> which countries differ
  rows       compare per-row hashes inside those countries -> the exact deltas,
             classified as ADD / CHANGE / REMOVE / HELD, with the SQL to apply

Why hashes: football_lookup carries about 10,000 rows. Pulling them all into a
session to diff them is slow and wasteful, and it is the reason nobody checks
this table. Hashing lets Postgres and Python each summarise their own copy, and
only the disagreements ever travel.

The canonical form below MUST stay byte-identical to the SQL in
references/sql.md, or every row looks changed. If you edit one, edit both.
"""

import argparse
import collections
import hashlib
import json
import os
import sys

FIELDS = [
    "cur_name", "team", "lookup_name", "uefa_name", "uefa_name_2", "uefa_name_3",
    "efs_name", "api_name", "api_name_2", "country", "city", "metro_area",
    "county", "continent", "league", "level", "lat", "long",
]

HERE = os.path.dirname(os.path.abspath(__file__))
DEFAULT_STATE = os.path.expanduser("~/cl-lookup-sync")
DEFAULT_WORKBOOK = os.path.expanduser("~/mnt/Excel Files/Champions League-201516.xlsx")
PROTECTED_FILE = os.path.join(HERE, "..", "references", "protected_rows.json")


# ---------------------------------------------------------------- canonical form

def clean(v):
    """Match sync_lookup.py exactly: blank, #N/A and literal 0 all mean 'no value'."""
    if v in (None, ""):
        return None
    s = str(v).strip()
    return None if s in ("#N/A", "0", "") else s


def canon(row):
    out = []
    for f in FIELDS:
        v = row.get(f)
        if v is None:
            out.append("")
        elif f == "level":
            out.append(str(int(v)))
        elif f in ("lat", "long"):
            out.append("%.6f" % float(v))
        else:
            out.append(str(v))
    return "\x1f".join(out)


def rowhash(row):
    return hashlib.md5(canon(row).encode("utf-8")).hexdigest()


def group_hash(hashes):
    return hashlib.md5("".join(sorted(hashes)).encode("utf-8")).hexdigest()


# ---------------------------------------------------------------- state helpers

def state_path(args, name):
    os.makedirs(args.state, exist_ok=True)
    return os.path.join(args.state, name)


def load_rows(args):
    p = state_path(args, "lookup.json")
    if not os.path.exists(p):
        sys.exit("No lookup.json yet. Run `cl_lookup.py extract` first.")
    return json.load(open(p, encoding="utf-8"))["rows"]


def load_protected():
    p = os.path.abspath(PROTECTED_FILE)
    if not os.path.exists(p):
        return []
    return json.load(open(p, encoding="utf-8")).get("rows", [])


def sql_literal(v):
    if v is None:
        return "null"
    if isinstance(v, (int, float)):
        return repr(v)
    return "'" + str(v).replace("'", "''") + "'"


# ---------------------------------------------------------------- extract

def cmd_extract(args):
    try:
        import openpyxl
    except ImportError:
        sys.exit("openpyxl is not installed here. `pip install openpyxl`, or run this step on a host that has it.")

    cols = {
        "cur_name": "Cur. Name", "team": "Team", "lookup_name": "Lookup",
        "uefa_name": "UEFA Name", "uefa_name_2": "UEFA Name 2", "uefa_name_3": "UEFA Name 3",
        "efs_name": "EFS Name", "api_name": "API Name", "api_name_2": "API Name 2",
        "country": "Country", "city": "City", "metro_area": "Metro Area",
        "county": "County", "continent": "Continent", "league": "League",
        "level": "Level", "lat": "Lat", "long": "Long",
    }

    wb = openpyxl.load_workbook(args.workbook, read_only=True, data_only=True)
    if "Lookup" not in wb.sheetnames:
        sys.exit("This workbook has no 'Lookup' sheet. Wrong file?")
    ws = wb["Lookup"]
    it = ws.iter_rows(values_only=True)
    hdr = [str(h).strip() if h is not None else "" for h in next(it)]
    idx = {k: (hdr.index(v) if v in hdr else None) for k, v in cols.items()}
    missing = [v for k, v in cols.items() if idx[k] is None]
    if missing:
        # A renamed or deleted column would otherwise sync as a silent mass blanking.
        sys.exit("Lookup sheet is missing expected columns: %s. Stop and check the sheet." % missing)

    rows = []
    for r in it:
        def g(i):
            return r[i] if (i is not None and i < len(r)) else None
        rec = {}
        for k, i in idx.items():
            v = clean(g(i))
            if k == "level":
                try:
                    v = int(float(v)) if v is not None else None
                except (TypeError, ValueError):
                    v = None
            elif k in ("lat", "long"):
                try:
                    v = float(v) if v is not None else None
                except (TypeError, ValueError):
                    v = None
            rec[k] = v
        if not (rec.get("cur_name") or rec.get("team") or rec.get("lookup_name")):
            continue
        rows.append(rec)

    json.dump({"workbook": args.workbook, "rows": rows},
              open(state_path(args, "lookup.json"), "w", encoding="utf-8"), ensure_ascii=False)

    per = collections.defaultdict(list)
    for r in rows:
        per[r.get("country") or "~none~"].append(rowhash(r))
    packed = "|".join("%s~%d~%s" % (c, len(h), group_hash(h)[:10]) for c, h in sorted(per.items()))
    open(state_path(args, "workbook_countries.txt"), "w", encoding="utf-8").write(packed)

    print("workbook rows      : %d" % len(rows))
    print("countries          : %d" % len(per))
    print("state written to   : %s" % args.state)
    print()
    print("NEXT: run the country-hash query in references/sql.md (step 2) against Supabase,")
    print("save its single `packed` value to %s," % state_path(args, "supabase_countries.txt"))
    print("then run: cl_lookup.py countries")


# ---------------------------------------------------------------- countries

def parse_packed(text):
    out = {}
    for part in text.strip().split("|"):
        if not part:
            continue
        c, n, h = part.rsplit("~", 2)
        out[c] = (int(n), h)
    return out


def cmd_countries(args):
    wb_file = state_path(args, "workbook_countries.txt")
    sb_file = state_path(args, "supabase_countries.txt")
    if not os.path.exists(sb_file):
        sys.exit("Save the Supabase `packed` value to %s first." % sb_file)
    wbc = parse_packed(open(wb_file, encoding="utf-8").read())
    sbc = parse_packed(open(sb_file, encoding="utf-8").read())

    diff = [c for c in sorted(set(wbc) | set(sbc)) if wbc.get(c) != sbc.get(c)]
    json.dump(diff, open(state_path(args, "diff_countries.json"), "w", encoding="utf-8"))

    print("workbook countries : %d" % len(wbc))
    print("supabase countries : %d" % len(sbc))
    print("differing          : %d" % len(diff))
    for c in diff:
        print("  %-30s workbook=%s  supabase=%s" % (c, wbc.get(c), sbc.get(c)))

    if not diff:
        print("\nIn sync. Nothing to do.")
        return

    quoted = ", ".join("'" + c.replace("'", "''") + "'" for c in diff)
    print()
    print("NEXT: run the row-hash query in references/sql.md (step 3) with this country list:")
    print("  %s" % quoted)
    print("Save its `packed` value to %s," % state_path(args, "supabase_rows.txt"))
    print("then run: cl_lookup.py rows")


# ---------------------------------------------------------------- rows

def cmd_rows(args):
    rows = load_rows(args)
    countries = json.load(open(state_path(args, "diff_countries.json"), encoding="utf-8"))
    sb_file = state_path(args, "supabase_rows.txt")
    if not os.path.exists(sb_file):
        sys.exit("Save the Supabase row-hash `packed` value to %s first." % sb_file)

    # supabase side: country^team^hash8
    sb = collections.defaultdict(collections.Counter)
    for part in open(sb_file, encoding="utf-8").read().strip().split("|"):
        if not part:
            continue
        country, team, h = part.split("^", 2)
        sb[(country, team)][h] += 1

    wb = collections.defaultdict(collections.Counter)
    wb_rows = collections.defaultdict(list)
    for r in rows:
        if r.get("country") in countries:
            k = (r.get("country") or "", r.get("team") or "")
            h = rowhash(r)[:8]
            wb[k][h] += 1
            wb_rows[(k, h)].append(r)

    protected = load_protected()
    prot_keys = {(p["country"], p["team"]) for p in protected}

    adds, removes, held = [], [], []
    for k in sorted(set(wb) | set(sb)):
        gained = wb[k] - sb[k]
        lost = sb[k] - wb[k]
        if not gained and not lost:
            continue
        if k in prot_keys:
            held.append((k, gained, lost))
            continue
        for h, n in gained.items():
            for _ in range(n):
                adds.append((k, h))
        for h, n in lost.items():
            for _ in range(n):
                removes.append((k, h))

    # A key with one gain and one loss is an edit, not a delete plus an insert.
    paired = collections.defaultdict(lambda: [0, 0])
    for k, _ in adds:
        paired[k][0] += 1
    for k, _ in removes:
        paired[k][1] += 1

    changes = [k for k, (a, r) in paired.items() if a and r]
    pure_adds = [(k, h) for k, h in adds if k not in changes]
    pure_removes = [(k, h) for k, h in removes if k not in changes]

    print("countries examined : %d" % len(countries))
    print("CHANGE (edit)      : %d" % len(changes))
    print("ADD (new club)     : %d" % len(pure_adds))
    print("REMOVE (gone)      : %d" % len(pure_removes))
    print("HELD (protected)   : %d" % len(held))
    print()

    if held:
        print("--- HELD: Supabase is deliberately ahead of the workbook. Do not overwrite. ---")
        by_key = {(p["country"], p["team"]): p for p in protected}
        for k, _, _ in held:
            p = by_key[k]
            print("  %s / %s" % k)
            print("      %s: supabase=%r  workbook=%r" % (p["field"], p["supabase"], p["workbook"]))
            print("      ruling: %s" % p["ruling"])
            print("      to retire this hold: %s" % p["fix_workbook"])
        print()

    stmts = []

    if changes:
        print("--- CHANGE: the workbook value wins. ---")
        for k in sorted(changes):
            country, team = k
            for r in [x for (kk, _), lst in wb_rows.items() if kk == k for x in lst]:
                print("  %s / %s -> city=%r metro=%r league=%r level=%r"
                      % (country, team, r["city"], r["metro_area"], r["league"], r["level"]))
                sets = ", ".join('%s = %s' % (('"long"' if f == "long" else f), sql_literal(r[f]))
                                 for f in FIELDS)
                stmts.append(
                    "update public.football_lookup set %s, updated_at = now()\n"
                    " where country = %s and team = %s;"
                    % (sets, sql_literal(country), sql_literal(team)))
        print()

    if pure_adds:
        print("--- ADD: in the workbook, missing from Supabase. ---")
        seen = set()
        for k, h in pure_adds:
            for r in wb_rows[(k, h)]:
                if id(r) in seen:
                    continue
                seen.add(id(r))
                print("  %s / %s  city=%r metro=%r" % (k[0], k[1], r["city"], r["metro_area"]))
                cols = ", ".join('"long"' if f == "long" else f for f in FIELDS)
                vals = ", ".join(sql_literal(r[f]) for f in FIELDS)
                stmts.append(
                    "insert into public.football_lookup (%s)\nselect %s\n"
                    " where not exists (select 1 from public.football_lookup\n"
                    "                    where country = %s and team = %s);"
                    % (cols, vals, sql_literal(k[0]), sql_literal(k[1])))
        print()

    if pure_removes:
        print("--- REMOVE: in Supabase, absent from the workbook. ---")
        print("    Read these before deleting. A club dropped from the sheet by accident")
        print("    looks exactly like one retired on purpose, and refresh.py will start")
        print("    raising UNMATCHED alerts for any api-football team that needed it.")
        for k, h in pure_removes:
            print("  %s / %s  (supabase row hash %s)" % (k[0], k[1], h))
            stmts.append("-- review before running:\n"
                         "-- delete from public.football_lookup where country = %s and team = %s;"
                         % (sql_literal(k[0]), sql_literal(k[1])))
        print()

    out = state_path(args, "apply.sql")
    open(out, "w", encoding="utf-8").write("\n\n".join(stmts) + ("\n" if stmts else ""))
    print("SQL written to     : %s  (%d statement(s))" % (out, len(stmts)))
    if not stmts:
        print("Nothing to apply.")
    else:
        print("Review it, then run each statement through the Supabase MCP.")
        print("Afterwards run the verification query in references/sql.md (step 5).")


# ---------------------------------------------------------------- main

def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--state", default=DEFAULT_STATE,
                    help="scratch directory for intermediate files (default: %(default)s)")
    sub = ap.add_subparsers(dest="cmd", required=True)

    p = sub.add_parser("extract", help="read the Lookup sheet into lookup.json")
    p.add_argument("--workbook", default=os.environ.get("CL_WORKBOOK", DEFAULT_WORKBOOK))
    p.set_defaults(func=cmd_extract)

    p = sub.add_parser("countries", help="compare per-country hashes")
    p.set_defaults(func=cmd_countries)

    p = sub.add_parser("rows", help="compare per-row hashes and emit the SQL")
    p.set_defaults(func=cmd_rows)

    args = ap.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()

#!/usr/bin/env python3
"""
Extract data from MetroAreas.xlsx and generate JSON files for the website.

Usage:
    python scripts/extract.py [path/to/MetroAreas.xlsx]

If no path is given, looks for MetroAreas.xlsx in the parent directory.

Outputs:
    public/data/metros.json          - Main rankings (all 4,285 metros)
    public/data/regions.json         - Regional aggregates
    public/data/details/<slug>.json  - Per-metro detail files
"""

import json
import math
import os
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--quiet", "--break-system-packages"])
    import openpyxl


def slugify(name):
    """Convert metro name to URL-safe slug."""
    s = name.lower().strip()
    s = re.sub(r'[àáâãäå]', 'a', s)
    s = re.sub(r'[èéêë]', 'e', s)
    s = re.sub(r'[ìíîï]', 'i', s)
    s = re.sub(r'[òóôõö]', 'o', s)
    s = re.sub(r'[ùúûü]', 'u', s)
    s = re.sub(r'[ñ]', 'n', s)
    s = re.sub(r'[ç]', 'c', s)
    s = re.sub(r'[ß]', 'ss', s)
    s = re.sub(r'[^a-z0-9\s-]', '', s)
    s = re.sub(r'[\s]+', '-', s)
    s = re.sub(r'-+', '-', s)
    return s.strip('-')


def safe_int(val, default=0):
    if val is None:
        return default
    try:
        return int(val)
    except (ValueError, TypeError):
        return default


def safe_float(val, default=0.0):
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def safe_str(val):
    if val is None:
        return ''
    return str(val).strip()


def extract_metros(wb):
    """Extract main metro data from the Metro Areas sheet."""
    ws = wb["Metro Areas"]
    metros = []

    for row in ws.iter_rows(min_row=4, values_only=True):
        v = list(row)
        name = safe_str(v[5])
        score = safe_float(v[58])
        if not name:
            continue

        pop = safe_int(v[9])
        lat = safe_float(v[63])
        lon = safe_float(v[64])
        region = safe_str(v[65]) or safe_str(v[41])
        slug = slugify(name)

        metro = {
            'slug': slug,
            'name': name,
            'country': safe_str(v[0]),
            'subCountry': safe_str(v[1]),
            'language': safe_str(v[2]),
            'capital': safe_str(v[3]),
            'gawcClass': safe_str(v[4]),
            'primaryState': safe_str(v[6]),
            'pop': pop,
            'region': region,
            'continent': safe_str(v[41]),
            'score': round(score, 1),
            'lat': lat,
            'lon': lon,
            'primaryCity': safe_str(v[61]),
            'primaryCityCountry': safe_str(v[62]),
            'gdp': round(safe_float(v[69]), 1),
            'gdpPerCapita': round(safe_float(v[70])),
            # Raw dimension data
            'dims': {
                'majorLeagueTeams': safe_int(v[43]),
                'totalTeams': safe_int(v[42]),
                'majorSportingEvents': safe_int(v[44]),
                'companies': safe_int(v[45]),
                'marketCap': safe_float(v[46]),
                'culturalEvents': safe_int(v[47]),
                'universities': safe_int(v[48]),
                'topUniHospResearch': safe_int(v[49]),
                'museumsLandmarks': safe_int(v[50]),
                'portsExchangesInfra': safe_int(v[51]),
                'airportScore': safe_float(v[52]),
                'luxuryStars': safe_float(v[53]),
                'metroStations': safe_int(v[54]),
                'suburbStations': safe_int(v[55]),
                'trainHubs': safe_int(v[56]),
                'skyscrapers': safe_int(v[57]),
            },
            'naRank': safe_int(v[59]) if v[59] else None,
            'pctOfCountry': round(safe_float(v[67]) * 100, 1) if v[67] else 0,
        }
        metros.append(metro)

    # Sort by score descending and assign global rank
    metros.sort(key=lambda x: x['score'], reverse=True)
    for i, m in enumerate(metros):
        m['rank'] = i + 1

    return metros


def extract_teams(wb):
    """Extract teams grouped by metro."""
    ws = wb["Team List"]
    teams = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        v = list(row)
        metro = safe_str(v[6])
        if not metro:
            continue
        teams.setdefault(metro, []).append({
            'sport': safe_str(v[0]),
            'league': safe_str(v[1]),
            'team': safe_str(v[2]),
            'city': safe_str(v[5]),
            'country': safe_str(v[8]),
            'level': safe_str(v[9]),
            'major': safe_str(v[11]) == 'Y',
        })

    # Also get RegTeams
    if "RegTeams" in wb.sheetnames:
        ws2 = wb["RegTeams"]
        for row in ws2.iter_rows(min_row=2, values_only=True):
            v = list(row)
            metro = safe_str(v[6])
            if not metro:
                continue
            teams.setdefault(metro, []).append({
                'sport': safe_str(v[0]),
                'league': safe_str(v[1]),
                'team': safe_str(v[2]),
                'city': safe_str(v[5]),
                'country': safe_str(v[8]),
                'level': safe_str(v[9]),
                'major': safe_str(v[11]) == 'Y',
            })

    return teams


def extract_universities(wb):
    """Extract universities grouped by metro."""
    ws = wb["Universities"]
    unis = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        v = list(row)
        metro = safe_str(v[5])
        if not metro:
            continue
        unis.setdefault(metro, []).append({
            'rank': safe_int(v[2]),
            'name': safe_str(v[3]),
            'city': safe_str(v[4]),
            'country': safe_str(v[0]),
        })
    # Sort each metro's universities by rank
    for metro in unis:
        unis[metro].sort(key=lambda x: x['rank'] if x['rank'] else 9999)
    return unis


def extract_culture(wb):
    """Extract cultural assets and infrastructure grouped by metro."""
    ws = wb["Culture-Infra"]
    culture = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        v = list(row)
        metro = safe_str(v[6])
        if not metro:
            continue
        culture.setdefault(metro, []).append({
            'name': safe_str(v[4]),
            'city': safe_str(v[5]),
            'subtype': safe_str(v[8]),
            'type': safe_str(v[11]),
            'majorType': safe_str(v[10]),
        })
    return culture


def extract_skyscrapers(wb):
    """Extract skyscraper data grouped by metro."""
    ws = wb["Skyscrapers"]
    scrapers = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        v = list(row)
        metro = safe_str(v[4])
        if not metro:
            continue
        scrapers[metro] = {
            'city': safe_str(v[2]),
            'over150m': safe_int(v[9]),
            'over200m': safe_int(v[10]),
            'over300m': safe_int(v[11]),
        }
    return scrapers


def extract_luxury(wb):
    """Extract luxury hospitality grouped by metro."""
    ws = wb["Luxury Hospitality"]
    luxury = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        v = list(row)
        metro = safe_str(v[6])
        if not metro:
            continue
        luxury.setdefault(metro, []).append({
            'name': safe_str(v[4]),
            'city': safe_str(v[5]),
            'type': safe_str(v[8]),
        })
    return luxury


def extract_events(wb):
    """Extract sporting events (Golf, Tennis, F1) grouped by metro."""
    ws = wb["Golf-Tennis-F1"]
    events = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        v = list(row)
        metro = safe_str(v[6])
        if not metro:
            continue
        events.setdefault(metro, []).append({
            'sport': safe_str(v[0]),
            'event': safe_str(v[1]),
            'year': safe_str(v[2]),
            'venue': safe_str(v[3]),
        })
    return events


def extract_mktcap(wb):
    """Extract market cap data grouped by metro."""
    ws = wb["MktCap_Data"]
    mktcap = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        v = list(row)
        metro = safe_str(v[0])
        val = safe_float(v[1])
        if not metro or val == 0:
            continue
        mktcap.setdefault(metro, []).append(val)
    # Sort each metro's companies by value descending
    for metro in mktcap:
        mktcap[metro].sort(reverse=True)
    return mktcap


def extract_football(wb):
    """Extract football club data grouped by metro."""
    ws = wb["FootballClub_Data"]
    football = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        v = list(row)
        metro = safe_str(v[0])
        if not metro:
            continue
        football.setdefault(metro, []).append({
            'level': safe_int(v[1]),
            'major': safe_str(v[2]) == 'Y' if v[2] else False,
        })
    return football


def compute_regions(metros):
    """Compute regional aggregates."""
    regions = {}
    for m in metros:
        r = m['region']
        if not r:
            continue
        if r not in regions:
            regions[r] = {
                'name': r,
                'metros': 0,
                'above50': 0,
                'above20': 0,
                'totalScore': 0,
                'top3': [],
                'totalPop': 0,
                'totalMarketCap': 0,
                'scores': [],
            }
        reg = regions[r]
        reg['metros'] += 1
        reg['totalScore'] += m['score']
        reg['totalPop'] += m['pop']
        reg['totalMarketCap'] += m['dims']['marketCap']
        reg['scores'].append(m['score'])
        if m['score'] >= 50:
            reg['above50'] += 1
        if m['score'] >= 20:
            reg['above20'] += 1
        if len(reg['top3']) < 3:
            reg['top3'].append({
                'name': m['name'],
                'score': m['score'],
                'rank': m['rank'],
                'slug': m['slug'],
            })

    # Compute medians
    for r in regions.values():
        scores = sorted(r['scores'])
        n = len(scores)
        r['medianScore'] = round(scores[n // 2], 2) if n else 0
        r['totalMarketCap'] = round(r['totalMarketCap'])
        del r['scores']

    return list(regions.values())


def build_detail(metro_name, teams, unis, culture, scrapers, luxury, events, mktcap, football):
    """Build a detail JSON object for a single metro."""
    detail = {}

    if metro_name in teams:
        detail['teams'] = teams[metro_name]

    if metro_name in unis:
        detail['universities'] = unis[metro_name]

    if metro_name in culture:
        # Group by type
        by_type = {}
        for item in culture[metro_name]:
            t = item['type'] or 'Other'
            by_type.setdefault(t, []).append(item)
        detail['culture'] = by_type

    if metro_name in scrapers:
        detail['skyscrapers'] = scrapers[metro_name]

    if metro_name in luxury:
        detail['luxury'] = luxury[metro_name]

    if metro_name in events:
        detail['events'] = events[metro_name]

    if metro_name in mktcap:
        caps = mktcap[metro_name]
        detail['marketCap'] = {
            'total': round(sum(caps)),
            'count': len(caps),
            'top10': [round(v) for v in caps[:10]],
        }

    if metro_name in football:
        clubs = football[metro_name]
        detail['football'] = {
            'total': len(clubs),
            'byLevel': {},
        }
        for c in clubs:
            lvl = str(c['level'])
            detail['football']['byLevel'][lvl] = detail['football']['byLevel'].get(lvl, 0) + 1

    return detail


def main():
    # Find the Excel file
    script_dir = Path(__file__).parent
    site_dir = script_dir.parent

    if len(sys.argv) > 1:
        xlsx_path = Path(sys.argv[1])
    else:
        # Look in parent of site/ directory
        xlsx_path = site_dir.parent / "MetroAreas.xlsx"

    if not xlsx_path.exists():
        print(f"ERROR: Cannot find {xlsx_path}")
        print(f"Usage: python {sys.argv[0]} [path/to/MetroAreas.xlsx]")
        sys.exit(1)

    print(f"Reading {xlsx_path}...")
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)

    # Extract all data
    print("Extracting metro data...")
    metros = extract_metros(wb)
    print(f"  {len(metros)} metros")

    print("Extracting teams...")
    teams = extract_teams(wb)
    print(f"  {sum(len(v) for v in teams.values())} teams across {len(teams)} metros")

    print("Extracting universities...")
    unis = extract_universities(wb)
    print(f"  {sum(len(v) for v in unis.values())} universities")

    print("Extracting cultural assets...")
    culture_data = extract_culture(wb)
    print(f"  {sum(len(v) for v in culture_data.values())} assets")

    print("Extracting skyscrapers...")
    scrapers = extract_skyscrapers(wb)
    print(f"  {len(scrapers)} cities with skyscrapers")

    print("Extracting luxury hospitality...")
    luxury = extract_luxury(wb)
    print(f"  {sum(len(v) for v in luxury.values())} entries")

    print("Extracting sporting events...")
    events = extract_events(wb)
    print(f"  {sum(len(v) for v in events.values())} events")

    print("Extracting market cap data...")
    mktcap = extract_mktcap(wb)
    print(f"  {sum(len(v) for v in mktcap.values())} companies")

    print("Extracting football clubs...")
    football = extract_football(wb)
    print(f"  {sum(len(v) for v in football.values())} clubs")

    wb.close()

    # Compute regions
    print("Computing regional aggregates...")
    regions = compute_regions(metros)

    # Output directories
    data_dir = site_dir / "public" / "data"
    details_dir = data_dir / "details"
    data_dir.mkdir(parents=True, exist_ok=True)
    details_dir.mkdir(parents=True, exist_ok=True)

    # Write main metros.json (slim version for rankings table)
    print("Writing metros.json...")
    slim_metros = []
    for m in metros:
        slim_metros.append({
            'rank': m['rank'],
            'slug': m['slug'],
            'name': m['name'],
            'country': m['country'],
            'region': m['region'],
            'pop': m['pop'],
            'score': m['score'],
            'lat': m['lat'],
            'lon': m['lon'],
            'primaryCity': m['primaryCity'],
            'gdp': m['gdp'],
            # Key dimension summaries for the table
            'majorTeams': m['dims']['majorLeagueTeams'],
            'companies': m['dims']['companies'],
            'marketCap': m['dims']['marketCap'],
            'skyscrapers': m['dims']['skyscrapers'],
            'metroStations': m['dims']['metroStations'],
            'universities': m['dims']['universities'],
        })

    with open(data_dir / "metros.json", 'w') as f:
        json.dump(slim_metros, f, separators=(',', ':'))
    size = os.path.getsize(data_dir / "metros.json")
    print(f"  metros.json: {size:,} bytes ({size/1024:.0f} KB)")

    # Write regions.json
    print("Writing regions.json...")
    with open(data_dir / "regions.json", 'w') as f:
        json.dump(regions, f, separators=(',', ':'))

    # Write per-metro detail files
    print("Writing detail files...")
    detail_count = 0
    total_detail_size = 0
    slug_map = {}  # Track slugs to handle duplicates

    for m in metros:
        slug = m['slug']
        # Handle duplicate slugs
        if slug in slug_map:
            slug = f"{slug}-{m['country'].lower().replace(' ', '-')}"
            m['slug'] = slug
        slug_map[slug] = True

        detail = build_detail(
            m['name'], teams, unis, culture_data, scrapers,
            luxury, events, mktcap, football
        )

        # Add the full metro data to the detail file
        detail['metro'] = m

        detail_path = details_dir / f"{slug}.json"
        with open(detail_path, 'w') as f:
            json.dump(detail, f, separators=(',', ':'))

        fsize = os.path.getsize(detail_path)
        total_detail_size += fsize
        detail_count += 1

    print(f"  {detail_count} detail files, {total_detail_size/1024/1024:.1f} MB total")

    # Rewrite metros.json with corrected slugs
    for sm in slim_metros:
        matched = next((m for m in metros if m['name'] == sm['name'] and m['country'] == sm['country']), None)
        if matched:
            sm['slug'] = matched['slug']

    with open(data_dir / "metros.json", 'w') as f:
        json.dump(slim_metros, f, separators=(',', ':'))

    print("\nDone! Data files written to public/data/")
    print(f"  metros.json    ({len(slim_metros)} metros)")
    print(f"  regions.json   ({len(regions)} regions)")
    print(f"  details/       ({detail_count} files)")


if __name__ == "__main__":
    main()
